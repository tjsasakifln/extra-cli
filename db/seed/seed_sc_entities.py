#!/usr/bin/env python3
"""
Seed sc_public_entities from the Extra Construtora spreadsheet.

Reads 'Extra - alvos de licitacao. R-0.xlsx' and upserts all ~2,085
SC public entities into the sc_public_entities table with proper
IBGE code resolution via BrasilAPI and distance calculation.

Usage:
    python -m db.seed.seed_sc_entities [--dsn POSTGRES_DSN]

Environment:
    LOCAL_DATALAKE_DSN     PostgreSQL connection string
    IBGE_CACHE_PATH        Path to IBGE cache JSON (default: data/ibge_cache.json)
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from math import asin, cos, radians, sin, sqrt
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2
import psycopg2.extras
import requests

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

FLORIANOPOLIS_LAT = -27.5954
FLORIANOPOLIS_LNG = -48.5480
EARTH_RADIUS_KM = 6371.0

BRASIL_API_URL = "https://brasilapi.com.br/api/ibge/municipios/v1/SC"
IBGE_API_URL = "https://servicodados.ibge.gov.br/api/v1/localidades/estados/42/municipios"

SHEET_NAME = "Entes Públicos SC"
TABLE_NAME = "sc_public_entities"
UNIQUE_INDEX_NAME = "idx_spe_cnpj_unique"

EXPECTED_COUNT = 2085

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("seed_sc_entities")


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class EntityRow:
    """A single entity row from the spreadsheet, cleaned and ready for upsert."""

    razao_social: str
    cnpj_8: str
    municipio: str | None
    codigo_ibge: str | None
    natureza_juridica: str | None
    cod_natureza: str | None
    latitude: float | None
    longitude: float | None
    distancia_fk: float | None
    raio_200km: bool


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    p = argparse.ArgumentParser(description="Seed SC public entities from spreadsheet")
    p.add_argument(
        "--dsn",
        default=os.getenv("LOCAL_DATALAKE_DSN") or None,
        help="PostgreSQL connection DSN (required via --dsn or LOCAL_DATALAKE_DSN; no weak default)",
    )
    p.add_argument(
        "--xlsx",
        default=None,
        help="Path to spreadsheet (default: auto-detect in project root)",
    )
    p.add_argument(
        "--cache",
        default=os.getenv("IBGE_CACHE_PATH", ""),
        help="Path to IBGE cache JSON file",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Print rows without touching the database",
    )
    p.add_argument(
        "--no-ibge-fetch",
        action="store_true",
        help="Skip BrasilAPI lookup for missing IBGE codes",
    )
    return p.parse_args()


# ---------------------------------------------------------------------------
# IBGE code resolution
# ---------------------------------------------------------------------------


def normalize_municipio(name: str) -> str:
    """Normalize a municipality name for fuzzy matching against API responses.

    Steps:
        1. Unicode NFKD decomposition (accents become separate combining chars)
        2. Strip combining diacritical marks
        3. Lowercase, strip non-alphanumeric, collapse whitespace

    This ensures "Agrolândia" (spreadsheet) matches "Agrolândia" (API).
    """
    # Decompose unicode: "Agrolândia" -> "Agro\0302landia" (â decomposes to a + combining)
    name = unicodedata.normalize("NFKD", name)
    # Remove combining diacritical marks (category Mn = Mark, Nonspacing)
    name = "".join(c for c in name if unicodedata.category(c) != "Mn")
    name = name.lower().strip()
    name = re.sub(r"[^a-z0-9\s]", "", name)
    name = re.sub(r"\s+", " ", name)
    return name.strip()


def load_ibge_cache(cache_path: Path) -> dict[str, str]:
    """Load cached IBGE municipality-to-code mapping from a JSON file."""
    if cache_path and cache_path.exists():
        with open(cache_path) as f:
            return json.load(f)
    return {}


def save_ibge_cache(cache_path: Path, data: dict[str, str]) -> None:
    """Save IBGE municipality-to-code mapping to a JSON cache file."""
    if not cache_path:
        return
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    with open(cache_path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    log.info("IBGE cache saved to %s (%d entries)", cache_path, len(data))


def fetch_ibge_codes_from_brasilapi() -> dict[str, str]:
    """Fetch SC municipality IBGE codes from the BrasilAPI.

    Response format:
        [{"codigo_ibge": "4200102", "nome": "Abdon Batista"}, ...]

    Returns:
        Dict mapping normalized municipality name -> 7-digit IBGE code.

    Raises:
        requests.RequestException on failure.
    """
    log.info("Fetching IBGE codes from BrasilAPI: %s", BRASIL_API_URL)
    resp = requests.get(BRASIL_API_URL, timeout=15)
    resp.raise_for_status()
    data = resp.json()

    mapping: dict[str, str] = {}
    for entry in data:
        codigo = entry.get("codigo_ibge", "")
        nome = entry.get("nome", "")
        if codigo and nome:
            key = normalize_municipio(nome)
            mapping[key] = codigo

    log.info("BrasilAPI returned %d municipalities for SC", len(mapping))
    return mapping


def fetch_ibge_codes_from_ibge_api() -> dict[str, str]:
    """Fallback: fetch SC municipality IBGE codes from the official IBGE API.

    Response format:
        [{"id": 4200102, "nome": "Abdon Batista"}, ...]

    Returns:
        Dict mapping normalized municipality name -> 7-digit IBGE code.
    """
    log.info("Fetching IBGE codes from IBGE API (fallback): %s", IBGE_API_URL)
    resp = requests.get(IBGE_API_URL, timeout=15)
    resp.raise_for_status()
    data = resp.json()

    mapping: dict[str, str] = {}
    for entry in data:
        codigo = str(entry.get("id", ""))
        nome = entry.get("nome", "")
        if codigo and nome:
            key = normalize_municipio(nome)
            mapping[key] = codigo

    log.info("IBGE API returned %d municipalities for SC", len(mapping))
    return mapping


def build_ibge_mapping(
    cache_path: Path | None,
    skip_fetch: bool = False,
) -> dict[str, str]:
    """Build the IBGE code mapping, using cache and API as needed.

    Priority: cache -> BrasilAPI -> IBGE API (fallback).

    Args:
        cache_path: Path to the cache JSON file, or None to skip caching.
        skip_fetch: If True, only use cache; do not call external APIs.

    Returns:
        Dict mapping normalized municipality name -> IBGE code.
    """
    mapping = load_ibge_cache(cache_path) if cache_path else {}
    log.info("Loaded %d entries from IBGE cache", len(mapping))

    if not mapping and not skip_fetch:
        try:
            mapping = fetch_ibge_codes_from_brasilapi()
        except requests.RequestException as exc:
            log.warning("BrasilAPI failed: %s", exc)
            try:
                mapping = fetch_ibge_codes_from_ibge_api()
            except requests.RequestException as exc2:
                log.error("IBGE API also failed: %s", exc2)
                log.warning("Proceeding without IBGE code resolution")

        if mapping:
            save_ibge_cache(cache_path, mapping)

    return mapping


def resolve_ibge_code(
    municipio: str | None,
    ibge_mapping: dict[str, str],
) -> str | None:
    """Resolve an IBGE code from a municipality name using the mapping.

    Uses multi-strategy matching:
        1. Exact normalized name match
        2. Remove connecting words (de, da, do, das, dos) and retry
        3. Remove all spaces and retry

    Returns None if the municipality is unknown or None.
    """
    if not municipio:
        return None

    key = normalize_municipio(municipio)

    # Strategy 1: exact match
    if key in ibge_mapping:
        return ibge_mapping[key]

    # Strategy 2: remove connecting words (de, da, do, das, dos)
    stripped = re.sub(r"\b(de|da|do|das|dos)\b", "", key)
    stripped = re.sub(r"\s+", " ", stripped).strip()
    if stripped and stripped != key and stripped in ibge_mapping:
        log.debug("IBGE fuzzy match: '%s' -> '%s' (removed connecting words)", key, stripped)
        return ibge_mapping[stripped]

    # Strategy 3: remove all spaces (handles "grao para" vs "graopara")
    no_space = key.replace(" ", "")
    if no_space != key and no_space in ibge_mapping:
        log.debug("IBGE fuzzy match: '%s' -> '%s' (no spaces)", key, no_space)
        return ibge_mapping[no_space]

    # Strategy 4: try matching by first word only (catches minor variations)
    first_word = key.split()[0] if key.split() else ""
    for map_key, map_val in ibge_mapping.items():
        if map_key.startswith(first_word) and len(map_key) <= len(key) + 3:
            log.debug("IBGE fuzzy match: '%s' -> '%s' (prefix match)", key, map_key)
            return map_val

    return None


# ---------------------------------------------------------------------------
# Distance calculation
# ---------------------------------------------------------------------------


def haversine_km(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    """Calculate the great-circle distance in km between two lat/lng points.

    Uses the Haversine formula.
    """
    lat1_r, lng1_r = radians(lat1), radians(lng1)
    lat2_r, lng2_r = radians(lat2), radians(lng2)

    dlat = lat2_r - lat1_r
    dlng = lng2_r - lng1_r

    a = sin(dlat / 2) ** 2 + cos(lat1_r) * cos(lat2_r) * sin(dlng / 2) ** 2
    c = 2 * asin(sqrt(a))

    return c * EARTH_RADIUS_KM


# ---------------------------------------------------------------------------
# Spreadsheet reading
# ---------------------------------------------------------------------------


def find_spreadsheet(project_root: Path) -> Path:
    """Find private or public-fixture spreadsheet (no silent backup selection)."""
    import os
    env = os.environ.get("EXTRA_TARGET_SPREADSHEET") or os.environ.get("TARGET_SPREADSHEET_PATH")
    if env:
        path = Path(env).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"EXTRA_TARGET_SPREADSHEET not found: {path}")
        if any(tok in path.name.lower() for tok in (".backup", ".copy", ".tmp")):
            raise FileNotFoundError(f"Refusing backup/temp spreadsheet: {path.name}")
        return path
    preferred = project_root / "Extra - alvos de licitação. R-0.xlsx"
    if preferred.is_file():
        return preferred
    fixture = project_root / "fixtures" / "canonical_universe_r0.xlsx"
    if fixture.is_file():
        return fixture
    raise FileNotFoundError(
        "Private spreadsheet not found. Set EXTRA_TARGET_SPREADSHEET or use "
        "fixtures/canonical_universe_r0.xlsx. See docs/ops/private-assets.md."
    )


def read_spreadsheet(
    xlsx_path: Path,
    ibge_mapping: dict[str, str] | None = None,
) -> list[EntityRow]:
    """Read the spreadsheet and return a list of EntityRow instances.

    Args:
        xlsx_path: Path to the Excel file.
        ibge_mapping: Optional mapping of municipality -> IBGE code for
            resolving missing codes.

    Expected columns:
        0: Razao Social
        1: CNPJ (8 digitos)
        2: Municipio
        3: Codigo IBGE
        4: Natureza Juridica
        5: Cod Natureza
        6: Latitude
        7: Longitude
        8: Distancia FK
        9: Raio 200km
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {available}")

    ws = wb[SHEET_NAME]
    entities: list[EntityRow] = []
    ibge_resolved = 0
    ibge_pending = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header row

        if not row or not row[0]:
            continue  # empty row

        razao_social = str(row[0]).strip() if row[0] else ""
        cnpj_8 = str(row[1]).strip() if row[1] else ""
        municipio = str(row[2]).strip() if row[2] else None
        codigo_ibge = str(row[3]).strip() if row[3] else None

        natureza_juridica = str(row[4]).strip() if row[4] else None
        cod_natureza = str(row[5]).strip() if row[5] else None

        if not cnpj_8:
            log.warning("Row %d skipped: no CNPJ", i + 1)
            continue

        # Normalize CNPJ: keep only digits
        cnpj_8 = "".join(c for c in cnpj_8 if c.isdigit())

        # Normalize IBGE: keep only digits, truncate to 7 chars
        if codigo_ibge:
            codigo_ibge = "".join(c for c in codigo_ibge if c.isdigit())
            if len(codigo_ibge) > 7:
                codigo_ibge = codigo_ibge[:7]

        # Resolve missing IBGE codes via the municipality-to-code mapping
        if not codigo_ibge and ibge_mapping:
            resolved = resolve_ibge_code(municipio, ibge_mapping)
            if resolved:
                codigo_ibge = resolved
                ibge_resolved += 1
            else:
                ibge_pending += 1

        # Parse coordinates
        latitude = None
        longitude = None
        try:
            if row[6] is not None:
                latitude = float(row[6])
        except (ValueError, TypeError):
            pass
        try:
            if row[7] is not None:
                longitude = float(row[7])
        except (ValueError, TypeError):
            pass

        # Validate IBGE: must be exactly 7 digits to be a valid municipality code.
        # State-level entities (municipio = "SANTA CATARINA") often have "42"
        # (the state code) instead of a 7-digit code. Set to NULL in that case
        # since it cannot be used for municipality-constrained matching.
        if codigo_ibge and len(codigo_ibge) != 7:
            codigo_ibge = None
            if not ibge_mapping:
                ibge_pending += 1

        # Calculate distance from Florianopolis using Haversine
        # Always recalculate so the values are consistent (AC6)
        distancia_fk = None
        if latitude is not None and longitude is not None:
            distancia_fk = haversine_km(
                FLORIANOPOLIS_LAT,
                FLORIANOPOLIS_LNG,
                latitude,
                longitude,
            )

        raio_200km = False
        if distancia_fk is not None and distancia_fk <= 200.0:
            raio_200km = True

        entities.append(
            EntityRow(
                razao_social=razao_social,
                cnpj_8=cnpj_8,
                municipio=municipio,
                codigo_ibge=codigo_ibge,
                natureza_juridica=natureza_juridica,
                cod_natureza=cod_natureza,
                latitude=latitude,
                longitude=longitude,
                distancia_fk=distancia_fk,
                raio_200km=raio_200km,
            )
        )

    wb.close()

    if ibge_mapping is not None:
        log.info(
            "IBGE codes resolved via API: %d, still pending: %d",
            ibge_resolved,
            ibge_pending,
        )

    return entities


# ---------------------------------------------------------------------------
# Database operations
# ---------------------------------------------------------------------------


def ensure_unique_index(conn) -> None:
    """Create a unique index on cnpj_8 if it does not already exist.

    This enables ``ON CONFLICT (cnpj_8) DO UPDATE`` for idempotent upserts.
    The index creation is idempotent (IF NOT EXISTS).
    """
    cur = conn.cursor()
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_spe_cnpj_unique
        ON sc_public_entities (cnpj_8)
        """
    )
    cur.close()
    log.info("Unique index %s ensured on sc_public_entities(cnpj_8)", UNIQUE_INDEX_NAME)


def upsert_entities(
    conn,
    entities: list[EntityRow],
    dry_run: bool = False,
) -> dict[str, int]:
    """Upsert all entities into sc_public_entities.

    Uses ``INSERT ... ON CONFLICT (cnpj_8) DO UPDATE`` so the script is
    fully idempotent (AC7). ``is_active`` is always set to TRUE (AC5).

    Args:
        conn: Database connection.
        entities: List of EntityRow objects to upsert.
        dry_run: If True, log what would be done without inserting.

    Returns:
        Dict with 'inserted' and 'updated' counts.
    """
    counts: dict[str, int] = {"inserted": 0, "updated": 0}

    if dry_run:
        preview = entities[:5]
        for e in preview:
            log.info(
                "[DRY-RUN] Would upsert: %s | CNPJ %s | %s",
                e.razao_social[:60],
                e.cnpj_8,
                e.municipio or "(no municipio)",
            )
        remaining = len(entities) - len(preview)
        if remaining > 0:
            log.info("[DRY-RUN] ... and %d more entities", remaining)
        return counts

    cur = conn.cursor()

    # Pre-fetch existing CNPJ bases so we can distinguish INSERT from UPDATE.
    # (PostgreSQL 16 returns rowcount=1 for both paths in ON CONFLICT.)
    cur.execute(f"SELECT cnpj_8 FROM {TABLE_NAME}")
    existing_cnpjs: set[str] = {row[0] for row in cur.fetchall()}
    log.info("Existing entities in DB: %d", len(existing_cnpjs))

    for e in entities:
        was_update = e.cnpj_8 in existing_cnpjs
        cur.execute(
            f"""
            INSERT INTO {TABLE_NAME} (
                razao_social, cnpj_8, municipio, codigo_ibge,
                natureza_juridica, cod_natureza,
                latitude, longitude, distancia_fk, raio_200km,
                is_active
            ) VALUES (
                %(razao_social)s, %(cnpj_8)s, %(municipio)s, %(codigo_ibge)s,
                %(natureza_juridica)s, %(cod_natureza)s,
                %(latitude)s, %(longitude)s, %(distancia_fk)s, %(raio_200km)s,
                TRUE
            )
            ON CONFLICT (cnpj_8) DO UPDATE SET
                razao_social        = EXCLUDED.razao_social,
                municipio           = EXCLUDED.municipio,
                codigo_ibge         = COALESCE({TABLE_NAME}.codigo_ibge,
                                               EXCLUDED.codigo_ibge),
                natureza_juridica   = EXCLUDED.natureza_juridica,
                cod_natureza        = EXCLUDED.cod_natureza,
                latitude            = COALESCE({TABLE_NAME}.latitude,
                                               EXCLUDED.latitude),
                longitude           = COALESCE({TABLE_NAME}.longitude,
                                               EXCLUDED.longitude),
                distancia_fk        = EXCLUDED.distancia_fk,
                raio_200km          = EXCLUDED.raio_200km,
                is_active           = TRUE
            """,
            {
                "razao_social": e.razao_social,
                "cnpj_8": e.cnpj_8,
                "municipio": e.municipio,
                "codigo_ibge": e.codigo_ibge,
                "natureza_juridica": e.natureza_juridica,
                "cod_natureza": e.cod_natureza,
                "latitude": e.latitude,
                "longitude": e.longitude,
                "distancia_fk": e.distancia_fk,
                "raio_200km": e.raio_200km,
            },
        )

        if was_update:
            counts["updated"] += 1
        else:
            counts["inserted"] += 1
            existing_cnpjs.add(e.cnpj_8)  # track for subsequent rows

    conn.commit()
    cur.close()
    return counts


def verify_import(conn) -> dict[str, Any]:
    """Run integrity checks on the sc_public_entities table.

    Returns a dict with stats: total, null_cnpj, null_municipio, active,
    pending_ibge, raio_200km, distinct_municipios.
    """
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute(f"SELECT COUNT(*) AS total FROM {TABLE_NAME}")
    total = cur.fetchone()["total"]

    cur.execute(f"SELECT COUNT(*) AS cnt FROM {TABLE_NAME} WHERE cnpj_8 IS NULL")
    null_cnpj = cur.fetchone()["cnt"]

    cur.execute(f"SELECT COUNT(*) AS cnt FROM {TABLE_NAME} WHERE municipio IS NULL")
    null_municipio = cur.fetchone()["cnt"]

    cur.execute(f"SELECT COUNT(*) AS cnt FROM {TABLE_NAME} WHERE is_active = TRUE")
    active = cur.fetchone()["cnt"]

    cur.execute(f"SELECT COUNT(*) AS cnt FROM {TABLE_NAME} WHERE codigo_ibge IS NULL")
    pending_ibge = cur.fetchone()["cnt"]

    cur.execute(f"SELECT COUNT(*) AS cnt FROM {TABLE_NAME} WHERE raio_200km = TRUE")
    raio_count = cur.fetchone()["cnt"]

    cur.execute(f"SELECT COUNT(DISTINCT municipio) AS cnt FROM {TABLE_NAME}")
    distinct_municipios = cur.fetchone()["cnt"]

    cur.close()

    return {
        "total": total,
        "null_cnpj": null_cnpj,
        "null_municipio": null_municipio,
        "active": active,
        "pending_ibge": pending_ibge,
        "raio_200km": raio_count,
        "distinct_municipios": distinct_municipios,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    """Entry point: parse args, read spreadsheet, upsert into DB, verify."""
    args = parse_args()
    if not args.dsn:
        print("ERROR: --dsn or LOCAL_DATALAKE_DSN is required (no weak password default)", file=sys.stderr)
        sys.exit(2)
    project_root = Path(__file__).resolve().parent.parent.parent

    # -- Locate spreadsheet ---------------------------------------------------
    xlsx_path = Path(args.xlsx) if args.xlsx else find_spreadsheet(project_root)
    log.info("Spreadsheet: %s", xlsx_path)

    # -- IBGE cache path ------------------------------------------------------
    cache_path: Path | None
    if args.cache:
        cache_path = Path(args.cache)
    else:
        cache_dir = project_root / "data"
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_path = cache_dir / "ibge_cache.json"

    # -- Build IBGE municipality code mapping ---------------------------------
    ibge_mapping = build_ibge_mapping(cache_path, skip_fetch=args.no_ibge_fetch)

    # -- Read and process spreadsheet -----------------------------------------
    entities = read_spreadsheet(xlsx_path, ibge_mapping=ibge_mapping)
    log.info("Read %d entities from spreadsheet", len(entities))

    # Spreadsheet-level stats
    with_coords = sum(1 for e in entities if e.latitude is not None)
    with_ibge = sum(1 for e in entities if e.codigo_ibge)
    distinct_municipios = len({e.municipio for e in entities if e.municipio})
    log.info("  Municipalities: %d", distinct_municipios)
    log.info("  With coordinates: %d", with_coords)
    log.info("  With IBGE code: %d", with_ibge)

    # -- Dry-run? -------------------------------------------------------------
    if args.dry_run:
        upsert_entities(None, entities, dry_run=True)
        return

    # -- Connect to database --------------------------------------------------
    log.info("Connecting to database...")
    conn = psycopg2.connect(args.dsn)
    conn.autocommit = False

    try:
        # Step 1: Ensure unique index for ON CONFLICT support
        ensure_unique_index(conn)

        # Step 2: Upsert all entities
        counts = upsert_entities(conn, entities)
        log.info(
            "Upsert complete: %d inserted, %d updated",
            counts["inserted"],
            counts["updated"],
        )

        # Step 3: Verify integrity
        stats = verify_import(conn)
        log.info("=== Integrity Verification ===")
        log.info("Total entities:        %d", stats["total"])
        log.info("Active (is_active):    %d", stats["active"])
        log.info("NULL cnpj_8:           %d", stats["null_cnpj"])
        log.info("NULL municipio:        %d", stats["null_municipio"])
        log.info("Missing IBGE code:     %d", stats["pending_ibge"])
        log.info("Within 200km of FLN:   %d", stats["raio_200km"])
        log.info("Distinct municipios:   %d", stats["distinct_municipios"])

        # Step 4: Validate against acceptance criteria
        errors: list[str] = []
        if stats["total"] < EXPECTED_COUNT:
            errors.append(f"AC4 FAIL: Expected {EXPECTED_COUNT} entities, got {stats['total']}")
        if stats["null_cnpj"] > 0:
            errors.append(f"AC4 FAIL: Found {stats['null_cnpj']} entities with NULL cnpj_8")
        if stats["null_municipio"] > 0:
            errors.append(f"AC4 FAIL: Found {stats['null_municipio']} entities with NULL municipio")
        if stats["pending_ibge"] > 0:
            log.warning(
                "AC3: IBGE codes still pending for %d entities - these will skip IBGE-constrained matching",
                stats["pending_ibge"],
            )
        if stats["active"] < EXPECTED_COUNT:
            errors.append(f"AC5 FAIL: Expected {EXPECTED_COUNT} active, got {stats['active']}")

        if errors:
            for err in errors:
                log.error(err)
            log.warning("Seed completed with %d validation issue(s)", len(errors))
        else:
            log.info("All acceptance criteria passed!")

    except Exception:
        conn.rollback()
        log.exception("Fatal error during seed operation")
        sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
