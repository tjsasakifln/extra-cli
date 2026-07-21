#!/usr/bin/env python3
"""Seed sc_public_entities from the Extra Construtora spreadsheet.

Reads 'Extra - alvos de licitação. R-0.xlsx' and inserts all 2,085
SC public entities into the sc_public_entities table.

Usage:
    python db/seed/001_sc_entities.py [--dsn POSTGRES_DSN]
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Seed SC public entities from spreadsheet")
    p.add_argument(
        "--dsn",
        default=os.getenv("LOCAL_DATALAKE_DSN") or None,
        help="PostgreSQL connection DSN (required via --dsn or LOCAL_DATALAKE_DSN; no weak default)",
    )
    p.add_argument(
        "--xlsx",
        default=None,
        help="Path to spreadsheet (default: project root Extra - alvos de licitação. R-0.xlsx)",
    )
    p.add_argument("--dry-run", action="store_true", help="Print rows without inserting")
    p.add_argument("--truncate", action="store_true", help="TRUNCATE table before seeding")
    return p.parse_args()


def find_spreadsheet(project_root: Path) -> Path:
    """Find the private target spreadsheet (not expected in public git tree)."""
    env = os.environ.get("EXTRA_TARGET_SPREADSHEET") or os.environ.get("TARGET_SPREADSHEET_PATH")
    if env:
        path = Path(env).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"EXTRA_TARGET_SPREADSHEET not found: {path}")
        if any(tok in path.name.lower() for tok in (".backup", ".copy", ".tmp")):
            raise FileNotFoundError(f"Refusing backup/temp spreadsheet: {path.name}")
        return path
    preferred = project_root / "Extra - alvos de licitação. R-0.xlsx"
    if preferred.is_file():
        return preferred
    public_fixture = project_root / "fixtures" / "canonical_universe_r0.xlsx"
    if public_fixture.is_file():
        return public_fixture
    raise FileNotFoundError(
        "Private spreadsheet not found. Set EXTRA_TARGET_SPREADSHEET or pass --xlsx. "
        "Asset is not shipped in the public repository. See docs/ops/private-assets.md."
    )


def read_spreadsheet(xlsx_path: Path) -> list[dict]:
    """Read the 'Entes Públicos SC' sheet and return list of entity dicts."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if "Entes Públicos SC" not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Sheet 'Entes Públicos SC' not found. Available: {available}")

    ws = wb["Entes Públicos SC"]

    entities = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # Header row: Razão Social, CNPJ (8 dígitos), Município, ...
            continue

        if not row or not row[0]:
            continue  # Skip empty rows

        razao_social = str(row[0]).strip() if row[0] else None
        cnpj_8 = str(row[1]).strip() if row[1] else None
        municipio = str(row[2]).strip() if row[2] else None
        codigo_ibge = str(row[3]).strip() if row[3] else None
        natureza_juridica = str(row[4]).strip() if row[4] else None
        cod_natureza = str(row[5]).strip() if row[5] else None

        # Parsing coordinates
        latitude = None
        longitude = None
        try:
            if row[6]:
                latitude = float(row[6])
        except (ValueError, TypeError):
            pass
        try:
            if row[7]:
                longitude = float(row[7])
        except (ValueError, TypeError):
            pass

        # Distance from Florianópolis
        distancia = None
        try:
            if row[8]:
                distancia = float(row[8])
        except (ValueError, TypeError):
            pass

        # Raio 200km flag
        raio = False
        if row[9]:
            raio_str = str(row[9]).strip().upper()
            raio = "SIM" in raio_str or "TRUE" in raio_str

        # Skip if no CNPJ
        if not cnpj_8:
            continue

        # Normalize CNPJ: keep only digits
        cnpj_8 = "".join(c for c in cnpj_8 if c.isdigit())

        # Normalize IBGE code
        if codigo_ibge:
            codigo_ibge = "".join(c for c in codigo_ibge if c.isdigit())

        entities.append(
            {
                "razao_social": razao_social,
                "cnpj_8": cnpj_8,
                "municipio": municipio,
                "codigo_ibge": codigo_ibge,
                "natureza_juridica": natureza_juridica,
                "cod_natureza": cod_natureza,
                "latitude": latitude,
                "longitude": longitude,
                "distancia_fk": distancia,
                "raio_200km": raio,
            }
        )

    wb.close()
    return entities


def seed_database(
    conn: psycopg2.extensions.connection,
    entities: list[dict],
    truncate: bool = False,
    dry_run: bool = False,
) -> int:
    """Insert entities into sc_public_entities."""
    if dry_run:
        print(f"[DRY RUN] Would insert {len(entities)} entities")
        for e in entities[:5]:
            print(f"  {e['razao_social'][:60]:60s} | {e['cnpj_8']:14s} | {e['municipio']}")
        print(f"  ... and {len(entities) - 5} more")
        return len(entities)

    cur = conn.cursor()

    if truncate:
        cur.execute("TRUNCATE TABLE sc_public_entities RESTART IDENTITY CASCADE")
        print("[TRUNCATE] sc_public_entities cleared")

    inserted = 0
    for e in entities:
        try:
            cur.execute(
                """
                INSERT INTO sc_public_entities (
                    razao_social, cnpj_8, municipio, codigo_ibge,
                    natureza_juridica, cod_natureza,
                    latitude, longitude, distancia_fk, raio_200km
                ) VALUES (
                    %(razao_social)s, %(cnpj_8)s, %(municipio)s, %(codigo_ibge)s,
                    %(natureza_juridica)s, %(cod_natureza)s,
                    %(latitude)s, %(longitude)s, %(distancia_fk)s, %(raio_200km)s
                )
                ON CONFLICT DO NOTHING
                """,
                e,
            )
            if cur.rowcount and cur.rowcount > 0:
                inserted += 1
        except Exception as exc:
            print(f"[ERROR] {e['razao_social']}: {exc}", file=sys.stderr)

    conn.commit()
    cur.close()
    return inserted


def main():
    args = parse_args()
    if not args.dsn:
        print("ERROR: --dsn or LOCAL_DATALAKE_DSN is required (no weak password default)", file=sys.stderr)
        sys.exit(2)
    project_root = Path(__file__).resolve().parent.parent.parent
    xlsx_path = Path(args.xlsx) if args.xlsx else find_spreadsheet(project_root)

    print(f"📋 Reading spreadsheet: {xlsx_path}")
    entities = read_spreadsheet(xlsx_path)
    print(f"   {len(entities)} entities found")

    # Stats
    raio_count = sum(1 for e in entities if e["raio_200km"])
    with_coords = sum(1 for e in entities if e["latitude"] and e["longitude"])
    municipios = len(set(e["municipio"] for e in entities if e["municipio"]))
    print(f"   Within 200km: {raio_count}")
    print(f"   With coordinates: {with_coords}")
    print(f"   Distinct municipalities: {municipios}")

    if args.dry_run:
        seed_database(None, entities, dry_run=True)
        return

    print("\n🔌 Connecting to database...")
    conn = psycopg2.connect(args.dsn)
    conn.autocommit = False

    try:
        inserted = seed_database(conn, entities, truncate=args.truncate)
        print(f"✅ Inserted {inserted} entities into sc_public_entities")

        # Verify
        cur = conn.cursor()
        cur.execute("SELECT count(*) FROM sc_public_entities")
        total = cur.fetchone()[0]
        cur.execute("SELECT count(*) FROM sc_public_entities WHERE raio_200km = TRUE")
        raio_db = cur.fetchone()[0]
        cur.close()

        print(f"   DB total: {total} entities")
        print(f"   DB within 200km: {raio_db}")
    except Exception as exc:
        conn.rollback()
        print(f"❌ Error: {exc}", file=sys.stderr)
        sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
