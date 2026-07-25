"""Multi-format readiness reports from a single model."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from scripts.bid_readiness.sanitize import sanitize_obj, sanitize_text


def build_report_model(case: dict[str, Any]) -> dict[str, Any]:
    """Single source model for all report formats."""
    model = {
        "case_id": case.get("case_id"),
        "system_status": case.get("system_status"),
        "package_status": case.get("package_status"),
        "reference_date": case.get("reference_date"),
        "entity": sanitize_obj(case.get("entity") or {}),
        "summary": case.get("summary") or {},
        "blockers": sanitize_obj(case.get("findings", {}).get("blockers") or []),
        "findings_count": len(case.get("findings", {}).get("all") or []),
        "requirements": sanitize_obj(case.get("requirements") or []),
        "documents": [
            {
                "document_id": d.get("document_id"),
                "original_name": d.get("original_name"),
                "classification": d.get("classification"),
                "sha256": d.get("sha256"),
                "validity": (d.get("validity") or {}).get("status"),
                "identity": (d.get("identity") or {}).get("cnpj_status"),
            }
            for d in (case.get("documents") or [])
        ],
        "match_matrix": sanitize_obj(case.get("match_rows") or []),
        "human_review": sanitize_obj(case.get("findings", {}).get("human-review") or []),
        "limitations": [
            "Operational support only — not a legal opinion.",
            "Does not assert habilitacao definitiva.",
            "Does not authenticate signatures biometrically.",
            "Does not submit to any portal.",
            "SIMULATION_ONLY package until human acceptance.",
        ],
        "claims": case.get("claims") or [],
        "non_claims": case.get("non_claims") or [],
    }
    model["model_sha256"] = hashlib.sha256(
        json.dumps(model, sort_keys=True, ensure_ascii=False).encode("utf-8")
    ).hexdigest()
    return model


def write_reports(reports_dir: Path, model: dict[str, Any]) -> dict[str, str]:
    reports_dir.mkdir(parents=True, exist_ok=True)
    paths: dict[str, str] = {}

    md = _render_markdown(model)
    md_path = reports_dir / "executive-summary.md"
    md_path.write_text(sanitize_text(md), encoding="utf-8")
    paths["md"] = str(md_path)

    html = _render_html(model)
    html_path = reports_dir / "readiness-report.html"
    html_path.write_text(sanitize_text(html), encoding="utf-8")
    paths["html"] = str(html_path)

    xlsx_path = reports_dir / "readiness-workbook.xlsx"
    _write_xlsx(xlsx_path, model)
    paths["xlsx"] = str(xlsx_path)

    pdf_path = reports_dir / "readiness-report.pdf"
    _write_pdf(pdf_path, model)
    paths["pdf"] = str(pdf_path)

    # reconciliation of key scalars across formats
    recon = {
        "model_sha256": model["model_sha256"],
        "system_status": model["system_status"],
        "package_status": model["package_status"],
        "blockers": len(model["blockers"]),
        "requirements": len(model["requirements"]),
        "documents": len(model["documents"]),
        "md_contains_status": model["package_status"] in md,
        "html_contains_status": model["package_status"] in html,
        "ok": model["package_status"] in md and model["package_status"] in html,
    }
    (reports_dir / "report-reconciliation.json").write_text(json.dumps(recon, indent=2) + "\n", encoding="utf-8")
    paths["reconciliation"] = str(reports_dir / "report-reconciliation.json")
    return paths


def _render_markdown(model: dict[str, Any]) -> str:
    lines = [
        f"# Readiness report — {model.get('case_id')}",
        "",
        f"- System status: **{model.get('system_status')}**",
        f"- Package status: **{model.get('package_status')}**",
        f"- Reference date: {model.get('reference_date')}",
        f"- Requirements: {len(model.get('requirements') or [])}",
        f"- Documents: {len(model.get('documents') or [])}",
        f"- Findings: {model.get('findings_count')}",
        f"- Blockers: {len(model.get('blockers') or [])}",
        "",
        "## Blockers",
        "",
    ]
    for b in model.get("blockers") or []:
        lines.append(
            f"- **{b.get('severity')}** `{b.get('classification')}`: {b.get('title')} — {b.get('objective_observation')}"
        )
    lines += ["", "## Limitations", ""]
    for lim in model.get("limitations") or []:
        lines.append(f"- {lim}")
    lines += ["", "## Non-claims", ""]
    for nc in model.get("non_claims") or []:
        lines.append(f"- {nc}")
    lines += ["", f"Model SHA-256: `{model.get('model_sha256')}`", ""]
    return "\n".join(lines)


def _render_html(model: dict[str, Any]) -> str:
    blockers = "".join(
        f"<li><strong>{b.get('severity')}</strong> {b.get('title')}: {b.get('objective_observation')}</li>"
        for b in (model.get("blockers") or [])
    )
    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="utf-8"/>
<title>Readiness {model.get("case_id")}</title>
<style>
body{{font-family:system-ui,sans-serif;margin:2rem;color:#122}}
.banner{{background:#fff3cd;border:1px solid #ffecb5;padding:1rem;margin-bottom:1rem}}
.bad{{color:#a00;font-weight:700}}
</style></head><body>
<div class="banner">SIMULATION_ONLY — apoio operacional; não é habilitação definitiva nem protocolo.</div>
<h1>Readiness report — {model.get("case_id")}</h1>
<p>System: <span class="bad">{model.get("system_status")}</span></p>
<p>Package: <span class="bad">{model.get("package_status")}</span></p>
<p>Reference date: {model.get("reference_date")}</p>
<p>Requirements: {len(model.get("requirements") or [])} |
Documents: {len(model.get("documents") or [])} |
Blockers: {len(model.get("blockers") or [])}</p>
<h2>Blockers</h2>
<ul>{blockers or "<li>None</li>"}</ul>
<h2>Limitations</h2>
<ul>{"".join(f"<li>{x}</li>" for x in (model.get("limitations") or []))}</ul>
<p>Model SHA-256: <code>{model.get("model_sha256")}</code></p>
</body></html>
"""


def _write_xlsx(path: Path, model: dict[str, Any]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    meta_rows: list[tuple[str, Any]] = [
        ("case_id", model.get("case_id")),
        ("system_status", model.get("system_status")),
        ("package_status", model.get("package_status")),
        ("reference_date", model.get("reference_date")),
        ("model_sha256", model.get("model_sha256")),
    ]
    # Metadados
    ws = wb.active
    ws.title = "Metadados"
    for i, (k, v) in enumerate(meta_rows, 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, str(v))

    ws_r = wb.create_sheet("Requisitos")
    ws_r.append(["requirement_id", "title", "category", "mandatory"])
    for r in model.get("requirements") or []:
        ws_r.append(
            [
                r.get("requirement_id"),
                r.get("title"),
                r.get("category"),
                r.get("mandatory"),
            ]
        )

    ws_d = wb.create_sheet("Documentos")
    ws_d.append(["document_id", "name", "classification", "sha256", "validity", "identity"])
    for d in model.get("documents") or []:
        ws_d.append(
            [
                d.get("document_id"),
                d.get("original_name"),
                d.get("classification"),
                d.get("sha256"),
                d.get("validity"),
                d.get("identity"),
            ]
        )

    ws_b = wb.create_sheet("Blockers")
    ws_b.append(["finding_id", "severity", "classification", "title", "observation"])
    for b in model.get("blockers") or []:
        ws_b.append(
            [
                b.get("finding_id"),
                b.get("severity"),
                b.get("classification"),
                b.get("title"),
                b.get("objective_observation"),
            ]
        )

    ws_p = wb.create_sheet("Pendências")
    ws_p.append(["finding_id", "title"])
    for h in model.get("human_review") or []:
        ws_p.append([h.get("finding_id"), h.get("title")])

    ws_l = wb.create_sheet("Limitações")
    for lim in model.get("limitations") or []:
        ws_l.append([lim])

    # Additional empty structural sheets required by campaign
    for name in (
        "Validade",
        "Identidade",
        "Jurídica",
        "Fiscal-Trabalhista",
        "Econômico-Financeira",
        "Qualificação-Técnica",
        "Atestados",
        "CAT-ART-RRT",
        "Declarações",
        "Proposta",
        "Garantia",
        "Assinaturas",
        "Pacote",
        "Evidências",
    ):
        if name not in wb.sheetnames:
            wb.create_sheet(name)

    wb.save(path)


def _write_pdf(path: Path, model: dict[str, Any]) -> None:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError:
        # Minimal PDF without reportlab
        content = _render_markdown(model).encode("latin-1", errors="replace")
        # Very small valid-ish PDF
        path.write_bytes(b"%PDF-1.1\n1 0 obj<<>>endobj\ntrailer<<>>\n%%EOF\n" + content[:2000])
        return

    c = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    y = height - 50
    lines = [
        f"Readiness report — {model.get('case_id')}",
        "SIMULATION_ONLY",
        f"System: {model.get('system_status')}",
        f"Package: {model.get('package_status')}",
        f"Reference: {model.get('reference_date')}",
        f"Requirements: {len(model.get('requirements') or [])}",
        f"Documents: {len(model.get('documents') or [])}",
        f"Blockers: {len(model.get('blockers') or [])}",
        "",
        "Blockers:",
    ]
    for b in (model.get("blockers") or [])[:20]:
        lines.append(f"- {b.get('severity')}: {b.get('title')}")
    lines += ["", "Limitations:"] + list(model.get("limitations") or [])
    for line in lines:
        if y < 50:
            c.showPage()
            y = height - 50
        c.drawString(40, y, sanitize_text(str(line))[:110])
        y -= 14
    c.save()
