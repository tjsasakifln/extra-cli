"""Load and normalize requirements from json/yaml/csv/xlsx."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

REQUIRED_KEYS = (
    "requirement_id",
    "category",
    "title",
    "description",
    "mandatory",
    "condition",
    "source_document",
    "source_locator",
    "source_excerpt",
    "deadline",
    "required_document_type",
    "required_issuer",
    "required_holder",
    "required_signatory",
    "validity_rule",
    "technical_criteria",
    "financial_criteria",
    "submission_format",
    "original_file_required",
    "signature_required",
    "authentication_required",
    "human_interpretation_required",
)


def _default_requirement(raw: dict[str, Any], idx: int) -> dict[str, Any]:
    rid = str(raw.get("requirement_id") or raw.get("id") or f"REQ-{idx:03d}")
    out: dict[str, Any] = {k: raw.get(k) for k in REQUIRED_KEYS}
    out["requirement_id"] = rid
    out["category"] = str(raw.get("category") or "OUTRO")
    out["title"] = str(raw.get("title") or raw.get("name") or rid)
    out["description"] = str(raw.get("description") or "")
    mand = raw.get("mandatory", True)
    if isinstance(mand, str):
        mand = mand.strip().lower() in {"1", "true", "yes", "sim", "obrigatorio", "obrigatório"}
    out["mandatory"] = bool(mand)
    out["condition"] = raw.get("condition")
    out["source_document"] = raw.get("source_document") or raw.get("source") or "edital"
    out["source_locator"] = raw.get("source_locator") or raw.get("locator") or {}
    out["source_excerpt"] = raw.get("source_excerpt") or raw.get("excerpt")
    out["deadline"] = raw.get("deadline")
    out["required_document_type"] = raw.get("required_document_type") or raw.get("document_type")
    out["required_issuer"] = raw.get("required_issuer")
    out["required_holder"] = raw.get("required_holder")
    out["required_signatory"] = raw.get("required_signatory")
    out["validity_rule"] = raw.get("validity_rule") or {}
    out["technical_criteria"] = raw.get("technical_criteria") or {}
    out["financial_criteria"] = raw.get("financial_criteria") or {}
    out["submission_format"] = raw.get("submission_format") or {}
    out["original_file_required"] = bool(raw.get("original_file_required", False))
    out["signature_required"] = bool(raw.get("signature_required", False))
    out["authentication_required"] = bool(raw.get("authentication_required", False))
    out["human_interpretation_required"] = bool(raw.get("human_interpretation_required", False))
    # pass-through extras
    for k, v in raw.items():
        if k not in out:
            out[k] = v
    return out


def load_requirements(path: Path) -> list[dict[str, Any]]:
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        rows = data if isinstance(data, list) else data.get("requirements") or data.get("items") or []
    elif suffix in {".yaml", ".yml"}:
        import yaml

        data = yaml.safe_load(path.read_text(encoding="utf-8"))
        rows = data if isinstance(data, list) else data.get("requirements") or []
    elif suffix == ".csv":
        with path.open(encoding="utf-8", newline="") as fh:
            rows = list(csv.DictReader(fh))
    elif suffix == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers: list[str] = []
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c or f"col{j}") for j, c in enumerate(row)]
                continue
            rows.append({headers[j]: row[j] for j in range(len(headers))})
    else:
        raise ValueError(f"unsupported requirements format: {suffix}")

    return [_default_requirement(dict(r), i + 1) for i, r in enumerate(rows)]
