"""Generate Markdown, HTML, Excel and PDF reports from the same data model."""

from __future__ import annotations

import hashlib
import html
from pathlib import Path
from typing import Any

from scripts.edital_case import DISCLAIMER
from scripts.edital_case.store import read_json, utc_now, write_json


def _load_case(case_dir: Path) -> dict[str, Any]:
    def safe(name: str) -> Any:
        p = case_dir / name
        return read_json(p) if p.exists() else {}

    return {
        "manifest": safe("case-manifest.json"),
        "inventory": safe("inventory.json"),
        "checklist": safe("checklist.json"),
        "timeline": safe("timeline.json"),
        "missing": safe("missing-documents.json"),
        "findings": safe("findings.json"),
        "inconsistencies": safe("inconsistencies.json"),
        "requirements": safe("requirements.json"),
        "risks": safe("risk-register.json"),
        "recommendation": safe("recommendation.json"),
        "evidence": safe("evidence-matrix.json"),
    }


def build_model(case_dir: Path) -> dict[str, Any]:
    data = _load_case(case_dir)
    rec = data["recommendation"] or {}
    inv = data["inventory"] or {}
    model = {
        "case_id": (data["manifest"] or {}).get("case_id"),
        "generated_at": utc_now(),
        "recommendation": rec.get("recommendation") or "REVIEW",
        "reasons": rec.get("reasons") or [],
        "favorable": rec.get("favorable") or [],
        "impeditive": rec.get("impeditive") or [],
        "disclaimer": DISCLAIMER,
        "documents": inv.get("documents") or [],
        "document_count": inv.get("document_count") or len(inv.get("documents") or []),
        "missing": data["missing"],
        "checklist": data["checklist"],
        "timeline": data["timeline"],
        "findings": data["findings"],
        "inconsistencies": data["inconsistencies"],
        "requirements": data["requirements"],
        "risks": data["risks"],
        "evidence": data["evidence"],
        "counts": {
            "documents": inv.get("document_count") or len(inv.get("documents") or []),
            "checklist_items": (data["checklist"] or {}).get("item_count")
            or len((data["checklist"] or {}).get("items") or []),
            "findings": (data["findings"] or {}).get("count") or 0,
            "timeline_events": (data["timeline"] or {}).get("event_count") or 0,
            "missing": (data["missing"] or {}).get("missing_count") or 0,
            "risks": (data["risks"] or {}).get("count") or 0,
            "inconsistencies": (data["inconsistencies"] or {}).get("count") or 0,
        },
    }
    return model


def render_markdown(model: dict[str, Any]) -> str:
    rec = model["recommendation"]
    lines = [
        f"# Triagem Técnica de Edital — {model.get('case_id')}",
        "",
        f"- Gerado em: `{model.get('generated_at')}`",
        f"- Recomendação preliminar: **{rec}**",
        f"- Documentos: {model['counts']['documents']}",
        f"- Checklist items: {model['counts']['checklist_items']}",
        f"- Findings: {model['counts']['findings']}",
        "",
        "## Disclaimer",
        "",
        DISCLAIMER,
        "",
        "## Sumário executivo",
        "",
        f"Recomendação **{rec}** com base em análise automática rastreável.",
        "",
        "### Motivos",
        "",
    ]
    for r in model.get("reasons") or []:
        lines.append(f"- {r}")
    lines += ["", "### Fatores favoráveis", ""]
    for f in model.get("favorable") or []:
        lines.append(f"- {f}")
    lines += ["", "### Fatores impeditivos / atenção", ""]
    for f in model.get("impeditive") or []:
        lines.append(f"- {f}")

    lines += ["", "## Documentos analisados", ""]
    for d in model.get("documents") or []:
        cls = (d.get("classification") or {}).get("result")
        lines.append(
            f"- `{d.get('document_id')}` — {d.get('original_name')} — "
            f"tipo={cls} sha256=`{(d.get('sha256') or '')[:12]}…` "
            f"quality={d.get('quality_status')}"
        )

    lines += ["", "## Documentos ausentes / ambíguos", ""]
    refs = (model.get("missing") or {}).get("references") or []
    miss = [r for r in refs if r.get("status") in {"MISSING", "AMBIGUOUS"}]
    if not miss:
        lines.append("- Nenhum anexo MISSING/AMBIGUOUS detectado.")
    for r in miss[:50]:
        lines.append(
            f"- **{r.get('status')}** `{r.get('referenced_name')}` "
            f"← {(r.get('referenced_from') or {}).get('document_id')} "
            f"p.{(r.get('referenced_from') or {}).get('page')}"
        )

    lines += ["", "## Linha do tempo", ""]
    for e in (model.get("timeline") or {}).get("events") or []:
        lines.append(
            f"- **{e.get('kind')}**: raw=`{e.get('raw_value')}` "
            f"norm=`{e.get('normalized')}` doc=`{e.get('document_id')}` "
            f"p.{e.get('page')}"
        )
    for c in (model.get("timeline") or {}).get("conflicts") or []:
        lines.append(f"- ⚠ CONFLITO {c.get('kind')}: {c.get('values')}")

    lines += ["", "## Checklist", ""]
    for item in (model.get("checklist") or {}).get("items") or []:
        ev = item.get("evidence") or {}
        lines.append(
            f"### {item.get('id')} — {item.get('label')}"
        )
        lines.append(f"- status: **{item.get('status')}** (critical={item.get('critical')})")
        lines.append(
            f"- evidência: doc=`{ev.get('document_id')}` locator=`{ev.get('locator')}` "
            f"page=`{ev.get('page')}`"
        )
        if ev.get("excerpt"):
            lines.append(f"- trecho: _{ev.get('excerpt')[:240]}_")
        lines.append(f"- análise: {ev.get('analysis')}")
        lines.append("")

    lines += ["", "## Riscos", ""]
    for r in (model.get("risks") or {}).get("risks") or []:
        lines.append(f"- **{r.get('severity')}** {r.get('title')}")

    lines += ["", "## Inconsistências", ""]
    for inc in (model.get("inconsistencies") or {}).get("inconsistencies") or []:
        lines.append(f"- **{inc.get('class')}** {inc.get('field')}: {inc.get('values')}")

    lines += ["", "## Pendências humanas", ""]
    for item in (model.get("checklist") or {}).get("items") or []:
        if item.get("status") in {
            "NEEDS_HUMAN",
            "MISSING_EVIDENCE",
            "NOT_FOUND",
            "BLOCKER",
            "EXTRACTION_FAILED",
        }:
            lines.append(f"- [{item.get('status')}] {item.get('label')}")

    lines += ["", "## Evidências (matriz)", ""]
    for e in (model.get("evidence") or {}).get("entries") or []:
        if e.get("excerpt"):
            lines.append(
                f"- {e.get('finding_id')}: `{e.get('locator')}` — {str(e.get('excerpt'))[:160]}"
            )

    lines += ["", "---", "", DISCLAIMER, ""]
    return "\n".join(lines)


def render_html(model: dict[str, Any]) -> str:
    md_like = render_markdown(model)
    # simple HTML wrapper
    body = html.escape(md_like).replace("\n", "<br>\n")
    rec = html.escape(str(model.get("recommendation")))
    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8"/>
<title>Triagem Edital — {html.escape(str(model.get('case_id')))}</title>
<style>
body {{ font-family: Georgia, serif; max-width: 960px; margin: 2rem auto; padding: 0 1rem; color: #222; }}
.banner {{ background: #1a365d; color: #fff; padding: 1rem 1.25rem; border-radius: 8px; }}
.rec {{ font-size: 1.4rem; font-weight: bold; }}
.disclaimer {{ background: #fff3cd; border: 1px solid #ffc107; padding: 1rem; margin: 1rem 0; }}
pre {{ white-space: pre-wrap; font-family: inherit; }}
</style>
</head>
<body>
<div class="banner">
  <div>EDITAL TECHNICAL TRIAGE CASE PACK</div>
  <div class="rec">Recomendação: {rec}</div>
  <div>Case: {html.escape(str(model.get('case_id')))}</div>
</div>
<div class="disclaimer">{html.escape(DISCLAIMER)}</div>
<pre>{body}</pre>
</body>
</html>
"""


def render_excel(model: dict[str, Any], path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()

    def sheet(name: str):
        if name == "Metadados":
            ws = wb.active
            ws.title = name
            return ws
        return wb.create_sheet(name)

    ws = sheet("Metadados")
    meta_rows = [
        ("case_id", model.get("case_id")),
        ("generated_at", model.get("generated_at")),
        ("recommendation", model.get("recommendation")),
        ("documents", model["counts"]["documents"]),
        ("checklist_items", model["counts"]["checklist_items"]),
        ("findings", model["counts"]["findings"]),
        ("timeline_events", model["counts"]["timeline_events"]),
        ("missing", model["counts"]["missing"]),
        ("disclaimer", DISCLAIMER),
    ]
    ws.append(["campo", "valor"])
    for r in meta_rows:
        ws.append(list(r))

    ws = sheet("Documentos")
    ws.append(
        [
            "document_id",
            "original_name",
            "type",
            "sha256",
            "quality",
            "pages",
            "chars",
        ]
    )
    for d in model.get("documents") or []:
        ws.append(
            [
                d.get("document_id"),
                d.get("original_name"),
                (d.get("classification") or {}).get("result"),
                d.get("sha256"),
                d.get("quality_status"),
                d.get("page_count"),
                d.get("total_chars"),
            ]
        )

    ws = sheet("Checklist")
    ws.append(
        [
            "id",
            "label",
            "category",
            "critical",
            "status",
            "document_id",
            "locator",
            "page",
            "excerpt",
            "analysis",
            "confidence",
        ]
    )
    for item in (model.get("checklist") or {}).get("items") or []:
        ev = item.get("evidence") or {}
        ws.append(
            [
                item.get("id"),
                item.get("label"),
                item.get("category"),
                item.get("critical"),
                item.get("status"),
                ev.get("document_id"),
                ev.get("locator"),
                ev.get("page"),
                (ev.get("excerpt") or "")[:500],
                ev.get("analysis"),
                ev.get("confidence"),
            ]
        )

    ws = sheet("Prazos")
    ws.append(
        [
            "kind",
            "raw",
            "normalized",
            "document_id",
            "page",
            "locator",
            "excerpt",
            "confidence",
        ]
    )
    for e in (model.get("timeline") or {}).get("events") or []:
        ws.append(
            [
                e.get("kind"),
                e.get("raw_value"),
                e.get("normalized"),
                e.get("document_id"),
                e.get("page"),
                e.get("locator"),
                (e.get("excerpt") or "")[:400],
                e.get("confidence"),
            ]
        )

    ws = sheet("Requisitos")
    ws.append(
        [
            "id",
            "requisito",
            "categoria",
            "obrigatorio",
            "documento",
            "localizador",
            "texto_fonte",
            "status",
            "acao_humana",
        ]
    )
    for r in (model.get("requirements") or {}).get("rows") or []:
        ws.append(
            [
                r.get("id"),
                r.get("requisito"),
                r.get("categoria"),
                r.get("obrigatorio"),
                r.get("documento"),
                r.get("localizador"),
                (r.get("texto_fonte") or "")[:400],
                r.get("status"),
                r.get("acao_humana"),
            ]
        )

    ws = sheet("Riscos")
    ws.append(["risk_id", "severity", "title", "status", "excerpt"])
    for r in (model.get("risks") or {}).get("risks") or []:
        ev = r.get("evidence") or {}
        ws.append(
            [
                r.get("risk_id"),
                r.get("severity"),
                r.get("title"),
                r.get("status"),
                (ev.get("excerpt") or "")[:400],
            ]
        )

    ws = sheet("Inconsistências")
    ws.append(["field", "class", "values", "analysis"])
    for inc in (model.get("inconsistencies") or {}).get("inconsistencies") or []:
        ws.append(
            [
                inc.get("field"),
                inc.get("class"),
                str(inc.get("values"))[:500],
                inc.get("analysis"),
            ]
        )

    ws = sheet("Evidências")
    ws.append(["finding_id", "title", "document_id", "sha256", "locator", "page", "excerpt"])
    for e in (model.get("evidence") or {}).get("entries") or []:
        ws.append(
            [
                e.get("finding_id"),
                e.get("title"),
                e.get("document_id"),
                e.get("sha256"),
                e.get("locator"),
                e.get("page"),
                (e.get("excerpt") or "")[:500],
            ]
        )

    ws = sheet("Pendências")
    ws.append(["id", "label", "status", "critical", "analysis"])
    for item in (model.get("checklist") or {}).get("items") or []:
        if item.get("status") in {
            "NEEDS_HUMAN",
            "MISSING_EVIDENCE",
            "NOT_FOUND",
            "BLOCKER",
            "EXTRACTION_FAILED",
            "RISK",
        }:
            ws.append(
                [
                    item.get("id"),
                    item.get("label"),
                    item.get("status"),
                    item.get("critical"),
                    (item.get("evidence") or {}).get("analysis"),
                ]
            )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def render_pdf(model: dict[str, Any], path: Path) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    def p(text: str, style: str = "BodyText") -> None:
        safe = (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )
        story.append(Paragraph(safe, styles[style]))
        story.append(Spacer(1, 0.2 * cm))

    p(f"Triagem Técnica de Edital — {model.get('case_id')}", "Title")
    p(f"Recomendação preliminar: {model.get('recommendation')}", "Heading1")
    p(DISCLAIMER)
    p("1. Identificação", "Heading2")
    p(f"Case ID: {model.get('case_id')}")
    p(f"Gerado em: {model.get('generated_at')}")
    p(f"Documentos: {model['counts']['documents']}")
    p("2. Sumário executivo", "Heading2")
    for r in model.get("reasons") or []:
        p(f"• {r}")
    p("3. Documentos analisados", "Heading2")
    for d in model.get("documents") or []:
        p(
            f"• {d.get('original_name')} [{(d.get('classification') or {}).get('result')}] "
            f"sha={(d.get('sha256') or '')[:16]}"
        )
    p("4. Documentos ausentes", "Heading2")
    refs = (model.get("missing") or {}).get("references") or []
    miss = [r for r in refs if r.get("status") in {"MISSING", "AMBIGUOUS"}]
    if not miss:
        p("Nenhum MISSING/AMBIGUOUS.")
    for r in miss[:40]:
        p(f"• [{r.get('status')}] {r.get('referenced_name')}")
    p("5. Linha do tempo", "Heading2")
    for e in (model.get("timeline") or {}).get("events") or []:
        p(
            f"• {e.get('kind')}: {e.get('raw_value')} → {e.get('normalized')} "
            f"(doc {e.get('document_id')} p.{e.get('page')})"
        )
    p("6. Checklist (amostra)", "Heading2")
    for item in ((model.get("checklist") or {}).get("items") or [])[:40]:
        ev = item.get("evidence") or {}
        p(
            f"• [{item.get('status')}] {item.get('label')} "
            f"— {ev.get('locator')} — {(ev.get('excerpt') or '')[:120]}"
        )
    p("7. Riscos", "Heading2")
    for r in (model.get("risks") or {}).get("risks") or []:
        p(f"• [{r.get('severity')}] {r.get('title')}")
    p("8. Inconsistências", "Heading2")
    for inc in (model.get("inconsistencies") or {}).get("inconsistencies") or []:
        p(f"• [{inc.get('class')}] {inc.get('field')}")
    p("9. Disclaimer final", "Heading2")
    p(DISCLAIMER)
    doc.build(story)


def checksums_for_reports(reports_dir: Path) -> dict[str, str]:
    out = {}
    for p in sorted(reports_dir.glob("*")):
        if p.is_file():
            h = hashlib.sha256(p.read_bytes()).hexdigest()
            out[p.name] = h
    return out


def reconcile_reports(model: dict[str, Any], reports_dir: Path) -> dict[str, Any]:
    """Ensure quantitative counts are consistent across artifacts."""
    issues: list[str] = []
    md_path = reports_dir / "executive-summary.md"
    if md_path.exists():
        text = md_path.read_text(encoding="utf-8")
        # recommendation must appear
        if f"**{model['recommendation']}**" not in text and model["recommendation"] not in text:
            issues.append("markdown missing recommendation")
        if str(model["counts"]["documents"]) not in text:
            issues.append("markdown document count mismatch")
    xlsx = reports_dir / "triage-workbook.xlsx"
    if xlsx.exists():
        from openpyxl import load_workbook

        wb = load_workbook(str(xlsx), read_only=True)
        if "Checklist" in wb.sheetnames:
            ws = wb["Checklist"]
            rows = list(ws.iter_rows(values_only=True))
            # header + items
            n = max(0, len(rows) - 1)
            expected = model["counts"]["checklist_items"]
            if n != expected:
                issues.append(f"excel checklist rows {n} != model {expected}")
        wb.close()
    return {
        "ok": len(issues) == 0,
        "issues": issues,
        "counts": model["counts"],
        "checksums": checksums_for_reports(reports_dir),
        "generated_at": utc_now(),
    }


def generate_reports(case_dir: Path) -> dict[str, Any]:
    model = build_model(case_dir)
    reports = case_dir / "reports"
    reports.mkdir(parents=True, exist_ok=True)

    md = render_markdown(model)
    (reports / "executive-summary.md").write_text(md, encoding="utf-8")
    (reports / "triage-report.html").write_text(render_html(model), encoding="utf-8")
    render_excel(model, reports / "triage-workbook.xlsx")
    render_pdf(model, reports / "triage-report.pdf")

    write_json(reports / "report-model.json", model)
    recon = reconcile_reports(model, reports)
    write_json(reports / "reconciliation.json", recon)
    write_json(reports / "checksums.json", recon.get("checksums") or {})
    return {"model": model, "reconciliation": recon, "reports_dir": str(reports)}
