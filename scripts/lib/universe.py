"""Canonical target-universe resolution shared by readiness and radar flows.

The spreadsheet is the only authority for membership in the Extra target
universe. Database radius flags are diagnostic data, never a denominator.
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_SEED_PATH = Path("Extra - alvos de licitação. R-0.xlsx")
DEFAULT_RADIUS_KM = 200.0

# Backwards-compatible historical value. New code MUST derive the denominator
# from ``load_canonical_universe`` instead of this constant.
CANONICAL_UNIVERSE = 1093


def normalize_cnpj8(cnpj: str) -> str:
    """Return the first eight digits of a CNPJ-like value."""
    return "".join(char for char in str(cnpj) if char.isdigit())[:8]


def normalize_codigo_ibge(value: str | int | None) -> str:
    """Normalize a município IBGE code to exactly 7 digits.

    Accepts int or digit strings with optional punctuation/spaces.
    Returns ``""`` for empty/None/non-digit or wrong-length values after
    digit extraction (IBGE municipality codes are always 7 digits).
    """
    if value is None:
        return ""
    digits = "".join(char for char in str(value) if char.isdigit())
    if len(digits) == 7:
        return digits
    # common zero-padded form for legacy 6-digit inputs is rejected (ambiguous)
    return ""


def normalize_identity_text(value: str | None) -> str:
    """Normalize human identity text without erasing word boundaries."""
    decomposed = unicodedata.normalize("NFKD", value or "")
    ascii_text = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", re.sub(r"[^A-Za-z0-9]+", " ", ascii_text)).strip().upper()


def sha256_file(path: str | Path) -> str:
    """Hash a file without loading it entirely into memory."""
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@dataclass(frozen=True)
class CanonicalEntity:
    """One immutable seed row with an auditable radius decision."""

    entity_id: str
    seed_row: int
    razao_social: str
    cnpj8: str
    municipio: str
    codigo_ibge: str
    natureza_juridica: str
    latitude: float | None
    longitude: float | None
    distancia_km: float | None
    radius_decision: str
    within_radius: bool | None
    decision_method: str
    identity_key: str
    duplicate_root: bool = False
    suspicious_duplicate: bool = False
    db_entity_id: int | None = None
    db_match_method: str | None = None


@dataclass
class CanonicalUniverse:
    """Resolved seed snapshot used by every QW-01 metric and export."""

    seed_path: str
    seed_sha256: str
    radius_km: float
    entities: list[CanonicalEntity] = field(default_factory=list)
    duplicate_roots: list[str] = field(default_factory=list)
    suspicious_duplicate_keys: list[str] = field(default_factory=list)

    @property
    def included(self) -> list[CanonicalEntity]:
        return [entity for entity in self.entities if entity.within_radius is True]

    @property
    def excluded(self) -> list[CanonicalEntity]:
        return [entity for entity in self.entities if entity.within_radius is False]

    @property
    def unresolved(self) -> list[CanonicalEntity]:
        return [entity for entity in self.entities if entity.within_radius is None]

    @property
    def conservative_monitoring_population(self) -> list[CanonicalEntity]:
        """Included entities plus unresolved rows, as required by QW-01."""
        return [entity for entity in self.entities if entity.within_radius is not False]

    @property
    def resolution_coverage(self) -> float:
        if not self.entities:
            return 0.0
        resolved = len(self.entities) - len(self.unresolved)
        return _bounded_percent(resolved, len(self.entities))

    def by_entity_id(self) -> dict[str, CanonicalEntity]:
        return {entity.entity_id: entity for entity in self.entities}

    def included_by_cnpj8(self) -> dict[str, list[CanonicalEntity]]:
        grouped: dict[str, list[CanonicalEntity]] = defaultdict(list)
        for entity in self.included:
            grouped[entity.cnpj8].append(entity)
        return dict(grouped)

    def resolve_opportunity(
        self,
        cnpj: str | None,
        orgao_nome: str | None,
        municipio: str | None,
    ) -> tuple[CanonicalEntity | None, str]:
        """Resolve one opportunity without silently collapsing duplicate roots."""
        root = normalize_cnpj8(cnpj or "")
        candidates = self.included_by_cnpj8().get(root, [])
        if not candidates:
            return None, "cnpj_root_not_in_target_universe"
        if len(candidates) == 1:
            return candidates[0], "cnpj8_unique"

        normalized_name = normalize_identity_text(orgao_nome)
        normalized_city = normalize_identity_text(municipio)
        exact = [
            entity
            for entity in candidates
            if normalize_identity_text(entity.razao_social) == normalized_name
            and normalize_identity_text(entity.municipio) == normalized_city
        ]
        if len(exact) == 1:
            return exact[0], "cnpj8_name_municipality"

        by_name = [entity for entity in candidates if normalize_identity_text(entity.razao_social) == normalized_name]
        if len(by_name) == 1:
            return by_name[0], "cnpj8_name"

        by_city = [entity for entity in candidates if normalize_identity_text(entity.municipio) == normalized_city]
        if len(by_city) == 1:
            return by_city[0], "cnpj8_municipality"

        return None, "ambiguous_duplicate_cnpj_root"

    def summary(self) -> dict[str, Any]:
        return {
            "seed_path": self.seed_path,
            "seed_sha256": self.seed_sha256,
            "radius_km": self.radius_km,
            "total_seed_rows": len(self.entities),
            "resolved_rows": len(self.entities) - len(self.unresolved),
            "unresolved_rows": len(self.unresolved),
            "within_radius": len(self.included),
            "outside_radius": len(self.excluded),
            "conservative_monitoring_denominator": len(self.conservative_monitoring_population),
            "universe_resolution_coverage_percent": self.resolution_coverage,
            "duplicate_cnpj_roots": self.duplicate_roots,
            "suspicious_duplicate_keys": self.suspicious_duplicate_keys,
            "db_matched_rows": sum(entity.db_entity_id is not None for entity in self.entities),
            "identity_formula": "sha256(normalized_cnpj8|normalized_municipio|normalized_razao_social)",
            "radius_formula": (
                "seed column 'Raio 200km?': SIM = included, NAO = excluded; missing/unknown decision remains unresolved"
            ),
        }

    def to_snapshot(self) -> dict[str, Any]:
        return {"summary": self.summary(), "entities": [asdict(entity) for entity in self.entities]}


def load_canonical_universe(
    seed_path: str | Path = DEFAULT_SEED_PATH,
    radius_km: float = DEFAULT_RADIUS_KM,
    conn: Any | None = None,
) -> CanonicalUniverse:
    """Load and validate the authoritative seed, preserving every data row."""
    path = Path(seed_path).resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Canonical universe seed not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Entes Públicos SC" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("Seed is missing required sheet 'Entes Públicos SC'")
    worksheet = workbook["Entes Públicos SC"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    workbook.close()

    parsed_rows: list[dict[str, Any]] = []
    for seed_row, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        parsed_rows.append(_parse_seed_row(seed_row, row, radius_km))

    root_counts = Counter(row["cnpj8"] for row in parsed_rows)
    identity_counts = Counter(row["identity_key"] for row in parsed_rows)
    identity_occurrences: Counter[str] = Counter()
    db_entities = _load_db_entities(conn) if conn is not None else {}

    entities: list[CanonicalEntity] = []
    suspicious_keys: set[str] = set()
    for parsed in parsed_rows:
        identity_key = parsed["identity_key"]
        identity_occurrences[identity_key] += 1
        suspicious = identity_counts[identity_key] > 1
        if suspicious:
            suspicious_keys.add(identity_key)
        stable_input = identity_key
        if suspicious:
            stable_input = f"{identity_key}|occurrence={identity_occurrences[identity_key]}"
        entity_id = f"extra-{hashlib.sha256(stable_input.encode('utf-8')).hexdigest()[:20]}"

        db_match = db_entities.get(parsed["cnpj8"])
        entities.append(
            CanonicalEntity(
                entity_id=entity_id,
                seed_row=parsed["seed_row"],
                razao_social=parsed["razao_social"],
                cnpj8=parsed["cnpj8"],
                municipio=parsed["municipio"],
                codigo_ibge=parsed["codigo_ibge"],
                natureza_juridica=parsed["natureza_juridica"],
                latitude=parsed["latitude"],
                longitude=parsed["longitude"],
                distancia_km=parsed["distancia_km"],
                radius_decision=parsed["radius_decision"],
                within_radius=parsed["within_radius"],
                decision_method=parsed["decision_method"],
                identity_key=identity_key,
                duplicate_root=root_counts[parsed["cnpj8"]] > 1,
                suspicious_duplicate=suspicious,
                db_entity_id=db_match["id"] if db_match else None,
                db_match_method="cnpj8" if db_match else None,
            )
        )

    universe = CanonicalUniverse(
        seed_path=str(path),
        seed_sha256=sha256_file(path),
        radius_km=radius_km,
        entities=entities,
        duplicate_roots=sorted(root for root, count in root_counts.items() if count > 1),
        suspicious_duplicate_keys=sorted(suspicious_keys),
    )
    _validate_universe(universe)
    return universe


def get_canonical_universe(conn: Any | None = None, seed_path: str | Path = DEFAULT_SEED_PATH) -> int:
    """Return the runtime seed-derived denominator (legacy API name)."""
    return len(load_canonical_universe(seed_path=seed_path, conn=conn).included)


def _parse_seed_row(seed_row: int, row: tuple[Any, ...], radius_km: float) -> dict[str, Any]:
    values = list(row) + [None] * max(0, 10 - len(row))
    cnpj8 = normalize_cnpj8(str(values[1] or ""))
    if len(cnpj8) != 8:
        raise ValueError(f"Seed row {seed_row} has invalid CNPJ root: {values[1]!r}")
    razao_social = str(values[0]).strip()
    municipio = str(values[2] or "").strip()
    identity_key = "|".join((cnpj8, normalize_identity_text(municipio), normalize_identity_text(razao_social)))

    latitude = _optional_float(values[6])
    longitude = _optional_float(values[7])
    distancia_km = _optional_float(values[8])
    radius_flag = normalize_identity_text(str(values[9] or ""))
    if radius_flag.startswith("SIM"):
        within_radius: bool | None = True
        radius_decision = "included"
        decision_method = "seed_radius_flag"
    elif radius_flag.startswith("NAO"):
        within_radius = False
        radius_decision = "excluded"
        decision_method = "seed_radius_flag"
    elif distancia_km is not None:
        within_radius = distancia_km <= radius_km
        radius_decision = "included" if within_radius else "excluded"
        decision_method = "seed_distance_fallback"
    else:
        within_radius = None
        radius_decision = "unresolved"
        decision_method = "insufficient_seed_evidence"

    return {
        "seed_row": seed_row,
        "razao_social": razao_social,
        "cnpj8": cnpj8,
        "municipio": municipio,
        "codigo_ibge": str(values[3] or "").split(".")[0],
        "natureza_juridica": str(values[4] or "").strip(),
        "latitude": latitude,
        "longitude": longitude,
        "distancia_km": distancia_km,
        "radius_decision": radius_decision,
        "within_radius": within_radius,
        "decision_method": decision_method,
        "identity_key": identity_key,
    }


def _load_db_entities(conn: Any) -> dict[str, dict[str, Any]]:
    with conn.cursor() as cursor:
        cursor.execute(
            """
            SELECT id, cnpj_8, razao_social, municipio
            FROM sc_public_entities
            WHERE is_active IS TRUE
            """
        )
        return {
            normalize_cnpj8(str(cnpj8)): {
                "id": int(entity_id),
                "razao_social": razao_social,
                "municipio": municipio,
            }
            for entity_id, cnpj8, razao_social, municipio in cursor.fetchall()
        }


def _validate_universe(universe: CanonicalUniverse) -> None:
    if not universe.entities:
        raise ValueError("Canonical universe seed contains no data rows")
    ids = [entity.entity_id for entity in universe.entities]
    if len(ids) != len(set(ids)):
        raise ValueError("Canonical entity IDs are not unique")
    decisions = len(universe.included) + len(universe.excluded) + len(universe.unresolved)
    if decisions != len(universe.entities):
        raise ValueError("Radius decisions do not cover every seed row")
    if not 0.0 <= universe.resolution_coverage <= 100.0:
        raise ValueError("Universe resolution coverage is outside [0, 100]")


def _optional_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _bounded_percent(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round(min(100.0, max(0.0, numerator / denominator * 100.0)), 2)
