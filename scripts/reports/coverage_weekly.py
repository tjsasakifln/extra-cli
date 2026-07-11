#!/usr/bin/env python3
"""
Weekly Coverage Report — Extra Consultoria.

Generates:
  - Executive PDF (1-2 pages, Big Four aesthetic via ReportLab)
  - Detailed Excel (4 sheets via openpyxl)

Usage:
    python -m scripts.reports.coverage_weekly
    python -m scripts.reports.coverage_weekly --date 2026-07-10
    python -m scripts.reports.coverage_weekly --output-dir /custom/path
    python -m scripts.reports.coverage_weekly --snapshot-only   # only gen snapshot, skip PDF/Excel
    python -m scripts.reports.coverage_weekly --skip-snapshot    # skip snapshot gen

Design: Big Four / Management Consulting aesthetic (McKinsey, BCG, Deloitte).
Typography: Serif headings (Times) + sans-serif data (Helvetica).
Palette: Monochromatic (charcoal navy + bronze accent + neutral grays).
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Path setup — ensure scripts/ is on sys.path
# ---------------------------------------------------------------------------
_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

DSN = os.getenv(
    "LOCAL_DATALAKE_DSN",
    "postgresql://postgres:smartlic_local@127.0.0.1:54399/postgres",
)

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logger = logging.getLogger("coverage-weekly")
_handler = logging.StreamHandler()
_handler.setFormatter(logging.Formatter(
    "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
))
logger.addHandler(_handler)
logger.setLevel(logging.INFO)


# ---------------------------------------------------------------------------
# DB layer
# ---------------------------------------------------------------------------
def get_conn():
    import psycopg2
    return psycopg2.connect(DSN)


def query(sql: str, params: list | None = None) -> list[dict]:
    """Run query, return list of dicts."""
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        cols = [d[0] for d in cur.description] if cur.description else []
        rows = [dict(zip(cols, row)) for row in cur.fetchall()]
        cur.close()
        return rows
    finally:
        conn.close()


def fetch_coverage_data(report_date: date) -> dict[str, Any]:
    """Fetch all data needed for the report.

    Returns dict with keys:
      - summary: list of per-source coverage
      - gaps: list of uncovered entities
      - gaps_by_municipio: agg by city
      - gaps_by_natureza: agg by natureza juridica
      - trend: list of weekly snapshots
      - total_entities: total active entities
      - total_covered: entities covered by at least one source
      - total_uncovered: entities with zero coverage
    """
    data: dict[str, Any] = {}

    # --- Total entities ---
    rows = query(
        "SELECT COUNT(*) AS total FROM sc_public_entities WHERE is_active = TRUE"
    )
    data["total_entities"] = rows[0]["total"] if rows else 0

    # --- Total covered (at least one source) ---
    rows = query("""
        SELECT COUNT(DISTINCT entity_id) AS covered
        FROM entity_coverage ec
        WHERE ec.is_covered = TRUE
          AND EXISTS (SELECT 1 FROM sc_public_entities e WHERE e.id = ec.entity_id AND e.is_active = TRUE)
    """)
    data["total_covered"] = rows[0]["covered"] if rows else 0
    data["total_uncovered"] = data["total_entities"] - data["total_covered"]

    # --- Coverage by source ---
    data["summary"] = query("""
        SELECT
            ec.source,
            COUNT(*) AS total_tracked,
            SUM(CASE WHEN ec.is_covered THEN 1 ELSE 0 END) AS covered,
            ROUND(100.0 * SUM(CASE WHEN ec.is_covered THEN 1 ELSE 0 END) / NULLIF(COUNT(*), 0), 1) AS pct
        FROM entity_coverage ec
        WHERE EXISTS (
            SELECT 1 FROM sc_public_entities e WHERE e.id = ec.entity_id AND e.is_active = TRUE
        )
        GROUP BY ec.source
        ORDER BY pct DESC
    """)

    # --- Gaps (uncovered entities) ---
    data["gaps"] = query("""
        SELECT e.razao_social, e.municipio, e.natureza_juridica, e.cnpj_8, e.uf
        FROM sc_public_entities e
        WHERE e.is_active = TRUE
          AND NOT EXISTS (
              SELECT 1 FROM entity_coverage ec
              WHERE ec.entity_id = e.id AND ec.is_covered = TRUE
          )
        ORDER BY e.municipio, e.razao_social
    """)

    # --- Gaps by municipio ---
    data["gaps_by_municipio"] = query("""
        SELECT
            e.municipio,
            COUNT(*) AS total_entes,
            COUNT(*) FILTER (WHERE NOT EXISTS (
                SELECT 1 FROM entity_coverage ec
                WHERE ec.entity_id = e.id AND ec.is_covered = TRUE
            )) AS entes_descobertos,
            ROUND(100.0 * COUNT(*) FILTER (WHERE NOT EXISTS (
                SELECT 1 FROM entity_coverage ec
                WHERE ec.entity_id = e.id AND ec.is_covered = TRUE
            )) / NULLIF(COUNT(*), 0), 1) AS pct_gap
        FROM sc_public_entities e
        WHERE e.is_active = TRUE
        GROUP BY e.municipio
        ORDER BY entes_descobertos DESC
    """)

    # --- Gaps by natureza juridica ---
    data["gaps_by_natureza"] = query("""
        SELECT
            e.natureza_juridica,
            COUNT(*) AS total_entes,
            COUNT(*) FILTER (WHERE NOT EXISTS (
                SELECT 1 FROM entity_coverage ec
                WHERE ec.entity_id = e.id AND ec.is_covered = TRUE
            )) AS entes_descobertos,
            ROUND(100.0 * COUNT(*) FILTER (WHERE NOT EXISTS (
                SELECT 1 FROM entity_coverage ec
                WHERE ec.entity_id = e.id AND ec.is_covered = TRUE
            )) / NULLIF(COUNT(*), 0), 1) AS pct_gap
        FROM sc_public_entities e
        WHERE e.is_active = TRUE
        GROUP BY e.natureza_juridica
        ORDER BY entes_descobertos DESC
    """)

    # --- Trend (last 4 weeks) ---
    data["trend"] = query("""
        SELECT snapshot_date, source, total_entities, covered_entities, pct_covered
        FROM coverage_snapshots
        WHERE snapshot_date >= %s::DATE - INTERVAL '35 days'
        ORDER BY snapshot_date ASC, source
    """, [report_date.isoformat()])

    return data


def generate_snapshot(snap_date: date) -> int:
    """Generate a coverage snapshot using the DB function."""
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT generate_coverage_snapshot(%s)", [snap_date])
        result = cur.fetchone()
        conn.commit()
        n = result[0] if result else 0
        cur.close()
        return n
    except Exception as e:
        logger.warning("Could not generate snapshot (function may not exist yet): %s", e)
        return 0
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# PDF generation — Big Four aesthetic
# ---------------------------------------------------------------------------
def _build_pdf_styles():
    """Build ReportLab styles matching the Big Four aesthetic."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm

    INK = colors.HexColor("#1B2A3D")
    ACCENT = colors.HexColor("#8B7355")
    SIGNAL_RED = colors.HexColor("#B5342A")
    SIGNAL_GREEN = colors.HexColor("#1B7A3D")
    SIGNAL_AMBER = colors.HexColor("#B8860B")
    TEXT_COLOR = colors.HexColor("#2D3748")
    TEXT_SECONDARY = colors.HexColor("#5A6577")
    TEXT_MUTED = colors.HexColor("#8896A6")

    base = getSampleStyleSheet()
    s = {}

    s["cover_title"] = ParagraphStyle(
        "cov_title", parent=base["Normal"],
        fontName="Times-Bold", fontSize=26, textColor=INK,
        alignment=TA_LEFT, leading=32, spaceAfter=4 * mm,
    )
    s["cover_subtitle"] = ParagraphStyle(
        "cov_sub", parent=base["Normal"],
        fontName="Times-Roman", fontSize=14, textColor=TEXT_SECONDARY,
        alignment=TA_LEFT, spaceAfter=3 * mm, leading=18,
    )
    s["cover_info"] = ParagraphStyle(
        "cov_info", parent=base["Normal"],
        fontName="Helvetica", fontSize=9, textColor=TEXT_SECONDARY,
        alignment=TA_LEFT, leading=13, spaceAfter=1.5 * mm,
    )
    s["h1"] = ParagraphStyle(
        "h1_r", parent=base["Normal"],
        fontName="Times-Bold", fontSize=14, textColor=INK,
        spaceBefore=6 * mm, spaceAfter=3 * mm, leading=18,
    )
    s["h2"] = ParagraphStyle(
        "h2_r", parent=base["Normal"],
        fontName="Times-Bold", fontSize=11, textColor=INK,
        spaceBefore=4 * mm, spaceAfter=2 * mm, leading=14,
    )
    s["h3"] = ParagraphStyle(
        "h3_r", parent=base["Normal"],
        fontName="Times-Bold", fontSize=10, textColor=TEXT_COLOR,
        spaceBefore=3 * mm, spaceAfter=2 * mm, leading=13,
    )
    s["body"] = ParagraphStyle(
        "body_r", parent=base["Normal"],
        fontName="Times-Roman", fontSize=10, textColor=TEXT_COLOR,
        alignment=TA_JUSTIFY, leading=14, spaceAfter=2 * mm,
    )
    s["body_small"] = ParagraphStyle(
        "body_s", parent=base["Normal"],
        fontName="Times-Roman", fontSize=9, textColor=TEXT_SECONDARY,
        leading=12, spaceAfter=1.5 * mm,
    )
    s["bullet"] = ParagraphStyle(
        "bullet_r", parent=base["Normal"],
        fontName="Times-Roman", fontSize=10, textColor=TEXT_COLOR,
        leading=14, leftIndent=10, spaceAfter=1.5 * mm,
    )
    s["metric_value"] = ParagraphStyle(
        "mv_r", parent=base["Normal"],
        fontName="Times-Bold", fontSize=18, textColor=INK,
        alignment=TA_CENTER, leading=22,
    )
    s["metric_label"] = ParagraphStyle(
        "ml_r", parent=base["Normal"],
        fontName="Helvetica", fontSize=7, textColor=TEXT_MUTED,
        alignment=TA_CENTER, leading=9,
    )
    s["cell"] = ParagraphStyle(
        "cell_r", parent=base["Normal"],
        fontName="Helvetica", fontSize=8, textColor=TEXT_COLOR,
        leading=10, alignment=TA_LEFT,
    )
    s["cell_center"] = ParagraphStyle(
        "cell_c", parent=base["Normal"],
        fontName="Helvetica", fontSize=8, textColor=TEXT_COLOR,
        leading=10, alignment=TA_CENTER,
    )
    s["cell_right"] = ParagraphStyle(
        "cell_rr", parent=base["Normal"],
        fontName="Helvetica", fontSize=8, textColor=TEXT_COLOR,
        leading=10, alignment=TA_RIGHT,
    )
    s["cell_header"] = ParagraphStyle(
        "ch_r", parent=base["Normal"],
        fontName="Helvetica-Bold", fontSize=8, textColor=INK,
        leading=10, alignment=TA_LEFT,
    )
    s["cell_header_center"] = ParagraphStyle(
        "chc_r", parent=base["Normal"],
        fontName="Helvetica-Bold", fontSize=8, textColor=INK,
        leading=10, alignment=TA_CENTER,
    )
    s["cell_header_right"] = ParagraphStyle(
        "chr_r", parent=base["Normal"],
        fontName="Helvetica-Bold", fontSize=8, textColor=INK,
        leading=10, alignment=TA_RIGHT,
    )
    s["caption"] = ParagraphStyle(
        "cap_r", parent=base["Normal"],
        fontName="Helvetica", fontSize=7, textColor=TEXT_MUTED,
        leading=9,
    )
    s["recommendation"] = ParagraphStyle(
        "rec_r", parent=base["Normal"],
        fontName="Times-Roman", fontSize=10, textColor=TEXT_COLOR,
        leading=14, spaceAfter=2 * mm, leftIndent=5,
    )
    s["positive"] = ParagraphStyle(
        "pos_r", parent=base["Normal"],
        fontName="Times-Roman", fontSize=10, textColor=SIGNAL_GREEN,
        leading=14,
    )
    s["negative"] = ParagraphStyle(
        "neg_r", parent=base["Normal"],
        fontName="Times-Roman", fontSize=10, textColor=SIGNAL_RED,
        leading=14,
    )
    s["amber"] = ParagraphStyle(
        "amb_r", parent=base["Normal"],
        fontName="Times-Roman", fontSize=10, textColor=SIGNAL_AMBER,
        leading=14,
    )

    # Store colors for later use
    s["_colors"] = {
        "INK": INK,
        "ACCENT": ACCENT,
        "SIGNAL_RED": SIGNAL_RED,
        "SIGNAL_GREEN": SIGNAL_GREEN,
        "SIGNAL_AMBER": SIGNAL_AMBER,
        "TEXT_COLOR": TEXT_COLOR,
        "TEXT_SECONDARY": TEXT_SECONDARY,
        "TEXT_MUTED": TEXT_MUTED,
    }
    s["_pagesize"] = A4
    return s


def _metric_cell(value: str, label: str, styles: dict) -> object:
    """Create a KPI metric card."""
    from reportlab.platypus import Paragraph, Table, TableStyle

    inner = Table(
        [[Paragraph(value, styles["metric_value"])],
         [Paragraph(label, styles["metric_label"])]],
        colWidths=["*"],
    )
    inner.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    return inner


def _section_heading(title: str, styles: dict) -> list:
    """Create a section heading: thin bronze rule + serif title."""
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

    C = styles["_colors"]
    avail = 210 * mm - 2 * 2.2 * cm  # A4 width minus margins
    rule_t = Table([[""]], colWidths=[avail], rowHeights=[1])
    rule_t.setStyle(TableStyle([
        ("LINEBELOW", (0, 0), (0, 0), 0.6, C["ACCENT"]),
        ("TOPPADDING", (0, 0), (0, 0), 0),
        ("BOTTOMPADDING", (0, 0), (0, 0), 0),
    ]))
    return [rule_t, Spacer(1, 2 * mm), Paragraph(title, styles["h1"])]


def _three_rule_table(rows: list, col_widths: list, styles: dict,
                      repeat_rows: int = 1) -> object:
    """Table with Big Four 'three-rule' styling."""
    from reportlab.platypus import Table, TableStyle

    C = styles["_colors"]
    t = Table(rows, colWidths=col_widths, repeatRows=repeat_rows)
    n = len(rows)
    style_cmds = [
        ("LINEABOVE", (0, 0), (-1, 0), 1.2, C["INK"]),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, C["INK"]),
        ("LINEBELOW", (0, n - 1), (-1, n - 1), 0.8, C["TEXT_MUTED"]),
        *[("LINEBELOW", (0, i), (-1, i), 0.3, C["TEXT_MUTED"]) for i in range(1, n - 1)],
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]
    t.setStyle(TableStyle(style_cmds))
    return t


def _build_cover_page(styles: dict, report_date: date, week_number: int) -> list:
    """Build cover page."""
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import PageBreak, Paragraph, Spacer, Table, TableStyle

    C = styles["_colors"]
    el = []

    # Generous top whitespace
    el.append(Spacer(1, 60 * mm))

    # Bronze rule
    avail = 210 * mm - 2 * 2.2 * cm
    rule_t = Table([["", ""]], colWidths=[40 * mm, avail - 40 * mm])
    rule_t.setStyle(TableStyle([("LINEBELOW", (0, 0), (0, 0), 0.8, C["ACCENT"])]))
    el.append(rule_t)
    el.append(Spacer(1, 6 * mm))

    # Title
    el.append(Paragraph(
        "Relatório de<br/>Cobertura Semanal",
        styles["cover_title"],
    ))
    el.append(Paragraph("Extra Construtora — SC", styles["cover_subtitle"]))
    el.append(Spacer(1, 12 * mm))

    # Metadata
    from datetime import datetime
    week_start = report_date - timedelta(days=report_date.weekday())
    week_end = week_start + timedelta(days=6)
    el.append(Paragraph(
        f"Semana {week_number}, {report_date.year}",
        styles["cover_info"],
    ))
    el.append(Paragraph(
        f"{week_start.strftime('%d/%b')} — {week_end.strftime('%d/%b/%Y')}",
        styles["cover_info"],
    ))
    el.append(Spacer(1, 4 * mm))
    el.append(Paragraph(
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles["cover_info"],
    ))
    el.append(Spacer(1, 30 * mm))

    # Attribution
    el.append(Paragraph(
        "<b>Tiago Sasaki</b><br/>"
        "Consultor de Inteligência em Licitações<br/>"
        "(48) 9 8834-4559",
        styles["cover_info"],
    ))

    el.append(PageBreak())
    return el


def _build_kpi_section(data: dict, styles: dict) -> list:
    """Build KPI cards section."""
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import Spacer, Table, TableStyle

    C = styles["_colors"]
    el = []
    el.extend(_section_heading("Indicadores de Cobertura", styles))

    total = data["total_entities"]
    covered = data["total_covered"]
    pct = round(covered / total * 100, 1) if total > 0 else 0

    # Variation vs previous week (from trend)
    variacao_text = "N/D"
    variacao_color = C["TEXT_MUTED"]
    trend = data.get("trend", [])
    if trend:
        # Get last two distinct snapshots by date
        dates = sorted(set(t["snapshot_date"] for t in trend if t["snapshot_date"]), reverse=True)
        if len(dates) >= 2:
            # Compare overall coverage between latest and previous
            def _total_pct_for_date(d):
                entries = [t for t in trend if t["snapshot_date"] == d]
                if entries:
                    total_cov = sum(e["covered_entities"] for e in entries)
                    total_ent = sum(e["total_entities"] for e in entries)
                    return round(total_cov / total_ent * 100, 1) if total_ent > 0 else 0
                return None
            latest_pct = _total_pct_for_date(dates[0])
            prev_pct = _total_pct_for_date(dates[1])
            if latest_pct is not None and prev_pct is not None:
                diff = latest_pct - prev_pct
                if diff > 0:
                    variacao_text = f"+{diff:.1f}pp ▲"
                    variacao_color = C["SIGNAL_GREEN"]
                elif diff < 0:
                    variacao_text = f"{diff:.1f}pp ▼"
                    variacao_color = C["SIGNAL_RED"]
                else:
                    variacao_text = "0.0pp —"
                    variacao_color = C["TEXT_MUTED"]

    avail = 210 * mm - 2 * 2.2 * cm
    col_w = avail / 3

    metrics = Table(
        [[
            _metric_cell(f"{pct}%", "Cobertura Total", styles),
            _metric_cell(f"{covered}/{total}", "Entes Cobertos", styles),
            _metric_cell(variacao_text, "Variação vs Semana Ant.", styles),
        ]],
        colWidths=[col_w] * 3, rowHeights=[20 * mm],
    )
    # Override variation color
    if variacao_color != C["TEXT_MUTED"]:
        metrics.setStyle(TableStyle([
            ("LINEABOVE", (0, 0), (-1, 0), 0.6, C["INK"]),
            ("LINEBELOW", (0, 0), (-1, 0), 0.4, C["TEXT_MUTED"]),
            ("LINEBEFORE", (1, 0), (1, 0), 0.3, C["TEXT_MUTED"]),
            ("LINEBEFORE", (2, 0), (2, 0), 0.3, C["TEXT_MUTED"]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
    else:
        metrics.setStyle(TableStyle([
            ("LINEABOVE", (0, 0), (-1, 0), 0.6, C["INK"]),
            ("LINEBELOW", (0, 0), (-1, 0), 0.4, C["TEXT_MUTED"]),
            ("LINEBEFORE", (1, 0), (1, 0), 0.3, C["TEXT_MUTED"]),
            ("LINEBEFORE", (2, 0), (2, 0), 0.3, C["TEXT_MUTED"]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

    el.append(metrics)
    el.append(Spacer(1, 4 * mm))
    return el


def _build_source_table(data: dict, styles: dict) -> list:
    """Build coverage by source table."""
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import Paragraph, Spacer

    el = []
    el.extend(_section_heading("Cobertura por Fonte", styles))

    summary = data.get("summary", [])
    if not summary:
        el.append(Paragraph("Dados insuficientes para o período.", styles["body"]))
        return el

    # Source name mapping
    SOURCE_NAMES = {
        "pncp": "PNCP",
        "dom_sc": "DOM-SC",
        "pcp": "PCP",
        "compras_gov": "ComprasGov",
        "sc_compras": "SC Compras",
        "tce_sc": "TCE-SC",
        "transparencia": "Transparência",
    }

    header = [
        Paragraph("Fonte", styles["cell_header"]),
        Paragraph("Cobertos", styles["cell_header_center"]),
        Paragraph("Total", styles["cell_header_center"]),
        Paragraph("%", styles["cell_header_center"]),
    ]
    rows = [header]
    for s in summary:
        src_name = SOURCE_NAMES.get(s["source"], s["source"])
        rows.append([
            Paragraph(src_name, styles["cell"]),
            Paragraph(str(s["covered"]), styles["cell_center"]),
            Paragraph(str(s["total_tracked"]), styles["cell_center"]),
            Paragraph(f'{s["pct"]}%', styles["cell_center"]),
        ])

    avail = 210 * mm - 2 * 2.2 * cm
    col_w = [avail * 0.35, avail * 0.20, avail * 0.20, avail * 0.25]
    t = _three_rule_table(rows, col_w, styles)
    el.append(t)
    el.append(Spacer(1, 4 * mm))
    return el


def _build_top_gaps(data: dict, styles: dict) -> list:
    """Build top 10 municipalities with most gaps."""
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import Paragraph, Spacer

    el = []
    el.extend(_section_heading("Municípios com Mais Gaps", styles))

    gaps_by_muni = data.get("gaps_by_municipio", [])
    # Filter to only those with gaps
    gaps_with_issues = [g for g in gaps_by_muni if g.get("entes_descobertos", 0) > 0]
    if not gaps_with_issues:
        el.append(Paragraph(
            "Todos os entes cobertos!", styles["positive"],
        ))
        return el

    top10 = gaps_with_issues[:10]

    header = [
        Paragraph("#", styles["cell_header_center"]),
        Paragraph("Município", styles["cell_header"]),
        Paragraph("Descobertos", styles["cell_header_center"]),
        Paragraph("Total", styles["cell_header_center"]),
        Paragraph("% Gap", styles["cell_header_center"]),
    ]
    rows = [header]
    for i, g in enumerate(top10, 1):
        rows.append([
            Paragraph(str(i), styles["cell_center"]),
            Paragraph(g.get("municipio", "N/D")[:30], styles["cell"]),
            Paragraph(str(g["entes_descobertos"]), styles["cell_center"]),
            Paragraph(str(g["total_entes"]), styles["cell_center"]),
            Paragraph(f'{g["pct_gap"]}%', styles["cell_center"]),
        ])

    avail = 210 * mm - 2 * 2.2 * cm
    col_w = [avail * 0.08, avail * 0.37, avail * 0.20, avail * 0.15, avail * 0.20]
    t = _three_rule_table(rows, col_w, styles)
    el.append(t)
    el.append(Spacer(1, 4 * mm))
    return el


def _build_trend_section(data: dict, styles: dict) -> list:
    """Build 4-week trend section."""
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import Paragraph, Spacer

    el = []
    el.extend(_section_heading("Tendência 4 Semanas", styles))

    trend = data.get("trend", [])
    if not trend:
        el.append(Paragraph(
            "Dados históricos insuficientes. O snapshot semanal será gerado automaticamente "
            "pelo timer do sistema.", styles["body_small"],
        ))
        return el

    # Group by date (calculate overall coverage per date)
    from collections import defaultdict
    by_date: dict[str, dict] = defaultdict(lambda: {"total_entities": 0, "covered_entities": 0})
    for t in trend:
        d = str(t["snapshot_date"])
        by_date[d]["total_entities"] += t["total_entities"]
        by_date[d]["covered_entities"] += t["covered_entities"]

    sorted_dates = sorted(by_date.keys())
    # Take last 4
    last4 = sorted_dates[-4:] if len(sorted_dates) >= 4 else sorted_dates

    if len(last4) < 2:
        el.append(Paragraph(
            "Menos de 2 snapshots disponíveis. A tendência será exibida quando houver "
            "mais dados históricos.", styles["body_small"],
        ))
        return el

    header = [
        Paragraph("Semana", styles["cell_header_center"]),
        Paragraph("Cobertura Total", styles["cell_header_center"]),
        Paragraph("Entes Cobertos", styles["cell_header_center"]),
        Paragraph("Variação", styles["cell_header_center"]),
    ]
    rows = [header]
    prev_pct = None
    for d in last4:
        dt = datetime.strptime(str(d)[:10], "%Y-%m-%d") if isinstance(d, str) else d
        iso_cal = dt.isocalendar() if hasattr(dt, "isocalendar") else None
        week_label = f"S{iso_cal[1]}" if iso_cal else str(d)[:10]

        entry = by_date[d]
        pct = round(entry["covered_entities"] / entry["total_entities"] * 100, 1) if entry["total_entities"] > 0 else 0

        if prev_pct is not None:
            diff = pct - prev_pct
            if diff > 0:
                var_text = f"+{diff:.1f}pp ▲"
                var_style = styles["positive"]
            elif diff < 0:
                var_text = f"{diff:.1f}pp ▼"
                var_style = styles["negative"]
            else:
                var_text = "0.0pp —"
                var_style = styles["body_small"]
        else:
            var_text = "—"
            var_style = styles["body_small"]

        rows.append([
            Paragraph(week_label, styles["cell_center"]),
            Paragraph(f"{pct}%", styles["cell_center"]),
            Paragraph(f"{entry['covered_entities']}/{entry['total_entities']}", styles["cell_center"]),
            Paragraph(var_text, var_style),
        ])
        prev_pct = pct

    avail = 210 * mm - 2 * 2.2 * cm
    col_w = [avail * 0.20, avail * 0.25, avail * 0.30, avail * 0.25]
    t = _three_rule_table(rows, col_w, styles)
    el.append(t)
    el.append(Spacer(1, 4 * mm))
    return el


def _build_recommendations(data: dict, styles: dict) -> list:
    """Build automatic recommendations based on data."""
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer

    el = []
    el.extend(_section_heading("Recomendações", styles))

    summary = data.get("summary", [])
    total_uncovered = data.get("total_uncovered", 0)
    gaps_by_muni = data.get("gaps_by_municipio", [])

    recommendations = []
    total_entities = data.get("total_entities", 0)
    total_covered = data.get("total_covered", 0)
    pct = round(total_covered / total_entities * 100, 1) if total_entities > 0 else 0

    # 1. Overall coverage assessment
    if pct >= 90:
        recommendations.append(
            ("✅ Cobertura elevada", f"A cobertura total de {pct}% está em nível satisfatório. "
             f"Mantenha o monitoramento contínuo para os {total_uncovered} entes restantes.")
        )
    elif pct >= 70:
        recommendations.append(
            ("⚠️ Cobertura moderada", f"A cobertura de {pct}% indica que {total_uncovered} entes "
             f"ainda estão descobertos. Foco em expandir fontes de dados para municípios com maior "
             f"concentração de gaps.")
        )
    else:
        recommendations.append(
            ("🔴 Cobertura crítica", f"Com apenas {pct}% de cobertura, é prioritário expandir "
             f"as fontes de dados. {total_uncovered} entes estão completamente descobertos.")
        )

    # 2. Worst source
    if summary:
        sorted_src = sorted(summary, key=lambda x: x["pct"])
        worst = sorted_src[0]
        if worst["pct"] < 50:
            recommendations.append(
                (f"🔧 Fonte crítica: {worst['source']}",
                 f"Apenas {worst['pct']}% de cobertura nesta fonte. Verificar crawler "
                 f"e conectividade com a API.")
            )

    # 3. Municipal concentration
    if gaps_by_muni:
        gaps_with_issues = [g for g in gaps_by_muni if g.get("entes_descobertos", 0) > 0]
        if gaps_with_issues:
            top_muni = gaps_with_issues[0]
            total_gap_entes = sum(g["entes_descobertos"] for g in gaps_with_issues)
            recommendations.append(
                ("📍 Foco municipal", f"{top_muni['municipio']} lidera com "
                 f"{top_muni['entes_descobertos']} entes descobertos. {total_gap_entes} entes "
                 f"descobertos no total, distribuídos em {len(gaps_with_issues)} municípios.")
            )

    # 4. Specific actions
    if pct < 100:
        recommendations.append(
            ("📋 Ações recomendadas",
             "• Verificar conectividade dos crawlers com todas as fontes\n"
             "• Revisar matching de entidades com gap total\n"
             "• Priorizar ativação de novas fontes para municípios com gaps\n"
             "• Validar cobertura de TCE-SC e Transparência como fontes complementares")
        )

    for title, desc in recommendations:
        el.append(Paragraph(f"<b>{title}</b>", styles["h3"]))
        el.append(Paragraph(desc, styles["recommendation"]))
        el.append(Spacer(1, 2 * mm))

    return el


def generate_pdf(data: dict, output_path: str, report_date: date) -> str:
    """Generate executive PDF report."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate

    styles = _build_pdf_styles()
    C = styles["_colors"]
    MARGIN = 2.2 * cm
    PAGE_WIDTH, PAGE_HEIGHT = A4

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=MARGIN,
        bottomMargin=MARGIN + 15 * mm,
        title="Relatório de Cobertura Semanal",
        author="Extra Construtora — Tiago Sasaki",
    )

    # Week number
    week_number = report_date.isocalendar()[1]

    # Build story
    story = []
    story.extend(_build_cover_page(styles, report_date, week_number))
    story.extend(_build_kpi_section(data, styles))
    story.extend(_build_source_table(data, styles))
    story.extend(_build_top_gaps(data, styles))
    story.extend(_build_trend_section(data, styles))
    story.extend(_build_recommendations(data, styles))

    # Draw footer on each page
    FOOTER_LINE1 = "Tiago Sasaki — Consultor de Inteligência em Licitações"
    FOOTER_LINE2 = "Relatório confidencial preparado exclusivamente para o destinatário"

    def _draw_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        y = MARGIN - 10 * mm
        canvas_obj.setStrokeColor(C["TEXT_MUTED"])
        canvas_obj.setLineWidth(0.4)
        canvas_obj.line(MARGIN, y + 4 * mm, PAGE_WIDTH - MARGIN, y + 4 * mm)
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.setFillColor(C["TEXT_MUTED"])
        canvas_obj.drawCentredString(PAGE_WIDTH / 2, y + 1.5 * mm, FOOTER_LINE1)
        canvas_obj.drawCentredString(PAGE_WIDTH / 2, y - 2 * mm, FOOTER_LINE2)
        canvas_obj.restoreState()

    doc.build(story, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    return output_path


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def _excel_header_style(wb):
    """Return common header style for Excel."""
    from openpyxl.styles import Font, PatternFill, Alignment
    return {
        "font": Font(bold=True, color="FFFFFF", size=10),
        "fill": PatternFill(start_color="1B2A3D", end_color="1B2A3D", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def _excel_write_sheet(ws, headers, keys, data_rows):
    """Write headers + data rows to a worksheet with minimal styling."""
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    thin_border = Border(
        bottom=Side(style="thin", color="8B7355"),
    )
    hdr = _excel_header_style(ws)

    # Headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr["font"]
        cell.fill = hdr["fill"]
        cell.alignment = hdr["alignment"]

    # Data
    for row_idx, record in enumerate(data_rows, 2):
        for col_idx, key in enumerate(keys, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(key))
            cell.border = thin_border

    # Auto-width (basic)
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row_idx in range(2, min(len(data_rows) + 2, 20)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3


def generate_excel(data: dict, output_path: str) -> str:
    """Generate detailed Excel report with 4 sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    header_fill = PatternFill(start_color="1B2A3D", end_color="1B2A3D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="8B7355"),
    )

    # --- Sheet 1: Resumo ---
    ws = wb.active
    ws.title = "Resumo"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25

    total = data["total_entities"]
    covered = data["total_covered"]
    uncovered = data["total_uncovered"]
    pct = round(covered / total * 100, 1) if total > 0 else 0

    summary_data = [
        ("Indicador", "Valor", ""),
        ("Cobertura Total (%)", f"{pct}%", ""),
        ("Entes Cobertos", covered, ""),
        ("Entes Descobertos", uncovered, ""),
        ("Total de Entes", total, ""),
        ("", "", ""),
        ("Cobertura por Fonte", "", ""),
        ("Fonte", "Cobertos", "%"),
    ]
    for s in data.get("summary", []):
        summary_data.append((s["source"], s["covered"], f'{s["pct"]}%'))

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1 or row_idx == 8:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            if row_idx > 1:
                cell.border = thin_border

    # --- Sheet 2: Entes Descobertos ---
    ws2 = wb.create_sheet("Entes Descobertos")
    _excel_write_sheet(
        ws2,
        headers=["Razão Social", "Município", "Natureza Jurídica", "CNPJ-8", "UF"],
        keys=["razao_social", "municipio", "natureza_juridica", "cnpj_8", "uf"],
        data_rows=data.get("gaps", []),
    )

    # --- Sheet 3: Cobertura por Município ---
    ws3 = wb.create_sheet("Cobertura por Municipio")
    _excel_write_sheet(
        ws3,
        headers=["Município", "Total Entes", "Entes Descobertos", "% Gap"],
        keys=["municipio", "total_entes", "entes_descobertos", "pct_gap"],
        data_rows=[
            g for g in data.get("gaps_by_municipio", [])
            if g.get("entes_descobertos", 0) > 0
        ],
    )

    # --- Sheet 4: Cobertura por Natureza Jurídica ---
    ws4 = wb.create_sheet("Cobertura por Natureza")
    _excel_write_sheet(
        ws4,
        headers=["Natureza Jurídica", "Total Entes", "Entes Descobertos", "% Gap"],
        keys=["natureza_juridica", "total_entes", "entes_descobertos", "pct_gap"],
        data_rows=[
            g for g in data.get("gaps_by_natureza", [])
            if g.get("entes_descobertos", 0) > 0
        ],
    )

    # Freeze panes on data sheets
    for ws_obj in [ws2, ws3, ws4]:
        ws_obj.freeze_panes = "A2"

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Gera relatório semanal de cobertura (PDF + Excel).",
    )
    parser.add_argument(
        "--date", "-d",
        type=str,
        default=None,
        help="Data-base do relatório (YYYY-MM-DD). Default: hoje.",
    )
    parser.add_argument(
        "--output-dir", "-o",
        type=str,
        default=None,
        help="Diretório de saída. Default: output/reports/coverage/YYYY-MM-DD/",
    )
    parser.add_argument(
        "--snapshot-only",
        action="store_true",
        help="Apenas gerar snapshot, sem PDF/Excel.",
    )
    parser.add_argument(
        "--skip-snapshot",
        action="store_true",
        help="Pular geração de snapshot (usar dados existentes).",
    )
    parser.add_argument(
        "--dsn",
        default=DSN,
        help="PostgreSQL DSN (default: LOCAL_DATALAKE_DSN env ou conexão local).",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Log detalhado.",
    )
    return parser.parse_args()


def main():
    """Main entry point — generates weekly coverage report."""
    args = parse_args()
    global DSN
    DSN = args.dsn

    if args.verbose:
        logger.setLevel(logging.DEBUG)

    report_date = date.fromisoformat(args.date) if args.date else date.today()
    week_number = report_date.isocalendar()[1]
    start_time = time.time()

    logger.info("=" * 60)
    logger.info("Relatório Semanal de Cobertura")
    logger.info("Data-base: %s (Semana %d)", report_date.isoformat(), week_number)
    logger.info("=" * 60)

    # --- Optional: generate snapshot ---
    if not args.skip_snapshot and not args.snapshot_only:
        logger.info("Gerando snapshot de cobertura...")
        snap_start = time.time()
        n_sources = generate_snapshot(report_date)
        snap_elapsed = time.time() - snap_start
        logger.info("Snapshot gerado: %d fontes em %.2fs", n_sources, snap_elapsed)

    if args.snapshot_only:
        logger.info("Modo snapshot-only concluído.")
        return 0

    # --- Fetch data ---
    logger.info("Coletando dados de cobertura...")
    query_start = time.time()
    try:
        data = fetch_coverage_data(report_date)
    except Exception as e:
        logger.error("Erro ao coletar dados do banco: %s", e)
        # Create minimal data for fallback PDF
        data = {
            "total_entities": 0,
            "total_covered": 0,
            "total_uncovered": 0,
            "summary": [],
            "gaps": [],
            "gaps_by_municipio": [],
            "gaps_by_natureza": [],
            "trend": [],
        }
    query_elapsed = time.time() - query_start
    logger.info("Dados coletados em %.2fs", query_elapsed)

    # --- Compute metrics ---
    total = data["total_entities"]
    covered = data["total_covered"]
    uncovered = data["total_uncovered"]
    pct = round(covered / total * 100, 1) if total > 0 else 0
    logger.info("Cobertura: %d/%d entes (%.1f%%)", covered, total, pct)
    logger.info("Gaps totais: %d entes descobertos", uncovered)
    logger.info("Fontes: %d | Municípios com gaps: %d | Snapshots: %d",
                len(data.get("summary", [])),
                len([g for g in data.get("gaps_by_municipio", []) if g.get("entes_descobertos", 0) > 0]),
                len(data.get("trend", [])),
                )

    # --- Output path ---
    date_str = report_date.isoformat()
    if args.output_dir:
        base_dir = Path(args.output_dir)
    else:
        base_dir = _PROJECT_ROOT / "output" / "reports" / "coverage" / date_str

    base_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = str(base_dir / f"coverage-report-{date_str}.pdf")
    xlsx_path = str(base_dir / f"coverage-detail-{date_str}.xlsx")

    # --- Generate PDF ---
    logger.info("Gerando PDF executivo...")
    pdf_start = time.time()
    try:
        generate_pdf(data, pdf_path, report_date)
        pdf_elapsed = time.time() - pdf_start
        pdf_size = Path(pdf_path).stat().st_size
        logger.info("PDF gerado: %s (%.1fKB em %.2fs)", pdf_path, pdf_size / 1024, pdf_elapsed)
    except Exception as e:
        logger.error("Erro ao gerar PDF: %s", e)
        # Create minimal fallback PDF
        try:
            _build_fallback_pdf(pdf_path, str(e), report_date)
            pdf_size = Path(pdf_path).stat().st_size
            logger.warning("PDF fallback gerado: %s (%.1fKB)", pdf_path, pdf_size / 1024)
        except Exception as e2:
            logger.error("Erro ao gerar PDF fallback: %s", e2)

    # --- Generate Excel ---
    logger.info("Gerando Excel detalhado...")
    xlsx_start = time.time()
    try:
        generate_excel(data, xlsx_path)
        xlsx_elapsed = time.time() - xlsx_start
        xlsx_size = Path(xlsx_path).stat().st_size
        logger.info("Excel gerado: %s (%.1fKB em %.2fs)", xlsx_path, xlsx_size / 1024, xlsx_elapsed)
    except Exception as e:
        logger.error("Erro ao gerar Excel: %s", e)

    total_elapsed = time.time() - start_time
    logger.info("-" * 60)
    logger.info("Relatório concluído em %.2fs", total_elapsed)
    logger.info("  PDF:  %s", pdf_path)
    logger.info("  XLSX: %s", xlsx_path)
    logger.info("-" * 60)

    return 0


def _build_fallback_pdf(output_path: str, error_msg: str, report_date: date):
    """Generate a minimal fallback PDF when data fetch fails."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    styles = _build_pdf_styles()
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=2.2 * cm, rightMargin=2.2 * cm,
        topMargin=2.2 * cm, bottomMargin=2.2 * cm,
    )
    story = [
        Paragraph("Relatório de Cobertura Semanal", styles["cover_title"]),
        Spacer(1, 10),
        Paragraph(f"Data: {report_date.isoformat()}", styles["cover_info"]),
        Spacer(1, 20),
        Paragraph("Erro ao gerar relatório", styles["h2"]),
        Paragraph(f"Não foi possível coletar os dados do banco: {error_msg}", styles["body"]),
        Paragraph("Verifique a conexão com o banco de dados e tente novamente.", styles["body"]),
    ]
    doc.build(story)


if __name__ == "__main__":
    sys.exit(main())
