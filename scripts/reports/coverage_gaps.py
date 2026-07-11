#!/usr/bin/env python3
"""Coverage Gap Report — Extra Consultoria.

Exporta lista de entes descobertos (sem cobertura em nenhuma fonte)
para Excel estruturado, com abas por município e natureza jurídica.

Uso:
    python scripts/reports/coverage_gaps.py
    python scripts/reports/coverage_gaps.py --dsn "postgresql://..."
    python scripts/reports/coverage_gaps.py --output /tmp/gaps.xlsx
    python scripts/reports/coverage_gaps.py --municipio Florianópolis
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date
from pathlib import Path
from typing import Any

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

DSN = os.getenv(
    "LOCAL_DATALAKE_DSN",
    "postgresql://postgres:smartlic_local@127.0.0.1:54399/postgres",
)


def get_conn():
    import psycopg2
    return psycopg2.connect(DSN)


def query(conn, sql: str, params: list = None) -> list[dict]:
    cur = conn.cursor()
    cur.execute(sql, params)
    cols = [d[0] for d in cur.description] if cur.description else []
    rows = [dict(zip(cols, row)) for row in cur.fetchall()]
    cur.close()
    return rows


def fetch_all_gaps(conn, municipio: str | None = None) -> list[dict]:
    """Fetch uncovered entities from v_coverage_gaps view."""
    where = ""
    params: list[Any] = []
    if municipio:
        where = "WHERE LOWER(municipio) LIKE LOWER(%s)"
        params.append(f"%{municipio}%")

    rows = query(
        conn,
        f"""SELECT id, razao_social, cnpj_8, municipio,
                   natureza_juridica, raio_200km,
                   fontes_ativas, gap_total
            FROM v_coverage_gaps
            {where}
            ORDER BY municipio, razao_social""",
        params,
    )
    # Convert array columns to string for Excel compatibility
    for r in rows:
        if isinstance(r.get("fontes_ativas"), (list, tuple)):
            r["fontes_ativas"] = ", ".join(str(x) for x in r["fontes_ativas"])
        if r.get("fontes_ativas") is None:
            r["fontes_ativas"] = "nenhuma"
    return rows


def fetch_gaps_by_municipio(conn) -> list[dict]:
    """Fetch gap aggregation by municipio."""
    return query(
        conn,
        """SELECT municipio, total_entes, entes_descobertos,
                  pct_gap, pct_coberto
           FROM v_coverage_gaps_by_municipio
           ORDER BY entes_descobertos DESC""",
    )


def export_excel(
    gaps: list[dict],
    gaps_by_muni: list[dict],
    filepath: str,
) -> None:
    """Export gaps to styled Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERRO: openpyxl nao instalado. Instale com: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # -- Style definitions --
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    alt_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="CC9999"))

    def _write_sheet(ws, data: list[dict], headers: list[str], keys: list[str]):
        """Write headers + data to a worksheet."""
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_idx, record in enumerate(data, 2):
            for col_idx, key in enumerate(keys, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=record.get(key))
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
        # Auto-adjust column widths (approximate)
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_idx, record in enumerate(data, 2):
                val = str(record.get(keys[col_idx - 1], ""))
                max_len = max(max_len, len(val))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 60)

    # -- Sheet 1: All gaps (detailed) --
    ws1 = wb.active
    ws1.title = "Gaps Detalhados"
    _write_sheet(
        ws1,
        gaps,
        ["ID", "Órgão", "CNPJ-8", "Município", "Natureza Jurídica", "Raio 200km", "Fontes Ativas", "Gap Total"],
        ["id", "razao_social", "cnpj_8", "municipio", "natureza_juridica",
         "raio_200km", "fontes_ativas", "gap_total"],
    )

    # -- Sheet 2: By municipio (aggregated) --
    ws2 = wb.create_sheet("Gaps por Municipio")
    _write_sheet(
        ws2,
        gaps_by_muni,
        ["Município", "Total Entes", "Entes Descobertos", "% Gap", "% Coberto"],
        ["municipio", "total_entes", "entes_descobertos", "pct_gap", "pct_coberto"],
    )

    # -- Sheet 3: Summary stats --
    ws3 = wb.create_sheet("Resumo")
    total_gaps = len(gaps)
    total_muni = len(gaps_by_muni)
    within_200km = sum(1 for g in gaps if g.get("raio_200km"))
    naturezas = set(g.get("natureza_juridica") for g in gaps if g.get("natureza_juridica"))

    ws3.cell(row=1, column=1, value="Métrica").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Valor").font = Font(bold=True)
    ws3.cell(row=2, column=1, value="Data do Relatório")
    ws3.cell(row=2, column=2, value=str(date.today()))
    ws3.cell(row=3, column=1, value="Total de Gaps")
    ws3.cell(row=3, column=2, value=total_gaps)
    ws3.cell(row=4, column=1, value="Gaps Raio 200km")
    ws3.cell(row=4, column=2, value=within_200km)
    ws3.cell(row=5, column=1, value="Municípios Afetados")
    ws3.cell(row=5, column=2, value=total_muni)
    ws3.cell(row=6, column=1, value="Naturezas Jurídicas")
    ws3.cell(row=6, column=2, value=len(naturezas))

    wb.save(filepath)
    print(f"Excel exportado: {filepath}")


def main() -> int:
    global DSN
    parser = argparse.ArgumentParser(description="Coverage Gap Export")
    parser.add_argument("--dsn", default=DSN, help="PostgreSQL DSN")
    parser.add_argument("--output", help="Output path (default: output/reports/coverage/gaps-YYYY-MM-DD.xlsx)")
    parser.add_argument("--municipio", help="Filter by municipio (partial match)")
    args = parser.parse_args()

    if args.dsn:
        DSN = args.dsn

    # Ensure output directory
    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = _PROJECT_ROOT / "output" / "reports" / "coverage"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"gaps-{date.today().isoformat()}.xlsx"

    # Fetch data
    conn = get_conn()
    try:
        gaps = fetch_all_gaps(conn, municipio=args.municipio)
        gaps_by_muni = fetch_gaps_by_municipio(conn)
    finally:
        conn.close()

    if not gaps:
        print("Nenhum gap de cobertura encontrado.")
        return 0

    print(f"Total de gaps: {len(gaps)} entidades em {len(gaps_by_muni)} municipios")

    # Export to Excel
    export_excel(gaps, gaps_by_muni, str(output_path))

    return 0


if __name__ == "__main__":
    sys.exit(main())
