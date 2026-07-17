#!/usr/bin/env python3
"""Commercial B2G session report — artifact-first with honest gaps.

Loads real session artifacts (Compras SC, DOM/CIGA, DOE Dados SC, coverage,
reconciliation, org ranking) and optional DB samples. Never invents GO or
hides low coverage / stale sources / partial execution.

Usage:
  PYTHONPATH=. python3 -m scripts.reports.commercial_b2g_session \\
    --output-dir output/reports
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
from collections import Counter
from datetime import UTC, datetime
from html import escape
from pathlib import Path
from typing import Any

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from scripts.reports.run_metadata import (  # noqa: E402
    _git_sha_short,
    new_run_id,
)

# Statuses considered "open" for commercial pipeline (Compras SC labels).
OPEN_STATUS_MARKERS = (
    "recebimento de proposta",
    "aguardando abertura",
    "em andamento",
    "publicado",
    "aberto",
)
SEMI_OPEN_STATUS_MARKERS = (
    "aguardando homolog",
)
CLOSED_STATUS_MARKERS = (
    "homologado",
    "fracassado",
    "deserto",
    "cancelado",
    "revogado",
    "anulado",
)


def _load_json(path: Path) -> dict[str, Any] | list[Any] | None:
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def _load_jsonl(path: Path, *, limit: int | None = None) -> list[dict[str, Any]]:
    if not path.is_file():
        return []
    rows: list[dict[str, Any]] = []
    try:
        with path.open(encoding="utf-8", errors="replace") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if isinstance(obj, dict):
                    rows.append(obj)
                if limit is not None and len(rows) >= limit:
                    break
    except OSError:
        return []
    return rows


def _rel(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(_PROJECT_ROOT))
    except ValueError:
        return str(path)


def _sha256_file(path: Path) -> str | None:
    if not path.is_file():
        return None
    h = hashlib.sha256()
    try:
        with path.open("rb") as fh:
            for chunk in iter(lambda: fh.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()
    except OSError:
        return None


def _latest_dir(parent: Path, prefix: str) -> Path | None:
    if not parent.is_dir():
        return None
    candidates = [p for p in parent.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _status_bucket(status: str | None) -> str:
    s = (status or "").strip().lower()
    if not s:
        return "unknown"
    if any(m in s for m in CLOSED_STATUS_MARKERS):
        return "closed"
    if any(m in s for m in OPEN_STATUS_MARKERS):
        return "open"
    if any(m in s for m in SEMI_OPEN_STATUS_MARKERS):
        return "semi_open"
    return "other"


def _truncate(text: str | None, n: int = 220) -> str | None:
    if text is None:
        return None
    t = re.sub(r"\s+", " ", str(text)).strip()
    if len(t) <= n:
        return t
    return t[: n - 1] + "…"


def _pick_sc_compras_run(prefer: str = "incremental") -> dict[str, Any]:
    """Return artifact + records for the preferred Compras SC run."""
    root = _PROJECT_ROOT / "output" / "sc_compras"
    out: dict[str, Any] = {
        "available": False,
        "prefer": prefer,
        "artifact_path": None,
        "jsonl_path": None,
        "artifact": None,
        "records": [],
        "live_fetch": None,
        "mode": None,
        "run_id": None,
    }
    if not root.is_dir():
        out["reason"] = "output/sc_compras missing"
        return out

    # Prefer incremental, then smoke, by mtime within prefix.
    for prefix in (f"sc_compras-{prefer}", "sc_compras-incremental", "sc_compras-smoke"):
        d = _latest_dir(root, prefix)
        if d is None:
            continue
        art_path = d / "artifact.json"
        jsonl = d / "licitacoes.jsonl"
        art = _load_json(art_path)
        if not isinstance(art, dict):
            continue
        records = _load_jsonl(jsonl)
        out.update(
            {
                "available": True,
                "artifact_path": _rel(art_path),
                "jsonl_path": _rel(jsonl) if jsonl.is_file() else None,
                "artifact": art,
                "records": records,
                "live_fetch": bool(art.get("live_fetch")),
                "mode": art.get("mode"),
                "run_id": art.get("run_id"),
                "status": art.get("status"),
                "artifact_sha256": art.get("artifact_sha256") or _sha256_file(art_path),
            }
        )
        # Also load smoke for status diversity if primary is incremental
        if prefer == "incremental" and prefix.startswith("sc_compras-incremental"):
            smoke = _latest_dir(root, "sc_compras-smoke")
            if smoke is not None:
                s_art = _load_json(smoke / "artifact.json")
                s_rows = _load_jsonl(smoke / "licitacoes.jsonl")
                out["smoke_side"] = {
                    "run_id": (s_art or {}).get("run_id") if isinstance(s_art, dict) else None,
                    "live_fetch": (s_art or {}).get("live_fetch") if isinstance(s_art, dict) else None,
                    "records_count": len(s_rows),
                    "status_counts": dict(Counter(r.get("status") for r in s_rows)),
                    "artifact_path": _rel(smoke / "artifact.json"),
                }
        break
    if not out["available"]:
        out["reason"] = "no sc_compras run directories found"
    return out


def _summarize_opportunities(records: list[dict[str, Any]]) -> dict[str, Any]:
    open_rows: list[dict[str, Any]] = []
    semi: list[dict[str, Any]] = []
    closed: list[dict[str, Any]] = []
    recent: list[dict[str, Any]] = []
    orgs: Counter[str] = Counter()
    modalities: Counter[str] = Counter()
    status_counts: Counter[str] = Counter()
    missing_docs = 0
    missing_value = 0
    missing_municipio = 0

    for r in records:
        status = r.get("status")
        status_counts[str(status or "null")] += 1
        bucket = _status_bucket(str(status) if status else None)
        orgao = r.get("orgao_razao_social") or r.get("orgao") or "—"
        orgs[str(orgao)] += 1
        mod = r.get("modalidade_nome") or r.get("modalidade") or "—"
        modalities[str(mod)] += 1
        docs = r.get("documentos")
        if not docs:
            missing_docs += 1
        if r.get("valor_total_estimado") is None and r.get("valor_total") is None:
            missing_value += 1
        if not r.get("municipio"):
            missing_municipio += 1

        row = {
            "source_id": r.get("source_id") or r.get("pncp_id") or r.get("api_id"),
            "status": status,
            "status_bucket": bucket,
            "orgao": orgao,
            "orgao_cnpj": r.get("orgao_cnpj"),
            "municipio": r.get("municipio"),
            "uf": r.get("uf") or "SC",
            "modalidade": mod,
            "objeto": _truncate(r.get("objeto_compra") or r.get("objeto"), 280),
            "valor_total_estimado": r.get("valor_total_estimado"),
            "data_publicacao": r.get("data_publicacao"),
            "data_abertura": r.get("data_abertura"),
            "data_encerramento": r.get("data_encerramento"),
            "link_oficial": r.get("link_pncp") or r.get("link") or r.get("url"),
            "documentos_count": len(docs) if isinstance(docs, list) else 0,
        }
        if bucket == "open":
            open_rows.append(row)
        elif bucket == "semi_open":
            semi.append(row)
        elif bucket == "closed":
            closed.append(row)
        recent.append(row)

    # Sort recent by publication date desc
    def _key(x: dict[str, Any]) -> str:
        return str(x.get("data_publicacao") or "")

    recent_sorted = sorted(recent, key=_key, reverse=True)
    open_rows = sorted(open_rows, key=_key, reverse=True)

    return {
        "total_records": len(records),
        "status_counts": dict(status_counts),
        "open_count": len(open_rows),
        "semi_open_count": len(semi),
        "closed_count": len(closed),
        "open_opportunities": open_rows[:50],
        "semi_open_sample": semi[:20],
        "recently_published": recent_sorted[:30],
        "buyer_orgs_top": [{"orgao": k, "n": v} for k, v in orgs.most_common(20)],
        "modalities": [{"modalidade": k, "n": v} for k, v in modalities.most_common(20)],
        "gaps_in_sample": {
            "missing_documentos": missing_docs,
            "missing_valor_estimado": missing_value,
            "missing_municipio": missing_municipio,
            "pct_missing_documentos": round(100.0 * missing_docs / len(records), 1) if records else None,
            "pct_missing_valor": round(100.0 * missing_value / len(records), 1) if records else None,
            "pct_missing_municipio": round(100.0 * missing_municipio / len(records), 1) if records else None,
        },
    }


def _load_dom_acts(*, sample_limit: int = 40) -> dict[str, Any]:
    root = _PROJECT_ROOT / "output" / "ciga_dom"
    summary = _load_json(root / "latest_summary.json")
    evidence = _load_json(root / "latest_evidence.json")
    freshness = _load_json(root / "freshness_manifest.json")
    run_dir = None
    if isinstance(summary, dict) and summary.get("run_id"):
        cand = root / str(summary["run_id"])
        if cand.is_dir():
            run_dir = cand
    if run_dir is None:
        run_dir = _latest_dir(root, "ciga-dom-")

    jsonl_path = None
    if run_dir is not None:
        jsonl_path = run_dir / "publications.jsonl"
    records = _load_jsonl(jsonl_path, limit=8000) if jsonl_path else []

    act_counts: Counter[str] = Counter()
    conf_labels: Counter[str] = Counter()
    munis: Counter[str] = Counter()
    recent_acts: list[dict[str, Any]] = []
    procurement_like = 0

    procurement_cats = {
        "aviso_licitacao",
        "edital",
        "homologacao",
        "extrato_contrato",
        "termo_aditivo",
        "ata_registro_precos",
        "dispensa",
        "inexigibilidade",
        "resultado",
        "outros_atos_contratacao",
        "chamamento_publico",
        "credenciamento",
        "rescisao",
        "anulacao",
        "revogacao",
        "apostilamento",
        "errata",
        "suspensao",
    }

    for r in records:
        cat = str(r.get("act_category") or "unknown")
        act_counts[cat] += 1
        conf_labels[str(r.get("act_confidence_label") or "unknown")] += 1
        mun = r.get("municipio")
        if mun:
            munis[str(mun)] += 1
        if cat in procurement_cats:
            procurement_like += 1
        if len(recent_acts) < sample_limit and cat in procurement_cats:
            recent_acts.append(
                {
                    "data": r.get("data"),
                    "municipio": r.get("municipio"),
                    "orgao": r.get("orgao") or r.get("entidade"),
                    "titulo": _truncate(r.get("titulo"), 200),
                    "categoria_dom": r.get("categoria_dom"),
                    "act_category": cat,
                    "act_confidence": r.get("act_confidence"),
                    "act_confidence_label": r.get("act_confidence_label"),
                    "act_needs_human_review": r.get("act_needs_human_review"),
                    "url": r.get("url"),
                    "codigo": r.get("codigo"),
                }
            )

    live_fetch = None
    if isinstance(evidence, dict):
        # evidence criteria may include live_ciga_api
        crit = evidence.get("criteria") or {}
        live_fetch = bool(crit.get("live_ciga_api")) if crit else True

    return {
        "available": bool(records) or isinstance(summary, dict),
        "run_id": (summary or {}).get("run_id") if isinstance(summary, dict) else None,
        "live_fetch": live_fetch,
        "status": (summary or {}).get("status") if isinstance(summary, dict) else None,
        "package_id": (summary or {}).get("package_id") if isinstance(summary, dict) else None,
        "counts": (summary or {}).get("counts") if isinstance(summary, dict) else None,
        "municipalities": (summary or {}).get("municipalities") if isinstance(summary, dict) else None,
        "records_loaded": len(records),
        "act_category_counts": dict(act_counts.most_common(30)),
        "act_confidence_labels": dict(conf_labels),
        "procurement_like_count": procurement_like,
        "municipalities_observed_top": [{"municipio": k, "n": v} for k, v in munis.most_common(20)],
        "recent_procurement_acts": recent_acts,
        "freshness_manifest": {
            "path": "output/ciga_dom/freshness_manifest.json",
            "status": (freshness or {}).get("status") if isinstance(freshness, dict) else None,
            "latest_resource_modified": (freshness or {}).get("latest_resource_modified")
            if isinstance(freshness, dict)
            else None,
            "package_year_month": (freshness or {}).get("package_year_month")
            if isinstance(freshness, dict)
            else None,
            "resource_count": (freshness or {}).get("resource_count")
            if isinstance(freshness, dict)
            else None,
        },
        "evidence_path": "output/ciga_dom/latest_evidence.json",
        "summary_path": "output/ciga_dom/latest_summary.json",
        "jsonl_path": _rel(jsonl_path) if jsonl_path and jsonl_path.is_file() else None,
        "mode": (summary or {}).get("mode") if isinstance(summary, dict) else None,
    }


def _load_doe_acts() -> dict[str, Any]:
    root = _PROJECT_ROOT / "output" / "dados_abertos_sc"
    smokes = sorted(root.glob("smoke-*.json"), key=lambda p: p.stat().st_mtime, reverse=True) if root.is_dir() else []
    if not smokes:
        return {"available": False, "reason": "no DOE smoke artifact"}
    path = smokes[0]
    data = _load_json(path)
    if not isinstance(data, dict):
        return {"available": False, "reason": f"invalid JSON {path.name}"}

    samples = data.get("sample_records") or []
    acts = []
    for r in samples:
        if not isinstance(r, dict):
            continue
        acts.append(
            {
                "data_publicacao": r.get("data_publicacao") or r.get("data_edicao"),
                "orgao": r.get("orgao") or r.get("unidade"),
                "titulo": _truncate(r.get("titulo") or r.get("assunto"), 200),
                "tipo_ato": r.get("tipo_ato"),
                "categoria": r.get("categoria"),
                "act_category_from_counts_only": True,
                "link_extrato": r.get("link_extrato"),
                "link_edicao": r.get("link_edicao"),
                "numero_publicacao": r.get("numero_publicacao"),
                "numero_edicao": r.get("numero_edicao"),
            }
        )

    return {
        "available": True,
        "run_id": data.get("run_id"),
        "live_fetch": bool(data.get("live_fetch")),
        "status": data.get("status"),
        "portal": data.get("portal"),
        "package": data.get("package"),
        "mode": data.get("mode"),
        "artifact_path": _rel(path),
        "counts": data.get("counts"),
        "act_categories": (data.get("counts") or {}).get("act_categories") if isinstance(data.get("counts"), dict) else None,
        "sample_recent_acts": acts,
        "selected_resources": data.get("selected_resources"),
        "note": (
            "Amostra DOE é smoke (1 resource); act_categories refletem rows normalizadas "
            "do resource processado — não universo DOE completo."
        ),
    }


def _load_coverage_bundle() -> dict[str, Any]:
    metrics = _load_json(_PROJECT_ROOT / "output/coverage/next30d-metrics-final.json") or {}
    coverage_gate = _load_json(_PROJECT_ROOT / "output/coverage/coverage-gate-report.json") or {}
    coverage_truth = _load_json(
        _PROJECT_ROOT / "output/sc_compras/coverage-truth/coverage-truth-2026-07-16.json"
    ) or {}
    freshness = _load_json(_PROJECT_ROOT / "output/readiness/freshness-gate.json") or {}
    runtime = _load_json(_PROJECT_ROOT / "output/sc_compras/runtime-next30d.json") or {}

    editais_pct = metrics.get("editais_crude_pct")
    if editais_pct is None and isinstance(coverage_gate, dict):
        editais_pct = coverage_gate.get("editais_crude_pct")

    mon = coverage_truth.get("monitoring_coverage") if isinstance(coverage_truth, dict) else None
    gaps = coverage_truth.get("gaps") if isinstance(coverage_truth, dict) else None
    source_health = coverage_truth.get("source_health") if isinstance(coverage_truth, dict) else None

    return {
        "editais_crude_pct": editais_pct,
        "covered_200km": metrics.get("covered_200km"),
        "editais_denominator": metrics.get("editais_denominator"),
        "pncp_supplier_contracts": metrics.get("pncp_supplier_contracts"),
        "pncp_raw_bids": metrics.get("pncp_raw_bids"),
        "pilot_status": metrics.get("pilot_status"),
        "go_no_go_3y": metrics.get("go_no_go_3y"),
        "go_no_go_path": metrics.get("go_no_go_path"),
        "metrics_note": metrics.get("note"),
        "metrics_path": "output/coverage/next30d-metrics-final.json",
        "coverage_truth": {
            "path": "output/sc_compras/coverage-truth/coverage-truth-2026-07-16.json",
            "monitoring_pct_display": (mon or {}).get("pct_display") if isinstance(mon, dict) else None,
            "monitoring_pct": (mon or {}).get("pct") if isinstance(mon, dict) else None,
            "freshness": coverage_truth.get("freshness") if isinstance(coverage_truth, dict) else None,
            "bid_presence": coverage_truth.get("bid_presence") if isinstance(coverage_truth, dict) else None,
            "contract_presence": coverage_truth.get("contract_presence")
            if isinstance(coverage_truth, dict)
            else None,
            "gaps_total": (gaps or {}).get("total_gap_combinations") if isinstance(gaps, dict) else None,
            "gaps_sample": (gaps or {}).get("sample") if isinstance(gaps, dict) else None,
            "source_health_notes": {
                k: (v.get("status") or v.get("_note") or v.get("blocker_reason"))
                for k, v in (source_health or {}).items()
                if isinstance(v, dict)
            }
            if isinstance(source_health, dict)
            else {},
        },
        "freshness_gate": {
            "path": "output/readiness/freshness-gate.json",
            "all_critical_fresh": (freshness.get("overall") or {}).get("all_critical_sources_fresh")
            if isinstance(freshness, dict)
            else None,
            "failing_sources": (freshness.get("overall") or {}).get("failing_sources")
            if isinstance(freshness, dict)
            else None,
            "generated_at": freshness.get("generated_at") if isinstance(freshness, dict) else None,
        },
        "runtime_next30d": {
            "path": "output/sc_compras/runtime-next30d.json",
            "summary": runtime.get("summary") if isinstance(runtime, dict) else None,
            "results": runtime.get("results") if isinstance(runtime, dict) else None,
        },
        "code_coverage_gate": {
            "path": "output/coverage/coverage-gate-report.json",
            "result": coverage_gate.get("result") if isinstance(coverage_gate, dict) else None,
            "note": "Gate de cobertura de código (pytest-cov), NÃO cobertura comercial de editais.",
        },
    }


def _load_reconciliation() -> dict[str, Any]:
    path = _PROJECT_ROOT / "output/reports/reconcile-next30d.json"
    data = _load_json(path)
    if not isinstance(data, dict):
        alt = _PROJECT_ROOT / "output/reports/reconcile-pdf-excel-2026-07-17.json"
        data = _load_json(alt)
        path = alt if isinstance(data, dict) else path
    if not isinstance(data, dict):
        return {"available": False, "reason": "no reconciliation artifact"}
    return {
        "available": True,
        "path": _rel(path),
        "verdict": data.get("verdict"),
        "consistent": data.get("consistent"),
        "critical_mismatches": data.get("critical_mismatches"),
        "soft_mismatches": data.get("soft_mismatches"),
        "pdf_path": data.get("pdf_path"),
        "excel_path": data.get("excel_path"),
        "reason": data.get("reason"),
        "pdf_run_id": (data.get("pdf_meta") or {}).get("run_id"),
        "excel_run_id": (data.get("excel_meta") or {}).get("run_id"),
    }


def _load_org_ranking(limit: int = 25) -> dict[str, Any]:
    path = _PROJECT_ROOT / "output/reports/org-ranking-next30d.json"
    data = _load_json(path)
    if not isinstance(data, dict):
        return {"available": False, "reason": "org-ranking missing"}
    ranking = data.get("ranking") or {}
    rows = ranking.get("rows") or []
    return {
        "available": ranking.get("status") == "OK" and bool(rows),
        "path": _rel(path),
        "status": ranking.get("status"),
        "source_table": ranking.get("source_table"),
        "valor_semantica": ranking.get("valor_semantica"),
        "uf_filter": ranking.get("uf_filter"),
        "count": ranking.get("count"),
        "rows": rows[:limit],
        "notes": ranking.get("notes"),
        "claims": ranking.get("claims"),
        "generated_at": data.get("generated_at"),
    }


def _db_commercial_sample(dsn: str | None, limit: int = 15) -> dict[str, Any]:
    if not dsn:
        return {"available": False, "reason": "no DSN", "live_fetch": False}
    try:
        import psycopg2
        from psycopg2.extras import RealDictCursor
    except ImportError:
        return {"available": False, "reason": "psycopg2 missing", "live_fetch": False}

    try:
        conn = psycopg2.connect(dsn, connect_timeout=5)
    except Exception as e:  # noqa: BLE001
        return {
            "available": False,
            "reason": f"connect failed: {type(e).__name__}: {e}",
            "live_fetch": False,
        }

    out: dict[str, Any] = {"available": True, "live_fetch": True}
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT COUNT(*) AS n FROM pncp_supplier_contracts")
        out["contracts_total"] = int((cur.fetchone() or {}).get("n") or 0)
        cur.execute(
            """
            SELECT COUNT(*) AS n FROM pncp_supplier_contracts
            WHERE uf = 'SC'
            """
        )
        out["contracts_sc"] = int((cur.fetchone() or {}).get("n") or 0)
        cur.execute(
            """
            SELECT orgao_nome, orgao_cnpj, COUNT(*) AS n,
                   SUM(valor_total) AS valor_sum
            FROM pncp_supplier_contracts
            WHERE uf = 'SC'
            GROUP BY 1, 2
            ORDER BY n DESC NULLS LAST
            LIMIT %s
            """,
            (limit,),
        )
        out["top_orgaos_sc"] = [dict(r) for r in cur.fetchall()]
        for row in out["top_orgaos_sc"]:
            if row.get("valor_sum") is not None:
                row["valor_sum"] = float(row["valor_sum"])
        cur.execute(
            """
            SELECT fornecedor_nome, fornecedor_cnpj, COUNT(*) AS n,
                   SUM(valor_total) AS valor_sum
            FROM pncp_supplier_contracts
            WHERE uf = 'SC' AND fornecedor_cnpj IS NOT NULL
            GROUP BY 1, 2
            ORDER BY n DESC NULLS LAST
            LIMIT %s
            """,
            (limit,),
        )
        out["winning_suppliers_sc"] = [dict(r) for r in cur.fetchall()]
        for row in out["winning_suppliers_sc"]:
            if row.get("valor_sum") is not None:
                row["valor_sum"] = float(row["valor_sum"])
        # Similar contracts sample (engineering-ish keywords if column exists)
        try:
            cur.execute(
                """
                SELECT orgao_nome, fornecedor_nome, objeto_contrato, valor_total,
                       data_publicacao, contrato_id, uf, municipio
                FROM pncp_supplier_contracts
                WHERE uf = 'SC'
                  AND (
                    objeto_contrato ILIKE %s OR objeto_contrato ILIKE %s
                    OR objeto_contrato ILIKE %s OR objeto_contrato ILIKE %s
                    OR objeto_contrato ILIKE %s OR objeto_contrato ILIKE %s
                  )
                ORDER BY data_publicacao DESC NULLS LAST
                LIMIT %s
                """,
                (
                    "%obra%",
                    "%paviment%",
                    "%construção%",
                    "%construcao%",
                    "%engenharia%",
                    "%reforma%",
                    limit,
                ),
            )
            out["similar_contracts_sc"] = [dict(r) for r in cur.fetchall()]
            for row in out["similar_contracts_sc"]:
                if row.get("valor_total") is not None:
                    row["valor_total"] = float(row["valor_total"])
                if row.get("objeto_contrato") is not None:
                    row["objeto"] = _truncate(str(row["objeto_contrato"]), 240)
                    row["objeto_contrato"] = row["objeto"]
        except Exception as e:  # noqa: BLE001
            out["similar_contracts_sc"] = []
            out["similar_contracts_error"] = f"{type(e).__name__}: {e}"
        try:
            cur.execute(
                """
                SELECT COUNT(*) AS n FROM pncp_raw_bids WHERE uf = 'SC'
                """
            )
            out["raw_bids_sc"] = int((cur.fetchone() or {}).get("n") or 0)
        except Exception:  # noqa: BLE001
            out["raw_bids_sc"] = None
        cur.close()
    except Exception as e:  # noqa: BLE001
        out["query_error"] = f"{type(e).__name__}: {e}"
    finally:
        conn.close()
    return out


def _detect_status_changes(
    primary: list[dict[str, Any]],
    secondary: list[dict[str, Any]],
) -> dict[str, Any]:
    """Detect status diffs for shared source_ids between two runs (if any)."""
    by_id_a = {
        str(r.get("source_id") or r.get("pncp_id") or r.get("api_id")): r.get("status")
        for r in primary
        if r.get("source_id") or r.get("pncp_id") or r.get("api_id")
    }
    by_id_b = {
        str(r.get("source_id") or r.get("pncp_id") or r.get("api_id")): r.get("status")
        for r in secondary
        if r.get("source_id") or r.get("pncp_id") or r.get("api_id")
    }
    shared = set(by_id_a) & set(by_id_b)
    changes = []
    for sid in sorted(shared):
        if by_id_a[sid] != by_id_b[sid]:
            changes.append(
                {
                    "source_id": sid,
                    "status_a": by_id_a[sid],
                    "status_b": by_id_b[sid],
                }
            )
    return {
        "comparable": bool(shared),
        "shared_ids": len(shared),
        "changes_count": len(changes),
        "changes": changes[:50],
        "note": (
            "Comparação entre runs incremental e smoke da mesma sessão; "
            "amostras parciais (page limits) — ausência de mudança NÃO prova estabilidade do portal."
            if shared
            else "Sem IDs compartilhados entre amostras; status changes não detectáveis nesta sessão."
        ),
    }


def build_disclaimers(
    *,
    coverage: dict[str, Any],
    sc: dict[str, Any],
    dom: dict[str, Any],
    doe: dict[str, Any],
    db: dict[str, Any],
    recon: dict[str, Any],
) -> list[str]:
    notes: list[str] = []
    pct = coverage.get("editais_crude_pct")
    if pct is not None and float(pct) < 50:
        notes.append(
            f"Cobertura bruta de editais ~{float(pct):.2f}% — amostra incompleta; "
            "não use como universo exaustivo de SC."
        )
    mon = (coverage.get("coverage_truth") or {}).get("monitoring_pct_display")
    if mon in (None, "unverified", "n/a", "N/A"):
        notes.append(
            "Monitoring coverage no coverage-truth está unverified/nulo — "
            "evidence ledger vazio para várias fontes."
        )
    failing = (coverage.get("freshness_gate") or {}).get("failing_sources") or []
    if failing:
        notes.append("Fontes críticas sem prova de freshness: " + ", ".join(str(s) for s in failing))
    if coverage.get("pilot_status") and coverage.get("pilot_status") != "success":
        notes.append(
            f"Piloto de contratos PNCP status={coverage.get('pilot_status')}; "
            "backfill nacional 90d/3y NÃO autorizado."
        )
    if sc.get("available"):
        gaps = (sc.get("opportunities") or {}).get("gaps_in_sample") or {}
        if gaps.get("pct_missing_documentos") == 100 or gaps.get("missing_documentos"):
            notes.append(
                "Compras SC: documentos do edital ausentes na listagem "
                f"({gaps.get('missing_documentos')}/{sc.get('opportunities', {}).get('total_records')} "
                "na amostra) — fetch_detail=false / partial execution."
            )
        if gaps.get("pct_missing_valor") and float(gaps["pct_missing_valor"]) >= 50:
            notes.append(
                f"Compras SC: {gaps.get('pct_missing_valor')}% da amostra sem valor_total_estimado."
            )
        if gaps.get("pct_missing_municipio") and float(gaps["pct_missing_municipio"]) >= 50:
            notes.append(
                f"Compras SC: {gaps.get('pct_missing_municipio')}% da amostra sem município "
                "(órgãos estaduais ou campo não preenchido na API de listagem)."
            )
        metrics = (sc.get("artifact") or {}).get("metrics") or {}
        api_total = metrics.get("api_total_elementos_reported")
        n = sc.get("opportunities", {}).get("total_records")
        if api_total and n and int(api_total) > int(n):
            notes.append(
                f"Compras SC: API reportou {api_total} elementos no ano; amostra normalizada={n} "
                f"(mode={sc.get('mode')}) — execução parcial por page limit."
            )
        if sc.get("live_fetch") is False:
            notes.append("Compras SC: live_fetch=false — dados de atestação/cache, não fetch ao vivo.")
    else:
        notes.append("Compras SC: artefato de sessão indisponível.")

    if dom.get("available"):
        if dom.get("mode") == "smoke":
            notes.append(
                f"DOM/CIGA mode=smoke (package={dom.get('package_id')}); "
                f"records_loaded={dom.get('records_loaded')} — não cobre todos os municípios SC."
            )
        mun = dom.get("municipalities") or {}
        if isinstance(mun, dict) and mun.get("gap_count"):
            notes.append(
                f"DOM/CIGA: gap_count municípios no universo={mun.get('gap_count')} "
                f"(observed={mun.get('observed_count')}, universe={mun.get('universe_count')})."
            )
    else:
        notes.append("DOM/CIGA: artefato de sessão indisponível.")

    if doe.get("available"):
        if doe.get("mode") == "smoke":
            notes.append(
                "DOE (Dados SC): execução smoke — act_categories e samples não representam "
                "todo o Diário Oficial do Estado."
            )
        if doe.get("live_fetch") is False:
            notes.append("DOE: live_fetch=false.")
    else:
        notes.append("DOE (Dados SC): artefato de sessão indisponível.")

    if not db.get("available"):
        notes.append(f"DB sample indisponível ({db.get('reason')}) — sem contratos similares live.")
    if recon.get("available") and recon.get("verdict") not in ("CONSISTENT", "PASS"):
        notes.append(f"Reconciliação PDF×Excel verdict={recon.get('verdict')}.")
    elif not recon.get("available"):
        notes.append("Reconciliação PDF×Excel não encontrada nesta sessão.")

    notes.append(
        "Datas em pncp_supplier_contracts.data_publicacao podem refletir dataAssinatura "
        "histórica (semântica legada) — não interpretar MIN/MAX como cobertura da janela de coleta."
    )
    notes.append("VPS/infra remota não deve ser tratada como operacional sem evidência.")
    notes.append(
        "live_fetch=true indica chamada HTTP real naquele run; ausência de live_fetch "
        "ou artefato de atestação não deve ser apresentado como coleta completa."
    )
    return notes


def _confidence_label(
    *,
    coverage: dict[str, Any],
    sc: dict[str, Any],
    dom: dict[str, Any],
    doe: dict[str, Any],
    db: dict[str, Any],
) -> str:
    score = 0
    if sc.get("available") and sc.get("live_fetch"):
        score += 2
    if dom.get("available"):
        score += 1
    if doe.get("available") and doe.get("live_fetch"):
        score += 1
    if db.get("available"):
        score += 1
    pct = coverage.get("editais_crude_pct")
    if pct is not None and float(pct) >= 50:
        score += 2
    elif pct is not None and float(pct) >= 20:
        score += 1
    failing = (coverage.get("freshness_gate") or {}).get("failing_sources") or []
    if failing:
        score -= 1
    if score >= 6:
        return "medium"
    if score >= 3:
        return "low_to_medium"
    return "low"


def build_session_report(*, dsn: str | None = None) -> dict[str, Any]:
    run_id = new_run_id(prefix="commercial-b2g")
    generated_at = datetime.now(UTC).isoformat()
    git_sha = _git_sha_short()

    sc_raw = _pick_sc_compras_run("incremental")
    opportunities = _summarize_opportunities(sc_raw.get("records") or [])
    sc_block = {
        "available": sc_raw.get("available"),
        "run_id": sc_raw.get("run_id"),
        "mode": sc_raw.get("mode"),
        "status": sc_raw.get("status"),
        "live_fetch": sc_raw.get("live_fetch"),
        "artifact_path": sc_raw.get("artifact_path"),
        "jsonl_path": sc_raw.get("jsonl_path"),
        "artifact_sha256": sc_raw.get("artifact_sha256"),
        "artifact": {
            "metrics": (sc_raw.get("artifact") or {}).get("metrics"),
            "sample_ids": (sc_raw.get("artifact") or {}).get("sample_ids"),
            "errors": (sc_raw.get("artifact") or {}).get("errors"),
            "started_at": (sc_raw.get("artifact") or {}).get("started_at"),
            "completed_at": (sc_raw.get("artifact") or {}).get("completed_at"),
            "evidence": {
                k: (sc_raw.get("artifact") or {}).get("evidence", {}).get(k)
                for k in (
                    "run_id",
                    "git_sha",
                    "git_branch",
                    "started_at",
                    "completed_at",
                    "command",
                    "output_hash",
                    "checkpoint_hash",
                )
            }
            if isinstance((sc_raw.get("artifact") or {}).get("evidence"), dict)
            else None,
        },
        "smoke_side": sc_raw.get("smoke_side"),
        "opportunities": opportunities,
        "reason": sc_raw.get("reason"),
    }

    # status changes between incremental and smoke if both exist
    smoke_records: list[dict[str, Any]] = []
    if sc_raw.get("smoke_side"):
        smoke_path = _PROJECT_ROOT / (sc_raw["smoke_side"].get("artifact_path") or "")
        smoke_dir = smoke_path.parent if smoke_path else None
        if smoke_dir and (smoke_dir / "licitacoes.jsonl").is_file():
            smoke_records = _load_jsonl(smoke_dir / "licitacoes.jsonl")
    status_changes = _detect_status_changes(sc_raw.get("records") or [], smoke_records)

    dom = _load_dom_acts()
    doe = _load_doe_acts()
    coverage = _load_coverage_bundle()
    recon = _load_reconciliation()
    org_ranking = _load_org_ranking()
    db = _db_commercial_sample(dsn)

    evidence_chain = [
        {
            "source": "sc_compras",
            "run_id": sc_block.get("run_id"),
            "live_fetch": sc_block.get("live_fetch"),
            "path": sc_block.get("artifact_path"),
            "sha256": sc_block.get("artifact_sha256"),
            "status": sc_block.get("status"),
        },
        {
            "source": "ciga_dom",
            "run_id": dom.get("run_id"),
            "live_fetch": dom.get("live_fetch"),
            "path": dom.get("summary_path"),
            "status": dom.get("status"),
        },
        {
            "source": "dados_abertos_sc_doe",
            "run_id": doe.get("run_id"),
            "live_fetch": doe.get("live_fetch"),
            "path": doe.get("artifact_path"),
            "status": doe.get("status"),
        },
        {
            "source": "coverage_metrics",
            "run_id": None,
            "live_fetch": False,
            "path": coverage.get("metrics_path"),
            "attestation": True,
        },
        {
            "source": "coverage_truth",
            "run_id": None,
            "live_fetch": False,
            "path": (coverage.get("coverage_truth") or {}).get("path"),
            "attestation": True,
        },
        {
            "source": "freshness_gate",
            "run_id": None,
            "live_fetch": False,
            "path": (coverage.get("freshness_gate") or {}).get("path"),
            "attestation": True,
        },
        {
            "source": "org_ranking",
            "run_id": None,
            "live_fetch": False,
            "path": org_ranking.get("path"),
            "status": org_ranking.get("status"),
            "attestation": True,
        },
        {
            "source": "reconcile_pdf_excel",
            "run_id": recon.get("pdf_run_id") or recon.get("excel_run_id"),
            "live_fetch": False,
            "path": recon.get("path"),
            "status": recon.get("verdict"),
            "attestation": True,
        },
        {
            "source": "db_sample",
            "run_id": None,
            "live_fetch": db.get("live_fetch"),
            "available": db.get("available"),
            "reason": db.get("reason"),
        },
    ]

    disclaimers = build_disclaimers(
        coverage=coverage, sc=sc_block, dom=dom, doe=doe, db=db, recon=recon
    )
    confidence = _confidence_label(coverage=coverage, sc=sc_block, dom=dom, doe=doe, db=db)

    claims_allowed = [
        "Amostra Compras SC com live_fetch e status de portal (abertas/semi/fechadas) da sessão.",
        "Atos DOM/CIGA classificados (act_category) em modo smoke com confidence labels.",
        "Amostra DOE (Dados SC) com contagem de act_categories do resource processado.",
        "Cobertura bruta e gaps expostos sem inflar para 95%.",
        "Ranking de órgãos e fornecedores vencedores quando DB ou artefato disponível.",
    ]
    if coverage.get("go_no_go_path") == "GO":
        claims_allowed.append("Path proof de contratos PNCP (1d) documentado — não equivale a 90d nacional.")
    claims_forbidden = [
        "Universo completo de editais SC coberto",
        "Piloto nacional 90d de contratos concluído",
        "GO para backfill 3 anos não supervisionado",
        "VPS operacional",
        "CONTRATOS_95 / editais 95%",
        "DOM/CIGA ou DOE smoke como cobertura estadual completa",
        "Valores de contratos como estimado de editais abertos (semântica diferente)",
    ]

    report: dict[str, Any] = {
        "report": "commercial-b2g-session-sc",
        "schema_version": 1,
        "audience": "Extra Construtora / consultoria B2G SC",
        "run_id": run_id,
        "generated_at": generated_at,
        "git_sha": git_sha,
        "confidence": confidence,
        "live_fetch_summary": {
            "sc_compras": sc_block.get("live_fetch"),
            "ciga_dom": dom.get("live_fetch"),
            "doe_dados_sc": doe.get("live_fetch"),
            "db_sample": db.get("live_fetch"),
            "coverage_metrics": False,
            "note": "Flags honestas: métricas de cobertura/ranking/reconciliação são atestação de artefato.",
        },
        "evidence_chain": evidence_chain,
        "coverage": coverage,
        "sc_compras": sc_block,
        "status_changes": status_changes,
        "dom_ciga": dom,
        "doe_sc": doe,
        "org_ranking": org_ranking,
        "reconciliation": recon,
        "db_sample": db,
        "opportunities_open": opportunities.get("open_opportunities") or [],
        "opportunities_recent": opportunities.get("recently_published") or [],
        "claims_allowed": claims_allowed,
        "claims_forbidden": claims_forbidden,
        "disclaimers": disclaimers,
        "key_findings": _key_findings(sc_block, dom, doe, coverage, db, opportunities),
    }
    return report


def _key_findings(
    sc: dict[str, Any],
    dom: dict[str, Any],
    doe: dict[str, Any],
    coverage: dict[str, Any],
    db: dict[str, Any],
    opportunities: dict[str, Any],
) -> list[str]:
    findings: list[str] = []
    if sc.get("available"):
        findings.append(
            f"Compras SC ({sc.get('mode')}): {opportunities.get('open_count', 0)} abertas, "
            f"{opportunities.get('semi_open_count', 0)} semi-abertas, "
            f"{opportunities.get('total_records', 0)} na amostra "
            f"(live_fetch={sc.get('live_fetch')})."
        )
        api_total = ((sc.get("artifact") or {}).get("metrics") or {}).get("api_total_elementos_reported")
        if api_total:
            findings.append(f"API Compras SC reportou {api_total} editais no filtro ano da sessão.")
    if dom.get("available"):
        findings.append(
            f"DOM/CIGA: {dom.get('records_loaded')} publicações; "
            f"procurement-like={dom.get('procurement_like_count')}; "
            f"top act={list((dom.get('act_category_counts') or {}).items())[:3]}."
        )
    if doe.get("available"):
        cats = doe.get("act_categories") or {}
        top = sorted(cats.items(), key=lambda x: -x[1])[:5] if isinstance(cats, dict) else []
        findings.append(f"DOE smoke act_categories top: {top}.")
    if coverage.get("editais_crude_pct") is not None:
        findings.append(
            f"Cobertura bruta editais={coverage.get('editais_crude_pct')}% "
            f"({coverage.get('covered_200km')}/{coverage.get('editais_denominator')} 200km)."
        )
    if db.get("available"):
        findings.append(
            f"DB: contracts_total={db.get('contracts_total')}, contracts_sc={db.get('contracts_sc')}, "
            f"winning_suppliers_sample={len(db.get('winning_suppliers_sc') or [])}."
        )
    failing = (coverage.get("freshness_gate") or {}).get("failing_sources") or []
    if failing:
        findings.append(f"Freshness gate falhando: {failing}.")
    return findings


def write_json(report: dict[str, Any], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(report, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )
    return path


def write_csv_bundle(report: dict[str, Any], path: Path) -> Path:
    """Write multi-section CSV (open opps + acts + suppliers)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    open_opps = report.get("opportunities_open") or []
    recent = report.get("opportunities_recent") or []
    acts = (report.get("dom_ciga") or {}).get("recent_procurement_acts") or []
    suppliers = (report.get("db_sample") or {}).get("winning_suppliers_sc") or []

    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["section", "field", "value"])
        w.writerow(["meta", "run_id", report.get("run_id")])
        w.writerow(["meta", "generated_at", report.get("generated_at")])
        w.writerow(["meta", "confidence", report.get("confidence")])
        w.writerow(["meta", "git_sha", report.get("git_sha")])
        for d in report.get("disclaimers") or []:
            w.writerow(["disclaimer", "text", d])
        for f in report.get("key_findings") or []:
            w.writerow(["finding", "text", f])

        w.writerow([])
        w.writerow(
            [
                "open_opportunity",
                "source_id",
                "status",
                "orgao",
                "municipio",
                "modalidade",
                "objeto",
                "valor_total_estimado",
                "data_publicacao",
                "data_abertura",
                "link_oficial",
            ]
        )
        for r in open_opps:
            w.writerow(
                [
                    "open_opportunity",
                    r.get("source_id"),
                    r.get("status"),
                    r.get("orgao"),
                    r.get("municipio"),
                    r.get("modalidade"),
                    r.get("objeto"),
                    r.get("valor_total_estimado"),
                    r.get("data_publicacao"),
                    r.get("data_abertura"),
                    r.get("link_oficial"),
                ]
            )
        for r in recent:
            w.writerow(
                [
                    "recent_opportunity",
                    r.get("source_id"),
                    r.get("status"),
                    r.get("orgao"),
                    r.get("municipio"),
                    r.get("modalidade"),
                    r.get("objeto"),
                    r.get("valor_total_estimado"),
                    r.get("data_publicacao"),
                    r.get("data_abertura"),
                    r.get("link_oficial"),
                ]
            )
        w.writerow([])
        w.writerow(
            [
                "dom_act",
                "data",
                "municipio",
                "orgao",
                "act_category",
                "act_confidence_label",
                "titulo",
                "url",
            ]
        )
        for a in acts:
            w.writerow(
                [
                    "dom_act",
                    a.get("data"),
                    a.get("municipio"),
                    a.get("orgao"),
                    a.get("act_category"),
                    a.get("act_confidence_label"),
                    a.get("titulo"),
                    a.get("url"),
                ]
            )
        w.writerow([])
        w.writerow(["winning_supplier", "fornecedor_nome", "fornecedor_cnpj", "n", "valor_sum"])
        for s in suppliers:
            w.writerow(
                [
                    "winning_supplier",
                    s.get("fornecedor_nome"),
                    s.get("fornecedor_cnpj"),
                    s.get("n"),
                    s.get("valor_sum"),
                ]
            )
    return path


def write_xlsx(report: dict[str, Any], path: Path) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return None

    wb = Workbook()
    # Meta
    ws = wb.active
    if ws is None:  # pragma: no cover — openpyxl always provides active sheet
        ws = wb.create_sheet("Meta")
    ws.title = "Meta"
    header_fill = PatternFill("solid", fgColor="1B2A3D")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(["campo", "valor"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for k in ("run_id", "generated_at", "confidence", "git_sha", "report", "audience"):
        ws.append([k, report.get(k)])
    ws.append(["live_fetch_sc", (report.get("live_fetch_summary") or {}).get("sc_compras")])
    ws.append(["live_fetch_dom", (report.get("live_fetch_summary") or {}).get("ciga_dom")])
    ws.append(["live_fetch_doe", (report.get("live_fetch_summary") or {}).get("doe_dados_sc")])
    ws.append(["live_fetch_db", (report.get("live_fetch_summary") or {}).get("db_sample")])

    ws2 = wb.create_sheet("Disclaimers")
    ws2.append(["#", "disclaimer"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for i, d in enumerate(report.get("disclaimers") or [], 1):
        ws2.append([i, d])
        ws2.cell(i + 1, 2).alignment = Alignment(wrap_text=True)

    ws3 = wb.create_sheet("Oportunidades_abertas")
    cols = [
        "source_id",
        "status",
        "status_bucket",
        "orgao",
        "municipio",
        "modalidade",
        "objeto",
        "valor_total_estimado",
        "data_publicacao",
        "data_abertura",
        "link_oficial",
        "documentos_count",
    ]
    ws3.append(cols)
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in report.get("opportunities_open") or []:
        ws3.append([r.get(c) for c in cols])

    ws3b = wb.create_sheet("Oportunidades_recentes")
    ws3b.append(cols)
    for cell in ws3b[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in report.get("opportunities_recent") or []:
        ws3b.append([r.get(c) for c in cols])

    ws4 = wb.create_sheet("Atos_DOM")
    act_cols = [
        "data",
        "municipio",
        "orgao",
        "act_category",
        "act_confidence",
        "act_confidence_label",
        "act_needs_human_review",
        "titulo",
        "url",
        "codigo",
    ]
    ws4.append(act_cols)
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
    for a in (report.get("dom_ciga") or {}).get("recent_procurement_acts") or []:
        ws4.append([a.get(c) for c in act_cols])

    ws5 = wb.create_sheet("DOE_sample")
    doe_cols = [
        "data_publicacao",
        "orgao",
        "tipo_ato",
        "categoria",
        "titulo",
        "link_extrato",
        "link_edicao",
    ]
    ws5.append(doe_cols)
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
    for a in (report.get("doe_sc") or {}).get("sample_recent_acts") or []:
        ws5.append([a.get(c) for c in doe_cols])

    ws6 = wb.create_sheet("Orgaos_ranking")
    org_cols = ["orgao", "orgao_cnpj", "uf", "qtd", "valor_total", "valor_semantica", "source_table"]
    ws6.append(org_cols)
    for cell in ws6[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in (report.get("org_ranking") or {}).get("rows") or []:
        ws6.append([r.get(c) for c in org_cols])

    ws7 = wb.create_sheet("Fornecedores")
    sup_cols = ["fornecedor_nome", "fornecedor_cnpj", "n", "valor_sum"]
    ws7.append(sup_cols)
    for cell in ws7[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in (report.get("db_sample") or {}).get("winning_suppliers_sc") or []:
        ws7.append([r.get(c) for c in sup_cols])

    ws8 = wb.create_sheet("Contratos_similares")
    sim_cols = [
        "orgao_nome",
        "fornecedor_nome",
        "objeto",
        "valor_total",
        "data_publicacao",
        "contrato_id",
        "uf",
        "municipio",
    ]
    ws8.append(sim_cols)
    for cell in ws8[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in (report.get("db_sample") or {}).get("similar_contracts_sc") or []:
        ws8.append([r.get(c) for c in sim_cols])

    ws9 = wb.create_sheet("Evidence_chain")
    ws9.append(["source", "run_id", "live_fetch", "attestation", "path", "status", "sha256"])
    for cell in ws9[1]:
        cell.fill = header_fill
        cell.font = header_font
    for e in report.get("evidence_chain") or []:
        ws9.append(
            [
                e.get("source"),
                e.get("run_id"),
                e.get("live_fetch"),
                e.get("attestation"),
                e.get("path"),
                e.get("status"),
                e.get("sha256"),
            ]
        )

    ws10 = wb.create_sheet("Key_findings")
    ws10.append(["#", "finding"])
    for cell in ws10[1]:
        cell.fill = header_fill
        cell.font = header_font
    for i, f in enumerate(report.get("key_findings") or [], 1):
        ws10.append([i, f])
        ws10.cell(i + 1, 2).alignment = Alignment(wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def write_html(report: dict[str, Any], path: Path) -> Path:
    conf = escape(str(report.get("confidence") or ""))
    run_id = escape(str(report.get("run_id") or ""))
    gen = escape(str(report.get("generated_at") or ""))
    findings = report.get("key_findings") or []
    disclaimers = report.get("disclaimers") or []
    open_opps = report.get("opportunities_open") or []
    recent = report.get("opportunities_recent") or []
    acts = (report.get("dom_ciga") or {}).get("recent_procurement_acts") or []
    doe_acts = (report.get("doe_sc") or {}).get("sample_recent_acts") or []
    suppliers = (report.get("db_sample") or {}).get("winning_suppliers_sc") or []
    similar = (report.get("db_sample") or {}).get("similar_contracts_sc") or []
    orgs = (report.get("org_ranking") or {}).get("rows") or []
    lf = report.get("live_fetch_summary") or {}
    cov = report.get("coverage") or {}

    def rows_table(headers: list[str], data: list[dict[str, Any]], keys: list[str]) -> str:
        th = "".join(f"<th>{escape(h)}</th>" for h in headers)
        body = []
        for r in data:
            tds = "".join(f"<td>{escape(str(r.get(k) if r.get(k) is not None else '—'))}</td>" for k in keys)
            body.append(f"<tr>{tds}</tr>")
        if not body:
            body.append(f"<tr><td colspan='{len(headers)}'><em>Sem dados nesta seção</em></td></tr>")
        return f"<table><thead><tr>{th}</tr></thead><tbody>{''.join(body)}</tbody></table>"

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Relatório Comercial B2G SC — {run_id}</title>
  <style>
    :root {{ --ink:#1B2A3D; --bg:#F5F6F8; --card:#fff; --amber:#B8860B; --red:#B5342A; --green:#1B7A3D; --link:#0563C1; }}
    body {{ font-family: "Segoe UI", system-ui, sans-serif; margin:0; background:var(--bg); color:var(--ink); line-height:1.45; }}
    header {{ background:var(--ink); color:#fff; padding:1.5rem 2rem; }}
    header h1 {{ margin:0 0 .4rem; font-size:1.4rem; }}
    header .meta {{ opacity:.85; font-size:.9rem; }}
    main {{ max-width:1100px; margin:0 auto; padding:1.25rem; }}
    section {{ background:var(--card); border-radius:8px; padding:1rem 1.25rem; margin-bottom:1rem; box-shadow:0 1px 3px rgba(0,0,0,.06); }}
    h2 {{ margin-top:0; font-size:1.1rem; border-bottom:2px solid var(--bg); padding-bottom:.4rem; }}
    .badge {{ display:inline-block; padding:.15rem .55rem; border-radius:999px; font-size:.75rem; font-weight:600; background:#e8eef5; }}
    .badge.low {{ background:#fde8e6; color:var(--red); }}
    .badge.low_to_medium {{ background:#fff3cd; color:var(--amber); }}
    .badge.medium {{ background:#e6f4ea; color:var(--green); }}
    ul.disc li {{ margin:.35rem 0; }}
    table {{ width:100%; border-collapse:collapse; font-size:.82rem; }}
    th, td {{ border-bottom:1px solid #e5e7eb; padding:.4rem .45rem; text-align:left; vertical-align:top; }}
    th {{ background:var(--ink); color:#fff; position:sticky; top:0; }}
    tr:nth-child(even) td {{ background:#fafbfc; }}
    .flags span {{ display:inline-block; margin-right:.6rem; font-size:.85rem; }}
    .ok {{ color:var(--green); }} .no {{ color:var(--red); }}
    a {{ color:var(--link); }}
    footer {{ font-size:.8rem; color:#555; padding:1rem 0 2rem; text-align:center; }}
    .warn {{ border-left:4px solid var(--amber); padding-left:.75rem; }}
  </style>
</head>
<body>
<header>
  <h1>Relatório Comercial B2G — Santa Catarina</h1>
  <div class="meta">
    run_id=<code>{run_id}</code> · gerado={gen} · confidence=
    <span class="badge {escape(str(report.get('confidence') or ''))}">{conf}</span>
    · git={escape(str(report.get('git_sha') or 'unknown'))}
  </div>
</header>
<main>
<section>
  <h2>Flags live_fetch vs atestação</h2>
  <div class="flags">
    <span>sc_compras: <strong class="{'ok' if lf.get('sc_compras') else 'no'}">{lf.get('sc_compras')}</strong></span>
    <span>ciga_dom: <strong class="{'ok' if lf.get('ciga_dom') else 'no'}">{lf.get('ciga_dom')}</strong></span>
    <span>doe: <strong class="{'ok' if lf.get('doe_dados_sc') else 'no'}">{lf.get('doe_dados_sc')}</strong></span>
    <span>db: <strong class="{'ok' if lf.get('db_sample') else 'no'}">{lf.get('db_sample')}</strong></span>
    <span>coverage_metrics: <strong class="no">False (attestation)</strong></span>
  </div>
  <p class="warn">{escape(str(lf.get('note') or ''))}</p>
</section>

<section>
  <h2>Key findings</h2>
  <ul class="disc">
    {''.join(f'<li>{escape(str(f))}</li>' for f in findings) or '<li><em>Nenhum finding</em></li>'}
  </ul>
</section>

<section>
  <h2>Cobertura & freshness</h2>
  <ul class="disc">
    <li>editais_crude_pct: <strong>{escape(str(cov.get('editais_crude_pct')))}</strong>
        ({escape(str(cov.get('covered_200km')))}/{escape(str(cov.get('editais_denominator')))} 200km)</li>
    <li>pilot_status: {escape(str(cov.get('pilot_status')))} · go_no_go_3y={escape(str(cov.get('go_no_go_3y')))}
        · go_no_go_path={escape(str(cov.get('go_no_go_path')))}</li>
    <li>freshness failing: {escape(str((cov.get('freshness_gate') or {}).get('failing_sources')))}</li>
    <li>monitoring (coverage-truth): {escape(str((cov.get('coverage_truth') or {}).get('monitoring_pct_display')))}</li>
  </ul>
</section>

<section>
  <h2>Oportunidades abertas (Compras SC)</h2>
  {rows_table(
      ['ID','Status','Órgão','Modalidade','Publicação','Abertura','Objeto','Link'],
      open_opps,
      ['source_id','status','orgao','modalidade','data_publicacao','data_abertura','objeto','link_oficial'],
  )}
</section>

<section>
  <h2>Publicadas recentemente (amostra)</h2>
  {rows_table(
      ['ID','Status','Órgão','Modalidade','Publicação','Objeto','Link'],
      recent[:20],
      ['source_id','status','orgao','modalidade','data_publicacao','objeto','link_oficial'],
  )}
</section>

<section>
  <h2>Atos recentes DOM/CIGA (procurement-like)</h2>
  {rows_table(
      ['Data','Município','Órgão','Categoria','Conf.','Título'],
      acts[:25],
      ['data','municipio','orgao','act_category','act_confidence_label','titulo'],
  )}
</section>

<section>
  <h2>Atos DOE (sample smoke)</h2>
  {rows_table(
      ['Data','Órgão','Tipo','Título'],
      doe_acts,
      ['data_publicacao','orgao','tipo_ato','titulo'],
  )}
</section>

<section>
  <h2>Ranking de órgãos (artefato / contratos)</h2>
  {rows_table(
      ['Órgão','CNPJ','UF','Qtd','Valor','Semântica'],
      orgs[:20],
      ['orgao','orgao_cnpj','uf','qtd','valor_total','valor_semantica'],
  )}
</section>

<section>
  <h2>Fornecedores vencedores (DB SC)</h2>
  {rows_table(
      ['Fornecedor','CNPJ','N','Valor sum'],
      suppliers,
      ['fornecedor_nome','fornecedor_cnpj','n','valor_sum'],
  )}
</section>

<section>
  <h2>Contratos similares (DB, keywords engenharia/obra)</h2>
  {rows_table(
      ['Órgão','Fornecedor','Valor','Data','Objeto'],
      similar,
      ['orgao_nome','fornecedor_nome','valor_total','data_publicacao','objeto'],
  )}
</section>

<section>
  <h2>Disclaimers automáticos</h2>
  <ul class="disc warn">
    {''.join(f'<li>{escape(str(d))}</li>' for d in disclaimers)}
  </ul>
</section>

<section>
  <h2>Evidence chain (run_ids)</h2>
  {rows_table(
      ['Source','run_id','live_fetch','attestation','path','status'],
      report.get('evidence_chain') or [],
      ['source','run_id','live_fetch','attestation','path','status'],
  )}
</section>
</main>
<footer>
  Extra Consultoria · relatório honesto artifact-first · não constitui GO comercial 95% · {run_id}
</footer>
</body>
</html>
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html, encoding="utf-8")
    return path


def write_all(
    report: dict[str, Any],
    *,
    output_dir: Path,
    basename: str = "commercial-b2g-session-sc",
) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths: dict[str, str] = {}
    jp = write_json(report, output_dir / f"{basename}.json")
    paths["json"] = str(jp)
    cp = write_csv_bundle(report, output_dir / f"{basename}.csv")
    paths["csv"] = str(cp)
    xp = write_xlsx(report, output_dir / f"{basename}.xlsx")
    if xp:
        paths["xlsx"] = str(xp)
    hp = write_html(report, output_dir / f"{basename}.html")
    paths["html"] = str(hp)
    # sidecar meta for evidence
    meta = {
        "run_id": report.get("run_id"),
        "generated_at": report.get("generated_at"),
        "git_sha": report.get("git_sha"),
        "confidence": report.get("confidence"),
        "artifacts": paths,
        "live_fetch_summary": report.get("live_fetch_summary"),
        "evidence_chain": report.get("evidence_chain"),
        "disclaimers_count": len(report.get("disclaimers") or []),
    }
    mp = output_dir / f"{basename}.meta.json"
    mp.write_text(json.dumps(meta, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    paths["meta"] = str(mp)
    return paths


def main() -> int:
    ap = argparse.ArgumentParser(description="Commercial B2G session report (honest)")
    ap.add_argument(
        "--output-dir",
        default="output/reports",
        help="Directory for JSON/CSV/XLSX/HTML",
    )
    ap.add_argument(
        "--basename",
        default="commercial-b2g-session-sc",
        help="Output file basename",
    )
    ap.add_argument(
        "--dsn",
        default=os.getenv("DATABASE_URL") or os.getenv("LOCAL_DATALAKE_DSN"),
        help="Optional PostgreSQL DSN",
    )
    args = ap.parse_args()
    out_dir = Path(args.output_dir)
    if not out_dir.is_absolute():
        out_dir = _PROJECT_ROOT / out_dir
    report = build_session_report(dsn=args.dsn)
    paths = write_all(report, output_dir=out_dir, basename=args.basename)
    print(
        json.dumps(
            {
                "wrote": paths,
                "run_id": report["run_id"],
                "confidence": report["confidence"],
                "disclaimers": len(report["disclaimers"]),
                "open_opportunities": len(report.get("opportunities_open") or []),
                "key_findings": report.get("key_findings"),
            },
            indent=2,
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
