#!/usr/bin/env python3
"""Market Panorama Report — Extra Consultoria.

Gera relatório de panorama do mercado de licitações para SC com foco em
engenharia civil. Output: terminal + Excel + PDF.

Usage:
    python scripts/reports/panorama.py
    python scripts/reports/panorama.py --setor engenharia --uf SC --dias 90
    python scripts/reports/panorama.py --monthly
    python scripts/reports/panorama.py --output-pdf
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

DSN = os.getenv(
    "LOCAL_DATALAKE_DSN",
    "postgresql://postgres:smartlic_local@127.0.0.1:54399/postgres",
)


def get_conn():
    import psycopg2

    return psycopg2.connect(DSN)


def query(conn, sql: str, params: list = None) -> list[dict]:
    cur = conn.cursor()
    cur.execute(sql, params)
    cols = [d[0] for d in cur.description] if cur.description else []
    rows = [dict(zip(cols, row)) for row in cur.fetchall()]
    cur.close()
    return rows


# ---------------------------------------------------------------------------
# Report sections
# ---------------------------------------------------------------------------


def section_volume(conn, uf: str = "SC", dias: int = 90) -> dict:
    """Total bid volume and value by modality."""
    return query(
        conn,
        """SELECT
             modalidade_nome,
             COUNT(*) AS total_bids,
             SUM(valor_total_estimado) AS valor_total,
             AVG(valor_total_estimado) AS ticket_medio,
             MIN(data_publicacao) AS primeira,
             MAX(data_publicacao) AS ultima
           FROM pncp_raw_bids
           WHERE is_active = TRUE
             AND uf = %s
             AND data_publicacao >= CURRENT_DATE - %s
             AND valor_total_estimado > 0
           GROUP BY modalidade_nome
           ORDER BY total_bids DESC""",
        (uf, dias),
    )


def section_municipios(conn, uf: str = "SC", dias: int = 90, limit: int = 20) -> list[dict]:
    """Top municipalities by bid count."""
    return query(
        conn,
        """SELECT
             municipio,
             COUNT(*) AS total_bids,
             SUM(valor_total_estimado) AS valor_total,
             COUNT(DISTINCT orgao_cnpj) AS orgaos
           FROM pncp_raw_bids
           WHERE is_active = TRUE
             AND uf = %s
             AND data_publicacao >= CURRENT_DATE - %s
             AND municipio IS NOT NULL
           GROUP BY municipio
           ORDER BY total_bids DESC
           LIMIT %s""",
        (uf, dias, limit),
    )


def section_orgaos(conn, uf: str = "SC", dias: int = 90, limit: int = 20) -> list[dict]:
    """Top contracting authorities."""
    return query(
        conn,
        """SELECT
             orgao_razao_social,
             orgao_cnpj,
             COUNT(*) AS total_bids,
             SUM(valor_total_estimado) AS valor_total
           FROM pncp_raw_bids
           WHERE is_active = TRUE
             AND uf = %s
             AND data_publicacao >= CURRENT_DATE - %s
             AND orgao_razao_social IS NOT NULL
           GROUP BY orgao_razao_social, orgao_cnpj
           ORDER BY total_bids DESC
           LIMIT %s""",
        (uf, dias, limit),
    )


def section_sazonalidade(conn, uf: str = "SC", meses: int = 12) -> list[dict]:
    """Monthly bid distribution (seasonal analysis)."""
    return query(
        conn,
        """SELECT
             EXTRACT(YEAR FROM data_publicacao)::INT AS ano,
             EXTRACT(MONTH FROM data_publicacao)::INT AS mes,
             COUNT(*) AS total_bids,
             SUM(valor_total_estimado) AS valor_total
           FROM pncp_raw_bids
           WHERE is_active = TRUE
             AND uf = %s
             AND data_publicacao >= CURRENT_DATE - INTERVAL '%s months'
           GROUP BY ano, mes
           ORDER BY ano, mes""",
        (uf, meses),
    )


def section_source_distribution(conn, uf: str = "SC", dias: int = 90) -> list[dict]:
    """Distribution by data source."""
    return query(
        conn,
        """SELECT
             source,
             COUNT(*) AS total_bids,
             COUNT(DISTINCT matched_entity_id) AS entities_covered
           FROM pncp_raw_bids
           WHERE is_active = TRUE
             AND uf = %s
             AND data_publicacao >= CURRENT_DATE - %s
           GROUP BY source
           ORDER BY total_bids DESC""",
        (uf, dias),
    )


def section_coverage_gaps(conn) -> list[dict]:
    """Entities with NO coverage (most critical for monitoring)."""
    return query(
        conn,
        """SELECT tue.legal_name AS razao_social, tue.cnpj8 AS cnpj_8,
                  tue.municipality AS municipio, tue.legal_nature AS natureza_juridica,
                  tue.radius_decision, tue.distance_km AS distancia_fk
           FROM target_universe_entities tue
           WHERE tue.universe_run_id = (SELECT MAX(id) FROM target_universe_runs)
             AND tue.radius_decision = 'included'
             AND tue.db_entity_id NOT IN (
                 SELECT entity_id FROM entity_coverage WHERE is_covered = TRUE
             )
           ORDER BY tue.municipality, tue.legal_name""",
    )


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------


def print_terminal_report(sections: dict) -> None:
    """Print panorama to terminal."""
    print("\n" + "=" * 72)
    print("  PANORAMA DE LICITAÇÕES — Extra Construtora")
    print(f"  {date.today().strftime('%d/%m/%Y')}")
    print("=" * 72)

    # Volume
    print("\n── VOLUME POR MODALIDADE ──")
    print(f"{'Modalidade':25s} {'Qtd':>5s} {'Valor Total':>18s} {'Ticket Médio':>16s}")
    print("-" * 65)
    for r in sections.get("volume", []):
        print(
            f"{r['modalidade_nome'] or 'N/A':25s} {r['total_bids']:>5d} "
            f"R$ {r['valor_total'] or 0:>15,.0f} "
            f"R$ {r['ticket_medio'] or 0:>13,.0f}"
        )

    # Top municípios
    print("\n── TOP 10 MUNICÍPIOS ──")
    print(f"{'Município':25s} {'Licitações':>10s} {'Valor Total':>18s} {'Órgãos':>7s}")
    print("-" * 62)
    for r in sections.get("municipios", [])[:10]:
        print(f"{r['municipio'][:25]:25s} {r['total_bids']:>10d} R$ {r['valor_total'] or 0:>15,.0f} {r['orgaos']:>7d}")

    # Sazonalidade
    print("\n── SAZONALIDADE (últimos 12 meses) ──")
    meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    for r in sections.get("sazonalidade", []):
        bar = "█" * max(1, int(r["total_bids"] / 10))
        print(f"  {meses[int(r['mes']) - 1]:3s} {r['ano']}: {bar} ({r['total_bids']})")

    # Source distribution
    print("\n── DISTRIBUIÇÃO POR FONTE ──")
    for r in sections.get("sources", []):
        print(f"  {r['source']:15s}: {r['total_bids']:>5d} bids, {r['entities_covered']:>4d} entidades")

    # Coverage gaps
    gaps = sections.get("gaps", [])
    if gaps:
        print(f"\n── 🚨 ENTIDADES SEM COBERTURA (raio 200km): {len(gaps)} ──")
        for g in gaps[:10]:
            print(f"  • {g['razao_social'][:50]:50s} | {g['municipio'] or 'N/A':20s}")
        if len(gaps) > 10:
            print(f"  ... e mais {len(gaps) - 10}")

    print("\n" + "=" * 72)


def export_excel(sections: dict, filepath: str) -> None:
    """Export panorama sections to styled Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        print("⚠️  openpyxl not installed, skipping Excel export")
        return

    wb = openpyxl.Workbook()

    # -- Volume sheet --
    ws = wb.active
    ws.title = "Volume"
    _write_sheet(
        ws,
        sections.get("volume", []),
        ["Modalidade", "Qtd", "Valor Total", "Ticket Médio", "Período"],
        ["modalidade_nome", "total_bids", "valor_total", "ticket_medio"],
    )

    # -- Municípios sheet --
    ws2 = wb.create_sheet("Municípios")
    _write_sheet(
        ws2,
        sections.get("municipios", []),
        ["Município", "Licitações", "Valor Total", "Órgãos"],
        ["municipio", "total_bids", "valor_total", "orgaos"],
    )

    # -- Sazonalidade sheet --
    ws3 = wb.create_sheet("Sazonalidade")
    _write_sheet(
        ws3,
        sections.get("sazonalidade", []),
        ["Ano", "Mês", "Total Bids", "Valor Total"],
        ["ano", "mes", "total_bids", "valor_total"],
    )

    # -- Gaps sheet --
    if sections.get("gaps"):
        ws4 = wb.create_sheet("Gaps")
        _write_sheet(
            ws4,
            sections["gaps"],
            ["Órgão", "CNPJ", "Município", "Natureza Jurídica", "Raio 200km"],
            ["razao_social", "cnpj_8", "municipio", "natureza_juridica", "raio_200km"],
        )

    # Style
    header_fill = PatternFill(start_color="1B2A3D", end_color="1B2A3D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        bottom=Side(style="thin", color="8B7355"),
    )

    for ws_obj in [ws, ws2, ws3]:
        for cell in ws_obj[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in ws_obj.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border

    wb.save(filepath)
    print(f"✅ Excel salvo: {filepath}")


def _write_sheet(ws, data: list[dict], headers: list[str], keys: list[str]):
    """Write data to a worksheet with headers."""
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, record in enumerate(data, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(key))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def parse_args():
    p = argparse.ArgumentParser(description="Market Panorama Report")
    p.add_argument("--setor", default="engenharia", help="Sector filter")
    p.add_argument("--uf", default="SC", help="UF filter")
    p.add_argument("--dias", type=int, default=90, help="Days window")
    p.add_argument("--monthly", action="store_true", help="Monthly mode (12 months)")
    p.add_argument("--output-pdf", action="store_true", help="Generate PDF")
    p.add_argument("--output-excel", action="store_true", help="Generate Excel")
    p.add_argument("--dsn", default=DSN, help="PostgreSQL DSN")
    return p.parse_args()


def main():
    args = parse_args()
    global DSN
    DSN = args.dsn

    conn = get_conn()
    try:
        dias = 365 if args.monthly else args.dias

        sections = {
            "volume": section_volume(conn, args.uf, dias),
            "municipios": section_municipios(conn, args.uf, dias),
            "orgaos": section_orgaos(conn, args.uf, dias),
            "sazonalidade": section_sazonalidade(conn, args.uf, 12),
            "sources": section_source_distribution(conn, args.uf, dias),
            "gaps": section_coverage_gaps(conn),
        }
    finally:
        conn.close()

    # Always print terminal report
    print_terminal_report(sections)

    # Excel export
    if args.output_excel:
        output_dir = _PROJECT_ROOT / "output" / "excels"
        output_dir.mkdir(parents=True, exist_ok=True)
        filepath = output_dir / f"panorama-{args.uf}-{date.today().isoformat()}.xlsx"
        export_excel(sections, str(filepath))

    # PDF export
    if args.output_pdf:
        print("⚠️  PDF export via reportlab — integração com generate_report_b2g.py pendente")

    return 0


if __name__ == "__main__":
    sys.exit(main())
