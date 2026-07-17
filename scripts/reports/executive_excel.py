#!/usr/bin/env python3
"""
Relatorio Executivo — Extra Construtora (Excel).

Gera planilha Excel multi-sheet com dados rastreaveis do banco de dados,
incluindo campos de proveniencia (source, source_id, source_url, content_hash)
para auditoria e rastreabilidade.

Sheets:
  - Editais: todas as oportunidades com colunas para rastreabilidade
  - Orgaos: ranking de orgaos com contagens
  - Concorrentes: orgaos que mais contratam (da pncp_raw_bids)
  - Metadados: data de geracao, fontes, freshness, DSN usado
  - Legenda: explicacao de cada coluna, ranking, score

Design: mesmo padrao visual do intel_excel.py (charcoal navy header, subtil shading).

Usage:
    python scripts/reports/executive_excel.py
    python scripts/reports/executive_excel.py --output output/reports/executivo-extra-2026-07-15.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import UTC, date, datetime
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

DSN = os.getenv(
    "LOCAL_DATALAKE_DSN",
    "postgresql://test:test@127.0.0.1:5433/pncp_datalake",
)

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("ERROR: psycopg2 not installed. Run: pip install psycopg2-binary")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ============================================================
# Design tokens (matching intel_excel.py)
# ============================================================

INK = "1B2A3D"
BG_SUBTLE = "F5F6F8"
GREEN_TEXT = "1B7A3D"
RED_TEXT = "B5342A"
AMBER_TEXT = "B8860B"
BLUE_TEXT = "1565C0"
GRAY_TEXT = "808080"
LINK_BLUE = "0563C1"

CURRENCY_FMT = "[R$-416] #,##0.00"
DATE_FMT = "DD/MM/YYYY"
DATETIME_FMT = "DD/MM/YYYY HH:MM"
PCT_FMT = "0.0%"

_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")

# ============================================================
# Column definitions for Editais sheet
# ============================================================

EDITAIS_COLUMNS = [
    ("ID", 6, "center"),
    ("Ranking", 8, "center"),
    ("Score", 6, "center"),
    ("Confianca", 10, "center"),
    ("Source", 12, "left"),
    ("Source ID", 32, "left"),
    ("Source URL", 12, "center"),
    ("Content Hash", 28, "left"),
    ("Orgao", 35, "left"),
    ("UF", 5, "center"),
    ("Municipio", 20, "left"),
    ("Modalidade", 22, "left"),
    ("Objeto", 55, "left"),
    ("Valor Estimado", 18, "right"),
    ("Valor Homologado", 18, "right"),
    ("Data Publicacao", 14, "center"),
    ("Data Abertura", 14, "center"),
    ("Data Encerramento", 14, "center"),
    ("Status Canonico", 12, "center"),
    ("Qualidade Score", 12, "center"),
    ("Ranking Regras", 40, "left"),
    ("Fatores Positivos", 40, "left"),
    ("Fatores Negativos", 40, "left"),
    ("Bloqueadores", 40, "left"),
    ("Link", 12, "center"),
]

ORGAOS_COLUMNS = [
    ("Orgao", 45, "left"),
    ("UF", 5, "center"),
    ("Editais", 10, "center"),
    ("Classif. GO", 12, "center"),
    ("Valor Est. Total", 22, "right"),
    ("Valor Hom. Total", 22, "right"),
]

CONCORRENTES_COLUMNS = [
    ("Orgao", 45, "left"),
    ("UF", 5, "center"),
    ("Municipio", 22, "left"),
    ("Qtd Vencidas", 14, "center"),
    ("Ticket Medio", 22, "right"),
    ("Total Contratado", 22, "right"),
]

METADADOS_LABELS = [
    "DSN",
    "Data de Geracao",
    "Hora de Geracao",
    "Fontes de Dados",
    "Tabela Principal",
    "Total Oportunidades (ativas)",
    "Total GO",
    "Total NO_GO",
    "Total REVIEW",
    "Score Medio GO",
    "Valor Estimado Total",
    "Orgaos Ativos (SC)",
    "Confianca HIGH",
    "Confianca MEDIUM",
    "Confianca LOW",
    "pncp_supplier_contracts (vincendos 180d)",
    "pncp_raw_bids (total registros)",
    "Data da Ultima Atualizacao no DB",
    "Proxima Atualizacao Recomendada",
    # Run metadata (shared with PDF via scripts/reports/run_metadata.py)
    "Run ID",
    "Generated At (UTC)",
    "Git SHA",
    "Profile ID",
    "Filter UF",
    "Filter is_active",
    "Filter table_primary",
    "Filter vincendos_horizon_days",
    "Cutoff as_of_date",
    "Cutoff data_window",
    "Cutoff ultima_atualizacao_db",
    "Sample label",
    "Sample opportunity_intel_active",
    "Sample ranking_go",
    "Sample ranking_review",
    "Sample ranking_no_go",
    "Sample orgaos_sc",
    "Sample vincendos_180d",
    "Sample pncp_raw_bids_active",
]

# ============================================================
# DB layer
# ============================================================


def get_conn():
    return psycopg2.connect(DSN)


def query(conn, sql: str, params: list = None) -> list[dict]:
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    return rows


def scalar(conn, sql: str, params: list = None):
    cur = conn.cursor()
    cur.execute(sql, params)
    val = cur.fetchone()
    cur.close()
    return val[0] if val else None


# ============================================================
# Data loads
# ============================================================


def load_editais(conn) -> list[dict]:
    return query(
        conn,
        """SELECT id, source, source_id, source_url, content_hash,
                  orgao_nome, uf, municipio, modalidade, objeto,
                  valor_estimado, valor_homologado,
                  data_publicacao, data_abertura, data_encerramento,
                  ranking, ranking_score, ranking_confianca,
                  ranking_fatores, ranking_regras,
                  qualidade_score, qualidade_fatores,
                  link_edital, numero_controle_pncp, status_canonico
           FROM opportunity_intel
           WHERE is_active = true
           ORDER BY ranking, ranking_score DESC""",
    )


def load_orgaos(conn) -> list[dict]:
    return query(
        conn,
        """SELECT orgao_nome, uf,
                  COUNT(*) AS qtd_editais,
                  COUNT(*) FILTER (WHERE ranking = 'GO') AS qtd_go,
                  ROUND(SUM(valor_estimado)::numeric, 2) AS total_valor_estimado,
                  ROUND(SUM(valor_homologado)::numeric, 2) AS total_valor_homologado
           FROM opportunity_intel
           WHERE is_active = true AND orgao_nome IS NOT NULL
           GROUP BY orgao_nome, uf
           ORDER BY qtd_editais DESC""",
    )


def load_concorrentes(conn) -> list[dict]:
    return query(
        conn,
        """SELECT orgao_razao_social, uf, municipio,
                  COUNT(*) AS qtd_vencidas,
                  ROUND(AVG(valor_total_estimado)::numeric, 2) AS ticket_medio,
                  ROUND(SUM(valor_total_estimado)::numeric, 2) AS total_contratado
           FROM pncp_raw_bids
           WHERE is_active = true
           GROUP BY orgao_razao_social, uf, municipio
           ORDER BY total_contratado DESC
           LIMIT 50""",
    )


# ============================================================
# Helpers
# ============================================================


def _sanitize(value) -> str:
    if value is None:
        return ""
    s = str(value)
    return _ILLEGAL_XML_RE.sub("", s)


def _safe_float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _safe_int(value) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _parse_dt(value) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.replace(tzinfo=None)
    except (ValueError, AttributeError):
        pass
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _format_cnpj(cnpj: str) -> str:
    d = re.sub(r"\D", "", str(cnpj))
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    return str(cnpj)


def _ranking_fatores_json(record: dict) -> dict:
    fatores = record.get("ranking_fatores")
    if fatores is None:
        return {"positivos": [], "negativos": [], "bloqueadores": []}
    if isinstance(fatores, str):
        import json

        try:
            return json.loads(fatores)
        except (json.JSONDecodeError, TypeError):
            return {"positivos": [], "negativos": [], "bloqueadores": []}
    if isinstance(fatores, dict):
        return fatores
    return {"positivos": [], "negativos": [], "bloqueadores": []}


# ============================================================
# Styles factory
# ============================================================


def _make_styles():
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=INK, end_color=INK, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    alt_fill = PatternFill(start_color=BG_SUBTLE, end_color=BG_SUBTLE, fill_type="solid")
    no_fill = PatternFill(fill_type=None)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    return {
        "header_font": header_font,
        "header_fill": header_fill,
        "header_align": header_align,
        "alt_fill": alt_fill,
        "no_fill": no_fill,
        "thin_border": thin_border,
        "green_font": Font(color=GREEN_TEXT, bold=True),
        "red_font": Font(color=RED_TEXT, bold=True),
        "amber_font": Font(color=AMBER_TEXT, bold=True),
        "gray_font": Font(color=GRAY_TEXT),
        "link_font": Font(color=LINK_BLUE, underline="single"),
        "bold_font": Font(bold=True),
        "label_font": Font(bold=True, color=INK),
        "label_fill": PatternFill(start_color="E8EBF0", end_color="E8EBF0", fill_type="solid"),
    }


def _woc(ws, value=None, font=None, fill=None, alignment=None, border=None, number_format=None):
    c = WriteOnlyCell(ws, value=value)
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if alignment is not None:
        c.alignment = alignment
    if border is not None:
        c.border = border
    if number_format is not None:
        c.number_format = number_format
    return c


# ============================================================
# Sheet builders
# ============================================================


def _build_editais(wb: Workbook, items: list[dict], st: dict):
    ws = wb.create_sheet("Editais")

    for col_idx, (_, width, _) in enumerate(EDITAIS_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # Header row
    header_row = []
    for header, _, _ in EDITAIS_COLUMNS:
        header_row.append(
            _woc(
                ws,
                value=header,
                font=st["header_font"],
                fill=st["header_fill"],
                alignment=st["header_align"],
                border=st["thin_border"],
            )
        )
    ws.append(header_row)

    # Data rows
    for row_num, item in enumerate(items, start=2):
        is_alt = (row_num % 2) == 0
        row_fill = st["alt_fill"] if is_alt else st["no_fill"]

        fatores = _ranking_fatores_json(item)
        positivos = "; ".join(fatores.get("positivos", [])[:5])
        negativos = "; ".join(fatores.get("negativos", [])[:5])
        bloqueadores = "; ".join(fatores.get("bloqueadores", [])[:5])

        regras = item.get("ranking_regras", [])
        if isinstance(regras, (list, tuple)):
            regras_str = "; ".join(str(r) for r in regras[:10])
        else:
            regras_str = str(regras)[:300]

        ranking = item.get("ranking", "")
        ranking_font = (
            st["green_font"] if ranking == "GO" else st["red_font"] if ranking == "NO_GO" else st["amber_font"]
        )

        # Valor columns
        valor_est = _safe_float(item.get("valor_estimado"))
        valor_hom = _safe_float(item.get("valor_homologado"))

        dt_pub = _parse_dt(item.get("data_publicacao"))
        dt_ab = _parse_dt(item.get("data_abertura"))
        dt_enc = _parse_dt(item.get("data_encerramento"))

        link = item.get("link_edital", "")
        if link:
            safe_url = str(link).replace('"', "%22")
            link_val = f'=HYPERLINK("{safe_url}","Abrir")'
        else:
            link_val = ""

        def _dc(col_idx, value=None, font=None, fill=None, number_format=None):
            _, _, h_align = EDITAIS_COLUMNS[col_idx - 1]
            eff_fill = fill if fill is not None else row_fill
            return _woc(
                ws,
                value=value,
                font=font,
                fill=eff_fill,
                alignment=Alignment(horizontal=h_align, vertical="top", wrap_text=True),
                border=st["thin_border"],
                number_format=number_format,
            )

        row = [
            _dc(1, value=item.get("id")),
            _dc(2, value=ranking, font=ranking_font),
            _dc(3, value=item.get("ranking_score")),
            _dc(4, value=item.get("ranking_confianca", "")),
            _dc(5, value=item.get("source", "")),
            _dc(6, value=item.get("source_id", "")),
            _dc(7, value="Link" if link else ""),
            _dc(8, value=item.get("content_hash", "")),
            _dc(9, value=_sanitize(item.get("orgao_nome", ""))),
            _dc(10, value=item.get("uf", "")),
            _dc(11, value=_sanitize(item.get("municipio", ""))),
            _dc(12, value=_sanitize(item.get("modalidade", ""))),
            _dc(13, value=_sanitize(item.get("objeto", ""))),
            _dc(14, value=valor_est, number_format=CURRENCY_FMT if valor_est else None),
            _dc(15, value=valor_hom, number_format=CURRENCY_FMT if valor_hom else None),
            _dc(16, value=dt_pub, number_format=DATE_FMT if dt_pub else None),
            _dc(17, value=dt_ab, number_format=DATETIME_FMT if dt_ab else None),
            _dc(18, value=dt_enc, number_format=DATETIME_FMT if dt_enc else None),
            _dc(19, value=item.get("status_canonico", "")),
            _dc(20, value=item.get("qualidade_score")),
            _dc(21, value=regras_str),
            _dc(22, value=positivos),
            _dc(23, value=negativos),
            _dc(24, value=bloqueadores),
            _dc(25, value=link_val, font=st["link_font"]),
        ]
        ws.append(row)


def _build_orgaos(wb: Workbook, items: list[dict], st: dict):
    ws = wb.create_sheet("Orgaos")

    for col_idx, (_, width, _) in enumerate(ORGAOS_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    header_row = []
    for header, _, _ in ORGAOS_COLUMNS:
        header_row.append(
            _woc(
                ws,
                value=header,
                font=st["header_font"],
                fill=st["header_fill"],
                alignment=st["header_align"],
                border=st["thin_border"],
            )
        )
    ws.append(header_row)

    for row_num, item in enumerate(items, start=2):
        is_alt = (row_num % 2) == 0
        row_fill = st["alt_fill"] if is_alt else st["no_fill"]

        def _dc(col_idx, value=None, font=None, fill=None, number_format=None):
            _, _, h_align = ORGAOS_COLUMNS[col_idx - 1]
            eff_fill = fill if fill is not None else row_fill
            return _woc(
                ws,
                value=value,
                font=font,
                fill=eff_fill,
                alignment=Alignment(horizontal=h_align, vertical="top", wrap_text=True),
                border=st["thin_border"],
                number_format=number_format,
            )

        row = [
            _dc(1, value=_sanitize(item.get("orgao_nome", ""))),
            _dc(2, value=item.get("uf", "")),
            _dc(3, value=item.get("qtd_editais", 0)),
            _dc(4, value=item.get("qtd_go", 0)),
            _dc(5, value=_safe_float(item.get("total_valor_estimado")), number_format=CURRENCY_FMT),
            _dc(6, value=_safe_float(item.get("total_valor_homologado")), number_format=CURRENCY_FMT),
        ]
        ws.append(row)


def _build_concorrentes(wb: Workbook, items: list[dict], st: dict):
    ws = wb.create_sheet("Concorrentes")

    for col_idx, (_, width, _) in enumerate(CONCORRENTES_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    header_row = []
    for header, _, _ in CONCORRENTES_COLUMNS:
        header_row.append(
            _woc(
                ws,
                value=header,
                font=st["header_font"],
                fill=st["header_fill"],
                alignment=st["header_align"],
                border=st["thin_border"],
            )
        )
    ws.append(header_row)

    for row_num, item in enumerate(items, start=2):
        is_alt = (row_num % 2) == 0
        row_fill = st["alt_fill"] if is_alt else st["no_fill"]

        def _dc(col_idx, value=None, font=None, fill=None, number_format=None):
            _, _, h_align = CONCORRENTES_COLUMNS[col_idx - 1]
            eff_fill = fill if fill is not None else row_fill
            return _woc(
                ws,
                value=value,
                font=font,
                fill=eff_fill,
                alignment=Alignment(horizontal=h_align, vertical="top", wrap_text=True),
                border=st["thin_border"],
                number_format=number_format,
            )

        row = [
            _dc(1, value=_sanitize(item.get("orgao_razao_social", ""))),
            _dc(2, value=item.get("uf", "")),
            _dc(3, value=_sanitize(item.get("municipio", ""))),
            _dc(4, value=item.get("qtd_vencidas", 0)),
            _dc(5, value=_safe_float(item.get("ticket_medio")), number_format=CURRENCY_FMT),
            _dc(6, value=_safe_float(item.get("total_contratado")), number_format=CURRENCY_FMT),
        ]
        ws.append(row)


def _build_metadados(wb: Workbook, stats: dict, st: dict, run_meta: dict | None = None):
    ws = wb.create_sheet("Metadados")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 60

    now = datetime.now(UTC)
    filters = (run_meta or {}).get("filters") or {}
    cutoff = (run_meta or {}).get("cutoff") or {}
    sample = (run_meta or {}).get("sample_size") or {}

    valores = [
        DSN,
        now.strftime("%d/%m/%Y"),
        now.strftime("%H:%M:%S UTC"),
        "PNCP (Portal Nacional de Contratacoes Publicas), SC Public Entities",
        "opportunity_intel",
        str(stats.get("total", 0)),
        str(stats.get("go_count", 0)),
        str(stats.get("no_go_count", 0)),
        str(stats.get("review_count", 0)),
        f"{stats.get('score_medio_go', 0):.1f}",
        f"R$ {stats.get('valor_total', 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        str(stats.get("orgaos_ativos", 0)),
        str(stats.get("conf_high", 0)),
        str(stats.get("conf_medium", 0)),
        str(stats.get("conf_low", 0)),
        str(stats.get("vincendos_count", 0)),
        str(stats.get("raw_bids_count", 0)),
        str(stats.get("ultima_atualizacao", "N/I")),
        f"+7 dias ({datetime.now(UTC).date().isoformat()})",
        str((run_meta or {}).get("run_id", "")),
        str((run_meta or {}).get("generated_at", now.isoformat().replace("+00:00", "Z"))),
        str((run_meta or {}).get("git_sha", "unknown")),
        str((run_meta or {}).get("profile_id", "extra")),
        str(filters.get("uf", "SC")),
        str(filters.get("is_active", True)),
        str(filters.get("table_primary", "opportunity_intel")),
        str(filters.get("vincendos_horizon_days", 180)),
        str(cutoff.get("as_of_date", now.date().isoformat())),
        str(cutoff.get("data_window", "all_active")),
        str(cutoff.get("ultima_atualizacao_db", stats.get("ultima_atualizacao", "N/I"))),
        str(sample.get("label", "")),
        str(sample.get("opportunity_intel_active", stats.get("total", 0))),
        str(sample.get("ranking_go", stats.get("go_count", 0))),
        str(sample.get("ranking_review", stats.get("review_count", 0))),
        str(sample.get("ranking_no_go", stats.get("no_go_count", 0))),
        str(sample.get("orgaos_sc", stats.get("orgaos_ativos", 0))),
        str(sample.get("vincendos_180d", stats.get("vincendos_count", 0))),
        str(sample.get("pncp_raw_bids_active", stats.get("raw_bids_count", 0))),
    ]

    for label, value in zip(METADADOS_LABELS, valores):
        label_cell = _woc(ws, value=label, font=st["label_font"], fill=st["label_fill"], border=st["thin_border"])
        val_cell = _woc(ws, value=value, border=st["thin_border"], alignment=Alignment(wrap_text=True))
        ws.append([label_cell, val_cell])


def _build_legenda(wb: Workbook, st: dict):
    ws = wb.create_sheet("Legenda")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 70

    # Section: ranking
    legenda_items = [
        ("--- RANKINGS ---", ""),
        ("GO", "Recomendado para participacao. Score > 0 e sem bloqueadores."),
        ("NO_GO", "Nao recomendado. Possui bloqueadores (status terminal, data expirada, etc.)."),
        ("REVIEW", "Aguardando classificacao. Requer analise manual ou dados adicionais."),
        ("", ""),
        ("--- SCORE (0-100) ---", ""),
        ("90-100", "Excelente. Todos os criterios positivos atendidos."),
        ("70-89", "Bom. Maioria dos criterios atendidos, poucos negativos."),
        ("50-69", "Moderado. Criterios mistos, requer atencao."),
        ("1-49", "Baixo. Diversos criterios negativos."),
        ("0", "Classificado como NO_GO ou sem pontuacao."),
        ("", ""),
        ("--- CONFIACA (Confidence) ---", ""),
        ("HIGH", "Dados consistentes, multiplas fontes verificadas."),
        ("MEDIUM", "Dados parcialmente verificados, algumas inconsistencias."),
        ("LOW", "Dados com baixa confiabilidade, requer verificacao manual."),
        ("", ""),
        ("--- COLUNAS DE RASTREABILIDADE ---", ""),
        ("Source", "Fonte de origem dos dados (ex: pncp, dom_sc)."),
        ("Source ID", "Identificador unico na fonte de origem."),
        ("Source URL", "URL direta para o registro na fonte."),
        ("Content Hash", "Hash SHA256 do conteudo para deteccao de alteracoes."),
        ("", ""),
        ("--- VALORES ---", ""),
        ("Valor Estimado", "Previsao inicial do orgao contratante (NUNCA confundir com homologado)."),
        ("Valor Homologado", "Valor efetivamente contratado apos a licitacao."),
        ("", ""),
        ("--- QUALIDADE SCORE (0-100) ---", ""),
        ("80-100", "Dados completos, todos os campos obrigatorios preenchidos."),
        ("50-79", "Dados parcialmente completos."),
        ("0-49", "Dados com muitos campos ausentes."),
        ("", ""),
        ("--- GERACAO DO RELATORIO ---", ""),
        ("Data", f"{datetime.now(UTC).strftime('%d/%m/%Y %H:%M:%S')} UTC"),
        ("Script", "scripts/reports/executive_excel.py"),
        ("DSN", DSN),
    ]

    for label, desc in legenda_items:
        label_cell = _woc(ws, value=label, font=st["label_font"] if label else None, border=st["thin_border"])
        val_cell = _woc(ws, value=desc, border=st["thin_border"], alignment=Alignment(wrap_text=True))
        ws.append([label_cell, val_cell])


# ============================================================
# MAIN GENERATOR
# ============================================================


def generate_executive_excel(output_path: str, run_id: str | None = None, uf: str = "SC"):
    from scripts.reports.run_metadata import build_run_metadata, write_sidecar

    conn = get_conn()
    st = _make_styles()

    try:
        # Load data
        editais = load_editais(conn)
        orgaos = load_orgaos(conn)
        concorrentes = load_concorrentes(conn)
        vincendos = query(
            conn,
            """SELECT COUNT(*) as cnt FROM pncp_supplier_contracts
               WHERE is_active = true
                 AND data_fim BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '180 days'""",
        )
        raw_bids_count = scalar(conn, "SELECT COUNT(*) FROM pncp_raw_bids WHERE is_active = true") or 0
        ultima_atualizacao = (
            scalar(conn, "SELECT MAX(updated_at)::text FROM opportunity_intel WHERE is_active = true") or "N/I"
        )

        # Compute stats
        go_count = sum(1 for o in editais if o.get("ranking") == "GO")
        no_go_count = sum(1 for o in editais if o.get("ranking") == "NO_GO")
        review_count = sum(1 for o in editais if o.get("ranking") == "REVIEW")

        go_scores = [float(o.get("ranking_score", 0) or 0) for o in editais if o.get("ranking") == "GO"]
        score_medio_go = sum(go_scores) / max(1, len(go_scores))

        valor_total = sum(float(o.get("valor_estimado") or 0) for o in editais)
        orgaos_ativos = len({o.get("orgao_nome") for o in editais if o.get("orgao_nome") and o.get("uf") == "SC"})

        conf_high = sum(1 for o in editais if o.get("ranking_confianca") == "HIGH")
        conf_medium = sum(1 for o in editais if o.get("ranking_confianca") == "MEDIUM")
        conf_low = sum(1 for o in editais if o.get("ranking_confianca") == "LOW")
        vincendos_count = vincendos[0]["cnt"] if vincendos else 0

        stats = {
            "total": len(editais),
            "go_count": go_count,
            "no_go_count": no_go_count,
            "review_count": review_count,
            "score_medio_go": score_medio_go,
            "valor_total": valor_total,
            "orgaos_ativos": orgaos_ativos,
            "conf_high": conf_high,
            "conf_medium": conf_medium,
            "conf_low": conf_low,
            "vincendos_count": vincendos_count,
            "raw_bids_count": raw_bids_count,
            "ultima_atualizacao": ultima_atualizacao,
        }

        run_meta = build_run_metadata(
            run_id=run_id,
            artifact_kind="excel",
            script="scripts/reports/executive_excel.py",
            uf=uf,
            stats=stats,
        )
        stats["run_id"] = run_meta["run_id"]
        stats["sample_label"] = run_meta["sample_size"]["label"]
        stats["run_meta"] = run_meta

        # Build workbook
        wb = Workbook(write_only=True)

        _build_editais(wb, editais, st)
        _build_orgaos(wb, orgaos, st)
        _build_concorrentes(wb, concorrentes, st)
        _build_metadados(wb, stats, st, run_meta=run_meta)
        _build_legenda(wb, st)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        write_sidecar(output_path, run_meta)

        return output_path, stats

    finally:
        conn.close()


# ============================================================
# CLI
# ============================================================


def main():
    parser = argparse.ArgumentParser(
        description="Gera Excel de Relatorio Executivo para Extra Construtora (dados reais do banco)"
    )
    parser.add_argument(
        "--output",
        "-o",
        help="Caminho do .xlsx de saida (default: output/reports/executivo-extra-YYYY-MM-DD.xlsx)",
    )
    parser.add_argument(
        "--run-id",
        help="Run ID compartilhado com o PDF (reconcile_pdf_excel). Se omitido, gera um novo.",
    )
    parser.add_argument("--uf", default="SC", help="UF de referencia no metadata (default SC)")
    args = parser.parse_args()

    today_str = datetime.now(UTC).strftime("%Y-%m-%d")
    output_path = args.output or f"output/reports/executivo-extra-{today_str}.xlsx"

    result, stats = generate_executive_excel(output_path, run_id=args.run_id, uf=args.uf)
    size_kb = os.path.getsize(result) / 1024
    print(f"Excel gerado: {result} ({size_kb:.0f}KB)")
    print("  Sheets: Editais, Orgaos, Concorrentes, Metadados, Legenda")
    print(
        f"  Oportunidades: {stats['total']} (GO: {stats['go_count']}, "
        f"NO_GO: {stats['no_go_count']}, REVIEW: {stats['review_count']})"
    )
    print(f"  Orgaos ativos SC: {stats['orgaos_ativos']}")
    print(f"  Score medio GO: {stats['score_medio_go']:.0f}/100")
    print(f"  Run ID: {stats.get('run_id')} sample={stats.get('sample_label')}")
    print(f"  Sidecar: {result}.meta.json")


if __name__ == "__main__":
    main()
