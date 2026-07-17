#!/usr/bin/env python3
"""Reconcile commercial PDF × Excel outputs for the next-30d window.

Discovers (or accepts) latest Excel under ``output/excels`` and PDF under
``output/``, compares run metadata when present (cutoff date, UF, profile
version), and writes a JSON report to ``output/reports/reconcile-next30d.json``.

Exit codes:
    0  consistent filters, or both artifacts missing (with note)
    1  conflicting filters / critical metadata mismatch
    2  usage / I/O error
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

DEFAULT_EXCEL_DIR = _PROJECT_ROOT / "output" / "excels"
DEFAULT_PDF_DIR = _PROJECT_ROOT / "output"
DEFAULT_OUTPUT = _PROJECT_ROOT / "output" / "reports" / "reconcile-next30d.json"

# Critical filter keys — mismatch → exit 1
CRITICAL_FILTER_KEYS = ("uf", "is_active", "table_primary", "vincendos_horizon_days")
CRITICAL_CUTOFF_KEYS = ("as_of_date", "data_window")


def _find_latest(directory: Path, suffixes: tuple[str, ...], *, recursive: bool = False) -> Path | None:
    """Return newest file matching any of *suffixes* under *directory*."""
    if not directory.is_dir():
        return None
    candidates: list[Path] = []
    iterator = directory.rglob("*") if recursive else directory.glob("*")
    for p in iterator:
        if not p.is_file():
            continue
        if p.name.endswith(".meta.json"):
            continue
        if p.suffix.lower() in suffixes:
            candidates.append(p)
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _meta_path(artifact: Path) -> Path:
    return artifact.with_suffix(artifact.suffix + ".meta.json")


def _load_json(path: Path) -> dict[str, Any] | None:
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def _extract_meta_from_excel(path: Path) -> dict[str, Any] | None:
    """Best-effort parse of Metadados sheet when sidecar is missing."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    if "Metadados" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["Metadados"]
    kv: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if not row or row[0] is None:
            continue
        kv[str(row[0]).strip()] = "" if row[1] is None else str(row[1]).strip()
    wb.close()

    run_id = kv.get("Run ID") or kv.get("run_id")
    if not run_id and not (kv.get("Filter UF") or kv.get("Cutoff as_of_date")):
        return None

    def _int(key: str, default: int = 0) -> int:
        raw = kv.get(key, "")
        try:
            return int(str(raw).replace(",", "").strip() or default)
        except ValueError:
            return default

    profile_id = kv.get("Profile ID") or kv.get("profile_id") or "extra"
    profile_version = kv.get("Profile Version") or kv.get("profile_version")

    return {
        "run_id": run_id or None,
        "generated_at": kv.get("Generated At (UTC)") or kv.get("Data de Geracao"),
        "profile_id": profile_id,
        "profile_version": profile_version,
        "git_sha": kv.get("Git SHA") or "unknown",
        "filters": {
            "uf": kv.get("Filter UF") or kv.get("UF") or None,
            "is_active": str(kv.get("Filter is_active", "True")).lower() in ("true", "1", "yes"),
            "table_primary": kv.get("Filter table_primary") or "opportunity_intel",
            "vincendos_horizon_days": _int("Filter vincendos_horizon_days", 180) or 180,
        },
        "cutoff": {
            "as_of_date": kv.get("Cutoff as_of_date") or kv.get("as_of_date") or None,
            "data_window": kv.get("Cutoff data_window") or "all_active",
            "ultima_atualizacao_db": kv.get("Cutoff ultima_atualizacao_db")
            or kv.get("Data da Ultima Atualizacao no DB")
            or "N/I",
        },
        "artifact_kind": "excel",
        "source": "excel_metadados_sheet",
    }


def resolve_metadata(artifact: Path, kind: str) -> tuple[dict[str, Any] | None, str]:
    """Return (meta, source) where source is sidecar|embedded|missing."""
    side = _load_json(_meta_path(artifact))
    if side:
        return side, "sidecar"
    if kind == "excel":
        embedded = _extract_meta_from_excel(artifact)
        if embedded:
            return embedded, "excel_metadados"
    # PDF binary parse intentionally avoided — generators must write sidecar.
    return None, "missing"


def _norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip().lower()


def _profile_version(meta: dict[str, Any] | None) -> Any:
    if not meta:
        return None
    if meta.get("profile_version") is not None:
        return meta.get("profile_version")
    # Some generators nest under filters or claims.
    filters = meta.get("filters") or {}
    if filters.get("profile_version") is not None:
        return filters.get("profile_version")
    # Fallback: profile_id alone is not a version, but record it for soft check.
    return None


def compare_filters(
    pdf_meta: dict[str, Any] | None,
    excel_meta: dict[str, Any] | None,
) -> dict[str, Any]:
    """Compare cutoff / UF / profile version between two metadata dicts."""
    checks: list[dict[str, Any]] = []

    def _check(name: str, left: Any, right: Any, *, critical: bool = True) -> None:
        # Empty vs empty is not a conflict when both lack the field.
        if _norm(left) == "" and _norm(right) == "":
            checks.append(
                {
                    "name": name,
                    "pdf": left,
                    "excel": right,
                    "match": True,
                    "critical": critical,
                    "note": "both_absent",
                }
            )
            return
        ok = _norm(left) == _norm(right)
        checks.append(
            {
                "name": name,
                "pdf": left,
                "excel": right,
                "match": ok,
                "critical": critical,
            }
        )

    pdf_filters = (pdf_meta or {}).get("filters") or {}
    excel_filters = (excel_meta or {}).get("filters") or {}
    pdf_cutoff = (pdf_meta or {}).get("cutoff") or {}
    excel_cutoff = (excel_meta or {}).get("cutoff") or {}

    for key in CRITICAL_CUTOFF_KEYS:
        _check(f"cutoff.{key}", pdf_cutoff.get(key), excel_cutoff.get(key), critical=True)

    for key in CRITICAL_FILTER_KEYS:
        _check(f"filters.{key}", pdf_filters.get(key), excel_filters.get(key), critical=True)

    _check(
        "profile_version",
        _profile_version(pdf_meta),
        _profile_version(excel_meta),
        critical=True,
    )
    _check(
        "profile_id",
        (pdf_meta or {}).get("profile_id"),
        (excel_meta or {}).get("profile_id"),
        critical=False,
    )
    _check(
        "run_id",
        (pdf_meta or {}).get("run_id"),
        (excel_meta or {}).get("run_id"),
        critical=False,
    )
    _check(
        "git_sha",
        (pdf_meta or {}).get("git_sha"),
        (excel_meta or {}).get("git_sha"),
        critical=False,
    )

    critical_fail = [c for c in checks if c["critical"] and not c["match"]]
    soft_fail = [c for c in checks if not c["critical"] and not c["match"]]

    if critical_fail:
        verdict = "CONFLICT"
    elif soft_fail:
        verdict = "CONSISTENT_SOFT"
    else:
        verdict = "CONSISTENT"

    return {
        "verdict": verdict,
        "checks": checks,
        "critical_mismatches": len(critical_fail),
        "soft_mismatches": len(soft_fail),
        "conflicting_fields": [c["name"] for c in critical_fail],
    }


def reconcile(
    pdf_path: Path | None,
    excel_path: Path | None,
) -> dict[str, Any]:
    report: dict[str, Any] = {
        "schema_version": 1,
        "reconciled_at": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "pdf_path": str(pdf_path) if pdf_path else None,
        "excel_path": str(excel_path) if excel_path else None,
        "pdf_exists": bool(pdf_path and pdf_path.is_file()),
        "excel_exists": bool(excel_path and excel_path.is_file()),
        "pdf_meta_path": str(_meta_path(pdf_path)) if pdf_path else None,
        "excel_meta_path": str(_meta_path(excel_path)) if excel_path else None,
    }

    # Both missing → exit 0 with note (cannot conflict).
    if not report["pdf_exists"] and not report["excel_exists"]:
        report["verdict"] = "BOTH_MISSING"
        report["consistent"] = True
        report["reason"] = (
            "Neither PDF nor Excel found under discovery paths — "
            "nothing to reconcile; treated as consistent with note."
        )
        report["checks"] = []
        report["exit_hint"] = 0
        return report

    # One present, one missing → not a filter conflict; consistent with note.
    if not report["pdf_exists"] or not report["excel_exists"]:
        missing = "PDF" if not report["pdf_exists"] else "Excel"
        present = "Excel" if not report["pdf_exists"] else "PDF"
        report["verdict"] = "PAIR_INCOMPLETE"
        report["consistent"] = True
        report["reason"] = (
            f"{missing} missing; only {present} present — cannot detect filter conflict. "
            "Pair incomplete (not a conflict)."
        )
        report["checks"] = []
        report["exit_hint"] = 0
        return report

    if pdf_path is None or excel_path is None:
        report["verdict"] = "BOTH_MISSING"
        report["consistent"] = True
        report["reason"] = "Internal path resolution failed — treated as both missing."
        report["checks"] = []
        report["exit_hint"] = 0
        return report

    pdf_meta, pdf_src = resolve_metadata(pdf_path, "pdf")
    excel_meta, excel_src = resolve_metadata(excel_path, "excel")
    report["pdf_meta_source"] = pdf_src
    report["excel_meta_source"] = excel_src

    # Both artifacts exist but no metadata → cannot prove conflict; note only.
    if pdf_meta is None and excel_meta is None:
        report["verdict"] = "METADATA_ABSENT"
        report["consistent"] = True
        report["reason"] = (
            "PDF and Excel exist but neither has run metadata "
            "(*.meta.json / Metadados sheet). Filters not comparable; "
            "treated as consistent with note (regenerate with run_metadata sidecars)."
        )
        report["checks"] = []
        report["exit_hint"] = 0
        return report

    # Only one side has metadata → incomplete comparison, not a hard conflict.
    if pdf_meta is None or excel_meta is None:
        missing_side = "pdf" if pdf_meta is None else "excel"
        report["verdict"] = "METADATA_PARTIAL"
        report["consistent"] = True
        report["reason"] = (
            f"Metadata only on one side ({missing_side} missing). "
            "Cannot confirm conflicting filters; treated as consistent with note."
        )
        report["pdf_meta"] = pdf_meta
        report["excel_meta"] = excel_meta
        report["checks"] = []
        report["exit_hint"] = 0
        return report

    comparison = compare_filters(pdf_meta, excel_meta)
    report.update(comparison)
    report["pdf_meta"] = {
        k: pdf_meta.get(k)
        for k in (
            "run_id",
            "filters",
            "cutoff",
            "sample_size",
            "git_sha",
            "profile_id",
            "profile_version",
        )
    }
    report["excel_meta"] = {
        k: excel_meta.get(k)
        for k in (
            "run_id",
            "filters",
            "cutoff",
            "sample_size",
            "git_sha",
            "profile_id",
            "profile_version",
        )
    }

    if comparison["verdict"] == "CONFLICT":
        report["consistent"] = False
        report["reason"] = (
            "Conflicting filters/metadata between PDF and Excel: "
            + ", ".join(comparison["conflicting_fields"])
        )
        report["exit_hint"] = 1
    else:
        report["consistent"] = True
        if comparison["verdict"] == "CONSISTENT_SOFT":
            report["reason"] = (
                "Critical filters match; soft fields differ "
                "(run_id / git_sha / profile_id)."
            )
        else:
            report["reason"] = "PDF and Excel share cutoff, UF and profile version filters"
        report["exit_hint"] = 0

    return report


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Reconcile PDF × Excel commercial outputs (next-30d)",
    )
    p.add_argument("--pdf", type=Path, help="Path to executive PDF")
    p.add_argument("--excel", type=Path, help="Path to executive Excel")
    p.add_argument(
        "--excel-dir",
        type=Path,
        default=DEFAULT_EXCEL_DIR,
        help=f"Directory to discover latest Excel (default: {DEFAULT_EXCEL_DIR})",
    )
    p.add_argument(
        "--pdf-dir",
        type=Path,
        default=DEFAULT_PDF_DIR,
        help=f"Directory to discover latest PDF recursively (default: {DEFAULT_PDF_DIR})",
    )
    p.add_argument(
        "--output",
        "-o",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"JSON report path (default: {DEFAULT_OUTPUT})",
    )
    p.add_argument(
        "--no-discover",
        action="store_true",
        help="Do not auto-discover; require --pdf/--excel explicitly",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    pdf = args.pdf
    excel = args.excel

    if not args.no_discover:
        if pdf is None:
            # Prefer executive reports under output/reports, then any PDF under output/
            reports_dir = _PROJECT_ROOT / "output" / "reports"
            pdf = _find_latest(reports_dir, (".pdf",), recursive=False)
            if pdf is None:
                pdf = _find_latest(args.pdf_dir, (".pdf",), recursive=True)
        if excel is None:
            excel = _find_latest(args.excel_dir, (".xlsx", ".xlsm"), recursive=False)
            # Fallback: executive xlsx under output/reports (paired with PDF)
            if excel is None:
                reports_dir = _PROJECT_ROOT / "output" / "reports"
                excel = _find_latest(reports_dir, (".xlsx",), recursive=False)

    report = reconcile(pdf, excel)

    text = json.dumps(report, ensure_ascii=False, indent=2) + "\n"
    out: Path = args.output
    try:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(text, encoding="utf-8")
    except OSError as exc:
        print(f"ERROR: cannot write {out}: {exc}", file=sys.stderr)
        return 2

    print(f"Reconciliation report: {out}")
    print(
        f"verdict={report.get('verdict')} consistent={report.get('consistent')} "
        f"pdf_exists={report.get('pdf_exists')} excel_exists={report.get('excel_exists')} "
        f"reason={report.get('reason')}"
    )

    return int(report.get("exit_hint", 0))


if __name__ == "__main__":
    sys.exit(main())
