#!/usr/bin/env python3
"""Deliverable A + E stubs — ranking de órgãos e GO/REVIEW/NO_GO com dados reais.

Gera JSON (+ opcional Excel) a partir de opportunity_intel / pncp_raw_bids.
Rótulos honestos de sample size; não inventa contratos vincendos (Deliverable C).

Usage:
    PYTHONPATH=. DATABASE_URL=... python scripts/reports/deliverable_orgaos_ranking.py
    python scripts/reports/deliverable_orgaos_ranking.py --uf SC --output output/reports/deliverable-a-e.json
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import UTC, datetime
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

DSN = os.getenv(
    "LOCAL_DATALAKE_DSN",
    os.getenv("DATABASE_URL", "postgresql://test:test@127.0.0.1:5433/pncp_datalake"),
)

from scripts.reports.run_metadata import (  # noqa: E402
    CLAIMS_ALLOWED,
    CLAIMS_FORBIDDEN,
    build_run_metadata,
    new_run_id,
    write_sidecar,
)


def get_conn():
    import psycopg2
    import psycopg2.extras

    return psycopg2.connect(DSN, cursor_factory=psycopg2.extras.RealDictCursor)


def query(conn, sql: str, params=None) -> list[dict]:
    cur = conn.cursor()
    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    return rows


def scalar(conn, sql: str, params=None):
    cur = conn.cursor()
    cur.execute(sql, params)
    row = cur.fetchone()
    cur.close()
    if row is None:
        return None
    if isinstance(row, dict):
        return next(iter(row.values()))
    return row[0]


def load_deliverable_a(conn, uf: str, limit: int = 20) -> list[dict]:
    """Ranking de órgãos compradores (Deliverable A)."""
    return query(
        conn,
        """
        SELECT orgao_nome AS orgao,
               uf,
               COUNT(*) AS qtd_editais,
               COUNT(*) FILTER (WHERE ranking = 'GO') AS qtd_go,
               COUNT(*) FILTER (WHERE ranking = 'REVIEW') AS qtd_review,
               COUNT(*) FILTER (WHERE ranking = 'NO_GO') AS qtd_no_go,
               ROUND(SUM(COALESCE(valor_estimado, 0))::numeric, 2) AS valor_estimado_total
        FROM opportunity_intel
        WHERE is_active = true
          AND (%s IS NULL OR uf = %s)
          AND orgao_nome IS NOT NULL
        GROUP BY orgao_nome, uf
        ORDER BY qtd_editais DESC, valor_estimado_total DESC NULLS LAST
        LIMIT %s
        """,
        (uf, uf, limit),
    )


def load_deliverable_e(conn, uf: str) -> dict[str, list[dict]]:
    """Stubs GO / REVIEW / NO_GO com registros reais (Deliverable E)."""
    rows = query(
        conn,
        """
        SELECT id, source, source_id, orgao_nome, uf, municipio, modalidade,
               LEFT(COALESCE(objeto, ''), 200) AS objeto,
               valor_estimado, ranking, ranking_score, ranking_confianca,
               status_canonico, data_publicacao, data_encerramento
        FROM opportunity_intel
        WHERE is_active = true
          AND (%s IS NULL OR uf = %s)
        ORDER BY ranking, ranking_score DESC NULLS LAST, id
        """,
        (uf, uf),
    )
    buckets: dict[str, list[dict]] = {"GO": [], "REVIEW": [], "NO_GO": [], "OTHER": []}
    for r in rows:
        # Serialize dates
        for k, v in list(r.items()):
            if hasattr(v, "isoformat"):
                r[k] = v.isoformat()
            elif v is not None and not isinstance(v, (str, int, float, bool)):
                r[k] = str(v)
        key = r.get("ranking") or "OTHER"
        if key not in buckets:
            buckets["OTHER"].append(r)
        else:
            buckets[key].append(r)
    return buckets


def load_deliverable_c(conn) -> dict:
    """Contratos vincendos (Deliverable C) — honest empty when table empty."""
    try:
        rows = query(
            conn,
            """
            SELECT COUNT(*) AS cnt
            FROM pncp_supplier_contracts
            WHERE is_active = true
              AND data_fim BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '180 days'
            """,
        )
        n = int(rows[0]["cnt"]) if rows else 0
        sample = query(
            conn,
            """
            SELECT contrato_id, orgao_nome, fornecedor_nome,
                   valor_total, data_inicio, data_fim
            FROM pncp_supplier_contracts
            WHERE is_active = true
              AND data_fim BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '180 days'
            ORDER BY data_fim ASC
            LIMIT 10
            """,
        )
        for r in sample:
            for k, v in list(r.items()):
                if hasattr(v, "isoformat"):
                    r[k] = v.isoformat()
                elif v is not None and not isinstance(v, (str, int, float, bool)):
                    r[k] = str(v)
    except Exception as exc:  # table missing or schema drift
        return {
            "status": "UNAVAILABLE",
            "count": 0,
            "rows": [],
            "error": str(exc),
            "claim": "Não afirmar contratos vincendos sem pncp_supplier_contracts populada.",
        }

    if n == 0:
        return {
            "status": "INSUFFICIENT",
            "count": 0,
            "rows": [],
            "claim": (
                "pncp_supplier_contracts sem linhas no horizonte 180d — "
                "Deliverable C bloqueado honestamente; não derivar de pncp_raw_bids."
            ),
        }
    return {
        "status": "OK",
        "count": n,
        "rows": sample,
        "claim": f"{n} contratos com data_fim nos próximos 180 dias (amostra até 10).",
    }


def build_payload(uf: str, limit: int = 20) -> dict:
    conn = get_conn()
    try:
        orgaos = load_deliverable_a(conn, uf, limit)
        buckets = load_deliverable_e(conn, uf)
        vincendos = load_deliverable_c(conn)
        n_opps = scalar(
            conn,
            """SELECT COUNT(*) FROM opportunity_intel
               WHERE is_active = true AND (%s IS NULL OR uf = %s)""",
            (uf, uf),
        ) or 0
        n_bids = scalar(
            conn,
            """SELECT COUNT(*) FROM pncp_raw_bids
               WHERE is_active = true AND (%s IS NULL OR uf = %s)""",
            (uf, uf),
        ) or 0
        ultima = scalar(
            conn,
            "SELECT MAX(updated_at)::text FROM opportunity_intel WHERE is_active = true",
        ) or "N/I"
    finally:
        conn.close()

    # Normalize numeric types for JSON
    for o in orgaos:
        for k, v in list(o.items()):
            if hasattr(v, "isoformat"):
                o[k] = v.isoformat()
            elif v is not None and not isinstance(v, (str, int, float, bool)):
                try:
                    o[k] = float(v)
                except (TypeError, ValueError):
                    o[k] = str(v)

    stats = {
        "total": int(n_opps),
        "go_count": len(buckets.get("GO", [])),
        "review_count": len(buckets.get("REVIEW", [])),
        "no_go_count": len(buckets.get("NO_GO", [])),
        "orgaos_ativos": len(orgaos),
        "raw_bids_count": int(n_bids),
        "vincendos_count": int(vincendos.get("count") or 0),
        "ultima_atualizacao": ultima,
    }
    run_id = new_run_id(prefix="deliv-ae")
    meta = build_run_metadata(
        run_id=run_id,
        artifact_kind="deliverable_a_e",
        script="scripts/reports/deliverable_orgaos_ranking.py",
        uf=uf or "ALL",
        stats=stats,
    )

    return {
        "deliverable": "A+E (+C status)",
        "title": "Ranking de órgãos + stubs GO/REVIEW/NO_GO",
        "run_metadata": meta,
        "deliverable_a_orgaos_ranking": {
            "status": "OK" if orgaos else "INSUFFICIENT",
            "uf_filter": uf,
            "count": len(orgaos),
            "rows": orgaos,
            "note": (
                "Ranking por volume de editais ativos em opportunity_intel. "
                f"Sample label={meta['sample_size']['label']}."
            ),
        },
        "deliverable_e_go_review_no_go": {
            "status": "OK" if n_opps else "INSUFFICIENT",
            "counts": {
                "GO": len(buckets.get("GO", [])),
                "REVIEW": len(buckets.get("REVIEW", [])),
                "NO_GO": len(buckets.get("NO_GO", [])),
                "OTHER": len(buckets.get("OTHER", [])),
            },
            "rows": buckets,
            "note": (
                "Classificação lida da coluna ranking; com fixtures pode ser 100% REVIEW. "
                "Não relabelar como GO sem motor de ranking reexecutado."
            ),
        },
        "deliverable_c_contratos_vincendos": vincendos,
        "claims": {
            "allowed": CLAIMS_ALLOWED,
            "forbidden": CLAIMS_FORBIDDEN,
        },
        "generated_at": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
    }


def maybe_write_excel(payload: dict, path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl missing — skip Excel for deliverable A/E")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Orgaos"
    headers = [
        "orgao",
        "uf",
        "qtd_editais",
        "qtd_go",
        "qtd_review",
        "qtd_no_go",
        "valor_estimado_total",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B2A3D")
    for row in payload["deliverable_a_orgaos_ranking"]["rows"]:
        ws.append([row.get(h) for h in headers])

    ws2 = wb.create_sheet("GO_REVIEW_NO_GO")
    ws2.append(
        [
            "ranking",
            "id",
            "orgao_nome",
            "uf",
            "municipio",
            "modalidade",
            "valor_estimado",
            "ranking_score",
            "objeto",
        ]
    )
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B2A3D")
    for rank in ("GO", "REVIEW", "NO_GO", "OTHER"):
        for r in payload["deliverable_e_go_review_no_go"]["rows"].get(rank, []):
            ws2.append(
                [
                    rank,
                    r.get("id"),
                    r.get("orgao_nome"),
                    r.get("uf"),
                    r.get("municipio"),
                    r.get("modalidade"),
                    r.get("valor_estimado"),
                    r.get("ranking_score"),
                    r.get("objeto"),
                ]
            )

    ws3 = wb.create_sheet("Metadados")
    meta = payload["run_metadata"]
    ws3.append(["Run ID", meta.get("run_id")])
    ws3.append(["Sample label", (meta.get("sample_size") or {}).get("label")])
    ws3.append(["UF", (meta.get("filters") or {}).get("uf")])
    ws3.append(["as_of_date", (meta.get("cutoff") or {}).get("as_of_date")])
    ws3.append(["Deliverable C", payload["deliverable_c_contratos_vincendos"].get("status")])
    ws3.append(["C claim", payload["deliverable_c_contratos_vincendos"].get("claim")])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Excel deliverable A/E: {path}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Deliverable A+E (orgãos + GO/REVIEW/NO_GO)")
    parser.add_argument("--uf", default="SC", help="UF filter (default SC; empty for all)")
    parser.add_argument("--limit", type=int, default=20)
    parser.add_argument(
        "--output",
        "-o",
        default=None,
        help="JSON output path",
    )
    parser.add_argument(
        "--excel",
        default=None,
        help="Optional Excel path for A/E",
    )
    parser.add_argument("--dsn", default=None)
    args = parser.parse_args()

    global DSN
    if args.dsn:
        DSN = args.dsn
    elif os.getenv("DATABASE_URL") and not os.getenv("LOCAL_DATALAKE_DSN"):
        DSN = os.environ["DATABASE_URL"]

    uf = args.uf.strip() or None
    payload = build_payload(uf=uf or "SC", limit=args.limit)
    # If user passed empty uf for all, fix filter label
    if uf is None:
        payload["run_metadata"]["filters"]["uf"] = "ALL"
        payload["deliverable_a_orgaos_ranking"]["uf_filter"] = None

    today = datetime.now(UTC).strftime("%Y-%m-%d")
    out = Path(args.output or f"output/reports/deliverable-a-e-{today}.json")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_sidecar(out, payload["run_metadata"])
    print(f"JSON: {out}")
    print(
        f"  A orgaos={payload['deliverable_a_orgaos_ranking']['count']} "
        f"E counts={payload['deliverable_e_go_review_no_go']['counts']} "
        f"C={payload['deliverable_c_contratos_vincendos']['status']} "
        f"sample={payload['run_metadata']['sample_size']['label']}"
    )

    excel_path = Path(args.excel or f"output/reports/deliverable-a-e-{today}.xlsx")
    maybe_write_excel(payload, excel_path)
    write_sidecar(excel_path, payload["run_metadata"])
    return 0


if __name__ == "__main__":
    sys.exit(main())
