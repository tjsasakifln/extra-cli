#!/usr/bin/env python3
"""Resolve 604 entities with 'N/D' coordinates in the seed spreadsheet.

Strategy:
  1. Read all rows from the seed spreadsheet
  2. For rows with lat='N/D', group by unique municipality name
  3. Geocode each unique municipality via Nominatim (OSM) API
  4. Map coordinates back to all rows sharing that municipality
  5. Update the spreadsheet with resolved coordinates and distances
  6. Update sc_public_entities table in PostgreSQL
  7. Re-run consulting_readiness.py to verify

Usage:
    python scripts/fix/resolve_unresolved_entities.py
"""

from __future__ import annotations

import json
import math
import os
import re
import subprocess
import sys
import time
import urllib.request
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
SEED_PATH = PROJECT_ROOT / "Extra - alvos de licitação. R-0.xlsx"
BACKUP_PATH = PROJECT_ROOT / "Extra - alvos de licitação. R-0.backup.xlsx"

FLORIANOPOLIS_LAT = -27.5954
FLORIANOPOLIS_LON = -48.5480
EARTH_RADIUS_KM = 6371.0
RADIUS_KM = 200.0

# Column indices (0-based, matching consulting_readiness.py)
COL_RAZAO = 0
COL_MUNICIPIO = 2
COL_IBGE = 3
COL_LATITUDE = 6
COL_LONGITUDE = 7
COL_DISTANCIA = 8
COL_RAIO200 = 9

USER_AGENT = "ExtraConsultoriaB2G/1.0 (research@extraconsultoria.com.br)"
NOMINATIM_DELAY = 1.1  # seconds between requests (respect rate limit)

_DSN = os.getenv(
    "LOCAL_DATALAKE_DSN",
    "postgresql://postgres:smartlic_local@127.0.0.1:54399/postgres",
)


# ---------------------------------------------------------------------------
# Haversine
# ---------------------------------------------------------------------------


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Great-circle distance in km."""
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2
    return EARTH_RADIUS_KM * 2 * math.asin(math.sqrt(a))


# ---------------------------------------------------------------------------
# Geocoding
# ---------------------------------------------------------------------------


def _fetch_json(url: str) -> Any:
    """Fetch and decode JSON from URL (handles gzip)."""
    import gzip

    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT, "Accept-Encoding": "gzip"})
    resp = urllib.request.urlopen(req, timeout=15)
    raw = resp.read()
    try:
        return json.loads(raw.decode("utf-8"))
    except UnicodeDecodeError:
        return json.loads(gzip.decompress(raw).decode("utf-8"))


def geocode_municipality(name: str) -> tuple[float, float] | None:
    """Geocode a SC municipality via Nominatim OSM API."""
    # Clean and encode
    q = urllib.request.quote(f"{name}, SC, Brasil")
    url = f"https://nominatim.openstreetmap.org/search?q={q}&format=json&limit=1"
    try:
        data = _fetch_json(url)
        if data and len(data) > 0:
            lat = float(data[0]["lat"])
            lon = float(data[0]["lon"])
            return lat, lon
    except Exception as e:
        print(f"    Nominatim error for '{name}': {e}", file=sys.stderr)
    return None


def _normalize_muni(name: str) -> str:
    """Normalize municipality name for comparison."""
    return re.sub(r"\s+", " ", name.strip().upper())


# ---------------------------------------------------------------------------
# IBGE API lookup (via municipality code)
# ---------------------------------------------------------------------------


def ibge_lookup_by_code(codigo_ibge: int) -> tuple[float, float] | None:
    """Lookup municipality via IBGE API — does NOT return lat/lon directly,
    but we try to use an alternative endpoint that might have coordinates."""
    # The standard IBGE API does not return lat/lon for municipalities.
    # This is a placeholder — we rely on Nominatim as primary source.
    return None


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------


def _get_db_conn():
    """Connect to PostgreSQL."""
    import psycopg2

    conn = psycopg2.connect(_DSN, connect_timeout=10)
    return conn


def load_db_municipality_coords(conn) -> dict[str, tuple[float, float]]:
    """Load coordinates from sc_municipalities table, keyed by normalized name."""
    cur = conn.cursor()
    cur.execute(
        "SELECT municipio, latitude, longitude FROM sc_municipalities "
        "WHERE latitude IS NOT NULL AND longitude IS NOT NULL"
    )
    result: dict[str, tuple[float, float]] = {}
    for row in cur.fetchall():
        name = _normalize_muni(row[0])
        lat, lon = row[1], row[2]
        if lat is not None and lon is not None:
            result[name] = (float(lat), float(lon))
    cur.close()
    return result


def update_sc_public_entities(
    conn,
    updates: list[tuple[float, float, float, str, str]],
) -> int:
    """Update latitude, longitude, distancia_fk, raio_200km for entities.

    Args:
        updates: list of (lat, lon, dist_km, within_flag, cnpj8)

    Returns:
        Number of rows updated.
    """
    cur = conn.cursor()
    count = 0
    for lat, lon, dist_km, within_flag, cnpj8 in updates:
        cur.execute(
            """UPDATE sc_public_entities
               SET latitude = %s,
                   longitude = %s,
                   distancia_fk = %s,
                   raio_200km = %s
               WHERE cnpj_8 = %s AND is_active = TRUE""",
            (lat, lon, dist_km, within_flag, cnpj8),
        )
        count += cur.rowcount
    conn.commit()
    cur.close()
    return count


# ---------------------------------------------------------------------------
# Main resolution logic
# ---------------------------------------------------------------------------

import openpyxl


def resolve() -> dict[str, Any]:
    """Main resolution pipeline."""
    # ── Step 0: Verify backup exists ───────────────────────────────────────
    if not BACKUP_PATH.exists():
        print("❌ Backup not found. Run backup first.")
        print(f"   Expected: {BACKUP_PATH}")
        sys.exit(1)

    print(f"📂 Seed file: {SEED_PATH}")
    print(f"📂 Backup:    {BACKUP_PATH}")

    # ── Step 1: Read spreadsheet ──────────────────────────────────────────
    print("\n📖 Reading spreadsheet...")
    wb = openpyxl.load_workbook(str(SEED_PATH), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    total_rows = len(rows)
    print(f"   Total rows: {total_rows}")

    # Identify unresolved entities (lat == 'N/D')
    unresolved_rows: list[dict[str, Any]] = []
    for i, row in enumerate(rows, start=2):  # 2-indexed for Excel row
        if not row or not row[0]:
            continue
        lat_raw = row[COL_LATITUDE] if len(row) > COL_LATITUDE else None
        if lat_raw == "N/D":
            muni = str(row[COL_MUNICIPIO]).strip() if len(row) > COL_MUNICIPIO and row[COL_MUNICIPIO] else ""
            cnpj8 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            unresolved_rows.append(
                {
                    "excel_row": i,
                    "razao": str(row[0]).strip(),
                    "cnpj8": cnpj8,
                    "municipio": muni,
                    "ibge": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                }
            )

    print(f"   Unresolved entities: {len(unresolved_rows)}")

    if not unresolved_rows:
        print("✅ No unresolved entities found. Nothing to do.")
        return {
            "resolved": 0,
            "total_unresolved": 0,
            "still_unresolved": 0,
            "failed_municipalities": [],
            "muni_coords": {},
        }

    # ── Step 2: Build unique municipality list ────────────────────────────
    unique_munis: dict[str, list[dict]] = {}
    for ent in unresolved_rows:
        muni = _normalize_muni(ent["municipio"])
        if muni:
            unique_munis.setdefault(muni, []).append(ent)

    print(f"\n🌍 Unique municipalities to geocode: {len(unique_munis)}")

    # ── Step 3: Load DB municipality coords (partial fallback) ────────────
    db_coords: dict[str, tuple[float, float]] = {}
    try:
        conn = _get_db_conn()
        db_coords = load_db_municipality_coords(conn)
        print(f"   DB sc_municipalities with coords: {len(db_coords)}")
        conn.close()
    except Exception as e:
        print(f"   ⚠️  Could not load DB coordinates: {e}")
        print("   Continuing with geocoding only...")

    # ── Step 4: Resolve coordinates for each unique municipality ──────────
    muni_coords: dict[str, tuple[float, float]] = {}
    muni_source: dict[str, str] = {}

    # First, check DB
    for muni_norm in list(unique_munis.keys()):
        if muni_norm in db_coords:
            muni_coords[muni_norm] = db_coords[muni_norm]
            muni_source[muni_norm] = "DB sc_municipalities"
            print(f"  ✅ DB: {muni_norm} → ({db_coords[muni_norm][0]:.4f}, {db_coords[muni_norm][1]:.4f})")

    # Now geocode the rest via Nominatim
    to_geocode = [m for m in unique_munis if m not in muni_coords]
    print(f"\n📍 Geocoding {len(to_geocode)} municipalities via Nominatim...")

    for idx, muni_norm in enumerate(to_geocode):
        # Use the original name (with proper accents) for geocoding
        original_name = unique_munis[muni_norm][0]["municipio"]
        print(f"   [{idx + 1}/{len(to_geocode)}] {original_name}...", end=" ", flush=True)

        coords = geocode_municipality(original_name)
        if coords:
            muni_coords[muni_norm] = coords
            muni_source[muni_norm] = "Nominatim"
            print(f"✅ ({coords[0]:.4f}, {coords[1]:.4f})")
        else:
            print("❌ FAILED")
            # Try again with normalized name
            time.sleep(NOMINATIM_DELAY)
            coords = geocode_municipality(muni_norm.title())
            if coords:
                muni_coords[muni_norm] = coords
                muni_source[muni_norm] = "Nominatim (retry)"
                print(f"   ✅ Retry: ({coords[0]:.4f}, {coords[1]:.4f})")
            else:
                print(f"   ❌ Could not geocode '{original_name}'")

        # Rate limit
        if idx < len(to_geocode) - 1:
            time.sleep(NOMINATIM_DELAY)

    # Report results
    resolved_munis = sum(1 for m in unique_munis if m in muni_coords)
    failed_munis = sum(1 for m in unique_munis if m not in muni_coords)
    print("\n📊 Geocoding summary:")
    print(f"   Resolved municipalities: {resolved_munis}/{len(unique_munis)}")
    if failed_munis > 0:
        print(f"   Failed municipalities: {failed_munis}")
        for m in unique_munis:
            if m not in muni_coords:
                print(f"     - {unique_munis[m][0]['municipio']}")

    # ── Step 5: Update spreadsheet cells ──────────────────────────────────
    print("\n📝 Updating spreadsheet...")
    update_count = 0
    skipped_no_coords = 0
    updates_for_db: list[tuple[float, float, float, str, str]] = []

    # Reload workbook in non-read_only mode for writing
    wb = openpyxl.load_workbook(str(SEED_PATH))
    ws = wb.active

    for ent in unresolved_rows:
        muni_norm = _normalize_muni(ent["municipio"])
        coords = muni_coords.get(muni_norm)

        if coords is None:
            skipped_no_coords += 1
            continue

        lat, lon = coords
        dist = haversine_km(FLORIANOPOLIS_LAT, FLORIANOPOLIS_LON, lat, lon)
        within = "SIM ✓" if dist <= RADIUS_KM else "NÃO ✗"
        dist_rounded = round(dist, 1)

        row_idx = ent["excel_row"]

        # Update cells
        ws.cell(row=row_idx, column=COL_LATITUDE + 1, value=lat)
        ws.cell(row=row_idx, column=COL_LONGITUDE + 1, value=lon)
        ws.cell(row=row_idx, column=COL_DISTANCIA + 1, value=dist_rounded)
        ws.cell(row=row_idx, column=COL_RAIO200 + 1, value=within)

        update_count += 1

        # Prepare DB update
        if ent["cnpj8"]:
            within_bool = "TRUE" if dist <= RADIUS_KM else "FALSE"
            updates_for_db.append((lat, lon, dist_rounded, within_bool, ent["cnpj8"]))

    # Save the spreadsheet
    wb.save(str(SEED_PATH))
    wb.close()
    print(f"   Rows updated in spreadsheet: {update_count}")
    print(f"   Skipped (no coords found):   {skipped_no_coords}")

    # ── Step 6: Update PostgreSQL sc_public_entities ───────────────────────
    print("\n🗄️  Updating PostgreSQL sc_public_entities...")
    try:
        conn = _get_db_conn()
        db_updated = update_sc_public_entities(conn, updates_for_db)
        conn.close()
        print(f"   Rows updated in DB: {db_updated}")
    except Exception as e:
        print(f"   ⚠️  DB update failed: {e}")
        print("   (Spreadsheet was already updated; DB can be patched later)")

    # ── Step 7: Summary ────────────────────────────────────────────────────
    still_unresolved = total_rows - update_count
    still_unresolved = max(0, len(unresolved_rows) - update_count)

    print(f"\n{'=' * 60}")
    print("  RESOLUTION SUMMARY")
    print(f"{'=' * 60}")
    print(f"  Total seed rows:      {total_rows}")
    print(f"  Initially unresolved: {len(unresolved_rows)}")
    print(f"  Resolved now:         {update_count}")
    print(f"  Still unresolved:     {still_unresolved}")
    print(f"{'=' * 60}")

    return {
        "total_rows": total_rows,
        "total_unresolved": len(unresolved_rows),
        "resolved": update_count,
        "still_unresolved": still_unresolved,
        "failed_municipalities": [m for m in unique_munis if m not in muni_coords],
        "muni_coords": {
            m: {"lat": c[0], "lon": c[1], "source": muni_source.get(m, "unknown")} for m, c in muni_coords.items()
        },
    }


def main() -> int:
    result = resolve()

    if result["still_unresolved"] > 0:
        print(f"\n⚠️  {result['still_unresolved']} entities still unresolved.")
        print("   These municipalities could not be geocoded:")
        for m in result["failed_municipalities"]:
            print(f"     - {m}")
        print("\n   Possible manual fixes:")
        print("   1. Add coordinates manually in the spreadsheet")
        print("   2. Try a different geocoding service")
        return 2

    print(f"\n✅ All {result['total_unresolved']} entities resolved successfully!")

    # ── Re-run consulting_readiness.py ────────────────────────────────────
    print("\n🔄 Re-running consulting_readiness.py to verify...")
    readiness_script = PROJECT_ROOT / "scripts" / "consulting_readiness.py"
    completed = subprocess.run([sys.executable, str(readiness_script)], check=False)
    exit_code = completed.returncode
    if exit_code == 0:
        print("\n✅ Readiness check PASSED")
    else:
        print(f"\n⚠️  Readiness check exit code: {exit_code} (expected 0=pass, 2=fail)")

    return exit_code


if __name__ == "__main__":
    sys.exit(main())
