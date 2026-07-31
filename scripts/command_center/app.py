"""FastAPI application for EXTRA Command Center."""

from __future__ import annotations

import json
import time
from typing import Any

from fastapi import Depends, FastAPI, Header, HTTPException, Query, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from scripts.command_center import CAMPAIGN_ID, TERMINAL_READY, __version__
from scripts.command_center.artifact_reader import list_recent_artifacts, read_artifact
from scripts.command_center.capabilities.registry import get_registry
from scripts.command_center.config import Settings, git_sha, load_settings
from scripts.command_center.job_runner import JobRunner, StartJobRequest
from scripts.command_center.overview import build_overview, build_search
from scripts.command_center.redaction import env_presence, redact_text
from scripts.command_center.security import issue_csrf_token, require_csrf, resolve_under_roots, safe_join
from scripts.command_center.store import Store

STARTED_AT = time.time()

# Server-owned human-confirmation defaults — never trust client payload for these.
DEFAULT_REVIEW_CONFIRMATION_PHRASE = (
    "Confirmo que revisei os dados e autorizo apenas a inclusão deste item na fila manual."
)
_CLIENT_FORBIDDEN_DECISION_KEYS = frozenset({"sensitive", "confirmation_phrase"})


def _neutralize_cell(value: Any) -> str:
    """Coerce spreadsheet cells to plain display strings; neutralize formula/script risks."""
    if value is None:
        return ""
    text = str(value)
    if text.startswith("=") or text.startswith("+") or text.startswith("-") or text.startswith("@"):
        text = "'" + text
    # Strip control chars that break HTML/layout; keep content readable.
    return "".join(ch if ch.isprintable() or ch in "\t\n" else " " for ch in text)[:500]


class ExecuteBody(BaseModel):
    capability_id: str = Field(..., min_length=1, max_length=200)
    params: dict[str, Any] = Field(default_factory=dict)
    confirmation: str | None = None
    actor: str = "local-user"


class DecisionBody(BaseModel):
    item_id: str
    decision: str
    rationale: str | None = None
    confirmation: str | None = None
    actor: str = "local-user"
    return_by: str | None = None
    artifact_hashes: dict[str, str] = Field(default_factory=dict)
    artifact_version: str | None = None
    payload: dict[str, Any] = Field(default_factory=dict)


class BundleBody(BaseModel):
    path: str
    include_logs: bool = False


class RegenerateBody(BaseModel):
    job_id: str
    item_id: str | None = None
    corrections: list[dict[str, Any]] = Field(default_factory=list)
    actor: str = "local-user"
    note: str | None = None


def _backend_decision_sensitivity(
    store: Store,
    registry: Any,
    item_id: str,
) -> tuple[bool, str]:
    """Sensitivity and phrase are determined only by backend (review item / capability).

    Client payload fields ``sensitive`` and ``confirmation_phrase`` are ignored.
    ACCEPT is always treated as sensitive for human review items.
    """
    review = store.get_review(item_id)
    cap = None
    if review and review.get("capability_id"):
        cap = registry.get(str(review["capability_id"]))
    if cap is None and item_id.upper().startswith("DOD"):
        return True, DEFAULT_REVIEW_CONFIRMATION_PHRASE
    if cap is not None:
        phrase = (cap.confirmation_phrase or DEFAULT_REVIEW_CONFIRMATION_PHRASE).strip()
        if not phrase:
            phrase = DEFAULT_REVIEW_CONFIRMATION_PHRASE
        # Human review ACCEPT is always sensitive; phrase may come from capability.
        return True, phrase
    # Default: pending review / unknown item → sensitive with default phrase
    return True, DEFAULT_REVIEW_CONFIRMATION_PHRASE


class PrefBody(BaseModel):
    key: str
    value: str


class ReviewEnqueueBody(BaseModel):
    title: str
    source: str = "manual"
    evidence: str = "Sem evidência anexada."
    limitations: str = "Limitações não informadas."
    risks: str = "Riscos não informados."
    job_id: str | None = None
    capability_id: str | None = None
    actor: str = "local-user"
    payload: dict[str, Any] = Field(default_factory=dict)


def create_app(settings: Settings | None = None) -> FastAPI:
    settings = settings or load_settings()
    settings.data_dir.mkdir(parents=True, exist_ok=True)
    settings.logs_dir.mkdir(parents=True, exist_ok=True)
    store = Store(settings.db_path)
    registry = get_registry()
    runner = JobRunner(settings, store, registry)

    app = FastAPI(
        title="EXTRA Command Center",
        version=__version__,
        docs_url="/api/docs",
        redoc_url=None,
    )
    app.state.settings = settings
    app.state.store = store
    app.state.registry = registry
    app.state.runner = runner

    app.add_middleware(
        CORSMiddleware,
        allow_origins=list(settings.cors_origins),
        allow_credentials=True,
        allow_methods=["GET", "POST", "OPTIONS"],
        allow_headers=["*"],
    )

    def csrf_dep(
        request: Request,
        x_cc_csrf: str | None = Header(default=None, alias="X-CC-CSRF"),
    ) -> None:
        # bind header into request for require_csrf
        require_csrf(
            request,
            cookie_name=settings.csrf_cookie_name,
            header_name=settings.csrf_header_name,
            x_csrf=x_cc_csrf,
        )

    @app.middleware("http")
    async def security_headers(request: Request, call_next: Any) -> Response:
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "no-referrer"
        response.headers["Cache-Control"] = "no-store"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        response.headers["Cross-Origin-Opener-Policy"] = "same-origin"
        # Local SPA: allow self scripts/styles; no remote CDNs required for the built app.
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self'; "
            "style-src 'self' 'unsafe-inline'; "
            "img-src 'self' data: blob:; "
            "font-src 'self' data:; "
            "connect-src 'self'; "
            "frame-ancestors 'none'; "
            "base-uri 'self'; "
            "form-action 'self'"
        )
        return response

    @app.get("/api/health")
    def health() -> dict[str, Any]:
        spa_ok = bool(settings.spa_dist and settings.spa_dist.exists())
        db_ok = True
        try:
            store.list_jobs(limit=1)
        except Exception:
            db_ok = False
        status = "ok" if db_ok else "degraded"
        return {
            "ok": db_ok,
            "status": status,
            "degraded": not db_ok,
            "service": "extra-command-center",
            "version": __version__,
            "campaign": CAMPAIGN_ID,
            "terminal_status": TERMINAL_READY,
            "sha": git_sha(),
            "uptime_sec": round(time.time() - STARTED_AT, 2),
            "host": settings.host,
            "port": settings.port,
            "services": {"sqlite": db_ok, "spa_dist": spa_ok},
            "bind": f"{settings.host}:{settings.port}",
            "public_bind": settings.host not in {"127.0.0.1", "localhost", "::1"},
            "sqlite": {"path": str(settings.db_path), "ok": settings.db_path.parent.exists()},
            "spa": {"configured": spa_ok, "path": str(settings.spa_dist) if settings.spa_dist else None},
            "jobs": store.job_counts(),
            "capabilities": {
                "total": len(registry.list()),
                "available": len(registry.available_ids()),
                "by_category": registry.categories_summary(),
            },
            "roots": [str(r) for r in settings.allowed_artifact_roots],
            "env": {
                "LOCAL_DATALAKE_DSN": env_presence("LOCAL_DATALAKE_DSN"),
                "OPENAI_API_KEY": env_presence("OPENAI_API_KEY"),
                "DATABASE_URL": env_presence("DATABASE_URL"),
            },
            "max_concurrent_jobs": settings.max_concurrent_jobs,
        }

    @app.get("/api/csrf")
    def csrf(response: Response) -> dict[str, str]:
        token = issue_csrf_token(response, settings.csrf_cookie_name)
        return {"csrf_token": token, "header": settings.csrf_header_name}

    @app.get("/api/overview")
    def overview() -> dict[str, Any]:
        return build_overview(settings, store, registry)

    @app.get("/api/capabilities")
    def capabilities(category: str | None = None) -> dict[str, Any]:
        items = registry.public_list()
        if category:
            items = [c for c in items if c["category"] == category]
        return {"capabilities": items}

    @app.get("/api/capabilities/{capability_id}")
    def capability_detail(capability_id: str) -> dict[str, Any]:
        cap = registry.get(capability_id)
        if not cap:
            raise HTTPException(404, "Capability não encontrada")
        return cap.public_dict()

    @app.post("/api/jobs")
    def start_job(body: ExecuteBody, _: None = Depends(csrf_dep)) -> dict[str, Any]:
        # Hard deny arbitrary command fields
        forbidden = {"command", "shell", "argv", "cmd", "executable"}
        if forbidden.intersection(body.params.keys()):
            raise HTTPException(400, "Parâmetros de comando arbitrário são proibidos.")
        try:
            rec = runner.start(
                StartJobRequest(
                    capability_id=body.capability_id,
                    params=body.params,
                    confirmation=body.confirmation,
                    actor=body.actor,
                )
            )
        except ValueError as exc:
            raise HTTPException(400, redact_text(str(exc))) from exc
        return {"job": rec.to_public()}

    @app.get("/api/jobs")
    def list_jobs(
        limit: int = 50,
        status: str | None = None,
        workspace_id: str | None = None,
        client_id: str | None = None,
    ) -> dict[str, Any]:
        jobs = store.list_jobs(
            limit=limit,
            status=status,
            workspace_id=workspace_id,
            client_id=client_id,
        )
        return {"jobs": [j.to_public() for j in jobs]}

    @app.get("/api/jobs/{job_id}")
    def get_job(job_id: str) -> dict[str, Any]:
        rec = store.get_job(job_id)
        if not rec:
            raise HTTPException(404, "Job não encontrado")
        return {"job": rec.to_public()}

    @app.post("/api/jobs/{job_id}/cancel")
    def cancel_job(job_id: str, _: None = Depends(csrf_dep)) -> dict[str, Any]:
        try:
            rec = runner.cancel(job_id)
        except KeyError as exc:
            raise HTTPException(404, "Job não encontrado") from exc
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
        return {"job": rec.to_public()}

    @app.get("/api/jobs/{job_id}/logs")
    def job_logs(job_id: str, after_id: int = 0, limit: int = 500) -> dict[str, Any]:
        if not store.get_job(job_id):
            raise HTTPException(404, "Job não encontrado")
        return {"logs": store.get_logs(job_id, after_id=after_id, limit=limit)}

    @app.get("/api/jobs/{job_id}/events")
    def job_events(job_id: str) -> StreamingResponse:
        if not store.get_job(job_id):
            raise HTTPException(404, "Job não encontrado")

        def gen() -> Any:
            q = runner.subscribe(job_id)
            try:
                while True:
                    try:
                        event = q.get(timeout=15)
                    except Exception:
                        yield "event: ping\ndata: {}\n\n"
                        continue
                    if event is None:
                        yield "event: end\ndata: {}\n\n"
                        break
                    yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
            finally:
                runner.unsubscribe(job_id, q)

        return StreamingResponse(gen(), media_type="text/event-stream")

    @app.get("/api/artifacts")
    def artifacts(path: str | None = None, recent: bool = False) -> dict[str, Any]:
        if recent or not path:
            return {"recent": list_recent_artifacts(settings)}
        return read_artifact(path, settings)

    @app.get("/api/artifacts/download")
    def download_artifact(path: str = Query(...)) -> FileResponse:
        resolved = resolve_under_roots(path, settings.allowed_artifact_roots)
        if not resolved.is_file():
            raise HTTPException(404, "Arquivo não encontrado")
        return FileResponse(resolved, filename=resolved.name)

    @app.get("/api/search")
    def search(q: str = "") -> dict[str, Any]:
        return build_search(q, store, registry, settings)

    @app.get("/api/audit")
    def audit(limit: int = 100) -> dict[str, Any]:
        return {"audit": store.list_audit(limit=limit)}

    @app.get("/api/decisions")
    def decisions(
        limit: int = 50,
        current_hashes: str | None = Query(
            None,
            description="JSON object of current artifact hashes to re-evaluate ACCEPT obsolescence",
        ),
        item_id: str | None = None,
    ) -> dict[str, Any]:
        from scripts.command_center.review_rules import invalidate_decisions_for_hashes

        items = store.list_decisions(limit=limit)
        if item_id:
            items = [d for d in items if d.get("item_id") == item_id]
        marked: list[str] = []
        if current_hashes:
            try:
                hashes = json.loads(current_hashes)
            except json.JSONDecodeError as exc:
                raise HTTPException(400, "current_hashes deve ser JSON object") from exc
            if not isinstance(hashes, dict):
                raise HTTPException(400, "current_hashes deve ser objeto")
            # persist obsolescence for matching ACCEPT rows
            if item_id:
                marked = store.mark_accepts_obsolete_for_item(item_id, current_hashes={str(k): str(v) for k, v in hashes.items()})
                items = store.list_decisions(limit=limit)
                items = [d for d in items if d.get("item_id") == item_id]
            else:
                # ephemeral view without item scope
                items = invalidate_decisions_for_hashes(items, {str(k): str(v) for k, v in hashes.items()})
        return {"decisions": items, "marked_obsolete": marked}

    @app.get("/api/reviews")
    def reviews(
        status: str | None = "pending",
        limit: int = 50,
        offset: int = 0,
    ) -> dict[str, Any]:
        """Human review queue — side-effect free. Does not create or mutate reviews."""
        lim = max(1, min(int(limit), 200))
        off = max(0, int(offset))
        items = store.list_reviews(status=status, limit=lim, offset=off)
        total = store.count_reviews(status=status)
        return {
            "reviews": items,
            "page_count": len(items),
            "total_count": total,
            "limit": lim,
            "offset": off,
            # Backward-compatible alias for older UI builds
            "count": total,
        }

    @app.post("/api/reviews/reconcile")
    def reconcile_reviews_api(
        _: None = Depends(csrf_dep),
        actor: str = "local-user",
    ) -> dict[str, Any]:
        """Idempotent reconciliation of BLOCKED_HUMAN jobs into the review queue."""
        result = store.reconcile_blocked_human_reviews(actor=actor)
        store.audit(actor, "review.reconcile.batch", result)
        return {"ok": True, **result}

    @app.post("/api/reviews")
    def enqueue_review_api(body: ReviewEnqueueBody, _: None = Depends(csrf_dep)) -> dict[str, Any]:
        """Explicitly enqueue a review item (local queue only — does not decide)."""
        title = body.title.strip()
        if not title:
            raise HTTPException(400, "title obrigatório")
        rid = store.enqueue_review(
            title=title,
            source=body.source,
            evidence=body.evidence,
            limitations=body.limitations,
            risks=body.risks,
            job_id=body.job_id,
            capability_id=body.capability_id,
            payload=body.payload,
        )
        store.audit(body.actor, "review.enqueue", {"id": rid})
        return {"ok": True, "id": rid}

    @app.post("/api/decisions")
    def post_decision(body: DecisionBody, _: None = Depends(csrf_dep)) -> dict[str, Any]:
        from scripts.command_center.review_rules import validate_decision_request

        decision = body.decision.upper().strip()
        if decision not in {"ACCEPT", "REJECT", "DEFER"}:
            raise HTTPException(400, "Decisão deve ser ACCEPT, REJECT ou DEFER.")
        review = store.get_review(body.item_id)
        title = (review or {}).get("title")
        presented_hashes = {}
        if review and isinstance(review.get("payload"), dict):
            presented_hashes = dict(review["payload"].get("artifact_hashes") or {})
            if review["payload"].get("content_hash") and "source" not in presented_hashes:
                presented_hashes["source"] = str(review["payload"]["content_hash"])
        bound_hashes = dict(body.artifact_hashes or {})
        if not bound_hashes and body.payload:
            bound_hashes = dict(body.payload.get("artifact_hashes") or {})
        # For ACCEPT, if client omitted hashes but review has them, require explicit echo
        rule_errors = validate_decision_request(
            decision=decision,
            rationale=body.rationale,
            return_by=body.return_by or (body.payload or {}).get("return_by"),
            artifact_hashes=bound_hashes if decision == "ACCEPT" else bound_hashes,
            presented_hashes=presented_hashes if decision == "ACCEPT" else None,
            title=str(title) if title else None,
        )
        if rule_errors:
            raise HTTPException(400, "; ".join(rule_errors))
        # Sensitivity + phrase: server-only (ignore client payload claims)
        sensitive, expected_phrase = _backend_decision_sensitivity(store, registry, body.item_id)
        if sensitive and decision == "ACCEPT":
            if not body.confirmation or body.confirmation.strip() != expected_phrase:
                raise HTTPException(
                    400,
                    f"Confirmação textual obrigatória para ACCEPT. Digite exatamente: {expected_phrase}",
                )
        # Never call dod accept automatically
        if body.item_id.upper().startswith("DOD") and decision == "ACCEPT":
            store.audit(
                body.actor,
                "decision.dod_accept_blocked",
                {"item_id": body.item_id, "note": "UI records intent only; use canonical controller manually"},
            )
            return {
                "ok": False,
                "blocked": True,
                "message": (
                    "O Command Center não aceita itens do DOD automaticamente. "
                    "A intenção foi auditada; use o controller canônico com gates."
                ),
            }
        # Strip client-controlled sensitivity fields before persistence
        safe_payload = {
            k: v for k, v in (body.payload or {}).items() if k not in _CLIENT_FORBIDDEN_DECISION_KEYS
        }
        safe_payload["sensitive"] = sensitive
        safe_payload["confirmation_phrase"] = expected_phrase if decision == "ACCEPT" else None
        safe_payload["sensitivity_source"] = "backend"
        safe_payload["artifact_hashes"] = bound_hashes
        safe_payload["artifact_version"] = body.artifact_version
        safe_payload["return_by"] = body.return_by or safe_payload.get("return_by")
        safe_payload["no_auto_outreach"] = True
        decision_id = store.save_decision(
            item_id=body.item_id,
            decision=decision,
            actor=body.actor,
            rationale=body.rationale,
            confirmation=body.confirmation,
            payload=safe_payload,
        )
        store.mark_review_decided(body.item_id, decision)
        store.audit(
            body.actor,
            "decision.save",
            {
                "decision_id": decision_id,
                "item_id": body.item_id,
                "decision": decision,
                "sensitive": sensitive,
                "artifact_hashes": bound_hashes,
            },
        )
        return {
            "ok": True,
            "decision_id": decision_id,
            "sensitive": sensitive,
            "confirmation_phrase": expected_phrase if decision == "ACCEPT" else None,
            "artifact_hashes": bound_hashes,
            "obsolete": False,
        }

    @app.get("/api/reviews/{item_id}/confirmation")
    def review_confirmation_requirements(item_id: str) -> dict[str, Any]:
        """Expose server-owned confirmation requirements for the review UI."""
        sensitive, phrase = _backend_decision_sensitivity(store, registry, item_id)
        review = store.get_review(item_id)
        return {
            "item_id": item_id,
            "sensitive": sensitive,
            "confirmation_phrase": phrase,
            "found": review is not None,
            "title": (review or {}).get("title"),
        }

    @app.get("/api/preferences/{key}")
    def get_pref(key: str) -> dict[str, Any]:
        return {"key": key, "value": store.get_pref(key)}

    @app.post("/api/preferences")
    def set_pref(body: PrefBody, _: None = Depends(csrf_dep)) -> dict[str, Any]:
        if body.key in {"env", "secrets", ".env"}:
            raise HTTPException(400, "Preferência proibida.")
        store.set_pref(body.key, body.value)
        return {"ok": True}

    @app.get("/api/workspaces")
    def list_workspaces() -> dict[str, Any]:
        """Local consulting workspace + projects (CC SQLite only).

        Also exposes fixed consulting filters (Extra / CONFENGE suppliers /
        CONFENGE agencies / process docs) for run isolation in the UI.
        """
        from scripts.command_center.store import CONSULTING_WORKSPACES

        ws = store.ensure_default_workspace()
        projects = store.list_projects(ws["id"])
        stored = store.list_workspaces()
        # Prefer fixed consulting filters for operational UI; keep stored CRM-lite rows
        consulting = list(CONSULTING_WORKSPACES)
        return {
            "workspace": ws,
            "workspaces": consulting + [w for w in stored if w.get("id") not in {c["id"] for c in consulting}],
            "projects": projects,
            "consulting_filters": consulting,
        }

    @app.post("/api/workspaces")
    def create_workspace(body: dict[str, Any], _: None = Depends(csrf_dep)) -> dict[str, Any]:
        name = str(body.get("name") or "").strip()
        if len(name) < 2:
            raise HTTPException(400, "Nome do workspace obrigatório.")
        ws = store.create_workspace(name, meta=body.get("meta") if isinstance(body.get("meta"), dict) else None)
        return {"ok": True, "workspace": ws}

    @app.post("/api/projects")
    def create_project(body: dict[str, Any], _: None = Depends(csrf_dep)) -> dict[str, Any]:
        wid = str(body.get("workspace_id") or "").strip()
        name = str(body.get("name") or "").strip()
        front = str(body.get("client_front") or "").strip()
        if not wid or len(name) < 2 or not front:
            raise HTTPException(400, "workspace_id, name e client_front são obrigatórios.")
        proj = store.create_project(workspace_id=wid, name=name, client_front=front)
        return {"ok": True, "project": proj}

    @app.get("/api/workflows")
    def workflows() -> dict[str, Any]:
        from scripts.command_center.workflows.catalog import list_workflows

        items = []
        for wf in list_workflows():
            items.append(
                {
                    "id": wf.id,
                    "title": wf.title,
                    "subtitle": wf.subtitle,
                    "client_id": wf.client_id,
                    "client_label": wf.client_label,
                    "outcome": wf.outcome,
                    "description": wf.description,
                    "steps": wf.steps,
                    "expected_deliverables": wf.expected_deliverables,
                    "params": [
                        {
                            "name": p.name,
                            "label": p.label,
                            "type": p.type,
                            "required": p.required,
                            "default": p.default,
                            "choices": p.choices,
                            "description": p.description,
                            "advanced": p.advanced,
                        }
                        for p in wf.params
                    ],
                    "limitations": wf.limitations,
                    "no_outreach": wf.no_outreach,
                    "href": f"/work/start/{wf.id}",
                    "capability_id": wf.id if wf.id.startswith("workflow.") else None,
                }
            )
        return {"workflows": items}

    @app.get("/api/workflows/{workflow_id}")
    def workflow_detail(workflow_id: str) -> dict[str, Any]:
        from scripts.command_center.workflows.catalog import get_workflow

        wf = get_workflow(workflow_id)
        if wf is None:
            raise HTTPException(404, "Fluxo não encontrado")
        return {
            "id": wf.id,
            "title": wf.title,
            "subtitle": wf.subtitle,
            "client_label": wf.client_label,
            "outcome": wf.outcome,
            "description": wf.description,
            "steps": wf.steps,
            "expected_deliverables": wf.expected_deliverables,
            "params": [
                {
                    "name": p.name,
                    "label": p.label,
                    "type": p.type,
                    "required": p.required,
                    "default": p.default,
                    "choices": p.choices,
                    "description": p.description,
                    "advanced": p.advanced,
                }
                for p in wf.params
            ],
            "limitations": wf.limitations,
            "preflight": {
                "objective": wf.outcome,
                "client": wf.client_label,
                "steps": wf.steps,
                "expected_deliverables": wf.expected_deliverables,
                "limitations": wf.limitations,
                "no_auto_outreach": True,
                "effects": "Somente geração local de arquivos sob data/command_center e output permitidos.",
            },
            "execution_modes": {
                "REAL": "Pipelines canônicos do extra-cli (preflight tipado; fail-closed).",
                "FIXTURE": "Demonstração explícita — não é evidência comercial/LIVE.",
            },
        }

    @app.get("/api/workflows/{workflow_id}/preflight")
    def workflow_preflight(
        workflow_id: str,
        data_mode: str = Query("REAL"),
    ) -> dict[str, Any]:
        """Typed operational preflight for REAL mode (or informational for FIXTURE)."""
        from scripts.command_center.adapters import preflight_workflow, resolve_data_mode
        from scripts.command_center.workflows.catalog import get_workflow

        wf = get_workflow(workflow_id)
        if wf is None:
            raise HTTPException(404, "Fluxo não encontrado")
        mode = resolve_data_mode({"data_mode": data_mode})
        if mode.value == "FIXTURE":
            return {
                "status": "READY",
                "safe_to_run": True,
                "data_mode": "FIXTURE",
                "checks": [
                    {
                        "name": "demo_mode",
                        "ok": True,
                        "detail": "Modo demonstração explícito — fixture permitida",
                        "required": True,
                    }
                ],
                "limitations": list(wf.limitations)
                + ["FIXTURE não prova LIVE; não alimenta outreach real."],
                "message": "Preflight FIXTURE READY (demonstração).",
                "capability_id": workflow_id,
            }
        out = settings.jobs_dir / "_preflight" / workflow_id
        result = preflight_workflow(workflow_id, {"data_mode": "REAL"}, out_dir=out)
        payload = result.to_dict()
        payload["data_mode"] = "REAL"
        return payload

    @app.post("/api/export-bundle")
    def export_bundle(body: BundleBody, _: None = Depends(csrf_dep)) -> dict[str, Any]:

        from scripts.command_center.export_bundle import build_export_bundle

        root = resolve_under_roots(body.path, settings.allowed_artifact_roots)
        try:
            if root.is_file() and root.name == "run-manifest.json":
                run_dir = root.parent
            elif root.is_dir():
                run_dir = root
            else:
                raise HTTPException(400, "Informe o diretório do run ou o run-manifest.json")
            result = build_export_bundle(run_dir, include_logs=body.include_logs)
        except FileNotFoundError as exc:
            raise HTTPException(404, str(exc)) from exc
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
        return result

    @app.get("/api/artifacts/preview-xlsx")
    def preview_xlsx(
        path: str = Query(...),
        sheet: str | None = None,
        offset: int = 0,
        limit: int = 100,
    ) -> dict[str, Any]:
        """In-browser XLSX preview: sheets, headers, windowed rows (read-only)."""

        from openpyxl import load_workbook

        resolved = resolve_under_roots(path, settings.allowed_artifact_roots)
        suffix = resolved.suffix.lower()
        if suffix == ".xls":
            raise HTTPException(
                400,
                "Arquivos .xls (Excel 97–2003) não são suportados na pré-visualização. "
                "Converta para .xlsx e tente novamente.",
            )
        if suffix != ".xlsx":
            raise HTTPException(400, "Somente arquivos .xlsx são aceitos na pré-visualização.")
        if not resolved.is_file():
            raise HTTPException(404, "Arquivo não encontrado")
        max_bytes = settings.max_artifact_read_bytes * 4
        if resolved.stat().st_size > max_bytes:
            raise HTTPException(413, "Planilha grande demais para pré-visualização; use o download.")

        offset = max(0, int(offset))
        limit = max(1, min(int(limit), 200))
        max_cols = 64
        # Scan at most this many data rows for total_rows estimate (windowed preview).
        max_scan = min(50_000, offset + limit + 1)

        wb = None
        try:
            wb = load_workbook(resolved, read_only=True, data_only=True)
            sheet_names = list(wb.sheetnames)
            target = sheet if sheet in sheet_names else sheet_names[0]
            ws = wb[target]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                raw_headers = next(rows_iter)
            except StopIteration:
                return {
                    "kind": "xlsx",
                    "path": str(resolved),
                    "name": resolved.name,
                    "sheets": sheet_names,
                    "sheet": target,
                    "headers": [],
                    "rows": [],
                    "offset": 0,
                    "limit": limit,
                    "total_rows": 0,
                    "previewable": True,
                    "downloadable": True,
                }

            headers = [
                _neutralize_cell(c) for c in list(raw_headers)[:max_cols]
            ]
            # Pad empty header names for stable keys
            headers = [h if h else f"col_{i+1}" for i, h in enumerate(headers)]

            page: list[list[str]] = []
            total_seen = 0
            for idx, row in enumerate(rows_iter):
                total_seen += 1
                if idx < offset:
                    if total_seen >= max_scan:
                        break
                    continue
                if len(page) < limit:
                    cells = list(row)[:max_cols]
                    page.append([_neutralize_cell(c) for c in cells])
                if total_seen >= max_scan and len(page) >= limit:
                    break

            total_rows = total_seen
            truncated = total_seen >= max_scan
            return {
                "kind": "xlsx",
                "path": str(resolved),
                "name": resolved.name,
                "sheets": sheet_names,
                "sheet": target,
                "headers": headers,
                "rows": [dict(zip(headers, r + [""] * max(0, len(headers) - len(r)), strict=False)) for r in page],
                "offset": offset,
                "limit": limit,
                "total_rows": total_rows,
                "total_rows_truncated": truncated,
                "previewable": True,
                "downloadable": True,
            }
        finally:
            if wb is not None:
                wb.close()

    @app.get("/api/jobs/{job_id}/manifest")
    def job_manifest(job_id: str) -> dict[str, Any]:
        from pathlib import Path

        from scripts.command_center.run_manifest import load_manifest, validate_manifest

        rec = store.get_job(job_id)
        if rec is None:
            raise HTTPException(404, "Job não encontrado")
        candidates: list[Path] = []
        for m in rec.manifests or []:
            candidates.append(Path(m))
        job_dir = settings.jobs_dir / job_id
        candidates.append(job_dir / "deliverables" / "run-manifest.json")
        candidates.append(job_dir / "run-manifest.json")
        for c in candidates:
            if c.is_file():
                data = load_manifest(c)
                return {
                    "path": str(c),
                    "valid": not validate_manifest(data),
                    "errors": validate_manifest(data),
                    "manifest": data,
                }
        raise HTTPException(404, "run-manifest não encontrado para este job")

    @app.get("/api/runs/compare")
    def compare_runs(
        current: str = Query(..., description="Path to current run-manifest.json"),
        previous: str | None = Query(None, description="Optional previous run-manifest.json"),
        workflow_id: str | None = Query(None),
    ) -> dict[str, Any]:
        """What changed since the last cycle — deterministic row/artifact diff."""
        from pathlib import Path

        from scripts.command_center.run_compare import compare_manifests, find_previous_manifest

        curr = resolve_under_roots(current, settings.allowed_artifact_roots)
        if not curr.is_file():
            raise HTTPException(404, "Manifest atual não encontrado")
        prev_path: Path | None = None
        if previous:
            prev_path = resolve_under_roots(previous, settings.allowed_artifact_roots)
            if not prev_path.is_file():
                raise HTTPException(404, "Manifest anterior não encontrado")
        else:
            prev_path = find_previous_manifest(
                workflow_id=workflow_id,
                current_manifest=curr,
                jobs_dir=settings.jobs_dir,
            )
            # also search data_dir jobs
            if prev_path is None:
                prev_path = find_previous_manifest(
                    workflow_id=workflow_id,
                    current_manifest=curr,
                    jobs_dir=settings.data_dir / "jobs",
                )
        if prev_path is None:
            return {
                "ok": True,
                "has_previous": False,
                "message": "Não há execução anterior comparável para este fluxo.",
                "current": str(curr),
            }
        try:
            diff = compare_manifests(prev_path, curr)
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
        return {
            "ok": True,
            "has_previous": True,
            "previous_path": str(prev_path),
            "current_path": str(curr),
            "diff": diff,
        }

    @app.get("/api/runs/recent-by-workflow")
    def recent_by_workflow(workflow_id: str = Query(...), limit: int = 10) -> dict[str, Any]:
        """List recent finished jobs for a workflow (for compare UI)."""
        from pathlib import Path

        from scripts.command_center.run_manifest import load_manifest

        limit = max(1, min(limit, 50))
        items: list[dict[str, Any]] = []
        for job in store.list_jobs(limit=200):
            if job.capability_id != workflow_id:
                continue
            if job.status not in {"SUCCEEDED", "SUCCEEDED_WITH_WARNINGS", "PARTIAL"}:
                continue
            man_path = None
            for m in job.manifests or []:
                if Path(m).is_file():
                    man_path = m
                    break
            if not man_path:
                cand = settings.jobs_dir / job.job_id / "deliverables" / "run-manifest.json"
                if cand.is_file():
                    man_path = str(cand)
            if not man_path:
                continue
            run_id = job.run_id
            try:
                mf = load_manifest(Path(man_path))
                run_id = mf.get("run_id") or run_id
            except (OSError, ValueError, KeyError, TypeError):
                run_id = job.run_id
            items.append(
                {
                    "job_id": job.job_id,
                    "run_id": run_id,
                    "status": job.status,
                    "finished_at": job.finished_at,
                    "manifest_path": man_path,
                    "action": job.action,
                }
            )
            if len(items) >= limit:
                break
        return {"workflow_id": workflow_id, "runs": items}

    @app.post("/api/reviews/regenerate")
    def regenerate_after_correction(body: RegenerateBody, _: None = Depends(csrf_dep)) -> dict[str, Any]:
        """Apply human corrections, regenerate deliverables, obsolete prior ACCEPTs."""
        from pathlib import Path

        from scripts.command_center.regenerate import regenerate_workflow_version
        from scripts.command_center.run_manifest import load_manifest

        rec = store.get_job(body.job_id)
        if rec is None:
            raise HTTPException(404, "Job original não encontrado")
        workflow_id = rec.capability_id
        if not str(workflow_id).startswith("workflow."):
            raise HTTPException(400, "Regeneração guiada só para workflow.*")

        parent_manifest = None
        for m in rec.manifests or []:
            if Path(m).is_file():
                parent_manifest = Path(m)
                break
        if parent_manifest is None:
            cand = settings.jobs_dir / body.job_id / "deliverables" / "run-manifest.json"
            if cand.is_file():
                parent_manifest = cand
        parent_run_id = rec.run_id
        prior_source = None
        if parent_manifest:
            try:
                pm = load_manifest(parent_manifest)
                parent_run_id = pm.get("run_id") or parent_run_id
            except (OSError, ValueError, TypeError, KeyError):
                pass
            for name in ("opportunities.json", "suppliers.json", "public_agencies.json", "documents-index.json"):
                p = parent_manifest.parent / name
                if p.is_file():
                    prior_source = p
                    break

        new_job_id = str(__import__("uuid").uuid4())
        out_dir = settings.jobs_dir / new_job_id / "deliverables"
        try:
            result = regenerate_workflow_version(
                workflow_id=workflow_id,
                params=dict(rec.params or {}),
                out_dir=out_dir,
                code_sha=rec.code_sha or git_sha(),
                job_id=new_job_id,
                parent_run_id=parent_run_id,
                corrections=list(body.corrections or []),
                prior_source=prior_source,
            )
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc

        # Persist as a new job record (history preserved; old job untouched)
        from datetime import UTC, datetime

        from scripts.command_center.status_normalize import JobState
        from scripts.command_center.store import JobRecord

        finished = datetime.now(UTC).isoformat()
        new_rec = JobRecord(
            job_id=new_job_id,
            capability_id=workflow_id,
            action=f"Regeneração: {rec.action}",
            params=dict(rec.params or {}),
            status=JobState.SUCCEEDED.value,
            human_message=result.get("message") or "Entregável regenerado após correção humana.",
            attention="ok",
            finished_at=finished,
            started_at=finished,
            exit_code=0,
            canonical_command=list(rec.canonical_command or []),
            artifacts=list(result.get("artifacts") or []),
            manifests=[result["manifest_path"]] if result.get("manifest_path") else [],
            run_id=result.get("run_id"),
            code_sha=rec.code_sha,
            output_paths=list(result.get("artifacts") or []),
            stdout_path=str(settings.jobs_dir / new_job_id / "stdout.log"),
            stderr_path=str(settings.jobs_dir / new_job_id / "stderr.log"),
        )
        (settings.jobs_dir / new_job_id).mkdir(parents=True, exist_ok=True)
        Path(new_rec.stdout_path or "").write_text(
            json.dumps({"regenerated_from": body.job_id, "run_id": result.get("run_id")}, indent=2) + "\n",
            encoding="utf-8",
        )
        store.create_job(new_rec)
        store.audit(
            body.actor,
            "job.regenerated",
            {
                "parent_job_id": body.job_id,
                "new_job_id": new_job_id,
                "parent_run_id": parent_run_id,
                "run_id": result.get("run_id"),
                "corrections": body.corrections,
            },
        )

        # Invalidate prior ACCEPT decisions tied to item when hashes changed
        marked: list[str] = []
        new_hashes = result.get("content_hashes") or {}
        if body.item_id and new_hashes:
            marked = store.mark_accepts_obsolete_for_item(
                body.item_id,
                current_hashes={str(k): str(v) for k, v in new_hashes.items()},
                reason="regenerated_after_correction",
            )
            # enqueue new review bound to new hashes
            store.enqueue_review(
                title=f"Re-revisar após correção: {rec.action}",
                source=workflow_id,
                evidence=f"Nova versão run_id={result.get('run_id')}; parent={parent_run_id}",
                limitations="Versão regenerada após correção humana; decisão anterior fica no histórico.",
                risks="Aceitar sem re-ler o entregável pode validar conteúdo não revisado.",
                job_id=new_job_id,
                capability_id=workflow_id,
                payload={
                    "parent_job_id": body.job_id,
                    "parent_run_id": parent_run_id,
                    "artifact_hashes": new_hashes,
                    "content_hash": new_hashes.get("source") or new_hashes.get("manifest"),
                    "corrections": body.corrections,
                },
            )

        return {
            "ok": True,
            "parent_job_id": body.job_id,
            "job_id": new_job_id,
            "run_id": result.get("run_id"),
            "parent_run_id": parent_run_id,
            "manifest_path": result.get("manifest_path"),
            "artifacts": result.get("artifacts") or [],
            "content_hashes": new_hashes,
            "marked_obsolete_decisions": marked,
            "version_meta_path": result.get("version_meta_path"),
        }

    @app.get("/api/onboarding")
    def onboarding() -> dict[str, Any]:
        import shutil
        import sys

        node = shutil.which("node")
        return {
            "python": {"ok": True, "version": sys.version.split()[0]},
            "node": {"ok": bool(node), "path": node},
            "spa_built": bool(settings.spa_dist and settings.spa_dist.exists()),
            "data_dir": str(settings.data_dir),
            "env": {
                "LOCAL_DATALAKE_DSN": env_presence("LOCAL_DATALAKE_DSN"),
                "OPENAI_API_KEY": env_presence("OPENAI_API_KEY"),
            },
            "capabilities_available": len(registry.available_ids()),
            "capabilities_total": len(registry.list()),
            "required_to_open_ui": ["python", "fastapi"],
            "notes": [
                "DSN e secrets nunca são exibidos — apenas status configurada/ausente/inválida.",
                "Capabilities ausentes degradam sem quebrar a UI.",
            ],
        }

    # SPA static (single URL mode)
    if settings.spa_dist and settings.spa_dist.exists():
        assets = settings.spa_dist / "assets"
        if assets.exists():
            app.mount("/assets", StaticFiles(directory=str(assets)), name="assets")

        @app.get("/")
        def index() -> FileResponse:
            return FileResponse(settings.spa_dist / "index.html")

        @app.get("/{full_path:path}")
        def spa_fallback(full_path: str) -> FileResponse:
            if full_path.startswith("api/"):
                raise HTTPException(404)
            # Same containment as artifacts: resolve and prove path stays under dist.
            try:
                candidate = safe_join(settings.spa_dist, full_path)
            except HTTPException:
                return FileResponse(settings.spa_dist / "index.html")
            if candidate.is_file():
                return FileResponse(candidate)
            return FileResponse(settings.spa_dist / "index.html")
    else:

        @app.get("/")
        def index_fallback() -> JSONResponse:
            return JSONResponse(
                {
                    "service": "extra-command-center",
                    "message": "API online. Frontend ainda não buildado (apps/command-center/dist).",
                    "health": "/api/health",
                    "docs": "/api/docs",
                }
            )

    return app


def _loopback_sockets(port: int) -> list[Any]:
    """Listen on IPv4 and IPv6 loopback only (never 0.0.0.0 / ::).

    Browsers often resolve ``localhost`` to ``::1``. Binding only ``127.0.0.1``
    makes ``http://localhost:8765`` fail to load while ``http://127.0.0.1:8765`` works.
    """
    import socket

    sockets: list[Any] = []
    targets: list[tuple[int, str]] = [
        (socket.AF_INET, "127.0.0.1"),
        (socket.AF_INET6, "::1"),
    ]
    for family, host in targets:
        try:
            sock = socket.socket(family, socket.SOCK_STREAM)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            if family == socket.AF_INET6 and hasattr(socket, "IPPROTO_IPV6"):
                # Keep v6-only on this socket; v4 has its own listener.
                sock.setsockopt(socket.IPPROTO_IPV6, socket.IPV6_V6ONLY, 1)
            sock.bind((host, port))
            sock.listen(2048)
            sockets.append(sock)
        except OSError as exc:
            # IPv6 may be disabled; IPv4-only is still usable.
            import sys

            print(f"WARN: could not bind loopback {host}:{port}: {exc}", file=sys.stderr)
    return sockets


def main() -> None:
    import uvicorn

    settings = load_settings()
    app = create_app(settings)
    host = settings.host
    port = settings.port

    # Dual loopback when staying on the default local-only host (never public binds).
    use_dual = host in {"127.0.0.1", "localhost", "::1"}
    if use_dual:
        socks = _loopback_sockets(port)
        if not socks:
            raise SystemExit(f"Failed to bind any loopback interface on port {port}")
        print(f"==> EXTRA Command Center listening on loopback port {port}")
        print(f"    http://127.0.0.1:{port}")
        print(f"    http://localhost:{port}")
        config = uvicorn.Config(app, log_level="info", access_log=True)
        server = uvicorn.Server(config)
        server.run(sockets=socks)
        return

    # Explicit host (still refused if public unless CC_ALLOW_PUBLIC_BIND=1 via load_settings)
    print(f"==> EXTRA Command Center http://{host}:{port}")
    uvicorn.run(app, host=host, port=port, log_level="info")


if __name__ == "__main__":
    main()
