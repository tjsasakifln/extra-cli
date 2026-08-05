"""Multi-sheet Excel export for commercial reajuste queue (supplier-first v2)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from scripts.commercial.reajuste_14133.export.reports import (
    FIELD_DICTIONARY,
    lead_flat_row,
    supplier_flat_row,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl required: pip install openpyxl") from exc


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FORMAT = "#,##0.00"
DATE_FORMAT = "YYYY-MM-DD"

# openpyxl rejects ASCII control chars (except tab/lf/cr)
_ILLEGAL_XML = __import__("re").compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _cell_value(v: Any) -> Any:
    if isinstance(v, str):
        return _ILLEGAL_XML.sub("", v)
    if isinstance(v, (list, dict)):
        return _ILLEGAL_XML.sub("", str(v))
    return v


def _write_sheet(ws, rows: list[dict[str, Any]], *, freeze: bool = True) -> None:
    if not rows:
        ws.append(["(sem registros)"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in rows:
        ws.append([_cell_value(row.get(h)) for h in headers])
    for i, h in enumerate(headers, start=1):
        width = min(48, max(12, len(str(h)) + 2))
        if h in {
            "objeto",
            "argumento_comercial",
            "mensagem_abordagem",
            "evidencias_favoraveis",
            "lacunas",
            "riscos",
        }:
            width = 48
        ws.column_dimensions[get_column_letter(i)].width = width
    money_cols = {
        i
        for i, h in enumerate(headers, start=1)
        if any(x in h for x in ("valor", "saldo", "teto", "base_potencialmente", "portfolio"))
    }
    date_cols = {
        i for i, h in enumerate(headers, start=1) if "data" in h or h in {"vigencia_final"}
    }
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, cell in enumerate(r, start=1):
            if i in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FORMAT
            if i in date_cols and cell.value:
                cell.number_format = DATE_FORMAT
    if freeze:
        ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_workbook(out_dir: Path, run: dict[str, Any]) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "leads_fornecedores_reajuste_14133.xlsx"
    # also keep legacy name
    legacy = out_dir / "leads_reajuste_14133.xlsx"
    wb = Workbook()

    def flat_contracts(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [lead_flat_row(x) for x in items]

    def flat_suppliers(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [supplier_flat_row(x) for x in items]

    portfolios = run.get("supplier_portfolios") or []
    all_leads = run.get("leads") or []

    ready = [p for p in portfolios if p.get("outreach_status") == "OUTREACH_READY"]
    ready_nv = [
        p
        for p in portfolios
        if p.get("outreach_status") == "OUTREACH_READY_WITHOUT_VALUE_ESTIMATE"
    ]
    doc_req = [
        p for p in portfolios if p.get("outreach_status") == "DOCUMENT_REQUEST_CANDIDATE"
    ]
    not_ready = [
        p for p in portfolios if p.get("outreach_status") == "NOT_READY_FOR_OUTREACH"
    ]
    sul = [p for p in portfolios if p.get("sul_priority")]
    already = [lead for lead in all_leads if lead.get("classificacao") == "ALREADY_ADJUSTED"]
    conflict = [
        lead
        for lead in all_leads
        if lead.get("classificacao") == "LEGAL_REGIME_CONFLICT"
        or lead.get("regime_legal") == "LEGAL_REGIME_CONFLICT"
    ]
    outliers = [
        lead
        for lead in all_leads
        if lead.get("value_quality_status")
        in {"VALUE_OUTLIER_REQUIRES_REVIEW", "VALUE_CONFLICT", "VALUE_UNUSABLE"}
    ]

    sheets: list[tuple[str, list[dict[str, Any]]]] = [
        ("OUTREACH_READY", flat_suppliers(ready)),
        ("READY_NO_VALUE", flat_suppliers(ready_nv)),
        ("DOCUMENT_REQUEST", flat_suppliers(doc_req)),
        ("SUL_SC_PRIORITY", flat_suppliers(sul)),
        ("NACIONAL", flat_suppliers(portfolios)),
        ("SUPPLIER_PORTFOLIOS", flat_suppliers(portfolios)),
        ("CONTRACT_LEVEL", flat_contracts(all_leads[:2000])),
        ("ALREADY_ADJUSTED", flat_contracts(already)),
        ("LEGAL_REGIME_CONFLICT", flat_contracts(conflict)),
        ("VALUE_OUTLIERS", flat_contracts(outliers)),
        ("NOT_READY", flat_suppliers(not_ready)),
    ]

    funnel = run.get("funnel") or {}
    dq = [{"metric": k, "value": v} for k, v in funnel.items()]
    metrics = run.get("metrics") or {}
    dq.extend({"metric": f"metrics.{k}", "value": v} for k, v in metrics.items())
    for k, v in (run.get("distributions") or {}).items():
        dq.append({"metric": f"dist.{k}", "value": str(v)})
    sheets.append(("DATA_QUALITY", dq))

    meth = [
        {"item": "escopo", "valor": "Reajuste em sentido estrito — Lei 14.133/2021"},
        {"item": "unidade_comercial", "valor": "Fornecedor (CNPJ) com contratos vinculados"},
        {
            "item": "exclui",
            "valor": "Reequilíbrio, repactuação, atualização por atraso, aditivo quantitativo",
        },
        {
            "item": "data_base",
            "valor": "Orçamento estimado confirmado; proxy só TEMPORAL_CANDIDATE_BY_PROXY",
        },
        {
            "item": "outreach",
            "valor": "OUTREACH_READY / READY_WITHOUT_VALUE / DOCUMENT_REQUEST / NOT_READY",
        },
        {
            "item": "pdf_rule",
            "valor": "PDF binário localizado ≠ TEXT_EXTRACTED ≠ gate documental",
        },
        {"item": "as_of", "valor": run.get("as_of")},
        {"item": "module_version", "valor": run.get("module_version")},
        {"item": "source_mode", "valor": run.get("source_mode")},
        {"item": "terminal_status", "valor": run.get("terminal_status")},
        {"item": "git_sha", "valor": run.get("git_sha")},
        {
            "item": "pagination",
            "valor": (run.get("params") or {}).get("pagination") or "keyset",
        },
    ]
    sheets.append(("METHODOLOGY", meth))
    fd = [{"campo": k, "descricao": v} for k, v in FIELD_DICTIONARY]
    fd.extend(
        [
            {
                "campo": "outreach_status",
                "descricao": "Gate comercial distinto da classificação jurídica",
            },
            {
                "campo": "value_quality_status",
                "descricao": "VALUE_CONFIRMED|PLAUSIBLE|OUTLIER|CONFLICT|UNUSABLE",
            },
            {
                "campo": "temporal_layer",
                "descricao": "TEMPORAL_CANDIDATE_BY_PROXY vs TEMPORAL_ELIGIBILITY_CONFIRMED",
            },
        ]
    )
    sheets.append(("FIELD_DICTIONARY", fd))

    first = True
    for name, rows in sheets:
        if first:
            ws = wb.active
            ws.title = name[:31]
            first = False
        else:
            ws = wb.create_sheet(name[:31])
        _write_sheet(ws, rows)

    wb.save(path)
    try:
        import shutil

        shutil.copy2(path, legacy)
    except OSError:
        pass
    return path
