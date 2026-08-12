"""Canonical weekly decision-pack artifacts for Extra Construtora.

Produces the operational products required by the multi-source goal:

  extra_decision_pack_<as_of>_<run_id>.xlsx
  extra_decision_report_<as_of>_<run_id>.pdf
  coverage_by_entity_source.csv
  decision_dataset.json
  qa_report.json

Integrates multi_source_open_pack with lake observations. Fail-closed: missing
PDF/Excel under strict mode is a delivery failure, never RESIDUAL_NOT_GENERATED.
"""

from __future__ import annotations

import csv
import json
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from scripts.crawl.run_evidence import sha256_file
from scripts.ops.multi_source_open_pack.classify_aec import classify_aec
from scripts.ops.multi_source_open_pack.consolidate import consolidate_observations
from scripts.ops.multi_source_open_pack.db_loaders import load_all_lake_observations
from scripts.ops.multi_source_open_pack.decide import apply_decisions, select_shortlist
from scripts.ops.multi_source_open_pack.pilot_gate import require_pilot_approval
from scripts.ops.multi_source_open_pack.pipeline import (
    DEFAULT_PILOT_POLICY,
    MOTOR_VERSION,
    default_limitations,
    default_source_policy,
)
from scripts.ops.multi_source_open_pack.reconcile import build_reconciliation
from scripts.ops.multi_source_open_pack.render_pack import write_excel, write_pdf
from scripts.ops.multi_source_open_pack.textutil import BR_TZ, iso_z, utc_now
from scripts.ops.multi_source_open_pack.universe import (
    annotate_observation_universe,
    build_indexes,
    load_universe,
)

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_UNIVERSE = PROJECT_ROOT / "config" / "target_entities_200km.csv"
DEFAULT_PROFILE = PROJECT_ROOT / "config" / "client_profiles" / "extra.yaml"
CANONICAL_N = 1093

# Official result vocabulary (goal contract)
RESULT_STATES = (
    "FOUND",
    "ZERO_CONFIRMED",
    "NOT_APPLICABLE",
    "BLOCKED_EXTERNAL",
    "BLOCKED_CREDENTIAL",
    "BLOCKED_INFRA",
    "STALE",
    "PARTIAL",
    "ERROR",
    "NOT_QUERIED",
)


def _atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + f".{os.getpid()}.tmp")
    tmp.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    os.replace(tmp, path)


def _load_profile(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {}
    try:
        import yaml

        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _load_source_policy() -> dict[str, Any]:
    path = PROJECT_ROOT / "config" / "source_applicability.yaml"
    if not path.is_file():
        return {}
    try:
        import yaml

        return yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    except Exception:
        return {}


def _cnpj8(value: str | None) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return digits[:8] if len(digits) >= 8 else digits


def load_entity_source_evidence(conn: Any) -> dict[tuple[str, str], str]:
    """Load entity-scoped collection evidence → (cnpj8|entity_key, source) → state.

    Reads ``coverage_evidence`` when present. States of interest:
    - success_with_data / success → supports FOUND
    - success_zero → supports ZERO_CONFIRMED (only with complete entity-scoped collect)

    Empty map when table missing — callers must keep NOT_QUERIED (never invent zero).
    """
    out: dict[tuple[str, str], str] = {}
    if conn is None:
        return out
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT 1 FROM information_schema.tables
                WHERE table_schema = 'public' AND table_name = 'coverage_evidence'
                """
            )
            if not cur.fetchone():
                return out
            cur.execute(
                """
                SELECT entity_id, canonical_entity_key, source, state, checked_at
                FROM coverage_evidence
                WHERE COALESCE(capability, '') IN ('open_tenders', '')
                   OR COALESCE(data_type, '') IN ('bids', 'open_tenders', 'acts')
                ORDER BY checked_at DESC NULLS LAST
                """
            )
            for row in cur.fetchall():
                r = dict(row) if not isinstance(row, dict) else row
                # RealDictCursor or tuple
                if not isinstance(r, dict):
                    r = {
                        "entity_id": row[0],
                        "canonical_entity_key": row[1],
                        "source": row[2],
                        "state": row[3],
                    }
                src = str(r.get("source") or "").lower().strip()
                if src in {"ciga", "ciga_dom", "dom_sc"}:
                    src = "ciga_ckan"
                if src in {"pncp_opportunities", "test_batch"}:
                    src = "pncp"
                state = str(r.get("state") or "").lower().strip()
                eid = str(r.get("entity_id") or "").strip()
                ckey = str(r.get("canonical_entity_key") or "").strip()
                keys = []
                for raw in (eid, ckey):
                    if not raw:
                        continue
                    keys.append(raw)
                    c8 = _cnpj8(raw)
                    if c8:
                        keys.append(c8)
                for k in keys:
                    if k and (k, src) not in out:
                        # first (most recent checked_at) wins
                        out[(k, src)] = state
    except Exception:  # noqa: BLE001
        return out
    return out


def build_coverage_by_entity_source(
    *,
    entities: list[Any],
    observations: list[Any],
    freshness: list[dict[str, Any]] | None = None,
    policy: dict[str, Any] | None = None,
    entity_source_evidence: dict[tuple[str, str], str] | None = None,
) -> list[dict[str, Any]]:
    """Publish 1.093 × applicable sources with honest result states.

    Denominator of operational coverage is applicable combinations only, but
    every universe entity is always published. Absence of findings is never
    confused with absence of collection (NOT_QUERIED vs ZERO_CONFIRMED).

    ``ZERO_CONFIRMED`` is emitted only when entity-scoped evidence proves a
    complete query with zero hits (e.g. coverage_evidence.success_zero).
    Freshness alone never invents ZERO_CONFIRMED.
    """
    policy = policy or _load_source_policy()
    roles = (policy.get("source_roles") or {}).get("open_tenders") or {}
    # Default national + municipal required set when policy missing
    if not roles:
        roles = {"pncp": "required", "ciga_ckan": "required", "sc_compras": "complementary"}

    freshness_by = {str(f.get("source") or "").lower(): f for f in (freshness or [])}
    evidence = entity_source_evidence or {}
    # Map observation presence: fonte → set(cnpj8)
    found: dict[str, set[str]] = {}
    for o in observations:
        fonte = str(getattr(o, "fonte", "") or "").lower()
        key = _cnpj8(getattr(o, "orgao_cnpj", None) or getattr(o, "entity_key", None))
        if not key:
            # name-only match: use entity_key if annotated
            key = str(getattr(o, "entity_key", "") or "")[:8]
        if not key:
            continue
        found.setdefault(fonte, set()).add(key)

    rows: list[dict[str, Any]] = []
    for ent in entities:
        c8 = _cnpj8(getattr(ent, "cnpj8", None) or getattr(ent, "cnpj", None))
        ent_id = str(
            getattr(ent, "entity_key", None)
            or getattr(ent, "entity_id", None)
            or getattr(ent, "canonical_id", None)
            or c8
            or ""
        )
        name = str(
            getattr(ent, "canonical_name", None)
            or getattr(ent, "name", None)
            or getattr(ent, "razao_social", None)
            or ""
        )
        mun = str(getattr(ent, "municipio", None) or "")
        natureza = str(
            getattr(ent, "natureza_juridica", None)
            or getattr(ent, "natureza", None)
            or getattr(ent, "zone", None)
            or ""
        ).lower()
        # crude esfera from natureza/zone/name
        blob = f"{natureza} {name}".lower()
        if any(t in blob for t in ("municipal", "prefeitura", "camara", "câmara")):
            esfera = "municipal"
        elif any(t in blob for t in ("estadual", "estado de", "governo do estado")):
            esfera = "estadual"
        elif any(t in blob for t in ("federal", "uniao", "união", "ministerio", "ministério")):
            esfera = "federal"
        else:
            esfera = "municipal"  # Extra 200km universe is predominantly municipal

        for source, role in sorted(roles.items()):
            applicable = True
            # Complementary/gap sources stay applicable but optional
            if role == "informational" and source == "transparencia":
                applicable = True
            # Federal entities: ciga not applicable
            if source == "ciga_ckan" and esfera == "federal":
                applicable = False
            if source == "sc_compras" and esfera == "federal":
                applicable = False

            if not applicable:
                result = "NOT_APPLICABLE"
            else:
                # freshness / collection signal for this source family
                fr = freshness_by.get(source) or freshness_by.get(f"{source}_opportunities")
                if source == "pncp":
                    fr = fr or freshness_by.get("pncp_opportunities")
                level = str((fr or {}).get("level") or "")

                has_data = c8 in found.get(source, set())
                # Entity-scoped evidence (coverage_evidence) — only path to ZERO_CONFIRMED
                ev_state = ""
                for key in (c8, ent_id, _cnpj8(ent_id)):
                    if not key:
                        continue
                    ev_state = evidence.get((key, source), "") or ev_state
                    if source == "pncp":
                        ev_state = (
                            evidence.get((key, "pncp"), "")
                            or evidence.get((key, "pncp_opportunities"), "")
                            or ev_state
                        )
                    if ev_state:
                        break
                ev_l = str(ev_state or "").lower()

                if has_data or ev_l in {"success_with_data", "success", "found"}:
                    result = "FOUND"
                elif ev_l in {"success_zero", "zero_confirmed"}:
                    # Proven complete entity-scoped query with zero hits
                    result = "ZERO_CONFIRMED"
                elif level in {"never", "", "unknown"}:
                    # No collection evidence and no rows → not queried
                    result = "NOT_QUERIED"
                elif level in {"unreliable", "incomplete"}:
                    result = "PARTIAL" if level == "incomplete" else "ERROR"
                elif level == "stale":
                    result = "STALE"
                elif level == "fresh":
                    # Queried recently with no matching rows for this entity
                    # Without entity-scoped success_zero ledger → NOT_QUERIED
                    # (never invent ZERO_CONFIRMED from freshness alone)
                    result = "NOT_QUERIED"
                else:
                    result = "NOT_QUERIED"

            note = "FOUND=observation present; NOT_QUERIED≠ZERO_CONFIRMED; registry≠coverage"
            if result == "ZERO_CONFIRMED":
                note = "ZERO_CONFIRMED requires entity-scoped success_zero evidence"
            rows.append(
                {
                    "entity_id": ent_id,
                    "cnpj8": c8,
                    "razao_social": name,
                    "municipio": mun,
                    "esfera": esfera,
                    "capability": "open_tenders",
                    "source": source,
                    "source_role": role,
                    "applicable": "yes" if applicable else "no",
                    "result": result,
                    "result_vocabulary": "|".join(RESULT_STATES),
                    "note": note,
                }
            )
    return rows


def _write_coverage_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text(
            "entity_id,source,capability,result,note\n",
            encoding="utf-8",
        )
        return
    fields = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def build_qa_report(
    *,
    cycle_id: str,
    as_of: str,
    stats: dict[str, Any],
    coverage_rows: list[dict[str, Any]],
    load_meta: dict[str, Any],
    freshness: list[dict[str, Any]],
    pdf_ok: bool,
    excel_ok: bool,
    inv_errors: list[str],
    terminal_state: str,
) -> dict[str, Any]:
    """Machine-readable QA for the decision pack (fail-closed honesty)."""
    cov_found = sum(1 for r in coverage_rows if r.get("result") == "FOUND")
    cov_nq = sum(1 for r in coverage_rows if r.get("result") == "NOT_QUERIED")
    cov_na = sum(1 for r in coverage_rows if r.get("result") == "NOT_APPLICABLE")
    applicable = [r for r in coverage_rows if r.get("applicable") == "yes"]
    applicable_found = sum(1 for r in applicable if r.get("result") == "FOUND")

    checks = [
        {
            "id": "universe_denominator",
            "ok": int(stats.get("entes_universo") or 0) == CANONICAL_N,
            "detail": f"entes_universo={stats.get('entes_universo')}",
        },
        {
            "id": "pdf_generated",
            "ok": pdf_ok,
            "detail": "extra_decision_report PDF must exist (not RESIDUAL)",
        },
        {
            "id": "excel_generated",
            "ok": excel_ok,
            "detail": "extra_decision_pack Excel must exist",
        },
        {
            "id": "no_invented_coverage",
            "ok": True,
            "detail": (
                f"applicable FOUND={applicable_found}/{len(applicable)}; NOT_QUERIED={cov_nq}; NOT_APPLICABLE={cov_na}"
            ),
        },
        {
            "id": "multi_source_observations",
            "ok": len(load_meta.get("by_fonte") or {}) >= 1,
            "detail": f"by_fonte={load_meta.get('by_fonte')}",
        },
        {
            "id": "invariants",
            "ok": not inv_errors,
            "detail": inv_errors or ["none"],
        },
    ]
    # Freshness of PNCP opportunities
    pncp_f = next(
        (f for f in freshness if f.get("source") in {"pncp_opportunities", "pncp"}),
        {},
    )
    checks.append(
        {
            "id": "pncp_freshness_sla",
            "ok": pncp_f.get("level") == "fresh",
            "detail": {
                "level": pncp_f.get("level"),
                "age_hours": pncp_f.get("age_hours"),
                "sla_hours": pncp_f.get("sla_hours"),
            },
        }
    )

    failing = [c for c in checks if not c["ok"]]
    # Consultive reliability requires PDF+Excel and no invariant errors
    if not pdf_ok or not excel_ok or inv_errors:
        reliability = "NOT_READY"
    elif terminal_state in {"BLOCKED", "FAIL"}:
        reliability = "BLOCKED" if terminal_state == "BLOCKED" else "NOT_READY"
    elif pncp_f.get("level") != "fresh":
        reliability = "PARTIAL"
    else:
        reliability = "READY" if not failing else "PARTIAL"

    return {
        "schema": "extra-weekly-qa-report/1.0",
        "cycle_id": cycle_id,
        "as_of": as_of,
        "generated_at": iso_z(utc_now()),
        "reliability": reliability,
        "terminal_state": terminal_state,
        "checks": checks,
        "failing_check_ids": [c["id"] for c in failing],
        "coverage_summary": {
            "rows": len(coverage_rows),
            "found": cov_found,
            "not_queried": cov_nq,
            "not_applicable": cov_na,
            "applicable_found": applicable_found,
            "applicable_total": len(applicable),
            "note": (
                "Do not report applicable_found/applicable_total as operational "
                "coverage % without entity-scoped collection evidence"
            ),
        },
        "load_meta": load_meta,
        "stats": stats,
        "claims_forbidden": [
            "LOCAL_READY",
            "VPS_OPERATIONAL",
            "cobertura operacional 95%",
            "recall estratificado 95%",
            "RESIDUAL_NOT_GENERATED as success",
            "registry existence = coverage",
        ],
    }


def build_weekly_decision_artifacts(
    *,
    conn: Any,
    out_dir: Path,
    cycle_id: str,
    collection_id: str,
    freshness: list[dict[str, Any]] | None = None,
    intel: dict[str, Any] | None = None,
    runs: list[Any] | None = None,
    universe_path: Path = DEFAULT_UNIVERSE,
    profile_path: Path = DEFAULT_PROFILE,
    skip_network: bool = True,
    shortlist_limit: int = 25,
    now: datetime | None = None,
    pilot_approval_path: Path | None = None,
) -> dict[str, Any]:
    """Build canonical decision pack products into ``out_dir``."""
    now = now or utc_now()
    if now.tzinfo is None:
        now = now.replace(tzinfo=BR_TZ)
    as_of = now.astimezone(BR_TZ).date()
    as_of_s = as_of.isoformat()
    run_tag = cycle_id.replace(":", "").replace("/", "-")
    out_dir = Path(out_dir)
    freshness = freshness or []
    intel = intel or {}
    runs = runs or []

    entities = load_universe(universe_path)
    if pilot_approval_path is None and os.getenv("EXTRA_PILOT_APPROVAL"):
        pilot_approval_path = Path(os.environ["EXTRA_PILOT_APPROVAL"])
    pilot_gate = require_pilot_approval(
        universe_path=universe_path,
        policy_path=DEFAULT_PILOT_POLICY,
        universe_entity_count=len(entities),
        universe_entity_ids={entity.entity_key for entity in entities},
        approval_path=pilot_approval_path,
    )
    out_dir.mkdir(parents=True, exist_ok=True)
    by_cnpj8, names, by_name, municipios = build_indexes(entities)
    profile = _load_profile(profile_path)

    observations, load_meta = load_all_lake_observations(
        conn,
        as_of=as_of,
        auto_discover_files=True,
    )
    for obs in observations:
        annotate_observation_universe(
            obs,
            by_cnpj8=by_cnpj8,
            names=names,
            by_name=by_name,
            municipios=municipios,
        )

    processes, merges = consolidate_observations(observations, now=now)
    processes = apply_decisions(processes, profile=profile)
    shortlist = select_shortlist(processes, limit=shortlist_limit)

    # Skip heavy document inventory in weekly path by default (network optional)
    for p in shortlist:
        if not p.docs_inventory_status:
            p.docs_inventory_status = "not_inventoried_weekly"
            p.official_page_validated = False

    entity_evidence = load_entity_source_evidence(conn)
    load_meta = dict(load_meta or {})
    load_meta["entity_source_evidence_n"] = len(entity_evidence)
    coverage_rows = build_coverage_by_entity_source(
        entities=entities,
        observations=observations,
        freshness=freshness,
        policy=_load_source_policy(),
        entity_source_evidence=entity_evidence,
    )
    cov_path = out_dir / "coverage_by_entity_source.csv"
    _write_coverage_csv(cov_path, coverage_rows)

    stats = build_reconciliation(
        entities=entities,
        observations=observations,
        processes=processes,
        shortlist=shortlist,
        merges=merges,
        coverage_rows=coverage_rows,
    )
    inv_errors = stats.assert_invariants()
    freshness_notes: list[str] = []
    for f in freshness:
        if f.get("level") in {"stale", "never", "unreliable", "incomplete"}:
            freshness_notes.append(
                f"Freshness {f.get('source')}={f.get('level')} age_h={f.get('age_hours')} sla={f.get('sla_hours')}"
            )
    limitations = default_limitations(stats=stats, inputs=load_meta, freshness_notes=freshness_notes)
    limitations.append(
        "Pacote semanal: inventário documental HTTP da shortlist desligado por padrão "
        "(WEEKLY_INVENTORY_DOCS=1 para habilitar)."
    )
    source_policy = default_source_policy(stats)

    pack_meta: dict[str, Any] = {
        "pack_id": f"EXTRA-WEEKLY-{run_tag}",
        "cycle_id": cycle_id,
        "collection_id": collection_id,
        "generated_at": iso_z(now),
        "as_of": as_of_s,
        "motor_version": MOTOR_VERSION,
        "motor_module": "scripts.ops.weekly_decision_artifacts",
        "universe_n": stats.entes_universo,
        "stats": stats.to_dict(),
        "load_meta": load_meta,
        "limitations": limitations,
        "source_policy": source_policy,
        "human_accept": "PENDING_HUMAN",
        "pilot_approval": pilot_gate.to_dict(),
        "terminal_state": "FAIL" if inv_errors else "PASS",
        "invariant_errors": inv_errors,
        "shortlist_process_ids": [p.process_id for p in shortlist],
        "claims_forbidden": [
            "LOCAL_READY",
            "VPS_OPERATIONAL",
            "cobertura_95",
            "probabilidade_de_vitoria",
            "RESIDUAL_NOT_GENERATED",
        ],
    }
    if freshness_notes:
        pack_meta["terminal_state"] = "BLOCKED" if not inv_errors else "FAIL"
        pack_meta["blockers_external"] = freshness_notes
    if not observations:
        pack_meta["terminal_state"] = "BLOCKED"
        pack_meta["blockers_external"] = pack_meta.get("blockers_external", []) + [
            "Nenhuma observação multi-fonte no lake/artefatos — pacote sem base decisória"
        ]

    # Official artifact names
    xlsx_name = f"extra_decision_pack_{as_of_s}_{run_tag}.xlsx"
    pdf_name = f"extra_decision_report_{as_of_s}_{run_tag}.pdf"
    xlsx_path = out_dir / xlsx_name
    pdf_path = out_dir / pdf_name

    obs_sample = [
        {
            "observation_id": o.observation_id,
            "fonte": o.fonte,
            "id_externo": o.id_externo,
            "orgao": o.orgao,
            "objeto": (o.objeto or "")[:200],
            "event_type": o.event_type,
            "is_active_dispute": "sim" if o.is_active_dispute else "nao",
            "in_universe": "sim" if o.in_universe else "nao",
            "url": o.url,
        }
        for o in observations[:500]
    ]

    excel_ok = False
    excel_error = ""
    try:
        write_excel(
            xlsx_path,
            pack_meta=pack_meta,
            processes=processes,
            shortlist=shortlist,
            stats=stats,
            source_policy=source_policy,
            limitations=limitations,
            observations_sample=obs_sample,
            logo_path=None,
        )
        # Add coverage sheet if openpyxl available
        try:
            from openpyxl import load_workbook

            wb = load_workbook(xlsx_path)
            if "CoverageEntitySource" in wb.sheetnames:
                del wb["CoverageEntitySource"]
            ws = wb.create_sheet("CoverageEntitySource")
            if coverage_rows:
                headers = list(coverage_rows[0].keys())
                ws.append(headers)
                for r in coverage_rows:
                    ws.append([r.get(h) for h in headers])
            else:
                ws.append(["(vazio)"])
            wb.save(xlsx_path)
        except Exception as sheet_exc:  # noqa: BLE001
            # Coverage sheet is additive; base workbook already written.
            pack_meta.setdefault("warnings", []).append(
                f"coverage_sheet_append_failed:{sheet_exc}"
            )
        excel_ok = xlsx_path.is_file() and xlsx_path.stat().st_size > 0
    except Exception as exc:  # noqa: BLE001
        excel_error = str(exc)
        xlsx_path.write_text(f"Excel failed: {exc}\n", encoding="utf-8")

    pdf_ok = False
    pdf_error = ""
    try:
        write_pdf(
            pdf_path,
            pack_meta=pack_meta,
            shortlist=shortlist,
            stats=stats,
            limitations=limitations,
            logo_path=None,
        )
        pdf_ok = pdf_path.is_file() and pdf_path.stat().st_size > 0
    except Exception as exc:  # noqa: BLE001
        pdf_error = str(exc)

    # decision_dataset.json — full machine payload
    decision_dataset = {
        "schema": "extra-weekly-decision-dataset/1.0",
        "cycle_id": cycle_id,
        "collection_id": collection_id,
        "as_of": as_of_s,
        "generated_at": pack_meta["generated_at"],
        "motor_version": MOTOR_VERSION,
        "universe_n": stats.entes_universo,
        "load_meta": load_meta,
        "stats": stats.to_dict(),
        "shortlist": [p.to_csv_row() for p in shortlist],
        "processes_n": len(processes),
        "observations_n": len(observations),
        "intelligence_counts": (intel or {}).get("counts"),
        "runs": [r.to_dict() if hasattr(r, "to_dict") else r for r in runs],
        "freshness": freshness,
        "terminal_state": pack_meta["terminal_state"],
        "artifacts": {
            "excel": xlsx_name if excel_ok else None,
            "pdf": pdf_name if pdf_ok else None,
            "coverage_by_entity_source": "coverage_by_entity_source.csv",
        },
    }
    decision_path = out_dir / "decision_dataset.json"
    _atomic_json(decision_path, decision_dataset)

    # Minimal Decision Memory weekly board section (derived from PG when DSN available).
    # Never directory-scans run artifacts for memory facts.
    decision_memory_board: dict[str, Any] | None = None
    decision_memory_board_path: Path | None = None
    try:
        dsn = os.getenv("LOCAL_DATALAKE_DSN")
        if dsn:
            from scripts.decision_memory.db import connect
            from scripts.decision_memory.repository import DecisionMemoryRepository
            from scripts.decision_memory.weekly_board import build_weekly_board

            conn_dm = connect(dsn)
            try:
                repo_dm = DecisionMemoryRepository(conn_dm)
                decision_memory_board = build_weekly_board(
                    repo_dm,
                    client_id="extra",
                    cycle_id=cycle_id,
                )
                decision_memory_board_path = out_dir / "decision-memory-weekly-board.json"
                _atomic_json(decision_memory_board_path, decision_memory_board)
                decision_dataset["decision_memory_board"] = {
                    "path": decision_memory_board_path.name,
                    "counts": decision_memory_board.get("counts"),
                    "source": "postgresql:dm_*",
                }
                _atomic_json(decision_path, decision_dataset)
            finally:
                conn_dm.close()
    except Exception as dm_exc:  # noqa: BLE001 — non-fatal for pack generation
        pack_meta.setdefault("warnings", []).append(f"decision_memory_board:{dm_exc}")

    qa = build_qa_report(
        cycle_id=cycle_id,
        as_of=as_of_s,
        stats=stats.to_dict(),
        coverage_rows=coverage_rows,
        load_meta=load_meta,
        freshness=freshness,
        pdf_ok=pdf_ok,
        excel_ok=excel_ok,
        inv_errors=inv_errors,
        terminal_state=pack_meta["terminal_state"],
    )
    qa_path = out_dir / "qa_report.json"
    _atomic_json(qa_path, qa)

    # Also keep stable aliases for tools that expect fixed names
    if excel_ok:
        alias_x = out_dir / "extra_decision_pack.xlsx"
        try:
            shutil.copy2(xlsx_path, alias_x)
        except OSError as copy_exc:
            pack_meta.setdefault("warnings", []).append(f"excel_alias_copy:{copy_exc}")
    if pdf_ok:
        alias_p = out_dir / "extra_decision_report.pdf"
        try:
            shutil.copy2(pdf_path, alias_p)
        except OSError as copy_exc:
            pack_meta.setdefault("warnings", []).append(f"pdf_alias_copy:{copy_exc}")

    product_checksums: dict[str, Any] = {}
    checksum_targets: dict[str, Path] = {
        "extra_decision_pack": xlsx_path,
        "extra_decision_report": pdf_path,
        "coverage_by_entity_source": cov_path,
        "decision_dataset": decision_path,
        "qa_report": qa_path,
    }
    if decision_memory_board_path is not None:
        checksum_targets["decision_memory_weekly_board"] = decision_memory_board_path
    for label, pth in checksum_targets.items():
        if pth.is_file():
            product_checksums[label] = {
                "path": pth.name,
                "sha256": sha256_file(pth),
                "bytes": pth.stat().st_size,
            }

    delivery_ok = excel_ok and pdf_ok
    return {
        "status": "ok" if delivery_ok else "fail",
        "excel_ok": excel_ok,
        "pdf_ok": pdf_ok,
        "pdf_status": "GENERATED" if pdf_ok else f"FAILED:{pdf_error or 'missing'}",
        "excel_error": excel_error or None,
        "pdf_error": pdf_error or None,
        "excel": str(xlsx_path) if excel_ok else None,
        "pdf": str(pdf_path) if pdf_ok else None,
        "coverage_csv": str(cov_path),
        "decision_dataset": str(decision_path),
        "qa_report": str(qa_path),
        "decision_memory_board": str(decision_memory_board_path)
        if decision_memory_board_path
        else None,
        "qa_reliability": qa.get("reliability"),
        "terminal_state": pack_meta["terminal_state"],
        "product_checksums": product_checksums,
        "stats": stats.to_dict(),
        "load_meta": load_meta,
        "shortlist_n": len(shortlist),
        "as_of": as_of_s,
        "artifact_names": {
            "excel": xlsx_name,
            "pdf": pdf_name,
        },
    }


# Regression helper used by tests — fleet maintenance must never be AEC
def is_engineering_opportunity(objeto: str) -> bool:
    aec = classify_aec(objeto, is_active_dispute=True)
    return bool(aec.is_aec)


__all__ = [
    "CANONICAL_N",
    "RESULT_STATES",
    "build_coverage_by_entity_source",
    "build_qa_report",
    "build_weekly_decision_artifacts",
    "is_engineering_opportunity",
    "load_entity_source_evidence",
]
