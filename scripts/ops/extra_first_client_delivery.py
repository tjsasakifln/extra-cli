#!/usr/bin/env python3
"""EXTRA-FIRST-CLIENT-DECISION-DELIVERY-01 — composição fina do pacote cliente.

Reutiliza o pacote semanal real (WEEKLY_INPUT) + perfil extra.yaml.
Não reimplementa crawlers, classificadores ou triagem completa de edital.
Não inventa respostas de elicitation. Nunca emite GO com campos críticos PENDING.
Não autoaceita human-review.

Comando canônico (Makefile):

  make extra-first-client-delivery \\
    WEEKLY_INPUT=/caminho/do/run-real \\
    DELIVERY_OUT=/caminho/externo

DSN isolado é opcional (CLIENT_READY_DSN) e só usado se fornecido; o núcleo
do pacote de decisão opera sobre os artefatos imutáveis do weekly pack.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import yaml

from scripts.ops.commercial_executive_render import br_currency, br_date

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_PROFILE = PROJECT_ROOT / "config" / "client_profiles" / "extra.yaml"
PACKAGE_TITLE = "Extra Construtora — Primeira Rodada de Decisão B2G"
SCHEMA = "extra-first-client-decision-delivery/1.0"

# Campos que bloqueiam GO (R3/R4)
CRITICAL_PENDING_KEYS = (
    "capital_giro",
    "capacidade_garantia",
    "capacidade_simultanea",
    "cats_atestados",
    "margem_minima",
)

# Até 10 perguntas de intake (R3)
INTAKE_PRIORITY = (
    ("capacidade_simultanea", "Quantas obras ou contratos simultâneos a Extra sustenta com segurança?"),
    ("capital_giro", "Qual o capital de giro disponível para propostas e garantias?"),
    ("capacidade_garantia", "Qual a capacidade de emissão de garantias (proposta/contrato)?"),
    ("cats_atestados", "Quais são as principais CATs/atestados vigentes (objeto, valor, órgão, validade)?"),
    ("equipe", "Qual a equipe técnica disponível (engenharia, orçamentista, administrativo)?"),
    ("equipamentos", "Quais equipamentos próprios relevantes para obras/serviços?"),
    ("certidoes", "Qual o status consolidado das certidões (federal/estadual/municipal/trabalhista/FGTS)?"),
    ("margem_minima", "Qual a margem mínima aceitável por tipo de objeto?"),
    ("risco_aceitavel", "Qual o apetite de risco (prazos curtos, garantias, órgãos com histórico ruim)?"),
    ("priority_organs_competitors", "Quais órgãos e concorrentes devem ser prioritários nesta rotina?"),
)

GENERIC_URL_MARKERS = (
    "pncp.gov.br/app/editais?",
    "pncp.gov.br/app/editais/?",
    "/app/editais?",
    "consulta",
    "search?",
    "busca?",
)

TERMINAL_STATUS_MARKERS = (
    "encerrad",
    "revogad",
    "suspens",
    "cancelad",
    "homologad",
    "fracassad",
    "deserto",
    "anulad",
    "concluid",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def utc_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    raw = str(value).strip()
    if not raw:
        return None
    # ISO or datetime
    try:
        return date.fromisoformat(raw[:10])
    except ValueError:
        return None


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def is_generic_url(url: str | None) -> bool:
    if not url or not str(url).strip():
        return True
    u = str(url).strip().lower()
    if u in {"#", "n/a", "na", "none", "null"}:
        return True
    # Specific PNCP compra path is OK: /app/editais/<cnpj>/<ano>/<seq>
    if re.search(r"pncp\.gov\.br/app/editais/\d{14}/\d{4}/\d+", u):
        return False
    if re.search(r"pncp\.gov\.br/api/", u):
        return False
    for marker in GENERIC_URL_MARKERS:
        if marker in u:
            return True
    # bare domain home pages
    if u.rstrip("/").endswith("pncp.gov.br") or u.rstrip("/").endswith("pncp.gov.br/app/editais"):
        return True
    return False


def is_terminal_status(status: str | None) -> bool:
    s = (status or "").lower()
    return any(m in s for m in TERMINAL_STATUS_MARKERS)


def normalize_text(s: str) -> str:
    t = (s or "").lower()
    # strip accents lightly for term match
    repl = str.maketrans(
        "áàâãäéèêëíìîïóòôõöúùûüç",
        "aaaaaeeeeiiiiooooouuuuc",
    )
    return t.translate(repl)


def match_terms(objeto: str, terms: list[str]) -> list[str]:
    o = normalize_text(objeto)
    found: list[str] = []
    for term in terms:
        nt = normalize_text(str(term))
        if nt and nt in o:
            found.append(str(term))
    return found


def new_run_id() -> str:
    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    suffix = hashlib.sha256(stamp.encode()).hexdigest()[:10]
    return f"extra-first-{stamp}-{suffix}"


# ---------------------------------------------------------------------------
# Profile / intake
# ---------------------------------------------------------------------------


def load_profile(path: Path) -> dict[str, Any]:
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    data["_path"] = str(path)
    data["_sha256"] = sha256_file(path)
    return data


def elicitation_block(profile: dict[str, Any]) -> dict[str, Any]:
    return dict(profile.get("elicitation") or {})


def field_is_pending(val: Any) -> bool:
    if val is None:
        return True
    if isinstance(val, dict):
        status = str(val.get("status") or "").upper()
        if status in {"PENDING", "ELICIT", "TODO", "NULL", ""}:
            return True
        inner = val.get("value")
        if inner is None:
            return True
        if isinstance(inner, (list, dict, str)) and not inner:
            return True
        return False
    if isinstance(val, (list, dict, str)) and not val:
        return True
    return False


def critical_pending_fields(profile: dict[str, Any]) -> list[str]:
    elic = elicitation_block(profile)
    pending: list[str] = []
    for key in CRITICAL_PENDING_KEYS:
        val = elic.get(key, profile.get(key))
        if field_is_pending(val):
            pending.append(key)
    return pending


def go_blocked_by_profile(profile: dict[str, Any]) -> bool:
    return len(critical_pending_fields(profile)) > 0


def build_intake(profile: dict[str, Any]) -> dict[str, Any]:
    elic = elicitation_block(profile)
    questions: list[dict[str, Any]] = []
    for key, question in INTAKE_PRIORITY:
        if key == "priority_organs_competitors":
            profile_fields = ["priority_organs", "known_competitors"]
            status = "PENDING"
            note = "Calibrar órgãos e concorrentes prioritários com Leonardo/Tiago"
        else:
            profile_fields = [f"elicitation.{key}"]
            raw = elic.get(key)
            if isinstance(raw, dict):
                status = str(raw.get("status") or "PENDING").upper()
                note = raw.get("note") or ""
            else:
                status = "PENDING" if field_is_pending(raw) else "SET"
                note = ""
        questions.append(
            {
                "id": f"Q{len(questions) + 1:02d}",
                "key": key,
                "question": question,
                "profile_fields": profile_fields,
                "status": status,
                "note": note,
                "answer": None,  # never invent
            }
        )
        if len(questions) >= 10:
            break

    patch: dict[str, Any] = {
        "profile_id": profile.get("profile_id"),
        "version": profile.get("version"),
        "version_notes": "candidate patch — NÃO aplicar automaticamente; validar com Tiago",
        "elicitation": {},
    }
    for q in questions:
        if q["key"] == "priority_organs_competitors":
            patch["priority_organs"] = None
            patch["known_competitors"] = None
            continue
        patch["elicitation"][q["key"]] = {
            "status": "PENDING",
            "value": None,
            "note": q.get("note") or "preencher após entrevista",
        }

    return {
        "schema": "extra-operational-intake/1.0",
        "client": profile.get("display_name") or "Extra Construtora",
        "profile_id": profile.get("profile_id"),
        "profile_version": profile.get("version"),
        "profile_sha256": profile.get("_sha256"),
        "generated_at": utc_now(),
        "questions": questions,
        "critical_pending": critical_pending_fields(profile),
        "go_blocked": True if go_blocked_by_profile(profile) else False,
        "rules": [
            "Nunca inventar resposta",
            "Nunca deduzir capital/capacidade/margem a partir de contratos públicos",
            "Nunca modificar extra.yaml automaticamente",
            "Nunca promover PENDING→SET sem validação de Tiago",
        ],
        "profile_patch_candidate": patch,
    }


def intake_to_markdown(intake: dict[str, Any]) -> str:
    lines = [
        f"# Intake operacional — {intake.get('client')}",
        "",
        f"- Perfil: `{intake.get('profile_id')}` v{intake.get('profile_version')}",
        f"- Hash do perfil: `{intake.get('profile_sha256')}`",
        f"- Gerado em: {intake.get('generated_at')}",
        f"- GO bloqueado por PENDING críticos: **{'sim' if intake.get('go_blocked') else 'não'}**",
        "",
        "## Perguntas (máx. 10) — respostas em branco de propósito",
        "",
        "Preencher com Leonardo/Extra. Não inventar valores.",
        "",
    ]
    for q in intake.get("questions") or []:
        lines.append(f"### {q['id']} — {q['question']}")
        lines.append(f"- Campo(s) do perfil: `{', '.join(q['profile_fields'])}`")
        lines.append(f"- Status atual: **{q['status']}**")
        if q.get("note"):
            lines.append(f"- Nota: {q['note']}")
        lines.append("- Resposta: _pendente_")
        lines.append("")
    lines.extend(
        [
            "## Regras",
            "",
        ]
    )
    for r in intake.get("rules") or []:
        lines.append(f"- {r}")
    lines.append("")
    return "\n".join(lines)


def profile_patch_yaml(intake: dict[str, Any]) -> str:
    body = {
        "profile_patch_candidate": intake.get("profile_patch_candidate"),
        "warning": "PLACEHOLDER — valores nulos; aplicar somente após validação humana",
    }
    return yaml.safe_dump(body, allow_unicode=True, sort_keys=False)


# ---------------------------------------------------------------------------
# Weekly pack validation
# ---------------------------------------------------------------------------


@dataclass
class WeeklyValidation:
    ok: bool
    weekly_dir: str
    cycle_id: str | None = None
    collection_id: str | None = None
    exit_code: int | None = None
    cut_date: str | None = None
    freshness: list[dict[str, Any]] = field(default_factory=list)
    source_health: list[dict[str, Any]] = field(default_factory=list)
    limitations: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    product_checksums_ok: bool = False
    opportunities_path: str | None = None
    orgaos_path: str | None = None
    source_health_path: str | None = None
    manifest_path: str | None = None


def validate_weekly_pack(weekly_dir: Path) -> WeeklyValidation:
    weekly_dir = weekly_dir.resolve()
    result = WeeklyValidation(ok=False, weekly_dir=str(weekly_dir))
    if not weekly_dir.is_dir():
        result.errors.append(f"WEEKLY_INPUT não é diretório: {weekly_dir}")
        return result

    manifest_path = weekly_dir / "manifest.json"
    checksums_path = weekly_dir / "checksums.json"
    if not manifest_path.is_file():
        result.errors.append("manifest.json ausente no weekly pack")
        return result
    if not checksums_path.is_file():
        result.errors.append("checksums.json ausente no weekly pack")
        return result

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        result.errors.append(f"manifest.json inválido: {exc}")
        return result

    try:
        checksums = json.loads(checksums_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        result.errors.append(f"checksums.json inválido: {exc}")
        return result

    result.manifest_path = str(manifest_path)
    result.cycle_id = manifest.get("cycle_id")
    result.collection_id = manifest.get("collection_id")
    result.exit_code = manifest.get("exit_code")
    result.cut_date = (
        (manifest.get("finished_at") or manifest.get("started_at") or "")[:10] or None
    )
    result.freshness = list(manifest.get("freshness") or [])
    result.source_health = list(manifest.get("source_health") or [])
    result.limitations = [str(x) for x in (manifest.get("limitations") or [])]

    # Validate declared artifact checksums (basename match — pack may be relocated)
    artifacts = checksums.get("artifacts") or {}
    if not artifacts:
        result.errors.append("checksums.json sem artifacts")
        return result

    mismatches: list[str] = []
    missing: list[str] = []
    for key, meta in artifacts.items():
        if not isinstance(meta, dict):
            mismatches.append(f"{key}: meta inválida")
            continue
        rel = meta.get("path") or ""
        name = Path(rel).name if rel else f"{key}"
        # common aliases
        candidates = [weekly_dir / name]
        if key.endswith("_csv"):
            candidates.append(weekly_dir / f"{key.replace('_csv', '')}.csv")
        expected = meta.get("sha256")
        found = next((c for c in candidates if c.is_file()), None)
        if found is None:
            # empty contracts/competitors may be 0-byte
            if expected == hashlib.sha256(b"").hexdigest():
                # treat as present empty
                continue
            missing.append(name)
            continue
        actual = sha256_file(found)
        if expected and actual != expected:
            mismatches.append(f"{name}: sha256 divergente")

    if missing:
        result.errors.append(f"arquivos declarados ausentes: {', '.join(missing)}")
    if mismatches:
        result.errors.append(f"checksum divergente: {'; '.join(mismatches)}")

    opp = weekly_dir / "opportunities.csv"
    if not opp.is_file():
        result.errors.append("opportunities.csv ausente")
    else:
        result.opportunities_path = str(opp)
    orgaos = weekly_dir / "orgaos.csv"
    if orgaos.is_file():
        result.orgaos_path = str(orgaos)
    sh = weekly_dir / "source_health.csv"
    if sh.is_file():
        result.source_health_path = str(sh)

    result.product_checksums_ok = not missing and not mismatches
    result.ok = result.product_checksums_ok and bool(result.opportunities_path) and not result.errors
    return result


def load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------------------
# Shortlist / decision rules
# ---------------------------------------------------------------------------


def build_pncp_specific_url(numero_controle: str | None, link: str | None) -> tuple[str | None, bool]:
    """Return (url, is_specific). Prefer explicit link if specific; else derive from control number."""
    if link and not is_generic_url(link):
        return str(link).strip(), True
    nc = (numero_controle or "").strip()
    # format: cnpj-seq-numero/ano  e.g. 83102541000158-1-000167/2025
    m = re.match(r"^(\d{14})-(\d+)-(\d+)/(\d{4})$", nc)
    if m:
        cnpj, _seq, num, year = m.groups()
        # PNCP public path uses year + sequential
        url = f"https://pncp.gov.br/app/editais/{cnpj}/{year}/{int(num)}"
        return url, True
    if link:
        return str(link).strip(), False
    return None, False


def evaluate_opportunity(
    row: dict[str, Any],
    *,
    profile: dict[str, Any],
    as_of: date,
    cycle_id: str | None,
    collection_id: str | None,
    cut_date: str | None,
    go_blocked: bool,
    critical_pending: list[str],
) -> dict[str, Any]:
    objeto = row.get("objeto") or ""
    status = row.get("status_canonico") or row.get("status") or ""
    deadline = parse_date(row.get("data_encerramento"))
    days_remaining = (deadline - as_of).days if deadline else None
    valor = parse_float(row.get("valor_estimado"))
    # absence of value stays null — never coerce to 0
    if row.get("valor_estimado") in (None, ""):
        valor = None

    positive_terms = list(profile.get("positive_terms") or [])
    negative_terms = list(profile.get("negative_terms") or [])
    # also pull desired object type terms
    for ot in profile.get("desired_object_types") or []:
        positive_terms.extend(ot.get("terms") or [])

    pos_hits = match_terms(objeto, positive_terms)
    neg_hits = match_terms(objeto, negative_terms)

    url, url_specific = build_pncp_specific_url(
        row.get("numero_controle_pncp") or row.get("source_id"),
        row.get("link_edital") or row.get("source_url"),
    )

    hard_blocks: list[str] = []
    if deadline is None:
        hard_blocks.append("DEADLINE_UNKNOWN")
    elif days_remaining is not None and days_remaining < 0:
        hard_blocks.append("DEADLINE_PASSED")
    if is_terminal_status(status):
        hard_blocks.append("TERMINAL_OR_SUSPENDED")
    if not (row.get("numero_controle_pncp") or row.get("source_id")):
        hard_blocks.append("IDENTITY_MISSING")
    if not url_specific:
        hard_blocks.append("URL_NOT_SPECIFIC")
    if neg_hits and not pos_hits:
        hard_blocks.append("NEGATIVE_OBJECT_MATCH")

    # client fit
    if pos_hits and not neg_hits:
        client_fit = "ADERENTE"
    elif pos_hits and neg_hits:
        client_fit = "MISTO"
    elif neg_hits:
        client_fit = "INCOMPATIVEL"
    else:
        client_fit = "INDETERMINADO"

    # recommendation
    recommendation = "REVIEW"
    reason_parts: list[str] = []
    if "DEADLINE_PASSED" in hard_blocks:
        recommendation = "NO_GO"
        reason_parts.append("prazo de encerramento já passou")
    if "TERMINAL_OR_SUSPENDED" in hard_blocks:
        recommendation = "NO_GO"
        reason_parts.append("status terminal/suspenso")
    if "IDENTITY_MISSING" in hard_blocks:
        recommendation = "NO_GO"
        reason_parts.append("sem identidade específica")
    if "NEGATIVE_OBJECT_MATCH" in hard_blocks:
        recommendation = "NO_GO"
        reason_parts.append("objeto com termos negativos sem aderência positiva")
    if recommendation == "REVIEW" and client_fit == "INCOMPATIVEL":
        recommendation = "NO_GO"
        reason_parts.append("objeto incompatível com perfil Extra")
    if recommendation == "REVIEW" and not pos_hits and client_fit == "INDETERMINADO":
        # keep as candidate only if engineering-ish via ranking REVIEW and SC
        uf = (row.get("uf") or "").upper()
        if uf and uf != "SC":
            recommendation = "NO_GO"
            reason_parts.append("fora de SC sem aderência explícita")

    # NEVER GO while critical pending
    if recommendation == "GO" or str(row.get("ranking") or "").upper() == "GO":
        if go_blocked:
            recommendation = "REVIEW"
            reason_parts.append(
                "GO proibido: campos críticos da Extra ainda PENDING ("
                + ", ".join(critical_pending)
                + ")"
            )

    # Score never becomes GO
    if recommendation not in {"REVIEW", "NO_GO"}:
        recommendation = "REVIEW"

    evidence_status = "DOCUMENTS_NOT_CONFIRMED"
    if url_specific:
        evidence_status = "URL_SPECIFIC_DOCUMENTS_NOT_CONFIRMED"
    if "URL_NOT_SPECIFIC" in hard_blocks:
        evidence_status = "URL_GENERIC_OR_MISSING"

    missing_client = list(critical_pending) if recommendation == "REVIEW" else []

    if recommendation == "REVIEW":
        next_action = "Selecionar para leitura de edital/anexos e validar capacidade Extra"
        if not reason_parts:
            reason_parts.append(
                "prazo futuro útil e objeto com aderência ou identidade suficiente para tempo humano"
            )
    else:
        next_action = "Descartar desta rodada; manter apenas como contexto se útil"
        if not reason_parts:
            reason_parts.append("não atende critérios defensáveis desta rodada")

    ranking_score = parse_float(row.get("ranking_score"))

    return {
        "opportunity_id": row.get("numero_controle_pncp")
        or row.get("source_id")
        or row.get("id"),
        "numero_controle": row.get("numero_controle_pncp") or row.get("source_id"),
        "objeto": objeto,
        "orgao": row.get("orgao_nome"),
        "orgao_cnpj": row.get("orgao_cnpj"),
        "municipio": row.get("municipio"),
        "uf": row.get("uf"),
        "modalidade": row.get("modalidade"),
        "data_limite": deadline.isoformat() if deadline else None,
        "dias_restantes": days_remaining,
        "valor": valor,
        "valor_semantica": row.get("valor_semantica") or row.get("valor_tipo") or None,
        "url_oficial": url,
        "url_especifica": url_specific,
        "origem": row.get("source") or "pncp",
        "collection_id": collection_id or row.get("cycle_collection_id"),
        "cycle_run_id": cycle_id or row.get("cycle_run_id"),
        "data_coleta": (row.get("ingested_at") or row.get("updated_at") or "")[:10] or None,
        "data_corte": cut_date,
        "freshness": None,  # filled by pack-level
        "data_confidence": row.get("ranking_confianca") or "unknown",
        "client_fit": client_fit,
        "termos_positivos": pos_hits,
        "termos_negativos": neg_hits,
        "hard_blocks": hard_blocks,
        "informacoes_extra_necessarias": missing_client,
        "recommendation": recommendation,
        "recommendation_reason": "; ".join(reason_parts),
        "evidence_status": evidence_status,
        "next_action": next_action,
        "owner": "Tiago / Leonardo",
        "decision_deadline": None,
        "ranking_source": row.get("ranking") or row.get("ranking_effective"),
        "ranking_score": ranking_score,
        "limitations": [],
        "status_canonico": status,
        "posicao": None,
        "motivo_posicao": None,
    }


def build_shortlist(
    rows: list[dict[str, Any]],
    *,
    profile: dict[str, Any],
    as_of: date,
    cycle_id: str | None,
    collection_id: str | None,
    cut_date: str | None,
    max_n: int = 10,
) -> dict[str, Any]:
    go_blocked = go_blocked_by_profile(profile)
    critical = critical_pending_fields(profile)

    evaluated = [
        evaluate_opportunity(
            r,
            profile=profile,
            as_of=as_of,
            cycle_id=cycle_id,
            collection_id=collection_id,
            cut_date=cut_date,
            go_blocked=go_blocked,
            critical_pending=critical,
        )
        for r in rows
    ]

    # Defensible shortlist: REVIEW with future deadline and identity
    review = [
        e
        for e in evaluated
        if e["recommendation"] == "REVIEW"
        and e.get("dias_restantes") is not None
        and e["dias_restantes"] >= 0
        and e.get("numero_controle")
        and "IDENTITY_MISSING" not in (e.get("hard_blocks") or [])
    ]

    # Prefer engineering adherence, nearer deadlines, higher source ranking score
    def sort_key(e: dict[str, Any]) -> tuple:
        fit_rank = {"ADERENTE": 0, "MISTO": 1, "INDETERMINADO": 2, "INCOMPATIVEL": 3}.get(
            e.get("client_fit") or "", 9
        )
        days = e.get("dias_restantes") if e.get("dias_restantes") is not None else 9999
        score = e.get("ranking_score") if e.get("ranking_score") is not None else -1.0
        return (fit_rank, days, -float(score))

    review_sorted = sorted(review, key=sort_key)
    shortlist = review_sorted[:max_n]

    # If fewer than 5, keep all and mark insufficiency
    insufficient = len(shortlist) < 5

    # Context: best engineering NO_GO by deadline for transparency (not recommended)
    context = [
        e
        for e in evaluated
        if e.get("client_fit") in {"ADERENTE", "MISTO"} and e["recommendation"] == "NO_GO"
    ]
    context = sorted(context, key=sort_key)[:10]

    from datetime import timedelta

    # Assign positions
    for i, e in enumerate(shortlist, start=1):
        e["posicao"] = i
        e["motivo_posicao"] = (
            f"Aderência={e.get('client_fit')}; "
            f"dias_restantes={e.get('dias_restantes')}; "
            f"termos+={len(e.get('termos_positivos') or [])}; "
            f"{e.get('recommendation_reason')}"
        )
        dr = e.get("dias_restantes")
        if dr is not None:
            delta = min(3, max(0, int(dr) - 1))
            e["decision_deadline"] = (as_of + timedelta(days=delta)).isoformat()
        e["freshness"] = "pack_level"
        if not e.get("url_especifica"):
            e["limitations"] = list(e.get("limitations") or []) + ["DOCUMENTS_NOT_CONFIRMED"]

    go_count = sum(1 for e in shortlist if e["recommendation"] == "GO")
    # enforce zero GO
    if go_count:
        for e in shortlist:
            if e["recommendation"] == "GO":
                e["recommendation"] = "REVIEW"
                e["recommendation_reason"] = (
                    (e.get("recommendation_reason") or "")
                    + "; GO forçado a REVIEW por política PENDING"
                ).strip("; ")
        go_count = 0

    return {
        "as_of": as_of.isoformat(),
        "candidates_total": len(evaluated),
        "blocked_total": sum(1 for e in evaluated if e["recommendation"] == "NO_GO"),
        "review_defensible_total": len(review),
        "shortlist": shortlist,
        "shortlist_count": len(shortlist),
        "review_count": sum(1 for e in shortlist if e["recommendation"] == "REVIEW"),
        "no_go_count": sum(1 for e in shortlist if e["recommendation"] == "NO_GO"),
        "go_count": go_count,
        "insufficient": insufficient,
        "insufficiency_reason": (
            f"Apenas {len(shortlist)} oportunidade(s) defensável(is) com prazo futuro "
            f"e identidade específica em {len(evaluated)} candidatas do weekly pack. "
            "Isto NÃO significa zero de mercado — significa que este recorte/coleta "
            "não sustenta shortlist completa de 5–10 itens acionáveis."
            if insufficient
            else None
        ),
        "context_recent_no_go": context,
        "evaluated_all": evaluated,
        "go_blocked_by_profile": go_blocked,
        "critical_pending": critical,
    }


def select_deep_dive(shortlist_result: dict[str, Any]) -> dict[str, Any] | None:
    """Pick one opportunity for dossiê — not only by score."""
    items = list(shortlist_result.get("shortlist") or [])
    if not items:
        return None
    # Prefer ADERENTE + specific URL + nearer deadline
    ranked = sorted(
        items,
        key=lambda e: (
            0 if e.get("client_fit") == "ADERENTE" else 1,
            0 if e.get("url_especifica") else 1,
            e.get("dias_restantes") if e.get("dias_restantes") is not None else 9999,
            -(e.get("ranking_score") or 0),
        ),
    )
    chosen = ranked[0]
    return {
        "opportunity_id": chosen.get("opportunity_id"),
        "numero_controle": chosen.get("numero_controle"),
        "motivo_selecao": (
            "Selecionada por combinação de aderência ao perfil, URL/identidade específica, "
            "prazo útil e utilidade para calibrar o serviço — não apenas maior score."
        ),
        "recommendation": chosen.get("recommendation"),
        "url_oficial": chosen.get("url_oficial"),
        "documents_available": False,  # set later if edital_case succeeds
    }


def build_decision_ledger(shortlist: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ledger = []
    for e in shortlist:
        ledger.append(
            {
                "opportunity_id": e.get("opportunity_id"),
                "recommendation": e.get("recommendation"),
                "recommendation_reason": e.get("recommendation_reason"),
                "evidence_status": e.get("evidence_status"),
                "missing_client_inputs": ";".join(e.get("informacoes_extra_necessarias") or []),
                "hard_blocks": ";".join(e.get("hard_blocks") or []),
                "next_action": e.get("next_action"),
                "owner": e.get("owner") or "Tiago / Leonardo",
                "decision_deadline": e.get("decision_deadline"),
                "client_decision": "",  # empty — agent must not simulate Leonardo
                "client_notes": "",
                "tiago_review": "",
                "reviewed_at": "",
            }
        )
    return ledger


# ---------------------------------------------------------------------------
# Renderers (PDF / Excel / MD)
# ---------------------------------------------------------------------------


def write_readme(
    path: Path,
    *,
    run_id: str,
    package_title: str,
    shortlist_result: dict[str, Any],
    weekly: WeeklyValidation,
    terminal_state: str,
) -> None:
    lines = [
        f"# {package_title}",
        "",
        "Pacote de decisão comercial para Leonardo / Extra Construtora.",
        "Não requer conhecimento do código, DOD ou arquitetura interna.",
        "",
        "## Como usar (≤ 20 minutos)",
        "",
        "1. Leia `01-resumo-executivo.md` (ou o PDF).",
        "2. Abra `02-oportunidades-priorizadas.xlsx` → aba Oportunidades.",
        "3. Registre decisões em `03-decision-ledger.csv` (coluna `client_decision`).",
        "4. Use `04-intake-operacional-extra.md` na reunião de onboarding.",
        "5. Siga `06-roteiro-reuniao.md` (30–45 min).",
        "6. Tiago preenche `human-review.json` (ACCEPTED / ACCEPTED_WITH_LIMITATIONS / REJECTED).",
        "",
        "## Conteúdo",
        "",
        "| Arquivo | Função |",
        "|---------|--------|",
        "| `01-resumo-executivo.pdf` / `.md` | Narrativa executiva |",
        "| `02-oportunidades-priorizadas.xlsx` | Shortlist operacional |",
        "| `03-decision-ledger.csv` / `.json` | Registro de decisões |",
        "| `04-intake-operacional-extra.*` | Dados da Extra ainda PENDING |",
        "| `05-limitacoes-e-confiabilidade.md` | O que não pode ser afirmado |",
        "| `06-roteiro-reuniao.md` | Roteiro 30–45 min |",
        "| `07-dossie-*` | Aprofundamento ou bloqueio honesto |",
        "| `manifest.json` / `checksums.json` | Integridade |",
        "| `human-review.json` | Aceite humano (PENDING_HUMAN) |",
        "",
        "## Identidade desta rodada",
        "",
        f"- **run_id:** `{run_id}`",
        f"- **weekly cycle:** `{weekly.cycle_id}`",
        f"- **collection_id:** `{weekly.collection_id}`",
        f"- **data de corte (weekly):** `{weekly.cut_date}`",
        f"- **estado terminal técnico:** `{terminal_state}`",
        f"- **shortlist:** {shortlist_result.get('shortlist_count')} "
        f"(REVIEW={shortlist_result.get('review_count')}, "
        f"NO_GO={shortlist_result.get('no_go_count')}, "
        f"GO={shortlist_result.get('go_count')})",
        "",
        "## Pergunta âncora para Leonardo",
        "",
        "> Destas oportunidades, quais devemos investigar e quais capacidades da Extra "
        "preciso considerar para parar de tratá-las apenas como REVIEW?",
        "",
        "## Aviso",
        "",
        "- Scores **não** são probabilidade de vitória.",
        "- Ausência de valor **não** é zero.",
        "- GO comercial exige dados operacionais da Extra ainda PENDING.",
        "",
    ]
    if shortlist_result.get("insufficient"):
        lines.extend(
            [
                "## Insuficiência de shortlist",
                "",
                shortlist_result.get("insufficiency_reason") or "Shortlist incompleta.",
                "",
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_limitations_md(
    path: Path,
    *,
    weekly: WeeklyValidation,
    shortlist_result: dict[str, Any],
    extra_limitations: list[str],
) -> None:
    lines = [
        "# Limitações e confiabilidade",
        "",
        "## Fontes e freshness (weekly)",
        "",
    ]
    if weekly.freshness:
        for f in weekly.freshness:
            lines.append(f"- `{json.dumps(f, ensure_ascii=False)}`")
    else:
        lines.append("- Freshness detalhada apenas no manifest do weekly de origem.")
    lines.extend(["", "## Source health", ""])
    if weekly.source_health:
        for s in weekly.source_health:
            lines.append(f"- `{json.dumps(s, ensure_ascii=False)}`")
    else:
        lines.append("- Ver `source_health.csv` do weekly de origem.")
    lines.extend(
        [
            "",
            "## Limitações do weekly de origem",
            "",
        ]
    )
    for lim in weekly.limitations:
        lines.append(f"- {lim}")
    if weekly.exit_code not in (0, None):
        lines.append(
            f"- Weekly exit_code={weekly.exit_code} — ciclo parcial; contratos ou outra "
            "fonte podem estar bloqueados (não interpretar ausência como zero de mercado)."
        )
    lines.extend(
        [
            "",
            "## Limitações desta entrega de decisão",
            "",
            "- Nenhuma oportunidade recebe **GO** enquanto capital, garantia, capacidade "
            "simultânea, CATs/atestados e margem estiverem PENDING no perfil.",
            "- URL específica ≠ documentos oficiais baixados e hasheados.",
            "- Ranking/score **não** é probabilidade de vitória.",
            "- Ausência de valor permanece nula (não vira R$ 0).",
            "- Pacote gerado para revisão humana; aceite só por Tiago em human-review.json.",
        ]
    )
    for lim in extra_limitations:
        lines.append(f"- {lim}")
    if shortlist_result.get("insufficient"):
        lines.extend(
            [
                "",
                "## Insuficiência de shortlist",
                "",
                shortlist_result.get("insufficiency_reason") or "",
            ]
        )
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_meeting_script(
    path: Path,
    *,
    shortlist_result: dict[str, Any],
    deep_dive: dict[str, Any] | None,
    intake: dict[str, Any],
) -> None:
    sl = shortlist_result.get("shortlist") or []
    lines = [
        "# Roteiro de reunião — Extra Construtora (30–45 min)",
        "",
        "Objetivo: sair com decisões registráveis, não apenas apresentação do sistema.",
        "",
        "## 0–5 min — Contexto",
        "",
        "- Lembrar: primeira rodada de decisão B2G, não parecer jurídico nem proposta pronta.",
        "- Explicar que GO comercial depende de dados operacionais ainda PENDING.",
        f"- Candidatas no recorte: {shortlist_result.get('candidates_total')}; "
        f"shortlist: {shortlist_result.get('shortlist_count')}; "
        f"GO: {shortlist_result.get('go_count')} (deve ser 0).",
        "",
        "## 5–15 min — Shortlist",
        "",
    ]
    if not sl:
        lines.append(
            "- Não há oportunidades defensáveis com prazo futuro neste recorte. "
            "Discutir: refresh de coleta, filtros geográficos/objeto, e se a rotina semanal "
            "ainda entrega valor com o pipeline atual."
        )
    else:
        for e in sl:
            lines.append(
                f"- **#{e.get('posicao')}** `{e.get('numero_controle')}` — "
                f"{e.get('recommendation')}: {(e.get('objeto') or '')[:100]}… "
                f"| {e.get('orgao')} | dias={e.get('dias_restantes')} | "
                f"próxima ação: {e.get('next_action')}"
            )
    lines.extend(
        [
            "",
            "Perguntas:",
            "1. Quais 1–3 itens merecem tempo humano esta semana?",
            "2. Quais descartamos de forma explícita?",
            "",
            "## 15–25 min — Oportunidade sugerida para aprofundamento",
            "",
        ]
    )
    if deep_dive:
        lines.extend(
            [
                f"- Selecionada: `{deep_dive.get('numero_controle')}`",
                f"- Motivo: {deep_dive.get('motivo_selecao')}",
                f"- URL: {deep_dive.get('url_oficial') or '—'}",
                "- Decisão: manter REVIEW / pedir documentos / descartar.",
            ]
        )
    else:
        lines.append("- Nenhuma oportunidade elegível para dossiê nesta rodada (ver 07-dossie-edital-NOT_AVAILABLE.md).")
    lines.extend(
        [
            "",
            "## 25–35 min — Intake operacional",
            "",
            "Cobrir as perguntas de `04-intake-operacional-extra.md` (sem inventar respostas):",
            "",
        ]
    )
    for q in (intake.get("questions") or [])[:10]:
        lines.append(f"- {q['id']}: {q['question']}")
    lines.extend(
        [
            "",
            "## 35–45 min — Decisões e responsáveis",
            "",
            "Registrar no ledger (`03-decision-ledger.csv`):",
            "",
            "| Decisão | Responsável | Prazo |",
            "|---------|-------------|-------|",
            "| Oportunidades a investigar (IDs) | Leonardo | 48h |",
            "| Oportunidades descartadas (IDs) | Leonardo | 48h |",
            "| Envio de dados de intake (capital, CATs, etc.) | Leonardo/Extra | 7 dias |",
            "| Próxima rodada semanal / refresh | Tiago | 7 dias |",
            "| Aceite do pacote (human-review) | Tiago | 24h |",
            "",
            "Encerramento: confirmar se o serviço já oferece valor suficiente para iniciar a rotina contratada.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_executive_md(
    path: Path,
    *,
    run_id: str,
    package_title: str,
    profile: dict[str, Any],
    weekly: WeeklyValidation,
    shortlist_result: dict[str, Any],
    deep_dive: dict[str, Any] | None,
    intake: dict[str, Any],
    terminal_state: str,
) -> None:
    sl = shortlist_result.get("shortlist") or []
    lines = [
        f"# {package_title}",
        "",
        f"**run_id:** `{run_id}`  ",
        f"**Perfil:** {profile.get('display_name')} v{profile.get('version')}  ",
        f"**Weekly:** `{weekly.cycle_id}` · collection `{weekly.collection_id}`  ",
        f"**Corte weekly:** {weekly.cut_date}  ",
        f"**Estado:** `{terminal_state}`",
        "",
        "## 1. Propósito da rodada",
        "",
        "Entregar a Leonardo uma primeira rodada de decisão B2G defensável: o que merece "
        "análise aprofundada, o que descartar, e quais dados operacionais da Extra ainda "
        "faltam para decisões GO futuras.",
        "",
        "## 2. Situação atual",
        "",
        f"- Candidatas no pacote semanal: **{shortlist_result.get('candidates_total')}**",
        f"- Bloqueadas (NO_GO): **{shortlist_result.get('blocked_total')}**",
        f"- Defensáveis (REVIEW com prazo futuro): **{shortlist_result.get('review_defensible_total')}**",
        f"- Shortlist final: **{shortlist_result.get('shortlist_count')}** "
        f"(REVIEW={shortlist_result.get('review_count')}, GO={shortlist_result.get('go_count')})",
        f"- GO bloqueado por perfil PENDING: **{'sim' if shortlist_result.get('go_blocked_by_profile') else 'não'}**",
        f"- Campos críticos PENDING: {', '.join(shortlist_result.get('critical_pending') or []) or '—'}",
        "",
        "## 3. Metodologia resumida",
        "",
        "1. Partir do ciclo semanal real (manifest + checksums validados).",
        "2. Avaliar cada oportunidade: prazo futuro, identidade, status, URL, aderência ao perfil.",
        "3. Aplicar hard blocks; score nunca compensa eliminatória.",
        "4. Proibir GO enquanto elicitation crítica estiver PENDING.",
        "5. Montar shortlist 5–10 (ou declarar insuficiência).",
        "6. Gerar intake, ledger, roteiro e dossiê (ou bloqueio honesto).",
        "",
        "## 4. Confiabilidade e limitações",
        "",
        f"- Exit code do weekly: `{weekly.exit_code}`",
        "- Ver detalhe em `05-limitacoes-e-confiabilidade.md`.",
        "- Ausência de contratos/concorrentes no weekly **não** é zero de mercado se a fonte falhou.",
        "",
        "## 5. Oportunidades priorizadas",
        "",
    ]
    if shortlist_result.get("insufficient"):
        lines.append(f"> **Insuficiência:** {shortlist_result.get('insufficiency_reason')}")
        lines.append("")
    if not sl:
        lines.append("Nenhuma oportunidade defensável com prazo futuro neste recorte.")
    else:
        lines.append("| # | Rec. | Controle | Órgão | Dias | Valor | Motivo |")
        lines.append("|---|------|----------|-------|------|-------|--------|")
        for e in sl:
            val = br_currency(e["valor"]) if e.get("valor") is not None else "—"
            lines.append(
                f"| {e.get('posicao')} | {e.get('recommendation')} | "
                f"`{e.get('numero_controle')}` | {(e.get('orgao') or '')[:40]} | "
                f"{e.get('dias_restantes')} | {val} | "
                f"{(e.get('recommendation_reason') or '')[:80]} |"
            )
    lines.extend(
        [
            "",
            "### 5b. Contexto recente de engenharia (não recomendado)",
            "",
            "Objetos aderentes ao perfil, porém com hard block (tipicamente prazo encerrado). "
            "Úteis para calibrar órgãos/tipos de obra — **não** para GO/REVIEW nesta rodada.",
            "",
        ]
    )
    ctx = list(shortlist_result.get("context_recent_no_go") or [])[:10]
    if not ctx:
        lines.append("_Sem contexto de engenharia recente._")
    else:
        lines.append("| # | Controle | Órgão | Limite | Valor | Motivo NO_GO |")
        lines.append("|---|----------|-------|--------|-------|--------------|")
        for i, e in enumerate(ctx, start=1):
            val = br_currency(e["valor"]) if e.get("valor") is not None else "—"
            lines.append(
                f"| C{i} | `{e.get('numero_controle')}` | {(e.get('orgao') or '')[:36]} | "
                f"{e.get('data_limite')} | {val} | {(e.get('recommendation_reason') or '')[:60]} |"
            )
    lines.extend(
        [
            "",
            "## 6. Oportunidade sugerida para aprofundamento",
            "",
        ]
    )
    if deep_dive:
        lines.extend(
            [
                f"- **Controle:** `{deep_dive.get('numero_controle')}`",
                f"- **Motivo:** {deep_dive.get('motivo_selecao')}",
                f"- **URL:** {deep_dive.get('url_oficial') or '—'}",
                f"- **Documentos oficiais no dossiê:** "
                f"{'sim' if deep_dive.get('documents_available') else 'não (ver 07)'}",
            ]
        )
    else:
        lines.append("- Nenhum caso elegível; ver `07-dossie-edital-NOT_AVAILABLE.md`.")
    lines.extend(
        [
            "",
            "## 7. Dados da Extra ainda necessários",
            "",
        ]
    )
    for q in (intake.get("questions") or [])[:10]:
        lines.append(f"- **{q['id']}** ({q['key']}): {q['question']}")
    lines.extend(
        [
            "",
            "## 8. Decisões solicitadas ao cliente",
            "",
            "1. Quais 1–3 oportunidades investigar?",
            "2. Quais descartar explicitamente?",
            "3. Quais órgãos/tipos de obra/faixas de valor priorizar?",
            "4. Quais respostas de intake podem ser fornecidas em até 7 dias?",
            "5. O serviço já oferece valor suficiente para iniciar a rotina contratada?",
            "",
            "## 9. Próximos sete dias",
            "",
            "- Tiago: revisar pacote (human-review) e conduzir reunião.",
            "- Leonardo: preencher decisões no ledger + intake prioritário.",
            "- Operação: refresh de coleta se shortlist insuficiente por prazo/stale.",
            "- Próxima rodada: reavaliar com perfil parcialmente preenchido.",
            "",
            "## 10. Apêndice de fontes",
            "",
            f"- Weekly pack: `{weekly.weekly_dir}`",
            f"- cycle_id: `{weekly.cycle_id}`",
            f"- collection_id: `{weekly.collection_id}`",
            f"- Perfil: `config/client_profiles/extra.yaml` sha256=`{profile.get('_sha256')}`",
            "- PNCP como origem das oportunidades do recorte.",
            "",
            "### O que foi encontrado",
            f"{shortlist_result.get('candidates_total')} candidatas; "
            f"{shortlist_result.get('shortlist_count')} na shortlist.",
            "",
            "### O que merece atenção",
            "Itens REVIEW da shortlist e campos PENDING do intake.",
            "",
            "### O que não pode ser afirmado",
            "Probabilidade de vitória; capacidade financeira/técnica da Extra; "
            "completude do mercado quando fonte falhou.",
            "",
            "### Qual decisão precisa ser tomada",
            "Priorizar 1–3 oportunidades e liberar dados de intake.",
            "",
            "### Qual informação precisa ser fornecida",
            "Ver seção 7 e `04-intake-operacional-extra.md`.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_executive_pdf(
    path: Path,
    *,
    run_id: str,
    package_title: str,
    profile: dict[str, Any],
    weekly: WeeklyValidation,
    shortlist_result: dict[str, Any],
    deep_dive: dict[str, Any] | None,
    intake: dict[str, Any],
    terminal_state: str,
) -> int:
    """Compact executive PDF (≈6–12 pages). Returns estimated page count."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="CoverTitle",
            parent=styles["Title"],
            fontSize=16,
            leading=20,
            alignment=TA_CENTER,
            spaceAfter=12,
        )
    )
    styles.add(
        ParagraphStyle(
            name="BodyJust",
            parent=styles["BodyText"],
            fontSize=10,
            leading=13,
            alignment=TA_JUSTIFY,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="H2BR",
            parent=styles["Heading2"],
            fontSize=12,
            spaceBefore=10,
            spaceAfter=6,
            textColor=colors.HexColor("#0B3D5C"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="Small",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
        )
    )

    story: list[Any] = []
    story.append(Paragraph(package_title, styles["CoverTitle"]))
    story.append(
        Paragraph(
            f"run_id: {run_id}<br/>Perfil: {profile.get('display_name')} v{profile.get('version')}<br/>"
            f"Weekly: {weekly.cycle_id}<br/>Collection: {weekly.collection_id}<br/>"
            f"Corte: {weekly.cut_date}<br/>Estado: {terminal_state}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.4 * cm))

    sections = [
        (
            "1. Propósito da rodada",
            "Entregar a Leonardo uma primeira rodada de decisão B2G defensável e acionável: "
            "escolher o que investigar, o que descartar e quais dados da Extra ainda faltam "
            "para decisões GO futuras. Isto não é parecer jurídico nem proposta pronta.",
        ),
        (
            "2. Situação atual",
            f"Candidatas no weekly: {shortlist_result.get('candidates_total')}. "
            f"Bloqueadas: {shortlist_result.get('blocked_total')}. "
            f"Defensáveis com prazo futuro: {shortlist_result.get('review_defensible_total')}. "
            f"Shortlist: {shortlist_result.get('shortlist_count')} "
            f"(REVIEW={shortlist_result.get('review_count')}, GO={shortlist_result.get('go_count')}). "
            f"Campos críticos PENDING: {', '.join(shortlist_result.get('critical_pending') or []) or 'nenhum'}.",
        ),
        (
            "3. Metodologia resumida",
            "Validação do weekly (manifest+checksums) → avaliação por prazo, identidade, status, "
            "URL e aderência ao perfil Extra → hard blocks → proibição de GO com PENDING críticos "
            "→ shortlist 5–10 ou insuficiência explícita → intake + ledger + roteiro.",
        ),
        (
            "4. Confiabilidade e limitações",
            f"Weekly exit_code={weekly.exit_code}. Scores não são probabilidade de vitória. "
            "Ausência de valor não é zero. Falha de fonte não é zero de mercado. "
            "Detalhes em 05-limitacoes-e-confiabilidade.md.",
        ),
    ]
    for title, body in sections:
        story.append(Paragraph(title, styles["H2BR"]))
        story.append(Paragraph(body, styles["BodyJust"]))

    story.append(Paragraph("5. Oportunidades priorizadas (shortlist defensável)", styles["H2BR"]))
    if shortlist_result.get("insufficient"):
        story.append(
            Paragraph(
                f"<b>Insuficiência:</b> {shortlist_result.get('insufficiency_reason')}",
                styles["BodyJust"],
            )
        )
    sl = shortlist_result.get("shortlist") or []
    if not sl:
        story.append(
            Paragraph(
                "Nenhuma oportunidade defensável com prazo futuro neste recorte. "
                "Isso não é “zero de mercado”: o recorte/coleta atual não sustenta recomendações "
                "REVIEW com prazo útil. A reunião deve focar em (a) intake operacional, "
                "(b) priorização de órgãos/tipos de obra e (c) decisão de rotina de refresh semanal.",
                styles["BodyJust"],
            )
        )
    else:
        data = [["#", "Rec.", "Controle", "Órgão", "Dias", "Valor"]]
        for e in sl:
            data.append(
                [
                    str(e.get("posicao")),
                    str(e.get("recommendation")),
                    str(e.get("numero_controle") or "")[:28],
                    (e.get("orgao") or "")[:32],
                    str(e.get("dias_restantes")),
                    br_currency(e["valor"]) if e.get("valor") is not None else "—",
                ]
            )
        t = Table(data, colWidths=[1 * cm, 1.8 * cm, 4.2 * cm, 4.5 * cm, 1.5 * cm, 3 * cm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D5C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(t)
        story.append(Spacer(1, 0.3 * cm))
        for e in sl:
            story.append(
                Paragraph(
                    f"<b>#{e.get('posicao')} {e.get('numero_controle')}</b> — "
                    f"{(e.get('objeto') or '')[:180]} "
                    f"<i>({e.get('recommendation_reason')})</i>",
                    styles["Small"],
                )
            )

    story.append(PageBreak())
    story.append(
        Paragraph(
            "5b. Contexto recente de engenharia (NÃO é shortlist recomendada)",
            styles["H2BR"],
        )
    )
    story.append(
        Paragraph(
            "Itens com aderência de objeto ao perfil Extra, mas com hard block (em geral prazo "
            "já encerrado). Servem para calibrar o tipo de obra/órgão — <b>não</b> para decidir GO/REVIEW "
            "nesta rodada.",
            styles["BodyJust"],
        )
    )
    ctx = list(shortlist_result.get("context_recent_no_go") or [])[:10]
    if not ctx:
        story.append(Paragraph("Sem contexto de engenharia recente no recorte.", styles["BodyJust"]))
    else:
        for i, e in enumerate(ctx, start=1):
            val = br_currency(e["valor"]) if e.get("valor") is not None else "—"
            story.append(
                Paragraph(
                    f"<b>C{i}. {e.get('numero_controle')}</b> — {e.get('orgao')} "
                    f"({e.get('municipio')}/{e.get('uf')}) · limite {br_date(e.get('data_limite'))} · "
                    f"{val}<br/>{(e.get('objeto') or '')[:220]}<br/>"
                    f"<i>NO_GO: {e.get('recommendation_reason')}</i> · "
                    f"URL: {e.get('url_oficial') or '—'}",
                    styles["Small"],
                )
            )
            story.append(Spacer(1, 0.15 * cm))

    story.append(Paragraph("6. Oportunidade sugerida para aprofundamento", styles["H2BR"]))
    if deep_dive:
        story.append(
            Paragraph(
                f"Controle <b>{deep_dive.get('numero_controle')}</b>. "
                f"{deep_dive.get('motivo_selecao')} "
                f"URL: {deep_dive.get('url_oficial') or '—'}. "
                f"Documentos no dossiê: {'sim' if deep_dive.get('documents_available') else 'não'}.",
                styles["BodyJust"],
            )
        )
    else:
        story.append(
            Paragraph(
                "Nenhum caso elegível para dossiê com documentos oficiais e prazo futuro útil "
                "nesta rodada. Ver 07-dossie-edital-NOT_AVAILABLE.md.",
                styles["BodyJust"],
            )
        )

    story.append(PageBreak())
    story.append(Paragraph("7. Dados da Extra ainda necessários (intake)", styles["H2BR"]))
    story.append(
        Paragraph(
            "Enquanto capital, garantia, capacidade simultânea, CATs/atestados e margem "
            "permanecerem PENDING, <b>nenhuma</b> oportunidade pode receber GO. As perguntas abaixo "
            "não têm respostas inventadas.",
            styles["BodyJust"],
        )
    )
    for q in (intake.get("questions") or [])[:10]:
        story.append(
            Paragraph(
                f"• <b>{q['id']}</b> [{', '.join(q.get('profile_fields') or [])}]: {q['question']}",
                styles["BodyJust"],
            )
        )

    story.append(Paragraph("8. Decisões solicitadas ao cliente", styles["H2BR"]))
    story.append(
        Paragraph(
            "1) Quais 1–3 oportunidades investigar (se houver shortlist)? "
            "2) Quais descartar explicitamente? "
            "3) Quais órgãos, tipos de obra e faixas de valor devem ser prioritários na rotina? "
            "4) Quais respostas de intake podem ser enviadas em até 7 dias? "
            "5) O serviço já justifica iniciar a rotina contratada mesmo com shortlist parcial?",
            styles["BodyJust"],
        )
    )

    story.append(Paragraph("9. Próximos sete dias", styles["H2BR"]))
    story.append(
        Paragraph(
            "Dia 0–1: Tiago revisa o pacote (human-review.json) e agenda reunião.<br/>"
            "Dia 1–2: reunião com Leonardo (roteiro 06); registrar ledger.<br/>"
            "Dia 2–5: Extra envia respostas prioritárias de intake (capital, CATs, capacidade).<br/>"
            "Dia 3–7: operação executa refresh do ciclo semanal; nova shortlist com prazos futuros.<br/>"
            "Dia 7: segunda rodada de decisão com perfil parcialmente preenchido.",
            styles["BodyJust"],
        )
    )

    story.append(PageBreak())
    story.append(Paragraph("9b. Como ler este pacote em 20 minutos", styles["H2BR"]))
    story.append(
        Paragraph(
            "Minutos 0–3: abrir 00-LEIA-ME e este resumo (propósito + contagens).<br/>"
            "Minutos 3–8: Excel → abas Resumo, Oportunidades e Contexto_Recente.<br/>"
            "Minutos 8–12: intake 04 (perguntas PENDING) e limitações 05.<br/>"
            "Minutos 12–18: roteiro 06 e preencher mentalmente as decisões do ledger.<br/>"
            "Minutos 18–20: Tiago preenche human-review.json (sem autoaceite).",
            styles["BodyJust"],
        )
    )
    story.append(
        Paragraph(
            "Se a shortlist estiver insuficiente, a decisão útil da reunião não é “forçar GO”, "
            "e sim: (1) liberar dados de intake, (2) definir órgãos/tipos prioritários, "
            "(3) autorizar refresh semanal, (4) confirmar se a rotina contratada começa assim.",
            styles["BodyJust"],
        )
    )
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph("Checklist de valor comercial", styles["H2BR"]))
    story.append(
        Paragraph(
            "□ Leonardo consegue apontar 0–3 frentes de investigação ou declarar “ainda não há prazo útil”.<br/>"
            "□ Extra entende quais dados operacionais desbloqueiam GO futuro.<br/>"
            "□ Tiago consegue conduzir a reunião sem falar de DOD, PR ou arquitetura.<br/>"
            "□ Ninguém trata score como probabilidade de vitória.<br/>"
            "□ Ninguém interpreta CSV vazio de contratos/concorrentes como ausência de mercado.",
            styles["BodyJust"],
        )
    )

    story.append(PageBreak())
    story.append(Paragraph("10. Apêndice de fontes e claims", styles["H2BR"]))
    story.append(
        Paragraph(
            f"<b>Weekly:</b> {weekly.cycle_id}<br/>"
            f"<b>Collection:</b> {weekly.collection_id}<br/>"
            f"<b>Corte:</b> {weekly.cut_date} · exit_code={weekly.exit_code}<br/>"
            f"<b>Perfil:</b> {profile.get('display_name')} v{profile.get('version')} "
            f"(sha256={profile.get('_sha256')})<br/>"
            f"<b>Origem:</b> PNCP via pacote semanal validado (manifest + checksums).<br/>"
            f"<b>run_id deste pacote:</b> {run_id}",
            styles["BodyJust"],
        )
    )
    story.append(Paragraph("<b>Claims permitidos</b>", styles["H2BR"]))
    story.append(
        Paragraph(
            "• Contagens e shortlist derivadas do weekly pack validado.<br/>"
            "• Intake derivado apenas de campos PENDING do perfil.<br/>"
            "• GO bloqueado por política de elicitation (não por “score baixo”).",
            styles["BodyJust"],
        )
    )
    story.append(Paragraph("<b>Claims proibidos</b>", styles["H2BR"]))
    story.append(
        Paragraph(
            "• Probabilidade de vitória.<br/>"
            "• Capacidade financeira/técnica inventada da Extra.<br/>"
            "• Zero de mercado a partir de fonte falha ou CSV vazio.<br/>"
            "• Aceite humano automático.<br/>"
            "• GO com campos críticos PENDING.",
            styles["BodyJust"],
        )
    )
    story.append(Spacer(1, 0.4 * cm))
    story.append(
        Paragraph(
            "Pergunta âncora para Leonardo:<br/><i>Destas oportunidades, quais devemos investigar "
            "e quais capacidades da Extra preciso considerar para parar de tratá-las apenas "
            "como REVIEW?</i>",
            styles["BodyJust"],
        )
    )
    story.append(Spacer(1, 0.3 * cm))
    story.append(
        Paragraph(
            "Limitações do weekly de origem (amostra): "
            + "; ".join((weekly.limitations or [])[:6]),
            styles["Small"],
        )
    )

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        title=package_title,
        author="Extra Consultoria",
    )
    # track pages
    page_count = {"n": 0}

    def _page(canvas: Any, _doc: Any) -> None:
        page_count["n"] += 1
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawString(1.6 * cm, 1 * cm, f"{package_title} · {run_id}")
        canvas.drawRightString(A4[0] - 1.6 * cm, 1 * cm, f"{page_count['n']}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_page, onLaterPages=_page)
    return max(page_count["n"], 1)


def write_excel(
    path: Path,
    *,
    run_id: str,
    profile: dict[str, Any],
    weekly: WeeklyValidation,
    shortlist_result: dict[str, Any],
    intake: dict[str, Any],
    ledger: list[dict[str, Any]],
    source_health_rows: list[dict[str, str]],
    orgaos_rows: list[dict[str, str]],
) -> dict[str, Any]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0B3D5C")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    def style_header(ws: Any) -> None:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.freeze_panes = "A2"

    # Resumo
    ws = wb.active
    ws.title = "Resumo"
    kpis = [
        ("run_id", run_id),
        ("profile_version", profile.get("version")),
        ("profile_sha256", profile.get("_sha256")),
        ("weekly_cycle_id", weekly.cycle_id),
        ("collection_id", weekly.collection_id),
        ("cut_date", weekly.cut_date),
        ("candidates_total", shortlist_result.get("candidates_total")),
        ("blocked_total", shortlist_result.get("blocked_total")),
        ("shortlist_count", shortlist_result.get("shortlist_count")),
        ("review_count", shortlist_result.get("review_count")),
        ("no_go_count", shortlist_result.get("no_go_count")),
        ("go_count", shortlist_result.get("go_count")),
        ("insufficient", shortlist_result.get("insufficient")),
        ("go_blocked_by_profile", shortlist_result.get("go_blocked_by_profile")),
    ]
    ws.append(["campo", "valor"])
    for k, v in kpis:
        ws.append([k, v if not isinstance(v, (list, dict)) else json.dumps(v, ensure_ascii=False)])
    style_header(ws)

    # Oportunidades
    ws_o = wb.create_sheet("Oportunidades")
    headers = [
        "posicao",
        "recommendation",
        "numero_controle",
        "objeto",
        "orgao",
        "orgao_cnpj",
        "municipio",
        "uf",
        "modalidade",
        "data_limite",
        "dias_restantes",
        "valor",
        "valor_semantica",
        "url_oficial",
        "client_fit",
        "hard_blocks",
        "next_action",
        "recommendation_reason",
        "evidence_status",
        "collection_id",
        "cycle_run_id",
        "data_corte",
    ]
    ws_o.append(headers)
    for e in shortlist_result.get("shortlist") or []:
        ws_o.append(
            [
                e.get("posicao"),
                e.get("recommendation"),
                e.get("numero_controle"),
                e.get("objeto"),
                e.get("orgao"),
                e.get("orgao_cnpj"),
                e.get("municipio"),
                e.get("uf"),
                e.get("modalidade"),
                e.get("data_limite"),
                e.get("dias_restantes"),
                e.get("valor"),
                e.get("valor_semantica"),
                e.get("url_oficial"),
                e.get("client_fit"),
                ";".join(e.get("hard_blocks") or []),
                e.get("next_action"),
                e.get("recommendation_reason"),
                e.get("evidence_status"),
                e.get("collection_id"),
                e.get("cycle_run_id"),
                e.get("data_corte"),
            ]
        )
    style_header(ws_o)

    # Evidências
    ws_e = wb.create_sheet("Evidências")
    ws_e.append(
        ["opportunity_id", "url_oficial", "url_especifica", "evidence_status", "termos_positivos", "termos_negativos"]
    )
    for e in shortlist_result.get("shortlist") or []:
        ws_e.append(
            [
                e.get("opportunity_id"),
                e.get("url_oficial"),
                e.get("url_especifica"),
                e.get("evidence_status"),
                ";".join(e.get("termos_positivos") or []),
                ";".join(e.get("termos_negativos") or []),
            ]
        )
    style_header(ws_e)

    # Dados_Pendentes_Extra
    ws_p = wb.create_sheet("Dados_Pendentes_Extra")
    ws_p.append(["id", "key", "question", "profile_fields", "status", "answer"])
    for q in intake.get("questions") or []:
        ws_p.append(
            [
                q.get("id"),
                q.get("key"),
                q.get("question"),
                ",".join(q.get("profile_fields") or []),
                q.get("status"),
                q.get("answer"),
            ]
        )
    style_header(ws_p)

    # Órgãos
    ws_org = wb.create_sheet("Órgãos")
    if orgaos_rows:
        org_headers = list(orgaos_rows[0].keys())
        ws_org.append(org_headers)
        for r in orgaos_rows:
            ws_org.append([r.get(h) for h in org_headers])
        style_header(ws_org)
    else:
        # derive from shortlist
        ws_org.append(["orgao", "orgao_cnpj", "municipio", "uf", "qtd_shortlist"])
        agg: dict[str, dict[str, Any]] = {}
        for e in shortlist_result.get("shortlist") or []:
            key = e.get("orgao_cnpj") or e.get("orgao") or "?"
            if key not in agg:
                agg[key] = {
                    "orgao": e.get("orgao"),
                    "orgao_cnpj": e.get("orgao_cnpj"),
                    "municipio": e.get("municipio"),
                    "uf": e.get("uf"),
                    "qtd": 0,
                }
            agg[key]["qtd"] += 1
        for v in agg.values():
            ws_org.append([v["orgao"], v["orgao_cnpj"], v["municipio"], v["uf"], v["qtd"]])
        style_header(ws_org)

    # Concorrentes — honest empty if not in weekly
    ws_c = wb.create_sheet("Concorrentes")
    ws_c.append(["nota"])
    ws_c.append(
        [
            "Sem lista de concorrentes neste weekly (arquivo vazio ou fonte bloqueada). "
            "Não interpretar como ausência de concorrência no mercado."
        ]
    )
    style_header(ws_c)

    # Valores
    ws_v = wb.create_sheet("Valores")
    ws_v.append(["opportunity_id", "valor", "valor_semantica", "nota"])
    valor_rows = list(shortlist_result.get("shortlist") or [])
    if not valor_rows:
        valor_rows = list(shortlist_result.get("context_recent_no_go") or [])[:15]
    for e in valor_rows:
        nota = "valor ausente (null) — não é zero" if e.get("valor") is None else ""
        if e.get("recommendation") == "NO_GO":
            nota = (nota + " | contexto NO_GO — não é shortlist").strip(" |")
        ws_v.append([e.get("opportunity_id"), e.get("valor"), e.get("valor_semantica"), nota])
    style_header(ws_v)

    # Contexto recente (engenharia NO_GO — transparência, não recomendação)
    ws_ctx = wb.create_sheet("Contexto_Recente")
    ws_ctx.append(
        [
            "numero_controle",
            "recommendation",
            "orgao",
            "municipio",
            "uf",
            "data_limite",
            "dias_restantes",
            "valor",
            "client_fit",
            "hard_blocks",
            "objeto",
            "url_oficial",
            "nota",
        ]
    )
    for e in shortlist_result.get("context_recent_no_go") or []:
        ws_ctx.append(
            [
                e.get("numero_controle"),
                e.get("recommendation"),
                e.get("orgao"),
                e.get("municipio"),
                e.get("uf"),
                e.get("data_limite"),
                e.get("dias_restantes"),
                e.get("valor"),
                e.get("client_fit"),
                ";".join(e.get("hard_blocks") or []),
                e.get("objeto"),
                e.get("url_oficial"),
                "NÃO é shortlist defensável — apenas contexto de calibração",
            ]
        )
    style_header(ws_ctx)

    # Source_Health
    ws_s = wb.create_sheet("Source_Health")
    if source_health_rows:
        sh = list(source_health_rows[0].keys())
        ws_s.append(sh)
        for r in source_health_rows:
            ws_s.append([r.get(h) for h in sh])
        style_header(ws_s)
    else:
        ws_s.append(["source", "note"])
        for s in weekly.source_health:
            ws_s.append([str(s), ""])
        style_header(ws_s)

    # Limitações
    ws_l = wb.create_sheet("Limitações")
    ws_l.append(["limitacao"])
    for lim in weekly.limitations:
        ws_l.append([lim])
    if shortlist_result.get("insufficient"):
        ws_l.append([shortlist_result.get("insufficiency_reason")])
    ws_l.append(["GO proibido com campos críticos PENDING"])
    ws_l.append(["Score não é probabilidade de vitória"])
    style_header(ws_l)

    # Metadados
    ws_m = wb.create_sheet("Metadados")
    ws_m.append(["campo", "valor"])
    for k, v in kpis:
        ws_m.append([k, v if not isinstance(v, (list, dict)) else json.dumps(v, ensure_ascii=False)])
    ws_m.append(["package_title", PACKAGE_TITLE])
    ws_m.append(["schema", SCHEMA])
    style_header(ws_m)

    # Ledger sheet (extra convenience)
    ws_led = wb.create_sheet("Decision_Ledger")
    if ledger:
        lh = list(ledger[0].keys())
        ws_led.append(lh)
        for r in ledger:
            ws_led.append([r.get(h) for h in lh])
        style_header(ws_led)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    return {
        "run_id": run_id,
        "shortlist_count": shortlist_result.get("shortlist_count"),
        "review_count": shortlist_result.get("review_count"),
        "go_count": shortlist_result.get("go_count"),
        "candidates_total": shortlist_result.get("candidates_total"),
        "cut_date": weekly.cut_date,
        "profile_version": profile.get("version"),
        "collection_id": weekly.collection_id,
        "cycle_run_id": weekly.cycle_id,
    }


def reconcile_package_counts(
    *,
    shortlist_result: dict[str, Any],
    excel_meta: dict[str, Any],
    pdf_meta: dict[str, Any],
    weekly: WeeklyValidation,
    profile: dict[str, Any],
) -> dict[str, Any]:
    divergences: list[str] = []
    checks = {
        "shortlist_count": shortlist_result.get("shortlist_count"),
        "go_count": shortlist_result.get("go_count"),
        "review_count": shortlist_result.get("review_count"),
        "candidates_total": shortlist_result.get("candidates_total"),
        "cut_date": weekly.cut_date,
        "profile_version": profile.get("version"),
        "collection_id": weekly.collection_id,
        "cycle_run_id": weekly.cycle_id,
    }
    for key, expected in checks.items():
        if excel_meta.get(key) != expected:
            divergences.append(f"excel.{key}={excel_meta.get(key)} != {expected}")
        if key in pdf_meta and pdf_meta.get(key) != expected:
            divergences.append(f"pdf.{key}={pdf_meta.get(key)} != {expected}")
    if shortlist_result.get("go_count") != 0 and shortlist_result.get("go_blocked_by_profile"):
        divergences.append("go_count deve ser 0 com perfil PENDING crítico")
    status = "PASS" if not divergences else "FAIL"
    return {"status": status, "divergences": divergences, "checks": checks}


def build_human_review(run_id: str, checksums: dict[str, str]) -> dict[str, Any]:
    return {
        "status": "PENDING_HUMAN",
        "reviewed_by": None,
        "reviewed_at": None,
        "package_run_id": run_id,
        "package_checksums": checksums,
        "decision": None,
        "limitations_accepted": [],
        "notes": None,
        "allowed_decisions": ["ACCEPTED", "ACCEPTED_WITH_LIMITATIONS", "REJECTED"],
        "rule": "Somente Tiago pode preencher decision/reviewed_by; agente não autoaceita",
    }


def assert_not_auto_accepted(human_review: dict[str, Any]) -> None:
    if human_review.get("status") != "PENDING_HUMAN":
        raise ValueError("human-review não pode iniciar fora de PENDING_HUMAN")
    if human_review.get("reviewed_by") is not None:
        raise ValueError("reviewed_by deve ser null no pacote gerado por agente")
    if human_review.get("decision") is not None:
        raise ValueError("decision deve ser null no pacote gerado por agente")


def reject_stale_acceptance(
    previous: dict[str, Any] | None, current_checksums: dict[str, str]
) -> bool:
    """Return True if previous acceptance must be rejected (checksum mismatch)."""
    if not previous:
        return False
    if previous.get("decision") not in {"ACCEPTED", "ACCEPTED_WITH_LIMITATIONS"}:
        return False
    prev_cs = previous.get("package_checksums") or {}
    return prev_cs != current_checksums


def write_dossie_not_available(
    path: Path,
    *,
    deep_dive: dict[str, Any] | None,
    reason: str,
    how_to_obtain: list[str],
) -> None:
    lines = [
        "# Dossiê de edital — NÃO DISPONÍVEL",
        "",
        "## Bloqueio",
        "",
        reason,
        "",
    ]
    if deep_dive:
        lines.extend(
            [
                "## Oportunidade candidata (sem documentos oficiais no pacote)",
                "",
                f"- Controle: `{deep_dive.get('numero_controle')}`",
                f"- URL: {deep_dive.get('url_oficial') or '—'}",
                f"- Motivo da seleção lógica: {deep_dive.get('motivo_selecao')}",
                "",
            ]
        )
    lines.extend(["## Como obter os documentos", ""])
    for h in how_to_obtain:
        lines.append(f"- {h}")
    lines.extend(
        [
            "",
            "## O que NÃO foi feito",
            "",
            "- Não foi usado corpus fabricado.",
            "- Não foi desbloqueado PR #133 artificialmente.",
            "- Não foi emitido parecer jurídico nem GO.",
            "",
            "O restante da entrega (shortlist, intake, ledger, roteiro) permanece utilizável.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


def run_delivery(
    *,
    weekly_input: Path,
    delivery_out: Path,
    profile_path: Path = DEFAULT_PROFILE,
    as_of: date | None = None,
    client_ready_dsn: str | None = None,
    allow_partial: bool = True,
) -> dict[str, Any]:
    """Produce the client package. Returns result dict; raises SystemExit on hard fail."""
    as_of = as_of or date.today()
    delivery_out = delivery_out.resolve()
    delivery_out.mkdir(parents=True, exist_ok=True)

    # Isolation: never write outside delivery_out for client artifacts
    weekly = validate_weekly_pack(Path(weekly_input))
    if not weekly.ok:
        result = {
            "terminal_state": "FAILED_VALIDATION",
            "exit_code": 2,
            "errors": weekly.errors,
            "run_id": None,
            "delivery_out": str(delivery_out),
        }
        (delivery_out / "FAILED_VALIDATION.json").write_text(
            json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
        print(json.dumps(result, indent=2, ensure_ascii=False))
        raise SystemExit(2)

    if not profile_path.is_file():
        print("profile missing", profile_path, file=sys.stderr)
        raise SystemExit(2)

    profile = load_profile(profile_path)
    run_id = new_run_id()
    rows = load_csv_rows(Path(weekly.opportunities_path or ""))
    # Empty bank of opportunities → not READY as full decision shortlist, but can still emit package with insufficiency
    if not rows and not allow_partial:
        result = {
            "terminal_state": "FAILED_VALIDATION",
            "exit_code": 2,
            "errors": ["opportunities.csv vazio — não gerar pacote READY"],
            "run_id": run_id,
        }
        print(json.dumps(result, indent=2, ensure_ascii=False))
        raise SystemExit(2)

    shortlist_result = build_shortlist(
        rows,
        profile=profile,
        as_of=as_of,
        cycle_id=weekly.cycle_id,
        collection_id=weekly.collection_id,
        cut_date=weekly.cut_date,
    )
    intake = build_intake(profile)
    ledger = build_decision_ledger(shortlist_result.get("shortlist") or [])
    deep_dive = select_deep_dive(shortlist_result)

    # Optional DSN is recorded but not required for weekly-based composition
    dsn_note = None
    if client_ready_dsn:
        # refuse obvious production writes — we only accept local-looking DSNs for optional paths
        low = client_ready_dsn.lower()
        if any(x in low for x in ("ec-prod", "prod-db", "production")):
            result = {
                "terminal_state": "FAILED_VALIDATION",
                "exit_code": 2,
                "errors": ["CLIENT_READY_DSN parece produção — recusado"],
                "run_id": run_id,
            }
            print(json.dumps(result, indent=2, ensure_ascii=False))
            raise SystemExit(2)
        dsn_note = "DSN isolado fornecido (não usado para escrita em produção)"

    # Write core artifacts
    write_readme(
        delivery_out / "00-LEIA-ME.md",
        run_id=run_id,
        package_title=PACKAGE_TITLE,
        shortlist_result=shortlist_result,
        weekly=weekly,
        terminal_state="PENDING_BUILD",
    )
    write_executive_md(
        delivery_out / "01-resumo-executivo.md",
        run_id=run_id,
        package_title=PACKAGE_TITLE,
        profile=profile,
        weekly=weekly,
        shortlist_result=shortlist_result,
        deep_dive=deep_dive,
        intake=intake,
        terminal_state="PENDING_BUILD",
    )
    pdf_pages = write_executive_pdf(
        delivery_out / "01-resumo-executivo.pdf",
        run_id=run_id,
        package_title=PACKAGE_TITLE,
        profile=profile,
        weekly=weekly,
        shortlist_result=shortlist_result,
        deep_dive=deep_dive,
        intake=intake,
        terminal_state="PENDING_BUILD",
    )
    pdf_meta = {
        "run_id": run_id,
        "shortlist_count": shortlist_result.get("shortlist_count"),
        "go_count": shortlist_result.get("go_count"),
        "review_count": shortlist_result.get("review_count"),
        "candidates_total": shortlist_result.get("candidates_total"),
        "cut_date": weekly.cut_date,
        "profile_version": profile.get("version"),
        "collection_id": weekly.collection_id,
        "cycle_run_id": weekly.cycle_id,
        "page_estimate": pdf_pages,
    }

    sh_rows: list[dict[str, str]] = []
    if weekly.source_health_path:
        sh_rows = load_csv_rows(Path(weekly.source_health_path))
    org_rows: list[dict[str, str]] = []
    if weekly.orgaos_path:
        org_rows = load_csv_rows(Path(weekly.orgaos_path))

    excel_meta = write_excel(
        delivery_out / "02-oportunidades-priorizadas.xlsx",
        run_id=run_id,
        profile=profile,
        weekly=weekly,
        shortlist_result=shortlist_result,
        intake=intake,
        ledger=ledger,
        source_health_rows=sh_rows,
        orgaos_rows=org_rows,
    )

    # ledger files
    ledger_json_path = delivery_out / "03-decision-ledger.json"
    ledger_json_path.write_text(
        json.dumps(
            {
                "schema": "extra-decision-ledger/1.0",
                "run_id": run_id,
                "items": ledger,
            },
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    with (delivery_out / "03-decision-ledger.csv").open("w", newline="", encoding="utf-8") as f:
        if ledger:
            w = csv.DictWriter(f, fieldnames=list(ledger[0].keys()))
            w.writeheader()
            w.writerows(ledger)
        else:
            w = csv.DictWriter(
                f,
                fieldnames=[
                    "opportunity_id",
                    "recommendation",
                    "recommendation_reason",
                    "evidence_status",
                    "missing_client_inputs",
                    "hard_blocks",
                    "next_action",
                    "owner",
                    "decision_deadline",
                    "client_decision",
                    "client_notes",
                    "tiago_review",
                    "reviewed_at",
                ],
            )
            w.writeheader()

    (delivery_out / "04-intake-operacional-extra.json").write_text(
        json.dumps(intake, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    (delivery_out / "04-intake-operacional-extra.md").write_text(
        intake_to_markdown(intake), encoding="utf-8"
    )
    (delivery_out / "profile-patch-candidate.yaml").write_text(
        profile_patch_yaml(intake), encoding="utf-8"
    )

    extra_lim = []
    if weekly.exit_code not in (0, None):
        extra_lim.append(
            f"Weekly de origem com exit_code={weekly.exit_code} (parcial). "
            "Contratos/concorrentes podem estar indisponíveis."
        )
    if shortlist_result.get("insufficient"):
        extra_lim.append("Shortlist abaixo de 5 itens defensáveis.")
    if dsn_note:
        extra_lim.append(dsn_note)
    write_limitations_md(
        delivery_out / "05-limitacoes-e-confiabilidade.md",
        weekly=weekly,
        shortlist_result=shortlist_result,
        extra_limitations=extra_lim,
    )
    write_meeting_script(
        delivery_out / "06-roteiro-reuniao.md",
        shortlist_result=shortlist_result,
        deep_dive=deep_dive,
        intake=intake,
    )

    # Dossiê: without official docs → NOT_AVAILABLE (do not invent corpus / PR133)
    dossie_status = "NOT_AVAILABLE"
    if deep_dive and deep_dive.get("url_oficial"):
        write_dossie_not_available(
            delivery_out / "07-dossie-edital-NOT_AVAILABLE.md",
            deep_dive=deep_dive,
            reason=(
                "Não há documentos oficiais (PDF/edital/anexos) baixados e hasheados para o caso "
                f"`{deep_dive.get('numero_controle')}`. A URL específica, quando presente, aponta "
                "para a página pública, mas o pipeline de ingestão documental (edital_case) não "
                "foi executado com corpus oficial nesta rodada. PR #133 (bid submission readiness) "
                "permanece fora de escopo."
            ),
            how_to_obtain=[
                f"Abrir a URL oficial: {deep_dive.get('url_oficial')}",
                "Baixar edital e anexos do portal PNCP/órgão.",
                "Executar: python -m scripts.edital_case run --help e ingest dos PDFs oficiais.",
                "Reexecutar este pacote com o case_dir populado quando os documentos existirem.",
            ],
        )
    else:
        write_dossie_not_available(
            delivery_out / "07-dossie-edital-NOT_AVAILABLE.md",
            deep_dive=deep_dive,
            reason=(
                "Nenhuma oportunidade defensável com prazo futuro e documentos oficiais "
                "suficientes para dossiê técnico nesta rodada."
            ),
            how_to_obtain=[
                "Atualizar coleta PNCP (extra-weekly) até obter prazos futuros.",
                "Selecionar oportunidade com URL específica e baixar edital/anexos.",
                "Rodar scripts.edital_case com documentos oficiais hasheados.",
            ],
        )
    if deep_dive:
        deep_dive["documents_available"] = False
        deep_dive["dossie_status"] = dossie_status

    # Reconcile
    recon = reconcile_package_counts(
        shortlist_result=shortlist_result,
        excel_meta=excel_meta,
        pdf_meta=pdf_meta,
        weekly=weekly,
        profile=profile,
    )
    if recon["status"] != "PASS":
        result = {
            "terminal_state": "FAILED_VALIDATION",
            "exit_code": 2,
            "errors": recon["divergences"],
            "run_id": run_id,
            "reconcile": recon,
        }
        (delivery_out / "FAILED_VALIDATION.json").write_text(
            json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
        print(json.dumps(result, indent=2, ensure_ascii=False))
        raise SystemExit(2)

    # Terminal state
    if shortlist_result.get("insufficient") or weekly.exit_code not in (0, None):
        terminal_state = "BUNDLE_READY_FOR_HUMAN_MERGE"
        # PARTIAL is allowed only when limitation is visible — already written
        package_quality = "PARTIAL_VISIBLE_LIMITATIONS"
    else:
        terminal_state = "BUNDLE_READY_FOR_HUMAN_MERGE"
        package_quality = "COMPLETE_PENDING_HUMAN"

    # Checksums of delivery artifacts
    artifact_names = [
        "00-LEIA-ME.md",
        "01-resumo-executivo.pdf",
        "01-resumo-executivo.md",
        "02-oportunidades-priorizadas.xlsx",
        "03-decision-ledger.csv",
        "03-decision-ledger.json",
        "04-intake-operacional-extra.md",
        "04-intake-operacional-extra.json",
        "05-limitacoes-e-confiabilidade.md",
        "06-roteiro-reuniao.md",
        "07-dossie-edital-NOT_AVAILABLE.md",
        "profile-patch-candidate.yaml",
    ]
    checksums: dict[str, str] = {}
    for name in artifact_names:
        p = delivery_out / name
        if p.is_file():
            checksums[name] = sha256_file(p)

    human_review = build_human_review(run_id, checksums)
    assert_not_auto_accepted(human_review)
    (delivery_out / "human-review.json").write_text(
        json.dumps(human_review, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    checksums["human-review.json"] = sha256_file(delivery_out / "human-review.json")

    checksums_doc = {
        "schema": "extra-first-client-checksums/1.0",
        "run_id": run_id,
        "generated_at": utc_now(),
        "artifacts": {
            name: {"path": name, "sha256": h, "bytes": (delivery_out / name).stat().st_size}
            for name, h in checksums.items()
        },
    }
    (delivery_out / "checksums.json").write_text(
        json.dumps(checksums_doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    # Sanitize shortlist for manifest (no private dump)
    manifest = {
        "schema": SCHEMA,
        "package_title": PACKAGE_TITLE,
        "run_id": run_id,
        "generated_at": utc_now(),
        "terminal_state": terminal_state,
        "package_quality": package_quality,
        "baseline_note": "verified_code_sha is git commit of tooling; package is external",
        "source_weekly": {
            "cycle_id": weekly.cycle_id,
            "collection_id": weekly.collection_id,
            "cut_date": weekly.cut_date,
            "exit_code": weekly.exit_code,
            "weekly_dir_basename": Path(weekly.weekly_dir).name,
            "product_checksums_ok": weekly.product_checksums_ok,
        },
        "profile": {
            "profile_id": profile.get("profile_id"),
            "version": profile.get("version"),
            "version_date": profile.get("version_date"),
            "sha256": profile.get("_sha256"),
            "critical_pending": shortlist_result.get("critical_pending"),
            "go_blocked": shortlist_result.get("go_blocked_by_profile"),
        },
        "counts": {
            "candidates_total": shortlist_result.get("candidates_total"),
            "blocked_total": shortlist_result.get("blocked_total"),
            "review_defensible_total": shortlist_result.get("review_defensible_total"),
            "shortlist_count": shortlist_result.get("shortlist_count"),
            "review_count": shortlist_result.get("review_count"),
            "no_go_count": shortlist_result.get("no_go_count"),
            "go_count": shortlist_result.get("go_count"),
            "insufficient": shortlist_result.get("insufficient"),
        },
        "deep_dive": deep_dive,
        "dossie_status": dossie_status,
        "reconcile": recon,
        "pdf_pages_estimate": pdf_pages,
        "human_review_status": "PENDING_HUMAN",
        "artifacts": list(checksums.keys()) + ["manifest.json", "checksums.json"],
        "claims_allowed": [
            "Shortlist e contagens derivadas do weekly pack validado",
            "Intake derivado de campos PENDING do perfil",
            "GO bloqueado por política de elicitation",
        ],
        "claims_forbidden": [
            "Probabilidade de vitória",
            "Capacidade financeira/técnica inventada da Extra",
            "Zero de mercado a partir de fonte falha ou arquivo vazio",
            "Aceite humano automático",
            "GO com PENDING críticos",
        ],
        "as_of": as_of.isoformat(),
    }
    (delivery_out / "manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    # Rewrite LEIA-ME / executive with final terminal state (content-level)
    write_readme(
        delivery_out / "00-LEIA-ME.md",
        run_id=run_id,
        package_title=PACKAGE_TITLE,
        shortlist_result=shortlist_result,
        weekly=weekly,
        terminal_state=terminal_state,
    )
    write_executive_md(
        delivery_out / "01-resumo-executivo.md",
        run_id=run_id,
        package_title=PACKAGE_TITLE,
        profile=profile,
        weekly=weekly,
        shortlist_result=shortlist_result,
        deep_dive=deep_dive,
        intake=intake,
        terminal_state=terminal_state,
    )
    # refresh checksums for rewritten md files
    for name in ("00-LEIA-ME.md", "01-resumo-executivo.md"):
        checksums[name] = sha256_file(delivery_out / name)
    checksums_doc["artifacts"] = {
        name: {"path": name, "sha256": h, "bytes": (delivery_out / name).stat().st_size}
        for name, h in checksums.items()
    }
    (delivery_out / "checksums.json").write_text(
        json.dumps(checksums_doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    # shortlist sidecar for operators (not required but useful; keep outside private data)
    (delivery_out / "shortlist.json").write_text(
        json.dumps(
            {
                "run_id": run_id,
                "shortlist": shortlist_result.get("shortlist"),
                "insufficient": shortlist_result.get("insufficient"),
                "insufficiency_reason": shortlist_result.get("insufficiency_reason"),
                "context_recent_no_go": shortlist_result.get("context_recent_no_go"),
            },
            indent=2,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )

    result = {
        "terminal_state": terminal_state,
        "package_quality": package_quality,
        "exit_code": 0,
        "run_id": run_id,
        "delivery_out": str(delivery_out),
        "source_weekly_run_id": weekly.cycle_id,
        "collection_id": weekly.collection_id,
        "cut_date": weekly.cut_date,
        "counts": manifest["counts"],
        "deep_dive": deep_dive,
        "dossie_status": dossie_status,
        "human_review": "PENDING_HUMAN",
        "reconcile": recon["status"],
        "pdf_pages_estimate": pdf_pages,
    }
    print(json.dumps(result, indent=2, ensure_ascii=False))
    print(f"DELIVERY_OUT={delivery_out}")
    print(f"RUN_ID={run_id}")
    print(f"TERMINAL_STATE={terminal_state}")
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Extra Construtora — Primeira Rodada de Decisão B2G (composição fina)"
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    run_p = sub.add_parser("run", help="Gera o pacote de decisão a partir de WEEKLY_INPUT real")
    run_p.add_argument("--weekly-input", required=True, type=Path, help="Dir do weekly pack real")
    run_p.add_argument(
        "--delivery-out",
        required=True,
        type=Path,
        help="Diretório externo de saída (fora do Git se for entrega cliente)",
    )
    run_p.add_argument(
        "--profile",
        type=Path,
        default=DEFAULT_PROFILE,
        help="Caminho do perfil extra.yaml",
    )
    run_p.add_argument(
        "--as-of",
        type=str,
        default=None,
        help="Data de referência YYYY-MM-DD (default: hoje)",
    )
    run_p.add_argument(
        "--client-ready-dsn",
        default=None,
        help="DSN isolado opcional (nunca produção)",
    )

    val_p = sub.add_parser("validate-weekly", help="Valida manifest/checksums do weekly pack")
    val_p.add_argument("--weekly-input", required=True, type=Path)

    args = parser.parse_args(argv)

    if args.cmd == "validate-weekly":
        v = validate_weekly_pack(args.weekly_input)
        print(json.dumps(asdict(v), indent=2, ensure_ascii=False))
        return 0 if v.ok else 2

    if args.cmd == "run":
        as_of = date.fromisoformat(args.as_of) if args.as_of else date.today()
        try:
            run_delivery(
                weekly_input=args.weekly_input,
                delivery_out=args.delivery_out,
                profile_path=args.profile,
                as_of=as_of,
                client_ready_dsn=args.client_ready_dsn,
            )
        except SystemExit as e:
            return int(e.code or 1)
        return 0

    return 1


if __name__ == "__main__":
    raise SystemExit(main())
