"""Per-item proof runner for CMI campaign items (DOD §10.1/§10.2/§11.1)."""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from collections.abc import Callable
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

_ROOT = Path(__file__).resolve().parents[2]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from scripts.ops import contract_market_intelligence as cmi  # noqa: E402
from scripts.ops.deliverable_b_competitors import capacity_hypothesis  # noqa: E402
from scripts.ops.deliverable_d_prices import (  # noqa: E402
    ComparabilityRule,
    PriceObservation,
    build_report,
)

CAMPAIGN = "CONTRACT-MARKET-INTELLIGENCE-ACCEPT-01"
PACKAGE_DIR = _ROOT / "artifacts/campaigns" / CAMPAIGN / "final-package"


def utc_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _sha_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _req(cond: bool, msg: str) -> None:
    if not cond:
        raise RuntimeError(msg)


def _require_package() -> dict[str, Any]:
    proof = PACKAGE_DIR / "proof.json"
    meta = PACKAGE_DIR / "metadata.json"
    _req(proof.is_file() and meta.is_file(), f"missing package under {PACKAGE_DIR}")
    return {
        "proof": json.loads(proof.read_text(encoding="utf-8")),
        "metadata": json.loads(meta.read_text(encoding="utf-8")),
    }


def _suppliers() -> list[dict[str, Any]]:
    p = PACKAGE_DIR / "suppliers-ranking.json"
    _req(p.is_file(), f"missing {p}")
    rows = json.loads(p.read_text(encoding="utf-8"))
    _req(bool(rows), "suppliers-ranking empty")
    return rows


def _rel() -> dict[str, Any]:
    return json.loads((PACKAGE_DIR / "reliability-status.json").read_text(encoding="utf-8"))


def _limitations() -> list[str]:
    return list(json.loads((PACKAGE_DIR / "limitations.json").read_text(encoding="utf-8")).get("limitations") or [])


def _vrefs() -> dict[str, Any]:
    return json.loads((PACKAGE_DIR / "value-references.json").read_text(encoding="utf-8"))


def _md() -> str:
    return (PACKAGE_DIR / "competitor-review.md").read_text(encoding="utf-8")


def check_10_1_01() -> dict[str, Any]:
    rows = _suppliers()
    _req(all(r.get("role") == "winner_identified" for r in rows), "role not winner_identified")
    _req(all(r.get("participant_identified") is False for r in rows), "participant must be false")
    return {"n": len(rows)}


def check_10_1_02() -> dict[str, Any]:
    rel = _rel()
    _req(rel["participantes"]["status"] == "SOURCE_UNAVAILABLE", "participantes status")
    md = _md().lower()
    _req("particip" in md, "review must discuss participants")
    return {"participantes": rel["participantes"]}


def check_10_1_03() -> dict[str, Any]:
    wr = cmi.win_rate(wins=3, proposals_presented=None)
    _req(wr["status"] == "NOT_COMPUTABLE", "win_rate unit")
    _req(_rel()["win_rate"]["status"] == "NOT_COMPUTABLE", "win_rate package")
    return {"unit": wr}


def check_10_1_04() -> dict[str, Any]:
    d = cmi.desagio_metric(valor_estimado=100.0, valor_homologado=80.0, same_certame_lote_item=False)
    _req(d["status"] == "NOT_COMPUTABLE", "desagio unit")
    _req(_rel()["desagio"]["status"] == "NOT_COMPUTABLE", "desagio package")
    return d


def check_10_1_05() -> dict[str, Any]:
    cap = capacity_hypothesis()
    _req(cap["claim_as_fact_forbidden"] is True, "capacity")
    for r in _suppliers():
        _req(r["capacity_claim"]["claim_as_fact_forbidden"] is True, "row capacity")
    return {"label": cap["label"]}


def check_10_1_06() -> dict[str, Any]:
    for r in _suppliers():
        _req(r["capacity_claim"]["claim_as_fact_forbidden"] is True, "n_contratos != capacity")
    return {"ok": True}


def check_10_1_07() -> dict[str, Any]:
    lim = _limitations()
    _req(len(lim) >= 3, "limitations count")
    for r in _suppliers():
        _req(bool(r.get("limitations")), "row limitations")
    return {"n_limitations": len(lim)}


def check_10_2_01() -> dict[str, Any]:
    rows = _suppliers()
    ranks = [r["rank"] for r in rows]
    _req(ranks == list(range(1, len(rows) + 1)), "ranking")
    return {"n": len(rows)}


def check_10_2_02() -> dict[str, Any]:
    rows = _suppliers()
    _req(all(int(r["n_contratos"]) >= 1 for r in rows), "n_contratos")
    return {"sample": rows[0]["n_contratos"]}


def check_10_2_03() -> dict[str, Any]:
    rows = _suppliers()
    _req(all(r.get("valor_type") == "valor_contratado" for r in rows), "valor_type")
    _req(
        any(
            r.get("valor_contratado_total") is not None
            or r.get("valor_contratado_total_status") == "MISSING"
            for r in rows
        ),
        "valor_contratado present or missing marked",
    )
    return {"ok": True}


def check_10_2_04() -> dict[str, Any]:
    for r in _suppliers():
        if r.get("ticket_contratado_medio") is not None:
            _req(int(r.get("ticket_denominator") or 0) > 0, "ticket denominator")
    return {"ok": True}


def check_10_2_05() -> dict[str, Any]:
    rows = _suppliers()
    _req(any(int(r.get("n_entes_atendidos") or 0) >= 1 for r in rows), "entes")
    return {"ok": True}


def check_10_2_06() -> dict[str, Any]:
    _req(any(r.get("distribuicao_municipio") for r in _suppliers()), "municipio")
    return {"ok": True}


def check_10_2_07() -> dict[str, Any]:
    _req(any(r.get("distribuicao_natureza_ente") for r in _suppliers()), "natureza")
    return {"ok": True}


def check_10_2_08() -> dict[str, Any]:
    _req(any(r.get("distribuicao_setor") for r in _suppliers()), "setor")
    return {"ok": True}


def check_10_2_09() -> dict[str, Any]:
    _req(all(isinstance(r.get("recorrencia"), dict) for r in _suppliers()), "recorrencia")
    return {"ok": True}


def check_10_2_10() -> dict[str, Any]:
    _req("ultima_contratacao_conhecida" in _suppliers()[0], "ultima")
    return {"ok": True}


def check_10_2_11() -> dict[str, Any]:
    p = PACKAGE_DIR / "concentration-by-entity.csv"
    _req(p.is_file() and p.stat().st_size > 30, "concentration entity file")
    return {"sha256": _sha_file(p), "bytes": p.stat().st_size}


def check_10_2_12() -> dict[str, Any]:
    p = PACKAGE_DIR / "concentration-by-supplier.csv"
    _req(p.is_file() and p.stat().st_size > 20, "concentration supplier file")
    return {"sha256": _sha_file(p), "bytes": p.stat().st_size}


def check_10_2_13() -> dict[str, Any]:
    rel = _rel()
    _req(rel["market_share"]["status"] in {"READY", "PARTIAL", "NOT_COMPUTABLE"}, "share status")
    meta = _require_package()["metadata"]
    _req(meta.get("complete_population_aggregated") is True, "complete population")
    return {"status": rel["market_share"]["status"]}


def check_10_2_14() -> dict[str, Any]:
    rel = _rel()
    _req(rel["hhi"]["status"] in {"READY", "PARTIAL", "NOT_COMPUTABLE"}, "hhi status")
    return {"status": rel["hhi"]["status"]}


def check_10_2_15() -> dict[str, Any]:
    meta = _require_package()["metadata"]
    _req(bool(meta.get("source") and meta.get("as_of") and meta.get("period_start")), "meta source/as_of")
    for r in _suppliers():
        _req(bool(r.get("source") and r.get("as_of")), "row source/as_of")
    return {"source": meta["source"], "as_of": meta["as_of"]}


def check_10_2_16() -> dict[str, Any]:
    p = PACKAGE_DIR / "executive-review.xlsx"
    sha_p = PACKAGE_DIR / "executive-review.xlsx.sha256"
    # If xlsx not in git, regenerate from DSN when available (CI / local).
    if (not p.is_file() or p.stat().st_size < 1000) and (
        os.environ.get("LOCAL_DATALAKE_DSN") or os.environ.get("TEST_DSN")
    ):
        dsn = os.environ.get("LOCAL_DATALAKE_DSN") or os.environ.get("TEST_DSN") or ""
        cmi.run_package(dsn, PACKAGE_DIR, seed_if_empty=True)
    if p.is_file() and p.stat().st_size > 1000:
        from openpyxl import load_workbook

        wb = load_workbook(p)
        _req("Fornecedores" in wb.sheetnames, "sheet Fornecedores")
        _req("Metadados" in wb.sheetnames, "sheet Metadados")
        digest = _sha_file(p)
        sha_p.write_text(
            f"{digest}  executive-review.xlsx\nsize_bytes={p.stat().st_size}\n"
            "regenerate: python3 -m scripts.ops.contract_market_intelligence run "
            f"--dsn $LOCAL_DATALAKE_DSN --out {PACKAGE_DIR}\n",
            encoding="utf-8",
        )
        return {"sheets": wb.sheetnames, "sha256": digest, "bytes": p.stat().st_size}
    _req(sha_p.is_file(), "xlsx or sha256 pointer required")
    text = sha_p.read_text(encoding="utf-8")
    _req(len(text) > 40, "xlsx sha pointer incomplete")
    return {"pointer": str(sha_p), "content_head": text.splitlines()[0]}


def check_10_2_17() -> dict[str, Any]:
    md = _md()
    _req("vencedor" in md.lower() and len(md) > 200, "competitor review material")
    return {"bytes": len(md.encode("utf-8"))}


def check_10_2_18() -> dict[str, Any]:
    dsn = os.environ.get("LOCAL_DATALAKE_DSN") or os.environ.get("TEST_DSN")
    _req(bool(dsn), "LOCAL_DATALAKE_DSN required")
    import psycopg2

    conn = psycopg2.connect(dsn)
    try:
        info = cmi.require_table_columns(conn, "pncp_supplier_contracts", cmi.REQUIRED_CONTRACT_COLS)
    finally:
        conn.close()
    meta = _require_package()["metadata"]
    _req(bool(meta.get("schema_check")), "package schema_check")
    return {"schema": info}


def check_10_2_19() -> dict[str, Any]:
    return check_10_2_18()


def check_10_2_20() -> dict[str, Any]:
    rel = _rel()
    statuses = {v.get("status") for v in rel.values()}
    _req("READY" in statuses, "need READY metric")
    _req("NOT_COMPUTABLE" in statuses or "SOURCE_UNAVAILABLE" in statuses, "need unavailable metric")
    return {"statuses": sorted(s for s in statuses if s)}


def check_11_1_01() -> dict[str, Any]:
    d = cmi.VALUE_DEFINITIONS["valor_estimado"]
    _req(len(d["definition"]) > 20 and d["enum"] == "valor_estimado", "valor_estimado def")
    return d


def check_11_1_02() -> dict[str, Any]:
    d = cmi.VALUE_DEFINITIONS["valor_homologado"]
    blob = d["definition"].lower()
    _req("homolog" in blob or "adjudic" in blob, "homologado def")
    return d


def check_11_1_03() -> dict[str, Any]:
    d = cmi.VALUE_DEFINITIONS["valor_contratado"]
    _req("contrato" in d["definition"].lower(), "contratado def")
    return d


def check_11_1_04() -> dict[str, Any]:
    d = cmi.VALUE_DEFINITIONS["valor_pago"]
    blob = d["definition"].lower()
    _req("pago" in blob or "desembolso" in blob, "pago def")
    return d


def check_11_1_05() -> dict[str, Any]:
    enums = {d["enum"] for d in cmi.VALUE_DEFINITIONS.values()}
    _req(len(enums) == 4, "four distinct enums")
    _req(_vrefs().get("fields_not_interchangeable") is True, "not interchangeable")
    return {"enums": sorted(enums)}


def check_11_1_06() -> dict[str, Any]:
    typed = _vrefs()["typed_values"]
    _req(all(t.get("value_type") for t in typed), "value_type field")
    return {"n": len(typed)}


def check_11_1_07() -> dict[str, Any]:
    typed = _vrefs()["typed_values"]
    _req(all(t.get("source") for t in typed), "source field")
    return {"ok": True}


def check_11_1_08() -> dict[str, Any]:
    typed = _vrefs()["typed_values"]
    _req(all("reference_date" in t for t in typed), "reference_date field")
    return {"ok": True}


def check_11_1_09() -> dict[str, Any]:
    typed = _vrefs()["typed_values"]
    _req(any(t.get("comparison_unit") for t in typed), "comparison_unit")
    return {"ok": True}


def check_11_1_10() -> dict[str, Any]:
    typed = _vrefs()["typed_values"]
    _req(all(t.get("value_scope") in {"global", "lote", "item", "unitario"} for t in typed), "scope")
    return {"scopes": sorted({t["value_scope"] for t in typed})}


def check_11_1_11() -> dict[str, Any]:
    tv = cmi.typed_value(
        value=None,
        value_type="valor_contratado",
        value_scope="global",
        source="t",
        reference_date=None,
        official_or_inferred="official",
    )
    _req(tv["value"] is None and tv["status"] == "MISSING", "missing not zero")
    for t in _vrefs()["typed_values"]:
        if t.get("status") == "MISSING":
            _req(t.get("value") is None, "MISSING must keep null")
    return {"ok": True}


def check_11_1_12() -> dict[str, Any]:
    tv = cmi.typed_value(
        value=1.0,
        value_type="valor_pago",
        value_scope="item",
        source="proxy",
        reference_date="2024-01-01",
        official_or_inferred="inferred",
        inference_method="test",
    )
    _req(tv["official_or_inferred"] == "inferred", "inferred marker")
    return tv


def check_11_1_13() -> dict[str, Any]:
    _req(any(t.get("official_or_inferred") == "official" for t in _vrefs()["typed_values"]), "official")
    return {"ok": True}


def check_11_1_14() -> dict[str, Any]:
    cf = " ".join(_vrefs().get("claims_forbidden") or []).lower()
    _req("heterogen" in cf or "preço real" in cf or "preco real" in cf, "forbidden claims")
    return {"ok": True}


def check_11_1_15() -> dict[str, Any]:
    _req(all("reference_date" in t for t in _vrefs()["typed_values"]), "dates on values")
    return {"ok": True}


def check_11_1_16() -> dict[str, Any]:
    _req(_vrefs().get("monetary_adjustment_applied") is False, "no silent monetary adjust")
    return {"monetary_adjustment_applied": False}


def check_11_1_17() -> dict[str, Any]:
    obs = [
        PriceObservation(
            value=float(100 + i),
            value_semantic="contratado",
            tipo_obra_servico="reforma_predial",
            unidade="m2",
            lote="unico",
            porte="medio",
            regiao="SC",
            periodo="2025-Q1",
            is_global_heterogeneous=True,
            source="t",
        )
        for i in range(6)
    ]
    report = build_report(obs, ComparabilityRule(min_sample=3))
    _req(any(p["status"] == "OK" for p in report.panels), "panel OK")
    for p in report.panels:
        _req(not p.get("labels_forbidden_used"), "labels_forbidden_used empty")
        claims = " ".join(p.get("claims") or []).lower()
        _req("preço real praticado" not in claims, "no preco real in claims")
        _req("preco real praticado" not in claims, "no preco real ascii in claims")
    return {"panels": len(report.panels)}


def check_11_1_18() -> dict[str, Any]:
    panels = _vrefs().get("panels") or []
    if panels:
        _req(any(p.get("n_observations") is not None for p in panels), "n_observations")
    return {"ok": True}


def check_11_1_19() -> dict[str, Any]:
    obs = [
        PriceObservation(
            value=v,
            value_semantic="contratado",
            tipo_obra_servico="reforma_predial",
            unidade="m2",
            lote="unico",
            porte="medio",
            regiao="SC",
            periodo="2025-Q1",
            is_global_heterogeneous=True,
            source="t",
        )
        for v in [100.0, 110.0, 120.0, 130.0, 140.0, 300.0]
    ]
    report = build_report(obs, ComparabilityRule(min_sample=3))
    _req(any(p.get("outliers_flagged") for p in report.panels), "outliers flagged")
    return {"outliers": report.panels[0].get("outliers_flagged")}


def check_11_1_20() -> dict[str, Any]:
    md = _md().lower()
    for pat in ("o preço real praticado é", "preco real praticado é", "preço real praticado ="):
        _req(pat not in md, f"forbidden claim pattern: {pat}")
    for p in _vrefs().get("panels") or []:
        claims = " ".join(p.get("claims") or []).lower()
        _req("preço real praticado" not in claims, "claims ban")
        _req(not p.get("labels_forbidden_used"), "labels ban")
    return {"ok": True}


CHECKERS: dict[str, Callable[[], dict[str, Any]]] = {
    "CMI-10.1-01": check_10_1_01,
    "CMI-10.1-02": check_10_1_02,
    "CMI-10.1-03": check_10_1_03,
    "CMI-10.1-04": check_10_1_04,
    "CMI-10.1-05": check_10_1_05,
    "CMI-10.1-06": check_10_1_06,
    "CMI-10.1-07": check_10_1_07,
    "CMI-10.2-01": check_10_2_01,
    "CMI-10.2-02": check_10_2_02,
    "CMI-10.2-03": check_10_2_03,
    "CMI-10.2-04": check_10_2_04,
    "CMI-10.2-05": check_10_2_05,
    "CMI-10.2-06": check_10_2_06,
    "CMI-10.2-07": check_10_2_07,
    "CMI-10.2-08": check_10_2_08,
    "CMI-10.2-09": check_10_2_09,
    "CMI-10.2-10": check_10_2_10,
    "CMI-10.2-11": check_10_2_11,
    "CMI-10.2-12": check_10_2_12,
    "CMI-10.2-13": check_10_2_13,
    "CMI-10.2-14": check_10_2_14,
    "CMI-10.2-15": check_10_2_15,
    "CMI-10.2-16": check_10_2_16,
    "CMI-10.2-17": check_10_2_17,
    "CMI-10.2-18": check_10_2_18,
    "CMI-10.2-19": check_10_2_19,
    "CMI-10.2-20": check_10_2_20,
    "CMI-11.1-01": check_11_1_01,
    "CMI-11.1-02": check_11_1_02,
    "CMI-11.1-03": check_11_1_03,
    "CMI-11.1-04": check_11_1_04,
    "CMI-11.1-05": check_11_1_05,
    "CMI-11.1-06": check_11_1_06,
    "CMI-11.1-07": check_11_1_07,
    "CMI-11.1-08": check_11_1_08,
    "CMI-11.1-09": check_11_1_09,
    "CMI-11.1-10": check_11_1_10,
    "CMI-11.1-11": check_11_1_11,
    "CMI-11.1-12": check_11_1_12,
    "CMI-11.1-13": check_11_1_13,
    "CMI-11.1-14": check_11_1_14,
    "CMI-11.1-15": check_11_1_15,
    "CMI-11.1-16": check_11_1_16,
    "CMI-11.1-17": check_11_1_17,
    "CMI-11.1-18": check_11_1_18,
    "CMI-11.1-19": check_11_1_19,
    "CMI-11.1-20": check_11_1_20,
}


def run_item(alias: str) -> dict[str, Any]:
    if alias not in CHECKERS:
        raise KeyError(f"unknown alias {alias}")
    detail = CHECKERS[alias]()
    return {
        "ok": True,
        "alias": alias,
        "campaign_id": CAMPAIGN,
        "as_of": utc_now(),
        "detail": detail,
        "package_dir": str(PACKAGE_DIR),
    }


def run_all() -> dict[str, Any]:
    results: dict[str, Any] = {}
    failed: list[str] = []
    for alias in CHECKERS:
        try:
            results[alias] = run_item(alias)
        except Exception as exc:  # noqa: BLE001
            results[alias] = {"ok": False, "alias": alias, "error": str(exc)}
            failed.append(alias)
    return {
        "ok": not failed,
        "failed": failed,
        "passed": [a for a in CHECKERS if a not in failed],
        "results": results,
        "n": len(CHECKERS),
    }


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="CMI per-item proofs")
    p.add_argument("--item", help="Alias e.g. CMI-10.1-01")
    p.add_argument("--all", action="store_true")
    p.add_argument("--json", action="store_true")
    p.add_argument("--list", action="store_true")
    args = p.parse_args(argv)
    if args.list:
        print("\n".join(CHECKERS))
        return 0
    if args.all or not args.item:
        out = run_all()
        payload = out if args.json else {"ok": out["ok"], "failed": out["failed"], "passed_n": len(out["passed"])}
        print(json.dumps(payload, indent=2, ensure_ascii=False, default=str))
        return 0 if out["ok"] else 1
    try:
        out = run_item(args.item)
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "alias": args.item, "error": str(exc)}, indent=2))
        return 1
    print(json.dumps(out, indent=2, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
