#!/usr/bin/env python3
"""Canonical decision pack entry point (EXTRA-DECISION-LOOP-01).

  make extra-decision-pack
  python -m scripts.ops.decision_pack --strict

Reuses weekly_cycle patterns (DSN, freshness, universe scope) and produces a
complete explainable decision package with profile hash, snapshot delta,
human review queue, PDF+Excel reconciliation, and checksums.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import traceback
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from scripts.crawl.run_evidence import get_git_meta, new_run_id, sha256_file  # noqa: E402
from scripts.opportunity_intel.decision_engine import decide_opportunity  # noqa: E402
from scripts.opportunity_intel.human_review import (  # noqa: E402
    calibrate,
    export_review_sample,
    import_review_labels,
    load_labels,
)
from scripts.opportunity_intel.profile_resolve import (  # noqa: E402
    resolve_extra_profile,
    write_profile_status,
)
from scripts.opportunity_intel.snapshot import (  # noqa: E402
    build_snapshot,
    compute_delta,
    load_latest_snapshot,
    pick_reconfirm_targets,
    reconfirm_batch,
    save_snapshot,
    select_active_opportunities,
)
from scripts.ops.weekly_cycle import (  # noqa: E402
    EXIT_OK,
    EXIT_TECH,
    EXIT_UNRELIABLE,
    _connect,
    _iso,
    _q,
    _resolve_dsn,
    _table_exists,
    stage_freshness,
)

COLLECTOR_VERSION = "decision-pack/1.0"
DEFAULT_OUT = _PROJECT_ROOT / "output" / "decision"


def _utc() -> datetime:
    return datetime.now(UTC)


def _atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + f".{os.getpid()}.tmp")
    tmp.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    os.replace(tmp, path)


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fields: list[str] = []
    seen: set[str] = set()
    for r in rows:
        for k in r:
            if k not in seen:
                seen.add(k)
                fields.append(k)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    k: (
                        json.dumps(r.get(k), ensure_ascii=False, default=str)
                        if isinstance(r.get(k), (dict, list))
                        else ("" if r.get(k) is None else str(r.get(k)))
                    )
                    for k in fields
                }
            )


def load_opportunities(conn: Any, *, limit: int = 500) -> list[dict[str, Any]]:
    if not _table_exists(conn, "opportunity_intel"):
        return []
    return _q(
        conn,
        """
        SELECT id, source, source_id, numero_controle_pncp, orgao_cnpj, orgao_nome,
               municipio, uf, objeto, modalidade, valor_estimado, valor_semantica,
               status_canonico, ranking, ranking_score, ranking_confianca,
               data_publicacao, data_abertura, data_encerramento,
               link_edital, run_id, crawl_batch_id, proveniencia,
               last_seen_source_run_id, ingested_at, updated_at, is_active
        FROM opportunity_intel
        WHERE COALESCE(is_active, TRUE)
          AND status_canonico IN ('open', 'upcoming')
          AND (
            uf = 'SC'
            OR LEFT(REGEXP_REPLACE(COALESCE(orgao_cnpj, ''), '[^0-9]', '', 'g'), 8)
               IN (
                 SELECT e.cnpj_8 FROM sc_public_entities e
                 WHERE e.is_active IS TRUE AND e.raio_200km IS TRUE
               )
          )
        ORDER BY
          CASE ranking WHEN 'GO' THEN 0 WHEN 'REVIEW' THEN 1 ELSE 2 END,
          ranking_score DESC NULLS LAST,
          data_encerramento NULLS LAST
        LIMIT %s
        """,
        (limit,),
    )


def _generate_pdf(path: Path, brief_md: str, meta: dict[str, Any]) -> tuple[bool, str]:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "TitleBR",
            parent=styles["Heading1"],
            fontSize=14,
            spaceAfter=8,
        )
        body = ParagraphStyle(
            "BodyBR",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
        )
        doc = SimpleDocTemplate(
            str(path),
            pagesize=A4,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )
        story: list[Any] = []
        story.append(Paragraph("Brief executivo de decisão — Extra Construtora", title_style))
        story.append(
            Paragraph(
                f"run_id={meta.get('run_id')} | cutoff={meta.get('cutoff')} | "
                f"profile_hash={meta.get('profile_hash') or ''}",
                body,
            )
        )
        story.append(Spacer(1, 6))
        # markdown-ish plain: escape and break paragraphs
        for block in brief_md.split("\n\n"):
            text = (
                block.replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace("\n", "<br/>")
            )
            if text.strip():
                story.append(Paragraph(text, body))
                story.append(Spacer(1, 4))
        story.append(Spacer(1, 8))
        story.append(
            Paragraph(
                "ALERTA: triagem automatizada — não garante vitória; não substitui "
                "análise jurídica/contábil/técnica final. Aceite humano: PENDING_HUMAN.",
                body,
            )
        )
        doc.build(story)
        return True, ""
    except Exception as exc:  # noqa: BLE001
        return False, str(exc)


def _read_pdf_text(pdf_path: Path) -> str:
    """Extract text from product PDF for reconciliation (real file read)."""
    if not pdf_path.is_file():
        return ""
    # Reject non-PDF placeholders (e.g. failed generation written as .txt content)
    head = pdf_path.read_bytes()[:8]
    if not head.startswith(b"%PDF"):
        return ""
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(str(pdf_path))
        parts: list[str] = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        return "\n".join(parts)
    except Exception:  # noqa: BLE001
        return ""


def _extract_meta_from_pdf_text(text: str) -> dict[str, str | None]:
    """Parse run_id / cutoff / profile_hash stamped into the brief PDF body."""
    import re

    run_id = None
    cutoff = None
    profile_hash = None
    m = re.search(r"run_id=([A-Za-z0-9._\-]+)", text)
    if m:
        run_id = m.group(1)
    m = re.search(r"cutoff=([0-9]{4}-[0-9]{2}-[0-9]{2})", text)
    if m:
        cutoff = m.group(1)
    m = re.search(r"profile_hash=([0-9a-fA-F]{8,})", text)
    if m:
        profile_hash = m.group(1)
    # full hash may be truncated in PDF with ellipsis — keep prefix for compare
    return {"run_id": run_id, "cutoff": cutoff, "profile_hash": profile_hash}


def _read_excel_product_meta(xlsx_path: Path) -> dict[str, Any]:
    """Read Metadados sheet + data-row counts from decision pack Excel."""
    out: dict[str, Any] = {
        "meta": {},
        "sheet_row_counts": {},
        "error": None,
    }
    if not xlsx_path.is_file():
        out["error"] = "excel_missing"
        return out
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
        meta: dict[str, str] = {}
        if "Metadados" in wb.sheetnames:
            ws = wb["Metadados"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                key = str(row[0]).strip()
                val = "" if row[1] is None else str(row[1]).strip()
                meta[key] = val
        out["meta"] = meta
        for sheet in ("PARTICIPAR", "REVIEW", "NAO_PARTICIPAR"):
            if sheet not in wb.sheetnames:
                out["sheet_row_counts"][sheet] = 0
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                out["sheet_row_counts"][sheet] = 0
                continue
            # header + optional "(vazio)" single cell
            data_rows = rows[1:]
            if len(data_rows) == 1 and data_rows[0] and str(data_rows[0][0]) == "(vazio)":
                out["sheet_row_counts"][sheet] = 0
            else:
                out["sheet_row_counts"][sheet] = len(data_rows)
        # Claims sheet optional
        if "Claims" in wb.sheetnames:
            ws = wb["Claims"]
            rows = list(ws.iter_rows(values_only=True))
            data_rows = rows[1:] if rows else []
            if len(data_rows) == 1 and data_rows[0] and str(data_rows[0][0]) == "(vazio)":
                out["sheet_row_counts"]["Claims"] = 0
            else:
                out["sheet_row_counts"]["Claims"] = len(data_rows)
        wb.close()
    except Exception as exc:  # noqa: BLE001
        out["error"] = str(exc)
    return out


def reconcile_pdf_excel(
    *,
    run_id: str,
    profile_hash: str,
    cutoff: str,
    pdf_path: Path,
    xlsx_path: Path,
    counts: dict[str, int],
    claims_count: int | None = None,
) -> dict[str, Any]:
    """Fail-closed product reconcile: read PDF text + Excel sheets (no mirror dicts)."""
    div: list[str] = []
    pdf_meta: dict[str, str | None] = {"run_id": None, "cutoff": None, "profile_hash": None}
    excel_info = _read_excel_product_meta(xlsx_path)

    if not pdf_path.is_file():
        div.append("pdf_missing")
    else:
        pdf_text = _read_pdf_text(pdf_path)
        if not pdf_text.strip():
            div.append("pdf_unreadable_or_not_pdf")
        else:
            pdf_meta = _extract_meta_from_pdf_text(pdf_text)
            if pdf_meta.get("run_id") != run_id:
                div.append(f"pdf_run_id_mismatch:{pdf_meta.get('run_id')}!={run_id}")
            if pdf_meta.get("cutoff") != cutoff:
                div.append(f"pdf_cutoff_mismatch:{pdf_meta.get('cutoff')}!={cutoff}")
            ph = pdf_meta.get("profile_hash") or ""
            # PDF may stamp truncated hash prefix
            if not profile_hash.startswith(ph) and ph not in profile_hash:
                div.append(f"pdf_profile_hash_mismatch:{ph}!={profile_hash[:16]}")

    if not xlsx_path.is_file():
        div.append("excel_missing")
    elif excel_info.get("error"):
        div.append(f"excel_read_error:{excel_info['error']}")
    else:
        emeta = excel_info.get("meta") or {}
        if str(emeta.get("run_id") or "") != run_id:
            div.append(f"excel_run_id_mismatch:{emeta.get('run_id')}!={run_id}")
        if str(emeta.get("profile_hash") or "") != profile_hash:
            div.append(
                f"excel_profile_hash_mismatch:{str(emeta.get('profile_hash') or '')[:16]}"
                f"!={profile_hash[:16]}"
            )
        if str(emeta.get("cutoff") or "") != cutoff:
            div.append(f"excel_cutoff_mismatch:{emeta.get('cutoff')}!={cutoff}")

        sheet_counts = excel_info.get("sheet_row_counts") or {}
        expected_map = {
            "PARTICIPAR": int(counts.get("participar") or 0),
            "REVIEW": int(counts.get("review") or 0),
            "NAO_PARTICIPAR": int(counts.get("nao_participar") or 0),
        }
        excel_counts_read: dict[str, int] = {
            "participar": int(sheet_counts.get("PARTICIPAR") or 0),
            "review": int(sheet_counts.get("REVIEW") or 0),
            "nao_participar": int(sheet_counts.get("NAO_PARTICIPAR") or 0),
        }
        excel_counts_read["total"] = (
            excel_counts_read["participar"]
            + excel_counts_read["review"]
            + excel_counts_read["nao_participar"]
        )
        for sheet, expected in expected_map.items():
            got = int(sheet_counts.get(sheet) or 0)
            if got != expected:
                div.append(f"excel_sheet_count_mismatch_{sheet}:{got}!={expected}")
        expected_total = int(counts.get("total") or 0)
        if excel_counts_read["total"] != expected_total:
            div.append(
                f"excel_total_mismatch:{excel_counts_read['total']}!={expected_total}"
            )
        if claims_count is not None and "Claims" in sheet_counts:
            got_claims = int(sheet_counts.get("Claims") or 0)
            if got_claims != int(claims_count):
                div.append(f"excel_claims_count_mismatch:{got_claims}!={claims_count}")

    same_run = (
        pdf_meta.get("run_id") == run_id
        and str((excel_info.get("meta") or {}).get("run_id") or "") == run_id
        and "pdf_run_id_mismatch" not in " ".join(div)
        and "excel_run_id_mismatch" not in " ".join(div)
    )
    status = "PASS" if not div else "FAIL"
    return {
        "status": status,
        "same_run_id": same_run,
        "run_id": run_id,
        "profile_hash": profile_hash,
        "cutoff": cutoff,
        "divergences": div,
        "pdf": str(pdf_path) if pdf_path.is_file() else None,
        "excel": str(xlsx_path) if xlsx_path.is_file() else None,
        "pdf_meta_read": pdf_meta,
        "excel_meta_read": excel_info.get("meta"),
        "excel_sheet_row_counts": excel_info.get("sheet_row_counts"),
        "counts_expected": counts,
    }


def build_decision_claims(
    decisions: list[dict[str, Any]],
    *,
    run_id: str,
    profile_hash: str,
    cutoff: str,
) -> list[dict[str, Any]]:
    """One reconcilable claim per decision — must match Excel Claims sheet rows."""
    claims: list[dict[str, Any]] = []
    for d in decisions:
        oid = d.get("opportunity_id") or d.get("id")
        rec = d.get("recommendation")
        claims.append(
            {
                "claim_id": f"dec-{oid}",
                "run_id": run_id,
                "profile_hash": profile_hash,
                "cutoff": cutoff,
                "opportunity_id": oid,
                "recommendation": rec,
                "internal_ranking": d.get("internal_ranking"),
                "confidence": d.get("confidence"),
                "orgao_nome": d.get("orgao_nome"),
                "statement": (
                    f"Oportunidade {oid} → {rec} "
                    f"(score={d.get('ranking_score')}, conf={d.get('confidence')}) "
                    f"órgão={d.get('orgao_nome')}"
                ),
                "hard_blockers": d.get("hard_blockers") or [],
                "missing_information": d.get("missing_information") or [],
            }
        )
    return claims


def reconcile_claims_to_excel(
    claims: list[dict[str, Any]],
    xlsx_path: Path,
) -> dict[str, Any]:
    """Assert every claim_id appears as a Claims sheet row (and counts match)."""
    div: list[str] = []
    if not xlsx_path.is_file():
        return {"status": "FAIL", "divergences": ["excel_missing"], "matched": 0}
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
        if "Claims" not in wb.sheetnames:
            wb.close()
            return {"status": "FAIL", "divergences": ["claims_sheet_missing"], "matched": 0}
        ws = wb["Claims"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return {"status": "FAIL", "divergences": ["claims_sheet_empty"], "matched": 0}
        headers = [str(h) if h is not None else "" for h in rows[0]]
        if "claim_id" not in headers:
            return {"status": "FAIL", "divergences": ["claims_header_missing_claim_id"], "matched": 0}
        idx = headers.index("claim_id")
        rec_idx = headers.index("recommendation") if "recommendation" in headers else None
        excel_ids: dict[str, str | None] = {}
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            if str(row[0]) == "(vazio)":
                continue
            cid = str(row[idx])
            rec = str(row[rec_idx]) if rec_idx is not None and rec_idx < len(row) else None
            excel_ids[cid] = rec
        claim_ids = {str(c.get("claim_id")) for c in claims}
        missing = sorted(claim_ids - set(excel_ids))
        extra = sorted(set(excel_ids) - claim_ids)
        if missing:
            div.append(f"claims_missing_in_excel:{len(missing)}")
        if extra:
            div.append(f"claims_extra_in_excel:{len(extra)}")
        if len(claims) != len(excel_ids):
            div.append(f"claims_count_mismatch:{len(excel_ids)}!={len(claims)}")
        # recommendation parity for matched ids
        mismatches = 0
        for c in claims:
            cid = str(c.get("claim_id"))
            if cid in excel_ids and excel_ids[cid] is not None:
                if str(c.get("recommendation")) != str(excel_ids[cid]):
                    mismatches += 1
        if mismatches:
            div.append(f"claims_recommendation_mismatch:{mismatches}")
        return {
            "status": "PASS" if not div else "FAIL",
            "divergences": div,
            "matched": len(claim_ids & set(excel_ids)),
            "n_claims": len(claims),
            "n_excel": len(excel_ids),
        }
    except Exception as exc:  # noqa: BLE001
        return {"status": "FAIL", "divergences": [f"claims_reconcile_error:{exc}"], "matched": 0}


def build_brief(
    *,
    run_id: str,
    decisions: list[dict[str, Any]],
    delta: dict[str, Any],
    profile: dict[str, Any],
    freshness: list[dict[str, Any]],
    gaps: list[str],
) -> str:
    part = [d for d in decisions if d.get("recommendation") == "PARTICIPAR"]
    rev = [d for d in decisions if d.get("recommendation") == "REVIEW"]
    no = [d for d in decisions if d.get("recommendation") == "NÃO_PARTICIPAR"]
    lines = [
        f"# Executive decision brief — {run_id}",
        "",
        f"- **Cutoff:** {profile.get('resolved_at', '')[:10]}",
        f"- **Profile:** {profile.get('profile_id')} v{profile.get('version')} "
        f"hash `{(profile.get('profile_hash') or '')[:16]}…`",
        f"- **Decisões:** PARTICIPAR={len(part)} | REVIEW={len(rev)} | NÃO_PARTICIPAR={len(no)}",
        "",
        "## O que fazer esta semana",
        "",
    ]
    if part:
        lines.append("1. Validar humanamente os PARTICIPAR abaixo (habilitação/margem).")
    else:
        lines.append("1. Nenhum PARTICIPAR automático — focar na fila REVIEW prioritária.")
    lines.append("2. Completar campos pendentes do perfil Extra (ver profile_status).")
    lines.append("3. Rotular amostra de revisão humana (extra-review-export).")
    lines += ["", "## PARTICIPAR", ""]
    if not part:
        lines.append("_Nenhuma — resultado legítimo se fundamentado._")
    for d in part[:20]:
        lines.append(
            f"- **{d.get('orgao_nome')}** — {(d.get('objeto') or '')[:100]} "
            f"(score={d.get('ranking_score')}, conf={d.get('confidence')})"
        )
    lines += ["", "## REVIEW (ação de análise)", ""]
    for d in rev[:25]:
        miss = ", ".join((d.get("missing_information") or [])[:4])
        lines.append(
            f"- **{d.get('orgao_nome')}** — {(d.get('objeto') or '')[:90]} "
            f"| missing: {miss or '—'}"
        )
    lines += ["", "## NÃO_PARTICIPAR (amostra de motivos)", ""]
    for d in no[:15]:
        blockers = ", ".join((d.get("hard_blockers") or d.get("risks") or [])[:3])
        lines.append(f"- {d.get('orgao_nome')}: {blockers or d.get('recommendation')}")
    lines += ["", "## Mudanças desde snapshot anterior", ""]
    counts = delta.get("counts") or {}
    lines.append(
        f"- novos={counts.get('new', 0)}, prazo_alterado={counts.get('deadline_changed', 0)}, "
        f"removidos={counts.get('removed', 0)}, ainda_abertos_reconfirmados="
        f"{counts.get('still_open_reconfirmed', 0)}"
    )
    lines += ["", "## Saúde das fontes", ""]
    for f in freshness[:12]:
        lines.append(f"- `{f.get('source')}`: {f.get('level')} (age_h={f.get('age_hours')})")
    lines += ["", "## Gaps / limitações", ""]
    for g in gaps:
        lines.append(f"- {g}")
    lines += [
        "",
        "## Aceite humano",
        "",
        "Status: **PENDING_HUMAN**",
        "",
        DISCLAIMER_SHORT,
        "",
    ]
    return "\n".join(lines)


DISCLAIMER_SHORT = (
    "Triagem automatizada — não é garantia de vitória; não substitui "
    "análise jurídica, contábil ou técnica final."
)


def run_decision_pack(
    *,
    dsn: str | None = None,
    out_dir: Path | None = None,
    limit: int = 500,
    strict: bool = False,
    offline_reconfirm: bool = False,
    skip_db: bool = False,
    reconfirm_top: int = 20,
    reconfirm_max: int = 40,
) -> tuple[int, dict[str, Any]]:
    run_id = new_run_id("decision")
    started = _iso()
    out = out_dir or (DEFAULT_OUT / run_id)
    out.mkdir(parents=True, exist_ok=True)
    git = get_git_meta()
    report: dict[str, Any] = {
        "schema": "extra-decision-pack/1.0",
        "run_id": run_id,
        "started_at": started,
        "collector_version": COLLECTOR_VERSION,
        "git": git,
        "strict": strict,
    }

    # --- Profile ---
    resolved = resolve_extra_profile()
    profile_meta = {
        "profile_id": resolved.profile_id,
        "version": resolved.version,
        "profile_hash": resolved.profile_hash,
        "resolved_at": resolved.resolved_at,
        "pending_critical": resolved.pending_critical,
    }
    write_profile_status(resolved, out)
    report["profile"] = resolved.to_public_dict()

    # --- Data ---
    rows: list[dict[str, Any]] = []
    freshness: list[dict[str, Any]] = []
    conn = None
    if not skip_db:
        try:
            conn = _connect(_resolve_dsn(dsn))
            rows = load_opportunities(conn, limit=limit)
            fr = stage_freshness(conn)
            freshness = list((fr.detail or {}).get("sources") or [])
            if not freshness:
                freshness.append(
                    {
                        "source": "opportunity_intel",
                        "level": "unknown",
                        "note": "no opportunity_runs freshness rows",
                    }
                )
        except Exception as exc:  # noqa: BLE001
            report["db_error"] = str(exc)
            if strict:
                report["status"] = "TECH_FAIL"
                _atomic_json(out / "decision_manifest.json", report)
                return EXIT_TECH, report

    active = select_active_opportunities(rows)
    # provisional actionable = high score REVIEW/GO candidates without hard terminal status
    provisional_actionable = [
        r
        for r in active
        if str(r.get("ranking")) in {"GO", "REVIEW"} and float(r.get("ranking_score") or 0) >= 55
    ]
    targets = pick_reconfirm_targets(
        active,
        actionable=provisional_actionable,
        top_n=reconfirm_top,
    )[:reconfirm_max]

    offline = offline_reconfirm or os.getenv("DECISION_RECONFIRM_OFFLINE") == "1"
    reconfirm_map = reconfirm_batch(
        targets,
        run_id=run_id,
        collection_id=run_id,
        offline=offline,
        sleep_s=0.05 if not offline else 0.0,
    )

    # Decisions
    decisions: list[dict[str, Any]] = []
    for r in active:
        oid = r.get("id") or r.get("source_id")
        rc = reconfirm_map.get(oid) or {"outcome": "not_attempted"}
        # only mark high-conf open when reconfirmed
        if r.get("status_canonico") in {"open", "upcoming"} and rc.get("outcome") != "ok":
            r = dict(r)
            r["freshness"] = rc.get("outcome") or "unconfirmed"
        dec = decide_opportunity(
            r,
            profile=resolved.data,
            profile_meta=profile_meta,
            reconfirm=rc,
        )
        d = dec.to_dict()
        d.update(
            {
                "opportunity_id": oid,
                "id": oid,
                "source_id": r.get("source_id"),
                "orgao_nome": r.get("orgao_nome"),
                "orgao_cnpj": r.get("orgao_cnpj"),
                "objeto": r.get("objeto"),
                "modalidade": r.get("modalidade"),
                "valor_estimado": r.get("valor_estimado"),
                "status_canonico": r.get("status_canonico"),
                "data_encerramento": str(r.get("data_encerramento") or "") or None,
                "link_edital": r.get("link_edital"),
                "source": r.get("source"),
                "uf": r.get("uf"),
                "municipio": r.get("municipio"),
            }
        )
        # flatten dimensions for CSV
        for name, dim in (d.get("dimensions") or {}).items():
            if isinstance(dim, dict):
                d[f"dim_{name}"] = dim.get("score")
        decisions.append(d)

    # Snapshot + delta
    snap = build_snapshot(
        active,
        run_id=run_id,
        collection_id=run_id,
        profile_hash=resolved.profile_hash,
        reconfirm_map=reconfirm_map,
    )
    prev = load_latest_snapshot()
    delta = compute_delta(prev, snap)
    snap_path = save_snapshot(snap)
    _atomic_json(out / "snapshot.json", snap.to_dict())
    _atomic_json(out / "snapshot_delta.json", delta)

    # Partition CSVs
    part_rows = [d for d in decisions if d["recommendation"] == "PARTICIPAR"]
    rev_rows = [d for d in decisions if d["recommendation"] == "REVIEW"]
    no_rows = [d for d in decisions if d["recommendation"] == "NÃO_PARTICIPAR"]

    _write_csv(out / "actionable_opportunities.csv", part_rows)
    _write_csv(out / "review_opportunities.csv", rev_rows)
    _write_csv(out / "discarded_opportunities.csv", no_rows)
    _write_csv(out / "all_decisions.csv", decisions)

    # snapshot_changes.csv
    change_rows = []
    for kind in (
        "new",
        "changed",
        "deadline_changed",
        "suspended",
        "revoked",
        "closed",
        "removed",
        "still_open_reconfirmed",
    ):
        for oid in delta.get(kind) or []:
            change_rows.append({"change_type": kind, "opportunity_id": oid})
    _write_csv(out / "snapshot_changes.csv", change_rows)

    # human review queue + export
    export_meta = export_review_sample(
        decisions,
        out / "human_review_queue.csv",
        target=min(40, max(len(decisions), 1)),
    )
    _write_csv(
        out / "source_health.csv",
        freshness
        or [
            {
                "source": "opportunity_intel",
                "level": "unknown",
                "note": "freshness detail unavailable",
            }
        ],
    )

    # calibrate (usually PENDING_HUMAN)
    cal = calibrate(load_labels())
    _atomic_json(out / "calibration.json", cal.to_dict())

    counts = {
        "participar": len(part_rows),
        "review": len(rev_rows),
        "nao_participar": len(no_rows),
        "total": len(decisions),
        "reconfirm_targets": len(targets),
        "reconfirm_ok": sum(1 for v in reconfirm_map.values() if v.get("outcome") == "ok"),
    }

    gaps = [
        "Aceite de produto permanece PENDING_HUMAN.",
        "PARTICIPAR exige reconfirmação oficial ok + perfil sem gaps materiais.",
        DISCLAIMER_SHORT,
    ]
    if resolved.pending_critical:
        gaps.append(
            "Perfil com campos críticos pendentes: " + ", ".join(resolved.pending_critical[:8])
        )
    if counts["participar"] == 0:
        gaps.append("Nenhum PARTICIPAR neste run — legítimo se fundamentado.")
    if offline:
        gaps.append("Reconfirmação em modo offline/fixture — não reivindicar HTTP live ok.")

    brief = build_brief(
        run_id=run_id,
        decisions=decisions,
        delta=delta,
        profile=profile_meta | {"resolved_at": resolved.resolved_at},
        freshness=freshness,
        gaps=gaps,
    )
    md_path = out / "executive_decision_brief.md"
    md_path.write_text(brief + "\n", encoding="utf-8")

    # Claims (reconcilable with Excel Claims sheet)
    cutoff = resolved.resolved_at[:10]
    claims = build_decision_claims(
        decisions,
        run_id=run_id,
        profile_hash=resolved.profile_hash,
        cutoff=cutoff,
    )
    _write_csv(out / "claims_provenance.csv", claims)

    # Excel
    xlsx_path = out / "extra_decision_pack.xlsx"
    excel_ok = False
    excel_note = ""
    try:
        from openpyxl import Workbook

        wb = Workbook()
        meta = wb.active
        meta.title = "Metadados"
        meta.append(["key", "value"])
        for k, v in [
            ("run_id", run_id),
            ("profile_id", resolved.profile_id),
            ("profile_version", resolved.version),
            ("profile_hash", resolved.profile_hash),
            ("cutoff", cutoff),
            ("git_sha", (git or {}).get("git_sha")),
            ("participar", counts["participar"]),
            ("review", counts["review"]),
            ("nao_participar", counts["nao_participar"]),
            ("total", counts["total"]),
            ("claims_count", len(claims)),
            ("human_accept", "PENDING_HUMAN"),
        ]:
            meta.append([k, str(v)])
        # Stable empty headers so reconcile can read sheets with 0 data rows
        empty_headers = {
            "PARTICIPAR": [
                "opportunity_id",
                "recommendation",
                "orgao_nome",
                "objeto",
                "ranking_score",
            ],
            "REVIEW": [
                "opportunity_id",
                "recommendation",
                "orgao_nome",
                "objeto",
                "ranking_score",
            ],
            "NAO_PARTICIPAR": [
                "opportunity_id",
                "recommendation",
                "orgao_nome",
                "objeto",
                "ranking_score",
            ],
            "Claims": [
                "claim_id",
                "run_id",
                "profile_hash",
                "cutoff",
                "opportunity_id",
                "recommendation",
                "internal_ranking",
                "confidence",
                "orgao_nome",
                "statement",
                "hard_blockers",
                "missing_information",
            ],
        }
        for name, rows_ in [
            ("PARTICIPAR", part_rows),
            ("REVIEW", rev_rows),
            ("NAO_PARTICIPAR", no_rows),
            ("Claims", claims),
            ("Delta", change_rows),
            ("SourceHealth", freshness),
            ("Reconfirm", list(reconfirm_map.values())),
        ]:
            ws = wb.create_sheet(name[:31])
            if not rows_:
                headers = empty_headers.get(name)
                if headers:
                    ws.append(headers)
                else:
                    ws.append(["(vazio)"])
                continue
            headers = list(rows_[0].keys())
            ws.append(headers)
            for r in rows_:
                ws.append(
                    [
                        json.dumps(r.get(h), ensure_ascii=False, default=str)
                        if isinstance(r.get(h), (dict, list))
                        else r.get(h)
                        for h in headers
                    ]
                )
        wb.save(xlsx_path)
        excel_ok = True
    except Exception as exc:  # noqa: BLE001
        excel_note = str(exc)
        xlsx_path.write_text(f"excel failed: {exc}\n", encoding="utf-8")

    pdf_path = out / "executive_decision_brief.pdf"
    pdf_ok, pdf_err = _generate_pdf(
        pdf_path,
        brief,
        {
            "run_id": run_id,
            "cutoff": cutoff,
            "profile_hash": resolved.profile_hash,
        },
    )
    if not pdf_ok:
        pdf_path.write_text(f"pdf failed: {pdf_err}\n{brief}", encoding="utf-8")

    reconcile = reconcile_pdf_excel(
        run_id=run_id,
        profile_hash=resolved.profile_hash,
        cutoff=cutoff,
        pdf_path=pdf_path,
        xlsx_path=xlsx_path,
        counts=counts,
        claims_count=len(claims),
    )
    claims_reconcile = reconcile_claims_to_excel(claims, xlsx_path)
    if claims_reconcile.get("status") != "PASS":
        reconcile["status"] = "FAIL"
        reconcile.setdefault("divergences", []).extend(
            claims_reconcile.get("divergences") or ["claims_reconcile_fail"]
        )
    reconcile["claims_reconcile"] = claims_reconcile
    if not pdf_ok or not excel_ok:
        reconcile["status"] = "FAIL"
        if not pdf_ok:
            reconcile["divergences"].append(f"pdf_error:{pdf_err}")
        if not excel_ok:
            reconcile["divergences"].append(f"excel_error:{excel_note}")
    _atomic_json(out / "reconcile.json", reconcile)

    # checksums (products only)
    artifacts = {
        "brief_md": md_path,
        "brief_pdf": pdf_path,
        "excel": xlsx_path,
        "actionable_csv": out / "actionable_opportunities.csv",
        "review_csv": out / "review_opportunities.csv",
        "discarded_csv": out / "discarded_opportunities.csv",
        "claims_csv": out / "claims_provenance.csv",
        "snapshot_changes_csv": out / "snapshot_changes.csv",
        "human_review_queue_csv": out / "human_review_queue.csv",
        "source_health_csv": out / "source_health.csv",
        "profile_status_json": out / "profile_status.json",
        "snapshot_json": out / "snapshot.json",
        "delta_json": out / "snapshot_delta.json",
    }
    checksums: dict[str, Any] = {}
    for label, pth in artifacts.items():
        if pth.is_file():
            checksums[label] = {
                "path": str(pth),
                "sha256": sha256_file(pth),
                "bytes": pth.stat().st_size,
            }
    checksums_path = out / "checksums.json"
    _atomic_json(
        checksums_path,
        {
            "schema": "extra-decision-checksums/1.0",
            "run_id": run_id,
            "profile_hash": resolved.profile_hash,
            "artifacts": checksums,
        },
    )

    finished = _iso()
    gate_fail = reconcile["status"] != "PASS" or not excel_ok or not pdf_ok
    status = "OK" if not gate_fail else "UNRELIABLE"
    if report.get("db_error") and not decisions:
        status = "TECH_FAIL"

    manifest = {
        **report,
        "finished_at": finished,
        "status": status,
        "out_dir": str(out),
        "profile_id": resolved.profile_id,
        "profile_version": resolved.version,
        "profile_hash": resolved.profile_hash,
        "cutoff": resolved.resolved_at[:10],
        "counts": counts,
        "snapshot_path": str(snap_path),
        "delta_counts": delta.get("counts"),
        "export_review": export_meta,
        "calibration": cal.to_dict(),
        "reconcile": reconcile,
        "products": {k: str(v) for k, v in artifacts.items()},
        "checksums_file": str(checksums_path),
        "human_acceptance": "PENDING_HUMAN",
        "gaps": gaps,
        "offline_reconfirm": offline,
        "mapping": {
            "GO": "PARTICIPAR",
            "REVIEW": "REVIEW",
            "NO_GO": "NÃO_PARTICIPAR",
        },
    }
    _atomic_json(out / "decision_manifest.json", manifest)

    if conn is not None:
        try:
            conn.close()
        except Exception as exc:  # noqa: BLE001
            report.setdefault("warnings", []).append(f"conn_close:{exc}")

    # §29 rastreabilidade: decision pack always records errors[] + report→run
    try:
        from scripts.ops.run_execution_ledger import record_execution_safe

        errs: list[str] = list(report.get("warnings") or []) if isinstance(report, dict) else []
        if status in {"TECH_FAIL", "UNRELIABLE"}:
            errs.append(f"status={status}")
        for g in (gaps or [])[:20]:
            errs.append(f"gap:{g}" if not str(g).startswith("gap:") else str(g))
        rec_status = "ok"
        if status == "TECH_FAIL":
            rec_status = "failed"
        elif status == "UNRELIABLE":
            rec_status = "partial"
        report_paths = [str(out / "decision_manifest.json")]
        for key in ("executive_decision_brief.md", "extra_decision_pack.xlsx", "checksums.json"):
            pth = out / key
            if pth.is_file():
                report_paths.append(str(pth))
        record_execution_safe(
            command=["python", "-m", "scripts.ops.decision_pack"],
            status=rec_status,
            errors=errs,
            exit_code=(
                EXIT_TECH
                if status == "TECH_FAIL"
                else (EXIT_UNRELIABLE if status == "UNRELIABLE" else EXIT_OK)
            ),
            report_paths=report_paths,
            run_id=run_id,
            meta={
                "entrypoint": "decision_pack",
                "status": status,
                "offline_reconfirm": offline,
                "counts": counts,
            },
        )
    except Exception:  # noqa: BLE001
        pass

    if status == "TECH_FAIL":
        return EXIT_TECH, manifest
    if status == "UNRELIABLE" and strict:
        return EXIT_UNRELIABLE, manifest
    if status == "UNRELIABLE":
        return EXIT_UNRELIABLE, manifest
    return EXIT_OK, manifest


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Extra decision pack (EXTRA-DECISION-LOOP-01)")
    p.add_argument("--dsn", default=None)
    p.add_argument("--out", type=Path, default=None)
    p.add_argument("--limit", type=int, default=500)
    p.add_argument("--strict", action="store_true")
    p.add_argument("--offline-reconfirm", action="store_true")
    p.add_argument("--skip-db", action="store_true")
    p.add_argument("--reconfirm-top", type=int, default=20)
    p.add_argument("--reconfirm-max", type=int, default=40)
    sub = p.add_subparsers(dest="cmd")
    sub.add_parser("run", help="default run")
    exp = sub.add_parser("review-export", help="export human review sample from last/all decisions")
    exp.add_argument("--decisions-csv", type=Path, required=True)
    exp.add_argument("--out", type=Path, required=True)
    imp = sub.add_parser("review-import", help="import human labels")
    imp.add_argument("--csv", type=Path, required=True)
    cal = sub.add_parser("calibrate", help="calibrate metrics from human labels")
    cal.add_argument("--out", type=Path, default=None)
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        if args.cmd == "review-export":
            rows = []
            with args.decisions_csv.open(encoding="utf-8", newline="") as fh:
                rows = list(csv.DictReader(fh))
            # normalize keys for export
            norm = []
            for r in rows:
                norm.append(
                    {
                        **r,
                        "recommendation": r.get("recommendation") or r.get("system_recommendation"),
                        "id": r.get("opportunity_id") or r.get("id"),
                    }
                )
            meta = export_review_sample(norm, args.out)
            print(json.dumps(meta, ensure_ascii=False, indent=2))
            return 0
        if args.cmd == "review-import":
            res = import_review_labels(args.csv)
            print(json.dumps(res, ensure_ascii=False, indent=2))
            return 0
        if args.cmd == "calibrate":
            cal_res = calibrate()
            if args.out:
                _atomic_json(args.out, cal_res.to_dict())
            print(json.dumps(cal_res.to_dict(), ensure_ascii=False, indent=2))
            return 0 if cal_res.status == "OK" else 2

        code, manifest = run_decision_pack(
            dsn=args.dsn,
            out_dir=args.out,
            limit=args.limit,
            strict=args.strict,
            offline_reconfirm=args.offline_reconfirm,
            skip_db=args.skip_db,
            reconfirm_top=args.reconfirm_top,
            reconfirm_max=args.reconfirm_max,
        )
        print(
            json.dumps(
                {
                    "exit_code": code,
                    "status": manifest.get("status"),
                    "run_id": manifest.get("run_id"),
                    "out_dir": manifest.get("out_dir"),
                    "counts": manifest.get("counts"),
                    "human_acceptance": manifest.get("human_acceptance"),
                    "reconcile": (manifest.get("reconcile") or {}).get("status"),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return code
    except Exception as exc:  # noqa: BLE001
        print(f"TECH_FAIL: {exc}", file=sys.stderr)
        traceback.print_exc()
        return EXIT_TECH


if __name__ == "__main__":
    raise SystemExit(main())
