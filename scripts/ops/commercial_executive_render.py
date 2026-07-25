#!/usr/bin/env python3
"""Renderizadores executivos (PDF + XLSX) para o pack comercial Extra Construtora.

Regras:
- Corpo principal em português de negócio (sem JSON bruto, DSNs, caminhos internos)
- Valores em R$ formato BR; datas dd/mm/aaaa; CNPJ como texto
- Apêndice de auditoria separado
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any


def br_currency(value: Any) -> str:
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "—"
    # 1234567.89 → R$ 1.234.567,89
    s = f"{v:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def br_date(value: Any) -> str:
    if value is None or value == "":
        return "—"
    raw = str(value)[:10]
    try:
        d = date.fromisoformat(raw)
        return d.strftime("%d/%m/%Y")
    except ValueError:
        return str(value)


def br_int(value: Any) -> str:
    try:
        v = int(value)
    except (TypeError, ValueError):
        return "—"
    return f"{v:,}".replace(",", ".")


def cnpj_text(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    if digits:
        return digits
    return "—"


def _safe(s: Any, n: int = 200) -> str:
    t = str(s or "").replace("\n", " ").strip()
    return (t[: n - 1] + "…") if len(t) > n else t


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def build_executive_pdf(
    path: Path,
    *,
    pack: dict[str, Any],
    products: dict[str, Any],
    as_of: str,
    client_name: str = "Extra Empreiteira e Construtora",
    confenge_label: str = "CONFENGE / Extra Construtora",
) -> int:
    """Professional multi-section PDF for client meeting. Returns page count estimate."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        KeepTogether,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    a = products.get("A") or {}
    b = products.get("B") or {}
    c = products.get("C") or {}
    d = products.get("D") or {}
    e = products.get("E") or {}

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="CoverTitle",
            parent=styles["Title"],
            fontSize=20,
            leading=24,
            alignment=TA_CENTER,
            spaceAfter=12,
            textColor=colors.HexColor("#0B3D5C"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="CoverSub",
            parent=styles["Normal"],
            fontSize=12,
            alignment=TA_CENTER,
            spaceAfter=6,
            textColor=colors.HexColor("#333333"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="H1BR",
            parent=styles["Heading1"],
            fontSize=14,
            textColor=colors.HexColor("#0B3D5C"),
            spaceBefore=12,
            spaceAfter=8,
        )
    )
    styles.add(
        ParagraphStyle(
            name="H2BR",
            parent=styles["Heading2"],
            fontSize=11,
            textColor=colors.HexColor("#1A5F7A"),
            spaceBefore=8,
            spaceAfter=4,
        )
    )
    styles.add(
        ParagraphStyle(
            name="BodyBR",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
            alignment=TA_JUSTIFY,
            spaceAfter=4,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SmallBR",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#444444"),
        )
    )
    styles.add(
        ParagraphStyle(
            name="Cell",
            parent=styles["Normal"],
            fontSize=7,
            leading=9,
            alignment=TA_LEFT,
        )
    )

    story: list[Any] = []

    # --- Cover ---
    story.append(Spacer(1, 3 * cm))
    story.append(Paragraph(confenge_label, styles["CoverSub"]))
    story.append(Paragraph("Relatório Executivo de Inteligência B2G", styles["CoverTitle"]))
    story.append(Paragraph(client_name, styles["CoverSub"]))
    story.append(Spacer(1, 8 * mm))
    story.append(
        Paragraph(
            f"Data de referência: <b>{br_date(as_of)}</b><br/>"
            "Escopo: contratos e editais de <b>engenharia civil / obras</b> "
            "aderentes ao perfil da Extra Construtora (SC e raio operacional).<br/>"
            "Status do release candidate: <b>PENDING_HUMAN</b> — aguarda aceite comercial.",
            styles["CoverSub"],
        )
    )
    story.append(Spacer(1, 12 * mm))
    story.append(
        Paragraph(
            "Documento preparado para reunião com o cliente. "
            "Não substitui análise jurídica, contábil ou técnica de proposta.",
            styles["SmallBR"],
        )
    )
    story.append(PageBreak())

    # --- 3. Resumo executivo ---
    n_a = len(a.get("rows") or [])
    n_b = len(b.get("rows") or [])
    n_c = len(c.get("rows") or [])
    n_e = len(e.get("recommendations") or [])
    e_status = str(e.get("status") or "")
    eng_pop = (a.get("population") or {}).get("n_contracts_eligible_engineering") or (
        a.get("population") or {}
    ).get("n_contracts_eligible")

    story.append(Paragraph("1. Resumo executivo", styles["H1BR"]))
    if e_status == "SUCCESS_ZERO_ENGINEERING_OPPORTUNITIES" or n_e == 0:
        opp_msg = (
            "Na evidência congelada de editais abertos <b>não há oportunidades "
            "classificadas como engenharia aderente</b> "
            "(resultado honesto: SUCCESS_ZERO_ENGINEERING_OPPORTUNITIES). "
            "Isso evita poluir a recomendação com objetos irrelevantes "
            "(ex.: materiais hospitalares, TI, frota)."
        )
    else:
        opp_msg = (
            f"Foram priorizadas <b>{n_e}</b> oportunidades abertas de engenharia "
            "com recomendação GO / REVIEW / NO_GO fundamentada."
        )
    story.append(
        Paragraph(
            f"Este relatório consolida inteligência comercial setorial para a "
            f"<b>{client_name}</b>, restrita a obras e serviços de engenharia. "
            f"Órgãos prioritários: <b>{n_a}</b> (por atividade de engenharia, "
            f"não por volume geral de compras). Concorrentes com evidência setorial: "
            f"<b>{n_b}</b>. Contratos vincendos de engenharia: <b>{n_c}</b>. "
            f"{opp_msg}",
            styles["BodyBR"],
        )
    )
    if eng_pop:
        story.append(
            Paragraph(
                f"Base histórica filtrada: cerca de <b>{br_int(eng_pop)}</b> contratos "
                "com objeto de engenharia no recorte analisado.",
                styles["BodyBR"],
            )
        )

    # --- 4. Oportunidades ---
    story.append(Paragraph("2. Principais oportunidades (editais)", styles["H1BR"]))
    recs = list(e.get("recommendations") or [])
    if not recs:
        story.append(
            Paragraph(
                "<b>SUCCESS_ZERO_ENGINEERING_OPPORTUNITIES</b> — nenhum edital aberto "
                "na evidência congelada passou no classificador setorial "
                "(ENGINEERING_HIGH_CONFIDENCE / ENGINEERING_REVIEW). "
                "Não foram inseridos editais irrelevantes apenas para preencher quantidade.",
                styles["BodyBR"],
            )
        )
        excluded = e.get("excluded_non_engineering") or e.get("excluded_not_open")
        if excluded:
            story.append(
                Paragraph(
                    f"Itens da evidência bruta excluídos por não engenharia / categoria: "
                    f"{br_int(excluded)}.",
                    styles["SmallBR"],
                )
            )
    else:
        rows_e = [["Objeto", "Órgão", "Prazo", "Valor", "Recomendação"]]
        for r in recs[:12]:
            rows_e.append(
                [
                    Paragraph(_safe(r.get("titulo") or r.get("objeto"), 80), styles["Cell"]),
                    Paragraph(_safe(r.get("orgao"), 40), styles["Cell"]),
                    Paragraph(br_date(r.get("prazo") or r.get("data_encerramento")), styles["Cell"]),
                    Paragraph(br_currency(r.get("valor_estimado") or r.get("valor")), styles["Cell"]),
                    Paragraph(str(r.get("ranking") or r.get("recomendacao") or "—"), styles["Cell"]),
                ]
            )
        t = Table(rows_e, colWidths=[6.5 * cm, 4 * cm, 2.2 * cm, 2.5 * cm, 2.3 * cm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D5C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FA")]),
                ]
            )
        )
        story.append(t)

    # --- 5. Órgãos ---
    story.append(Paragraph("3. Órgãos prioritários (mercado de engenharia)", styles["H1BR"]))
    story.append(
        Paragraph(
            "Ranking por <b>atividade de engenharia aderente</b> "
            "(quantidade e valor de contratos de obras/serviços de engenharia), "
            "não pelo volume geral de compras do órgão.",
            styles["BodyBR"],
        )
    )
    org_rows = [["#", "Órgão", "UF", "Qtd eng.", "Valor eng.", "Ticket mediano"]]
    for r in (a.get("rows") or [])[:15]:
        org_rows.append(
            [
                str(r.get("rank") or ""),
                Paragraph(_safe(r.get("orgao"), 45), styles["Cell"]),
                str(r.get("uf") or ""),
                br_int(r.get("qtd_contratacoes") or r.get("qtd_engenharia")),
                br_currency(r.get("valor_total") or r.get("valor_engenharia")),
                br_currency(r.get("ticket_medio") or r.get("ticket_mediano")),
            ]
        )
    if len(org_rows) == 1:
        story.append(Paragraph("Nenhum órgão com atividade de engenharia suficiente no recorte.", styles["BodyBR"]))
    else:
        t = Table(org_rows, colWidths=[1 * cm, 7 * cm, 1.2 * cm, 2 * cm, 3.5 * cm, 3 * cm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D5C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FA")]),
                ]
            )
        )
        story.append(t)

    # --- 6. Concorrentes ---
    story.append(Paragraph("4. Concorrentes relevantes", styles["H1BR"]))
    story.append(
        Paragraph(
            "Somente empresas com evidência de atuação em segmentos concorrentes da Extra. "
            "Classes: concorrente direto, adjacente, fornecedor/material, mineração/insumos, "
            "não confirmada, excluir.",
            styles["BodyBR"],
        )
    )
    comp_rows = [["#", "Empresa", "CNPJ", "Classe", "Contratos", "Valor"]]
    for r in (b.get("rows") or [])[:15]:
        comp_rows.append(
            [
                str(r.get("rank") or ""),
                Paragraph(_safe(r.get("nome"), 40), styles["Cell"]),
                Paragraph(cnpj_text(r.get("cnpj")), styles["Cell"]),
                Paragraph(_safe(r.get("classe_concorrente") or r.get("competitor_class") or "direto", 24), styles["Cell"]),
                br_int(r.get("n_contratos")),
                br_currency(r.get("valor_contratado_total")),
            ]
        )
    if len(comp_rows) == 1:
        story.append(Paragraph("Nenhum concorrente com evidência setorial suficiente.", styles["BodyBR"]))
    else:
        t = Table(comp_rows, colWidths=[1 * cm, 5.5 * cm, 3.2 * cm, 2.5 * cm, 1.8 * cm, 3.5 * cm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D5C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FA")]),
                ]
            )
        )
        story.append(t)

    # --- 7. Vincendos ---
    story.append(Paragraph("5. Contratos vincendos (engenharia)", styles["H1BR"]))
    story.append(
        Paragraph(
            "Somente contratos de engenharia aderentes. Ordenados por proximidade do vencimento, "
            "recorrência com evidência, valor e aderência. "
            "<b>Sem percentual inventado de probabilidade.</b>",
            styles["BodyBR"],
        )
    )
    c_rows = [["Órgão", "Objeto", "Término", "Valor", "Confiança"]]
    for r in (c.get("rows") or [])[:12]:
        c_rows.append(
            [
                Paragraph(_safe(r.get("orgao"), 30), styles["Cell"]),
                Paragraph(_safe(r.get("objeto"), 50), styles["Cell"]),
                br_date(r.get("termino_efetivo") or r.get("termino")),
                br_currency(r.get("valor")),
                str(r.get("confianca") or "—"),
            ]
        )
    if len(c_rows) == 1:
        story.append(
            Paragraph(
                "Nenhum contrato vincendo de engenharia na janela após filtro setorial "
                "(success_zero honesto se a query completa retornou vazio).",
                styles["BodyBR"],
            )
        )
    else:
        t = Table(c_rows, colWidths=[4 * cm, 6.5 * cm, 2.2 * cm, 3 * cm, 2 * cm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D5C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FA")]),
                ]
            )
        )
        story.append(t)

    # --- 8. Inteligência de valores ---
    story.append(Paragraph("6. Inteligência de mercado (valores)", styles["H1BR"]))
    d_status = str(d.get("status") or "")
    if d_status in {"INSUFFICIENT_COMPARABLE_DATA", "INSUFFICIENT_SAMPLE", "NOT_READY", "EMPTY"}:
        story.append(
            Paragraph(
                f"Status: <b>{d_status}</b>. Valores globais de contrato heterogêneos "
                "não são apresentados como preço unitário. "
                "Quando não há comparabilidade semântica suficiente "
                "(categoria, subtipo, unidade, porte, região, período, modalidade), "
                "o sistema retorna <b>INSUFFICIENT_COMPARABLE_DATA</b>.",
                styles["BodyBR"],
            )
        )
    ok_panels = [p for p in (d.get("panels") or []) if p.get("status") == "OK"]
    if ok_panels:
        p_rows = [["Grupo", "n", "Mediana", "P25", "P75", "Status"]]
        for p in ok_panels[:10]:
            dims = p.get("dimensions") or {}
            g = " | ".join(f"{k}={v}" for k, v in list(dims.items())[:3])
            p_rows.append(
                [
                    Paragraph(_safe(g, 40), styles["Cell"]),
                    br_int(p.get("n_observations")),
                    br_currency(p.get("median")),
                    br_currency(p.get("p25")),
                    br_currency(p.get("p75")),
                    str(p.get("status")),
                ]
            )
        t = Table(p_rows, colWidths=[6 * cm, 1.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D5C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ]
            )
        )
        story.append(t)
    else:
        story.append(
            Paragraph(
                "Nenhum painel de preços unitários comparáveis aprovado no recorte. "
                "Medianas de grupos inválidos foram eliminadas.",
                styles["BodyBR"],
            )
        )

    # --- 9. Riscos ---
    story.append(Paragraph("7. Riscos e limitações", styles["H1BR"]))
    risks = [
        "Capacidade operacional da Extra (capital, CATs, equipe, garantias) permanece PENDING — "
        "recomendações GO automáticas são degradadas para REVIEW quando campos críticos faltam.",
        "Evidência de editais abertos pode estar desatualizada em relação ao PNCP ao vivo; "
        "validar URL oficial e status antes de decidir.",
        "Concorrentes listados são observáveis por contratos históricos, não ranking de market share.",
        "Valores de contrato são CONTRATADO (magnitude), não preço unitário de medições.",
        "Raio geográfico e universo partem da planilha canônica SC 200 km.",
    ]
    for r in risks:
        story.append(Paragraph(f"• {r}", styles["BodyBR"]))

    # --- 10. Plano de ação ---
    story.append(Paragraph("8. Plano de ação (7 / 30 / 90 dias)", styles["H1BR"]))
    story.append(Paragraph("<b>Próximos 7 dias</b>", styles["H2BR"]))
    story.append(
        Paragraph(
            "• Revisar órgãos top do ranking de engenharia e mapear contatos de engenharia/compras.<br/>"
            "• Validar se há editais novos no PNCP nos segmentos pavimentação, drenagem, reforma predial e edificações.<br/>"
            "• Completar elicitação de CATs e capital de giro (campos PENDING).",
            styles["BodyBR"],
        )
    )
    story.append(Paragraph("<b>Próximos 30 dias</b>", styles["H2BR"]))
    story.append(
        Paragraph(
            "• Monitorar contratos vincendos de engenharia e sinais de relicitação com evidência.<br/>"
            "• Aprofundar dossiês dos 5 órgãos com maior valor de engenharia recorrente.<br/>"
            "• Calibrar faixa de valor e margem mínima com a diretoria da Extra.",
            styles["BodyBR"],
        )
    )
    story.append(Paragraph("<b>Próximos 90 dias</b>", styles["H2BR"]))
    story.append(
        Paragraph(
            "• Ciclo recorrente de inteligência setorial (semanal) com aceite humano do pack.<br/>"
            "• Expandir base de concorrentes diretos com exemplos de contratos sustentando a classe.<br/>"
            "• Construir painéis de valores somente onde houver unidade semântica comparável.",
            styles["BodyBR"],
        )
    )

    # --- 11. Fontes (apêndice) ---
    story.append(PageBreak())
    story.append(Paragraph("Apêndice A — Fontes e metodologia", styles["H1BR"]))
    story.append(
        Paragraph(
            "Fontes: PNCP (contratos e editais), evidências capturadas de ciclo open-tenders, "
            "perfil versionado da Extra Construtora. "
            "Classificação setorial auditável "
            "(ENGINEERING_HIGH_CONFIDENCE | ENGINEERING_REVIEW | NON_ENGINEERING | "
            "AMBIGUOUS | EXCLUDED_CATEGORY) com termos positivos/negativos, categoria, "
            "confiança e evidência textual. "
            "Deliverable E exclui NON_ENGINEERING e EXCLUDED_CATEGORY. "
            "Deliverable D rejeita comparações semanticamente inválidas "
            "(INSUFFICIENT_COMPARABLE_DATA).",
            styles["BodyBR"],
        )
    )
    story.append(Paragraph("Apêndice B — Auditoria técnica (uso interno)", styles["H1BR"]))
    story.append(
        Paragraph(
            f"Identidade do pack, hashes e metadados de reconciliação constam nos arquivos "
            f"de empacotamento (checksums, pack-manifest, ARTIFACT-IDENTITY). "
            f"Data de geração do relatório: {datetime.now().strftime('%d/%m/%Y %H:%M')}. "
            f"Detalhes de run e schema ficam fora do corpo principal deste PDF.",
            styles["SmallBR"],
        )
    )

    def _footer(canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.drawString(2 * cm, 1.2 * cm, f"{client_name} — Inteligência B2G engenharia")
        canvas.drawRightString(A4[0] - 2 * cm, 1.2 * cm, f"Pág. {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.8 * cm,
        title=f"Relatório Executivo B2G — {client_name}",
        author="CONFENGE / Extra Consultoria",
    )
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    # estimate pages from flowables
    pages = max(4, min(20, 3 + (n_a > 0) + (n_b > 0) + (n_c > 0) + 2))
    return pages


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def build_executive_xlsx(
    path: Path,
    *,
    pack: dict[str, Any],
    products: dict[str, Any],
    as_of: str,
    meta: dict[str, Any] | None = None,
) -> None:
    """Workbook navegável para reunião (Dashboard + abas executivas + técnica)."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    a = products.get("A") or {}
    b = products.get("B") or {}
    c = products.get("C") or {}
    d = products.get("D") or {}
    e = products.get("E") or {}
    meta = meta or {}

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0B3D5C")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws: Any, ncol: int) -> None:
        for col in range(1, ncol + 1):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    def autosize(ws: Any, widths: list[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def write_rows(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
        ws.append(headers)
        for row in rows:
            ws.append(row)
        style_header(ws, len(headers))
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in r:
                cell.alignment = wrap
                cell.border = thin
                # Force CNPJ-like columns as text
                if cell.column is not None and headers[cell.column - 1].upper() in {
                    "CNPJ",
                    "ÓRGÃO CNPJ",
                    "ORGAO_CNPJ",
                    "CONTRATADO_CNPJ",
                }:
                    cell.number_format = "@"
                    if cell.value is not None:
                        cell.value = str(cell.value)

    # --- Dashboard ---
    dash = wb.active
    dash.title = "Dashboard"
    dash["A1"] = "Inteligência B2G — Extra Construtora"
    dash["A1"].font = Font(bold=True, size=14, color="0B3D5C")
    dash["A2"] = f"Data de referência: {br_date(as_of)}"
    dash["A3"] = "Escopo: engenharia civil / obras aderentes ao perfil Extra"
    dash["A4"] = f"Status RC: PENDING_HUMAN | Pack: {pack.get('run_id') or '—'}"

    n_a = len(a.get("rows") or [])
    n_b = len(b.get("rows") or [])
    n_c = len(c.get("rows") or [])
    n_e = len(e.get("recommendations") or [])
    e_status = str(e.get("status") or "")

    kpis = [
        ("KPI", "Valor"),
        ("Órgãos prioritários (engenharia)", n_a),
        ("Concorrentes com evidência", n_b),
        ("Contratos vincendos engenharia", n_c),
        ("Oportunidades engenharia abertas", n_e),
        ("Status editais (E)", e_status or "—"),
        ("Status valores (D)", str(d.get("status") or "—")),
    ]
    for i, (k, v) in enumerate(kpis, start=6):
        dash.cell(i, 1, k)
        dash.cell(i, 2, v)
        if i == 6:
            dash.cell(i, 1).fill = header_fill
            dash.cell(i, 1).font = header_font
            dash.cell(i, 2).fill = header_fill
            dash.cell(i, 2).font = header_font
    dash.column_dimensions["A"].width = 40
    dash.column_dimensions["B"].width = 36

    # Chart data from top organs
    dash["A15"] = "Top órgãos por qtd contratos engenharia"
    dash["A15"].font = Font(bold=True, color="0B3D5C")
    dash["A16"] = "Órgão"
    dash["B16"] = "Qtd"
    for i, r in enumerate((a.get("rows") or [])[:8], start=17):
        dash.cell(i, 1, _safe(r.get("orgao"), 40))
        dash.cell(i, 2, int(r.get("qtd_contratacoes") or 0))
    if n_a:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Órgãos prioritários (engenharia)"
        chart.y_axis.title = "Contratos"
        data = Reference(dash, min_col=2, min_row=16, max_row=16 + min(8, n_a))
        cats = Reference(dash, min_col=1, min_row=17, max_row=16 + min(8, n_a))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.style = 10
        dash.add_chart(chart, "D6")

    dash["A27"] = "Próximos passos (7/30/90 dias)"
    dash["A27"].font = Font(bold=True)
    dash["A28"] = "7d: validar órgãos top + novos editais engenharia no PNCP + CATs PENDING"
    dash["A29"] = "30d: monitorar vincendos + dossiês dos 5 órgãos de maior valor"
    dash["A30"] = "90d: ciclo semanal setorial + painéis de preço só com unidade comparável"

    # --- Oportunidades ---
    ws = wb.create_sheet("Oportunidades")
    headers = [
        "Objeto",
        "Órgão",
        "Prazo",
        "Valor estimado",
        "Localização",
        "Segmento",
        "Aderência",
        "Impedimentos",
        "Documentos",
        "URL oficial",
        "Recomendação",
        "Motivo",
        "Dados faltantes",
        "Classificação setorial",
    ]
    rows = []
    for r in e.get("recommendations") or []:
        clf = r.get("sector_classification") or {}
        rows.append(
            [
                r.get("titulo") or r.get("objeto") or "",
                r.get("orgao") or "",
                br_date(r.get("prazo") or r.get("data_encerramento")),
                br_currency(r.get("valor_estimado") or r.get("valor")),
                r.get("localizacao") or r.get("municipio") or r.get("uf") or "",
                r.get("segmento") or clf.get("subcategory") or "",
                r.get("aderencia") or clf.get("label") or "",
                "; ".join(r.get("fatores_impeditivos_ou_riscos") or r.get("impedimentos") or [])
                if isinstance(r.get("fatores_impeditivos_ou_riscos") or r.get("impedimentos"), list)
                else (r.get("impedimentos") or ""),
                "; ".join(r.get("documentos") or r.get("referencias_oficiais") or [])
                if isinstance(r.get("documentos") or r.get("referencias_oficiais"), list)
                else "",
                (r.get("openness") or {}).get("official_url")
                if isinstance(r.get("openness"), dict)
                else (r.get("url") or r.get("official_url") or ""),
                r.get("ranking") or r.get("recomendacao") or "",
                r.get("motivo") or r.get("score_notes") or clf.get("reason") or "",
                r.get("dados_faltantes") or "",
                clf.get("label") or "",
            ]
        )
    if not rows:
        rows = [
            [
                "SUCCESS_ZERO_ENGINEERING_OPPORTUNITIES",
                "—",
                "—",
                "—",
                "—",
                "—",
                "—",
                "Nenhum edital de engenharia na evidência congelada",
                "—",
                "—",
                "NO_GO",
                "Filtro setorial fail-closed",
                "Aguardar novos editais de engenharia",
                "—",
            ]
        ]
    write_rows(ws, headers, rows)
    autosize(ws, [40, 28, 12, 16, 14, 16, 14, 28, 20, 30, 14, 36, 24, 22])
    # hyperlinks
    url_col = headers.index("URL oficial") + 1
    for i, r in enumerate(e.get("recommendations") or [], start=2):
        url = ""
        if isinstance(r.get("openness"), dict):
            url = r["openness"].get("official_url") or ""
        url = url or r.get("url") or r.get("official_url") or ""
        if url and str(url).startswith("http"):
            cell = ws.cell(i, url_col)
            cell.hyperlink = str(url)
            cell.style = "Hyperlink"

    # --- Órgãos ---
    ws = wb.create_sheet("Órgãos")
    headers = [
        "#",
        "Órgão",
        "CNPJ",
        "UF",
        "Qtd contratos engenharia",
        "Valor engenharia",
        "Ticket mediano",
        "Tipos de obra",
        "Fontes",
        "Limitação",
    ]
    rows = []
    for r in a.get("rows") or []:
        rows.append(
            [
                r.get("rank"),
                r.get("orgao"),
                cnpj_text(r.get("orgao_cnpj")),
                r.get("uf"),
                r.get("qtd_contratacoes"),
                br_currency(r.get("valor_total")),
                br_currency(r.get("ticket_medio")),
                ", ".join(r.get("tipos_obra") or r.get("segmentos") or [])
                if isinstance(r.get("tipos_obra") or r.get("segmentos"), list)
                else (r.get("tipos_obra") or ""),
                ", ".join(r.get("fontes") or []) if isinstance(r.get("fontes"), list) else "",
                r.get("data_quality_limitation") or "",
            ]
        )
    write_rows(ws, headers, rows or [["", "Sem órgãos de engenharia no recorte", "", "", "", "", "", "", "", ""]])
    autosize(ws, [4, 40, 20, 6, 12, 18, 16, 24, 20, 30])

    # --- Concorrentes ---
    ws = wb.create_sheet("Concorrentes")
    headers = [
        "#",
        "Empresa",
        "CNPJ",
        "Classe",
        "N contratos",
        "Valor contratado",
        "UFs",
        "Exemplos de contratos",
        "Justificativa",
    ]
    rows = []
    for r in b.get("rows") or []:
        exemplos = r.get("exemplos_contratos") or r.get("sample_contracts") or []
        if isinstance(exemplos, list):
            ex_txt = " | ".join(_safe(x if isinstance(x, str) else x.get("objeto"), 60) for x in exemplos[:3])
        else:
            ex_txt = str(exemplos or "")
        rows.append(
            [
                r.get("rank"),
                r.get("nome"),
                cnpj_text(r.get("cnpj")),
                r.get("classe_concorrente") or r.get("competitor_class") or "concorrente_direto",
                r.get("n_contratos"),
                br_currency(r.get("valor_contratado_total")),
                ", ".join(r.get("ufs") or []) if isinstance(r.get("ufs"), list) else "",
                ex_txt,
                r.get("selection_justification") or "",
            ]
        )
    write_rows(ws, headers, rows or [["", "Sem concorrentes setoriais", "", "", "", "", "", "", ""]])
    autosize(ws, [4, 32, 20, 16, 10, 16, 12, 40, 30])

    # --- Vincendos ---
    ws = wb.create_sheet("Vincendos")
    headers = [
        "Órgão",
        "Objeto",
        "Contratado",
        "CNPJ",
        "Término",
        "Valor",
        "Segmento",
        "Aderência",
        "Confiança",
        "Recorrência (evidência)",
        "Classificação",
    ]
    rows = []
    for r in c.get("rows") or []:
        clf = r.get("sector_classification") or {}
        rel = r.get("relicitacao") or {}
        rows.append(
            [
                r.get("orgao"),
                r.get("objeto"),
                r.get("contratado") or r.get("fornecedor"),
                cnpj_text(r.get("contratado_cnpj") or r.get("fornecedor_cnpj")),
                br_date(r.get("termino_efetivo") or r.get("termino")),
                br_currency(r.get("valor")),
                clf.get("subcategory") or r.get("segmento") or "",
                clf.get("label") or "",
                r.get("confianca") or "",
                rel.get("evidence_class") or rel.get("sinais") or "sem % inventado",
                clf.get("label") or "",
            ]
        )
    write_rows(ws, headers, rows or [["", "Nenhum vincendo de engenharia na janela", "", "", "", "", "", "", "", "", ""]])
    autosize(ws, [28, 40, 24, 20, 12, 14, 14, 18, 10, 20, 18])

    # --- Valores ---
    ws = wb.create_sheet("Valores")
    headers = [
        "Grupo",
        "Categoria",
        "Subtipo",
        "Unidade",
        "Porte",
        "Região",
        "n",
        "Mediana",
        "P25",
        "P75",
        "Min",
        "Max",
        "Status",
        "Limitações",
    ]
    rows = []
    for p in d.get("panels") or []:
        dims = p.get("dimensions") or {}
        rows.append(
            [
                p.get("group_key"),
                dims.get("tipo_obra_servico"),
                dims.get("lote"),
                dims.get("unidade"),
                dims.get("porte"),
                dims.get("regiao"),
                p.get("n_observations"),
                br_currency(p.get("median")) if p.get("median") is not None else "—",
                br_currency(p.get("p25")) if p.get("p25") is not None else "—",
                br_currency(p.get("p75")) if p.get("p75") is not None else "—",
                br_currency(p.get("min_value")) if p.get("min_value") is not None else "—",
                br_currency(p.get("max_value")) if p.get("max_value") is not None else "—",
                p.get("status"),
                "; ".join(p.get("limitations") or []) if isinstance(p.get("limitations"), list) else "",
            ]
        )
    if not rows:
        rows = [["—", "—", "—", "—", "—", "—", 0, "—", "—", "—", "—", "—", str(d.get("status") or "EMPTY"), "Sem painéis"]]
    write_rows(ws, headers, rows)
    autosize(ws, [28, 16, 12, 14, 10, 10, 6, 14, 12, 12, 12, 12, 22, 36])

    # --- Metodologia ---
    ws = wb.create_sheet("Metodologia")
    ws["A1"] = "Metodologia"
    ws["A1"].font = Font(bold=True, size=12, color="0B3D5C")
    lines = [
        "Classificador setorial auditável com labels ENGINEERING_HIGH_CONFIDENCE, ENGINEERING_REVIEW, NON_ENGINEERING, AMBIGUOUS, EXCLUDED_CATEGORY.",
        "Vocabulário positivo/negativo/exclusão no perfil config/client_profiles/extra.yaml.",
        "Genéricos isolados (serviço, manutenção, construção, projeto) não geram aderência.",
        "A: órgãos ranqueados por contratos de engenharia (não volume geral).",
        "B: concorrentes só com evidência setorial e exemplos de contratos.",
        "C: vincendos de engenharia; sem probabilidade inventada.",
        "D: INSUFFICIENT_COMPARABLE_DATA quando unidade/semântica não permite preço unitário.",
        "E: apenas HIGH_CONFIDENCE/REVIEW; caso contrário SUCCESS_ZERO_ENGINEERING_OPPORTUNITIES.",
        "Capacidades PENDING da Extra não são inventadas; GO automático é degradado.",
    ]
    for i, line in enumerate(lines, start=3):
        ws.cell(i, 1, line)
        ws.cell(i, 1).alignment = wrap
    ws.column_dimensions["A"].width = 110

    # --- Fontes ---
    ws = wb.create_sheet("Fontes")
    ws["A1"] = "Fontes"
    ws["A1"].font = Font(bold=True, size=12, color="0B3D5C")
    fontes = [
        "PNCP — Portal Nacional de Contratações Públicas (contratos e editais)",
        "Evidência capturada do ciclo open-tenders (editais abertos)",
        "Perfil operacional Extra Construtora (versionado)",
        "Dump autenticado de contratos em ambiente isolado (não produção)",
    ]
    for i, f in enumerate(fontes, start=3):
        ws.cell(i, 1, f)
    ws.column_dimensions["A"].width = 90

    # --- Dados técnicos (optional / last) ---
    ws = wb.create_sheet("Dados técnicos")
    ws.sheet_state = "hidden"
    ws.append(["chave", "valor"])
    tech = {
        "run_id": pack.get("run_id"),
        "as_of": as_of,
        "git_sha": pack.get("git_sha") or meta.get("git_sha"),
        "campaign_id": pack.get("campaign_id") or meta.get("campaign_id"),
        "deliverable_a_status": a.get("status"),
        "deliverable_b_status": b.get("status"),
        "deliverable_c_status": c.get("status"),
        "deliverable_d_status": d.get("status"),
        "deliverable_e_status": e.get("status"),
        "population": str((a.get("population") or pack.get("population") or {})),
    }
    for k, v in tech.items():
        ws.append([k, str(v) if v is not None else ""])
    style_header(ws, 2)
    autosize(ws, [28, 80])

    wb.save(path)
