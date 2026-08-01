"""Render dos 6 entregáveis cliente EXTRA-MS-OPEN."""

from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path
from typing import Any

from scripts.ops.multi_source_open_pack.models import CanonicalProcess, ReconciliationStats
from scripts.ops.multi_source_open_pack.reconcile import format_reconciliation_labels
from scripts.ops.multi_source_open_pack.textutil import br_currency, br_date, excel_safe

CLIENT_ARTIFACTS = (
    "00-LEIA-ME.md",
    "01-resumo-executivo-multifonte.pdf",
    "02-oportunidades-multifonte-dados.xlsx",
    "oportunidades-multifonte.csv",
    "manifest.json",
    "checksums.json",
)

NAVY_900 = "061a33"
NAVY_800 = "0a294b"
GREEN_700 = "2d6f2d"
LIME = "ced62a"
INK = "071a31"
TEXT = "26374a"
MUTED = "647182"
LINE = "dce3e8"
SOFT = "f5f7f6"
WHITE = "FFFFFF"


def write_readme(
    path: Path,
    *,
    pack_id: str,
    stats: ReconciliationStats,
    generated_at: str,
    as_of: str,
    limitations: list[str],
    motor_version: str,
) -> None:
    labels = format_reconciliation_labels(stats)
    lines = [
        f"# {pack_id}",
        "",
        "Pacote multi-fonte de **inteligência decisória B2G** para Extra Empreiteira e Construtora.",
        "Identidade visual: CONFENGE — navy #061a33 + lime #ced62a.",
        "",
        "## Finalidade",
        "Apoiar o gestor a decidir **onde disputar**, **por quê**, **com quais riscos**,",
        "**contra quem (quando houver base)** e **qual esforço alocar nesta semana**.",
        "Não é dump de publicações nem probabilidade de vitória.",
        "",
        "## Como interpretar os números (dimensões distintas)",
        "",
        "| Conceito | Significado |",
        "|----------|-------------|",
        "| **Ente comprador** | Um dos entes do universo canônico (ex.: 1.093) |",
        "| **Observação bruta** | Registro/publicação em uma fonte (PNCP, DOM, SC Compras) |",
        "| **Processo canônico** | Contratação deduplicada consolidando fontes/publicações |",
        "| **Oportunidade acionável** | Processo aberto + no universo + AEC + sem blocker terminal |",
        "| **Shortlist** | Subconjunto priorizado para esforço desta semana |",
        "",
        "Contagens de observações **não** são contagens de entes.",
        "Contagens de processos são **≤** observações após merge.",
        "",
        "## Arquivos cliente (exatamente estes 6)",
        "- `01-resumo-executivo-multifonte.pdf` — síntese decisória",
        "- `02-oportunidades-multifonte-dados.xlsx` — instrumento analítico (abas)",
        "- `oportunidades-multifonte.csv` — 1 linha por processo canônico",
        "- `manifest.json` / `checksums.json` — integridade + reconciliação",
        "",
        f"Motor: `{motor_version}` · gerado: `{generated_at}` · as_of: `{as_of}`",
        "",
        "## Reconciliação",
        "",
    ]
    for lab, val in labels:
        lines.append(f"- **{lab}:** {val}")
    lines += [
        "",
        "## Critérios de decisão",
        "- `GO` — somente com perfil crítico completo e validação humana (Tiago); "
        "não emitido com campos PENDING.",
        "- `REVIEW` — candidata AEC aberta no universo; exige deep dive.",
        "- `NO_GO` — blocker comprovado (prazo, terminal, fora AEC, fora universo, etc.).",
        "- Score **não** é probabilidade de vitória.",
        "",
        "## Limitações materiais",
        "",
    ]
    for lim in limitations:
        lines.append(f"- {lim}")
    lines += [
        "",
        "Aceite humano: **PENDING_HUMAN** (Tiago). Ausência de manifestação ≠ aceite.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def write_csv(path: Path, processes: list[CanonicalProcess]) -> None:
    rows = [p.to_csv_row() for p in processes]
    if not rows:
        # still write header
        fieldnames = list(CanonicalProcess(
            process_id="",
            merge_key="",
            merge_method="",
            merge_confidence=0,
            fontes=[],
            observation_ids=[],
            id_externo_principal="",
            orgao="",
            orgao_cnpj="",
            municipio="",
            uf="",
            objeto="",
            modalidade="",
            valor_estimado=None,
            data_publicacao="",
            data_encerramento="",
            deadline_dt=None,
            url_oficial="",
            urls_all=[],
            status_processo="",
            event_types=[],
            is_active_dispute=False,
            in_universe=False,
            match_universo="",
            distance_km=None,
            distance_method="",
            entity_key="",
            calendar_days_remaining=None,
            business_days_remaining=None,
        ).to_csv_row().keys())
        with path.open("w", encoding="utf-8", newline="") as f:
            csv.DictWriter(f, fieldnames=fieldnames).writeheader()
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_excel(
    path: Path,
    *,
    pack_meta: dict[str, Any],
    processes: list[CanonicalProcess],
    shortlist: list[CanonicalProcess],
    stats: ReconciliationStats,
    source_policy: list[dict[str, str]],
    limitations: list[str],
    observations_sample: list[dict[str, Any]] | None = None,
    logo_path: Path | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    fill_header = PatternFill("solid", fgColor=NAVY_900)
    fill_soft = PatternFill("solid", fgColor=SOFT)
    fill_go = PatternFill("solid", fgColor="c6efce")
    fill_review = PatternFill("solid", fgColor="fff2cc")
    fill_nogo = PatternFill("solid", fgColor="f4cccc")
    font_h = Font(name="Calibri", bold=True, color=WHITE, size=11)
    font_title = Font(name="Calibri", bold=True, color=NAVY_900, size=16)
    font_sub = Font(name="Calibri", color=TEXT, size=10)
    font_cell = Font(name="Calibri", color=INK, size=9)
    thin = Border(
        left=Side(style="thin", color=LINE),
        right=Side(style="thin", color=LINE),
        top=Side(style="thin", color=LINE),
        bottom=Side(style="thin", color=LINE),
    )

    def style_header(ws, cols: list[str]) -> None:
        for j, c in enumerate(cols, start=1):
            cell = ws.cell(1, j, c)
            cell.font = font_h
            cell.fill = fill_header
            cell.alignment = Alignment(wrap_text=True)

    def autosize(ws, cols: list[str], max_w: int = 40) -> None:
        for j, c in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(j)].width = min(max_w, max(12, len(c) + 2))

    # --- Resumo ---
    ws = wb.active
    ws.title = "Resumo"
    if logo_path and logo_path.is_file():
        try:
            from openpyxl.drawing.image import Image as XLImage

            img = XLImage(str(logo_path))
            img.width = 180
            img.height = 48
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 40
        except Exception as logo_exc:  # noqa: BLE001
            # Logo is optional branding; workbook generation continues without it.
            ws["A2"] = f"(logo opcional não aplicado: {type(logo_exc).__name__})"
    ws["A3"] = "CONFENGE — Inteligência decisória B2G"
    ws["A3"].font = font_title
    ws["A4"] = "Pacote multi-fonte — Extra Construtora (processos canônicos)"
    ws["A4"].font = Font(name="Calibri", bold=True, color=NAVY_800, size=12)
    ws["A5"] = (
        f"Gerado: {pack_meta['generated_at']} | as_of: {pack_meta['as_of']} | "
        f"pack_id: {pack_meta['pack_id']} | motor: {pack_meta.get('motor_version', '')}"
    )
    ws["A5"].font = font_sub
    ws["A7"] = "Indicador (dimensão explícita)"
    ws["B7"] = "Valor"
    ws["A7"].font = font_h
    ws["B7"].font = font_h
    ws["A7"].fill = fill_header
    ws["B7"].fill = fill_header
    for i, (lab, val) in enumerate(format_reconciliation_labels(stats), start=8):
        ws[f"A{i}"] = lab
        ws[f"B{i}"] = val
        ws[f"A{i}"].font = font_cell
        ws[f"B{i}"].font = font_cell
        if i % 2 == 0:
            ws[f"A{i}"].fill = fill_soft
            ws[f"B{i}"].fill = fill_soft
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A8"

    # --- Shortlist ---
    wsl = wb.create_sheet("Shortlist")
    sl_cols = [
        "process_id",
        "recommendation",
        "score",
        "orgao",
        "municipio",
        "distance_km",
        "objeto",
        "category_label",
        "valor_estimado",
        "data_encerramento",
        "dias_uteis_restantes",
        "url_oficial",
        "fontes",
        "blockers",
        "next_action",
        "pending",
        "inclusion_reason",
    ]
    style_header(wsl, sl_cols)
    for i, p in enumerate(shortlist, start=2):
        row = p.to_csv_row()
        for j, c in enumerate(sl_cols, start=1):
            val = row.get(c, "")
            cell = wsl.cell(i, j, excel_safe(val))
            cell.font = font_cell
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            rec = row.get("recommendation", "")
            if rec == "GO":
                cell.fill = fill_go
            elif rec == "REVIEW":
                cell.fill = fill_review
            elif rec == "NO_GO":
                cell.fill = fill_nogo
            if c == "url_oficial" and val and str(val).startswith("http"):
                cell.hyperlink = str(val)
                cell.style = "Hyperlink"
        wsl.row_dimensions[i].height = 48
    autosize(wsl, sl_cols)
    wsl.auto_filter.ref = f"A1:{get_column_letter(len(sl_cols))}{max(1, len(shortlist)+1)}"
    wsl.freeze_panes = "A2"

    # --- Processos canônicos ---
    wp = wb.create_sheet("Processos_canonicos")
    # decision layer first
    ordered = sorted(
        processes,
        key=lambda p: (
            0 if p.layer == "decision" else 1,
            0 if p.decision and p.decision.recommendation == "GO" else 1,
            0 if p.decision and p.decision.recommendation == "REVIEW" else 2,
            -(p.decision.score if p.decision else 0),
        ),
    )
    p_cols = list(ordered[0].to_csv_row().keys()) if ordered else ["process_id"]
    style_header(wp, p_cols)
    for i, p in enumerate(ordered, start=2):
        row = p.to_csv_row()
        for j, c in enumerate(p_cols, start=1):
            val = row.get(c, "")
            cell = wp.cell(i, j, excel_safe(val))
            cell.font = font_cell
            cell.border = thin
            if c == "url_oficial" and val and str(val).startswith("http"):
                cell.hyperlink = str(val)
                cell.style = "Hyperlink"
            rec = row.get("recommendation", "")
            if rec == "NO_GO" and j == 1:
                cell.fill = fill_nogo
            elif rec == "REVIEW" and p.layer == "decision":
                cell.fill = fill_review
            elif i % 2 == 0:
                cell.fill = fill_soft
    autosize(wp, p_cols, max_w=28)
    if ordered:
        wp.auto_filter.ref = f"A1:{get_column_letter(len(p_cols))}{len(ordered)+1}"
    wp.freeze_panes = "A2"

    # --- Documentos ---
    wd = wb.create_sheet("Documentos_links")
    d_cols = [
        "process_id",
        "orgao",
        "doc_type",
        "title",
        "url",
        "fonte",
        "download_status",
        "parse_status",
        "official_page_validated",
    ]
    style_header(wd, d_cols)
    ri = 2
    for p in ordered:
        if not p.documents:
            wd.cell(ri, 1, p.process_id)
            wd.cell(ri, 2, excel_safe(p.orgao))
            wd.cell(ri, 3, "nenhum")
            wd.cell(ri, 4, "sem documento linkado")
            wd.cell(ri, 7, p.docs_inventory_status)
            ri += 1
            continue
        for doc in p.documents:
            wd.cell(ri, 1, p.process_id)
            wd.cell(ri, 2, excel_safe(p.orgao))
            wd.cell(ri, 3, doc.doc_type)
            wd.cell(ri, 4, excel_safe(doc.title))
            cell = wd.cell(ri, 5, excel_safe(doc.url))
            if doc.url.startswith("http"):
                cell.hyperlink = doc.url
                cell.style = "Hyperlink"
            wd.cell(ri, 6, doc.fonte)
            wd.cell(ri, 7, doc.download_status)
            wd.cell(ri, 8, doc.parse_status)
            wd.cell(ri, 9, "sim" if p.official_page_validated else "nao")
            ri += 1
    autosize(wd, d_cols)
    wd.freeze_panes = "A2"

    # --- Analise edital / requisitos / riscos (honest placeholders + decision data) ---
    for sheet_name, builder in (
        (
            "Analise_edital",
            lambda p: [
                p.process_id,
                p.orgao,
                p.objeto[:200],
                p.modalidade,
                p.valor_estimado if p.valor_estimado is not None else "",
                "estimado",
                p.requirements_summary,
                p.docs_inventory_status,
                "extração completa depende de download/OCR do edital",
            ],
        ),
        (
            "Orgaos",
            lambda p: [
                p.process_id,
                p.orgao,
                p.orgao_cnpj,
                p.municipio,
                p.distance_km if p.distance_km is not None else "",
                p.buyer_analysis,
            ],
        ),
        (
            "Concorrentes",
            lambda p: [
                p.process_id,
                p.orgao,
                p.competitors_probable,
                "nao_inventado",
            ],
        ),
        (
            "Riscos_acoes",
            lambda p: [
                p.process_id,
                p.decision.recommendation if p.decision else "",
                p.risks_summary,
                p.decision.next_action if p.decision else "",
                p.decision.owner_suggested if p.decision else "",
                p.decision.action_deadline if p.decision else "",
                "|".join(p.decision.pending) if p.decision else "",
            ],
        ),
    ):
        w = wb.create_sheet(sheet_name)
        headers = {
            "Analise_edital": [
                "process_id",
                "orgao",
                "objeto",
                "modalidade",
                "valor_estimado",
                "valor_semantica",
                "requisitos_resumo",
                "docs_status",
                "nota",
            ],
            "Orgaos": [
                "process_id",
                "orgao",
                "orgao_cnpj",
                "municipio",
                "distance_km",
                "analise_orgao",
            ],
            "Concorrentes": [
                "process_id",
                "orgao",
                "concorrentes_provaveis",
                "evidencia",
            ],
            "Riscos_acoes": [
                "process_id",
                "recommendation",
                "riscos",
                "next_action",
                "owner",
                "action_deadline",
                "pending",
            ],
        }[sheet_name]
        style_header(w, headers)
        focus = shortlist or [p for p in ordered if p.layer == "decision"][:50]
        for i, p in enumerate(focus, start=2):
            vals = builder(p)
            for j, v in enumerate(vals, start=1):
                w.cell(i, j, excel_safe(v)).font = font_cell
        autosize(w, headers)
        w.freeze_panes = "A2"

    # --- Observacoes brutas (secondary) ---
    wo = wb.create_sheet("Observacoes_brutas")
    o_cols = [
        "observation_id",
        "fonte",
        "id_externo",
        "orgao",
        "objeto",
        "event_type",
        "is_active_dispute",
        "in_universe",
        "url",
    ]
    style_header(wo, o_cols)
    for i, r in enumerate((observations_sample or [])[:5000], start=2):
        for j, c in enumerate(o_cols, start=1):
            wo.cell(i, j, excel_safe(r.get(c, ""))).font = font_cell
    autosize(wo, o_cols)
    wo.freeze_panes = "A2"

    # --- Cobertura / Politica / Limitacoes / Metodologia ---
    wc = wb.create_sheet("Cobertura_fontes")
    style_header(wc, ["indicador", "valor"])
    for i, (lab, val) in enumerate(format_reconciliation_labels(stats), start=2):
        wc.cell(i, 1, lab)
        wc.cell(i, 2, val)
    wc.freeze_panes = "A2"

    wpol = wb.create_sheet("Politica_fontes")
    ph = ["fonte", "papel_open_tenders", "esferas", "status_no_pack", "notas"]
    style_header(wpol, ph)
    for i, row in enumerate(source_policy, start=2):
        for j, h in enumerate(ph, start=1):
            wpol.cell(i, j, row.get(h, "")).font = font_cell

    wl = wb.create_sheet("Limitacoes")
    wl["A1"] = "Limitações, non-claims e metodologia"
    wl["A1"].font = font_title
    for i, line in enumerate(limitations, start=3):
        wl[f"A{i}"] = f"• {line}"
        wl[f"A{i}"].font = font_sub
    wl.column_dimensions["A"].width = 110

    wm = wb.create_sheet("Metodologia")
    method_lines = [
        "Modelo semântico: ente → consulta de fonte → observação → processo canônico → evento → documento → avaliação.",
        "Dedup: PNCP control → CNPJ+processo+ano → platform id → órgão+modalidade+objeto → fingerprint.",
        "Classificação AEC: scripts.ops.sector_classifier (hierarquia aec-hierarchy) — engenharia_hint NÃO é autoridade.",
        "Decisão: GO|REVIEW|NO_GO com hard gates; GO bloqueado com campos críticos PENDING do perfil Extra.",
        "Distância: geodésica do seed universo (Florianópolis); não rotulada como rodoviária.",
        "CSV = 1 linha por processo canônico. Fora-universo apenas layer=secondary_reference.",
        f"Scoring version: {pack_meta.get('scoring_version', '')}",
        f"Taxonomy version: {pack_meta.get('taxonomy_version', '')}",
    ]
    wm["A1"] = "Metodologia e versões"
    wm["A1"].font = font_title
    for i, line in enumerate(method_lines, start=3):
        wm[f"A{i}"] = f"• {line}"
        wm[f"A{i}"].font = font_sub
    wm.column_dimensions["A"].width = 110

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_pdf(
    path: Path,
    *,
    pack_meta: dict[str, Any],
    shortlist: list[CanonicalProcess],
    stats: ReconciliationStats,
    limitations: list[str],
    logo_path: Path | None = None,
) -> None:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_JUSTIFY
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        HRFlowable,
        Image,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    navy = colors.HexColor(f"#{NAVY_900}")
    navy800 = colors.HexColor(f"#{NAVY_800}")
    lime = colors.HexColor(f"#{LIME}")
    text_c = colors.HexColor(f"#{TEXT}")
    soft = colors.HexColor(f"#{SOFT}")
    muted = colors.HexColor(f"#{MUTED}")
    green = colors.HexColor(f"#{GREEN_700}")

    styles = getSampleStyleSheet()
    for name, kwargs in {
        "BrandTitle": dict(fontName="Helvetica-Bold", fontSize=16, leading=20, textColor=navy, spaceAfter=4),
        "BrandSub": dict(fontName="Helvetica", fontSize=10, leading=13, textColor=navy800, spaceAfter=6),
        "H1C": dict(fontName="Helvetica-Bold", fontSize=12, leading=15, textColor=navy, spaceBefore=10, spaceAfter=5),
        "BodyC": dict(
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=text_c,
            alignment=TA_JUSTIFY,
            spaceAfter=4,
        ),
        "SmallC": dict(fontName="Helvetica", fontSize=8, leading=10, textColor=muted, spaceAfter=2),
        "CellC": dict(fontName="Helvetica", fontSize=7, leading=9, textColor=text_c),
        "Eyebrow": dict(fontName="Helvetica-Bold", fontSize=8, textColor=green, spaceAfter=2),
    }.items():
        styles.add(ParagraphStyle(name=name, **kwargs))

    story: list[Any] = []
    if logo_path and logo_path.is_file():
        try:
            logo = Image(str(logo_path), width=5.2 * cm, height=1.35 * cm)
            ht = Table(
                [
                    [
                        logo,
                        Paragraph(
                            "CONFENGE<br/><font size='9' color='#647182'>"
                            "Inteligência técnica para obras públicas</font>",
                            styles["BrandSub"],
                        ),
                    ]
                ],
                colWidths=[6 * cm, 11 * cm],
            )
            ht.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("BACKGROUND", (0, 0), (-1, -1), soft),
                        ("LINEBELOW", (0, 0), (-1, -1), 3, lime),
                    ]
                )
            )
            story.append(ht)
        except Exception:
            story.append(Paragraph("CONFENGE", styles["BrandTitle"]))
    else:
        story.append(Paragraph("CONFENGE", styles["BrandTitle"]))

    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph("Consultoria B2G · Extra Construtora", styles["Eyebrow"]))
    story.append(Paragraph("Relatório executivo decisório — multi-fonte", styles["BrandTitle"]))
    story.append(
        Paragraph(
            f"Cliente: <b>Extra Empreiteira e Construtora</b> · "
            f"Universo: <b>{stats.entes_universo} entes</b> (200 km) · "
            f"Referência: <b>{br_date(pack_meta['as_of'])}</b><br/>"
            f"pack_id: {pack_meta['pack_id']} · gerado: {pack_meta['generated_at']} · "
            f"motor: {pack_meta.get('motor_version', '')}",
            styles["BodyC"],
        )
    )
    story.append(HRFlowable(width="100%", thickness=1, color=lime, spaceBefore=4, spaceAfter=8))

    story.append(Paragraph("1. Estado do mercado monitorado e cobertura reconciliada", styles["H1C"]))
    story.append(
        Paragraph(
            f"O motor distingue <b>entes</b> ({stats.entes_universo}), "
            f"<b>observações brutas</b> ({stats.observacoes_brutas}: "
            f"PNCP {stats.observacoes_por_fonte.get('pncp', 0)} observações, "
            f"CIGA/DOM {stats.observacoes_por_fonte.get('ciga_ckan', 0)} publicações, "
            f"SC Compras {stats.observacoes_por_fonte.get('sc_compras', 0)} observações), "
            f"<b>processos canônicos</b> ({stats.processos_canonicos} após {stats.merges_realizados} merges), "
            f"dos quais <b>{stats.processos_abertos} abertos</b> com disputa ativa, "
            f"<b>{stats.processos_no_universo} no universo</b>, "
            f"<b>{stats.processos_aec} AEC</b>, "
            f"<b>{stats.oportunidades_acionaveis} acionáveis</b> e "
            f"<b>{stats.shortlist} na shortlist</b>. "
            f"Observações no universo ({stats.observacoes_no_universo}) "
            f"<b>não</b> equivalem a {stats.observacoes_no_universo} entes.",
            styles["BodyC"],
        )
    )

    kpi = [["Indicador dimensional", "Valor"]]
    for lab, val in format_reconciliation_labels(stats)[:16]:
        kpi.append([lab, str(val)])
    kt = Table(kpi, colWidths=[12 * cm, 4 * cm])
    kt.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), navy),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, soft]),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor(f"#{LINE}")),
            ]
        )
    )
    story.append(kt)

    story.append(Paragraph("2. Shortlist ordenada — decisões e prazos", styles["H1C"]))
    if not shortlist:
        story.append(
            Paragraph(
                "Nenhuma oportunidade atendeu aos gates da shortlist (universo + AEC + disputa aberta). "
                "Ver aba Processos_canonicos e exclusões no manifest.",
                styles["BodyC"],
            )
        )
    else:
        data = [["#", "Decisão", "Órgão / dist.", "Objeto / categoria", "Prazo", "Ação"]]
        for i, p in enumerate(shortlist[:15], start=1):
            d = p.decision
            url = p.url_oficial or ""
            link = f'<link href="{url}">abrir</link>' if url.startswith("http") else "—"
            dist = f"{p.distance_km:.0f} km" if p.distance_km is not None else "dist. n/d"
            data.append(
                [
                    str(i),
                    Paragraph(
                        f"<b>{d.recommendation if d else '?'}</b><br/>score {d.score if d else '—'}{link and '<br/>'+link}",
                        styles["CellC"],
                    ),
                    Paragraph(
                        f"{(p.orgao or '')[:42]}<br/>{p.municipio}/{p.uf} · {dist}",
                        styles["CellC"],
                    ),
                    Paragraph(
                        f"{(p.objeto or '')[:100]}<br/><i>{(d.category_label if d else '')}</i>",
                        styles["CellC"],
                    ),
                    Paragraph(
                        f"{br_date(p.data_encerramento)}<br/>"
                        f"{p.business_days_remaining if p.business_days_remaining is not None else '—'} dúteis",
                        styles["CellC"],
                    ),
                    Paragraph(
                        f"{(d.next_action if d else '')[:90]}<br/>"
                        f"<b>{br_currency(p.valor_estimado) if p.valor_estimado else 'valor n/d'}</b>",
                        styles["CellC"],
                    ),
                ]
            )
        ot = Table(data, colWidths=[0.7 * cm, 2.2 * cm, 3.5 * cm, 5.5 * cm, 2.2 * cm, 3.5 * cm])
        ot.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), navy),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{LINE}")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, soft]),
                    ("LINEBELOW", (0, 0), (-1, 0), 2, lime),
                ]
            )
        )
        story.append(ot)

    story.append(Paragraph("3. Alocação de esforço e decisões requeridas", styles["H1C"]))
    story.append(
        Paragraph(
            f"Concentrar equipe/orçamento nas <b>{stats.shortlist}</b> linhas da shortlist "
            f"(GO={stats.go}, REVIEW={stats.review}, NO_GO={stats.no_go} no total de processos). "
            "1) Validar páginas oficiais e baixar editais das top 5–10. "
            "2) Preencher campos PENDING do perfil Extra (CATs, capital, equipe) antes de qualquer GO. "
            "3) Não tratar publicações DOM de contrato/credenciamento como oportunidade aberta. "
            "4) Processos fora do universo ficam apenas como referência secundária.",
            styles["BodyC"],
        )
    )

    story.append(Paragraph("4. Limitações que alteram a decisão", styles["H1C"]))
    for lim in limitations[:12]:
        story.append(Paragraph(f"• {lim}", styles["SmallC"]))

    story.append(Spacer(1, 6 * mm))
    story.append(HRFlowable(width="100%", thickness=2, color=navy, spaceBefore=4, spaceAfter=4))
    story.append(
        Paragraph(
            "CONFENGE · confenge.com.br · Documento para decisão comercial. "
            "Não é parecer jurídico nem probabilidade de vitória. "
            "Aceite: PENDING_HUMAN (Tiago).",
            styles["SmallC"],
        )
    )

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(navy)
        canvas.rect(0, A4[1] - 8, A4[0], 8, fill=1, stroke=0)
        canvas.setFillColor(lime)
        canvas.rect(0, 0, A4[0], 6, fill=1, stroke=0)
        canvas.setFillColor(muted)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(1.5 * cm, 1.0 * cm, f"{pack_meta['pack_id']} · p.{doc.page}")
        canvas.drawRightString(A4[0] - 1.5 * cm, 1.0 * cm, "CONFENGE × Extra")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=1.4 * cm,
        rightMargin=1.4 * cm,
        topMargin=1.3 * cm,
        bottomMargin=1.5 * cm,
        title="Inteligência decisória multi-fonte — Extra",
        author="CONFENGE",
        subject=f"PENDING_HUMAN | {pack_meta['pack_id']}",
    )
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)


def write_checksums_and_manifest(
    out_dir: Path,
    *,
    pack_meta: dict[str, Any],
    artifact_paths: dict[str, Path],
) -> tuple[Path, Path]:
    checksums: dict[str, Any] = {
        "schema": "extra-ms-open-pack/2.0",
        "artifacts": {},
    }
    for name, p in artifact_paths.items():
        if name in {"manifest.json", "checksums.json"}:
            continue
        if not p.is_file():
            continue
        data = p.read_bytes()
        checksums["artifacts"][name] = {
            "sha256": hashlib.sha256(data).hexdigest(),
            "bytes": len(data),
        }
    cpath = out_dir / "checksums.json"
    cpath.write_text(json.dumps(checksums, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    pack_meta["checksums"] = checksums
    mpath = out_dir / "manifest.json"
    mpath.write_text(
        json.dumps(pack_meta, indent=2, ensure_ascii=False, default=str) + "\n",
        encoding="utf-8",
    )
    # refresh checksums including manifest? Spec: checksums of client product files;
    # include the four content files; manifest/checksums self-ref ok as in v1.
    return mpath, cpath
