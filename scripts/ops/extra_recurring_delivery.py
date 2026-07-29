#!/usr/bin/env python3
"""EXTRA recurring delivery — compare weekly packs, emit deltas + reports.

CLI-first, fail-closed. Compares current weekly run vs previous and writes
external delivery artifacts (weekly/monthly reports, EventDelta, urgent alerts).

Does NOT: dashboard, frontend, SaaS, queue, streaming, multi-channel notify.

Canonical entry:

  python3 -m scripts.ops.extra_recurring_delivery run \\
    --current-run /path/weekly-current \\
    --delivery-out /path/external \\
    [--previous-run /path/weekly-prev] \\
    [--previous-monthly /path/monthly-prev] \\
    [--profile config/client_profiles/extra.yaml] \\
    [--expiry-window-days 180] \\
    [--as-of YYYY-MM-DD]

Reuses:
  - scripts.ops.extra_first_client_delivery.validate_weekly_pack (WeeklyInput gate)
  - scripts.ops.strategic_monthly_monitor (variation, window helpers, report builders)
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from dataclasses import asdict, dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from scripts.ops.extra_first_client_delivery import (
    load_csv_rows,
    load_profile,
    parse_date,
    parse_float,
    sha256_file,
    validate_weekly_pack,
)
from scripts.ops.strategic_monthly_monitor import (
    compute_variation,
    contracts_in_window,
)

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_PROFILE = PROJECT_ROOT / "config" / "client_profiles" / "extra.yaml"
SCHEMA = "extra-recurring-delivery/1.0"

EXIT_OK = 0
EXIT_TECH = 1
EXIT_FAIL_CLOSED = 2
EXIT_BLOCKED = 3  # source weekly not consultive (exit_code != 0)

# Frozen EventDelta.event_type allow-list
ALLOWED_EVENT_TYPES = frozenset(
    {
        "NEW_TENDER",
        "DEADLINE_CHANGED",
        "STATUS_CHANGED",
        "SUSPENDED",
        "REVOKED",
        "REOPENED",
        "RECTIFIED",
        "CONTRACT_ENTERED_EXPIRY_WINDOW",
        "NEW_WINNER",
        "WINNER_CONCENTRATION_CHANGED",
        "SOURCE_DEGRADED",
        "FRESHNESS_BREACH",
    }
)

URGENT_EVENT_TYPES = frozenset(
    {
        "SUSPENDED",
        "REVOKED",
        "FRESHNESS_BREACH",
        "SOURCE_DEGRADED",
        "DEADLINE_CHANGED",
        "CONTRACT_ENTERED_EXPIRY_WINDOW",
        "REOPENED",
    }
)

# Freshness levels ordered worst→best for degradation detection
_FRESHNESS_RANK = {
    "never_crawled": 0,
    "unknown": 1,
    "stale": 2,
    "degraded": 2,
    "aging": 3,
    "fresh": 4,
    "ok": 4,
}

_STATUS_EVENT_MAP = (
    (("suspens", "suspend"), "SUSPENDED"),
    (("revog", "cancelad", "anulad", "revok"), "REVOKED"),
    (("reabert", "reopen"), "REOPENED"),
    (("retific", "rectif"), "RECTIFIED"),
)


# ---------------------------------------------------------------------------
# Frozen interfaces
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class EventDelta:
    """Frozen delta contract between weekly runs."""

    entity_type: str
    entity_id: str
    event_type: str
    previous_value: Any
    current_value: Any
    detected_at: str
    source_run_id: str | None
    previous_run_id: str | None
    official_url: str | None
    severity: str
    action_required: str

    def __post_init__(self) -> None:
        if self.event_type not in ALLOWED_EVENT_TYPES:
            raise ValueError(f"event_type not allowed: {self.event_type}")

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class WeeklyInput:
    """Frozen weekly pack surface used by recurring delivery.

    Built from validate_weekly_pack + CSVs. Do not invent fields outside
    the weekly_cycle product layout.
    """

    path: Path
    cycle_id: str | None
    collection_id: str | None
    exit_code: int | None
    opportunities: list[dict[str, Any]] = field(default_factory=list)
    contracts: list[dict[str, Any]] = field(default_factory=list)
    competitors: list[dict[str, Any]] = field(default_factory=list)
    orgaos: list[dict[str, Any]] = field(default_factory=list)
    source_health: list[dict[str, Any]] = field(default_factory=list)
    manifest: dict[str, Any] = field(default_factory=dict)
    checksums: dict[str, Any] = field(default_factory=dict)
    validation_ok: bool = False
    validation_errors: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def utc_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def detected_at_for(as_of: date) -> str:
    """Deterministic timestamp for idempotent outputs (as_of noon UTC)."""
    return f"{as_of.isoformat()}T12:00:00Z"


def opp_id(row: dict[str, Any]) -> str:
    for k in (
        "numero_controle_pncp",
        "source_id",
        "id",
        "edital_id",
    ):
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def contract_id(row: dict[str, Any]) -> str:
    for k in ("contrato_id", "contract_id", "id", "source_id"):
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def winner_key(row: dict[str, Any]) -> str:
    for k in ("fornecedor_cnpj", "contratado_cnpj", "cnpj"):
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    name = row.get("fornecedor_nome") or row.get("contratado") or row.get("fornecedor") or ""
    return str(name).strip()


def opp_status(row: dict[str, Any]) -> str:
    return str(
        row.get("status_canonico") or row.get("status") or row.get("situation") or ""
    ).strip()


def opp_deadline(row: dict[str, Any]) -> str | None:
    d = parse_date(row.get("data_encerramento") or row.get("prazo_fim") or row.get("deadline"))
    return d.isoformat() if d else None


def opp_url(row: dict[str, Any]) -> str | None:
    for k in ("link_edital", "official_url", "source_url", "url"):
        v = row.get(k)
        if v and str(v).strip():
            return str(v).strip()
    return None


def classify_status_event(old: str, new: str) -> str:
    """Map status transition to allowed event_type (specific > generic)."""
    new_l = (new or "").lower()
    old_l = (old or "").lower()
    for markers, etype in _STATUS_EVENT_MAP:
        if any(m in new_l for m in markers) and not any(m in old_l for m in markers):
            return etype
    # reopened: was suspended/revoked, now open
    if any(m in old_l for m in ("suspens", "revog", "cancelad")) and any(
        m in new_l for m in ("open", "abert", "public")
    ):
        return "REOPENED"
    return "STATUS_CHANGED"


def severity_for(event_type: str, *, days_remaining: int | None = None) -> str:
    if event_type in {"REVOKED", "FRESHNESS_BREACH"}:
        return "critical"
    if event_type in {"SUSPENDED", "SOURCE_DEGRADED", "REOPENED"}:
        return "high"
    if event_type == "DEADLINE_CHANGED":
        if days_remaining is not None and days_remaining <= 7:
            return "high"
        return "medium"
    if event_type == "CONTRACT_ENTERED_EXPIRY_WINDOW":
        if days_remaining is not None and days_remaining <= 30:
            return "high"
        return "medium"
    if event_type in {"NEW_TENDER", "RECTIFIED", "NEW_WINNER", "WINNER_CONCENTRATION_CHANGED"}:
        return "medium"
    return "low"


def action_for(event_type: str) -> str:
    return {
        "NEW_TENDER": "Avaliar elegibilidade e prazos; incluir no radar semanal.",
        "DEADLINE_CHANGED": "Recalcular cronograma de proposta; alertar equipe comercial.",
        "STATUS_CHANGED": "Verificar status oficial e atualizar pipeline.",
        "SUSPENDED": "Confirmar suspensão na fonte oficial; pausar esforço de proposta.",
        "REVOKED": "Arquivar oportunidade; não gastar esforço comercial.",
        "REOPENED": "Reavaliar viabilidade e prazo; pode reabrir disputa.",
        "RECTIFIED": "Ler retificação completa; revisar requisitos e orçamento.",
        "CONTRACT_ENTERED_EXPIRY_WINDOW": "Priorizar órgão para prospecção pré-vencimento.",
        "NEW_WINNER": "Registrar novo player no mapa de concorrência do órgão.",
        "WINNER_CONCENTRATION_CHANGED": "Revisar concentração de mercado no panorama mensal.",
        "SOURCE_DEGRADED": "Investigar coleta; não confiar em ausência de oportunidades.",
        "FRESHNESS_BREACH": "Reexecutar coleta da fonte antes de decisão consultiva.",
    }.get(event_type, "Revisar delta no relatório semanal.")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    """Write CSV product. Never emit zero-byte files (header-only or row_count=0)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        headers = fieldnames or ["row_count"]
        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
            w.writeheader()
            if headers == ["row_count"]:
                w.writerow({"row_count": 0})
        return
    headers = fieldnames or list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({h: _csv_cell(r.get(h)) for h in headers})


def _csv_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, default=str)
    return str(v)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=str) + "\n",
        encoding="utf-8",
    )


def event_to_row(e: EventDelta) -> dict[str, Any]:
    d = e.as_dict()
    d["previous_value"] = (
        json.dumps(d["previous_value"], ensure_ascii=False, default=str)
        if isinstance(d["previous_value"], (dict, list))
        else d["previous_value"]
    )
    d["current_value"] = (
        json.dumps(d["current_value"], ensure_ascii=False, default=str)
        if isinstance(d["current_value"], (dict, list))
        else d["current_value"]
    )
    return d


# ---------------------------------------------------------------------------
# WeeklyInput loader (fail-closed)
# ---------------------------------------------------------------------------


def load_weekly_input(path: Path, *, require_ok: bool = True) -> WeeklyInput:
    """Load and validate a weekly pack as WeeklyInput.

    Fail-closed: missing dir / invalid structure → raises SystemExit(2) when
    require_ok is True.
    """
    path = path.resolve()
    if not path.exists():
        raise SystemExit(EXIT_FAIL_CLOSED)
    if not path.is_dir():
        raise SystemExit(EXIT_FAIL_CLOSED)

    validation = validate_weekly_pack(path)
    wi = WeeklyInput(
        path=path,
        cycle_id=validation.cycle_id,
        collection_id=validation.collection_id,
        exit_code=validation.exit_code,
        validation_ok=validation.ok,
        validation_errors=list(validation.errors),
    )

    manifest_path = path / "manifest.json"
    checksums_path = path / "checksums.json"
    if manifest_path.is_file():
        try:
            wi.manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            wi.validation_errors.append("manifest.json JSON inválido")
            wi.validation_ok = False
    if checksums_path.is_file():
        try:
            wi.checksums = json.loads(checksums_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            wi.validation_errors.append("checksums.json JSON inválido")
            wi.validation_ok = False

    # Minimum structure for recurring: opportunities.csv + manifest + checksums
    opp_path = path / "opportunities.csv"
    if not opp_path.is_file():
        wi.validation_errors.append("opportunities.csv ausente")
        wi.validation_ok = False
    else:
        wi.opportunities = load_csv_rows(opp_path)

    for name, attr in (
        ("contracts.csv", "contracts"),
        ("competitors.csv", "competitors"),
        ("orgaos.csv", "orgaos"),
        ("source_health.csv", "source_health"),
    ):
        p = path / name
        if p.is_file():
            setattr(wi, attr, load_csv_rows(p))

    # Prefer source_health from CSV; fallback to manifest lists
    if not wi.source_health:
        wi.source_health = list(
            wi.manifest.get("source_health") or wi.manifest.get("freshness") or []
        )

    if require_ok and not wi.validation_ok:
        raise SystemExit(EXIT_FAIL_CLOSED)
    return wi


# ---------------------------------------------------------------------------
# Delta detection
# ---------------------------------------------------------------------------


def _index_opps(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for r in rows:
        eid = opp_id(r)
        if eid:
            out[eid] = r
    return out


def detect_tender_deltas(
    current: WeeklyInput,
    previous: WeeklyInput | None,
    *,
    as_of: date,
    detected_at: str,
) -> list[EventDelta]:
    events: list[EventDelta] = []
    cur = _index_opps(current.opportunities)
    prev = _index_opps(previous.opportunities) if previous else {}
    src = current.cycle_id
    prev_run = previous.cycle_id if previous else None

    if previous is None:
        return events  # FIRST_RUN — no NEW_TENDER noise (all would be "new")

    for eid, row in cur.items():
        if eid not in prev:
            events.append(
                EventDelta(
                    entity_type="tender",
                    entity_id=eid,
                    event_type="NEW_TENDER",
                    previous_value=None,
                    current_value={
                        "status": opp_status(row),
                        "deadline": opp_deadline(row),
                        "orgao": row.get("orgao_nome") or row.get("orgao"),
                        "objeto": (row.get("objeto") or "")[:200],
                    },
                    detected_at=detected_at,
                    source_run_id=src,
                    previous_run_id=prev_run,
                    official_url=opp_url(row),
                    severity=severity_for("NEW_TENDER"),
                    action_required=action_for("NEW_TENDER"),
                )
            )
            continue

        old = prev[eid]
        old_dl = opp_deadline(old)
        new_dl = opp_deadline(row)
        if old_dl != new_dl and (old_dl or new_dl):
            days_rem = None
            if new_dl:
                nd = parse_date(new_dl)
                if nd:
                    days_rem = (nd - as_of).days
            events.append(
                EventDelta(
                    entity_type="tender",
                    entity_id=eid,
                    event_type="DEADLINE_CHANGED",
                    previous_value=old_dl,
                    current_value=new_dl,
                    detected_at=detected_at,
                    source_run_id=src,
                    previous_run_id=prev_run,
                    official_url=opp_url(row),
                    severity=severity_for("DEADLINE_CHANGED", days_remaining=days_rem),
                    action_required=action_for("DEADLINE_CHANGED"),
                )
            )

        old_st = opp_status(old)
        new_st = opp_status(row)
        if old_st and new_st and old_st.upper() != new_st.upper():
            etype = classify_status_event(old_st, new_st)
            events.append(
                EventDelta(
                    entity_type="tender",
                    entity_id=eid,
                    event_type=etype,
                    previous_value=old_st,
                    current_value=new_st,
                    detected_at=detected_at,
                    source_run_id=src,
                    previous_run_id=prev_run,
                    official_url=opp_url(row),
                    severity=severity_for(etype),
                    action_required=action_for(etype),
                )
            )
    return events


def _contracts_as_window_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Normalize weekly contracts.csv to strategic_monthly_monitor shape."""
    out: list[dict[str, Any]] = []
    for r in rows:
        out.append(
            {
                "id": contract_id(r),
                "contract_id": contract_id(r),
                "orgao": r.get("orgao_nome") or r.get("orgao") or "",
                "contratado": r.get("fornecedor_nome") or r.get("contratado") or "",
                "fornecedor": r.get("fornecedor_nome") or "",
                "vigencia_fim": r.get("data_fim") or r.get("vigencia_fim") or r.get("termino"),
                "termino": r.get("data_fim") or r.get("vigencia_fim"),
            }
        )
    return out


def detect_contract_window_deltas(
    current: WeeklyInput,
    previous: WeeklyInput | None,
    *,
    as_of: date,
    expiry_window_days: int,
    detected_at: str,
) -> tuple[list[EventDelta], list[dict[str, Any]]]:
    """Contracts in [0, expiry_window_days] that newly entered the window."""
    cur_norm = _contracts_as_window_rows(current.contracts)
    cur_in = contracts_in_window(
        cur_norm, as_of=as_of, min_days=0, max_days=expiry_window_days
    )
    prev_ids: set[str] = set()
    if previous is not None:
        prev_norm = _contracts_as_window_rows(previous.contracts)
        prev_in = contracts_in_window(
            prev_norm, as_of=as_of, min_days=0, max_days=expiry_window_days
        )
        prev_ids = {str(x.get("contract_id") or "") for x in prev_in}

    events: list[EventDelta] = []
    expiring_rows: list[dict[str, Any]] = []
    src = current.cycle_id
    prev_run = previous.cycle_id if previous else None

    for row in cur_in:
        cid = str(row.get("contract_id") or "")
        expiring_rows.append(dict(row))
        if previous is None:
            continue  # FIRST_RUN: list expiring but no ENTERED delta spam
        if cid and cid not in prev_ids:
            days = row.get("days_to_end")
            days_i = int(days) if isinstance(days, (int, float)) else None
            events.append(
                EventDelta(
                    entity_type="contract",
                    entity_id=cid,
                    event_type="CONTRACT_ENTERED_EXPIRY_WINDOW",
                    previous_value="outside_window_or_absent",
                    current_value={
                        "vigencia_fim": row.get("vigencia_fim"),
                        "days_to_end": days,
                        "orgao": row.get("orgao"),
                        "contratado": row.get("contratado"),
                    },
                    detected_at=detected_at,
                    source_run_id=src,
                    previous_run_id=prev_run,
                    official_url=None,
                    severity=severity_for(
                        "CONTRACT_ENTERED_EXPIRY_WINDOW", days_remaining=days_i
                    ),
                    action_required=action_for("CONTRACT_ENTERED_EXPIRY_WINDOW"),
                )
            )
    return events, expiring_rows


def _winner_counts(contracts: list[dict[str, Any]], competitors: list[dict[str, Any]]) -> dict[str, float]:
    counts: dict[str, float] = {}
    if competitors:
        for r in competitors:
            k = winner_key(r)
            if not k:
                continue
            n = parse_float(r.get("n_contratos") or r.get("count")) or 0.0
            if n <= 0:
                n = 1.0
            counts[k] = counts.get(k, 0.0) + float(n)
        return counts
    for r in contracts:
        k = winner_key(r)
        if k:
            counts[k] = counts.get(k, 0.0) + 1.0
    return counts


def _concentration(counts: dict[str, float]) -> dict[str, Any]:
    total = sum(counts.values())
    if total <= 0:
        return {"total": 0, "top1_share": None, "hhi": None, "top5": []}
    ordered = sorted(counts.items(), key=lambda x: (-x[1], x[0]))
    top1 = ordered[0][1] / total if ordered else 0.0
    hhi = sum((v / total) ** 2 for v in counts.values())
    top5 = [{"winner": k, "count": v, "share": round(v / total, 4)} for k, v in ordered[:5]]
    return {
        "total": total,
        "top1_share": round(top1, 4),
        "hhi": round(hhi, 6),
        "top5": top5,
    }


def detect_winner_deltas(
    current: WeeklyInput,
    previous: WeeklyInput | None,
    *,
    detected_at: str,
    concentration_delta_threshold: float = 0.05,
) -> tuple[list[EventDelta], list[dict[str, Any]]]:
    events: list[EventDelta] = []
    orgaos_winners_rows: list[dict[str, Any]] = []
    if previous is None:
        return events, orgaos_winners_rows

    cur_w = _winner_counts(current.contracts, current.competitors)
    prev_w = _winner_counts(previous.contracts, previous.competitors)
    src = current.cycle_id
    prev_run = previous.cycle_id

    for w in cur_w:
        if w not in prev_w:
            events.append(
                EventDelta(
                    entity_type="winner",
                    entity_id=w,
                    event_type="NEW_WINNER",
                    previous_value=None,
                    current_value={"count": cur_w[w]},
                    detected_at=detected_at,
                    source_run_id=src,
                    previous_run_id=prev_run,
                    official_url=None,
                    severity=severity_for("NEW_WINNER"),
                    action_required=action_for("NEW_WINNER"),
                )
            )
            orgaos_winners_rows.append(
                {
                    "winner": w,
                    "change": "NEW",
                    "previous_count": 0,
                    "current_count": cur_w[w],
                }
            )

    cur_c = _concentration(cur_w)
    prev_c = _concentration(prev_w)
    cur_share = cur_c.get("top1_share")
    prev_share = prev_c.get("top1_share")
    if (
        isinstance(cur_share, (int, float))
        and isinstance(prev_share, (int, float))
        and abs(cur_share - prev_share) >= concentration_delta_threshold
    ):
        events.append(
            EventDelta(
                entity_type="market",
                entity_id="winner_concentration",
                event_type="WINNER_CONCENTRATION_CHANGED",
                previous_value=prev_c,
                current_value=cur_c,
                detected_at=detected_at,
                source_run_id=src,
                previous_run_id=prev_run,
                official_url=None,
                severity=severity_for("WINNER_CONCENTRATION_CHANGED"),
                action_required=action_for("WINNER_CONCENTRATION_CHANGED"),
            )
        )
        orgaos_winners_rows.append(
            {
                "winner": "(concentration)",
                "change": "CONCENTRATION",
                "previous_top1_share": prev_share,
                "current_top1_share": cur_share,
                "previous_hhi": prev_c.get("hhi"),
                "current_hhi": cur_c.get("hhi"),
            }
        )
    return events, orgaos_winners_rows


def detect_source_health_deltas(
    current: WeeklyInput,
    previous: WeeklyInput | None,
    *,
    detected_at: str,
) -> tuple[list[EventDelta], dict[str, Any]]:
    events: list[EventDelta] = []
    cur_rows = list(current.source_health or [])
    prev_by_src: dict[str, dict[str, Any]] = {}
    if previous:
        for r in previous.source_health or []:
            s = str(r.get("source") or "")
            if s:
                prev_by_src[s] = r

    health_payload: dict[str, Any] = {
        "sources": [],
        "degraded": [],
        "freshness_breaches": [],
        "first_run": previous is None,
    }
    src_run = current.cycle_id
    prev_run = previous.cycle_id if previous else None

    for r in cur_rows:
        source = str(r.get("source") or "unknown")
        level = str(r.get("level") or "unknown").lower()
        try:
            age = float(r["age_hours"]) if r.get("age_hours") not in (None, "") else None
        except (TypeError, ValueError):
            age = None
        try:
            sla = float(r["sla_hours"]) if r.get("sla_hours") not in (None, "") else None
        except (TypeError, ValueError):
            sla = None

        entry = {
            "source": source,
            "level": level,
            "age_hours": age,
            "sla_hours": sla,
            "last_status": r.get("last_status"),
        }
        health_payload["sources"].append(entry)

        breach = False
        if age is not None and sla is not None and age > sla:
            breach = True
        if level in {"stale", "never_crawled", "degraded"}:
            breach = True
        if breach:
            health_payload["freshness_breaches"].append(entry)
            events.append(
                EventDelta(
                    entity_type="source",
                    entity_id=source,
                    event_type="FRESHNESS_BREACH",
                    previous_value=prev_by_src.get(source),
                    current_value=entry,
                    detected_at=detected_at,
                    source_run_id=src_run,
                    previous_run_id=prev_run,
                    official_url=None,
                    severity=severity_for("FRESHNESS_BREACH"),
                    action_required=action_for("FRESHNESS_BREACH"),
                )
            )

        if previous and source in prev_by_src:
            prev_level = str(prev_by_src[source].get("level") or "unknown").lower()
            cur_rank = _FRESHNESS_RANK.get(level, 1)
            prev_rank = _FRESHNESS_RANK.get(prev_level, 1)
            if cur_rank < prev_rank:
                health_payload["degraded"].append(
                    {"source": source, "from": prev_level, "to": level}
                )
                events.append(
                    EventDelta(
                        entity_type="source",
                        entity_id=source,
                        event_type="SOURCE_DEGRADED",
                        previous_value=prev_level,
                        current_value=level,
                        detected_at=detected_at,
                        source_run_id=src_run,
                        previous_run_id=prev_run,
                        official_url=None,
                        severity=severity_for("SOURCE_DEGRADED"),
                        action_required=action_for("SOURCE_DEGRADED"),
                    )
                )

    return events, health_payload


def detect_all_deltas(
    current: WeeklyInput,
    previous: WeeklyInput | None,
    *,
    as_of: date,
    expiry_window_days: int,
) -> dict[str, Any]:
    detected_at = detected_at_for(as_of)
    first_run = previous is None

    tender_events = detect_tender_deltas(
        current, previous, as_of=as_of, detected_at=detected_at
    )
    contract_events, expiring = detect_contract_window_deltas(
        current,
        previous,
        as_of=as_of,
        expiry_window_days=expiry_window_days,
        detected_at=detected_at,
    )
    winner_events, orgaos_winners = detect_winner_deltas(
        current, previous, detected_at=detected_at
    )
    source_events, source_health = detect_source_health_deltas(
        current, previous, detected_at=detected_at
    )

    all_events = tender_events + contract_events + winner_events + source_events
    # stable order for idempotency
    all_events.sort(key=lambda e: (e.event_type, e.entity_type, e.entity_id))

    by_type: dict[str, int] = {t: 0 for t in sorted(ALLOWED_EVENT_TYPES)}
    for e in all_events:
        by_type[e.event_type] = by_type.get(e.event_type, 0) + 1

    status = "FIRST_RUN" if first_run else ("SUCCESS_ZERO" if not all_events else "OK")
    return {
        "status": status,
        "first_run": first_run,
        "events": all_events,
        "counts_by_type": by_type,
        "total_events": len(all_events),
        "expiring_contracts": expiring,
        "orgaos_winners_delta": orgaos_winners,
        "source_health": source_health,
        "detected_at": detected_at,
    }


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------


def metrics_from_weekly(wi: WeeklyInput, expiring_n: int, events_n: int) -> dict[str, Any]:
    return {
        "opportunities_total": len(wi.opportunities),
        "contracts_total": len(wi.contracts),
        "competitors_total": len(wi.competitors),
        "orgaos_total": len(wi.orgaos),
        "expiring_in_window": expiring_n,
        "events_total": events_n,
    }


def build_weekly_report_md(
    *,
    current: WeeklyInput,
    previous: WeeklyInput | None,
    delta: dict[str, Any],
    as_of: date,
    variation: dict[str, Any],
) -> str:
    events: list[EventDelta] = delta["events"]
    counts = delta["counts_by_type"]
    lines = [
        f"# Relatório semanal de recorrência — {current.cycle_id or 'unknown'}",
        "",
        f"- **Gerado (as_of):** {as_of.isoformat()}",
        f"- **Current run:** `{current.cycle_id}` (exit={current.exit_code})",
        f"- **Previous run:** `{previous.cycle_id if previous else '— (FIRST_RUN)'}`",
        f"- **Status deltas:** **{delta['status']}**",
        f"- **Total eventos:** {delta['total_events']}",
        "",
        "> Alertas urgentes são artefatos **separados** (`urgent-alerts.*`).",
        "> Este relatório consolidado existe mesmo com zero alertas.",
        "",
        "## Contagem por tipo de evento",
        "",
        "| event_type | count |",
        "|---|---:|",
    ]
    for t in sorted(ALLOWED_EVENT_TYPES):
        lines.append(f"| `{t}` | {counts.get(t, 0)} |")

    lines += ["", "## Variação de métricas", ""]
    if variation.get("has_previous"):
        fields = variation.get("fields") or {}
        lines.append("| métrica | previous | current | delta | delta_% |")
        lines.append("|---|---:|---:|---:|---:|")
        for k, v in sorted(fields.items()):
            if not isinstance(v, dict):
                continue
            lines.append(
                f"| `{k}` | {v.get('previous')} | {v.get('current')} | "
                f"{v.get('delta')} | {v.get('delta_pct')} |"
            )
    else:
        lines.append(f"- {variation.get('note') or 'Sem período anterior.'}")

    lines += ["", "## Eventos (até 40)", ""]
    if not events:
        if delta["first_run"]:
            lines.append("- FIRST_RUN: nenhum delta vs período anterior (baseline estabelecida).")
        else:
            lines.append("- SUCCESS_ZERO: runs idênticos em dimensões monitoradas — sem deltas.")
    else:
        lines.append("| event_type | entity | severity | previous → current |")
        lines.append("|---|---|---|---|")
        for e in events[:40]:
            lines.append(
                f"| `{e.event_type}` | `{e.entity_id[:48]}` | {e.severity} | "
                f"{_short(e.previous_value)} → {_short(e.current_value)} |"
            )

    lines += [
        "",
        "## Contratos na janela de expiração",
        "",
        f"- **n:** {len(delta.get('expiring_contracts') or [])}",
        "",
        "## Saúde das fontes",
        "",
    ]
    sh = delta.get("source_health") or {}
    for s in sh.get("sources") or []:
        lines.append(
            f"- `{s.get('source')}`: **{s.get('level')}** "
            f"(age_h={s.get('age_hours')}, sla={s.get('sla_hours')})"
        )
    if not sh.get("sources"):
        lines.append("- (sem source_health no pack)")

    lines += [
        "",
        "## Limitações",
        "",
        "- Deltas são diff de snapshots weekly (não stream em tempo real).",
        "- FIRST_RUN não emite NEW_TENDER para todo o universo (evita ruído).",
        "- Não reivindica cobertura 95%, VPS_OPERATIONAL ou PROJECT_DONE.",
        "",
    ]
    return "\n".join(lines)


def _short(v: Any, n: int = 40) -> str:
    if v is None:
        return "—"
    s = v if isinstance(v, str) else json.dumps(v, ensure_ascii=False, default=str)
    s = s.replace("|", "/").replace("\n", " ")
    return s if len(s) <= n else s[: n - 1] + "…"


def build_monthly_report_md(
    *,
    current: WeeklyInput,
    previous: WeeklyInput | None,
    previous_monthly: dict[str, Any] | None,
    delta: dict[str, Any],
    as_of: date,
    variation: dict[str, Any],
    monthly_comparison: dict[str, Any],
) -> str:
    month_start = as_of.replace(day=1).isoformat()
    lines = [
        f"# Relatório mensal comparativo — {as_of.strftime('%Y-%m')}",
        "",
        f"- **Período:** {month_start} → {as_of.isoformat()}",
        f"- **Current weekly:** `{current.cycle_id}`",
        f"- **Previous weekly:** `{previous.cycle_id if previous else '—'}`",
        f"- **Previous monthly artifact:** "
        f"{'yes' if previous_monthly else 'no'}",
        "",
        "## Resumo",
        "",
        f"- Oportunidades atuais: **{len(current.opportunities)}**",
        f"- Contratos atuais: **{len(current.contracts)}**",
        f"- Eventos na janela semanal: **{delta['total_events']}**",
        f"- Contratos em janela de expiração: **{len(delta.get('expiring_contracts') or [])}**",
        "",
        "## Comparativo",
        "",
    ]
    for k, v in (monthly_comparison.get("fields") or {}).items():
        if isinstance(v, dict):
            lines.append(
                f"- **{k}:** {v.get('previous')} → {v.get('current')} "
                f"(Δ={v.get('delta')}, %{v.get('delta_pct')})"
            )
    if not monthly_comparison.get("has_previous"):
        lines.append(f"- {monthly_comparison.get('note')}")

    lines += [
        "",
        "## Panorama de vencedores (top)",
        "",
    ]
    cur_w = _winner_counts(current.contracts, current.competitors)
    conc = _concentration(cur_w)
    for t in conc.get("top5") or []:
        lines.append(f"- `{t['winner']}`: n={t['count']} share={t['share']}")
    if not conc.get("top5"):
        lines.append("- (sem dados de vencedores)")

    lines += [
        "",
        "## Notas",
        "",
        "- Relatório consolidado mensal **não** substitui alertas urgentes.",
        f"- Variação semanal: has_previous={variation.get('has_previous')}.",
        "",
    ]
    return "\n".join(lines)


def build_meeting_support_md(
    *,
    current: WeeklyInput,
    delta: dict[str, Any],
    as_of: date,
    urgent: list[EventDelta],
) -> str:
    events: list[EventDelta] = delta["events"]
    lines = [
        f"# Apoio de reunião — recorrência Extra ({as_of.isoformat()})",
        "",
        f"**Weekly:** `{current.cycle_id}` · status deltas: **{delta['status']}**",
        "",
        "## Pauta sugerida (15–20 min)",
        "",
        "1. Saúde das fontes e freshness (bloqueia decisão se breach)",
        "2. Alertas urgentes da semana",
        "3. Novos editais e mudanças de prazo/status",
        "4. Contratos entrando na janela de expiração",
        "5. Mudanças de concorrência / concentração",
        "6. Decisões e donos para a próxima semana",
        "",
        f"## Alertas urgentes ({len(urgent)})",
        "",
    ]
    if not urgent:
        lines.append("- Nenhum alerta urgente. Prosseguir com relatório consolidado.")
    else:
        for e in urgent[:15]:
            lines.append(
                f"- **{e.severity.upper()}** `{e.event_type}` · `{e.entity_id[:60]}` — "
                f"{e.action_required}"
            )

    new_t = [e for e in events if e.event_type == "NEW_TENDER"]
    lines += ["", f"## Novos editais ({len(new_t)})", ""]
    if not new_t:
        lines.append("- Nenhum NEW_TENDER nesta comparação.")
    else:
        for e in new_t[:10]:
            lines.append(f"- `{e.entity_id}` — {_short(e.current_value, 80)}")

    exp = delta.get("expiring_contracts") or []
    lines += ["", f"## Contratos na janela ({len(exp)})", ""]
    for c in exp[:10]:
        lines.append(
            f"- `{c.get('contract_id')}` · {c.get('orgao')} · "
            f"fim={c.get('vigencia_fim')} · d={c.get('days_to_end')}"
        )
    if not exp:
        lines.append("- Nenhum contrato na janela configurada.")

    lines += [
        "",
        "## Perguntas para a Extra",
        "",
        "1. Quais órgãos priorizar entre os contratos a vencer?",
        "2. Há retificações/suspensões que mudam a carteira ativa?",
        "3. Alguma fonte degradada exige recoleta antes da próxima decisão?",
        "",
    ]
    return "\n".join(lines)


def build_weekly_xlsx(
    path: Path,
    *,
    events: list[EventDelta],
    expiring: list[dict[str, Any]],
    metrics: dict[str, Any],
    counts: dict[str, int],
) -> bool:
    try:
        from openpyxl import Workbook
    except ImportError:
        path.write_text("openpyxl unavailable\n", encoding="utf-8")
        return False

    wb = Workbook()
    meta = wb.active
    meta.title = "Metadados"
    meta.append(["key", "value"])
    for k, v in metrics.items():
        meta.append([k, v])
    meta.append(["schema", SCHEMA])

    ws_c = wb.create_sheet("Contagens")
    ws_c.append(["event_type", "count"])
    for t, n in sorted(counts.items()):
        ws_c.append([t, n])

    ws_e = wb.create_sheet("Eventos")
    headers = [
        "entity_type",
        "entity_id",
        "event_type",
        "previous_value",
        "current_value",
        "detected_at",
        "source_run_id",
        "previous_run_id",
        "official_url",
        "severity",
        "action_required",
    ]
    ws_e.append(headers)
    for e in events:
        row = event_to_row(e)
        ws_e.append([row.get(h) for h in headers])

    ws_x = wb.create_sheet("Expiring")
    if expiring:
        xh = list(expiring[0].keys())
        ws_x.append(xh)
        for r in expiring:
            ws_x.append([_csv_cell(r.get(h)) for h in xh])
    else:
        ws_x.append(["(vazio)"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return True


def select_urgent(events: list[EventDelta]) -> list[EventDelta]:
    out: list[EventDelta] = []
    for e in events:
        if e.event_type in URGENT_EVENT_TYPES or e.severity in {"high", "critical"}:
            out.append(e)
    return out


def load_previous_monthly(path: Path | None) -> dict[str, Any] | None:
    if path is None:
        return None
    path = path.resolve()
    if not path.exists():
        return None
    if path.is_file() and path.suffix == ".json":
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return None
    if path.is_dir():
        for name in (
            "monthly-comparison.json",
            "monthly-report.json",
            "monthly-monitor-report.json",
            "period_metrics.json",
        ):
            p = path / name
            if p.is_file():
                try:
                    return json.loads(p.read_text(encoding="utf-8"))
                except json.JSONDecodeError:
                    continue
        # metrics embedded in monthly-comparison
        mc = path / "monthly-comparison.json"
        if mc.is_file():
            try:
                return json.loads(mc.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                return None
    return None


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


def run_delivery(
    *,
    current_run: Path,
    delivery_out: Path,
    previous_run: Path | None = None,
    previous_monthly: Path | None = None,
    profile_path: Path | None = None,
    expiry_window_days: int = 180,
    as_of: date | None = None,
) -> dict[str, Any]:
    as_of = as_of or date.today()
    delivery_out = delivery_out.resolve()
    delivery_out.mkdir(parents=True, exist_ok=True)

    # Fail-closed current
    if not current_run or not Path(current_run).exists():
        raise SystemExit(EXIT_FAIL_CLOSED)
    current = load_weekly_input(Path(current_run), require_ok=True)
    # D1 parity: non-zero weekly exit is never a consultive recurring pack
    # (no SUCCESS_ZERO / OK that could be read as market absence).
    if current.exit_code not in (0, None):
        print(
            json.dumps(
                {
                    "ok": False,
                    "status": "BLOCKED_EXTERNAL",
                    "role": "current",
                    "cycle_id": current.cycle_id,
                    "exit_code": current.exit_code,
                    "error": (
                        f"current weekly exit_code={current.exit_code} is not consultive; "
                        "refuse recurring delivery package"
                    ),
                },
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        raise SystemExit(EXIT_BLOCKED)

    previous: WeeklyInput | None = None
    if previous_run is not None:
        prev_path = Path(previous_run)
        if prev_path.exists():
            # Previous may be slightly soft on checksums but must be a pack dir
            previous = load_weekly_input(prev_path, require_ok=True)
            if previous.exit_code not in (0, None):
                print(
                    json.dumps(
                        {
                            "ok": False,
                            "status": "BLOCKED_EXTERNAL",
                            "role": "previous",
                            "cycle_id": previous.cycle_id,
                            "exit_code": previous.exit_code,
                            "error": (
                                f"previous weekly exit_code={previous.exit_code} is not "
                                "comparable/consultive; refuse delta package"
                            ),
                        },
                        ensure_ascii=False,
                    ),
                    file=sys.stderr,
                )
                raise SystemExit(EXIT_BLOCKED)
        # Explicit previous path that does not exist is fail-closed (not silent FIRST_RUN).
        else:
            raise SystemExit(EXIT_FAIL_CLOSED)

    profile: dict[str, Any] = {}
    if profile_path and Path(profile_path).is_file():
        profile = load_profile(Path(profile_path))
    elif DEFAULT_PROFILE.is_file():
        profile = load_profile(DEFAULT_PROFILE)

    delta = detect_all_deltas(
        current,
        previous,
        as_of=as_of,
        expiry_window_days=expiry_window_days,
    )
    events: list[EventDelta] = delta["events"]
    urgent = select_urgent(events)

    cur_metrics = metrics_from_weekly(
        current, len(delta["expiring_contracts"]), delta["total_events"]
    )
    prev_metrics = (
        metrics_from_weekly(previous, 0, 0) if previous else None
    )
    # strip events_total from prev for cleaner variation (prev has no event count)
    if prev_metrics is not None:
        prev_metrics = {
            k: v for k, v in prev_metrics.items() if k != "events_total"
        }
        cur_for_var = {k: v for k, v in cur_metrics.items() if k != "events_total"}
    else:
        cur_for_var = {k: v for k, v in cur_metrics.items() if k != "events_total"}
    variation = compute_variation(cur_for_var, prev_metrics)

    prev_monthly = load_previous_monthly(previous_monthly)
    prev_monthly_metrics: dict[str, Any] | None = None
    if prev_monthly:
        prev_monthly_metrics = (
            prev_monthly.get("metrics")
            or prev_monthly.get("period_metrics")
            or prev_monthly.get("current_metrics")
            or (
                {
                    k: (v.get("current") if isinstance(v, dict) else v)
                    for k, v in (prev_monthly.get("fields") or {}).items()
                }
                if prev_monthly.get("fields")
                else None
            )
        )
    monthly_comparison = compute_variation(
        cur_for_var, prev_monthly_metrics if prev_monthly_metrics else prev_metrics
    )

    # --- Always write consolidated reports (even zero alerts) ---
    weekly_md = build_weekly_report_md(
        current=current,
        previous=previous,
        delta=delta,
        as_of=as_of,
        variation=variation,
    )
    (delivery_out / "weekly-report.md").write_text(weekly_md + "\n", encoding="utf-8")

    monthly_md = build_monthly_report_md(
        current=current,
        previous=previous,
        previous_monthly=prev_monthly,
        delta=delta,
        as_of=as_of,
        variation=variation,
        monthly_comparison=monthly_comparison,
    )
    (delivery_out / "monthly-report.md").write_text(monthly_md + "\n", encoding="utf-8")

    meeting_md = build_meeting_support_md(
        current=current, delta=delta, as_of=as_of, urgent=urgent
    )
    (delivery_out / "meeting-support.md").write_text(meeting_md + "\n", encoding="utf-8")

    event_rows = [event_to_row(e) for e in events]
    write_json(
        delivery_out / "weekly-delta.json",
        {
            "schema": f"{SCHEMA}/weekly-delta",
            "status": delta["status"],
            "first_run": delta["first_run"],
            "as_of": as_of.isoformat(),
            "current_run_id": current.cycle_id,
            "previous_run_id": previous.cycle_id if previous else None,
            "total_events": delta["total_events"],
            "counts_by_type": delta["counts_by_type"],
            "events": event_rows,
            "variation": variation,
        },
    )
    event_fields = [
        "entity_type",
        "entity_id",
        "event_type",
        "previous_value",
        "current_value",
        "detected_at",
        "source_run_id",
        "previous_run_id",
        "official_url",
        "severity",
        "action_required",
    ]
    write_csv(delivery_out / "weekly-delta.csv", event_rows, event_fields)
    write_csv(delivery_out / "tender-events.csv", event_rows, event_fields)
    write_csv(
        delivery_out / "expiring-contracts.csv",
        list(delta["expiring_contracts"]),
    )
    write_csv(
        delivery_out / "orgaos-winners-delta.csv",
        list(delta["orgaos_winners_delta"]),
    )

    # Urgent — separate artifacts; NEVER replace reports
    urgent_rows = [event_to_row(e) for e in urgent]
    write_json(
        delivery_out / "urgent-alerts.json",
        {
            "schema": f"{SCHEMA}/urgent-alerts",
            "count": len(urgent_rows),
            "as_of": as_of.isoformat(),
            "current_run_id": current.cycle_id,
            "alerts": urgent_rows,
            "note": "Separate from consolidated weekly/monthly reports.",
        },
    )
    write_csv(delivery_out / "urgent-alerts.csv", urgent_rows, event_fields)

    write_json(
        delivery_out / "monthly-comparison.json",
        {
            "schema": f"{SCHEMA}/monthly-comparison",
            "as_of": as_of.isoformat(),
            "current_run_id": current.cycle_id,
            "previous_run_id": previous.cycle_id if previous else None,
            "previous_monthly_loaded": prev_monthly is not None,
            "metrics_current": cur_for_var,
            "variation": monthly_comparison,
            "fields": monthly_comparison.get("fields") or {},
            "has_previous": monthly_comparison.get("has_previous"),
            "note": monthly_comparison.get("note"),
        },
    )
    write_json(delivery_out / "source-health.json", delta["source_health"])

    xlsx_ok = build_weekly_xlsx(
        delivery_out / "weekly-report.xlsx",
        events=events,
        expiring=list(delta["expiring_contracts"]),
        metrics={
            **cur_metrics,
            "as_of": as_of.isoformat(),
            "status": delta["status"],
            "current_run_id": current.cycle_id or "",
            "previous_run_id": (previous.cycle_id if previous else "") or "",
            "profile_id": profile.get("profile_id") or "",
        },
        counts=delta["counts_by_type"],
    )

    # Checksums of product artifacts (exclude manifest itself)
    product_names = [
        "weekly-report.md",
        "weekly-report.xlsx",
        "weekly-delta.json",
        "weekly-delta.csv",
        "tender-events.csv",
        "expiring-contracts.csv",
        "orgaos-winners-delta.csv",
        "urgent-alerts.json",
        "urgent-alerts.csv",
        "monthly-report.md",
        "monthly-comparison.json",
        "meeting-support.md",
        "source-health.json",
    ]
    checksums_artifacts: dict[str, Any] = {}
    for name in product_names:
        p = delivery_out / name
        if p.is_file():
            checksums_artifacts[name] = {
                "path": name,
                "sha256": sha256_file(p),
                "bytes": p.stat().st_size,
            }

    checksums_doc = {
        "schema": f"{SCHEMA}/checksums",
        "as_of": as_of.isoformat(),
        "current_run_id": current.cycle_id,
        "note": "Product artifacts only — does not include manifest.json",
        "artifacts": checksums_artifacts,
    }
    write_json(delivery_out / "checksums.json", checksums_doc)

    manifest = {
        "schema": SCHEMA,
        "generated_at": utc_now(),
        "as_of": as_of.isoformat(),
        "current_run": {
            "path": str(current.path),
            "cycle_id": current.cycle_id,
            "collection_id": current.collection_id,
            "exit_code": current.exit_code,
        },
        "previous_run": (
            {
                "path": str(previous.path),
                "cycle_id": previous.cycle_id,
                "collection_id": previous.collection_id,
                "exit_code": previous.exit_code,
            }
            if previous
            else None
        ),
        "status": delta["status"],
        "first_run": delta["first_run"],
        "total_events": delta["total_events"],
        "urgent_alerts": len(urgent),
        "counts_by_type": delta["counts_by_type"],
        "expiry_window_days": expiry_window_days,
        "xlsx_ok": xlsx_ok,
        "products": product_names + ["manifest.json", "checksums.json"],
        "claims_allowed": [
            "Weekly-over-weekly EventDelta on frozen allow-list",
            "Consolidated weekly + monthly reports independent of urgent alerts",
            "FIRST_RUN / SUCCESS_ZERO honest statuses",
        ],
        "claims_forbidden": [
            "Operational coverage 95%",
            "VPS_OPERATIONAL / PROJECT_DONE",
            "Real-time multi-channel notifications",
        ],
    }
    write_json(delivery_out / "manifest.json", manifest)

    # Fail-closed integrity: reports must exist even with zero alerts
    for required in (
        "weekly-report.md",
        "monthly-report.md",
        "urgent-alerts.json",
        "manifest.json",
        "checksums.json",
    ):
        if not (delivery_out / required).is_file():
            raise SystemExit(EXIT_FAIL_CLOSED)

    result = {
        "ok": True,
        "status": delta["status"],
        "delivery_out": str(delivery_out),
        "current_run_id": current.cycle_id,
        "previous_run_id": previous.cycle_id if previous else None,
        "total_events": delta["total_events"],
        "urgent_alerts": len(urgent),
        "first_run": delta["first_run"],
        "exit_code": EXIT_OK,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Extra recurring delivery — weekly deltas + reports (CLI-first)"
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    run_p = sub.add_parser("run", help="Compare weekly runs and write delivery pack")
    run_p.add_argument(
        "--current-run",
        required=True,
        type=Path,
        help="Dir do weekly pack atual (manifest+checksums+CSVs)",
    )
    run_p.add_argument(
        "--previous-run",
        type=Path,
        default=None,
        help="Dir do weekly pack anterior (opcional; ausente → FIRST_RUN)",
    )
    run_p.add_argument(
        "--previous-monthly",
        type=Path,
        default=None,
        help="Dir/arquivo do comparativo mensal anterior (opcional)",
    )
    run_p.add_argument(
        "--delivery-out",
        required=True,
        type=Path,
        help="Diretório externo de saída",
    )
    run_p.add_argument(
        "--profile",
        type=Path,
        default=DEFAULT_PROFILE,
        help="Perfil cliente YAML",
    )
    run_p.add_argument(
        "--expiry-window-days",
        type=int,
        default=180,
        help="Janela máxima (dias) para CONTRACT_ENTERED_EXPIRY_WINDOW (default 180)",
    )
    run_p.add_argument(
        "--as-of",
        type=str,
        default=None,
        help="Data de referência YYYY-MM-DD (default: hoje)",
    )

    val_p = sub.add_parser("validate-weekly", help="Valida WeeklyInput (pack)")
    val_p.add_argument("--current-run", required=True, type=Path)

    args = parser.parse_args(argv)

    if args.cmd == "validate-weekly":
        try:
            wi = load_weekly_input(args.current_run, require_ok=False)
        except SystemExit as e:
            return int(e.code or EXIT_FAIL_CLOSED)
        print(
            json.dumps(
                {
                    "ok": wi.validation_ok,
                    "cycle_id": wi.cycle_id,
                    "errors": wi.validation_errors,
                    "n_opportunities": len(wi.opportunities),
                    "n_contracts": len(wi.contracts),
                },
                indent=2,
                ensure_ascii=False,
            )
        )
        return EXIT_OK if wi.validation_ok else EXIT_FAIL_CLOSED

    if args.cmd == "run":
        as_of = date.fromisoformat(args.as_of) if args.as_of else date.today()
        try:
            run_delivery(
                current_run=args.current_run,
                delivery_out=args.delivery_out,
                previous_run=args.previous_run,
                previous_monthly=args.previous_monthly,
                profile_path=args.profile,
                expiry_window_days=args.expiry_window_days,
                as_of=as_of,
            )
        except SystemExit as e:
            code = int(e.code) if e.code is not None else EXIT_TECH
            return code
        except Exception as exc:  # noqa: BLE001
            print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
            return EXIT_TECH
        return EXIT_OK

    return EXIT_TECH


if __name__ == "__main__":
    raise SystemExit(main())
