#!/usr/bin/env python3
"""ARCH-RESET spikes E/G/H/J — local benchmarks (no production dependency changes).

Run with project venv or .spike-venv that has optional packages installed.
"""
from __future__ import annotations

import json
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "docs/ops/campaigns/ARCH-RESET-2026-07-20/spikes"


def _now() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")


def spike_e_dbt() -> dict[str, Any]:
    """dbt snapshots — honest rejection without experiment.

    The tiny synthetic event paths below are *illustrative limitations only*.
    They are NOT a dbt run, NOT SCD2 concordance, NOT a ≥200 corpus experiment.
    """
    illustrative_paths = [
        {"id": "T1", "events": ["open", "suspended", "reopened", "closed"]},
        {"id": "T2", "events": ["open", "revoked"]},
        {"id": "T3", "events": ["open", "deadline_change", "closed"]},
        {"id": "T4", "events": ["open", "hard_delete_from_source"]},
        {"id": "T5", "events": ["open", "no_updated_at_field"]},
    ]
    limitations = [
        "snapshot interval can miss intermediate statuses",
        "source without reliable updated_at cannot drive SCD correctly",
        "hard_delete requires dbt hard_deletes config and still loses legal event time",
        "second transform line risks dual truth vs operational migrations",
        "juridical effective time != snapshot observed time (bitemporality gap)",
    ]
    return {
        "spike": "E",
        "component": "dbt-core snapshots",
        "decision": "REJECTED_WITHOUT_EXPERIMENT",
        "honest": True,
        "experiment_run": False,
        "dbt_installed": False,
        "corpus_opportunities": 0,
        "min_corpus_required": 200,
        "reason": (
            "No isolated dbt project was executed against a ≥200 opportunity temporal corpus. "
            "illustrative_status_paths are design notes only — not experimental evidence. "
            "Dual-truth and temporal limitations remain reasons to avoid naive production adoption."
        ),
        "illustrative_status_paths": illustrative_paths,
        "limitations": limitations,
        "production_dep_added": False,
        "optional_later": "REFERENCE_ONLY for analytics warehouse if product needs SCD later",
        "generated_at_utc": _now(),
    }


def spike_g_parsers(tmpdir: Path) -> dict[str, Any]:
    """Benchmark available PDF readers on synthetic digital PDFs."""
    tmpdir.mkdir(parents=True, exist_ok=True)
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        return {"spike": "G", "decision": "BLOCKED", "error": f"reportlab missing: {exc}"}

    samples: list[Path] = []
    for i, text in enumerate(
        [
            "Edital simples valor R$ 1.250.000,50 prazo 2026-08-01",
            "Multi\ncoluna\nlinha1\tlinha2\nvalor 99000",
            "Tabela\nItem | Qtd | Valor\n1 | 10 | 1000",
        ],
        start=1,
    ):
        p = tmpdir / f"sample-{i}.pdf"
        c = canvas.Canvas(str(p), pagesize=A4)
        y = 800
        for line in text.split("\n"):
            c.drawString(72, y, line[:100])
            y -= 16
        c.save()
        samples.append(p)

    results: dict[str, Any] = {}
    # pypdf
    try:
        from pypdf import PdfReader

        t0 = time.perf_counter()
        texts = []
        for p in samples:
            r = PdfReader(str(p))
            texts.append("\n".join(page.extract_text() or "" for page in r.pages))
        results["pypdf"] = {
            "ok": True,
            "seconds": round(time.perf_counter() - t0, 4),
            "money_hits": sum("1.250.000" in t or "1250000" in t or "1.250.000,50" in t for t in texts),
            "date_hits": sum("2026-08-01" in t for t in texts),
            "pages": sum(len(PdfReader(str(p)).pages) for p in samples),
        }
    except Exception as exc:  # noqa: BLE001
        results["pypdf"] = {"ok": False, "error": str(exc)}

    # pdfplumber
    try:
        import pdfplumber

        t0 = time.perf_counter()
        texts = []
        for p in samples:
            with pdfplumber.open(str(p)) as pdf:
                texts.append("\n".join((page.extract_text() or "") for page in pdf.pages))
        results["pdfplumber"] = {
            "ok": True,
            "seconds": round(time.perf_counter() - t0, 4),
            "money_hits": sum("1.250.000" in t or "1.250.000,50" in t for t in texts),
            "date_hits": sum("2026-08-01" in t for t in texts),
        }
    except Exception as exc:  # noqa: BLE001
        results["pdfplumber"] = {"ok": False, "error": str(exc)}

    # fitz/PyMuPDF — AGPL gate
    try:
        import fitz  # type: ignore

        results["pymupdf_fitz"] = {
            "ok": True,
            "installed": True,
            "license_gate": "AGPL_OR_COMMERCIAL — production adoption blocked without ADR license accept",
            "note": "Present in some envs via scripts/intel_extract_docs.py; not approved for production expansion",
        }
        _ = fitz  # silence lint
    except Exception:
        results["pymupdf_fitz"] = {
            "ok": False,
            "installed": False,
            "license_gate": "AGPL_OR_COMMERCIAL — spike may evaluate offline only with explicit license decision",
        }

    return {
        "spike": "G",
        "component": "document_parsers",
        "decision": "DEFERRED_NO_CORPUS",
        "honest": True,
        "reason": (
            "KEEP_CURRENT_STACK is NOT proven by 3 synthetic digital PDFs. Required strata "
            "≥5 simple + ≥5 multicolumn + ≥5 tables + ≥5 scanned before engine adoption. "
            "This microbench is exploratory only. PyMuPDF blocked by AGPL without license ADR. "
            "No production dep change."
        ),
        "required_strata": {
            "digital_simple": 5,
            "multicolumn": 5,
            "tables": 5,
            "scanned": 5,
        },
        "corpus_counts": {
            "digital_simple": len(samples),
            "multicolumn": 0,
            "tables": 0,
            "scanned": 0,
        },
        "samples": [str(p.name) for p in samples],
        "engines": results,
        "production_dep_added": False,
        "generated_at_utc": _now(),
    }


def spike_h_identity() -> dict[str, Any]:
    """Entity resolution residual — deterministic first; Splink deferred."""
    from scripts.entity_identity.pncp_orgao_resolve import digits, pick_match

    # Deterministic root guard
    hits = [
        {"cnpj": "12345678000199", "razaoSocial": "PREFEITURA MUNICIPAL ALFA"},
        {"cnpj": "99999999000100", "razaoSocial": "PREFEITURA MUNICIPAL ALFA"},  # different root
    ]
    pick = pick_match("12345678", "PREFEITURA MUNICIPAL ALFA", hits)
    if pick is None:
        raise RuntimeError("expected deterministic pick_match for matching CNPJ-8 root")
    cnpj14, method, _ = pick
    if not cnpj14.startswith("12345678"):
        raise RuntimeError("pick_match returned wrong CNPJ root")
    if digits("99999999000100")[:8] == "12345678":
        raise RuntimeError("test fixture broken")

    return {
        "spike": "H",
        "component": "splink_residual",
        "decision": "REJECTED_SPIKE_FOR_NOW",
        "reason": (
            "Deterministic CNPJ-8/14 path already hard-fails conflicting roots "
            "(pick_match). No 300-pair gold residual corpus available to validate "
            "Splink auto-link ≥99% precision. Probabilistic merge must not write "
            "canonical entity without human review."
        ),
        "deterministic_proof": {
            "method": method,
            "cnpj14": cnpj14,
            "rejects_conflicting_root": True,
        },
        "splink": {
            "status": "REFERENCE_ONLY",
            "min_corpus_pairs": 300,
            "auto_link_precision_min": 0.99,
            "auto_write_canonical": False,
        },
        "production_dep_added": False,
        "generated_at_utc": _now(),
    }


def spike_j_reporting(tmpdir: Path) -> dict[str, Any]:
    """Compare openpyxl vs XlsxWriter write-only throughput on synthetic rows."""
    tmpdir.mkdir(parents=True, exist_ok=True)
    rows = [{"id": i, "objeto": f"Obra {i}", "valor": float(i) * 1000.5} for i in range(1, 2001)]
    out: dict[str, Any] = {"spike": "J", "n_rows": len(rows)}

    # openpyxl
    try:
        from openpyxl import Workbook

        path = tmpdir / "openpyxl.xlsx"
        t0 = time.perf_counter()
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("dados")
        ws.append(["id", "objeto", "valor"])
        for r in rows:
            ws.append([r["id"], r["objeto"], r["valor"]])
        wb.save(path)
        out["openpyxl"] = {
            "ok": True,
            "seconds": round(time.perf_counter() - t0, 4),
            "bytes": path.stat().st_size,
            "write_only": True,
        }
    except Exception as exc:  # noqa: BLE001
        out["openpyxl"] = {"ok": False, "error": str(exc)}

    # XlsxWriter
    try:
        import xlsxwriter

        path = tmpdir / "xlsxwriter.xlsx"
        t0 = time.perf_counter()
        wb = xlsxwriter.Workbook(str(path), {"constant_memory": True})
        ws = wb.add_worksheet("dados")
        ws.write_row(0, 0, ["id", "objeto", "valor"])
        for i, r in enumerate(rows, start=1):
            ws.write_row(i, 0, [r["id"], r["objeto"], r["valor"]])
        wb.close()
        out["xlsxwriter"] = {
            "ok": True,
            "seconds": round(time.perf_counter() - t0, 4),
            "bytes": path.stat().st_size,
            "constant_memory": True,
            "limitation": "cannot edit existing workbooks",
        }
    except Exception as exc:  # noqa: BLE001
        out["xlsxwriter"] = {"ok": False, "error": str(exc)}

    # PDF engines present
    try:
        import reportlab  # noqa: F401

        out["reportlab"] = {"installed": True, "role": "current_pdf_generator"}
    except ImportError:
        out["reportlab"] = {"installed": False}
    try:
        import fpdf  # noqa: F401

        out["fpdf2"] = {"installed": True}
    except ImportError:
        out["fpdf2"] = {
            "installed": False,
            "decision_hint": "no adoption without visual/regression parity on real pack",
        }

    ox = out.get("openpyxl") or {}
    xw = out.get("xlsxwriter") or {}
    if ox.get("ok") and xw.get("ok"):
        # Prefer keep openpyxl if already integrated and can read/edit; XlsxWriter only if faster enough and write-only OK
        faster = "xlsxwriter" if xw["seconds"] < ox["seconds"] else "openpyxl"
        out["decision"] = "KEEP_OPENPYXL_REPORTLAB"
        out["reason"] = (
            f"XlsxWriter may be faster for pure write ({faster} won microbench) but openpyxl is already "
            "integrated for read/edit paths; fpdf2 not installed and would require full PDF regression. "
            "No production dep change without pack-level parity."
        )
    else:
        out["decision"] = "KEEP_CURRENT_STACK"
        out["reason"] = "Incomplete optional engines; retain openpyxl+reportlab."
    out["production_dep_added"] = False
    out["generated_at_utc"] = _now()
    return out


def main() -> int:
    base = OUT
    base.mkdir(parents=True, exist_ok=True)
    tmp = base / "_tmp_bench"
    tmp.mkdir(exist_ok=True)

    reports = {
        "E_dbt": spike_e_dbt(),
        "G_parsers": spike_g_parsers(tmp / "pdfs"),
        "H_identity": spike_h_identity(),
        "J_reporting": spike_j_reporting(tmp / "xlsx"),
    }
    for key, rep in reports.items():
        folder = {
            "E_dbt": "DBT",
            "G_parsers": "DOCUMENTS",
            "H_identity": "IDENTITY",
            "J_reporting": "REPORTING",
        }[key]
        d = base / folder
        d.mkdir(parents=True, exist_ok=True)
        (d / "benchmark.json").write_text(json.dumps(rep, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        (d / "DECISION.md").write_text(
            f"# Spike {rep.get('spike')} — {rep.get('component')}\n\n"
            f"**Decision:** `{rep.get('decision')}`\n\n"
            f"**Reason:** {rep.get('reason')}\n\n"
            f"**Production dependency added:** {rep.get('production_dep_added')}\n\n"
            f"Evidence: `benchmark.json`\n",
            encoding="utf-8",
        )
    summary = {
        "campaign": "ARCH-RESET-2026-07-20",
        "generated_at_utc": _now(),
        "decisions": {k: v.get("decision") for k, v in reports.items()},
        "production_deps_added": False,
    }
    (base / "EGHJ-SUMMARY.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
