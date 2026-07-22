#!/usr/bin/env python3
"""Golden Path — Pipeline de validacao completa do DataLake Extra Consultoria.

Executa a sequencia completa e idempotente:
  1. Verifica conectividade com PostgreSQL
  2. Aplica migrations versionadas (db/migrations via scripts.ops.apply_migrations)
  3. Aplica seeds determinísticos (db/seed/001_sc_entities + 002_entity_aliases)
  4. Importa/valida a planilha-alvo canônica (Extra alvos de licitação)
  5. Crawl das fontes prioritarias (pncp, pcp, compras_gov) com
     timeout, retry 3x e backoff exponencial + jitter
  6. Valida freshness gate
  7. Gera relatorios (PDF executivo + Excel rastreavel)

Cada etapa e registrada em um ledger JSON para rastreabilidade.

Modo canônico (default: strict fail-closed):
  - Fonte essencial falhou ou vazia sem success_zero → exit != 0
  - Freshness gate FAIL (se não skip) → exit != 0
  - Relatório Excel/PDF FAIL (se não skip) → exit != 0
  - Zero fontes com dados e zero success_zero → exit != 0

Statuses de run: success | success_zero | partial | degraded | empty | failed

Usage:
    python scripts/golden_path.py
    python scripts/golden_path.py --sources pncp,pcp
    python scripts/golden_path.py --skip-reports
    python scripts/golden_path.py --no-strict   # legado permissivo (não canônico)
    python scripts/golden_path.py --verbose
    python scripts/golden_path.py --ledger-output custom_ledger.json

Exit codes:
    0 — success ou success_zero (todos gates obrigatórios OK)
    1 — failed / empty (sem dados utilizáveis)
    2 — partial (fontes essenciais falharam)
    3 — freshness gate reprovado (strict)
    4 — relatório obrigatório falhou (strict)
    5 — degraded (gates mistos / não essencial)
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import random
import subprocess
import sys
import time
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Rich console (optional dep)
# ---------------------------------------------------------------------------
try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table

    HAS_RICH = True
    _console = Console()
except ImportError:
    HAS_RICH = False
    _console = None

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
_SCRIPTS_DIR = _PROJECT_ROOT / "scripts"
_OUTPUT_DIR = _PROJECT_ROOT / "output"
_GOLDEN_PATH_DIR = _OUTPUT_DIR / "golden-path"
_GOLDEN_PATH_DIR.mkdir(parents=True, exist_ok=True)
_LEDGER_PATH = _GOLDEN_PATH_DIR / "ledger.json"

# ---------------------------------------------------------------------------
# Source definitions
# ---------------------------------------------------------------------------


@dataclass
class SourceDef:
    name: str
    essential: bool
    description: str
    timeout_s: int = 120
    max_retries: int = 3


SOURCES: list[SourceDef] = [
    SourceDef(
        name="pncp",
        essential=True,
        description="PNCP API (editais abertos — fonte crítica)",
        # Multi-day window crawl; 7d incremental can exceed 6 min under load.
        timeout_s=720,
        max_retries=2,
    ),
    SourceDef(
        name="pcp",
        essential=True,
        description="PCP (Portal de Compras Publicas)",
    ),
    SourceDef(
        name="compras_gov",
        essential=True,
        description="ComprasGov (compras federais SC)",
    ),
]

ESSENTIAL_SOURCE_NAMES: tuple[str, ...] = tuple(s.name for s in SOURCES if s.essential)


def essential_sources() -> list[SourceDef]:
    """Return the minimum essential source adapters for the golden path."""
    return [s for s in SOURCES if s.essential]


def assert_essential_sources_executed(
    source_records: list[SourceRecord],
    *,
    require_success: bool = False,
) -> tuple[bool, dict[str, Any]]:
    """Verify that every essential source was executed (not merely configured).

    Execution proof = a SourceRecord exists with attempts >= 1 and status in
    {success, success_zero, fail}. "fail" still proves the adapter ran;
    configuration-list smoke tests do not.
    """
    by_name = {r.name: r for r in source_records}
    missing: list[str] = []
    not_executed: list[str] = []
    executed: dict[str, dict[str, Any]] = {}
    for name in ESSENTIAL_SOURCE_NAMES:
        rec = by_name.get(name)
        if rec is None:
            missing.append(name)
            continue
        attempts = int(rec.attempts or 0)
        if attempts < 1:
            not_executed.append(name)
            continue
        if require_success and rec.status not in {"success", "success_zero"}:
            not_executed.append(name)
            continue
        executed[name] = {
            "status": rec.status,
            "attempts": attempts,
            "duration_ms": rec.duration_ms,
            "metrics": rec.metrics or {},
            "error": rec.error,
        }
    ok = not missing and not not_executed and len(executed) == len(ESSENTIAL_SOURCE_NAMES)
    return ok, {
        "essential": list(ESSENTIAL_SOURCE_NAMES),
        "executed": executed,
        "missing": missing,
        "not_executed": not_executed,
        "require_success": require_success,
    }


def assert_sources_persisted(
    source_records: list[SourceRecord],
    *,
    min_persisted_rows: int = 1,
) -> tuple[bool, dict[str, Any]]:
    """Prove crawl wrote data (inserted/updated/persisted metrics), not mere fetch.

    Counts rows from SourceRecord.metrics keys commonly emitted by monitor.py:
    inserted, updated, persisted. At least one source must show total writes
    >= min_persisted_rows.
    """
    per_source: dict[str, dict[str, Any]] = {}
    total_writes = 0
    writers: list[str] = []
    for rec in source_records:
        m = rec.metrics or {}
        inserted = int(m.get("inserted") or 0)
        updated = int(m.get("updated") or 0)
        persisted = int(m.get("persisted") or 0)
        writes = inserted + updated + persisted
        per_source[rec.name] = {
            "inserted": inserted,
            "updated": updated,
            "persisted": persisted,
            "writes": writes,
            "status": rec.status,
        }
        total_writes += writes
        if writes > 0:
            writers.append(rec.name)
    ok = total_writes >= min_persisted_rows and len(writers) >= 1
    return ok, {
        "min_persisted_rows": min_persisted_rows,
        "total_writes": total_writes,
        "writers": writers,
        "per_source": per_source,
    }


def assert_freshness_gate_executed(
    freshness: FreshnessRecord | None,
) -> tuple[bool, dict[str, Any]]:
    """Prove the freshness gate was executed (pass or fail), not skipped/absent.

    Execution proof requires status in {pass, fail} and structured details
    (JSON gate output). "fail" still proves the gate ran — a separate DOD item
    covers pass/SLA.
    """
    if freshness is None:
        return False, {"error": "no freshness record", "status": None}
    status = str(freshness.status or "")
    details = freshness.details if isinstance(freshness.details, dict) else {}
    has_structure = bool(details.get("overall") or details.get("critical_sources"))
    ok = status in {"pass", "fail"} and has_structure
    failing = []
    if isinstance(details.get("overall"), dict):
        failing = list(details["overall"].get("failing_sources") or [])
    return ok, {
        "status": status,
        "has_structure": has_structure,
        "failing_sources": failing,
        "error": freshness.error,
        "detail_keys": sorted(details.keys()) if details else [],
    }


def _backoff_delay(attempt: int, base: float = 4.0, max_delay: float = 60.0) -> float:
    """Exponential backoff with jitter."""
    delay = min(base * (2**attempt), max_delay)
    jitter = random.uniform(0, delay * 0.15)  # noqa: S311
    return delay + jitter


# ---------------------------------------------------------------------------
# Ledger types
# ---------------------------------------------------------------------------


@dataclass
class StepRecord:
    step: str
    status: str  # pass | fail | skipped
    duration_ms: float
    error: str | None = None
    details: dict[str, Any] | None = None


@dataclass
class SourceRecord:
    name: str
    status: str  # success | success_zero | fail | skipped
    duration_ms: float
    attempts: int
    metrics: dict[str, int] | None = None
    error: str | None = None


@dataclass
class ReportRecord:
    type: str  # excel | pdf
    status: str  # generated | skipped | fail
    path: str | None = None
    error: str | None = None


@dataclass
class FreshnessRecord:
    status: str  # pass | fail | skipped
    details: dict[str, Any] | None = None
    error: str | None = None


@dataclass
class RunRecord:
    run_id: str
    timestamp: str
    status: str  # success | success_zero | partial | degraded | empty | failed
    wall_clock_ms: float
    steps: list[StepRecord] = field(default_factory=list)
    sources: list[SourceRecord] = field(default_factory=list)
    reports: list[ReportRecord] = field(default_factory=list)
    freshness: FreshnessRecord | None = None
    # DoD §12.1 — always populated by _save_final_ledger
    meta: dict[str, Any] | None = None


def collect_run_metadata(*, dsn: str | None = None) -> dict[str, Any]:
    """Metadata for DoD §12.1 golden-path auditability (pure-ish, fail-soft)."""
    now = datetime.now(UTC)
    git_sha = "unknown"
    try:
        r = subprocess.run(
            ["git", "rev-parse", "HEAD"],  # noqa: S607
            capture_output=True,
            text=True,
            timeout=5,
            check=False,
            cwd=str(_PROJECT_ROOT),
        )
        if r.returncode == 0 and r.stdout.strip():
            git_sha = r.stdout.strip()[:40]
    except Exception as exc:
        logging.getLogger(__name__).debug("git_sha unavailable: %s", exc)
    mig_root = _PROJECT_ROOT / "db" / "migrations"
    mig_count = len(list(mig_root.glob("*.sql"))) if mig_root.is_dir() else 0
    return {
        "canonical_command": "python3 -m scripts.golden_path",
        "as_of": now.isoformat().replace("+00:00", "Z"),
        "reference_period": {
            "as_of": now.isoformat().replace("+00:00", "Z"),
            "window": "run-scoped",
        },
        "git_sha": git_sha,
        "dsn_present": bool(dsn),
        "migration_files_count": mig_count,
        "schema_version": f"migrations_count={mig_count}",
        "limitations": [
            "Metadata is local/run-scoped; does not prove VPS or LOCAL_READY.",
            "Coverage and freshness require a reachable DSN and applied migrations.",
        ],
    }


def run_coverage_calculation(
    dsn: str,
    *,
    project_root: Path | None = None,
    expected_denominator: int | None = None,
    capabilities: list[str] | None = None,
    require_gate: bool = False,
    output_dir: Path | None = None,
) -> StepRecord:
    """Calculate dual capability monitoring coverage (canonical).

    Computes independent metrics for open_tenders and historical_contracts using
    ``scripts.coverage.dual_capability_coverage``. Does **not** use
    ``entity_coverage.is_covered`` or ``any_row`` as coverage methods.

    Measurement can pass with low coverage. When ``require_gate`` is True, status
    fails if either capability gate (<95%) fails. Always records
    measurement_success vs coverage_gate_pass in details.
    """
    t0 = time.perf_counter()
    root = project_root or _PROJECT_ROOT
    details: dict[str, Any] = {
        "denominator": None,
        "numerator": None,
        "coverage_pct": None,
        "method": "dual_capability_coverage",
        "measurement_success": False,
        "coverage_gate_pass": False,
        "pipeline_success": False,
        "capabilities": {},
    }
    try:
        from scripts.coverage.dual_capability_coverage import (
            CAPABILITIES,
            FORBIDDEN_METHODS,
            compute_dual_coverage,
            write_reports,
        )
        from scripts.lib.universe import CANONICAL_UNIVERSE

        if expected_denominator is None:
            expected_denominator = int(CANONICAL_UNIVERSE)
        caps = list(capabilities) if capabilities else list(CAPABILITIES)
        report = compute_dual_coverage(
            dsn=dsn,
            project_root=root,
            capabilities=caps,
            expected_denominator=expected_denominator,
        )
        out = output_dir or (_OUTPUT_DIR / "coverage")
        if report.measurement_success:
            paths = write_reports(report, Path(out), capabilities=caps)
            details["artifact_paths"] = {k: str(v) for k, v in paths.items()}

        details["measurement_success"] = report.measurement_success
        details["coverage_gate_pass"] = report.coverage_gate_pass
        details["pipeline_success"] = report.pipeline_success
        details["universe"] = report.universe.to_dict()
        details["legacy_metric"] = report.legacy_metric
        details["limitations"] = report.limitations
        details["forbidden_methods"] = sorted(FORBIDDEN_METHODS)
        details["as_of"] = report.as_of
        details["error"] = report.error

        # Populate dual capability blocks
        for cap, res in report.capabilities.items():
            details["capabilities"][cap] = res.to_summary_dict()

        # Transition mirrors (open_tenders primary for backward-compatible den/num/pct fields)
        primary = report.capabilities.get("open_tenders") or next(iter(report.capabilities.values()), None)
        if primary is not None:
            details["denominator"] = primary.applicable_denominator
            details["numerator"] = primary.covered_numerator
            details["coverage_pct"] = primary.coverage_pct
            details["seed_sha256"] = report.universe.seed_sha256
            details["canonical_ids_sha256"] = report.universe.canonical_ids_sha256
            details["expected_denominator"] = expected_denominator

        if not report.measurement_success:
            return StepRecord(
                step="coverage_calculation",
                status="fail",
                duration_ms=(time.perf_counter() - t0) * 1000,
                error=report.error or "dual coverage measurement failed",
                details=details,
            )

        if require_gate and not report.coverage_gate_pass:
            return StepRecord(
                step="coverage_calculation",
                status="fail",
                duration_ms=(time.perf_counter() - t0) * 1000,
                error="coverage_gate_pass=false (measurement_success=true)",
                details=details,
            )

        # Calculation itself succeeds even if coverage is low — "calcula" not "atinge 95%"
        return StepRecord(
            step="coverage_calculation",
            status="pass",
            duration_ms=(time.perf_counter() - t0) * 1000,
            details=details,
        )
    except Exception as exc:
        details["error"] = str(exc)[:300]
        return StepRecord(
            step="coverage_calculation",
            status="fail",
            duration_ms=(time.perf_counter() - t0) * 1000,
            error=str(exc)[:300],
            details=details,
        )


def run_snapshot_reconciliation(
    dsn: str,
    *,
    project_root: Path | None = None,
    snapshot_dir: Path | None = None,
) -> StepRecord:
    """Reconcile current editais snapshot (pncp_raw_bids) against previous snapshot file.

    Produces structured delta: added / removed / changed (by content_hash), counts,
    and sha256 of current ID set. First run creates baseline (pass with baseline=true).
    Fail-closed on DB errors or empty current set when table exists with schema.
    """
    t0 = time.perf_counter()
    _ = project_root or _PROJECT_ROOT  # reserved for future seed-scoped snapshots
    out_dir = snapshot_dir or (_OUTPUT_DIR / "golden-path" / "snapshots")
    out_dir.mkdir(parents=True, exist_ok=True)
    prev_path = out_dir / "editais-snapshot-prev.json"
    curr_path = out_dir / "editais-snapshot-curr.json"
    details: dict[str, Any] = {
        "source_table": "pncp_raw_bids",
        "prev_path": str(prev_path),
        "curr_path": str(curr_path),
    }
    try:
        import psycopg2

        conn = psycopg2.connect(dsn, connect_timeout=10)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT pncp_id, COALESCE(content_hash, ''),
                           COALESCE(to_char(data_publicacao AT TIME ZONE 'UTC', 'YYYY-MM-DD'), '')
                    FROM pncp_raw_bids
                    WHERE is_active IS TRUE OR is_active IS NULL
                    """
                )
                rows = cur.fetchall()
        finally:
            conn.close()

        current: dict[str, dict[str, str]] = {}
        for pncp_id, chash, pub in rows:
            pid = str(pncp_id)
            current[pid] = {"content_hash": str(chash or ""), "data_publicacao": str(pub or "")}

        ids_sorted = sorted(current.keys())
        ids_payload = "\n".join(ids_sorted).encode("utf-8")
        ids_sha = hashlib.sha256(ids_payload).hexdigest()
        details["current_count"] = len(current)
        details["ids_sha256"] = ids_sha
        if len(current) == 0:
            details["error"] = "pncp_raw_bids returned zero active rows; cannot reconcile editais snapshot"
            return StepRecord(
                step="snapshot_reconciliation",
                status="fail",
                duration_ms=(time.perf_counter() - t0) * 1000,
                error=details["error"],
                details=details,
            )

        curr_doc = {
            "table": "pncp_raw_bids",
            "count": len(current),
            "ids_sha256": ids_sha,
            "records": current,
            "as_of": datetime.now(UTC).isoformat(),
        }
        curr_path.write_text(json.dumps(curr_doc, ensure_ascii=False, sort_keys=True) + "\n", encoding="utf-8")

        if not prev_path.is_file():
            # Baseline: copy current → prev
            prev_path.write_text(curr_path.read_text(encoding="utf-8"), encoding="utf-8")
            details["baseline"] = True
            details["added"] = 0
            details["removed"] = 0
            details["changed"] = 0
            details["unchanged"] = len(current)
            return StepRecord(
                step="snapshot_reconciliation",
                status="pass",
                duration_ms=(time.perf_counter() - t0) * 1000,
                details=details,
            )

        prev_doc = json.loads(prev_path.read_text(encoding="utf-8"))
        prev_records = prev_doc.get("records") or {}
        prev_ids = set(prev_records.keys())
        curr_ids = set(current.keys())
        added = sorted(curr_ids - prev_ids)
        removed = sorted(prev_ids - curr_ids)
        common = curr_ids & prev_ids
        changed = sorted(
            i
            for i in common
            if (prev_records.get(i) or {}).get("content_hash") != (current.get(i) or {}).get("content_hash")
        )
        details["baseline"] = False
        details["previous_count"] = len(prev_ids)
        details["added"] = len(added)
        details["removed"] = len(removed)
        details["changed"] = len(changed)
        details["unchanged"] = len(common) - len(changed)
        details["added_sample"] = added[:10]
        details["removed_sample"] = removed[:10]
        details["changed_sample"] = changed[:10]
        details["previous_ids_sha256"] = prev_doc.get("ids_sha256")

        # Rotate: curr becomes next prev
        prev_path.write_text(curr_path.read_text(encoding="utf-8"), encoding="utf-8")

        return StepRecord(
            step="snapshot_reconciliation",
            status="pass",
            duration_ms=(time.perf_counter() - t0) * 1000,
            details=details,
        )
    except Exception as exc:
        details["error"] = str(exc)[:400]
        return StepRecord(
            step="snapshot_reconciliation",
            status="fail",
            duration_ms=(time.perf_counter() - t0) * 1000,
            error=str(exc)[:300],
            details=details,
        )


def evaluate_run_outcome(
    source_records: list[SourceRecord],
    essential_names: set[str],
    freshness: FreshnessRecord | None,
    reports: list[ReportRecord],
    *,
    strict: bool = True,
    skip_freshness: bool = False,
    skip_reports: bool = False,
    skip_sources: bool = False,
    allow_zero: bool = False,
) -> tuple[str, int]:
    """Classify overall status and exit code (pure; unit-testable).

    Fail-closed when ``strict=True`` (canonical path):
    - essential source fail → partial / exit 2
    - essential source success_zero without allow_zero → empty / exit 1
    - no success and no success_zero → failed / exit 1
    - freshness fail (not skipped) → exit 3
    - mandatory report fail (not skipped) → exit 4
    - skip_sources (clean-env offline foundation) → treat sources as skipped
    """
    if skip_sources:
        freshness_status = freshness.status if freshness else "skipped"
        if skip_freshness:
            freshness_status = "skipped"
        report_fails = [r for r in reports if r.status == "fail" and not skip_reports]
        if freshness_status == "fail":
            return "failed", 3
        if report_fails:
            return "failed", 4
        return "success", 0

    if not source_records:
        return "failed", 1

    by_status: dict[str, list[SourceRecord]] = {}
    for rec in source_records:
        by_status.setdefault(rec.status, []).append(rec)

    successes = by_status.get("success", [])
    zeros = by_status.get("success_zero", [])
    fails = by_status.get("fail", [])

    essential_fail = [r for r in fails if r.name in essential_names]
    essential_zero = [r for r in zeros if r.name in essential_names]
    essential_ok = [r for r in successes if r.name in essential_names]
    non_essential_fail = [r for r in fails if r.name not in essential_names]

    freshness_status = freshness.status if freshness else "skipped"
    if skip_freshness:
        freshness_status = "skipped"
    report_fails = [r for r in reports if r.status == "fail" and not skip_reports]

    # --- non-strict legacy path (explicit opt-out only) ---
    if not strict:
        if not successes and not zeros:
            return "failed", 1
        if essential_fail:
            return "partial", 2
        if non_essential_fail:
            return "degraded", 0
        if successes:
            return "success", 0
        return "success_zero", 0

    # --- strict fail-closed ---
    if essential_fail:
        return "partial", 2

    if not successes and not zeros:
        return "failed", 1

    if essential_zero and not essential_ok and not allow_zero:
        # Empty essential sources are not global success in strict mode
        if freshness_status == "fail":
            return "empty", 3
        if report_fails:
            return "empty", 4
        return "empty", 1

    if not successes and zeros:
        overall = "success_zero"
    elif non_essential_fail:
        overall = "degraded"
    else:
        overall = "success"

    if freshness_status == "fail":
        return overall if overall == "empty" else "failed", 3

    if report_fails:
        return "failed", 4

    if overall == "degraded":
        # Non-essential failure with all mandatory gates green → still non-zero
        # so operators cannot treat degraded as full success.
        return "degraded", 5

    return overall, 0


# ---------------------------------------------------------------------------
# Ledger persistence
# ---------------------------------------------------------------------------


def _normalize_ledger_runs(runs: Any) -> list[dict]:
    """Return a flat list of run records; unwrap nested corruption safely."""
    # Prior bug: _save_final_ledger passed the whole ledger dict into _save_ledger,
    # producing {"version":1,"runs":{"version":1,"runs":[...]}}.
    while isinstance(runs, dict) and "runs" in runs:
        runs = runs["runs"]
    if not isinstance(runs, list):
        return []
    return [r for r in runs if isinstance(r, dict)]


def _load_ledger() -> dict:
    if _LEDGER_PATH.exists():
        try:
            with open(_LEDGER_PATH) as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            return {"version": 1, "runs": []}
        if not isinstance(data, dict):
            return {"version": 1, "runs": []}
        return {
            "version": data.get("version", 1),
            "runs": _normalize_ledger_runs(data.get("runs", [])),
        }
    return {"version": 1, "runs": []}


def _save_ledger(runs: list[dict], ledger_path: Path | None = None) -> None:
    path = ledger_path or _LEDGER_PATH
    data = _load_ledger()
    data["runs"] = _normalize_ledger_runs(runs)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

_log_file = _GOLDEN_PATH_DIR / f"gp-{datetime.now().strftime('%Y%m%d-%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(str(_log_file)),
    ],
)
_logger = logging.getLogger("golden-path")


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------


def _echo(msg: str = "", style: str = "default") -> None:
    if HAS_RICH and _console:
        style_map: dict[str, str] = {
            "info": "bold cyan",
            "ok": "bold green",
            "warn": "bold yellow",
            "error": "bold red",
            "header": "bold white on blue",
            "default": "",
        }
        s = style_map.get(style, "")
        if s:
            _console.print(msg, style=s)
        else:
            _console.print(msg)
    else:
        print(msg)


def _print_table(title: str, columns: list[str], rows: list[list[str]]) -> None:
    if HAS_RICH and _console:
        table = Table(title=title, title_style="bold")
        for col in columns:
            table.add_column(col)
        for row in rows:
            table.add_row(*row)
        _console.print(table)
    else:
        header = " | ".join(columns)
        sep = "-" * max(len(h) for h in columns) * 2
        print(f"\n=== {title} ===")
        print(header)
        print(sep)
        for row in rows:
            print(" | ".join(row))


# ---------------------------------------------------------------------------
# Step 1: DB Connectivity
# ---------------------------------------------------------------------------


def check_db(dsn: str) -> tuple[bool, float]:
    """Test PostgreSQL connectivity. Returns (ok, duration_ms)."""
    start = time.monotonic()
    try:
        import psycopg2

        conn = psycopg2.connect(dsn, connect_timeout=10)
        conn.close()
        dur = (time.monotonic() - start) * 1000
        return True, dur
    except Exception as exc:
        dur = (time.monotonic() - start) * 1000
        _logger.error("DB connectivity failed: %s", exc)
        return False, dur


def apply_migrations(dsn: str) -> tuple[bool, float, dict[str, list[str]]]:
    """Apply all versioned SQL migrations under db/migrations (idempotent).

    Uses ``scripts.ops.apply_migrations.apply_range`` in upgrade mode so
    re-runs skip already-applied versions. Returns (ok, duration_ms, summary).
    """
    start = time.monotonic()
    empty: dict[str, list[str]] = {"applied": [], "skipped": [], "repaired": []}
    try:
        from scripts.ops.apply_migrations import apply_range

        root = _PROJECT_ROOT / "db" / "migrations"
        if not root.is_dir():
            _logger.error("Migrations root missing: %s", root)
            dur = (time.monotonic() - start) * 1000
            return False, dur, empty
        summary = apply_range(
            dsn,
            root,
            max_num=None,
            min_num=1,
            allow_concurrent=False,
            mode="upgrade",
        )
        dur = (time.monotonic() - start) * 1000
        return True, dur, summary
    except Exception as exc:
        dur = (time.monotonic() - start) * 1000
        _logger.error("Migration apply failed: %s", exc)
        return False, dur, empty


_SEED_SCRIPTS: tuple[str, ...] = (
    "db/seed/001_sc_entities.py",
    "db/seed/002_entity_aliases.py",
)


def apply_seeds(dsn: str) -> tuple[bool, float, dict[str, list[str]]]:
    """Run deterministic seed scripts (entities + aliases).

    Mirrors ``scripts.ops.run_full_suite.apply_seeds``. Returns
    (ok, duration_ms, summary) with ``ran`` / ``missing`` / ``failed`` lists.
    """
    start = time.monotonic()
    summary: dict[str, list[str]] = {"ran": [], "missing": [], "failed": []}
    env = os.environ.copy()
    env["DATABASE_URL"] = dsn
    env["LOCAL_DATALAKE_DSN"] = dsn
    try:
        for rel in _SEED_SCRIPTS:
            path = _PROJECT_ROOT / rel
            if not path.is_file():
                summary["missing"].append(rel)
                _logger.warning("Seed script missing: %s", rel)
                continue
            # Trusted argv: fixed sys.executable + in-repo seed script path.
            proc = subprocess.run(  # noqa: S603
                [sys.executable, str(path)],
                cwd=str(_PROJECT_ROOT),
                env=env,
                capture_output=True,
                text=True,
                timeout=600,
                check=False,
            )
            if proc.returncode != 0:
                summary["failed"].append(rel)
                _logger.error(
                    "Seed failed %s rc=%s stderr=%s",
                    rel,
                    proc.returncode,
                    (proc.stderr or "")[-500:],
                )
                dur = (time.monotonic() - start) * 1000
                return False, dur, summary
            summary["ran"].append(rel)
        dur = (time.monotonic() - start) * 1000
        # Missing seed scripts is a hard failure for the canonical path.
        if summary["missing"]:
            return False, dur, summary
        return True, dur, summary
    except Exception as exc:
        dur = (time.monotonic() - start) * 1000
        _logger.error("Seed apply failed: %s", exc)
        return False, dur, summary


# ---------------------------------------------------------------------------
# Step 1d: Validate target spreadsheet (canonical — DoD §12.1)
# ---------------------------------------------------------------------------

# Preferred basename (never a .backup / .copy / temp variant).
CANONICAL_SPREADSHEET_BASENAME = "Extra - alvos de licitação. R-0.xlsx"
# Expected included (raio 200km) set size and ordered-ids hash from current DOD.
EXPECTED_CANONICAL_INCLUDED = 1093
EXPECTED_CANONICAL_IDS_SHA256 = "0b3f894d87ba71f2e0fa96887cb3075033488de1af1e6e55f97ccda0701fb396"
REQUIRED_SHEET_NAME = "Entes Públicos SC"
REQUIRED_HEADER_MARKERS = ("Razão Social", "CNPJ", "Município", "Raio")
_BACKUP_NAME_TOKENS = (".backup", ".copy", ".tmp", "~$", ".temp")


def _is_backup_or_temp_name(name: str) -> bool:
    lowered = name.lower()
    return any(tok in lowered for tok in _BACKUP_NAME_TOKENS)


def _ordered_ids_sha256(ids: list[str]) -> str:
    payload = "\n".join(sorted(str(i) for i in ids)).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def resolve_canonical_spreadsheet(
    project_root: Path,
    *,
    explicit_path: Path | None = None,
    allow_backup: bool = False,
) -> Path:
    """Select exactly one canonical target spreadsheet by explicit rules.

    Rules (fail-closed):
    1. Explicit path wins when provided and exists.
    2. Env ``EXTRA_TARGET_SPREADSHEET`` (or ``TARGET_SPREADSHEET_PATH``) if set.
    3. Prefer exact basename CANONICAL_SPREADSHEET_BASENAME under project root
       (local private asset — not required to be git-tracked).
    4. Else non-backup candidates matching ``Extra*alvos*.xlsx`` under root
       (and ``data/``). Zero → missing; >1 → ambiguous.
    5. Backup/copy/temp names are never selected silently.
    6. Backup-only is allowed only when ``allow_backup`` is True.

    Public clones must provide the private asset via env or ``--spreadsheet``.
    Sanitized seed CSV: ``config/target_entities_200km.csv`` (no client branding).
    """
    root = project_root.resolve()

    if explicit_path is not None:
        path = explicit_path.expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"Explicit spreadsheet not found: {path}")
        if _is_backup_or_temp_name(path.name) and not allow_backup:
            raise FileNotFoundError(f"Refusing backup/temp spreadsheet without allow_backup: {path.name}")
        return path

    for env_key in ("EXTRA_TARGET_SPREADSHEET", "TARGET_SPREADSHEET_PATH"):
        env_val = os.environ.get(env_key, "").strip()
        if not env_val:
            continue
        env_path = Path(env_val).expanduser().resolve()
        if not env_path.is_file():
            raise FileNotFoundError(
                f"{env_key} set but file not found: {env_path}. "
                "Provide a local private spreadsheet or unset the variable."
            )
        if _is_backup_or_temp_name(env_path.name) and not allow_backup:
            raise FileNotFoundError(
                f"Refusing backup/temp spreadsheet from {env_key} without allow_backup: {env_path.name}"
            )
        return env_path

    preferred = root / CANONICAL_SPREADSHEET_BASENAME
    if preferred.is_file():
        return preferred.resolve()

    # Public CI / OSS fixture (no client branding in path). Same universe identity.
    public_fixture = root / "fixtures" / "canonical_universe_r0.xlsx"
    if public_fixture.is_file():
        return public_fixture.resolve()

    def _collect(dir_path: Path) -> list[Path]:
        if not dir_path.is_dir():
            return []
        return sorted(dir_path.glob("Extra*alvos*.xlsx"))

    all_hits = _collect(root) + _collect(root / "data")
    # de-dupe by resolve
    seen: set[Path] = set()
    unique: list[Path] = []
    for p in all_hits:
        rp = p.resolve()
        if rp in seen:
            continue
        seen.add(rp)
        unique.append(rp)

    primary = [p for p in unique if not _is_backup_or_temp_name(p.name)]
    backups = [p for p in unique if _is_backup_or_temp_name(p.name)]

    if len(primary) == 1:
        return primary[0]
    if len(primary) > 1:
        names = ", ".join(p.name for p in primary)
        raise FileNotFoundError(f"Ambiguous target spreadsheets (multiple primary candidates): {names}")
    # no primary
    if backups and allow_backup:
        if len(backups) == 1:
            return backups[0]
        names = ", ".join(p.name for p in backups)
        raise FileNotFoundError(f"Ambiguous backup spreadsheets: {names}")
    if backups and not allow_backup:
        raise FileNotFoundError(
            "Only backup/temp spreadsheet(s) found; refusing silent selection. "
            "Provide canonical "
            f"'{CANONICAL_SPREADSHEET_BASENAME}' or set allow_backup."
        )
    raise FileNotFoundError(
        f"Canonical spreadsheet '{CANONICAL_SPREADSHEET_BASENAME}' not found under project root or data/. "
        "Public clones must set EXTRA_TARGET_SPREADSHEET or pass --spreadsheet to a local private asset "
        "(not git-tracked). See docs/ops/private-assets.md. "
        "Sanitized public seed: config/target_entities_200km.csv."
    )


def validate_target_spreadsheet(
    project_root: Path | None = None,
    *,
    explicit_path: Path | None = None,
    allow_backup: bool | None = None,
    expected_included: int = EXPECTED_CANONICAL_INCLUDED,
    expected_ids_sha256: str = EXPECTED_CANONICAL_IDS_SHA256,
) -> tuple[bool, float, dict[str, Any]]:
    """Locate and strongly validate the Extra target spreadsheet (planilha-alvo).

    Uses ``scripts.lib.universe.load_canonical_universe`` (project authority).
    Records path, SHA-256, sheet, physical rows vs canonical included entities
    as separate metrics, and fail-closed on identity mismatch.
    Does not mutate the database.
    """
    start = time.monotonic()
    root = (project_root or _PROJECT_ROOT).resolve()
    details: dict[str, Any] = {
        "physical_rows": None,
        "canonical_entities": None,
        "sheet_name": REQUIRED_SHEET_NAME,
        "path": None,
        "sha256": None,
        "canonical_ids_sha256": None,
        "selection_rule": None,
    }
    if allow_backup is None:
        allow_backup = os.getenv("EXTRA_GP_ALLOW_BACKUP_SPREADSHEET", "").strip() in {
            "1",
            "true",
            "TRUE",
            "yes",
            "YES",
        }
    try:
        from scripts.lib.universe import load_canonical_universe

        xlsx_path = resolve_canonical_spreadsheet(
            root,
            explicit_path=explicit_path,
            allow_backup=allow_backup,
        )
        details["path"] = str(xlsx_path)
        details["selection_rule"] = (
            "explicit"
            if explicit_path is not None
            else (
                "exact_basename"
                if xlsx_path.name == CANONICAL_SPREADSHEET_BASENAME
                else ("backup_explicit_allow" if allow_backup else "single_primary_glob")
            )
        )
        if _is_backup_or_temp_name(xlsx_path.name) and not allow_backup:
            details["error"] = f"backup path selected without allow: {xlsx_path.name}"
            dur = (time.monotonic() - start) * 1000
            return False, dur, details

        # Header / sheet preflight (fail-closed with clear errors)
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            if REQUIRED_SHEET_NAME not in wb.sheetnames:
                details["error"] = f"required sheet '{REQUIRED_SHEET_NAME}' missing; available={list(wb.sheetnames)}"
                dur = (time.monotonic() - start) * 1000
                return False, dur, details
            ws = wb[REQUIRED_SHEET_NAME]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            header_text = " ".join(str(c) for c in (header_row or []) if c is not None)
            missing_markers = [m for m in REQUIRED_HEADER_MARKERS if m.lower() not in header_text.lower()]
            # CNPJ marker may appear as "CNPJ (8 dígitos)"
            if missing_markers:
                details["error"] = f"required header markers missing: {missing_markers}"
                details["header"] = header_text[:300]
                dur = (time.monotonic() - start) * 1000
                return False, dur, details
            details["header"] = header_text[:300]
        finally:
            wb.close()

        universe = load_canonical_universe(seed_path=xlsx_path)
        physical = len(universe.entities)
        included = universe.included
        included_ids = [e.entity_id for e in included]
        ids_sha = _ordered_ids_sha256(included_ids)
        details["physical_rows"] = physical
        details["canonical_entities"] = len(included)
        details["outside_radius"] = len(universe.excluded)
        details["unresolved_rows"] = len(universe.unresolved)
        details["sha256"] = universe.seed_sha256
        details["canonical_ids_sha256"] = ids_sha
        details["seed_path_resolved"] = universe.seed_path
        details["expected_canonical_entities"] = expected_included
        details["expected_canonical_ids_sha256"] = expected_ids_sha256

        if physical <= 0:
            details["error"] = "spreadsheet has zero physical entity rows"
            dur = (time.monotonic() - start) * 1000
            return False, dur, details
        if len(included) != expected_included:
            details["error"] = (
                f"canonical included count mismatch: got {len(included)} "
                f"expected {expected_included} "
                f"(physical_rows={physical})"
            )
            dur = (time.monotonic() - start) * 1000
            return False, dur, details
        if ids_sha != expected_ids_sha256:
            details["error"] = f"canonical_ids_sha256 mismatch (got={ids_sha} expected={expected_ids_sha256})"
            dur = (time.monotonic() - start) * 1000
            return False, dur, details
        # set equality with itself is tautological; uniqueness already enforced
        if len(set(included_ids)) != len(included_ids):
            details["error"] = "duplicate canonical entity ids"
            dur = (time.monotonic() - start) * 1000
            return False, dur, details

        dur = (time.monotonic() - start) * 1000
        return True, dur, details
    except Exception as exc:
        details["error"] = str(exc)[:500]
        dur = (time.monotonic() - start) * 1000
        _logger.error("Target spreadsheet validation failed: %s", exc)
        return False, dur, details


# ---------------------------------------------------------------------------
# Step 2: Crawl Sources (with timeout, retry, backoff + jitter)
# ---------------------------------------------------------------------------


def crawl_source(
    source: SourceDef,
    dsn: str,
    output_json: Path,
) -> SourceRecord:
    """Execute a single source crawl with retry logic."""
    _echo(f"\n>>> Iniciando crawl: {source.name} — {source.description}", "info")

    attempts = 0
    last_error: str | None = None
    metrics: dict[str, int] = {}

    for attempt in range(1, source.max_retries + 1):
        start = time.monotonic()
        try:
            child_env = os.environ.copy()
            # Ensure project root is importable for `config.*` and `scripts.*`
            # when monitor.py is launched as a subprocess (not as -m package).
            root = str(_PROJECT_ROOT)
            existing = child_env.get("PYTHONPATH", "")
            child_env["PYTHONPATH"] = root if not existing else f"{root}{os.pathsep}{existing}"
            # PNCP needs a multi-day window (API is sparse per single day).
            # monitor.py defaults: incremental=7d, full=30d when dates omitted.
            cmd = [
                sys.executable,
                str(_SCRIPTS_DIR / "crawl" / "monitor.py"),
                "--source",
                source.name,
                "--mode",
                "incremental",  # 7d default window after monitor date-window fix
                "--dsn",
                dsn,
                "--output-json",
                str(output_json),
            ]
            if source.name == "pncp":
                # Rolling 2d window ending today — avoids permanent watermark short-circuit
                # on a fixed historical range while staying under wall-clock budget.
                from datetime import date, timedelta

                day_to = date.today()
                day_from = day_to - timedelta(days=2)
                cmd = [
                    sys.executable,
                    str(_SCRIPTS_DIR / "crawl" / "monitor.py"),
                    "--source",
                    "pncp",
                    "--mode",
                    "full",
                    "--date-from",
                    day_from.isoformat(),
                    "--date-to",
                    day_to.isoformat(),
                    "--dsn",
                    dsn,
                    "--output-json",
                    str(output_json),
                ]
            result = subprocess.run(  # noqa: S603
                cmd,
                cwd=str(_PROJECT_ROOT),
                capture_output=True,
                text=True,
                timeout=source.timeout_s,
                env=child_env,
            )
            dur = (time.monotonic() - start) * 1000
            attempts = attempt

            if result.returncode == 0:
                # Parse structured output from monitor.py --output-json
                resumed_live = False
                adapter_failed = False
                adapter_error: str | None = None
                if output_json.exists():
                    try:
                        data = json.loads(output_json.read_text())
                        summary = data.get("summary", {})
                        metrics = {
                            "fetched": summary.get("total_fetched", 0),
                            "transformed": summary.get("total_transformed", 0),
                            "inserted": summary.get("total_inserted", 0),
                            "updated": summary.get("total_updated", 0),
                            "matched": summary.get("total_matched", 0),
                            "persisted": summary.get("total_persisted_opportunities", 0),
                            "external_failures": summary.get("total_external_failures", 0),
                        }
                        # Honor per-source status even when process exit code is 0.
                        for row in data.get("results") or []:
                            if str(row.get("source") or "") != source.name:
                                continue
                            st = str(row.get("status") or "").lower()
                            if st in {"failed", "fail", "error"}:
                                adapter_failed = True
                                adapter_error = (row.get("error_message") or st)[:400]
                            meta = row.get("metadata") or {}
                            wm = meta.get("watermark") or {}
                            ck = (wm.get("checkpoint") or {}) if isinstance(wm, dict) else {}
                            pages = int(ck.get("pages_fetched") or 0)
                            if (
                                wm.get("status") == "committed"
                                and wm.get("db_committed")
                                and pages > 0
                                and int(metrics.get("fetched") or 0) == 0
                            ):
                                resumed_live = True
                                metrics["fetched"] = pages
                                metrics["resumed_from_watermark"] = 1
                                metrics["pages_fetched_watermark"] = pages
                        if int(summary.get("sources_failed") or 0) > 0 and not adapter_failed:
                            # Summary indicates failure for this run
                            for row in data.get("results") or []:
                                if str(row.get("status") or "").lower() in {
                                    "failed",
                                    "fail",
                                    "error",
                                }:
                                    adapter_failed = True
                                    adapter_error = (row.get("error_message") or "sources_failed")[:400]
                                    break
                    except (json.JSONDecodeError, KeyError, TypeError, ValueError):
                        pass

                if adapter_failed:
                    _echo(
                        f"  FAIL {source.name}: adapter reported failure (fetched={metrics.get('fetched', 0)})",
                        "warn",
                    )
                    return SourceRecord(
                        name=source.name,
                        status="fail",
                        duration_ms=dur,
                        attempts=attempts,
                        metrics=metrics,
                        error=adapter_error,
                    )

                fetched = int(metrics.get("fetched", 0) or 0)
                # Exit 0 with zero records is success_zero, never silent "success"
                # (unless watermark resume of a prior live committed crawl).
                if fetched == 0 and not resumed_live:
                    src_status = "success_zero"
                else:
                    src_status = "success"
                label = "ZERO" if src_status == "success_zero" else ("OK/RESUME" if resumed_live else "OK")
                _echo(
                    f"  {label} {source.name}: "
                    f"fetched={fetched}, "
                    f"inserted={metrics.get('inserted', 0)}, "
                    f"persisted={metrics.get('persisted', 0)}" + (" [watermark-resume]" if resumed_live else ""),
                    "ok" if src_status == "success" else "warn",
                )
                return SourceRecord(
                    name=source.name,
                    status=src_status,
                    duration_ms=dur,
                    attempts=attempts,
                    metrics=metrics,
                )

            # Non-zero exit
            last_error = (result.stderr or result.stdout or "").strip()[-300:]
            _echo(
                f"  Attempt {attempt}/{source.max_retries} failed: exit {result.returncode}",
                "warn",
            )

        except subprocess.TimeoutExpired:
            dur = (time.monotonic() - start) * 1000
            attempts = attempt
            last_error = f"timeout after {source.timeout_s}s"
            _echo(
                f"  Attempt {attempt}/{source.max_retries} timed out ({source.timeout_s}s)",
                "warn",
            )

        except Exception as exc:
            dur = (time.monotonic() - start) * 1000
            attempts = attempt
            last_error = str(exc)
            _echo(f"  Attempt {attempt}/{source.max_retries} error: {exc}", "warn")

        # Retry or fail
        if attempt < source.max_retries:
            delay = _backoff_delay(attempt - 1)
            _echo(f"  Retrying in {delay:.1f}s... (backoff+jitter)", "warn")
            time.sleep(delay)
        else:
            return SourceRecord(
                name=source.name,
                status="fail",
                duration_ms=dur,
                attempts=attempts,
                error=last_error,
            )

    return SourceRecord(
        name=source.name,
        status="fail",
        duration_ms=0,
        attempts=attempts,
        error=last_error,
    )


# ---------------------------------------------------------------------------
# Step 3: Freshness Gate
# ---------------------------------------------------------------------------


def run_freshness_gate(dsn: str) -> FreshnessRecord:
    """Execute freshness_gate.py and parse its output."""
    _echo("\n>>> Validando freshness gate...", "info")
    start = time.monotonic()
    try:
        result = subprocess.run(  # noqa: S603
            [sys.executable, str(_SCRIPTS_DIR / "freshness_gate.py")],
            cwd=str(_PROJECT_ROOT),
            capture_output=True,
            text=True,
            timeout=60,
            env={
                **os.environ,
                "LOCAL_DATALAKE_DSN": dsn,
                "PYTHONPATH": f"{_PROJECT_ROOT}:{os.environ.get('PYTHONPATH', '')}",
            },
        )
        dur = (time.monotonic() - start) * 1000

        # Read output JSON produced by freshness_gate.py
        gate_path = _OUTPUT_DIR / "readiness" / "freshness-gate.json"
        details: dict[str, Any] | None = None
        if gate_path.exists():
            details = json.loads(gate_path.read_text())

        if result.returncode == 0:
            _echo(f"  Freshness gate: PASS ({dur:.0f}ms)", "ok")
            return FreshnessRecord(
                status="pass",
                details=details,
            )
        else:
            _echo(f"  Freshness gate: FAIL (stale sources) ({dur:.0f}ms)", "warn")
            if details:
                failing = details.get("overall", {}).get("failing_sources", [])
                if failing:
                    _echo(f"    Sources stale: {', '.join(failing)}", "warn")
            return FreshnessRecord(
                status="fail",
                details=details,
                error=result.stderr[-300:] if result.stderr else None,
            )
    except Exception as exc:
        dur = (time.monotonic() - start) * 1000
        _echo(f"  Freshness gate error: {exc}", "warn")
        return FreshnessRecord(
            status="fail",
            error=str(exc),
        )


# ---------------------------------------------------------------------------
# Step 4: Reports (Excel + PDF)
# ---------------------------------------------------------------------------


def run_editais_report(dsn: str, *, out_dir: Path | None = None) -> StepRecord:
    """Generate domain-specific editais report (not panorama Excel/PDF)."""
    t0 = time.perf_counter()
    try:
        from scripts.reports.editais_report import write_editais_report

        result = write_editais_report(dsn, out_dir=out_dir)
        path = result.get("path")
        size = int(result.get("size") or 0)
        if not result.get("ok") or not path or not Path(path).is_file():
            return StepRecord(
                step="editais_report",
                status="fail",
                duration_ms=(time.perf_counter() - t0) * 1000,
                error=str(result.get("limitations") or "editais report missing"),
                details=result,
            )
        # Require editais-specific identity: filename + columns in sidecar/json
        if "relatorio-editais" not in Path(path).name:
            return StepRecord(
                step="editais_report",
                status="fail",
                duration_ms=(time.perf_counter() - t0) * 1000,
                error="editais report path must be domain-specific",
                details=result,
            )
        cols = result.get("columns") or []
        for required in ("pncp_id", "objeto_compra", "uf"):
            if required not in cols:
                return StepRecord(
                    step="editais_report",
                    status="fail",
                    duration_ms=(time.perf_counter() - t0) * 1000,
                    error=f"missing column {required}",
                    details=result,
                )
        _echo(f"  Editais report: {path} rows={result.get('row_count')} size={size}", "ok")
        return StepRecord(
            step="editais_report",
            status="pass",
            duration_ms=(time.perf_counter() - t0) * 1000,
            details=result,
        )
    except Exception as exc:  # noqa: BLE001
        return StepRecord(
            step="editais_report",
            status="fail",
            duration_ms=(time.perf_counter() - t0) * 1000,
            error=str(exc),
        )


def run_contratos_report(dsn: str, *, out_dir: Path | None = None) -> StepRecord:
    """Generate domain-specific contratos report (not panorama Excel/PDF)."""
    t0 = time.perf_counter()
    try:
        from scripts.reports.contratos_report import write_contratos_report

        result = write_contratos_report(dsn, out_dir=out_dir)
        path = result.get("path")
        size = int(result.get("size") or 0)
        if not result.get("ok") or not path or not Path(path).is_file():
            return StepRecord(
                step="contratos_report",
                status="fail",
                duration_ms=(time.perf_counter() - t0) * 1000,
                error=str(result.get("limitations") or "contratos report missing"),
                details=result,
            )
        if "relatorio-contratos" not in Path(path).name:
            return StepRecord(
                step="contratos_report",
                status="fail",
                duration_ms=(time.perf_counter() - t0) * 1000,
                error="contratos report path must be domain-specific",
                details=result,
            )
        cols = result.get("columns") or []
        for required in ("ente_id", "n_contratos", "valor_total"):
            if required not in cols:
                return StepRecord(
                    step="contratos_report",
                    status="fail",
                    duration_ms=(time.perf_counter() - t0) * 1000,
                    error=f"missing column {required}",
                    details=result,
                )
        _echo(
            f"  Contratos report: {path} rows={result.get('row_count')} size={size}",
            "ok",
        )
        return StepRecord(
            step="contratos_report",
            status="pass",
            duration_ms=(time.perf_counter() - t0) * 1000,
            details=result,
        )
    except Exception as exc:  # noqa: BLE001
        return StepRecord(
            step="contratos_report",
            status="fail",
            duration_ms=(time.perf_counter() - t0) * 1000,
            error=str(exc),
        )


def run_concorrentes_report(dsn: str, *, out_dir: Path | None = None) -> StepRecord:
    """Generate domain-specific concorrentes report (not panorama Excel/PDF)."""
    t0 = time.perf_counter()
    try:
        from scripts.reports.concorrentes_report import write_concorrentes_report

        result = write_concorrentes_report(dsn, out_dir=out_dir)
        path = result.get("path")
        size = int(result.get("size") or 0)
        if not result.get("ok") or not path or not Path(path).is_file():
            return StepRecord(
                step="concorrentes_report",
                status="fail",
                duration_ms=(time.perf_counter() - t0) * 1000,
                error=str(result.get("limitations") or "concorrentes report missing"),
                details=result,
            )
        if "relatorio-concorrentes" not in Path(path).name:
            return StepRecord(
                step="concorrentes_report",
                status="fail",
                duration_ms=(time.perf_counter() - t0) * 1000,
                error="concorrentes report path must be domain-specific",
                details=result,
            )
        cols = result.get("columns") or []
        for required in ("concorrente_id", "n_contratos"):
            if required not in cols:
                return StepRecord(
                    step="concorrentes_report",
                    status="fail",
                    duration_ms=(time.perf_counter() - t0) * 1000,
                    error=f"missing column {required}",
                    details=result,
                )
        _echo(
            f"  Concorrentes report: {path} rows={result.get('row_count')} size={size}",
            "ok",
        )
        return StepRecord(
            step="concorrentes_report",
            status="pass",
            duration_ms=(time.perf_counter() - t0) * 1000,
            details=result,
        )
    except Exception as exc:  # noqa: BLE001
        return StepRecord(
            step="concorrentes_report",
            status="fail",
            duration_ms=(time.perf_counter() - t0) * 1000,
            error=str(exc),
        )


def run_reports(dsn: str) -> list[ReportRecord]:
    """Generate panorama reports (Excel + PDF)."""
    reports: list[ReportRecord] = []

    # --- Excel ---
    _echo("\n>>> Gerando relatorio Excel...", "info")
    try:
        result = subprocess.run(  # noqa: S603
            [
                sys.executable,
                str(_SCRIPTS_DIR / "reports" / "panorama.py"),
                "--output-excel",
                "--dsn",
                dsn,
            ],
            cwd=str(_PROJECT_ROOT),
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode == 0:
            excel_files = sorted(
                (_OUTPUT_DIR / "excels").glob("panorama-*.xlsx"),
                key=lambda p: p.stat().st_mtime,
            )
            path = str(excel_files[-1]) if excel_files else None
            if not path or not Path(path).is_file() or Path(path).stat().st_size < 100:
                _echo("  Excel FAIL: no valid xlsx file produced", "warn")
                reports.append(ReportRecord(type="excel", status="fail", error="missing_or_empty_xlsx"))
            else:
                _echo(f"  Excel: {path}", "ok")
                reports.append(ReportRecord(type="excel", status="generated", path=path))
        else:
            err = (result.stderr or result.stdout or "")[-300:]
            _echo(f"  Excel report failed: {err}", "warn")
            reports.append(ReportRecord(type="excel", status="fail", error=err))
    except Exception as exc:
        _echo(f"  Excel report error: {exc}", "warn")
        reports.append(ReportRecord(type="excel", status="fail", error=str(exc)))

    # --- PDF ---
    _echo("\n>>> Gerando relatorio PDF...", "info")
    try:
        result = subprocess.run(  # noqa: S603
            [
                sys.executable,
                str(_SCRIPTS_DIR / "reports" / "panorama.py"),
                "--output-pdf",
                "--dsn",
                dsn,
            ],
            cwd=str(_PROJECT_ROOT),
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode == 0:
            pdf_files = sorted(
                (_OUTPUT_DIR / "pdfs").glob("panorama-*.pdf"),
                key=lambda p: p.stat().st_mtime,
            )
            path = str(pdf_files[-1]) if pdf_files else None
            if not path or not Path(path).is_file() or Path(path).stat().st_size < 100:
                _echo("  PDF FAIL: no valid pdf file produced", "warn")
                reports.append(ReportRecord(type="pdf", status="fail", error="missing_or_empty_pdf"))
            else:
                _echo(f"  PDF: {path}", "ok")
                reports.append(ReportRecord(type="pdf", status="generated", path=path))
        else:
            err = (result.stderr or result.stdout or "")[-300:]
            _echo(f"  PDF report failed: {err}", "warn")
            reports.append(ReportRecord(type="pdf", status="fail", error=err))
    except Exception as exc:
        _echo(f"  PDF report error: {exc}", "warn")
        reports.append(ReportRecord(type="pdf", status="fail", error=str(exc)))

    return reports


# ---------------------------------------------------------------------------
# Save final ledger
# ---------------------------------------------------------------------------


def _save_final_ledger(
    run_id: str,
    timestamp: str,
    overall: str,
    steps: list[StepRecord],
    sources: list[SourceRecord],
    reports: list[ReportRecord],
    freshness: FreshnessRecord | None,
    wall_start: float,
    ledger_path_str: str | None,
    *,
    dsn: str | None = None,
) -> None:
    wall_dur = (time.monotonic() - wall_start) * 1000
    meta = collect_run_metadata(dsn=dsn)
    for step in steps:
        if step.step == "validate_target_spreadsheet" and isinstance(step.details, dict):
            sha = step.details.get("sha256")
            if sha:
                meta["spreadsheet_sha256"] = sha
            cids = step.details.get("canonical_ids_sha256")
            if cids:
                meta["canonical_ids_sha256"] = cids
            break
    record = RunRecord(
        run_id=run_id,
        timestamp=timestamp,
        status=overall,
        wall_clock_ms=wall_dur,
        steps=steps,
        sources=sources,
        reports=reports,
        freshness=freshness,
        meta=meta,
    )
    data = _load_ledger()
    run_list = _normalize_ledger_runs(data.get("runs", []))
    run_list.append(asdict(record))
    path = Path(ledger_path_str) if ledger_path_str else _LEDGER_PATH
    _save_ledger(run_list, path)
    _echo(f"\nLedger salvo:  {path}")
    _echo(f"Log salvo:     {_log_file}")
    _echo(f"  meta.git_sha={meta.get('git_sha')} schema={meta.get('schema_version')}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Golden Path — Pipeline de validacao completa do DataLake Extra Consultoria",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "--sources",
        default="pncp,pcp,compras_gov",
        help="Comma-separated source names (default: pncp,pcp,compras_gov)",
    )
    p.add_argument(
        "--skip-freshness",
        action="store_true",
        help="Skip freshness gate validation",
    )
    p.add_argument(
        "--skip-sources",
        action="store_true",
        help="Skip essential source crawls (clean-env foundation / offline proof only)",
    )
    p.add_argument(
        "--skip-reports",
        action="store_true",
        help="Skip report generation (Excel + PDF)",
    )
    p.add_argument(
        "--skip-migrations",
        action="store_true",
        help="Skip applying db/migrations (not canonical; use only when schema already applied)",
    )
    p.add_argument(
        "--skip-seeds",
        action="store_true",
        help="Skip db/seed scripts (not canonical; use only when seed already loaded)",
    )
    p.add_argument(
        "--skip-spreadsheet",
        action="store_true",
        help="Skip target spreadsheet validation (not canonical)",
    )
    p.add_argument(
        "--validate-spreadsheet-only",
        action="store_true",
        help=("Run only canonical planilha-alvo validation and write ledger (skips DB/migrations/seeds/crawl/reports)"),
    )
    p.add_argument(
        "--execute-sources-only",
        action="store_true",
        help=(
            "Run only essential source crawls + ledger "
            "(requires DB; skips migrations/seeds/spreadsheet/freshness/reports)"
        ),
    )
    p.add_argument(
        "--execute-freshness-only",
        action="store_true",
        help=("Run only freshness gate + ledger (requires DB; skips crawl/reports/migrations/seeds)"),
    )
    p.add_argument(
        "--execute-coverage-only",
        action="store_true",
        help=(
            "Run dual capability coverage calculation + ledger "
            "(open_tenders + historical_contracts; requires DB + planilha)"
        ),
    )
    p.add_argument(
        "--execute-dual-coverage-only",
        action="store_true",
        help=(
            "Isolated dual-coverage reproof mode (same engine as coverage step). "
            "Use with --capability open_tenders|historical_contracts|both"
        ),
    )
    p.add_argument(
        "--capability",
        choices=["open_tenders", "historical_contracts", "both"],
        default="both",
        help="Capability scope for dual coverage modes (default: both)",
    )
    p.add_argument(
        "--require-coverage-gate",
        action="store_true",
        help="Fail coverage step when either capability gate is below 95 percent (measurement may still succeed)",
    )
    p.add_argument(
        "--execute-snapshot-only",
        action="store_true",
        help="Run only editais snapshot reconciliation + ledger (requires DB)",
    )
    p.add_argument(
        "--execute-reports-only",
        action="store_true",
        help="Run only panorama Excel+PDF generation + ledger (requires DB)",
    )
    p.add_argument(
        "--execute-editais-report-only",
        action="store_true",
        help="Run only domain editais report (CSV+JSON) + ledger (requires DB)",
    )
    p.add_argument(
        "--execute-contratos-report-only",
        action="store_true",
        help="Run only domain contratos report (CSV+JSON) + ledger (requires DB)",
    )
    p.add_argument(
        "--execute-concorrentes-report-only",
        action="store_true",
        help="Run only domain concorrentes report (CSV+JSON) + ledger (requires DB)",
    )
    p.add_argument(
        "--spreadsheet",
        default=None,
        help="Explicit path to target spreadsheet (must not be .backup unless allowed)",
    )
    p.add_argument(
        "--allow-backup-spreadsheet",
        action="store_true",
        help="Allow selecting a .backup/.copy spreadsheet (off by default; fail-closed)",
    )
    p.add_argument(
        "--strict",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Fail-closed on empty essential sources, freshness fail, report fail (default: true)",
    )
    p.add_argument(
        "--allow-zero",
        action="store_true",
        help="In strict mode, accept essential success_zero as valid global success_zero",
    )
    p.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="Verbose/debug output",
    )
    p.add_argument(
        "--dsn",
        default=None,
        help="PostgreSQL DSN (default: LOCAL_DATALAKE_DSN env var or config.settings)",
    )
    p.add_argument(
        "--ledger-output",
        default=str(_LEDGER_PATH),
        help="Output path for execution ledger JSON (default: output/golden-path/ledger.json)",
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()
    wall_start = time.monotonic()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Early path: spreadsheet-only validation (CLI proof for DoD planilha item)
    if args.validate_spreadsheet_only:
        run_id = f"gp-ss-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        timestamp = datetime.now(UTC).isoformat()
        steps: list[StepRecord] = []
        source_records: list[SourceRecord] = []
        report_records: list[ReportRecord] = []
        freshness_record: FreshnessRecord | None = None
        explicit = Path(args.spreadsheet) if args.spreadsheet else None
        _echo("\n[validate-spreadsheet-only] Validando planilha-alvo canônica...", "header")
        ss_ok, ss_dur, ss_details = validate_target_spreadsheet(
            _PROJECT_ROOT,
            explicit_path=explicit,
            allow_backup=args.allow_backup_spreadsheet or None,
        )
        steps.append(
            StepRecord(
                step="validate_target_spreadsheet",
                status="pass" if ss_ok else "fail",
                duration_ms=ss_dur,
                details=ss_details,
            )
        )
        overall = "success" if ss_ok else "failed"
        _save_final_ledger(
            run_id,
            timestamp,
            overall,
            steps,
            source_records,
            report_records,
            freshness_record,
            wall_start,
            args.ledger_output,
        )
        if not ss_ok:
            _echo(
                f"  Planilha-alvo INVALIDA: {ss_details.get('error', 'unknown')}",
                "error",
            )
            return 1
        _echo(
            "  Planilha-alvo OK "
            f"(physical={ss_details.get('physical_rows')} "
            f"canonical={ss_details.get('canonical_entities')} "
            f"sha256={str(ss_details.get('sha256', ''))[:12]}… "
            f"{ss_dur:.0f}ms)",
            "ok",
        )
        _echo(f"  Ledger: {args.ledger_output}", "ok")
        return 0

    # Early path: execute essential sources only (DoD fontes mínimas proof)
    if args.execute_sources_only:
        run_id = f"gp-src-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        timestamp = datetime.now(UTC).isoformat()
        steps: list[StepRecord] = []
        source_records: list[SourceRecord] = []
        report_records: list[ReportRecord] = []
        freshness_record: FreshnessRecord | None = None

        dsn = args.dsn or os.getenv("LOCAL_DATALAKE_DSN")
        if not dsn:
            try:
                import config.settings as _cfg

                dsn = _cfg.LOCAL_DATALAKE_DSN
            except ImportError:
                dsn = "postgresql://postgres:@127.0.0.1:54399/postgres"

        ok_db, dur_db = check_db(dsn)
        steps.append(
            StepRecord(
                step="db_connectivity",
                status="pass" if ok_db else "fail",
                duration_ms=dur_db,
            )
        )
        if not ok_db:
            _save_final_ledger(
                run_id,
                timestamp,
                "failed",
                steps,
                source_records,
                report_records,
                freshness_record,
                wall_start,
                args.ledger_output,
            )
            return 1

        source_names_arg = [s.strip() for s in args.sources.split(",") if s.strip()]
        selected = [s for s in SOURCES if s.name in source_names_arg]
        if not selected:
            selected = essential_sources()
        # Always ensure essentials included for this proof mode
        selected_names = {s.name for s in selected}
        for s in essential_sources():
            if s.name not in selected_names:
                selected.append(s)

        _echo("\n[execute-sources-only] Executando fontes mínimas...", "header")
        for src in selected:
            output_json = _GOLDEN_PATH_DIR / f"crawl-{src.name}-{run_id}.json"
            rec = crawl_source(src, dsn, output_json)
            source_records.append(rec)

        exec_ok, exec_details = assert_essential_sources_executed(source_records)
        steps.append(
            StepRecord(
                step="execute_essential_sources",
                status="pass" if exec_ok else "fail",
                duration_ms=sum(r.duration_ms for r in source_records),
                details=exec_details,
            )
        )
        persist_ok, persist_details = assert_sources_persisted(source_records)
        steps.append(
            StepRecord(
                step="persist_source_data",
                status="pass" if persist_ok else "fail",
                duration_ms=0.0,
                details=persist_details,
            )
        )
        overall = "success" if exec_ok else "failed"
        # If essentials executed but all failed adapters, still "executed" for this item;
        # exit non-zero only when not executed.
        _save_final_ledger(
            run_id,
            timestamp,
            overall,
            steps,
            source_records,
            report_records,
            freshness_record,
            wall_start,
            args.ledger_output,
        )
        if not exec_ok:
            _echo(f"  Fontes mínimas NÃO executadas: {exec_details}", "error")
            return 1
        _echo(
            "  Fontes mínimas executadas: "
            + ", ".join(f"{n}={v['status']}/a{v['attempts']}" for n, v in exec_details["executed"].items()),
            "ok",
        )
        _echo(f"  Ledger: {args.ledger_output}", "ok")
        return 0

    # Early path: freshness gate only (DoD executa freshness gate)
    if args.execute_freshness_only:
        run_id = f"gp-fr-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        timestamp = datetime.now(UTC).isoformat()
        steps: list[StepRecord] = []
        source_records: list[SourceRecord] = []
        report_records: list[ReportRecord] = []

        dsn = args.dsn or os.getenv("LOCAL_DATALAKE_DSN")
        if not dsn:
            try:
                import config.settings as _cfg

                dsn = _cfg.LOCAL_DATALAKE_DSN
            except ImportError:
                dsn = "postgresql://postgres:@127.0.0.1:54399/postgres"

        ok_db, dur_db = check_db(dsn)
        steps.append(
            StepRecord(
                step="db_connectivity",
                status="pass" if ok_db else "fail",
                duration_ms=dur_db,
            )
        )
        if not ok_db:
            _save_final_ledger(
                run_id,
                timestamp,
                "failed",
                steps,
                source_records,
                report_records,
                None,
                wall_start,
                args.ledger_output,
            )
            return 1

        _echo("\n[execute-freshness-only] Executando freshness gate...", "header")
        freshness_record = run_freshness_gate(dsn)
        exec_ok, exec_details = assert_freshness_gate_executed(freshness_record)
        steps.append(
            StepRecord(
                step="run_freshness_gate",
                status="pass" if exec_ok else "fail",
                duration_ms=0.0,
                details={
                    **exec_details,
                    "gate_status": freshness_record.status,
                    "gate_details": freshness_record.details,
                },
            )
        )
        overall = "success" if exec_ok else "failed"
        _save_final_ledger(
            run_id,
            timestamp,
            overall,
            steps,
            source_records,
            report_records,
            freshness_record,
            wall_start,
            args.ledger_output,
        )
        if not exec_ok:
            _echo(f"  Freshness gate NÃO executado: {exec_details}", "error")
            return 1
        _echo(
            f"  Freshness gate executado: status={freshness_record.status} "
            f"failing={exec_details.get('failing_sources')}",
            "ok" if freshness_record.status == "pass" else "warn",
        )
        _echo(f"  Ledger: {args.ledger_output}", "ok")
        # Exit 0 when gate ran (even if status=fail); non-zero only if not executed.
        return 0

    if args.execute_coverage_only or args.execute_dual_coverage_only:
        mode = "dual-coverage" if args.execute_dual_coverage_only else "coverage"
        run_id = f"gp-cov-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        timestamp = datetime.now(UTC).isoformat()
        steps: list[StepRecord] = []
        dsn = args.dsn or os.getenv("LOCAL_DATALAKE_DSN")
        if not dsn:
            try:
                import config.settings as _cfg

                dsn = _cfg.LOCAL_DATALAKE_DSN
            except ImportError:
                dsn = "postgresql://postgres:@127.0.0.1:54399/postgres"
        ok_db, dur_db = check_db(dsn)
        steps.append(StepRecord(step="db_connectivity", status="pass" if ok_db else "fail", duration_ms=dur_db))
        if not ok_db:
            _save_final_ledger(run_id, timestamp, "failed", steps, [], [], None, wall_start, args.ledger_output)
            return 1
        caps = ["open_tenders", "historical_contracts"] if args.capability == "both" else [args.capability]
        _echo(f"\n[execute-{mode}-only] Calculando cobertura dual {caps}...", "header")
        cov = run_coverage_calculation(
            dsn,
            project_root=_PROJECT_ROOT,
            capabilities=caps,
            require_gate=bool(args.require_coverage_gate),
        )
        steps.append(cov)
        d = cov.details or {}
        measurement_ok = bool(d.get("measurement_success")) if d else cov.status == "pass"
        gate_ok = bool(d.get("coverage_gate_pass"))
        overall = "success" if cov.status == "pass" else "failed"
        _save_final_ledger(run_id, timestamp, overall, steps, [], [], None, wall_start, args.ledger_output)
        if not measurement_ok or cov.status != "pass":
            _echo(f"  Cobertura FALHOU: {cov.error or cov.details}", "error")
            return 1
        for cap_name, cap_block in (d.get("capabilities") or {}).items():
            _echo(
                f"  {cap_name}: den={cap_block.get('applicable_denominator')} "
                f"num={cap_block.get('covered_numerator')} pct={cap_block.get('coverage_pct')} "
                f"gate={cap_block.get('gate_status')} presence={cap_block.get('data_presence_pct')} "
                f"fresh={cap_block.get('fresh_count')} stale={cap_block.get('stale_count')} "
                f"unknown={cap_block.get('unknown_count')} blocked={cap_block.get('blocked_count')}",
                "ok" if cap_block.get("gate_status") == "PASS" else "warn",
            )
        _echo(
            f"  measurement_success={measurement_ok} coverage_gate_pass={gate_ok} method={d.get('method')}",
            "ok",
        )
        _echo(f"  Ledger: {args.ledger_output}", "ok")
        if args.require_coverage_gate and not gate_ok:
            return 2
        return 0

    if args.execute_snapshot_only:
        run_id = f"gp-snap-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        timestamp = datetime.now(UTC).isoformat()
        steps: list[StepRecord] = []
        dsn = args.dsn or os.getenv("LOCAL_DATALAKE_DSN")
        if not dsn:
            try:
                import config.settings as _cfg

                dsn = _cfg.LOCAL_DATALAKE_DSN
            except ImportError:
                dsn = "postgresql://postgres:@127.0.0.1:54399/postgres"
        ok_db, dur_db = check_db(dsn)
        steps.append(StepRecord(step="db_connectivity", status="pass" if ok_db else "fail", duration_ms=dur_db))
        if not ok_db:
            _save_final_ledger(run_id, timestamp, "failed", steps, [], [], None, wall_start, args.ledger_output)
            return 1
        _echo("\n[execute-snapshot-only] Reconciliando snapshot de editais...", "header")
        snap = run_snapshot_reconciliation(dsn, project_root=_PROJECT_ROOT)
        steps.append(snap)
        overall = "success" if snap.status == "pass" else "failed"
        _save_final_ledger(run_id, timestamp, overall, steps, [], [], None, wall_start, args.ledger_output)
        if snap.status != "pass":
            _echo(f"  Snapshot FAIL: {snap.error or snap.details}", "error")
            return 1
        d = snap.details or {}
        _echo(
            f"  Snapshot OK count={d.get('current_count')} "
            f"added={d.get('added')} removed={d.get('removed')} changed={d.get('changed')} "
            f"baseline={d.get('baseline')}",
            "ok",
        )
        _echo(f"  Ledger: {args.ledger_output}", "ok")
        return 0

    if args.execute_reports_only:
        run_id = f"gp-rep-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        timestamp = datetime.now(UTC).isoformat()
        steps: list[StepRecord] = []
        dsn = args.dsn or os.getenv("LOCAL_DATALAKE_DSN")
        if not dsn:
            try:
                import config.settings as _cfg

                dsn = _cfg.LOCAL_DATALAKE_DSN
            except ImportError:
                dsn = "postgresql://postgres:@127.0.0.1:54399/postgres"
        ok_db, dur_db = check_db(dsn)
        steps.append(StepRecord(step="db_connectivity", status="pass" if ok_db else "fail", duration_ms=dur_db))
        if not ok_db:
            _save_final_ledger(run_id, timestamp, "failed", steps, [], [], None, wall_start, args.ledger_output)
            return 1
        _echo("\n[execute-reports-only] Gerando Excel + PDF...", "header")
        report_records = run_reports(dsn)
        ok_excel = any(r.type == "excel" and r.status == "generated" and r.path for r in report_records)
        ok_pdf = any(r.type == "pdf" and r.status == "generated" and r.path for r in report_records)
        steps.append(
            StepRecord(
                step="generate_reports",
                status="pass" if (ok_excel and ok_pdf) else "fail",
                duration_ms=0.0,
                details={
                    "excel": next((r.path for r in report_records if r.type == "excel"), None),
                    "pdf": next((r.path for r in report_records if r.type == "pdf"), None),
                    "records": [asdict(r) for r in report_records],
                },
            )
        )
        overall = "success" if (ok_excel and ok_pdf) else "failed"
        _save_final_ledger(run_id, timestamp, overall, steps, [], report_records, None, wall_start, args.ledger_output)
        if overall != "success":
            _echo(f"  Reports FAIL excel={ok_excel} pdf={ok_pdf}", "error")
            return 1
        _echo("  Reports OK (excel+pdf files present)", "ok")
        _echo(f"  Ledger: {args.ledger_output}", "ok")
        return 0

    if args.execute_editais_report_only:
        run_id = f"gp-editais-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        timestamp = datetime.now(UTC).isoformat()
        steps: list[StepRecord] = []
        wall_start = time.monotonic()
        dsn = args.dsn or os.getenv("LOCAL_DATALAKE_DSN")
        if not dsn:
            try:
                import config.settings as _cfg

                dsn = _cfg.LOCAL_DATALAKE_DSN
            except ImportError:
                dsn = "postgresql://postgres:@127.0.0.1:54399/postgres"
        ok_db, dur_db = check_db(dsn)
        steps.append(StepRecord(step="db_connectivity", status="pass" if ok_db else "fail", duration_ms=dur_db))
        if not ok_db:
            _save_final_ledger(run_id, timestamp, "failed", steps, [], [], None, wall_start, args.ledger_output)
            return 1
        _echo("\n[execute-editais-report-only] Gerando relatório de editais...", "header")
        editais_step = run_editais_report(dsn)
        steps.append(editais_step)
        overall = "success" if editais_step.status == "pass" else "failed"
        _save_final_ledger(run_id, timestamp, overall, steps, [], [], None, wall_start, args.ledger_output, dsn=dsn)
        if overall != "success":
            _echo(f"  Editais report FAIL: {editais_step.error}", "error")
            return 1
        _echo("  Editais report OK (domain CSV+JSON present)", "ok")
        _echo(f"  Ledger: {args.ledger_output}", "ok")
        return 0

    if args.execute_contratos_report_only:
        run_id = f"gp-contratos-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        timestamp = datetime.now(UTC).isoformat()
        steps: list[StepRecord] = []
        wall_start = time.monotonic()
        dsn = args.dsn or os.getenv("LOCAL_DATALAKE_DSN")
        if not dsn:
            try:
                import config.settings as _cfg

                dsn = _cfg.LOCAL_DATALAKE_DSN
            except ImportError:
                dsn = "postgresql://postgres:@127.0.0.1:54399/postgres"
        ok_db, dur_db = check_db(dsn)
        steps.append(StepRecord(step="db_connectivity", status="pass" if ok_db else "fail", duration_ms=dur_db))
        if not ok_db:
            _save_final_ledger(run_id, timestamp, "failed", steps, [], [], None, wall_start, args.ledger_output)
            return 1
        _echo("\n[execute-contratos-report-only] Gerando relatório de contratos...", "header")
        contratos_step = run_contratos_report(dsn)
        steps.append(contratos_step)
        overall = "success" if contratos_step.status == "pass" else "failed"
        _save_final_ledger(run_id, timestamp, overall, steps, [], [], None, wall_start, args.ledger_output, dsn=dsn)
        if overall != "success":
            _echo(f"  Contratos report FAIL: {contratos_step.error}", "error")
            return 1
        _echo("  Contratos report OK (domain CSV+JSON present)", "ok")
        _echo(f"  Ledger: {args.ledger_output}", "ok")
        return 0

    if args.execute_concorrentes_report_only:
        run_id = f"gp-concorrentes-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        timestamp = datetime.now(UTC).isoformat()
        steps: list[StepRecord] = []
        wall_start = time.monotonic()
        dsn = args.dsn or os.getenv("LOCAL_DATALAKE_DSN")
        if not dsn:
            try:
                import config.settings as _cfg

                dsn = _cfg.LOCAL_DATALAKE_DSN
            except ImportError:
                dsn = "postgresql://postgres:@127.0.0.1:54399/postgres"
        ok_db, dur_db = check_db(dsn)
        steps.append(StepRecord(step="db_connectivity", status="pass" if ok_db else "fail", duration_ms=dur_db))
        if not ok_db:
            _save_final_ledger(run_id, timestamp, "failed", steps, [], [], None, wall_start, args.ledger_output)
            return 1
        _echo("\n[execute-concorrentes-report-only] Gerando relatório de concorrentes...", "header")
        conc_step = run_concorrentes_report(dsn)
        steps.append(conc_step)
        overall = "success" if conc_step.status == "pass" else "failed"
        _save_final_ledger(run_id, timestamp, overall, steps, [], [], None, wall_start, args.ledger_output, dsn=dsn)
        if overall != "success":
            _echo(f"  Concorrentes report FAIL: {conc_step.error}", "error")
            return 1
        _echo("  Concorrentes report OK (domain CSV+JSON present)", "ok")
        _echo(f"  Ledger: {args.ledger_output}", "ok")
        return 0

    # ── Resolve DSN ──
    dsn = args.dsn or os.getenv("LOCAL_DATALAKE_DSN")
    if not dsn:
        try:
            import config.settings as _cfg

            dsn = _cfg.LOCAL_DATALAKE_DSN
        except ImportError:
            dsn = "postgresql://postgres:@127.0.0.1:54399/postgres"

    # ── Resolve sources ──
    source_names_arg = [s.strip() for s in args.sources.split(",")]
    selected_sources = [s for s in SOURCES if s.name in source_names_arg]
    if not selected_sources:
        _echo(
            f"ERROR: No valid sources in '{args.sources}'. Available: {', '.join(s.name for s in SOURCES)}",
            "error",
        )
        return 1

    run_id = f"gp-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    timestamp = datetime.now(UTC).isoformat()

    steps: list[StepRecord] = []
    source_records: list[SourceRecord] = []
    report_records: list[ReportRecord] = []
    freshness_record: FreshnessRecord | None = None

    # ── Header ──
    _echo("")
    panel_text = (
        "[bold]GOLDEN PATH[/bold] — Pipeline de Validacao do DataLake\n"
        f"Run: {run_id}  |  "
        f"Fontes: {', '.join(s.name for s in selected_sources)}  |  "
        f"DSN: {dsn[:60]}..."
    )
    if HAS_RICH:
        _echo(Panel.fit(panel_text))
    else:
        _echo(f"{'=' * 60}")
        _echo(f"  GOLDEN PATH — {run_id}")
        _echo(f"{'=' * 60}")

    # =========================================================================
    # Step 1: DB Connectivity
    # =========================================================================
    _echo("\n[1/4] Verificando conectividade PostgreSQL...", "header")
    ok, dur = check_db(dsn)
    steps.append(
        StepRecord(
            step="db_connectivity",
            status="pass" if ok else "fail",
            duration_ms=dur,
        )
    )
    if not ok:
        _echo("  PostgreSQL NAO respondeu. Execute 'make db-up' primeiro.", "error")
        _echo(f"  DSN: {dsn}", "warn")
        _save_final_ledger(
            run_id,
            timestamp,
            "failed",
            steps,
            source_records,
            report_records,
            freshness_record,
            wall_start,
            args.ledger_output,
        )
        return 1
    _echo(f"  PostgreSQL OK ({dur:.0f}ms)", "ok")

    # =========================================================================
    # Step 1b: Apply migrations (canonical — DoD §12.1)
    # =========================================================================
    if args.skip_migrations:
        _echo("\n[1b/5] Migrations SKIPPED (--skip-migrations)", "warn")
        steps.append(
            StepRecord(
                step="apply_migrations",
                status="skipped",
                duration_ms=0.0,
                details={"reason": "skip-migrations"},
            )
        )
    else:
        _echo("\n[1b/5] Aplicando migrations (db/migrations)...", "header")
        mig_ok, mig_dur, mig_summary = apply_migrations(dsn)
        steps.append(
            StepRecord(
                step="apply_migrations",
                status="pass" if mig_ok else "fail",
                duration_ms=mig_dur,
                details={
                    "applied": len(mig_summary.get("applied") or []),
                    "skipped": len(mig_summary.get("skipped") or []),
                    "repaired": len(mig_summary.get("repaired") or []),
                },
            )
        )
        if not mig_ok:
            _echo("  Migrations FALHARAM (fail-closed).", "error")
            _save_final_ledger(
                run_id,
                timestamp,
                "failed",
                steps,
                source_records,
                report_records,
                freshness_record,
                wall_start,
                args.ledger_output,
            )
            return 1
        _echo(
            "  Migrations OK "
            f"(applied={len(mig_summary.get('applied') or [])} "
            f"skipped={len(mig_summary.get('skipped') or [])} "
            f"repaired={len(mig_summary.get('repaired') or [])} "
            f"{mig_dur:.0f}ms)",
            "ok",
        )

    # =========================================================================
    # Step 1c: Apply seeds (canonical — DoD §12.1)
    # =========================================================================
    if args.skip_seeds:
        _echo("\n[1c/6] Seeds SKIPPED (--skip-seeds)", "warn")
        steps.append(
            StepRecord(
                step="apply_seeds",
                status="skipped",
                duration_ms=0.0,
                details={"reason": "skip-seeds"},
            )
        )
    else:
        _echo("\n[1c/6] Aplicando seeds (db/seed)...", "header")
        seed_ok, seed_dur, seed_summary = apply_seeds(dsn)
        steps.append(
            StepRecord(
                step="apply_seeds",
                status="pass" if seed_ok else "fail",
                duration_ms=seed_dur,
                details={
                    "ran": list(seed_summary.get("ran") or []),
                    "missing": list(seed_summary.get("missing") or []),
                    "failed": list(seed_summary.get("failed") or []),
                },
            )
        )
        if not seed_ok:
            _echo("  Seeds FALHARAM (fail-closed).", "error")
            _save_final_ledger(
                run_id,
                timestamp,
                "failed",
                steps,
                source_records,
                report_records,
                freshness_record,
                wall_start,
                args.ledger_output,
            )
            return 1
        _echo(
            f"  Seeds OK (ran={len(seed_summary.get('ran') or [])} {seed_dur:.0f}ms)",
            "ok",
        )

    # =========================================================================
    # Step 1d: Validate target spreadsheet (canonical — DoD §12.1)
    # =========================================================================
    if args.skip_spreadsheet:
        _echo("\n[1d/7] Planilha-alvo SKIPPED (--skip-spreadsheet)", "warn")
        steps.append(
            StepRecord(
                step="validate_target_spreadsheet",
                status="skipped",
                duration_ms=0.0,
                details={"reason": "skip-spreadsheet"},
            )
        )
    else:
        _echo("\n[1d/7] Validando planilha-alvo canônica...", "header")
        explicit = Path(args.spreadsheet) if args.spreadsheet else None
        ss_ok, ss_dur, ss_details = validate_target_spreadsheet(
            _PROJECT_ROOT,
            explicit_path=explicit,
            allow_backup=args.allow_backup_spreadsheet or None,
        )
        steps.append(
            StepRecord(
                step="validate_target_spreadsheet",
                status="pass" if ss_ok else "fail",
                duration_ms=ss_dur,
                details=ss_details,
            )
        )
        if not ss_ok:
            _echo(
                f"  Planilha-alvo INVALIDA: {ss_details.get('error', 'unknown')}",
                "error",
            )
            _save_final_ledger(
                run_id,
                timestamp,
                "failed",
                steps,
                source_records,
                report_records,
                freshness_record,
                wall_start,
                args.ledger_output,
            )
            return 1
        _echo(
            "  Planilha-alvo OK "
            f"(physical={ss_details.get('physical_rows')} "
            f"canonical={ss_details.get('canonical_entities')} "
            f"ids={str(ss_details.get('canonical_ids_sha256', ''))[:12]}… "
            f"{ss_dur:.0f}ms)",
            "ok",
        )

    # =========================================================================
    # Step 2: Crawl Sources
    # =========================================================================
    if args.skip_sources:
        _echo("\n[2/7] Crawl das fontes SKIPPED (--skip-sources)", "warn")
        source_records = []
        sources_success = []
        sources_zero = []
        sources_fail = []
        essential_fail = []
        essential_names: set[str] = set()
        exec_ok, exec_details = True, {"reason": "skip-sources", "executed": []}
        persist_ok, persist_details = True, {"reason": "skip-sources"}
    else:
        _echo("\n[2/7] Crawl das fontes de dados...", "header")
        _echo(f"  Fontes: {', '.join(s.name for s in selected_sources)}")
        _echo("  Timeout: 120s por fonte  |  Retries: 3x com backoff+jitter")

        for src in selected_sources:
            output_json = _GOLDEN_PATH_DIR / f"crawl-{src.name}-{run_id}.json"
            rec = crawl_source(src, dsn, output_json)
            source_records.append(rec)

        # Classify
        sources_success = [r for r in source_records if r.status == "success"]
        sources_zero = [r for r in source_records if r.status == "success_zero"]
        sources_fail = [r for r in source_records if r.status == "fail"]
        essential_names = {s.name for s in selected_sources if s.essential}
        essential_fail = [r for r in sources_fail if r.name in essential_names]

        exec_ok, exec_details = assert_essential_sources_executed(source_records)
        persist_ok, persist_details = assert_sources_persisted(source_records)

    steps.append(
        StepRecord(
            step="execute_essential_sources",
            status=("skipped" if args.skip_sources else ("pass" if exec_ok else "fail")),
            duration_ms=sum(r.duration_ms for r in source_records),
            details=exec_details,
        )
    )
    steps.append(
        StepRecord(
            step="persist_source_data",
            status=("skipped" if args.skip_sources else ("pass" if persist_ok else "fail")),
            duration_ms=0.0,
            details=persist_details,
        )
    )

    def _src_label(status: str) -> str:
        return {
            "success": "OK",
            "success_zero": "ZERO",
            "fail": "FAIL",
            "skipped": "SKIP",
        }.get(status, status.upper())

    _echo("")
    _print_table(
        "Resultados por Fonte",
        ["Fonte", "Status", "Tempo", "Tentativas", "Fetched", "Inserted", "Persistidas"],
        [
            [
                r.name,
                _src_label(r.status),
                f"{r.duration_ms:.0f}ms",
                str(r.attempts),
                str((r.metrics or {}).get("fetched", "-")),
                str((r.metrics or {}).get("inserted", "-")),
                str((r.metrics or {}).get("persisted", "-")),
            ]
            for r in source_records
        ],
    )

    if sources_fail or sources_zero:
        _echo(
            f"\n  {len(sources_success)} com dados, {len(sources_zero)} zero, {len(sources_fail)} falha(s)",
            "warn",
        )
        for r in sources_fail:
            _echo(f"    {r.name}: {r.error or 'erro desconhecido'}", "warn")
        for r in sources_zero:
            _echo(f"    {r.name}: success_zero (fetched=0)", "warn")
    else:
        _echo(f"\n  {len(sources_success)}/{len(selected_sources)} fontes OK", "ok")

    # =========================================================================
    # Step 3: Freshness Gate
    # =========================================================================
    _echo("\n[3/4] Freshness Gate...", "header")
    if args.skip_freshness:
        _echo("  SKIP (--skip-freshness)", "warn")
        freshness_record = FreshnessRecord(status="skipped")
        steps.append(
            StepRecord(
                step="run_freshness_gate",
                status="skipped",
                duration_ms=0.0,
                details={"reason": "skip-freshness"},
            )
        )
    else:
        freshness_record = run_freshness_gate(dsn)
        fr_ok, fr_details = assert_freshness_gate_executed(freshness_record)
        steps.append(
            StepRecord(
                step="run_freshness_gate",
                status="pass" if fr_ok else "fail",
                duration_ms=0.0,
                details={
                    **fr_details,
                    "gate_status": freshness_record.status,
                },
            )
        )

    # =========================================================================
    # Step 3b: Coverage calculation (DoD §12.1)
    # =========================================================================
    _echo("\n[3b/7] Calculando cobertura...", "header")
    cov_step = run_coverage_calculation(dsn, project_root=_PROJECT_ROOT)
    steps.append(cov_step)
    if cov_step.status == "pass":
        d = cov_step.details or {}
        for cap_name, cap_block in (d.get("capabilities") or {}).items():
            _echo(
                f"  {cap_name}: den={cap_block.get('applicable_denominator')} "
                f"num={cap_block.get('covered_numerator')} pct={cap_block.get('coverage_pct')} "
                f"gate={cap_block.get('gate_status')}",
                "ok" if cap_block.get("gate_status") == "PASS" else "warn",
            )
        if not d.get("capabilities"):
            _echo(
                f"  Cobertura den={d.get('denominator')} num={d.get('numerator')} pct={d.get('coverage_pct')}",
                "ok",
            )
        _echo(
            f"  measurement_success={d.get('measurement_success')} coverage_gate_pass={d.get('coverage_gate_pass')}",
            "ok",
        )
    else:
        _echo(f"  Cobertura FAIL: {cov_step.error}", "warn")

    # =========================================================================
    # Step 3c: Snapshot reconciliation (editais)
    # =========================================================================
    _echo("\n[3c/7] Reconciliando snapshot de editais...", "header")
    snap_step = run_snapshot_reconciliation(dsn, project_root=_PROJECT_ROOT)
    steps.append(snap_step)
    if snap_step.status == "pass":
        d = snap_step.details or {}
        _echo(
            f"  Snapshot count={d.get('current_count')} "
            f"+{d.get('added')}/-{d.get('removed')}/~{d.get('changed')} "
            f"baseline={d.get('baseline')}",
            "ok",
        )
    else:
        _echo(f"  Snapshot FAIL: {snap_step.error}", "warn")

    # =========================================================================
    # Step 4: Reports
    # =========================================================================
    _echo("\n[4/7] Geracao de relatorios...", "header")
    if args.skip_reports:
        _echo("  SKIP (--skip-reports)", "warn")
        report_records = [
            ReportRecord(type="excel", status="skipped"),
            ReportRecord(type="pdf", status="skipped"),
        ]
    else:
        report_records = run_reports(dsn)

    # Step 4b: domain-specific editais report (DoD §12.1 — not panorama)
    _echo("\n[4b/7] Relatório de editais (domínio)...", "header")
    editais_step = run_editais_report(dsn)
    steps.append(editais_step)
    if editais_step.status != "pass":
        _echo(f"  Editais report FAIL: {editais_step.error}", "error")

    # Step 4c: domain-specific contratos report (DoD §12.1 — not panorama)
    _echo("\n[4c/7] Relatório de contratos (domínio)...", "header")
    contratos_step = run_contratos_report(dsn)
    steps.append(contratos_step)
    if contratos_step.status != "pass":
        _echo(f"  Contratos report FAIL: {contratos_step.error}", "error")

    # Step 4d: domain-specific concorrentes report
    _echo("\n[4d/7] Relatório de concorrentes (domínio)...", "header")
    concorrentes_step = run_concorrentes_report(dsn)
    steps.append(concorrentes_step)
    if concorrentes_step.status != "pass":
        _echo(f"  Concorrentes report FAIL: {concorrentes_step.error}", "error")

    # =========================================================================
    # Summary
    # =========================================================================
    wall_dur = (time.monotonic() - wall_start) * 1000

    overall, exit_code = evaluate_run_outcome(
        source_records,
        essential_names,
        freshness_record,
        report_records,
        strict=bool(args.strict),
        skip_freshness=bool(args.skip_freshness),
        skip_reports=bool(args.skip_reports),
        skip_sources=bool(args.skip_sources),
        allow_zero=bool(args.allow_zero),
    )
    # Domain domain reports are mandatory for §12.1 (independent of panorama Excel/PDF).
    if editais_step.status != "pass" and exit_code == 0:
        overall = "failed"
        exit_code = 4
    if contratos_step.status != "pass" and exit_code == 0:
        overall = "failed"
        exit_code = 4
    if concorrentes_step.status != "pass" and exit_code == 0:
        overall = "failed"
        exit_code = 4

    freshness_str = freshness_record.status if freshness_record else "N/A"
    excel_status = report_records[0].status if report_records else "N/A"
    pdf_status = report_records[1].status if len(report_records) > 1 else "N/A"
    mode_str = "strict" if args.strict else "no-strict"

    summary_text = (
        f"[bold]GOLDEN PATH — RESUMO[/bold]\n\n"
        f"Run ID:          {run_id}\n"
        f"Mode:            {mode_str}\n"
        f"Status:          {overall.upper()}\n"
        f"Exit code:       {exit_code}\n"
        f"Fontes c/ dados: {len(sources_success)}/{len(selected_sources)}\n"
        f"Fontes zero:     {len(sources_zero)}\n"
        f"Fontes essenciais com falha: {len(essential_fail)}\n"
        f"Freshness gate:  {freshness_str}\n"
        f"Excel:           {excel_status}\n"
        f"PDF:             {pdf_status}\n"
        f"Wall clock:      {wall_dur:.0f}ms ({wall_dur / 1000:.1f}s)"
    )

    _echo("")
    if HAS_RICH:
        _echo(Panel.fit(summary_text))
    else:
        _echo(f"\n{'=' * 60}")
        _echo("  GOLDEN PATH — RESUMO")
        _echo(f"  Mode: {mode_str}")
        _echo(f"  Status: {overall.upper()}")
        _echo(f"  Exit: {exit_code}")
        _echo(f"  Wall clock: {wall_dur:.0f}ms ({wall_dur / 1000:.1f}s)")
        _echo(f"{'=' * 60}")

    # ── Persist ledger ──
    _save_final_ledger(
        run_id,
        timestamp,
        overall,
        steps,
        source_records,
        report_records,
        freshness_record,
        wall_start,
        args.ledger_output,
    )

    # ── Exit ──
    if exit_code == 0:
        if overall == "success_zero":
            _echo("\nGolden Path: success_zero (sem registros, gates OK).", "warn")
        else:
            _echo("\nGolden Path concluido com sucesso!", "ok")
        return 0

    messages = {
        1: "FALHA/EMPTY: sem dados utilizáveis ou essential zero sem --allow-zero",
        2: "PARCIAL: fontes essenciais falharam",
        3: "FRESHNESS FAIL: gate de freshness reprovado (strict)",
        4: "REPORT FAIL: Excel/PDF ou relatório de domínio (editais/contratos/concorrentes) falhou",
        5: "DEGRADED: fontes não essenciais falharam (strict)",
    }
    _echo(f"\n{messages.get(exit_code, f'Exit {exit_code}')}. Verifique o ledger.", "error")
    return exit_code


if __name__ == "__main__":
    sys.exit(main())
