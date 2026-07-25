"""Generate PDF, HTML, Markdown and XLSX audit workbook from same model."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from scripts.budget_audit.case_store import read_json, read_jsonl, utc_now, write_json
from scripts.budget_audit.export_safety import safe_cell_value
from scripts.budget_audit.hashing import sha256_file, sha256_text


def _load_case_model(case_dir: Path) -> dict[str, Any]:
    manifest = read_json(case_dir / "case-manifest.json")
    audits_dir = case_dir / "audits"
    reports: dict[str, Any] = {"manifest": manifest}
    for name in (
        "arithmetic",
        "workbook-integrity",
        "quantities",
        "unit-prices",
        "compositions",
        "bdi",
        "social-charges",
        "cross-sheet",
        "cross-workbook",
        "official-references",
        "findings",
        "risk-register",
    ):
        p = audits_dir / f"{name}.json"
        if p.is_file():
            reports[name.replace("-", "_")] = read_json(p)
    items = read_jsonl(case_dir / "normalized" / "budget-items.jsonl")
    reports["budget_items"] = items
    reports["bdi_components"] = read_jsonl(case_dir / "normalized" / "bdi-components.jsonl")
    # alias keys used by PDF sections
    if "workbook_integrity" not in reports and "workbook-integrity" in reports:
        reports["workbook_integrity"] = reports["workbook-integrity"]
    return reports


def _model_primary_totals(model: dict[str, Any]) -> dict[str, Any]:
    """Primary monetary observables shared across report formats."""
    items = model.get("budget_items") or []
    total_sale = 0.0
    total_direct = 0.0
    n_sale = n_direct = 0
    for it in items:
        ts = it.get("total_sale_price")
        td = it.get("total_direct_cost")
        if isinstance(ts, (int, float)):
            total_sale += float(ts)
            n_sale += 1
        if isinstance(td, (int, float)):
            total_direct += float(td)
            n_direct += 1
    findings = model.get("findings") or {}
    arith = model.get("arithmetic") or {}
    bdi = model.get("bdi") or {}
    return {
        "item_count": len(items),
        "finding_count": findings.get("finding_count") or 0,
        "severity_counts": findings.get("severity_counts") or {},
        "total_sale_price_sum": round(total_sale, 2) if n_sale else None,
        "total_direct_cost_sum": round(total_direct, 2) if n_direct else None,
        "items_with_sale_total": n_sale,
        "items_with_direct_total": n_direct,
        "arithmetic_check_count": arith.get("check_count") or 0,
        "arithmetic_status_counts": arith.get("status_counts") or {},
        "bdi_component_count": bdi.get("component_count")
        or len(model.get("bdi_components") or []),
        "bdi_sum_percent_points": bdi.get("sum_percent_points"),
    }


def build_executive_summary(model: dict[str, Any]) -> str:
    m = model.get("manifest") or {}
    findings = model.get("findings") or {}
    bdi = model.get("bdi") or {}
    totals = _model_primary_totals(model)
    lines = [
        "# Executive Summary — Budget Audit",
        "",
        f"- **Case ID:** {m.get('case_id')}",
        f"- **Generated:** {utc_now()}",
        f"- **Documents:** {len(m.get('documents') or [])}",
        f"- **Budget items:** {totals['item_count']}",
        f"- **Findings:** {totals['finding_count']}",
        f"- **Severity:** {totals['severity_counts']}",
        f"- **Arithmetic checks:** {totals['arithmetic_check_count']} ({totals['arithmetic_status_counts']})",
        f"- **total_sale_price_sum={totals['total_sale_price_sum']}**",
        f"- **total_direct_cost_sum={totals['total_direct_cost_sum']}**",
        f"- **BDI components:** {totals['bdi_component_count']}",
        f"- **bdi_sum_percent_points={totals['bdi_sum_percent_points']}**",
        "",
        "## Scope & limitations",
        "",
        "- Human remains responsible for price formation, margin, competitive strategy and professional seal.",
        "- System does not invent BDI, margins, win probability or internal costs.",
        "- Missing formula cache is never treated as zero.",
        "- Official references require explicit manifest (system, month, locality, regime).",
        "",
        "## Top findings",
        "",
    ]
    for f in (findings.get("findings") or [])[:15]:
        cells = ", ".join(f.get("cells") or []) or "—"
        lines.append(
            f"- **{f.get('severity')}** `{f.get('finding_id')}` {f.get('title')} — cells: {cells}"
        )
    lines.append("")
    lines.append("## Non-claims")
    lines.append("")
    for nc in (bdi.get("non_claims") or []):
        lines.append(f"- {nc}")
    lines.append("- Inexequibility conclusion")
    lines.append("- Legal compliance of BDI")
    lines.append("- Optimal bid suggestion")
    lines.append("")
    return "\n".join(lines)


def write_markdown(case_dir: Path, model: dict[str, Any] | None = None) -> Path:
    model = model or _load_case_model(case_dir)
    text = build_executive_summary(model)
    out = case_dir / "reports" / "executive-summary.md"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(text, encoding="utf-8")
    return out


def write_html(case_dir: Path, model: dict[str, Any] | None = None) -> Path:
    model = model or _load_case_model(case_dir)
    m = model.get("manifest") or {}
    findings = model.get("findings") or {}
    totals = _model_primary_totals(model)
    rows = []
    for f in findings.get("findings") or []:
        cells = ", ".join(f.get("cells") or [])
        rows.append(
            "<tr>"
            f"<td>{f.get('finding_id')}</td>"
            f"<td>{f.get('severity')}</td>"
            f"<td>{f.get('classification')}</td>"
            f"<td>{f.get('title')}</td>"
            f"<td>{cells}</td>"
            f"<td>{f.get('reported_value')}</td>"
            f"<td>{f.get('recomputed_value')}</td>"
            "</tr>"
        )
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8"/>
<title>Budget Audit — {m.get('case_id')}</title>
<style>
body {{ font-family: system-ui, sans-serif; margin: 2rem; color: #122; }}
h1,h2 {{ color: #0b3d5c; }}
table {{ border-collapse: collapse; width: 100%; margin: 1rem 0; }}
th, td {{ border: 1px solid #ccc; padding: 6px 8px; font-size: 13px; }}
th {{ background: #e8f1f8; }}
.meta {{ background: #f7fafc; padding: 1rem; border-radius: 8px; }}
.badge {{ display:inline-block; padding:2px 8px; border-radius:4px; background:#dde; }}
</style>
</head>
<body>
<h1>Budget Audit Report</h1>
<div class="meta">
<p><strong>Case:</strong> {m.get('case_id')}</p>
<p><strong>Generated:</strong> {utc_now()}</p>
<p><strong>Items:</strong> {totals['item_count']}</p>
<p><strong>Findings:</strong> {totals['finding_count']}
<span class="badge">{totals['severity_counts']}</span></p>
<p><strong>total_sale_price_sum:</strong> {totals['total_sale_price_sum']}</p>
<p><strong>total_direct_cost_sum:</strong> {totals['total_direct_cost_sum']}</p>
<p><strong>arithmetic_check_count:</strong> {totals['arithmetic_check_count']}</p>
</div>
<h2>Findings</h2>
<table>
<thead><tr><th>ID</th><th>Severity</th><th>Class</th><th>Title</th><th>Cells</th><th>Reported</th><th>Recomputed</th></tr></thead>
<tbody>
{''.join(rows) if rows else '<tr><td colspan="7">No findings</td></tr>'}
</tbody>
</table>
<h2>Limitations</h2>
<ul>
<li>Human engineer remains responsible for interpretation and professional seal.</li>
<li>No invented margins, win probability or optimal bid.</li>
<li>Missing formula cache never treated as zero.</li>
</ul>
</body>
</html>
"""
    out = case_dir / "reports" / "budget-audit-report.html"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html, encoding="utf-8")
    return out


def write_pdf(case_dir: Path, model: dict[str, Any] | None = None) -> Path:
    model = model or _load_case_model(case_dir)
    m = model.get("manifest") or {}
    findings = model.get("findings") or {}
    arith = model.get("arithmetic") or {}
    bdi = model.get("bdi") or {}
    compositions = model.get("compositions") or {}
    totals = _model_primary_totals(model)
    out = case_dir / "reports" / "budget-audit-report.pdf"
    out.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(str(out), pagesize=A4)
    styles = getSampleStyleSheet()
    story: list[Any] = []

    def h(text: str) -> None:
        story.append(Paragraph(text, styles["Heading2"]))

    def p(text: str) -> None:
        story.append(Paragraph(text.replace("\n", "<br/>"), styles["Normal"]))
        story.append(Spacer(1, 6))

    story.append(Paragraph("Budget Audit Report", styles["Title"]))
    story.append(Spacer(1, 12))
    h("1. Cover")
    p("Evidence-first engineering budget / composition / BDI audit.")
    h("2. Case identification")
    p(
        f"Case: {m.get('case_id')}<br/>Generated: {utc_now()}<br/>"
        f"Documents: {len(m.get('documents') or [])}"
    )
    h("3. Executive summary")
    p(
        f"Items: {totals['item_count']}<br/>"
        f"Findings: {totals['finding_count']} — {totals['severity_counts']}<br/>"
        f"Arithmetic checks: {totals['arithmetic_check_count']} — {totals['arithmetic_status_counts']}<br/>"
        f"Sum total_sale_price (items with value): {totals['total_sale_price_sum']}<br/>"
        f"Sum total_direct_cost (items with value): {totals['total_direct_cost_sum']}<br/>"
        f"BDI components: {totals['bdi_component_count']}; sum_pct_points={totals['bdi_sum_percent_points']}"
    )
    h("4. Scope and limitations")
    p(
        "Human remains responsible for price formation, margin, competitive strategy and seal. "
        "Missing formula cache is never treated as zero. BDI is not margin. "
        "No win probability or optimal bid."
    )
    h("5. Files and sheets")
    for d in m.get("documents") or []:
        p(
            f"Document {d.get('document_id')}: {d.get('original_name')} "
            f"sha256={d.get('sha256')} cells={d.get('cell_count')}"
        )
    h("6. Extraction quality")
    p(f"See workbook extraction-quality artifacts. Case phase={m.get('phase')}.")
    h("7. Budget summary")
    p(
        f"Normalized items={totals['item_count']}; "
        f"with sale totals={totals['items_with_sale_total']}; "
        f"with direct totals={totals['items_with_direct_total']}."
    )
    h("8. Formula integrity")
    integ = model.get("workbook_integrity") or {}
    p(f"Integrity issue_count={integ.get('issue_count')}; formula_issues tracked in audits.")
    h("9. Arithmetic reconciliation")
    p(
        f"Checks={arith.get('check_count')}; status_counts={arith.get('status_counts')}; "
        f"policy={arith.get('policy')}"
    )
    h("10. Quantities")
    p(f"Quantity audit: {(model.get('quantities') or {}).get('issue_count')} issues")
    h("11. Prices")
    p(f"Unit-price audit note: {(model.get('unit_prices') or {}).get('note')}")
    h("12. Compositions")
    p(
        f"Compositions={compositions.get('composition_count') if isinstance(compositions, dict) else 'n/a'}; "
        f"inputs={compositions.get('input_count') if isinstance(compositions, dict) else 'n/a'}; "
        f"issues={compositions.get('issue_count') if isinstance(compositions, dict) else 'n/a'}"
    )
    h("13. Social charges")
    p(f"Social charges: {(model.get('social_charges') or {}).get('component_count', 'see audits')}")
    h("14. BDI")
    p(
        f"Components={bdi.get('component_count')}; sum_percent_points={bdi.get('sum_percent_points')}; "
        f"non_claims={bdi.get('non_claims')}"
    )
    for c in (bdi.get("components") or model.get("bdi_components") or [])[:20]:
        p(
            f"- {c.get('original_name')}: raw={c.get('percentage')} "
            f"frac={c.get('fraction')} rule={c.get('interpretation_rule')}"
        )
    h("15. Schedule")
    p("Schedule reconciliation in audits/ when schedule items present.")
    h("16. ABC curve")
    p("ABC materiality classes are not error/overprice labels.")
    h("17. Official references")
    p(f"References: {(model.get('official_references') or {}).get('status', 'see audits')}")
    h("18. Risks")
    risks = model.get("risk_register") or {}
    p(f"Risk count={risks.get('risk_count')}; non_claims={risks.get('non_claims')}")
    h("19. Prioritized findings")
    for f in (findings.get("findings") or [])[:60]:
        cells = ", ".join(f.get("cells") or []) or "—"
        p(
            f"<b>{f.get('finding_id')}</b> [{f.get('severity')}] {f.get('classification')}: "
            f"{f.get('title')} — cells: {cells} — reported={f.get('reported_value')} "
            f"recomputed={f.get('recomputed_value')}"
        )
    h("20. Human pending items")
    p("Review HIGH findings, BDI methodology, and any NEEDS_ENGINEER_REVIEW items.")
    h("21. Evidence annex")
    p(
        f"Primary totals for reconciliation: item_count={totals['item_count']}, "
        f"finding_count={totals['finding_count']}, "
        f"total_sale_price_sum={totals['total_sale_price_sum']}, "
        f"total_direct_cost_sum={totals['total_direct_cost_sum']}, "
        f"arithmetic_check_count={totals['arithmetic_check_count']}."
    )
    # Explicit machine-readable footer (kept as plain paragraphs for recon)
    p(
        "RECON_FOOTER "
        f"item_count={totals['item_count']} "
        f"finding_count={totals['finding_count']} "
        f"total_sale_price_sum={totals['total_sale_price_sum']} "
        f"total_direct_cost_sum={totals['total_direct_cost_sum']} "
        f"arithmetic_check_count={totals['arithmetic_check_count']}"
    )
    doc.build(story)
    # Sidecar with same primary totals (PDF streams may compress body text)
    write_json(
        case_dir / "reports" / "budget-audit-report.pdf.json",
        {"format": "pdf", "derived_from": "case_model", "primary_totals": totals},
    )
    return out


def write_xlsx(case_dir: Path, model: dict[str, Any] | None = None) -> Path:
    model = model or _load_case_model(case_dir)
    m = model.get("manifest") or {}
    out = case_dir / "reports" / "budget-audit-workbook.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    def sheet(name: str):
        if name == "Metadados":
            ws = wb.active
            ws.title = name
            return ws
        return wb.create_sheet(name)

    totals = _model_primary_totals(model)
    ws = sheet("Metadados")
    ws.append(["key", "value"])
    for k, v in {
        "case_id": m.get("case_id"),
        "generated_at": utc_now(),
        "item_count": totals["item_count"],
        "finding_count": totals["finding_count"],
        "total_sale_price_sum": totals["total_sale_price_sum"],
        "total_direct_cost_sum": totals["total_direct_cost_sum"],
        "arithmetic_check_count": totals["arithmetic_check_count"],
        "bdi_component_count": totals["bdi_component_count"],
        "bdi_sum_percent_points": totals["bdi_sum_percent_points"],
    }.items():
        ws.append([safe_cell_value(k), safe_cell_value(v)])

    ws = sheet("Arquivos")
    ws.append(["document_id", "original_name", "sha256", "extension"])
    for d in m.get("documents") or []:
        ws.append(
            [
                safe_cell_value(d.get("document_id")),
                safe_cell_value(d.get("original_name")),
                safe_cell_value(d.get("sha256")),
                safe_cell_value(d.get("extension")),
            ]
        )

    ws = sheet("Itens")
    headers = [
        "item_id", "code", "description", "unit", "quantity",
        "unit_sale_price", "total_sale_price", "sheet", "row",
    ]
    ws.append(headers)
    for it in model.get("budget_items") or []:
        ws.append([safe_cell_value(it.get(h)) for h in headers])

    ws = sheet("BDI")
    ws.append(["component_id", "original_name", "percentage", "sheet", "row"])
    for c in model.get("bdi_components") or []:
        ws.append(
            [
                safe_cell_value(c.get("component_id")),
                safe_cell_value(c.get("original_name")),
                safe_cell_value(c.get("percentage")),
                safe_cell_value(c.get("sheet")),
                safe_cell_value(c.get("row")),
            ]
        )

    ws = sheet("Achados")
    ws.append(
        [
            "finding_id", "severity", "classification", "title",
            "cells", "reported_value", "recomputed_value", "difference",
        ]
    )
    for f in (model.get("findings") or {}).get("findings") or []:
        ws.append(
            [
                safe_cell_value(f.get("finding_id")),
                safe_cell_value(f.get("severity")),
                safe_cell_value(f.get("classification")),
                safe_cell_value(f.get("title")),
                safe_cell_value(", ".join(f.get("cells") or [])),
                safe_cell_value(f.get("reported_value")),
                safe_cell_value(f.get("recomputed_value")),
                safe_cell_value(f.get("difference")),
            ]
        )

    ws = sheet("Riscos")
    ws.append(["risk_id", "signal", "title"])
    for r in (model.get("risk_register") or {}).get("risks") or []:
        ws.append(
            [
                safe_cell_value(r.get("risk_id")),
                safe_cell_value(r.get("signal")),
                safe_cell_value(r.get("title")),
            ]
        )

    # Required sheet names (may be empty)
    for name in (
        "Abas", "Composições", "Insumos", "Encargos", "Cronograma", "ABC",
        "Fórmulas", "Reconciliação", "Referências", "Evidências", "Limitações",
    ):
        if name not in wb.sheetnames:
            sheet(name)

    wb["Limitações"].append(["limitation"])
    for line in [
        "Human engineer remains responsible",
        "No invented margins or win probability",
        "Missing cache never treated as zero",
        "Difference % is not proof of overprice",
    ]:
        wb["Limitações"].append([safe_cell_value(line)])

    wb.save(out)
    return out


def reconcile_reports(case_dir: Path) -> dict[str, Any]:
    """Ensure primary observables match across JSON / MD / HTML / XLSX / PDF derivatives."""
    model = _load_case_model(case_dir)
    totals = _model_primary_totals(model)
    finding_count = totals["finding_count"]
    item_count = totals["item_count"]
    sale_sum = totals["total_sale_price_sum"]
    direct_sum = totals["total_direct_cost_sum"]
    arith_n = totals["arithmetic_check_count"]

    md = case_dir / "reports" / "executive-summary.md"
    html = case_dir / "reports" / "budget-audit-report.html"
    pdf = case_dir / "reports" / "budget-audit-report.pdf"
    xlsx = case_dir / "reports" / "budget-audit-workbook.xlsx"

    issues: list[str] = []
    checks: dict[str, Any] = {"primary_totals": totals}

    def _require_tokens(label: str, text: str, tokens: list[str]) -> None:
        for tok in tokens:
            if tok is None:
                continue
            if str(tok) not in text:
                issues.append(f"{label}_missing_token:{tok}")

    if md.is_file():
        text = md.read_text(encoding="utf-8")
        _require_tokens("markdown", text, [str(finding_count), str(item_count)])
        if sale_sum is not None:
            # markdown may not print sale sum yet — require after we embed it
            if f"total_sale_price_sum={sale_sum}" not in text and str(sale_sum) not in text:
                issues.append("markdown_sale_sum_mismatch")
    else:
        issues.append("markdown_missing")

    if html.is_file():
        h = html.read_text(encoding="utf-8")
        _require_tokens("html", h, [str(finding_count), str(item_count)])
        if sale_sum is not None and str(sale_sum) not in h:
            issues.append("html_sale_sum_mismatch")
    else:
        issues.append("html_missing")

    if not pdf.is_file():
        issues.append("pdf_missing")
    else:
        # Full §33.1 PDF is multi-section; pre-fix stubs were ~3-4KB
        if pdf.stat().st_size < 6000:
            issues.append("pdf_too_small_stub")
        sidecar = case_dir / "reports" / "budget-audit-report.pdf.json"
        if not sidecar.is_file():
            issues.append("pdf_sidecar_missing")
        else:
            side = read_json(sidecar)
            st = (side.get("primary_totals") or {})
            if st.get("item_count") != item_count:
                issues.append("pdf_sidecar_item_count_mismatch")
            if st.get("finding_count") != finding_count:
                issues.append("pdf_sidecar_finding_count_mismatch")
            if sale_sum is not None and st.get("total_sale_price_sum") != sale_sum:
                issues.append("pdf_sidecar_sale_sum_mismatch")
            if st.get("arithmetic_check_count") != arith_n:
                issues.append("pdf_sidecar_arith_count_mismatch")

    if not xlsx.is_file():
        issues.append("xlsx_missing")
    else:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        try:
            # Metadados sheet should carry item_count / finding_count
            if "Metadados" in wb.sheetnames:
                meta = {str(r[0]): r[1] for r in wb["Metadados"].iter_rows(values_only=True) if r and r[0]}
                if meta.get("item_count") not in (item_count, str(item_count)):
                    issues.append("xlsx_item_count_mismatch")
                if meta.get("finding_count") not in (finding_count, str(finding_count)):
                    issues.append("xlsx_finding_count_mismatch")
                if sale_sum is not None:
                    if meta.get("total_sale_price_sum") not in (sale_sum, str(sale_sum)):
                        issues.append("xlsx_sale_sum_mismatch")
            else:
                issues.append("xlsx_metadados_missing")
            for required in ("Itens", "BDI", "Achados", "Limitações"):
                if required not in wb.sheetnames:
                    issues.append(f"xlsx_sheet_missing:{required}")
        finally:
            wb.close()

    checksums = {}
    for path in (md, html, pdf, xlsx):
        if path.is_file():
            checksums[path.name] = sha256_file(path)

    status = "PASS" if not issues else "FAIL"
    result = {
        "status": status,
        "finding_count": finding_count,
        "item_count": item_count,
        "total_sale_price_sum": sale_sum,
        "total_direct_cost_sum": direct_sum,
        "arithmetic_check_count": arith_n,
        "issues": issues,
        "checks": checks,
        "checksums": checksums,
        "model_hash": sha256_text(
            "|".join(
                str(x)
                for x in (
                    finding_count,
                    item_count,
                    sale_sum,
                    direct_sum,
                    arith_n,
                    totals["severity_counts"],
                )
            )
        ),
    }
    write_json(case_dir / "reports" / "report-reconciliation.json", result)
    return result


def generate_all_reports(case_dir: Path) -> dict[str, Any]:
    model = _load_case_model(case_dir)
    paths = {
        "markdown": str(write_markdown(case_dir, model)),
        "html": str(write_html(case_dir, model)),
        "pdf": str(write_pdf(case_dir, model)),
        "xlsx": str(write_xlsx(case_dir, model)),
    }
    recon = reconcile_reports(case_dir)
    return {"paths": paths, "reconciliation": recon}
