"""Safe workbook reading — formulas and cached values kept separate.

Never recalculates complex formulas pretending Excel equivalence.
Never treats missing cache as zero.
"""

from __future__ import annotations

import csv
import re
from collections.abc import Iterator
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from scripts.budget_audit.constants import (
    MAX_CELLS_PER_SHEET,
    MAX_SHEETS,
    MAX_TOTAL_CELLS,
)
from scripts.budget_audit.hashing import sha256_file

try:
    import openpyxl as _openpyxl_mod

    OPENPYXL_VERSION = getattr(_openpyxl_mod, "__version__", "unknown")
except Exception:  # pragma: no cover
    OPENPYXL_VERSION = "unknown"

_ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!")
_EXTERNAL_RE = re.compile(r"\[.+?\]|https?://|\\\\|'[A-Za-z]:\\")
_BROKEN_REF_RE = re.compile(r"#REF!", re.I)


class WorkbookReadError(RuntimeError):
    pass


def _cell_data_type(value: Any, number_format: str | None) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        fmt = (number_format or "").lower()
        if "%" in fmt:
            return "percent"
        if any(tok in fmt for tok in ("yy", "mm", "dd", "date", "h:")):
            return "date_serial_or_number"
        return "number"
    if isinstance(value, (datetime, date)):
        return "date"
    if isinstance(value, str):
        if value in _ERROR_TOKENS or any(value.startswith(t[:4]) for t in _ERROR_TOKENS):
            return "error"
        return "text"
    return type(value).__name__


def classify_formula(formula: str | None, cached_value: Any) -> str:
    if not formula:
        return "NOT_EVALUATED"
    f = formula if formula.startswith("=") else f"={formula}"
    if _BROKEN_REF_RE.search(f) or (isinstance(cached_value, str) and "#REF!" in cached_value):
        return "BROKEN_REFERENCE"
    if _EXTERNAL_RE.search(f):
        return "EXTERNAL_REFERENCE"
    if isinstance(cached_value, str) and any(tok in cached_value for tok in _ERROR_TOKENS):
        return "BROKEN_REFERENCE"
    if cached_value is None:
        return "MISSING_CACHE"
    return "VALID"


def _serialize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def iter_cells(ws, *, max_cells: int = MAX_CELLS_PER_SHEET) -> Iterator[dict[str, Any]]:
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None and not getattr(cell, "data_type", None) == "f":
                # skip completely empty non-formula cells to control size
                # but still we need hidden/merged awareness later
                continue
            count += 1
            if count > max_cells:
                raise WorkbookReadError(
                    f"sheet {ws.title!r} exceeds MAX_CELLS_PER_SHEET={max_cells}"
                )
            yield cell_to_record(cell, ws.title)


def cell_to_record(cell: Cell, sheet_name: str) -> dict[str, Any]:
    formula = None
    cached_value = None
    raw = cell.value

    if isinstance(raw, str) and raw.startswith("="):
        formula = raw
        # data_only workbook would have cached; here we are formula mode
        cached_value = None
        formula_status = classify_formula(formula, cached_value)
    elif cell.data_type == "f":
        formula = str(raw) if raw is not None else None
        cached_value = None
        formula_status = classify_formula(formula, cached_value)
    else:
        cached_value = raw
        formula_status = "NOT_EVALUATED" if formula else "VALID"

    number_format = cell.number_format
    data_type = _cell_data_type(cached_value if formula is None else cached_value, number_format)

    # display approximation — never invent
    displayed_value = _serialize_value(cached_value)

    return {
        "sheet": sheet_name,
        "coordinate": cell.coordinate,
        "row": cell.row,
        "column": cell.column,
        "column_letter": get_column_letter(cell.column) if cell.column else None,
        "formula": formula,
        "cached_value": _serialize_value(cached_value),
        "displayed_value": displayed_value,
        "data_type": data_type,
        "number_format": number_format,
        "formula_status": formula_status if formula else None,
        "is_hidden_row": bool(getattr(cell.parent.row_dimensions.get(cell.row), "hidden", False))
        if cell.parent is not None
        else False,
        "is_hidden_column": False,  # filled later if needed
    }


def read_xlsx(
    path: Path | str,
    *,
    document_id: str | None = None,
) -> dict[str, Any]:
    """Read XLSX/XLSM without executing macros. Dual pass for formulas + cached values."""
    path = Path(path)
    if not path.is_file():
        raise WorkbookReadError(f"file not found: {path}")

    # Formula pass
    try:
        wb_f = load_workbook(path, data_only=False, read_only=False, keep_vba=False)
    except Exception as exc:
        raise WorkbookReadError(f"failed to open workbook: {exc}") from exc

    if len(wb_f.sheetnames) > MAX_SHEETS:
        wb_f.close()
        raise WorkbookReadError(f"too many sheets: {len(wb_f.sheetnames)}")

    # Cached values pass
    try:
        wb_v = load_workbook(path, data_only=True, read_only=False, keep_vba=False)
    except Exception:
        wb_v = None

    sheets_meta: list[dict[str, Any]] = []
    cells: list[dict[str, Any]] = []
    formulas: list[dict[str, Any]] = []
    merged: list[dict[str, Any]] = []
    names: list[dict[str, Any]] = []
    hidden_content: list[dict[str, Any]] = []
    total_cells = 0
    warnings: list[str] = []

    for sheet_name in wb_f.sheetnames:
        ws = wb_f[sheet_name]
        ws_v = wb_v[sheet_name] if wb_v is not None and sheet_name in wb_v.sheetnames else None
        state = "visible"
        if ws.sheet_state and ws.sheet_state != "visible":
            state = str(ws.sheet_state)
            hidden_content.append({"type": "sheet", "sheet": sheet_name, "state": state})

        for merged_range in ws.merged_cells.ranges:
            merged.append({"sheet": sheet_name, "range": str(merged_range)})

        sheet_cell_count = 0
        # build value map from data_only
        value_map: dict[str, Any] = {}
        if ws_v is not None:
            for row in ws_v.iter_rows():
                for c in row:
                    if c.value is not None:
                        value_map[c.coordinate] = c.value

        for row in ws.iter_rows():
            for cell in row:
                has_formula = isinstance(cell.value, str) and str(cell.value).startswith("=")
                if cell.value is None and not has_formula:
                    continue
                sheet_cell_count += 1
                total_cells += 1
                if sheet_cell_count > MAX_CELLS_PER_SHEET:
                    raise WorkbookReadError(
                        f"sheet {sheet_name!r} exceeds cell limit"
                    )
                if total_cells > MAX_TOTAL_CELLS:
                    raise WorkbookReadError("workbook exceeds total cell limit")

                rec = cell_to_record(cell, sheet_name)
                # merge cached value from data_only pass
                if rec["formula"]:
                    cached = value_map.get(cell.coordinate)
                    rec["cached_value"] = _serialize_value(cached)
                    rec["displayed_value"] = rec["cached_value"]
                    rec["formula_status"] = classify_formula(rec["formula"], cached)
                    # check inconsistency: if both formula and we could re-evaluate simple cases later
                    formulas.append(
                        {
                            "sheet": sheet_name,
                            "coordinate": cell.coordinate,
                            "formula": rec["formula"],
                            "cached_value": rec["cached_value"],
                            "formula_status": rec["formula_status"],
                        }
                    )
                # hidden column
                col_dim = ws.column_dimensions.get(get_column_letter(cell.column))
                if col_dim is not None and col_dim.hidden:
                    rec["is_hidden_column"] = True
                    hidden_content.append(
                        {
                            "type": "column",
                            "sheet": sheet_name,
                            "column": get_column_letter(cell.column),
                        }
                    )
                if rec.get("is_hidden_row"):
                    hidden_content.append(
                        {
                            "type": "row",
                            "sheet": sheet_name,
                            "row": cell.row,
                        }
                    )
                cells.append(rec)

        sheets_meta.append(
            {
                "name": sheet_name,
                "state": state,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "cell_count": sheet_cell_count,
                "dimensions": ws.dimensions,
            }
        )

    # defined names
    try:
        for dn in wb_f.defined_names.definedName:
            names.append(
                {
                    "name": dn.name,
                    "attr_text": dn.attr_text,
                    "value": str(dn.attr_text),
                }
            )
    except Exception as exc:
        warnings.append(f"defined names incomplete: {exc}")

    has_vba = path.suffix.lower() == ".xlsm"
    if has_vba:
        warnings.append("XLSM detected: macros NOT executed (keep_vba=False)")

    wb_f.close()
    if wb_v is not None:
        wb_v.close()

    missing_cache = sum(1 for f in formulas if f["formula_status"] == "MISSING_CACHE")
    broken = sum(1 for f in formulas if f["formula_status"] == "BROKEN_REFERENCE")
    external = sum(1 for f in formulas if f["formula_status"] == "EXTERNAL_REFERENCE")

    quality = {
        "parser": "openpyxl",
        "parser_version": OPENPYXL_VERSION,
        "formula_mode": "stored_formula + cached_value dual-pass",
        "calculation_mode": "NOT_RECALCULATED",
        "sheet_count": len(sheets_meta),
        "cell_count": total_cells,
        "formula_count": len(formulas),
        "missing_cache_count": missing_cache,
        "broken_reference_count": broken,
        "external_reference_count": external,
        "merged_range_count": len(merged),
        "hidden_items": len(hidden_content),
        "warnings": warnings,
        "macros_executed": False,
        "external_links_followed": False,
    }

    return {
        "document_id": document_id or path.stem,
        "source_path": str(path),
        "sha256": sha256_file(path),
        "extension": path.suffix.lower(),
        "workbook": {
            "sheetnames": [s["name"] for s in sheets_meta],
            "sheet_count": len(sheets_meta),
            "has_vba_container": has_vba,
            "parser": "openpyxl",
            "parser_version": OPENPYXL_VERSION,
        },
        "sheets": sheets_meta,
        "cells": cells,
        "formulas": formulas,
        "merged_ranges": merged,
        "names": names,
        "hidden_content": hidden_content,
        "extraction_quality": quality,
        "warnings": warnings,
    }


def read_csv(path: Path | str, *, document_id: str | None = None) -> dict[str, Any]:
    """Read CSV as a single-sheet workbook model. Does not execute formulas."""
    path = Path(path)
    cells: list[dict[str, Any]] = []
    formulas: list[dict[str, Any]] = []
    warnings: list[str] = []
    with path.open(encoding="utf-8-sig", newline="") as fh:
        # detect delimiter
        sample = fh.read(4096)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(fh, dialect)
        for r_idx, row in enumerate(reader, start=1):
            for c_idx, val in enumerate(row, start=1):
                if val is None or val == "":
                    continue
                formula = None
                formula_status = None
                cached: Any = val
                if isinstance(val, str) and val.startswith("="):
                    # treat as formula text — do NOT evaluate
                    formula = val
                    cached = None
                    formula_status = "MISSING_CACHE"
                    warnings.append(
                        f"CSV formula-like value at row {r_idx} col {c_idx} not evaluated"
                    )
                    formulas.append(
                        {
                            "sheet": "CSV",
                            "coordinate": f"{get_column_letter(c_idx)}{r_idx}",
                            "formula": formula,
                            "cached_value": None,
                            "formula_status": formula_status,
                        }
                    )
                # try number parse only when not formula
                if formula is None:
                    try:
                        if "," in val and "." not in val:
                            cached = float(val.replace(".", "").replace(",", "."))
                        else:
                            cached = float(val) if re.fullmatch(r"-?\d+(\.\d+)?", val) else val
                    except ValueError:
                        cached = val
                cells.append(
                    {
                        "sheet": "CSV",
                        "coordinate": f"{get_column_letter(c_idx)}{r_idx}",
                        "row": r_idx,
                        "column": c_idx,
                        "column_letter": get_column_letter(c_idx),
                        "formula": formula,
                        "cached_value": cached if not isinstance(cached, str) or formula is None else None,
                        "displayed_value": val if formula else cached,
                        "data_type": "formula_text" if formula else _cell_data_type(cached, None),
                        "number_format": None,
                        "formula_status": formula_status,
                        "is_hidden_row": False,
                        "is_hidden_column": False,
                    }
                )

    quality = {
        "parser": "csv",
        "parser_version": "stdlib",
        "formula_mode": "text_only_no_eval",
        "calculation_mode": "NOT_EVALUATED",
        "sheet_count": 1,
        "cell_count": len(cells),
        "formula_count": len(formulas),
        "warnings": warnings,
        "macros_executed": False,
        "external_links_followed": False,
    }
    return {
        "document_id": document_id or path.stem,
        "source_path": str(path),
        "sha256": sha256_file(path),
        "extension": path.suffix.lower(),
        "workbook": {
            "sheetnames": ["CSV"],
            "sheet_count": 1,
            "has_vba_container": False,
            "parser": "csv",
            "parser_version": "stdlib",
        },
        "sheets": [
            {
                "name": "CSV",
                "state": "visible",
                "max_row": max((c["row"] for c in cells), default=0),
                "max_column": max((c["column"] for c in cells), default=0),
                "cell_count": len(cells),
            }
        ],
        "cells": cells,
        "formulas": formulas,
        "merged_ranges": [],
        "names": [],
        "hidden_content": [],
        "extraction_quality": quality,
        "warnings": warnings,
    }


def read_workbook(
    path: Path | str,
    *,
    document_id: str | None = None,
    extension: str | None = None,
) -> dict[str, Any]:
    path = Path(path)
    ext = (extension or path.suffix or "").lower()
    if ext and not ext.startswith("."):
        ext = f".{ext}"
    if ext in {".xlsx", ".xlsm"}:
        return read_xlsx(path, document_id=document_id)
    if ext == ".csv":
        return read_csv(path, document_id=document_id)
    if ext == ".xls":
        return {
            "document_id": document_id or path.stem,
            "source_path": str(path),
            "sha256": sha256_file(path),
            "extension": ext,
            "status": "CONVERSION_REQUIRED",
            "workbook": {"sheetnames": [], "sheet_count": 0},
            "sheets": [],
            "cells": [],
            "formulas": [],
            "merged_ranges": [],
            "names": [],
            "hidden_content": [],
            "extraction_quality": {
                "parser": None,
                "status": "CONVERSION_REQUIRED",
                "warnings": ["XLS legacy requires safe conversion before parse"],
            },
            "warnings": ["CONVERSION_REQUIRED"],
        }
    if ext == ".ods":
        return {
            "document_id": document_id or path.stem,
            "source_path": str(path),
            "sha256": sha256_file(path),
            "extension": ext,
            "status": "NOT_AVAILABLE",
            "workbook": {"sheetnames": [], "sheet_count": 0},
            "sheets": [],
            "cells": [],
            "formulas": [],
            "merged_ranges": [],
            "names": [],
            "hidden_content": [],
            "extraction_quality": {
                "parser": None,
                "status": "NOT_AVAILABLE",
                "warnings": ["ODS parser not bundled; install odfpy for support"],
            },
            "warnings": ["ODS_NOT_AVAILABLE"],
        }
    if ext == ".pdf":
        return {
            "document_id": document_id or path.stem,
            "source_path": str(path),
            "sha256": sha256_file(path),
            "extension": ext,
            "status": "LIMITED",
            "workbook": {"sheetnames": [], "sheet_count": 0},
            "sheets": [],
            "cells": [],
            "formulas": [],
            "merged_ranges": [],
            "names": [],
            "hidden_content": [],
            "extraction_quality": {
                "parser": None,
                "status": "LIMITED",
                "warnings": ["PDF spreadsheet mode limited — no cell model"],
            },
            "warnings": ["PDF_LIMITED"],
        }
    raise WorkbookReadError(f"unsupported extension: {ext}")
