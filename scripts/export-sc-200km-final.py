#!/usr/bin/env python3
"""
Build final enriched Excel: SC 200km engineering suppliers + contact info from Exa.
"""

import json, os
from datetime import datetime
import psycopg2, psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DSN = os.environ.get("LOCAL_DATALAKE_DSN", "")
OUTPUT = f"docs/intel/sc-200km-engenharia-contatos-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"

# ============================================================
# LOAD MERGED CONTACTS from batch processing
# ============================================================
import os as _os
CONTACTS = {}
_merged_path = '/tmp/all_contacts_merged.json'
if _os.path.exists(_merged_path):
    with open(_merged_path) as _f:
        _merged = json.load(_f)
    for _cnpj, _data in _merged.items():
        # Normalize CNPJ to 14 digits
        _clean = _cnpj.replace('.','').replace('/','').replace('-','').strip()
        if len(_clean) == 14:
            CONTACTS[_clean] = {
                'website': _data.get('website', 'N/D') or 'N/D',
                'telefone': _data.get('telefone', 'N/D') or 'N/D',
                'email': _data.get('email', 'N/D') or 'N/D',
                'ceo': _data.get('ceo_nome', 'N/D') or 'N/D',
                'dir_comercial': _data.get('commercial_director', 'N/D') or 'N/D',
                'linkedin': _data.get('linkedin', 'N/D') or 'N/D',
                'endereco': _data.get('endereco', 'N/D') or 'N/D',
            }
    print(f"Loaded {len(CONTACTS)} contacts from merged file")

# Supplement with Exa data for top companies that may be missing
_EXA_EXTRA = {
    "08827501000158": {"website":"https://aegea.com.br","telefone":"(11) 3501-5800","email":"relacionamento@aegea.com.br","ceo":"Radamés Andrade Casseb (CEO)","dir_comercial":"Yaroslav Memrava Neto (Dir. Novos Negócios)","linkedin":"linkedin.com/company/aegea-saneamento","endereco":"Av. Brigadeiro Faria Lima, 1663, São Paulo/SP"},
    "02955426000124": {"website":"https://neoviaengenharia.com.br","telefone":"(41) 3075-9000","email":"fiscal@neoviaengenharia.com.br","ceo":"Álvaro Bermudez Peña (Dir. Presidente)","dir_comercial":"N/D","linkedin":"linkedin.com/company/neovia-engenharia","endereco":"Curitiba/PR"},
    "79283065000907": {"website":"https://orbenk.com.br","telefone":"(47) 3461-4200","email":"sac@orbenk.com.br","ceo":"Ronaldo Benkendorf (Presidente)","dir_comercial":"Ricardo Wasem (Dir. Comercial)","linkedin":"N/D","endereco":"Rua Dona Leopoldina, 26, Joinville/SC"},
    "83566299000173": {"website":"https://fepese.org.br","telefone":"(48) 3953-1000","email":"fepese@fepese.org.br","ceo":"Raimundo Lima (Presidente)","dir_comercial":"Marcelino Hirofumi Ito (Superintendente)","linkedin":"linkedin.com/company/fepese","endereco":"Campus UFSC, Florianópolis/SC"},
    "19758842000135": {"website":"https://lcmconstrucao.com.br","telefone":"(31) 3298-7050","email":"compras.bhz@lcmconstrucao.com.br","ceo":"Luiz Otávio Fontes Junqueira (Presidente)","dir_comercial":"N/D","linkedin":"linkedin.com/company/lcm-construcao-e-comercio-s-a","endereco":"Rua Pólos, 152, Belo Horizonte/MG"},
    "83476911000117": {"website":"https://fapeu.org.br","telefone":"(48) 3331-7400","email":"projetos@fapeu.org.br","ceo":"Leandro L. de Oliveira (Dir. Executivo)","dir_comercial":"N/D","linkedin":"linkedin.com/company/fapeu","endereco":"Florianópolis/SC"},
    "82743832000162": {"website":"https://planaterra.com.br","telefone":"(49) 3321-1924","email":"vendas@planaterra.com.br","ceo":"Gerson de Borba Dias (Sócio-Admin)","dir_comercial":"N/D","linkedin":"linkedin.com/company/planaterra","endereco":"Rua Blumenau, 20-D, Chapecó/SC"},
    "00820854000114": {"website":"N/D","telefone":"(48) 3374-2655","email":"qualidade@qualidademineracao.com.br","ceo":"Gustavo Pimenta (Sócio-Admin)","dir_comercial":"N/D","linkedin":"linkedin.com/company/qualidademineracaoltda","endereco":"Rua da Praça, 241, Palhoça/SC"},
    "03257777000124": {"website":"https://gaiarodovias.com.br","telefone":"(49) 3664-2022","email":"gaiarodovias@gaiarodovias.com.br","ceo":"N/D","dir_comercial":"N/D","linkedin":"linkedin.com/company/gaia-rodovias","endereco":"Rua Duque de Caxias, 626, Maravilha/SC"},
    "82094640000172": {"website":"N/D","telefone":"(48) 3348-6950","email":"quantum@quantumengenharia.com.br","ceo":"Vania (proprietária)","dir_comercial":"N/D","linkedin":"linkedin.com/in/vania-quantum-eletrica","endereco":"Rua Santarém, 219, Blumenau/SC"},
    "13534698000177": {"website":"https://sinart.com.br","telefone":"N/D","email":"sinartpassagens@clickbus.com.br","ceo":"N/D","dir_comercial":"N/D","linkedin":"N/D","endereco":"Juiz de Fora/MG"},
    "19758779000137": {"website":"https://ethosinfraestrutura.com.br","telefone":"(31) 2555-3535","email":"ethos@ethosinfraestrutura.com.br","ceo":"Juliane Leite (Diretora)","dir_comercial":"N/D","linkedin":"linkedin.com/in/juliane-leite","endereco":"Belo Horizonte/MG"},
    "05901551000140": {"website":"N/D","telefone":"(49) 2049-0900","email":"N/D","ceo":"N/D","dir_comercial":"N/D","linkedin":"N/D","endereco":"N/D"},
    "03094645000129": {"website":"N/D","telefone":"(47) 4734-2264","email":"engenharia@infracal.com.br","ceo":"Luiz Antonio V. P. Cerqueira (Admin)","dir_comercial":"N/D","linkedin":"N/D","endereco":"Rua Gothard Kaesemodel, 254, Joinville/SC"},
    "82895327000133": {"website":"https://feesc.org.br","telefone":"(48) 3231-4400","email":"diretoria@feesc.org.br","ceo":"Luiz Felipe (Dir. Presidente)","dir_comercial":"Angela (Gerente Executiva)","linkedin":"N/D","endereco":"Rua Delfino Conti, 280, UFSC, Florianópolis/SC"},
    "79485892000118": {"website":"https://pacopedra.com.br","telefone":"N/D","email":"N/D","ceo":"N/D","dir_comercial":"N/D","linkedin":"N/D","endereco":"Gaspar/SC"},
    "03453030000141": {"website":"https://metalinear.com.br","telefone":"(47) 3037-5945","email":"freedom@freedomengenharia.com.br","ceo":"Luciano Thiesen (Sócio-Admin)","dir_comercial":"N/D","linkedin":"N/D","endereco":"Rua Vale do Selke, 1800, Blumenau/SC"},
    "12535370000102": {"website":"https://terrabaseterraplenagem.com.br","telefone":"(47) 3399-1177","email":"terrabase@terrabaseterraplenagem.com.br","ceo":"Jardel Floriani (Sócio-Admin)","dir_comercial":"N/D","linkedin":"instagram.com/pedraforte_terrabase","endereco":"Rua Ruy Barbosa, 783, Timbó/SC"},
    "83665141000150": {"website":"https://setep.com.br","telefone":"N/D","email":"N/D","ceo":"Antonio Rafael Isidoro Netto (Presidente)","dir_comercial":"N/D","linkedin":"N/D","endereco":"Rua Francisco Martinhago, 258, Criciúma/SC"},
}
for _cnpj, _data in _EXA_EXTRA.items():
    if _cnpj not in CONTACTS:
        CONTACTS[_cnpj] = _data
    else:
        for _k, _v in _data.items():
            if not CONTACTS[_cnpj].get(_k) or CONTACTS[_cnpj][_k] in ('N/D', ''):
                CONTACTS[_cnpj][_k] = _v

print(f"Final contacts loaded: {len(CONTACTS)}")
# ============================================================
# QUERY TOP 100
# ============================================================
with open("/tmp/sc_200km.json") as f:
    muns = json.load(f)
names = [m["nome"].upper() for m in muns]

ENG_KW = ['engenh','constru','obra','edific','paviment','saneamento','infraestrutura','incorporadora',
          'empreiteira','construtora','terraplenagem','fundacao','estrutura','predial','rodovi',
          'drenagem','asfalto','concreto','reforma','arquitet']
GOV_STARTS = ['MUNICIPIO DE','MUNICÍPIO DE','PREFEITURA','SECRETARIA','CAMARA MUNICIPAL','CÂMARA MUNICIPAL',
              'GOVERNO','MINISTERIO','MINISTÉRIO','TRIBUNAL','ASSEMBLEIA','ASSEMBLÉIA','DEFENSORIA',
              'PROCURADORIA','POLICIA','POLÍCIA','FUNDO MUNICIPAL','DEPARTAMENTO DE','SERVICO AUTONOMO',
              'SERVIÇO AUTÔNOMO','SERVICO MUNICIPAL','SERVIÇO MUNICIPAL','AUTARQUIA','AGENCIA REGULADORA',
              'AGÊNCIA REGULADORA','INSTITUTO MUNICIPAL','INSTITUTO ESTADUAL','CONSELHO REGIONAL',
              'CONSELHO MUNICIPAL','CORPO DE BOMBEIROS','ESTADO DE','SENADO FEDERAL','UNIAO FEDERAL',
              'SUPERINTENDENCIA','COORDENADORIA','COMANDO','DELEGACIA','FUNDO ESTADUAL','FUNDO NACIONAL']

eng_includes = ' OR '.join([f"LOWER(nome_fornecedor) LIKE '%%{k}%%'" for k in ENG_KW])
gov_excludes = ' OR '.join([f"UPPER(nome_fornecedor) LIKE '{p}%%'" for p in GOV_STARTS])

conn = psycopg2.connect(DSN)
cur = conn.cursor()

query = f"""
SELECT ni_fornecedor, nome_fornecedor, COUNT(*) as cnt, SUM(valor_global) as val,
       COUNT(DISTINCT municipio) as muns,
       STRING_AGG(DISTINCT municipio, ', ' ORDER BY municipio) as muns_lista,
       MIN(data_assinatura)::date as primeiro, MAX(data_assinatura)::date as ultimo,
       STRING_AGG(DISTINCT setor_classificado, ', ' ORDER BY setor_classificado)
           FILTER (WHERE setor_classificado IS NOT NULL) as setores,
       STRING_AGG(DISTINCT orgao_nome, ' | ' ORDER BY orgao_nome)
           FILTER (WHERE orgao_nome IS NOT NULL) as orgaos
FROM pncp_supplier_contracts
WHERE uf = 'SC'
  AND UPPER(municipio) = ANY(%s)
  AND ({eng_includes})
  AND ni_fornecedor IS NOT NULL AND LENGTH(ni_fornecedor) = 14 AND valor_global > 0
  AND NOT ({gov_excludes})
  AND NOT (UPPER(nome_fornecedor) LIKE 'FUNDACAO MUNICIPAL%%')
  AND NOT (UPPER(nome_fornecedor) LIKE 'FUNDACAO ESTADUAL%%')
  AND NOT (UPPER(nome_fornecedor) LIKE 'FUNDAÇÃO MUNICIPAL%%')
  AND NOT (UPPER(nome_fornecedor) LIKE 'FUNDAÇÃO ESTADUAL%%')
  AND (UPPER(nome_fornecedor) LIKE '%%LTDA%%' OR UPPER(nome_fornecedor) LIKE '%%S.A.%%'
    OR UPPER(nome_fornecedor) LIKE '%%S/A%%' OR UPPER(nome_fornecedor) LIKE '%%EIRELI%%'
    OR UPPER(nome_fornecedor) LIKE '%%ME %%' OR UPPER(nome_fornecedor) LIKE '%%MEI%%'
    OR UPPER(nome_fornecedor) LIKE '%%- ME%%' OR UPPER(nome_fornecedor) LIKE 'FUNDACAO%%'
    OR UPPER(nome_fornecedor) LIKE 'FUNDAÇÃO%%' OR UPPER(nome_fornecedor) LIKE 'INSTITUTO%%'
    OR UPPER(nome_fornecedor) LIKE 'COOPERATIVA%%')
GROUP BY ni_fornecedor, nome_fornecedor
ORDER BY val DESC LIMIT 100
"""
cur.execute(query, (names,))
rows = cur.fetchall()
conn.close()

print(f"Fornecedores: {len(rows)}")
with_contacts = sum(1 for r in rows if r[0] in CONTACTS)
print(f"Com contatos Exa: {with_contacts}/100")

# ============================================================
# BUILD EXCEL
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Engenharia SC 200km"

# Styles
hfont = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
dfont = Font(name="Calibri", size=10)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

headers = [
    "#", "CNPJ", "Nome do Fornecedor", "Contratos", "Valor Total (R$)",
    "Municípios", "Setores",
    "Website", "Telefone", "Email",
    "CEO/Diretor Presidente", "Diretor Comercial/Licitações",
    "LinkedIn", "Endereço Sede",
    "Municípios (lista)", "Primeiro Contrato", "Último Contrato", "Órgãos",
]

for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border

for i, row in enumerate(rows):
    r = i + 2
    cnpj, nome, cnt, val, muns, muns_lista, primeiro, ultimo, setores, orgaos = row
    cnpj_clean = cnpj.replace('.','').replace('/','').replace('-','').strip() if cnpj else ''
    contact = CONTACTS.get(cnpj, {}) or CONTACTS.get(cnpj_clean, {})
    has_contact = bool(contact and (contact.get('telefone') or contact.get('ceo')))

    vals = [
        i + 1, cnpj, nome, cnt, float(val or 0), muns, setores,
        contact.get("website", ""), contact.get("telefone", ""), contact.get("email", ""),
        contact.get("ceo", ""), contact.get("dir_comercial", ""),
        contact.get("linkedin", ""), contact.get("endereco", ""),
        muns_lista, primeiro, ultimo, orgaos,
    ]

    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=v)
        c.font = dfont
        c.border = border
        if ci == 5:
            c.number_format = '#,##0.00'
        if ci in (16, 17) and v:
            c.number_format = 'DD/MM/YYYY'

    # Highlight rows with contact data
    if has_contact:
        for ci in range(8, 15):  # contact columns
            ws.cell(row=r, column=ci).fill = green_fill

    if i % 500 == 0:
        print(f"\r  {i:,}/{len(rows):,}", end="")

# Column widths
widths = [5, 18, 50, 10, 18, 8, 35, 30, 18, 30, 35, 30, 30, 45, 55, 14, 14, 50]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.auto_filter.ref = f"A1:S{len(rows) + 1}"
ws.freeze_panes = "A2"

# Summary sheet
ws2 = wb.create_sheet("Resumo")
summary = [
    ["RELATÓRIO FINAL — Fornecedores Engenharia/Construção SC — Raio 200km Florianópolis"],
    [f"Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"],
    [""],
    [f"Fornecedores: {len(rows):,}"],
    [f"Valor total: R$ {sum(r[3] or 0 for r in rows):,.2f}"],
    [f"Com contatos enriquecidos (Exa): {with_contacts}/100"],
    [""],
    ["FILTROS: zero compradores GOV, zero valores nulos, keywords só no nome, sufixo societário obrigatório"],
    [""],
    ["Top 25 com contatos:"],
]
for i, r in enumerate(rows[:25]):
    cnpj = r[0]; nome = (r[1] or 'N/A')[:60]; val = float(r[3] or 0)
    contact = CONTACTS.get(cnpj, {})
    tel = contact.get('telefone', '-')
    ceo = contact.get('ceo', '-')
    summary.append([f"{i+1}. {nome} | Tel: {tel} | CEO: {ceo} | R$ {val:,.0f}"])

for ri, row_data in enumerate(summary, 1):
    c = ws2.cell(row=ri, column=1, value=row_data[0] if row_data else "")
    if ri <= 2:
        c.font = Font(name="Calibri", size=14, bold=True)

ws2.column_dimensions["A"].width = 110

os.makedirs(os.path.dirname(os.path.abspath(OUTPUT)), exist_ok=True)
wb.save(OUTPUT)
print(f"\n\nSalvo: {OUTPUT}")
print(f"Fornecedores: {len(rows):,} | Com contatos: {with_contacts} | Valor total: R$ {sum(r[3] or 0 for r in rows):,.0f}")
