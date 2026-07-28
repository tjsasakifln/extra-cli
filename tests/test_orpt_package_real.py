"""ORPT: real package generation (no Demo A/B fixture as proof)."""
from __future__ import annotations

import os
from pathlib import Path

import pytest


def _dsn() -> str:
    dsn = os.environ.get("LOCAL_DATALAKE_DSN") or os.environ.get("DATABASE_URL")
    if os.environ.get("REQUIRE_REAL_DB") == "1":
        if not dsn:
            pytest.fail("REQUIRE_REAL_DB=1 but DSN unset")
        import psycopg2

        try:
            conn = psycopg2.connect(dsn)
            conn.close()
        except Exception as exc:  # noqa: BLE001
            pytest.fail(f"REQUIRE_REAL_DB=1 but PostgreSQL unavailable: {exc}")
        return dsn
    if not dsn:
        pytest.skip("no DSN")
    return dsn


@pytest.mark.integration
def test_build_package_from_db_no_fixture_rows(tmp_path: Path):
    from openpyxl import load_workbook

    from scripts.ops.deliverable_package_final import build_package_from_db

    dsn = _dsn()
    report = build_package_from_db(dsn, tmp_path / "pkg")
    data = report if isinstance(report, dict) else report.__dict__
    pkg = data.get("package") if isinstance(data, dict) else None
    if pkg is None:
        from dataclasses import asdict

        pkg = asdict(report)["package"]
    meta = pkg.get("meta") or {}
    assert meta.get("fixture") is False
    pdf = Path(pkg["pdf_path"])
    xlsx = Path(pkg["excel_path"])
    if not pdf.is_file():
        pdf = Path.cwd() / pkg["pdf_path"]
    if not xlsx.is_file():
        xlsx = Path.cwd() / pkg["excel_path"]
    assert pdf.is_file(), pkg["pdf_path"]
    assert xlsx.is_file(), pkg["excel_path"]
    assert pdf.read_bytes()[:5] == b"%PDF-"
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        assert "Metadados" in wb.sheetnames
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    s = str(cell)
                    assert s not in {"Demo A", "Demo B"}
                    assert "fixture" not in s.lower() or "fixture=false" in s.lower()
    finally:
        wb.close()
    recon = data.get("reconcile") if isinstance(data, dict) else report.reconcile
    if not isinstance(recon, dict):
        from dataclasses import asdict

        recon = asdict(report)["reconcile"]
    assert recon.get("same_run_id") is True
    assert recon.get("status") == "PASS"


@pytest.mark.integration
def test_export_pack_fail_closed_sql(monkeypatch, tmp_path: Path):
    from scripts.reports.operational_export_pack import (
        OperationalExportError,
        build_pack,
    )

    dsn = _dsn()

    def _boom(*_a, **_k):
        raise OperationalExportError("injected")

    monkeypatch.setattr(
        "scripts.reports.operational_export_pack.universe_version",
        _boom,
    )
    with pytest.raises(OperationalExportError, match="injected"):
        build_pack(dsn, tmp_path)


@pytest.mark.integration
def test_acceptance_harness_runs(tmp_path: Path):
    from scripts.ops.operational_reporting_acceptance import run_acceptance

    dsn = _dsn()
    result = run_acceptance(dsn, tmp_path / "acc")
    assert "proofs" in result
    assert result["proofs"]["lists"]["ok"] is True
    assert result["proofs"]["export_pack"]["ok"] is True
    assert result["proofs"]["package_final"]["ok"] is True
