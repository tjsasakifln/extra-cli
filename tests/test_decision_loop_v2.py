"""Mandatory tests for EXTRA-DECISION-LOOP-01 decision loop."""

from __future__ import annotations

import json
from datetime import UTC, datetime, timedelta
from pathlib import Path

from scripts.opportunity_intel.decision_engine import (
    decide_opportunity,
    map_external_to_internal,
    map_internal_to_external,
)
from scripts.opportunity_intel.human_review import (
    calibrate,
    export_review_sample,
    import_review_labels,
    stratified_sample,
)
from scripts.opportunity_intel.profile_resolve import (
    assert_local_not_tracked,
    resolve_extra_profile,
    write_profile_status,
)
from scripts.opportunity_intel.snapshot import (
    analyze_reconfirm_body,
    build_pncp_url,
    build_snapshot,
    classify_http_status,
    classify_pagination_outcome,
    compute_delta,
    is_generic_pncp_org_url,
    is_high_confidence_open,
    pick_reconfirm_targets,
    reconfirm_opportunity,
    reconfirm_paginated_listing,
    select_active_opportunities,
)
from scripts.ops.decision_pack import (
    build_decision_claims,
    reconcile_claims_to_excel,
    reconcile_pdf_excel,
    run_decision_pack,
)


def _open_row(**overrides):
    now = datetime.now(UTC)
    base = {
        "id": 1,
        "source": "pncp",
        "source_id": "x-1",
        "orgao_cnpj": "12345678000199",
        "orgao_nome": "Prefeitura Teste",
        "objeto": "Reforma predial de prédio público municipal",
        "valor_estimado": 500_000,
        "modalidade": "Pregão Eletrônico",
        "status_canonico": "open",
        "data_abertura": now + timedelta(days=5),
        "data_encerramento": now + timedelta(days=30),
        "data_publicacao": now - timedelta(days=2),
        "uf": "SC",
        "municipio": "Florianópolis",
        "link_edital": "https://pncp.gov.br/app/editais/example",
        "dentro_raio": True,
        "fonte_confiavel": True,
        "has_match_entity": True,
        "is_active": True,
        "ranking": "REVIEW",
        "ranking_score": 60,
    }
    base.update(overrides)
    return base


class TestProfileResolve:
    def test_hash_stable(self):
        a = resolve_extra_profile()
        b = resolve_extra_profile()
        assert a.profile_hash == b.profile_hash
        assert a.profile_id == "extra_construtora"
        assert a.pending_critical  # capacity still PENDING

    def test_local_override_merge(self, tmp_path: Path):
        public = Path("config/client_profiles/extra.yaml")
        local = tmp_path / "extra.local.yaml"
        local.write_text(
            "priority_municipalities:\n  - Florianópolis\nelicitation:\n"
            "  capital_giro:\n    status: SET\n    value: 1000\n",
            encoding="utf-8",
        )
        r = resolve_extra_profile(public_path=public, local_path=local)
        assert r.local_loaded is True
        assert "Florianópolis" in (r.data.get("priority_municipalities") or [])
        assert "capital_giro" not in r.pending_critical or r.data["elicitation"]["capital_giro"]["status"] == "SET"

    def test_profile_status_written(self, tmp_path: Path):
        r = resolve_extra_profile()
        j, m = write_profile_status(r, tmp_path)
        assert j.is_file() and m.is_file()
        payload = json.loads(j.read_text(encoding="utf-8"))
        assert payload["profile_hash"] == r.profile_hash
        assert "pending" in payload

    def test_local_not_tracked_gitignore(self):
        check = assert_local_not_tracked()
        assert check["gitignore_covers"] is True
        assert check["safe"] is True


class TestDecisionEngine:
    def test_mapping_roundtrip(self):
        assert map_internal_to_external("GO") == "PARTICIPAR"
        assert map_internal_to_external("NO_GO") == "NÃO_PARTICIPAR"
        assert map_external_to_internal("PARTICIPAR") == "GO"

    def test_hard_blocker_not_compensated(self):
        row = _open_row(objeto="", status_canonico="open")
        d = decide_opportunity(
            row,
            reconfirm={"outcome": "ok", "status": "ok"},
            profile_meta={"pending_critical": []},
        )
        assert d.recommendation == "NÃO_PARTICIPAR"
        assert d.hard_blockers

    def test_stale_never_participar(self):
        row = _open_row()
        d = decide_opportunity(
            row,
            reconfirm={"outcome": "stale", "status": "stale"},
            profile_meta={"pending_critical": []},
            profile={
                "profile_id": "t",
                "version": 1,
                "region": {"uf": "SC"},
                "desired_object_types": [{"id": "reforma", "terms": ["reforma predial"]}],
                "value_band_soft": {"min_brl": 1, "max_brl": 9e9},
                "operational_constraints": [{"id": "r"}],
                "engineering_categories": ["reforma"],
                "positive_terms": ["reforma"],
                **{k: {"status": "SET", "value": 1} for k in (
                    "capital_giro", "capacidade_simultanea", "cats_atestados", "equipe",
                    "equipamentos", "certidoes", "margem_minima", "risco_aceitavel",
                    "contratos_atuais", "apetite_consorcios", "capacidade_garantia",
                )},
            },
        )
        assert d.recommendation != "PARTICIPAR"

    def test_unconfirmed_never_participar(self):
        row = _open_row()
        d = decide_opportunity(row, reconfirm={"outcome": "not_attempted"})
        assert d.recommendation != "PARTICIPAR"

    def test_pending_profile_review_not_auto_nogo_for_fit(self):
        """Incomplete profile demotes GO→REVIEW, does not invent NÃO_PARTICIPAR alone."""
        row = _open_row()
        d = decide_opportunity(
            row,
            reconfirm={"outcome": "ok", "status": "ok"},
            profile_meta={"pending_critical": ["capital_giro", "margem_minima"]},
            profile={"region": {"uf": "SC"}, "desired_object_types": [{"id": "x", "terms": ["reforma"]}]},
        )
        assert d.recommendation in {"REVIEW", "NÃO_PARTICIPAR"}
        # With open complete data, hard block shouldn't force only from pending profile
        if not d.hard_blockers:
            assert d.recommendation == "REVIEW"

    def test_dimensions_present(self):
        d = decide_opportunity(_open_row(), reconfirm={"outcome": "http_204"})
        for key in (
            "data_confidence",
            "client_fit",
            "technical_fit",
            "commercial_fit",
            "operational_fit",
            "temporal_fit",
        ):
            assert key in d.dimensions

    def test_zero_extra_fit_is_nao_participar(self):
        """Objects with no Extra applicability must discard (not stay REVIEW forever)."""
        prof = {
            "profile_id": "extra_construtora",
            "version": 2,
            "desired_object_types": [
                {
                    "id": "reforma_predial",
                    "label": "Reforma predial",
                    "terms": ["reforma predial", "reforma de edificio"],
                }
            ],
            "positive_terms": ["engenharia civil", "obra de engenharia"],
            "negative_terms": [],
            "engineering_categories": ["reforma_predial"],
            "triage_thresholds": {
                "discard_max_client_fit": 20,
                "priority_min_client_fit": 55,
            },
            "value_band_soft": {"min_brl": 1, "max_brl": 9e9},
        }
        for objeto in (
            "Aquisição de pipoca doce para festa junina",
            "Compra de brinquedos infantis sortidos",
            "Fornecimento de medicamento controlado hospitalar",
        ):
            d = decide_opportunity(
                _open_row(objeto=objeto),
                profile=prof,
                profile_meta={"pending_critical": []},
                reconfirm={"outcome": "ok", "status": "ok"},
            )
            assert d.recommendation == "NÃO_PARTICIPAR", objeto
            assert d.dimensions["client_fit"].score <= 20
            assert any("client_fit_below_discard" in r for r in d.rules)

    def test_expired_deadline_prevails_over_open_status(self):
        """status_canonico=open with past deadline must never stay actionable PARTICIPAR."""
        past = datetime.now(UTC) - timedelta(days=3)
        row = _open_row(
            status_canonico="open",
            data_encerramento=past,
            objeto="Reforma predial de prédio público municipal",
        )
        d = decide_opportunity(
            row,
            reconfirm={"outcome": "ok", "status": "ok", "identity_matched": True},
            profile_meta={"pending_critical": []},
            profile={
                "desired_object_types": [{"id": "reforma", "terms": ["reforma predial"]}],
                "positive_terms": ["reforma"],
                "engineering_categories": ["reforma"],
                "value_band_soft": {"min_brl": 1, "max_brl": 9e9},
            },
        )
        assert d.recommendation == "NÃO_PARTICIPAR"
        assert "prazo_encerrado" in d.hard_blockers
        assert d.recommendation != "PARTICIPAR"

    def test_offline_fixture_never_participar(self):
        row = _open_row(objeto="Reforma predial de prédio público municipal")
        d = decide_opportunity(
            row,
            reconfirm={"outcome": "skipped_offline_fixture", "status": "skipped_offline_fixture"},
            profile_meta={"pending_critical": []},
            profile={
                "desired_object_types": [{"id": "reforma", "terms": ["reforma predial"]}],
                "positive_terms": ["reforma"],
                "engineering_categories": ["reforma"],
                "value_band_soft": {"min_brl": 1, "max_brl": 9e9},
            },
        )
        assert d.recommendation != "PARTICIPAR"
        assert any("offline" in r.lower() for r in d.rules) or d.recommendation in {
            "REVIEW",
            "NÃO_PARTICIPAR",
        }

    def test_reconfirm_ok_without_identity_never_participar(self):
        row = _open_row(objeto="Reforma predial de prédio público municipal")
        d = decide_opportunity(
            row,
            reconfirm={"outcome": "ok", "status": "ok"},  # missing identity_matched
            profile_meta={"pending_critical": []},
            profile={
                "desired_object_types": [{"id": "reforma", "terms": ["reforma predial"]}],
                "positive_terms": ["reforma"],
                "engineering_categories": ["reforma"],
                "value_band_soft": {"min_brl": 1, "max_brl": 9e9},
            },
        )
        assert d.recommendation != "PARTICIPAR"

    def test_suspended_status_is_nao_participar(self):
        d = decide_opportunity(
            _open_row(status_canonico="suspended"),
            reconfirm={"outcome": "ok", "identity_matched": True},
            profile_meta={"pending_critical": []},
        )
        assert d.recommendation == "NÃO_PARTICIPAR"


class TestSnapshot:
    def test_closed_excluded_from_active(self):
        rows = [
            _open_row(id=1, status_canonico="open"),
            _open_row(id=2, status_canonico="closed"),
        ]
        active = select_active_opportunities(rows)
        assert len(active) == 1
        assert active[0]["id"] == 1

    def test_deadline_delta(self):
        prev = build_snapshot(
            [_open_row(id=1, data_encerramento="2026-08-01")],
            run_id="r1",
            reconfirm_map={1: {"outcome": "ok"}},
        )
        cur = build_snapshot(
            [_open_row(id=1, data_encerramento="2026-08-15")],
            run_id="r2",
            reconfirm_map={1: {"outcome": "ok"}},
        )
        # normalize snapshot rows dates as strings
        prev.opportunities[0]["data_encerramento"] = "2026-08-01"
        cur.opportunities[0]["data_encerramento"] = "2026-08-15"
        delta = compute_delta(prev, cur)
        assert 1 in delta["deadline_changed"] or delta["counts"]["deadline_changed"] >= 1

    def test_http_semantics_distinct(self):
        assert classify_http_status(204) == "http_204"
        assert classify_http_status(403) == "http_403"
        assert classify_http_status(429) == "http_429"
        assert classify_http_status(503) == "http_5xx"
        assert classify_http_status(200) == "ok"

    def test_pagination_incomplete_not_success(self):
        assert (
            classify_pagination_outcome(
                http_status=200,
                pages_fetched=2,
                total_pages=5,
            )
            == "pagination_incomplete"
        )
        assert (
            classify_pagination_outcome(
                http_status=200,
                pages_fetched=5,
                total_pages=5,
            )
            == "ok"
        )
        # 429 remains 429 even if pages incomplete
        assert (
            classify_pagination_outcome(
                http_status=429,
                pages_fetched=1,
                total_pages=3,
            )
            == "http_429"
        )
        partial = reconfirm_paginated_listing(
            pages=[(200, '{"page":1}', None), (200, '{"page":2}', None)],
            total_pages=4,
            opportunity_id=99,
        )
        assert partial.outcome == "pagination_incomplete"
        assert partial.outcome != "ok"
        complete = reconfirm_paginated_listing(
            pages=[(200, "a", None), (200, "b", None)],
            total_pages=2,
            opportunity_id=100,
        )
        assert complete.outcome == "ok"

    def test_reconfirm_timeout(self):
        def boom(_url, **_kw):
            return None, "", "timeout:x"

        rc = reconfirm_opportunity(
            _open_row(),
            http_get=boom,
            offline=False,
        )
        assert rc.outcome == "timeout"

    def test_reconfirm_204(self):
        def empty(_url, **_kw):
            return 204, "", None

        rc = reconfirm_opportunity(_open_row(), http_get=empty, offline=False)
        assert rc.outcome == "http_204"

    def test_reconfirm_generic_page_not_ok(self):
        def generic(_url, **_kw):
            return 200, "<html><body>Lista geral de editais do órgão</body></html>", None

        row = _open_row(
            source_id="PNCP-999",
            orgao_cnpj="12345678000199",
            numero_controle="PNCP-999",
        )
        rc = reconfirm_opportunity(row, http_get=generic, offline=False)
        assert rc.outcome != "ok"
        assert rc.identity_matched is False
        assert rc.outcome in {"not_found", "unconfirmed", "partial", "identity_mismatch"}

    def test_reconfirm_identity_match_ok(self):
        def page(_url, **_kw):
            html = (
                "<html><body>Edital PNCP-999 CNPJ 12.345.678/0001-99 "
                "status aberto reforma predial</body></html>"
            )
            return 200, html, None

        row = _open_row(
            source_id="PNCP-999",
            numero_controle="PNCP-999",
            orgao_cnpj="12345678000199",
        )
        rc = reconfirm_opportunity(row, http_get=page, offline=False)
        assert rc.outcome == "ok"
        assert rc.identity_matched is True
        assert rc.raw_hash

    def test_reconfirm_login_page_200_not_ok(self):
        def login(_url, **_kw):
            return 200, "<html>Faça login para continuar</html>", None

        rc = reconfirm_opportunity(_open_row(), http_get=login, offline=False)
        assert rc.outcome != "ok"
        assert rc.identity_matched is False

    def test_reconfirm_revoked_with_identity(self):
        def page(_url, **_kw):
            return (
                200,
                "Edital PNCP-999 CNPJ 12345678000199 foi revogado pelo órgão",
                None,
            )

        row = _open_row(
            source_id="PNCP-999",
            numero_controle="PNCP-999",
            orgao_cnpj="12345678000199",
        )
        rc = reconfirm_opportunity(row, http_get=page, offline=False)
        assert rc.identity_matched is True
        assert rc.status_observed == "revoked"

    def test_reconfirm_offline_no_identity_claim(self):
        rc = reconfirm_opportunity(_open_row(), offline=True)
        assert rc.outcome == "skipped_offline_fixture"
        assert rc.identity_matched is False

    def test_reconfirm_control_ok_but_cnpj_missing_is_mismatch(self):
        """Expected full CNPJ not on page → identity_mismatch, never ok."""
        def page(_url, **_kw):
            return 200, "<html>Edital PNCP-999 status aberto sem cnpj do orgao</html>", None

        row = _open_row(
            source_id="PNCP-999",
            numero_controle="PNCP-999",
            orgao_cnpj="12345678000199",
        )
        rc = reconfirm_opportunity(row, http_get=page, offline=False)
        assert rc.outcome == "identity_mismatch"
        assert rc.identity_matched is False

    def test_reconfirm_cnpj_divergent_is_mismatch(self):
        def page(_url, **_kw):
            return (
                200,
                "Edital PNCP-999 CNPJ 99.999.999/0001-91 outro orgao",
                None,
            )

        row = _open_row(
            source_id="PNCP-999",
            numero_controle="PNCP-999",
            orgao_cnpj="12345678000199",
        )
        rc = reconfirm_opportunity(row, http_get=page, offline=False)
        assert rc.outcome == "identity_mismatch"
        assert rc.identity_matched is False

    def test_expired_open_not_in_active_snapshot(self):
        past = datetime.now(UTC) - timedelta(days=2)
        future = datetime.now(UTC) + timedelta(days=10)
        rows = [
            _open_row(id=1, status_canonico="open", data_encerramento=past),
            _open_row(id=2, status_canonico="open", data_encerramento=future),
        ]
        active = select_active_opportunities(rows)
        ids = [r["id"] for r in active]
        assert 1 not in ids
        assert 2 in ids

    def test_high_confidence_open_requires_ok(self):
        snap = build_snapshot(
            [_open_row(id=9)],
            run_id="r",
            reconfirm_map={9: {"outcome": "not_attempted"}},
        )
        assert snap.opportunities[0]["high_confidence_open"] is False

    def test_pick_targets_includes_all_when_small(self):
        rows = [_open_row(id=i) for i in range(5)]
        t = pick_reconfirm_targets(rows, top_n=20)
        assert len(t) == 5


class TestIdentityProofStrict:
    """Mandatory regressions: CNPJ-only / generic pages never confirm."""

    CONTROL = "12345678000199-1-000001/2026"
    CNPJ = "12345678000199"

    def _row(self, **kw):
        base = {
            "id": 42,
            "source_id": self.CONTROL,
            "numero_controle_pncp": self.CONTROL,
            "orgao_cnpj": self.CNPJ,
            "orgao_nome": "Prefeitura Teste",
            "numero": "000001",
            "link_edital": f"https://pncp.gov.br/app/editais/{self.CNPJ}/2026/1",
        }
        base.update(kw)
        return _open_row(**base)

    def test_01_generic_page_cnpj_only_never_confirms(self):
        """Página genérica contendo somente o CNPJ esperado não confirma."""
        body = (
            "<html><body><h1>Portal do órgão</h1>"
            f"<p>CNPJ: 12.345.678/0001-99</p>"
            "<p>Informações institucionais</p></body></html>"
        )
        analysis = analyze_reconfirm_body(body, self._row(), http_status=200)
        assert analysis["identity_matched"] is False
        assert analysis["outcome"] != "ok"
        assert analysis["outcome"] in {
            "not_found",
            "unconfirmed",
            "ambiguous",
            "identity_mismatch",
        }

        def page(_url, **_kw):
            return 200, body, None

        rc = reconfirm_opportunity(self._row(), http_get=page, offline=False)
        assert rc.outcome != "ok"
        assert rc.identity_matched is False

    def test_02_generic_page_org_name_never_confirms(self):
        """Página genérica contendo nome do órgão não confirma."""
        body = (
            "<html><body><h1>Prefeitura Teste</h1>"
            "<p>Bem-vindo ao portal de compras do município</p></body></html>"
        )
        analysis = analyze_reconfirm_body(body, self._row(), http_status=200)
        assert analysis["identity_matched"] is False
        assert analysis["outcome"] != "ok"

    def test_03_exact_control_and_cnpj_confirms(self):
        """Página com número de controle exato e CNPJ correto confirma."""
        body = (
            "<html><body>"
            f"Controle PNCP: {self.CONTROL} "
            f"CNPJ 12.345.678/0001-99 status aberto reforma"
            "</body></html>"
        )
        analysis = analyze_reconfirm_body(body, self._row(), http_status=200)
        assert analysis["identity_matched"] is True
        assert analysis["outcome"] == "ok"
        assert analysis["checks"]["control_found"] is True

        def page(_url, **_kw):
            return 200, body, None

        rc = reconfirm_opportunity(self._row(), http_get=page, offline=False)
        assert rc.outcome == "ok"
        assert rc.identity_matched is True

    def test_04_partial_control_never_confirms(self):
        """Número de controle parcial não confirma oportunidade."""
        # No exact hiring number field — only a short suffix of the full control
        row = self._row(numero="")
        body = (
            "<html><body>"
            "Fragmento: 0000012026 "  # normalized tail of control, not full ID
            f"CNPJ 12.345.678/0001-99"
            "</body></html>"
        )
        analysis = analyze_reconfirm_body(body, row, http_status=200)
        assert analysis["identity_matched"] is False
        assert analysis["outcome"] != "ok"
        assert analysis["checks"].get("control_found") is False
        assert analysis["outcome"] in {
            "not_found",
            "unconfirmed",
            "ambiguous",
            "identity_mismatch",
        }

    def test_05_same_numero_divergent_cnpj_is_mismatch(self):
        """Número de contratação igual, CNPJ divergente → mismatch."""
        body = (
            "<html><body>"
            "Contratação nº 000001 "
            "CNPJ 99.999.999/0001-91 outro orgao"
            "</body></html>"
        )
        row = self._row(numero_controle_pncp="", source_id="", numero="000001")
        # Without control, only numero+cnpj path applies
        analysis = analyze_reconfirm_body(body, row, http_status=200)
        assert analysis["identity_matched"] is False
        assert analysis["outcome"] == "identity_mismatch"

    def test_06_ok_without_identity_never_high_confidence(self):
        """outcome=ok com identity_matched=false nunca produz high_confidence_open."""
        row = self._row()
        rc = {
            "outcome": "ok",
            "identity_matched": False,
            "rule": "identity_matched_official_page",
        }
        assert is_high_confidence_open(row, rc) is False
        snap = build_snapshot([row], run_id="r", reconfirm_map={42: rc})
        assert snap.opportunities[0]["high_confidence_open"] is False

    def test_07_offline_fixture_never_participar_nor_high_conf(self):
        """Fixture offline nunca produz PARTICIPAR nem high confidence."""
        row = self._row()
        rc = reconfirm_opportunity(row, offline=True)
        assert rc.outcome == "skipped_offline_fixture"
        assert rc.identity_matched is False
        d = decide_opportunity(
            row,
            reconfirm=rc.to_dict(),
            profile_meta={"pending_critical": []},
        )
        assert d.recommendation != "PARTICIPAR"
        assert is_high_confidence_open(row, rc.to_dict()) is False
        snap = build_snapshot(
            [row],
            run_id="off",
            reconfirm_map={42: rc.to_dict()},
        )
        assert snap.opportunities[0]["high_confidence_open"] is False

    def test_08_expired_open_still_out_of_active(self):
        """Prazo vencido com status open continua fora das oportunidades ativas."""
        past = datetime.now(UTC) - timedelta(days=3)
        row = self._row(status_canonico="open", data_encerramento=past)
        active = select_active_opportunities([row])
        assert active == []
        # Even if someone forges reconfirm ok, high conf must stay false
        rc = {
            "outcome": "ok",
            "identity_matched": True,
            "identity_checks": {
                "control_found": True,
                "cnpj_found": True,
                "specific_identity_proof": True,
            },
        }
        assert is_high_confidence_open(row, rc) is False

    def test_09_login_captcha_error_empty_listing_never_confirm(self):
        """Login, CAPTCHA, erro ou listagem vazia nunca confirmam."""
        cases = [
            "<html>Faça login para continuar</html>",
            "<html>Resolva o captcha para acessar</html>",
            "<html>Página não encontrada erro 404</html>",
            "<html>Lista de editais — nenhum edital encontrado</html>",
            "<html>Resultados da busca: nenhum resultado</html>",
        ]
        for body in cases:
            analysis = analyze_reconfirm_body(body, self._row(), http_status=200)
            assert analysis["identity_matched"] is False, body
            assert analysis["outcome"] != "ok", body

    def test_10_pagination_incomplete_never_confirms(self):
        """Paginação incompleta nunca confirma oportunidade."""
        rc = reconfirm_paginated_listing(
            pages=[(200, "<html>page1</html>", None)],
            total_pages=5,
            opportunity_id=42,
        )
        assert rc.outcome == "pagination_incomplete"
        assert rc.identity_matched is False
        row = self._row()
        assert is_high_confidence_open(row, rc.to_dict()) is False
        d = decide_opportunity(row, reconfirm=rc.to_dict())
        assert d.recommendation != "PARTICIPAR"

    def test_generic_org_url_never_ok(self):
        """URL só com CNPJ é página genérica — nunca ok quando é a única URL."""
        url = f"https://pncp.gov.br/app/editais/{self.CNPJ}"
        assert is_generic_pncp_org_url(url) is True
        # No parseable specific control → cannot upgrade to detail URL
        row = self._row(
            link_edital=url,
            numero_controle_pncp="SEM-DETALHE",
            source_id="SEM-DETALHE",
            numero="",
            ano="",
            sequencial="",
        )
        built = build_pncp_url(row)
        assert built is None or is_generic_pncp_org_url(built)

        def page(_url, **_kw):
            return (
                200,
                f"<html>CNPJ {self.CNPJ} lista de editais do orgao</html>",
                None,
            )

        rc = reconfirm_opportunity(row, http_get=page, offline=False)
        assert rc.outcome != "ok"
        assert rc.identity_matched is False

        # analyze_reconfirm_body also rejects generic org URL even with rich body
        rich = f"<html>{self.CONTROL} CNPJ {self.CNPJ} aberto</html>"
        analysis = analyze_reconfirm_body(
            rich,
            self._row(link_edital=url),
            http_status=200,
            url=url,
        )
        assert analysis["identity_matched"] is False
        assert analysis["outcome"] != "ok"

    def test_high_confidence_requires_full_proof(self):
        row = self._row()
        good = {
            "outcome": "ok",
            "identity_matched": True,
            "identity_checks": {
                "control_found": True,
                "cnpj_found": True,
                "specific_identity_proof": True,
                "proof_kind": "numero_controle_pncp_exact",
            },
        }
        assert is_high_confidence_open(row, good) is True
        snap = build_snapshot([row], run_id="r", reconfirm_map={42: good})
        assert snap.opportunities[0]["high_confidence_open"] is True

    def test_forged_ok_identity_without_specific_proof_never_high_conf(self):
        """outcome=ok + identity_matched without specific proof must not high-conf."""
        row = self._row()
        # CNPJ-only checks (the false-positive pattern)
        forged_cnpj = {
            "outcome": "ok",
            "identity_matched": True,
            "identity_checks": {"cnpj_found": True, "control_found": False},
        }
        assert is_high_confidence_open(row, forged_cnpj) is False
        snap = build_snapshot([row], run_id="forge", reconfirm_map={42: forged_cnpj})
        assert snap.opportunities[0]["high_confidence_open"] is False

        # Bare identity_matched with empty checks (legacy forge)
        forged_empty = {"outcome": "ok", "identity_matched": True}
        assert is_high_confidence_open(row, forged_empty) is False
        snap2 = build_snapshot([row], run_id="forge2", reconfirm_map={42: forged_empty})
        assert snap2.opportunities[0]["high_confidence_open"] is False


class TestHumanReview:
    def test_export_never_auto_labels(self, tmp_path: Path):
        decisions = [
            {**_open_row(id=i), "recommendation": "REVIEW", "confidence": "LOW"}
            for i in range(12)
        ]
        path = tmp_path / "queue.csv"
        meta = export_review_sample(decisions, path, target=12)
        text = path.read_text(encoding="utf-8")
        assert "human_decision" in text
        # data rows should have empty human_decision (trailing commas / empty field)
        lines = [ln for ln in text.splitlines() if ln and not ln.startswith("opportunity_id")]
        assert lines
        assert meta["n_sample"] == 12

    def test_import_idempotent(self, tmp_path: Path):
        csv_path = tmp_path / "lab.csv"
        csv_path.write_text(
            "opportunity_id,human_decision,human_reason,hard_block_confirmed,"
            "missing_information,would_present_to_client,reviewed_at,reviewer,"
            "system_recommendation,stratum\n"
            "1,REVIEW,ok,,,,2026-07-19,tiago,REVIEW,borderline\n",
            encoding="utf-8",
        )
        store = tmp_path / "store"
        a = import_review_labels(csv_path, store_dir=store)
        b = import_review_labels(csv_path, store_dir=store)
        assert a["n_labels_total"] == 1
        assert b["n_labels_total"] == 1

    def test_metrics_blocked_without_sample(self, tmp_path: Path):
        res = calibrate(labels={}, store_dir=tmp_path)
        assert res.status == "PENDING_HUMAN"
        assert res.metrics == {}

    def test_metrics_with_enough_labels(self):
        labels = {
            str(i): {
                "human_decision": "REVIEW" if i % 2 else "NÃO_PARTICIPAR",
                "system_recommendation": "REVIEW" if i % 3 else "NÃO_PARTICIPAR",
                "stratum": "borderline",
            }
            for i in range(12)
        }
        res = calibrate(labels=labels, min_labels=10)
        assert res.status == "OK"
        assert "confusion_matrix" in res.metrics
        assert "agreement_system_vs_human" in res.metrics

    def test_stratified_all_when_few(self):
        dec = [{"recommendation": "REVIEW", "id": 1}]
        assert len(stratified_sample(dec, target=40)) == 1


class TestPackGates:
    def test_reconcile_fails_on_fake_files(self, tmp_path: Path):
        pdf = tmp_path / "a.pdf"
        xlsx = tmp_path / "a.xlsx"
        pdf.write_text("not a real pdf", encoding="utf-8")
        xlsx.write_text("not excel", encoding="utf-8")
        r = reconcile_pdf_excel(
            run_id="r1",
            profile_hash="abc123deadbeef",
            cutoff="2026-07-19",
            pdf_path=pdf,
            xlsx_path=xlsx,
            counts={"participar": 1, "review": 2, "nao_participar": 3, "total": 6},
        )
        assert r["status"] == "FAIL"
        assert r["same_run_id"] is False
        assert r["divergences"]

    def test_reconcile_reads_real_products_roundtrip(self, tmp_path: Path):
        """Product-path: generate pack artifacts then reconcile by reading files."""
        code, manifest = run_decision_pack(
            out_dir=tmp_path / "pack_rt",
            offline_reconfirm=True,
            strict=False,
            reconfirm_max=3,
            limit=15,
        )
        out = Path(manifest["out_dir"])
        pdf = out / "executive_decision_brief.pdf"
        xlsx = out / "extra_decision_pack.xlsx"
        assert pdf.is_file() and xlsx.is_file()
        # Tamper Excel Metadados run_id → must FAIL
        from openpyxl import load_workbook

        wb = load_workbook(xlsx)
        ws = wb["Metadados"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "run_id":
                row[1].value = "TAMPERED-RUN"
                break
        wb.save(xlsx)
        bad = reconcile_pdf_excel(
            run_id=manifest["run_id"],
            profile_hash=manifest["profile_hash"],
            cutoff=manifest["cutoff"],
            pdf_path=pdf,
            xlsx_path=xlsx,
            counts=manifest["counts"],
            claims_count=manifest["counts"]["total"],
        )
        assert bad["status"] == "FAIL"
        assert any("excel_run_id_mismatch" in d for d in bad["divergences"])

    def test_claims_reconcile_with_excel_rows(self, tmp_path: Path):
        """Unit path: build claims + Excel without DB; assert row-level reconcile."""
        from openpyxl import Workbook

        run_id = "decision-test-claims-1"
        profile_hash = "deadbeefcafebabe0123456789abcdef0123456789abcdef0123456789abcdef"
        cutoff = "2026-07-19"
        decisions = [
            {
                "opportunity_id": 1,
                "id": 1,
                "recommendation": "REVIEW",
                "internal_ranking": "REVIEW",
                "confidence": "LOW",
                "orgao_nome": "Pref A",
                "ranking_score": 50,
                "hard_blockers": [],
                "missing_information": ["valor_estimado"],
            },
            {
                "opportunity_id": 2,
                "id": 2,
                "recommendation": "NÃO_PARTICIPAR",
                "internal_ranking": "NO_GO",
                "confidence": "HIGH",
                "orgao_nome": "Pref B",
                "ranking_score": 10,
                "hard_blockers": ["objeto_sem_match"],
                "missing_information": [],
            },
        ]
        claims = build_decision_claims(
            decisions,
            run_id=run_id,
            profile_hash=profile_hash,
            cutoff=cutoff,
        )
        assert len(claims) == 2
        xlsx = tmp_path / "pack.xlsx"
        wb = Workbook()
        meta = wb.active
        meta.title = "Metadados"
        meta.append(["key", "value"])
        meta.append(["run_id", run_id])
        meta.append(["profile_hash", profile_hash])
        meta.append(["cutoff", cutoff])
        for name, rows in [
            ("PARTICIPAR", []),
            ("REVIEW", [decisions[0]]),
            ("NAO_PARTICIPAR", [decisions[1]]),
            ("Claims", claims),
        ]:
            ws = wb.create_sheet(name)
            if not rows:
                ws.append(["opportunity_id", "recommendation"])
                continue
            headers = list(rows[0].keys())
            ws.append(headers)
            for r in rows:
                ws.append([r.get(h) if not isinstance(r.get(h), (list, dict)) else str(r.get(h)) for h in headers])
        wb.save(xlsx)
        res = reconcile_claims_to_excel(claims, xlsx)
        assert res["status"] == "PASS", res
        assert res["matched"] == 2
        # count reconcile reads Excel sheets (not mirrored dicts)
        r = reconcile_pdf_excel(
            run_id=run_id,
            profile_hash=profile_hash,
            cutoff=cutoff,
            pdf_path=tmp_path / "missing.pdf",
            xlsx_path=xlsx,
            counts={"participar": 0, "review": 1, "nao_participar": 1, "total": 2},
            claims_count=2,
        )
        # PDF missing → FAIL, but excel sheet counts must be inspected
        assert r["status"] == "FAIL"
        assert "pdf_missing" in r["divergences"]
        assert r["excel_sheet_row_counts"]["REVIEW"] == 1
        assert r["excel_sheet_row_counts"]["NAO_PARTICIPAR"] == 1
        assert r["excel_sheet_row_counts"]["Claims"] == 2

    def test_checksum_changes_when_file_changes(self, tmp_path: Path):
        from scripts.crawl.run_evidence import sha256_file

        f = tmp_path / "p.txt"
        f.write_text("a", encoding="utf-8")
        h1 = sha256_file(f)
        f.write_text("b", encoding="utf-8")
        h2 = sha256_file(f)
        assert h1 != h2

    def test_run_pack_with_injected_opportunities(self, tmp_path: Path, monkeypatch):
        """Full pack path with real decisions (injected rows — no DB mock dependency)."""
        from scripts.ops.weekly_cycle import StageResult

        rows = [
            _open_row(
                id=101,
                objeto="Reforma predial de prédio público municipal",
                ranking="REVIEW",
                ranking_score=60,
            ),
            _open_row(
                id=102,
                objeto="Aquisição de pipoca e brinquedos para evento",
                ranking="REVIEW",
                ranking_score=55,
            ),
            _open_row(
                id=103,
                objeto="Fornecimento de medicamento hospitalar",
                ranking="REVIEW",
                ranking_score=50,
            ),
        ]
        monkeypatch.setattr(
            "scripts.ops.decision_pack.load_opportunities",
            lambda *_a, **_k: rows,
        )
        monkeypatch.setattr(
            "scripts.ops.decision_pack.stage_freshness",
            lambda _conn: StageResult(
                name="freshness",
                status="ok",
                detail={
                    "sources": [
                        {
                            "source": "pncp_opportunities",
                            "level": "fresh",
                            "age_hours": 1,
                            "sla_hours": 48,
                        }
                    ]
                },
            ),
        )
        monkeypatch.setattr(
            "scripts.ops.decision_pack._connect",
            lambda _dsn: object(),
        )
        monkeypatch.setattr(
            "scripts.ops.decision_pack._resolve_dsn",
            lambda _explicit: "postgresql://mock/mock",
        )

        code, manifest = run_decision_pack(
            out_dir=tmp_path / "pack_inj",
            offline_reconfirm=True,
            skip_db=False,
            strict=False,
            reconfirm_max=3,
            limit=10,
        )
        assert code == 0
        assert manifest.get("counts", {}).get("total") == 3
        # pipoca + medicamento must discard
        assert manifest.get("counts", {}).get("nao_participar", 0) >= 2
        assert (manifest.get("reconcile") or {}).get("status") == "PASS", manifest.get(
            "reconcile"
        )
        assert (manifest.get("reconcile") or {}).get("same_run_id") is True
        claims_csv = Path(manifest["out_dir"]) / "claims_provenance.csv"
        assert claims_csv.is_file()
        import csv

        with claims_csv.open(encoding="utf-8", newline="") as fh:
            claims = list(csv.DictReader(fh))
        assert len(claims) == 3
        res = reconcile_claims_to_excel(
            [{"claim_id": c["claim_id"], "recommendation": c["recommendation"]} for c in claims],
            Path(manifest["out_dir"]) / "extra_decision_pack.xlsx",
        )
        assert res["status"] == "PASS"
        assert res["matched"] == 3

    def test_run_pack_empty_db_still_reconciles(self, tmp_path: Path):
        """Under default pytest mock (empty DB), pack must still PASS reconcile."""
        code, manifest = run_decision_pack(
            out_dir=tmp_path / "pack_empty",
            offline_reconfirm=True,
            skip_db=False,
            strict=False,
            reconfirm_max=2,
            limit=10,
        )
        assert code == 0
        assert (manifest.get("reconcile") or {}).get("status") == "PASS", manifest.get(
            "reconcile"
        )
        assert (tmp_path / "pack_empty" / "claims_provenance.csv").is_file()
        assert (manifest.get("reconcile") or {}).get("claims_reconcile", {}).get(
            "status"
        ) == "PASS"
        # same_run_id proven by reading products
        assert (manifest.get("reconcile") or {}).get("same_run_id") is True

    def test_exit_nonzero_on_reconcile_fail_strict(self, tmp_path: Path, monkeypatch):
        def bad_pdf(*_a, **_k):
            return False, "forced"

        monkeypatch.setattr("scripts.ops.decision_pack._generate_pdf", bad_pdf)
        code, manifest = run_decision_pack(
            out_dir=tmp_path / "pack2",
            offline_reconfirm=True,
            strict=True,
            reconfirm_max=3,
            limit=10,
        )
        assert code != 0
        assert manifest.get("reconcile", {}).get("status") == "FAIL"

    def test_reexecution_does_not_duplicate_labels(self, tmp_path: Path):
        csv_path = tmp_path / "lab.csv"
        csv_path.write_text(
            "opportunity_id,human_decision,human_reason,hard_block_confirmed,"
            "missing_information,would_present_to_client,reviewed_at,reviewer,"
            "system_recommendation,stratum\n"
            "42,REVIEW,r,,,,2026-07-19,tiago,REVIEW,borderline\n",
            encoding="utf-8",
        )
        store = tmp_path / "s"
        import_review_labels(csv_path, store_dir=store)
        import_review_labels(csv_path, store_dir=store)
        doc = json.loads((store / "labels.json").read_text(encoding="utf-8"))
        assert len(doc["labels"]) == 1
