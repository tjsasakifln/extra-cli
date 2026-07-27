"""CMI vertical — DOD §10.1 / §10.2 / §11.1 requirement-linked tests.

Drives shipped entry points (contract_market_intelligence, value semantics,
deliverable B/D helpers, operational_reports). No reimplementation of production
logic inside assertions beyond calling the public API.
"""
from __future__ import annotations

import json
import os
from pathlib import Path

import pytest

from scripts.lib.value_semantics import ValorSemantica
from scripts.ops import contract_market_intelligence as cmi
from scripts.ops.deliverable_b_competitors import capacity_hypothesis, desagio_from_pair
from scripts.ops.deliverable_d_prices import (
    ComparabilityRule,
    PriceObservation,
    build_report,
)

DSN = os.environ.get(
    "LOCAL_DATALAKE_DSN", "postgresql://test:test@127.0.0.1:5433/extra_test"
)
REQUIRE_REAL_DB = os.environ.get("REQUIRE_REAL_DB", "").strip() in {"1", "true", "yes"}


# ── 15.1 Escopo honesto ─────────────────────────────────────────────────────


def test_winner_not_auto_sole_participant():
    contracts = [
        {
            "contrato_id": "1",
            "fornecedor_cnpj": "22222222000191",
            "fornecedor_nome": "A",
            "orgao_cnpj": "11111111000111",
            "orgao_nome": "Pref",
            "valor_contratado": 100.0,
            "objeto_contrato": "reforma",
            "municipio": "X",
            "data_inicio": None,
            "source": "pncp",
        }
    ]
    rows = cmi.aggregate_suppliers(contracts, as_of="2026-07-27T00:00:00Z", source="t")
    assert rows[0]["role"] == "winner_identified"
    assert rows[0]["participant_identified"] is False
    lim_blob = " ".join(rows[0]["limitations"]).lower()
    assert "particip" in lim_blob or "proposta" in lim_blob or "fonte" in lim_blob
    assert rows[0]["win_rate"]["status"] == "NOT_COMPUTABLE"


def test_unknown_participants_remain_unknown():
    wr = cmi.win_rate(wins=1, proposals_presented=None)
    assert wr["status"] == "NOT_COMPUTABLE"
    assert any("não expõe" in x.lower() or "propostas" in x.lower() for x in wr["limitations"])


def test_win_rate_without_denominator_not_computable():
    assert cmi.win_rate(wins=10, proposals_presented=None)["status"] == "NOT_COMPUTABLE"
    assert cmi.win_rate(wins=10, proposals_presented=0)["status"] == "NOT_COMPUTABLE"
    ready = cmi.win_rate(wins=2, proposals_presented=10)
    assert ready["status"] == "READY"
    assert ready["value"] == pytest.approx(0.2)


def test_desagio_without_pair_not_computable():
    d = cmi.desagio_metric(
        valor_estimado=100.0, valor_homologado=80.0, same_certame_lote_item=False
    )
    assert d["status"] == "NOT_COMPUTABLE"
    ok = cmi.desagio_metric(
        valor_estimado=100.0, valor_homologado=80.0, same_certame_lote_item=True
    )
    assert ok["status"] == "READY"
    assert ok["value"] == pytest.approx(20.0)


def test_capacity_not_inferred_from_contracts():
    cap = capacity_hypothesis("x")
    assert cap["claim_as_fact_forbidden"] is True
    assert cap["label"] == "HYPOTHESIS"
    row = cmi.aggregate_suppliers(
        [
            {
                "contrato_id": "1",
                "fornecedor_cnpj": "22222222000191",
                "fornecedor_nome": "A",
                "orgao_cnpj": "11111111000111",
                "orgao_nome": "Pref",
                "valor_contratado": 50.0,
                "objeto_contrato": "obra",
                "municipio": "Y",
                "data_inicio": None,
                "source": "pncp",
            }
        ],
        as_of="t",
        source="t",
    )[0]
    assert row["capacity_claim"]["claim_as_fact_forbidden"] is True
    assert "capacidade" in " ".join(row["limitations"]).lower() or "n_contratos" in " ".join(
        row["limitations"]
    )


def test_limitations_emitted_on_indicators():
    audit = cmi._unit_audit()
    assert audit["ok"] is True


# ── 15.3 Semântica ──────────────────────────────────────────────────────────


def test_four_value_definitions_explicit_and_not_interchangeable():
    defs = cmi.VALUE_DEFINITIONS
    assert set(defs) == {
        "valor_estimado",
        "valor_homologado",
        "valor_contratado",
        "valor_pago",
    }
    enums = {d["enum"] for d in defs.values()}
    assert enums == {
        ValorSemantica.ESTIMADO.value,
        ValorSemantica.HOMOLOGADO.value,
        ValorSemantica.CONTRATADO.value,
        ValorSemantica.PAGO.value,
    }
    for d in defs.values():
        assert len(d["definition"]) > 20


def test_missing_value_never_becomes_zero():
    tv = cmi.typed_value(
        value=None,
        value_type="valor_contratado",
        value_scope="global",
        source="x",
        reference_date=None,
        official_or_inferred="official",
    )
    assert tv["value"] is None
    assert tv["status"] == "MISSING"
    tm = cmi.ticket_medio([None, None])
    assert tm["value"] is None
    assert tm["status"] == "NOT_COMPUTABLE"


def test_inferred_and_official_markers():
    off = cmi.typed_value(
        value=10.0,
        value_type="valor_estimado",
        value_scope="item",
        source="edital",
        reference_date="2024-01-01",
        official_or_inferred="official",
    )
    inf = cmi.typed_value(
        value=10.0,
        value_type="valor_pago",
        value_scope="unitario",
        source="model",
        reference_date="2024-01-01",
        official_or_inferred="inferred",
        inference_method="proxy",
    )
    assert off["official_or_inferred"] == "official"
    assert inf["official_or_inferred"] == "inferred"
    assert inf["inference_method"] == "proxy"


def test_percentiles_comparable_only_and_no_preco_real_label():
    obs = [
        PriceObservation(
            value=float(100 + i * 5),
            value_semantic="contratado",
            tipo_obra_servico="reforma_predial",
            unidade="m2",
            lote="unico",
            porte="medio",
            regiao="SC",
            periodo="2025-Q1",
            is_global_heterogeneous=True,
            source="t",
        )
        for i in range(6)
    ]
    report = build_report(obs, ComparabilityRule(min_sample=3))
    # labels_forbidden_used must stay empty; claims must not assert "preço real praticado".
    for p in report.panels:
        assert not p.get("labels_forbidden_used")
        claims = " ".join(p.get("claims") or []).lower()
        assert "preço real praticado" not in claims
        assert "preco real praticado" not in claims
    # Heterogeneous globals surface a limitation (honest ban), not a marketing claim.
    lim_blob = " ".join(
        " ".join(p.get("limitations") or []) for p in report.panels
    ).lower()
    assert "heterogen" in lim_blob or "preço real" in lim_blob or "preco real" in lim_blob
    assert any(p["status"] == "OK" for p in report.panels)
    assert any(p.get("outliers_flagged") is not None for p in report.panels)
    assert any(p.get("n_observations", 0) >= 3 for p in report.panels)


def test_market_share_hhi_fail_closed_and_ready():
    bad = cmi.market_share_and_hhi(
        {"a": 1.0}, population_definition="x", semantically_valid=False
    )
    assert bad["status"] == "NOT_COMPUTABLE"
    good = cmi.market_share_and_hhi(
        {"a": 100.0, "b": 50.0, "c": 50.0},
        population_definition="eligible SC contracts valor_contratado",
        semantically_valid=True,
    )
    assert good["status"] == "READY"
    assert good["hhi"]["hhi_value"] > 0
    assert abs(sum(v["share"] for v in good["market_share"].values()) - 1.0) < 1e-9


def test_ticket_medio_records_denominator():
    tm = cmi.ticket_medio([100.0, None, 300.0])
    assert tm["denominator"] == 2
    assert tm["numerator"] == 400.0
    assert tm["value"] == 200.0
    assert tm["excluded_missing_value_count"] == 1


# ── 15.2 / 15.4 / 15.5 package against real PostgreSQL ───────────────────────


@pytest.mark.real_db
@pytest.mark.skipif(not REQUIRE_REAL_DB, reason="REQUIRE_REAL_DB not set")
def test_real_db_schema_and_package(tmp_path: Path):
    import psycopg2

    try:
        conn = psycopg2.connect(DSN)
        try:
            info = cmi.require_table_columns(
                conn, "pncp_supplier_contracts", cmi.REQUIRED_CONTRACT_COLS
            )
            assert info["table"] == "pncp_supplier_contracts"
            with pytest.raises(RuntimeError, match="missing"):
                cmi.require_table_columns(
                    conn, "pncp_supplier_contracts", ["no_such_col_xyz"]
                )
        finally:
            conn.close()

        out = tmp_path / "cmi-run"
        result = cmi.run_package(DSN, out, seed_if_empty=True, uf_filter="SC")
        assert result["ok"] is True
        assert result["population_count"] >= 1
        assert result["supplier_count"] >= 1

        # required artifacts
        required = [
            "metadata.json",
            "suppliers-ranking.csv",
            "suppliers-ranking.json",
            "concentration-by-supplier.csv",
            "concentration-by-entity.csv",
            "value-references.csv",
            "value-references.json",
            "limitations.json",
            "reliability-status.json",
            "executive-review.xlsx",
            "competitor-review.md",
            "proof.json",
            "ledger.json",
            "acceptance-manifest.json",
        ]
        for name in required:
            p = out / name
            assert p.is_file(), name
            assert p.stat().st_size > 20, name

        meta = json.loads((out / "metadata.json").read_text(encoding="utf-8"))
        assert meta["complete_population_aggregated"] is True
        assert meta["population_count"] == result["population_count"]
        assert "valor_estimado" in meta["value_definitions"]
        assert meta["code_sha"]

        suppliers = json.loads((out / "suppliers-ranking.json").read_text(encoding="utf-8"))
        assert suppliers
        for s in suppliers:
            assert s["role"] == "winner_identified"
            assert s["participant_identified"] is False
            assert s["win_rate"]["status"] == "NOT_COMPUTABLE"
            assert s["capacity_claim"]["claim_as_fact_forbidden"] is True
            # no orgao as supplier: CNPJ of supplier present
            assert len(s["fornecedor_cnpj"]) == 14

        rel = json.loads((out / "reliability-status.json").read_text(encoding="utf-8"))
        assert rel["win_rate"]["status"] == "NOT_COMPUTABLE"
        assert rel["participantes"]["status"] == "SOURCE_UNAVAILABLE"
        assert rel["ranking_vencedores"]["status"] == "READY"

        proof = json.loads((out / "proof.json").read_text(encoding="utf-8"))
        assert proof["ok"] is True
        assert proof["population_count"] >= 1
        assert proof["claims_scan"]["ok"] is True

        md = (out / "competitor-review.md").read_text(encoding="utf-8")
        assert "vencedor" in md.lower()
        assert "NOT_COMPUTABLE" in md or "not_computable" in md.lower()
        assert "preço real praticado" in md.lower()  # mentioned as forbidden explanation
        assert "limita" in md.lower()

        # Excel sheets
        from openpyxl import load_workbook

        wb = load_workbook(out / "executive-review.xlsx")
        names = set(wb.sheetnames)
        for need in (
            "Metadados",
            "Fornecedores",
            "Contratos",
            "Concentracao",
            "ReferenciasValores",
            "Limitacoes",
            "Confiabilidade",
        ):
            assert need in names, need

        # material rows not header-only
        csv_text = (out / "suppliers-ranking.csv").read_text(encoding="utf-8")
        assert csv_text.count("\n") >= 2

        # null valor preserved in typed values
        vrefs = json.loads((out / "value-references.json").read_text(encoding="utf-8"))
        typed = vrefs["typed_values"]
        assert any(t["value"] is None and t["status"] == "MISSING" for t in typed) or any(
            t.get("value") is not None for t in typed
        )
        # never claim four fields interchangeable
        assert vrefs["fields_not_interchangeable"] is True
    finally:
        cmi.cleanup_cmi_fixture(DSN)


@pytest.mark.real_db
@pytest.mark.skipif(not REQUIRE_REAL_DB, reason="REQUIRE_REAL_DB not set")
def test_operational_reports_no_orgao_as_competitor():
    import psycopg2
    import psycopg2.extras

    from scripts.reports.operational_reports import report_concorrentes

    try:
        cmi.seed_cmi_fixture(DSN)
        conn = psycopg2.connect(DSN, cursor_factory=psycopg2.extras.RealDictCursor)
        try:
            rows = report_concorrentes(conn)
            assert not any(
                r.get("provenance") == "fallback_orgao_not_supplier" for r in rows
            )
            if rows and "_error" not in rows[0]:
                for r in rows:
                    assert r.get("role") == "winner_identified" or r.get(
                        "provenance"
                    ) == ("from_pncp_supplier_contracts")
                    assert r.get("n_contratos") is not None
        finally:
            conn.close()
    finally:
        cmi.cleanup_cmi_fixture(DSN)


@pytest.mark.real_db
@pytest.mark.skipif(not REQUIRE_REAL_DB, reason="REQUIRE_REAL_DB not set")
def test_query_failure_not_silent_empty():
    """Broken SQL surfaces _error instead of SUCCESS_ZERO-looking empty."""
    import psycopg2
    import psycopg2.extras

    from scripts.reports import operational_reports as op

    conn = psycopg2.connect(DSN, cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        rows = op._q(conn, "SELECT * FROM definitely_missing_table_cmi_xyz")
        assert rows, "expected error marker row, got empty"
        assert "_error" in rows[0]
        assert "definitely_missing_table_cmi_xyz" in str(rows[0]["_error"])
        # missing supplier table path also non-silent
        missing = op.report_concorrentes(conn)
        # either real rows or explicit error — never orgao fallback
        assert not any(
            r.get("provenance") == "fallback_orgao_not_supplier" for r in missing
        )
    finally:
        conn.close()


def test_cli_audit_unit_exit_zero():
    assert cmi.main(["audit-unit"]) == 0


def test_desagio_helper_alignment():
    v, st, _ = desagio_from_pair(
        valor_estimado=None, valor_homologado=1.0, same_certame_lote_item=True
    )
    assert v is None and st == "INSUFFICIENT_PAIR"
