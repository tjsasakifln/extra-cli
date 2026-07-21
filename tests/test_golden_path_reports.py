"""DoD §12.1 — golden path generates real Excel and PDF files."""

from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

import pytest

from scripts.golden_path import run_reports

pytestmark = pytest.mark.real_db


def test_help_documents_execute_reports_only() -> None:
    r = subprocess.run(
        [sys.executable, "-m", "scripts.golden_path", "--help"],
        capture_output=True,
        text=True,
        timeout=30,
        check=False,
    )
    assert r.returncode == 0
    assert "execute-reports-only" in (r.stdout + r.stderr)


def test_run_reports_produces_excel_and_pdf_files() -> None:
    dsn = os.getenv("LOCAL_DATALAKE_DSN", "postgresql://test:test@127.0.0.1:5433/extra_test")
    try:
        import psycopg2

        psycopg2.connect(dsn, connect_timeout=3).close()
    except Exception:
        pytest.skip("no local test-db")

    recs = run_reports(dsn)
    by_type = {r.type: r for r in recs}
    assert "excel" in by_type
    assert "pdf" in by_type
    excel = by_type["excel"]
    pdf = by_type["pdf"]
    assert excel.status == "generated", excel
    assert pdf.status == "generated", pdf
    assert excel.path and Path(excel.path).is_file()
    assert pdf.path and Path(pdf.path).is_file()
    assert Path(excel.path).stat().st_size >= 100
    assert Path(pdf.path).stat().st_size >= 100
    # Excel openable
    from openpyxl import load_workbook

    wb = load_workbook(excel.path, read_only=True)
    assert wb.sheetnames
    wb.close()
    # PDF magic
    assert Path(pdf.path).read_bytes()[:4] == b"%PDF"
