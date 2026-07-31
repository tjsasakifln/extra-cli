"""Workbench campaign tests — drive shipped workflow + review + deliverable paths."""

from __future__ import annotations

import time
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from scripts.command_center.app import create_app
from scripts.command_center.capabilities.registry import reset_registry
from scripts.command_center.config import Settings
from scripts.command_center.deliverables.excel_render import neutralize_formula_injection
from scripts.command_center.export_bundle import build_export_bundle
from scripts.command_center.review_rules import decision_is_obsolete, validate_decision_request
from scripts.command_center.run_manifest import load_manifest, validate_manifest
from scripts.command_center.workflows.runner import run_workflow

try:
    from pypdf import PdfReader as _PdfReader
except ImportError:  # pragma: no cover
    _PdfReader = None


def _pdf_ok(path: Path) -> bool:
    assert path.is_file() and path.stat().st_size > 100
    raw = path.read_bytes()[:8]
    assert raw.startswith(b"%PDF")
    if _PdfReader is not None:
        reader = _PdfReader(str(path))
        assert len(reader.pages) >= 1
        text = "".join((p.extract_text() or "") for p in reader.pages[:3])
        assert len(text.strip()) > 20
    return True


def _xlsx_ok(path: Path, *, min_sheets: int = 2) -> None:
    assert path.is_file()
    wb = load_workbook(path, read_only=True)
    assert len(wb.sheetnames) >= min_sheets
    assert "Resumo" in wb.sheetnames or "Metodologia e fontes" in wb.sheetnames
    wb.close()


@pytest.mark.parametrize(
    "workflow_id,params",
    [
        ("workflow.extra.opportunities", {"use_fixture": True, "max_shortlist": 3}),
        ("workflow.confenge.suppliers", {"use_fixture": True, "uf": "SC", "max_companies": 3}),
        ("workflow.confenge.public_agencies", {"use_fixture": True, "uf": "SC", "max_leads": 3}),
        ("workflow.process_documents", {"use_fixture": True, "query": "demo-processo-001"}),
    ],
)
def test_workflow_produces_manifest_pdf_xlsx(tmp_path: Path, workflow_id: str, params: dict) -> None:
    out = tmp_path / workflow_id
    events: list[dict] = []
    result = run_workflow(workflow_id, params, out_dir=out, code_sha="testsha", on_progress=events.append)
    assert result["status"] == "SUCCEEDED"
    mf_path = Path(result["manifest_path"])
    mf = load_manifest(mf_path)
    assert validate_manifest(mf) == []
    assert mf["discovery_source"] == "manifest"
    assert events, "structured progress events required"
    suffixes = {Path(a["path"]).suffix.lower() for a in mf["artifacts"]}
    assert ".pdf" in suffixes
    assert ".xlsx" in suffixes
    assert any(a.get("primary") for a in mf["artifacts"])
    for a in mf["artifacts"]:
        p = Path(a["path"])
        if p.suffix.lower() == ".pdf":
            _pdf_ok(p)
        if p.suffix.lower() == ".xlsx":
            _xlsx_ok(p)
    bundle = build_export_bundle(out)
    assert Path(bundle["bundle_path"]).is_file()
    assert bundle["checksums"]


def test_formula_injection_neutralized() -> None:
    assert neutralize_formula_injection("=1+1") == "'=1+1"
    assert neutralize_formula_injection("+cmd") == "'+cmd"
    assert neutralize_formula_injection("@sum") == "'@sum"
    assert neutralize_formula_injection("texto") == "texto"
    assert neutralize_formula_injection(12.5) == 12.5


def test_decision_rules_reject_defer_accept() -> None:
    assert validate_decision_request(decision="REJECT", rationale="x", title="T")
    assert not validate_decision_request(
        decision="REJECT", rationale="motivo suficiente de recusa", title="T"
    )
    assert validate_decision_request(
        decision="DEFER", rationale="adiar com justificativa longa", return_by=None
    )
    assert not validate_decision_request(
        decision="DEFER",
        rationale="adiar com justificativa longa",
        return_by="2026-08-01",
    )
    assert validate_decision_request(
        decision="ACCEPT",
        rationale="ok",
        presented_hashes={"source": "abc"},
        artifact_hashes={},
    )
    assert not validate_decision_request(
        decision="ACCEPT",
        rationale="ok",
        presented_hashes={"source": "abc"},
        artifact_hashes={"source": "abc"},
    )
    assert decision_is_obsolete(stored_hashes={"source": "a"}, current_hashes={"source": "b"})


@pytest.fixture()
def client(tmp_path: Path) -> TestClient:
    reset_registry()
    data = tmp_path / "cc-data"
    out = tmp_path / "output"
    out.mkdir()
    settings = Settings(
        host="127.0.0.1",
        port=8765,
        data_dir=data,
        open_browser=False,
        spa_dist=None,
        allowed_artifact_roots=(out.resolve(), data.resolve()),
        max_concurrent_jobs=2,
    )
    app = create_app(settings=settings)
    return TestClient(app)


def _csrf(c: TestClient) -> str:
    r = c.get("/api/csrf")
    assert r.status_code == 200
    token = r.json().get("csrf_token") or r.json().get("token")
    assert token
    return str(token)


def test_api_workflows_list_and_run_extra(client: TestClient) -> None:
    r = client.get("/api/workflows")
    assert r.status_code == 200
    ids = {w["id"] for w in r.json()["workflows"]}
    assert "workflow.extra.opportunities" in ids
    token = _csrf(client)
    phrase = "Confirmo a geração local de entregáveis (sem envio automático de mensagens)."
    start = client.post(
        "/api/jobs",
        headers={"X-CC-CSRF": token},
        json={
            "capability_id": "workflow.extra.opportunities",
            "params": {"use_fixture": True, "max_shortlist": 3},
            "confirmation": phrase,
        },
    )
    assert start.status_code == 200, start.text
    job_id = start.json()["job"]["job_id"]
    final = None
    for _ in range(80):
        j = client.get(f"/api/jobs/{job_id}").json()["job"]
        if j["status"] not in {"QUEUED", "VALIDATING", "RUNNING", "CANCELLING"}:
            final = j
            break
        time.sleep(0.1)
    assert final is not None, "job did not finish"
    assert final["status"] == "SUCCEEDED", final
    assert final.get("artifacts"), final
    man = client.get(f"/api/jobs/{job_id}/manifest")
    assert man.status_code == 200, man.text
    body = man.json()
    assert body["valid"] is True
    arts = body["manifest"]["artifacts"]
    pdfs = [a for a in arts if str(a["path"]).endswith(".pdf")]
    xlsxs = [a for a in arts if str(a["path"]).endswith(".xlsx")]
    assert pdfs and xlsxs
    # preview xlsx
    prev = client.get("/api/artifacts/preview-xlsx", params={"path": xlsxs[0]["path"]})
    assert prev.status_code == 200, prev.text
    assert prev.json()["sheets"]
    assert prev.json()["headers"]
    # pdf downloadable / kind
    art = client.get("/api/artifacts", params={"path": pdfs[0]["path"]})
    assert art.status_code == 200
    assert art.json()["kind"] == "pdf"
    # reviews enqueued
    reviews = client.get("/api/reviews", params={"status": "pending"}).json()["reviews"]
    assert len(reviews) >= 1
    item = reviews[0]
    payload = item.get("payload") or {}
    hashes = payload.get("artifact_hashes") or (
        {"source": payload["content_hash"]} if payload.get("content_hash") else {}
    )
    token = _csrf(client)
    # REJECT without rationale blocked
    bad = client.post(
        "/api/decisions",
        headers={"X-CC-CSRF": token},
        json={"item_id": item["id"], "decision": "REJECT", "rationale": "x"},
    )
    assert bad.status_code == 400
    # ACCEPT with matching hashes + confirmation
    conf = client.get(f"/api/reviews/{item['id']}/confirmation").json()
    phrase2 = conf["confirmation_phrase"]
    ok = client.post(
        "/api/decisions",
        headers={"X-CC-CSRF": token},
        json={
            "item_id": item["id"],
            "decision": "ACCEPT",
            "rationale": "Aceite após leitura das evidências da shortlist.",
            "confirmation": phrase2,
            "artifact_hashes": hashes,
        },
    )
    assert ok.status_code == 200, ok.text
    assert ok.json()["ok"] is True


def test_no_dod_auto_accept(client: TestClient) -> None:
    token = _csrf(client)
    conf = client.get("/api/reviews/DOD-ITEM-1/confirmation").json()
    r = client.post(
        "/api/decisions",
        headers={"X-CC-CSRF": token},
        json={
            "item_id": "DOD-ITEM-1",
            "decision": "ACCEPT",
            "rationale": "tentativa de autoaceite DOD deve falhar",
            "confirmation": conf["confirmation_phrase"],
        },
    )
    assert r.status_code == 200
    assert r.json().get("blocked") is True


def test_overview_points_to_workflows(client: TestClient) -> None:
    ov = client.get("/api/overview").json()
    hrefs = " ".join(a["href"] for a in ov["quick_actions"])
    assert "/work/start/" in hrefs
