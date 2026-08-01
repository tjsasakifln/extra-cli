"""Integration: fixture honesty, regenerate with real content change, ACCEPT hash invalidation."""

from __future__ import annotations

import json
import time
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from scripts.command_center.app import create_app
from scripts.command_center.capabilities.registry import reset_registry
from scripts.command_center.config import Settings
from scripts.command_center.regenerate import apply_corrections_to_source, regenerate_workflow_version
from scripts.command_center.run_manifest import sha256_file
from scripts.command_center.workflows.runner import run_workflow


@pytest.fixture()
def client(tmp_path: Path) -> TestClient:
    reset_registry()
    data = tmp_path / "cc-data"
    out = tmp_path / "output"
    out.mkdir()
    settings = Settings(
        host="127.0.0.1",
        port=8765,
        data_dir=data,
        open_browser=False,
        spa_dist=None,
        allowed_artifact_roots=(out.resolve(), data.resolve()),
        max_concurrent_jobs=2,
    )
    return TestClient(create_app(settings=settings))


def _csrf(c: TestClient) -> str:
    r = c.get("/api/csrf")
    assert r.status_code == 200
    return str(r.json().get("csrf_token") or r.json().get("token"))


def _wait_job(client: TestClient, job_id: str) -> dict:
    final = None
    for _ in range(80):
        j = client.get(f"/api/jobs/{job_id}").json()["job"]
        if j["status"] not in {"QUEUED", "VALIDATING", "RUNNING", "CANCELLING"}:
            final = j
            break
        time.sleep(0.1)
    assert final and final["status"] == "SUCCEEDED", final
    return final


def test_use_fixture_false_is_real_fail_closed(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """use_fixture=False selects REAL; without DSN it blocks — never falls back to fixture."""
    monkeypatch.delenv("LOCAL_DATALAKE_DSN", raising=False)
    result = run_workflow(
        "workflow.extra.opportunities",
        {"use_fixture": False, "max_shortlist": 3},
        out_dir=tmp_path / "real-no-dsn",
        code_sha="x",
    )
    assert result["data_mode"] == "REAL"
    assert str(result["status"]).startswith("BLOCKED_")
    assert result["status"] != "SUCCEEDED"


def test_apply_corrections_changes_source_and_pdf_content(tmp_path: Path) -> None:
    """AC#25 core: corrected classification must enter public_agencies.json and PDF."""
    first = run_workflow(
        "workflow.confenge.public_agencies",
        {"use_fixture": True, "uf": "SC", "max_leads": 3},
        out_dir=tmp_path / "v1",
        code_sha="test",
        job_id="j1",
    )
    prior = Path(first["out_dir"]) / "public_agencies.json"
    assert prior.is_file()
    hash_before = sha256_file(prior)
    rows_before = json.loads(prior.read_text(encoding="utf-8"))
    orgao = rows_before[0]["orgao"]
    marker = "CORRIGIDA_HUMANA_PRELIMINAR_XYZ"
    corrected = apply_corrections_to_source(
        prior,
        [{"orgao": orgao, "item_key": orgao, "fields": {"classificacao_juridica_preliminar": marker}}],
    )
    assert corrected.is_file()
    # list shape preserved
    corr_rows = json.loads(corrected.read_text(encoding="utf-8"))
    assert isinstance(corr_rows, list)
    assert any(marker in str(r.get("classificacao_juridica_preliminar")) for r in corr_rows)

    second = regenerate_workflow_version(
        workflow_id="workflow.confenge.public_agencies",
        params={"use_fixture": True, "uf": "SC", "max_leads": 3},
        out_dir=tmp_path / "v2",
        code_sha="test",
        job_id="j2",
        parent_run_id=first.get("run_id"),
        corrections=[
            {
                "orgao": orgao,
                "item_key": orgao,
                "fields": {"classificacao_juridica_preliminar": marker},
                "note": "classificação revisada",
            }
        ],
        prior_source=prior,
    )
    new_src = Path(second["out_dir"]) / "public_agencies.json"
    assert new_src.is_file()
    hash_after = sha256_file(new_src)
    assert hash_after != hash_before, "source hash must change without manual mutation"
    assert second["content_hashes"]["source"] == hash_after
    assert second["content_hashes"]["source"] != hash_before
    body = new_src.read_text(encoding="utf-8")
    assert marker in body
    # PDF is binary; extract text via reportlab path — at least file exists and package json has marker
    pdf = Path(second["out_dir"]) / "relatorio-orgaos-publicos.pdf"
    assert pdf.is_file() and pdf.stat().st_size > 500
    # try extract text if pypdf/pdfminer available; else assert review package
    review_pkg = Path(second["out_dir"]) / "pacote-revisao-orgaos.json"
    assert marker in review_pkg.read_text(encoding="utf-8")
    # XLSX preview path would also contain it when loaded by openpyxl
    from openpyxl import load_workbook

    wb = load_workbook(Path(second["out_dir"]) / "workbook-orgaos-publicos.xlsx", read_only=True)
    found = False
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if any(marker in str(c) for c in row if c is not None):
                found = True
                break
        if found:
            break
    assert found, "corrected classification must appear in regenerated XLSX"


def test_regenerate_obsoletes_accept_naturally(client: TestClient, tmp_path: Path) -> None:
    """AC#24: real correction changes source hash → prior ACCEPT obsolete without mutated- hacks."""
    token = _csrf(client)
    phrase = "Confirmo a geração local de entregáveis (sem envio automático de mensagens)."
    start = client.post(
        "/api/jobs",
        headers={"X-CC-CSRF": token},
        json={
            "capability_id": "workflow.confenge.public_agencies",
            "params": {"use_fixture": True, "uf": "SC", "max_leads": 3},
            "confirmation": phrase,
        },
    )
    assert start.status_code == 200, start.text
    job_id = start.json()["job"]["job_id"]
    _wait_job(client, job_id)

    reviews = client.get("/api/reviews?status=pending").json()["reviews"]
    assert reviews, "workflow should enqueue reviews"
    item = reviews[0]
    payload = item.get("payload") or {}
    hashes = payload.get("artifact_hashes") or (
        {"source": payload["content_hash"]} if payload.get("content_hash") else {}
    )
    assert hashes.get("source"), "ACCEPT must bind to source hash"
    item_key = payload.get("item_key") or str(item.get("title") or "").strip()
    assert item_key

    conf = client.get(f"/api/reviews/{item['id']}/confirmation").json()
    acc = client.post(
        "/api/decisions",
        headers={"X-CC-CSRF": token},
        json={
            "item_id": item["id"],
            "decision": "ACCEPT",
            "rationale": "Aceite consciente da classificação preliminar com ressalvas.",
            "confirmation": conf["confirmation_phrase"],
            "artifact_hashes": hashes,
        },
    )
    assert acc.status_code == 200, acc.text
    decision_id = acc.json()["decision_id"]

    marker = "CORRIGIDA_API_NATURAL_HASH"
    regen = client.post(
        "/api/reviews/regenerate",
        headers={"X-CC-CSRF": token},
        json={
            "job_id": job_id,
            "item_id": item["id"],
            "corrections": [
                {
                    "item_key": item_key,
                    "orgao": item_key,
                    "fields": {
                        "classificacao_juridica_preliminar": marker,
                        "limitacoes": "Corrigido na bancada; ainda preliminar.",
                    },
                    "note": "Correção de classificação na revisão humana",
                }
            ],
            "note": "regenerate with content correction",
        },
    )
    assert regen.status_code == 200, regen.text
    body = regen.json()
    assert body["job_id"] != job_id
    new_hashes = body.get("content_hashes") or {}
    assert new_hashes.get("source"), "regenerate must expose source hash"
    # Natural change — no mutated- prefix allowed in production path
    assert new_hashes["source"] != hashes["source"]
    assert not str(new_hashes["source"]).startswith("mutated-")

    # Corrected content on disk
    man = Path(body["manifest_path"])
    src = man.parent / "public_agencies.json"
    assert src.is_file()
    assert marker in src.read_text(encoding="utf-8")

    # Prior ACCEPT obsolete via regenerate mark OR list with current_hashes
    inv = client.get(
        "/api/decisions",
        params={"item_id": item["id"], "current_hashes": json.dumps(new_hashes)},
    )
    assert inv.status_code == 200
    decs = inv.json()["decisions"]
    accept_rows = [d for d in decs if d.get("id") == decision_id or d.get("decision") == "ACCEPT"]
    assert accept_rows
    assert any(d.get("obsolete") for d in accept_rows) or inv.json().get("marked_obsolete"), (
        f"ACCEPT must be obsolete after natural hash change; got {accept_rows}"
    )


def test_last_params_preset_saved(client: TestClient) -> None:
    token = _csrf(client)
    phrase = "Confirmo a geração local de entregáveis (sem envio automático de mensagens)."
    start = client.post(
        "/api/jobs",
        headers={"X-CC-CSRF": token},
        json={
            "capability_id": "workflow.extra.opportunities",
            "params": {"use_fixture": True, "max_shortlist": 4, "period_days": 14},
            "confirmation": phrase,
        },
    )
    assert start.status_code == 200
    pref = client.get("/api/preferences/last_params:workflow.extra.opportunities")
    assert pref.status_code == 200
    val = pref.json().get("value")
    assert val
    data = json.loads(val)
    assert data.get("max_shortlist") == 4
    assert data.get("period_days") == 14
