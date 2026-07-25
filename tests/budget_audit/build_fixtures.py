"""Build golden and operational workbooks for budget audit tests/cases."""

from __future__ import annotations

from datetime import UTC
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


def build_golden(path: Path) -> Path:
    """Controlled defects for detection proof."""
    wb = Workbook()

    # --- Orçamento Analítico ---
    ws = wb.active
    ws.title = "Orcamento Analitico"
    headers = [
        "Item", "Código", "Descrição", "Unidade", "Quantidade",
        "Custo Unit", "BDI %", "Preço Unit", "Preço Total",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Correct row: 10 * 100 = 1000
    ws.append([1, "SINAPI-001", "Concreto fck 25", "m³", 10, 80, 25, 100, 1000])
    # Wrong subtotal: 5 * 50 should be 250, written 300
    ws.append([2, "SINAPI-002", "Forma de madeira", "m²", 5, 40, 25, 50, 300])
    # Broken formula (#REF! token — never treat as zero)
    ws.append([3, "SINAPI-003", "Aço CA-50", "kg", 100, 8, 25, 10, None])
    ws["I4"] = "=#REF!"
    # Unit divergent row (same code as row1 different unit) — for matching tests
    ws.append([4, "SINAPI-001", "Concreto fck 25", "m²", 2, 80, 25, 100, 200])
    # Duplicate code intentional
    ws.append([5, "SINAPI-002", "Forma de madeira dup", "m²", 1, 40, 25, 50, 50])
    # Item without composition code match later
    ws.append([6, "NOCOMP-99", "Serviço sem composição", "vb", 1, 1000, 25, 1250, 1250])
    # Double BDI suspect: direct 100, bdi 25% => once 125, twice 156.25 — write twice
    ws.append([7, "DBL-001", "Item BDI duplo", "un", 1, 100, 25, 156.25, 156.25])
    # Negative quantity
    ws.append([8, "NEG-001", "Item quantidade negativa", "un", -2, 10, 25, 12.5, -25])
    # Zero quantity explicit
    ws.append([9, "ZERO-001", "Item quantidade zero", "un", 0, 10, 25, 12.5, 0])
    # Hidden row with material amount
    ws.append([10, "HID-001", "Item em linha oculta", "un", 3, 20, 25, 25, 75])
    ws.row_dimensions[11].hidden = True

    # --- BDI ---
    bdi = wb.create_sheet("BDI")
    bdi.append(["Componente", "%"])
    bdi.append(["Administração central", 5])
    bdi.append(["Riscos", 1.5])
    bdi.append(["Seguros", 0.5])
    bdi.append(["Garantias", 0.5])
    bdi.append(["Despesas financeiras", 1])
    bdi.append(["Tributos", 8])
    bdi.append(["Lucro", 6])
    bdi.append(["Tributos", 8])  # duplicate component
    bdi.append(["Total BDI declarado", 25])

    # --- Composições ---
    comp = wb.create_sheet("Composicoes")
    comp.append(["Código", "Descrição", "Unidade", "Quantidade", "Preço Unit", "Preço Total"])
    # correct: 2 * 10 = 20
    comp.append(["INS-01", "Cimento CP-II", "kg", 2, 10, 20])
    # wrong: 3 * 5 should be 15, written 18
    comp.append(["INS-02", "Areia média", "m³", 3, 5, 18])
    # missing coefficient
    comp.append(["INS-03", "Brita 1", "m³", None, 40, None])
    # negative coef
    comp.append(["INS-04", "Água", "L", -1, 0.01, -0.01])

    # --- Encargos ---
    enc = wb.create_sheet("Encargos Sociais")
    enc.append(["Componente", "%"])
    enc.append(["INSS", 20])
    enc.append(["FGTS", 8])
    enc.append(["RAT", 3])
    enc.append(["Outros", 250])  # out of scale

    # --- Cronograma ---
    cron = wb.create_sheet("Cronograma")
    cron.append(["Código", "Descrição", "Unidade", "Quantidade", "Preço Total", "Mês 1", "Mês 2"])
    cron.append(["SINAPI-001", "Concreto fck 25", "m³", 10, 1000, 600, 500])  # 1100 != 1000

    # --- ABC ---
    abc = wb.create_sheet("Curva ABC")
    abc.append(["Código", "Descrição", "Unidade", "Quantidade", "Preço Total"])
    abc.append(["SINAPI-001", "Concreto", "m³", 10, 1000])
    abc.append(["NOCOMP-99", "Serviço", "vb", 1, 1250])

    # External-like formula
    aux = wb.create_sheet("Auxiliar")
    aux.append(["Nota", "Valor"])
    aux["B2"] = "='[OtherWorkbook.xlsx]Sheet1'!A1"

    # Hidden sheet
    hid = wb.create_sheet("Oculta")
    hid.sheet_state = "hidden"
    hid.append(["segredo", 1])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def build_operational(path: Path, n_items: int = 60) -> Path:
    """Public-style engineering budget with 50+ items, BDI, compositions, formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orcamento Analitico"
    ws.append(
        [
            "Item", "Código", "Descrição", "Unidade", "Quantidade",
            "Custo Unit", "BDI %", "Preço Unit", "Preço Total",
        ]
    )
    services = [
        ("SINAPI-74001", "Concreto fck 20 MPa", "m³", 12.5, 420.0),
        ("SINAPI-74010", "Concreto fck 25 MPa", "m³", 8.0, 480.0),
        ("SINAPI-72321", "Forma madeira compensado", "m²", 120.0, 45.0),
        ("SINAPI-73128", "Armação aço CA-50", "kg", 850.0, 9.5),
        ("SINAPI-87215", "Alvenaria bloco cerâmico", "m²", 200.0, 68.0),
        ("SINAPI-88401", "Chapisco", "m²", 200.0, 12.0),
        ("SINAPI-88410", "Emboço", "m²", 200.0, 28.0),
        ("SINAPI-88420", "Reboco", "m²", 200.0, 22.0),
        ("SINAPI-91100", "Pintura PVA duas demãos", "m²", 180.0, 18.5),
        ("SINAPI-95001", "Piso cerâmico 45x45", "m²", 90.0, 55.0),
    ]
    bdi = 25.0
    for i in range(n_items):
        base = services[i % len(services)]
        code = f"{base[0]}-{i+1:03d}" if i >= len(services) else base[0]
        desc = f"{base[1]} (trecho {i+1})"
        unit = base[2]
        qty = round(base[3] * (1 + (i % 5) * 0.1), 3)
        custo = round(base[4] * (1 + (i % 7) * 0.02), 2)
        preco_unit = round(custo * (1 + bdi / 100.0), 2)
        # inject a few intentional errors
        if i == 15:
            total = round(qty * preco_unit + 150.0, 2)  # material error
        elif i == 30:
            total = round(qty * preco_unit + 0.02, 2)  # rounding-ish
        else:
            total = round(qty * preco_unit, 2)
        ws.append([i + 1, code, desc, unit, qty, custo, bdi, preco_unit, total])
        # formula for last col on some rows
        if i % 10 == 0:
            r = i + 2
            ws[f"I{r}"] = f"=E{r}*H{r}"

    bdi_ws = wb.create_sheet("BDI")
    bdi_ws.append(["Componente", "%"])
    for name, pct in [
        ("Administração central", 4.0),
        ("Riscos", 1.27),
        ("Seguros", 0.40),
        ("Garantias", 0.50),
        ("Despesas financeiras", 1.23),
        ("Tributos (ISS/PIS/COFINS)", 7.65),
        ("Lucro", 6.0),
        ("Total BDI", 21.05),
    ]:
        bdi_ws.append([name, pct])

    comp = wb.create_sheet("Composicoes")
    comp.append(["Código", "Descrição", "Unidade", "Quantidade", "Preço Unit", "Preço Total"])
    for i, (code, desc, unit, coef, price) in enumerate(
        [
            ("INS-CIM", "Cimento CP-II-32", "kg", 350.0, 0.85),
            ("INS-ARE", "Areia média", "m³", 0.6, 95.0),
            ("INS-BRI", "Brita 1", "m³", 0.8, 110.0),
            ("INS-ACO", "Aço CA-50 10mm", "kg", 1.0, 8.2),
            ("INS-MAD", "Madeira compensado 12mm", "m²", 1.05, 42.0),
        ],
        start=1,
    ):
        total = round(coef * price, 2)
        if i == 3:
            total = round(total + 5, 2)  # error
        comp.append([code, desc, unit, coef, price, total])

    enc = wb.create_sheet("Encargos Sociais")
    enc.append(["Componente", "%"])
    for name, pct in [("Grupo A", 35.0), ("Grupo B", 18.0), ("Grupo C", 12.0)]:
        enc.append([name, pct])

    cron = wb.create_sheet("Cronograma")
    cron.append(["Item", "Código", "Descrição", "Unidade", "Quantidade", "Preço Total"])
    for i in range(1, 13):
        cron.append([i, f"MES-{i:02d}", f"Parcela {i}", "vb", 1, 10000 + i * 100])

    abc = wb.create_sheet("Curva ABC")
    abc.append(["Código", "Descrição", "Unidade", "Quantidade", "Preço Total"])
    # filled from top items conceptually
    for i in range(1, 21):
        abc.append([f"ABC-{i:03d}", f"Item ABC {i}", "un", 1, 50000 - i * 1000])

    prop = wb.create_sheet("Proposta")
    prop.append(
        [
            "Item", "Código", "Descrição", "Unidade", "Quantidade",
            "Preço Unit", "Preço Total",
        ]
    )
    # proposal with discount on first 10
    for i in range(10):
        base = services[i % len(services)]
        qty = base[3]
        pu = round(base[4] * 1.25 * 0.95, 2)  # 5% discount vs sale
        prop.append([i + 1, base[0], base[1], base[2], qty, pu, round(qty * pu, 2)])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def build_reference_manifest(path: Path, items_path: Path) -> Path:
    import json
    from datetime import datetime

    from scripts.budget_audit.hashing import sha256_file

    items = []
    for i, (code, desc, unit, price) in enumerate(
        [
            ("SINAPI-74001", "Concreto fck 20 MPa", "m³", 430.0),
            ("SINAPI-74010", "Concreto fck 25 MPa", "m³", 490.0),
            ("SINAPI-72321", "Forma madeira compensado", "m²", 46.0),
            ("SINAPI-73128", "Armação aço CA-50", "kg", 9.8),
            ("SINAPI-001", "Concreto fck 25", "m³", 82.0),
        ]
    ):
        items.append(
            {
                "code": code,
                "description": desc,
                "unit": unit,
                "unit_price": price,
                "unit_direct_cost": price,
            }
        )
    items_path.write_text(
        "\n".join(json.dumps(x, ensure_ascii=False) for x in items) + "\n",
        encoding="utf-8",
    )
    # placeholder file for sha
    ref_blob = path.parent / "sinapi-sample-blob.bin"
    ref_blob.write_bytes(b"SINAPI-SAMPLE-REFERENCE-SNAPSHOT-NOT-OFFICIAL\n")
    manifest = {
        "system": "SINAPI",
        "publisher": "CAIXA/IBGE (sample snapshot for audit tooling — not live official feed)",
        "source_url": "https://www.caixa.gov.br/site/Paginas/sinapi.aspx",
        "acquired_at": datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "reference_month": "2026-01",
        "locality": "BR-SC",
        "tax_regime": "nao_desonerado",
        "file_name": ref_blob.name,
        "file_path": str(ref_blob.resolve()),
        "items_path": str(items_path.resolve()),
        "size": ref_blob.stat().st_size,
        "sha256": sha256_file(ref_blob),
        "license_or_access_note": "Sample fixture for tooling tests; not a redistributed official table",
        "parser_version": "1.0.0",
    }
    path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return path


if __name__ == "__main__":
    root = Path(__file__).resolve().parent / "fixtures"
    build_golden(root / "golden_case.xlsx")
    build_operational(root / "operational_public_style_budget.xlsx", 60)
    build_reference_manifest(
        root / "reference_manifest.json",
        root / "reference_items.jsonl",
    )
    print("fixtures ok", root)
