"""Opt-in PostgreSQL integration tests for migration 029 and QW-01 artifacts."""

from __future__ import annotations

import csv
import json
import os
from pathlib import Path
from uuid import uuid4

import psycopg2
import pytest
from openpyxl import load_workbook

from scripts.opportunity_intel.radar import run_radar
from scripts.opportunity_intel.schema import validate_qw01_schema

pytestmark = [
    pytest.mark.integration,
    pytest.mark.database,
    pytest.mark.skipif(
        os.getenv("REQUIRE_TEST_DB") != "1",
        reason="Set REQUIRE_TEST_DB=1 to run database tests",
    ),
]

DSN = os.getenv(
    "TEST_DSN",
    "postgresql://postgres:smartlic_local@127.0.0.1:54399/postgres",
)


@pytest.fixture()
def postgres_conn():
    if os.getenv("REQUIRE_TEST_DB") != "1":
        pytest.skip("Set REQUIRE_TEST_DB=1 to run QW-01 PostgreSQL integration tests")
    try:
        conn = psycopg2.connect(DSN)
    except psycopg2.Error as exc:
        pytest.fail(f"Required PostgreSQL is unavailable: {exc}")
    conn.autocommit = False
    try:
        yield conn
    finally:
        conn.rollback()
        conn.close()


def test_migration_029_schema_is_ready(postgres_conn) -> None:
    result = validate_qw01_schema(postgres_conn)
    assert result["backend"] == "postgresql"
    assert result["migration_029_ready"] is True


def test_success_zero_without_complete_pagination_is_rejected(postgres_conn) -> None:
    with pytest.raises(psycopg2.Error):
        with postgres_conn.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO coverage_evidence (
                    canonical_entity_key, source, data_type, queried_start, queried_end,
                    run_id, state, scope_key, pages_processed, freshness_status,
                    evidence_metadata
                ) VALUES (
                    %s, 'pncp', 'bids', CURRENT_DATE, CURRENT_DATE,
                    %s, 'success_zero', 'test-incomplete', 0, 'fresh',
                    '{"scope_complete": false}'::jsonb
                )
                """,
                (f"qw01-test-{uuid4().hex}", f"qw01-test-{uuid4().hex}"),
            )
    postgres_conn.rollback()


def test_partial_evidence_is_allowed_for_incomplete_scope(postgres_conn) -> None:
    with postgres_conn.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO coverage_evidence (
                canonical_entity_key, source, data_type, queried_start, queried_end,
                run_id, state, scope_key, pages_processed, freshness_status,
                evidence_metadata
            ) VALUES (
                %s, 'pncp', 'bids', CURRENT_DATE, CURRENT_DATE,
                %s, 'partial', 'test-partial', 1, 'unknown',
                '{"scope_complete": false}'::jsonb
            )
            """,
            (f"qw01-test-{uuid4().hex}", f"qw01-test-{uuid4().hex}"),
        )


def test_qw01_upsert_is_idempotent(postgres_conn) -> None:
    unique = uuid4().hex
    payload = [
        {
            "source_id": f"qw01-test-{unique}",
            "content_hash": f"qw01-test-hash-{unique}",
            "numero_controle_pncp": f"QW01-TEST-{unique}",
            "uf": "SC",
            "objeto": "Reforma predial para teste transacional",
            "status_canonico": "open",
            "link_anexos": None,
        }
    ]
    with postgres_conn.cursor() as cursor:
        cursor.execute(
            "SELECT action FROM upsert_qw01_pncp_opportunities(%s::jsonb)",
            (json.dumps(payload),),
        )
        first = cursor.fetchone()[0]
        cursor.execute(
            "SELECT action FROM upsert_qw01_pncp_opportunities(%s::jsonb)",
            (json.dumps(payload),),
        )
        second = cursor.fetchone()[0]
        cursor.execute(
            "SELECT count(*) FROM opportunity_intel WHERE numero_controle_pncp = %s",
            (payload[0]["numero_controle_pncp"],),
        )
        count = cursor.fetchone()[0]

    assert first == "insert"
    assert second == "update"
    assert count == 1


def test_one_command_emits_one_immutable_artifact_set(postgres_conn, tmp_path) -> None:
    # Close the fixture connection so the orchestrator owns an independent,
    # production-shaped PostgreSQL connection.
    postgres_conn.rollback()
    execution = run_radar(
        dsn=DSN,
        profile_path="config/client_profiles/extra.yaml",
        seed_path="fixtures/canonical_universe_r0.xlsx",
        window_days=45,
        output_root=tmp_path,
        update_mode="never",
    )
    run_dir = Path(execution.output_dir)
    expected = {
        "universe_snapshot.json",
        "coverage_manifest.json",
        "coverage_gaps.csv",
        "source_health.json",
        "source-applicability.csv",
        "radar_editais.csv",
        "radar_editais.xlsx",
        "run_manifest.json",
        "summary.md",
    }
    assert expected <= {path.name for path in run_dir.iterdir()}
    assert run_dir.name == execution.run_id

    for name in (
        "universe_snapshot.json",
        "coverage_manifest.json",
        "source_health.json",
        "run_manifest.json",
    ):
        document = json.loads((run_dir / name).read_text(encoding="utf-8"))
        assert document["metadata"]["run_id"] == execution.run_id

    with (run_dir / "coverage_gaps.csv").open(encoding="utf-8", newline="") as stream:
        reader = csv.DictReader(stream)
        rows = list(reader)
    assert reader.fieldnames is not None
    assert "run_id" in reader.fieldnames
    assert not rows or {row["run_id"] for row in rows} == {execution.run_id}

    workbook = load_workbook(run_dir / "radar_editais.xlsx", read_only=True, data_only=True)
    metadata = dict(workbook["Metadata"].iter_rows(values_only=True))
    workbook.close()
    assert metadata["run_id"] == execution.run_id
    assert execution.run_id in (run_dir / "summary.md").read_text(encoding="utf-8")
